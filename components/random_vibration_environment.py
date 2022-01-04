# -*- coding: utf-8 -*-
"""
This file defines a Random Vibration Environment where a specification is
defined and the controller solves for excitations that will cause the test
article to match the specified response.

This environment has a number of subprocesses, including CPSD and FRF
computation, data analysis, and signal generation.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from PyQt5 import QtWidgets,uic
from PyQt5.QtCore import QTimer
from .abstract_environment import AbstractEnvironment,AbstractUI,AbstractMetadata
from .utilities import DataAcquisitionParameters,VerboseMessageQueue,GlobalCommands,db2scale
from .ui_utilities import (
                           TransformationMatrixWindow,
                           load_python_module,load_specification,
                           multiline_plotter,
                           ACQUISITION_FRAMES_TO_DISPLAY,
                           PlotWindow
                           )
from .environments import (ControlTypes,environment_definition_ui_paths,
                           environment_prediction_ui_paths,
                           environment_run_ui_paths,
                           system_identification_ui_path,
                          )
import datetime
from multiprocessing.queues import Queue
import time
from enum import Enum
import multiprocessing as mp
import numpy as np
import inspect
import netCDF4 as nc4
import openpyxl

control_type = ControlTypes.RANDOM

maximum_name_length = 50

class RandomVibrationCommands(Enum):
    """Commands accepted by the Random Vibration Environment"""
    ANALYZE = 0
    INITIALIZE_CHANNELS = 1
    INITIALIZE_TEST_PARAMETERS = 2
    LOAD_SPECIFICATION = 3
    SHOW_SPECIFICATION = 4
    INITIALIZE_CONTROL_STRATEGY = 5
    INITIALIZE_LOGGING = 6
    LOGGING = 7
    START_TRANSFER_FUNCTION = 8
    STOP_TRANSFER_FUNCTION = 9
    START_CONTROL = 10
    STOP_CONTROL = 11
    STOP_ACQUISITION_AND_ANALYSIS = 12
    SHOW_TEST_PREDICTION = 13
    ADJUST_TEST_LEVEL = 14
    SHOW_FRF = 15

class RandomEnvironmentQueues:
    """A container class for the queues that random vibration will manage."""
    def __init__(self,
                 environment_name : str,
                 environment_command_queue : VerboseMessageQueue,
                 gui_update_queue : mp.queues.Queue,
                 controller_communication_queue : VerboseMessageQueue,
                 data_in_queue : mp.queues.Queue,
                 data_out_queue : mp.queues.Queue,
                 log_file_queue : VerboseMessageQueue
                 ):
        """A container class for the queues that random vibration will manage.
        
        The environment uses many queues to pass data between the various pieces.
        This class organizes those queues into one common namespace.
        

        Parameters
        ----------
        environment_name : str
            Name of the environment
        environment_command_queue : VerboseMessageQueue
            Queue that is read by the environment for environment commands
        gui_update_queue : mp.queues.Queue
            Queue where various subtasks put instructions for updating the
            widgets in the user interface
        controller_communication_queue : VerboseMessageQueue
            Queue that is read by the controller for global controller commands
        data_in_queue : mp.queues.Queue
            Multiprocessing queue that connects the acquisition subtask to the
            environment subtask.  Each environment will retrieve acquired data
            from this queue.
        data_out_queue : mp.queues.Queue
            Multiprocessing queue that connects the output subtask to the
            environment subtask.  Each environment will put data that it wants
            the controller to generate in this queue.
        log_file_queue : VerboseMessageQueue
            Queue for putting logging messages that will be read by the logging
            subtask and written to a file.
        """
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.data_analysis_command_queue = VerboseMessageQueue(log_file_queue,environment_name + ' Data Analysis Command Queue')
        self.signal_generation_command_queue = VerboseMessageQueue(log_file_queue,environment_name + ' Signal Generation Command Queue')
        self.frf_command_queue = VerboseMessageQueue(log_file_queue,environment_name + ' FRF Computation Command Queue')
        self.cpsd_command_queue = VerboseMessageQueue(log_file_queue,environment_name + ' CPSD Computation Command Queue')
        self.collector_command_queue = VerboseMessageQueue(log_file_queue,environment_name + ' Data Collector Command Queue')
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.data_for_frf_queue = mp.Queue()
        self.data_for_cpsd_queue = mp.Queue()
        self.updated_cpsd_queue = mp.Queue()
        self.updated_frf_queue = mp.Queue()
        self.cpsd_to_generate_queue = mp.Queue()
        self.log_file_queue = log_file_queue


class RandomVibrationParameters(AbstractMetadata):
    """Container to hold the signal processing parameters of the environment"""
    def __init__(self,sample_rate,samples_per_frame,test_level_ramp_time,
                 averaging_type,system_id_averages,averaging_coefficient,
                 frf_technique,frf_window,overlap_percentage,frf_voltage,update_tf_during_control,
                 cola_window,cola_overlap_percentage,cola_window_exponent,frames_in_cpsd,
                 cpsd_window,response_transformation_matrix,output_transformation_matrix,
                 control_python_script,control_python_function,control_python_function_type,
                 control_python_function_parameters,
                 specification_frequency_lines,specification_cpsd_matrix):
        """
        Container to hold the signal processing parameters of the environment

        Parameters
        ----------
        sample_rate : int
            Number of samples per second that the controller runs.
        samples_per_frame : int
            Number of samples per measurement frame.
        test_level_ramp_time : float
            Time taken to ramp between two test levels.
        averaging_type : TYPE
            Averaging type used to compute FRFs
        system_id_averages : int
            Number of averages to compute the FRFs
        averaging_coefficient : float
            Exponential averaging coefficient
        frf_technique : TYPE
            Technique to compute FRFs
        frf_window : TYPE
            Window function used to compute FRFs
        overlap_percentage : float
            Overlap percentage in FRF computation. 0-100
        frf_voltage : float
            RMS voltage of the signal used for system identification.
        update_tf_during_control : bool
            Whether or not to allow the controller to update the transfer functions
            during control.
        cola_window : TYPE
            Window function used for constant overlap and add.
        cola_overlap_percentage : float
            Percentage overlap during constant overlap and add. 0-100
        cola_window_exponent : float
            Exponent applied to the window function for constant overlap and add.
        frames_in_cpsd : int
            Number of frames used in CPSD computation.
        cpsd_window : TYPE
            Window used in CPSD computations.
        response_transformation_matrix : np.ndarray
            Response transformation.  None if Identity is used.
        output_transformation_matrix : TYPE
            Output transformation.  None if Identity is used.
        control_python_script : str
            Path to the Python file that contains the control script.
        control_python_function : str
            Function name of the control strategy within the script.
        control_python_function_type : TYPE
            Type of the Python data used for control Function/Generator/Etc.
        control_python_function_parameters : str
            Extra parameters passed to the Python function.
        specification_frequency_lines : np.ndarray
            Array of frequencies that the specification is defined at
        specification_cpsd_matrix : np.ndarray
            CPSD matrix that defines the specification.

        """
        
        self.sample_rate = sample_rate
        self.samples_per_frame = samples_per_frame
        self.test_level_ramp_time = test_level_ramp_time
        self.averaging_type = averaging_type
        self.system_id_averages = system_id_averages
        self.averaging_coefficient = averaging_coefficient
        self.frf_technique = frf_technique
        self.frf_window = frf_window
        self.frf_voltage = frf_voltage
        self.overlap = overlap_percentage/100
        self.update_tf_during_control = update_tf_during_control
        self.cola_window = cola_window
        self.cola_overlap = cola_overlap_percentage/100
        self.cola_window_exponent = cola_window_exponent
        self.frames_in_cpsd = frames_in_cpsd
        self.cpsd_window = cpsd_window
        self.response_transformation_matrix = response_transformation_matrix
        self.output_transformation_matrix = output_transformation_matrix
        self.control_python_script = control_python_script
        self.control_python_function = control_python_function
        self.control_python_function_type = control_python_function_type
        self.control_python_function_parameters = control_python_function_parameters
        self.specification_frequency_lines = specification_frequency_lines
        self.specification_cpsd_matrix = specification_cpsd_matrix
        
    @property
    def samples_per_acquire(self):
        """Property returning the samples per acquisition step given the overlap"""
        return int(self.samples_per_frame*(1-self.overlap))
    
    @property
    def frame_time(self):
        """Property returning the time per measurement frame"""
        return self.samples_per_frame/self.sample_rate
    
    @property
    def nyquist_frequency(self):
        """Property returning half the sample rate"""
        return self.sample_rate/2
    
    @property
    def fft_lines(self):
        """Property returning the frequency lines given the sampling parameters"""
        return self.samples_per_frame//2 + 1
    
    @property
    def frequency_spacing(self):
        """Property returning frequency line spacing given the sampling parameters"""
        return self.sample_rate/self.samples_per_frame
    
    @property
    def samples_per_output(self):
        """Property returning the samples per output given the COLA overlap"""
        return int(self.samples_per_frame*(1-self.cola_overlap))
    
    @property
    def overlapped_output_samples(self):
        """Property returning the number of output samples that are overlapped."""
        return self.samples_per_frame - self.samples_per_output
    
    def store_to_netcdf(self,netcdf_group_handle : nc4._netCDF4.Group):
        """
        Stores parameters to a netCDF group so they can be recovered.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            Reference to the netCDF4 group in which the environment data should
            be stored.

        """
        netcdf_group_handle.samples_per_frame = self.samples_per_frame
        netcdf_group_handle.test_level_ramp_time = self.test_level_ramp_time
        netcdf_group_handle.averaging_type = self.averaging_type
        netcdf_group_handle.system_id_averages = self.system_id_averages
        netcdf_group_handle.averaging_coefficient = self.averaging_coefficient
        netcdf_group_handle.frf_technique = self.frf_technique
        netcdf_group_handle.frf_window = self.frf_window
        netcdf_group_handle.overlap = self.overlap
        netcdf_group_handle.frf_voltage = self.frf_voltage
        netcdf_group_handle.update_tf_during_control = 1 if self.update_tf_during_control else 0
        netcdf_group_handle.cola_window = self.cola_window
        netcdf_group_handle.cola_overlap = self.cola_overlap
        netcdf_group_handle.cola_window_exponent = self.cola_window_exponent
        netcdf_group_handle.frames_in_cpsd = self.frames_in_cpsd
        netcdf_group_handle.cpsd_window = self.cpsd_window
        netcdf_group_handle.control_python_script = self.control_python_script
        netcdf_group_handle.control_python_function = self.control_python_function
        netcdf_group_handle.control_python_function_type = self.control_python_function_type
        netcdf_group_handle.control_python_function_parameters = self.control_python_function_parameters
        # Now set up specifications
        netcdf_group_handle.createDimension('fft_lines',self.fft_lines)
        netcdf_group_handle.createDimension('specification_channels',self.specification_cpsd_matrix.shape[-1])
        var = netcdf_group_handle.createVariable('specification_frequency_lines','f8',('fft_lines',))
        var[...] = self.specification_frequency_lines
        var = netcdf_group_handle.createVariable('specification_cpsd_matrix_real','f8',('fft_lines','specification_channels','specification_channels'))
        var[...] = self.specification_cpsd_matrix.real
        var = netcdf_group_handle.createVariable('specification_cpsd_matrix_imag','f8',('fft_lines','specification_channels','specification_channels'))
        var[...] = self.specification_cpsd_matrix.imag
        if not self.response_transformation_matrix is None:
            netcdf_group_handle.createDimension('response_transformation_rows',self.response_transformation_matrix.shape[0])
            netcdf_group_handle.createDimension('response_transformation_cols',self.response_transformation_matrix.shape[1])
            var = netcdf_group_handle.createVariable('response_transformation_matrix','f8',('response_transformation_rows','response_transformation_cols'))
            var[...] = self.response_transformation_matrix
        if not self.output_transformation_matrix is None:
            netcdf_group_handle.createDimension('output_transformation_rows',self.output_transformation_matrix.shape[0])
            netcdf_group_handle.createDimension('output_transformation_cols',self.output_transformation_matrix.shape[1])
            var = netcdf_group_handle.createVariable('output_transformation_matrix','f8',('output_transformation_rows','output_transformation_cols'))
            var[...] = self.output_transformation_matrix
    
    @classmethod
    def from_ui(cls,ui):
        """Creates a RandomVibrationParameters object from the user interface

        Parameters
        ----------
        ui : RandomVibrationUI
            A Random Vibration User Interface

        Returns
        -------
        test_parameters : RandomVibrationParameters
            Parameters corresponding to the data in the user interface.

        """
        if ui.python_control_module is None:
            control_module = None
            control_function = None
            control_function_type = None
            control_function_parameters = None
        else:
            control_module = ui.definition_widget.control_script_file_path_input.text()
            control_function = ui.definition_widget.control_function_input.itemText(ui.definition_widget.control_function_input.currentIndex())
            control_function_type = ui.definition_widget.control_function_generator_selector.currentIndex()
            control_function_parameters = ui.definition_widget.control_parameters_text_input.toPlainText()
        return cls(sample_rate=ui.definition_widget.sample_rate_display.value(),
                   samples_per_frame=ui.definition_widget.samples_per_frame_selector.value(),
                   test_level_ramp_time=ui.definition_widget.ramp_time_spinbox.value(),
                   averaging_type=ui.definition_widget.system_id_averaging_scheme_selector.itemText(ui.definition_widget.system_id_averaging_scheme_selector.currentIndex()),
                   system_id_averages=ui.definition_widget.system_id_frames_to_average_selector.value(),
                   averaging_coefficient=ui.definition_widget.system_id_averaging_coefficient_selector.value(),
                   frf_technique=ui.definition_widget.system_id_frf_technique_selector.itemText(ui.definition_widget.system_id_frf_technique_selector.currentIndex()),
                   frf_window=ui.definition_widget.system_id_transfer_function_computation_window_selector.itemText(ui.definition_widget.system_id_transfer_function_computation_window_selector.currentIndex()),
                   overlap_percentage=ui.definition_widget.system_id_overlap_percentage_selector.value(),
                   frf_voltage=ui.definition_widget.system_id_initial_drive_voltage_selector.value(),
                   update_tf_during_control=ui.definition_widget.update_transfer_function_during_control_selector.isChecked(),
                   cola_window=ui.definition_widget.cola_window_selector.itemText(ui.definition_widget.cola_window_selector.currentIndex()),
                   cola_overlap_percentage=ui.definition_widget.cola_overlap_percentage_selector.value(),
                   cola_window_exponent=ui.definition_widget.cola_exponent_selector.value(),
                   frames_in_cpsd=ui.definition_widget.cpsd_frames_selector.value(),
                   cpsd_window=ui.definition_widget.cpsd_computation_window_selector.itemText(ui.definition_widget.cpsd_computation_window_selector.currentIndex()),
                   response_transformation_matrix=ui.response_transformation_matrix,
                   output_transformation_matrix=ui.output_transformation_matrix,
                   control_python_script=control_module,
                   control_python_function=control_function,
                   control_python_function_type=control_function_type,
                   control_python_function_parameters=control_function_parameters,
                   specification_frequency_lines = ui.specification_frequency_lines,
                   specification_cpsd_matrix = ui.specification_cpsd_matrix)

from .random_vibration_data_analysis import random_data_analysis_process,RandomDataAnalysisMessages
from .random_vibration_signal_generation import random_signal_generation_process,RandomSignalGenerationMessages
from .frf_computation import frf_computation_process,FRFMessages
from .cpsd_computation import cpsd_computation_process,CPSDMessages
from .random_vibration_data_collector import random_data_collector_process,RandomDataCollectorMessages

class RandomVibrationUI(AbstractUI):
    """Class defining the user interface for a Random Vibration environment.
    
    This class will contain four main UIs, the environment definition,
    system identification, test prediction, and run.  The widgets corresponding
    to these interfaces are stored in TabWidgets in the main UI.
    
    This class defines all the call backs and user interface operations required
    for the Random Vibration environment."""
    def __init__(self,
                 environment_name : str,
                 definition_tabwidget : QtWidgets.QTabWidget,
                 system_id_tabwidget : QtWidgets.QTabWidget,
                 test_predictions_tabwidget : QtWidgets.QTabWidget,
                 run_tabwidget : QtWidgets.QTabWidget,
                 environment_command_queue : VerboseMessageQueue,
                 controller_communication_queue : VerboseMessageQueue,
                 log_file_queue : Queue):
        """
        Constructs a Random Vibration User Interface
        
        Given the tab widgets from the main interface as well as communication
        queues, this class assembles the user interface components specific to
        the Random Vibration Environment

        Parameters
        ----------
        definition_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Control
            Definition main tab
        system_id_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the System
            Identification main tab
        test_predictions_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Test Predictions
            main tab
        run_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Run
            main tab.
        environment_command_queue : VerboseMessageQueue
            Queue for sending commands to the Random Vibration Environment
        controller_communication_queue : VerboseMessageQueue
            Queue for sending global commands to the controller
        log_file_queue : Queue
            Queue where log file messages can be written.
    
        """
        super().__init__(environment_name,
             environment_command_queue,controller_communication_queue,log_file_queue)
        # Add the page to the control definition tabwidget
        self.definition_widget = QtWidgets.QWidget()
        uic.loadUi(environment_definition_ui_paths[control_type],self.definition_widget)
        definition_tabwidget.addTab(self.definition_widget,self.environment_name)
        # Add the page to the control prediction tabwidget
        self.prediction_widget = QtWidgets.QWidget()
        uic.loadUi(environment_prediction_ui_paths[control_type],self.prediction_widget)
        test_predictions_tabwidget.addTab(self.prediction_widget,self.environment_name)
        # Add the page to the system id tabwidget
        self.system_id_widget = QtWidgets.QWidget()
        uic.loadUi(system_identification_ui_path,self.system_id_widget)
        system_id_tabwidget.addTab(self.system_id_widget,self.environment_name)
        # Add the page to the run tabwidget
        self.run_widget = QtWidgets.QWidget()
        uic.loadUi(environment_run_ui_paths[control_type],self.run_widget)
        run_tabwidget.addTab(self.run_widget,self.environment_name)
        
        # Set up some persistent data
        self.data_acquisition_parameters = None
        self.environment_parameters = None
        self.plot_data_items = {}
        self.plot_windows = []
        self.run_start_time = None
        self.run_level_start_time = None
        self.run_timer = QTimer()
        self.response_transformation_matrix = None
        self.output_transformation_matrix = None
        self.python_control_module = None
        self.specification_frequency_lines = None
        self.specification_cpsd_matrix = None
        self.physical_control_names = None
        self.physical_output_names = None
        self.control_selector_widgets = None
        self.output_selector_widgets = None
        
        self.complete_ui()
        self.connect_callbacks()
        
        # Complete the profile commands
        self.command_map['Set Test Level'] = self.change_test_level_from_profile
        
    def retrieve_metadata(self, netcdf_handle : nc4._netCDF4.Dataset):
        """Collects environment parameters from a netCDF dataset.

        This function retrieves parameters from a netCDF dataset that was written
        by the controller during streaming.  It must populate the widgets
        in the user interface with the proper information.

        This function is the "read" counterpart to the store_to_netcdf 
        function in the RandomVibrationParameters class, which will write 
        parameters to the netCDF file to document the metadata.
        
        Note that the entire dataset is passed to this function, so the function
        should collect parameters pertaining to the environment from a Group
        in the dataset sharing the environment's name, e.g.
        
        ``group = netcdf_handle.groups[self.environment_name]``
        ``self.definition_widget.parameter_selector.setValue(group.parameter)``
        
        Parameters
        ----------
        netcdf_handle : nc4._netCDF4.Dataset :
            The netCDF dataset from which the data will be read.  It should have
            a group name with the enviroment's name.
        """
        # Get the group
        group = netcdf_handle.groups[self.environment_name]
        # Spinboxes
        self.definition_widget.samples_per_frame_selector.setValue(group.samples_per_frame)
        self.definition_widget.ramp_time_spinbox.setValue(group.test_level_ramp_time)
        self.definition_widget.system_id_frames_to_average_selector.setValue(group.system_id_averages)
        self.definition_widget.system_id_averaging_coefficient_selector.setValue(group.averaging_coefficient)
        self.definition_widget.system_id_initial_drive_voltage_selector.setValue(group.frf_voltage)
        self.definition_widget.system_id_overlap_percentage_selector.setValue(group.overlap*100)
        self.definition_widget.cola_overlap_percentage_selector.setValue(group.cola_overlap*100)
        self.definition_widget.cola_exponent_selector.setValue(group.cola_window_exponent)
        self.definition_widget.cpsd_frames_selector.setValue(group.frames_in_cpsd)
        # Checkboxes
        self.definition_widget.update_transfer_function_during_control_selector.setChecked(bool(group.update_tf_during_control))
        # Comboboxes
        self.definition_widget.system_id_averaging_scheme_selector.setCurrentIndex(self.definition_widget.system_id_averaging_scheme_selector.findText(group.averaging_type))
        self.definition_widget.system_id_frf_technique_selector.setCurrentIndex(self.definition_widget.system_id_frf_technique_selector.findText(group.frf_technique))
        self.definition_widget.system_id_transfer_function_computation_window_selector.setCurrentIndex(self.definition_widget.system_id_transfer_function_computation_window_selector.findText(group.frf_window))
        self.definition_widget.cola_window_selector.setCurrentIndex(self.definition_widget.cola_window_selector.findText(group.cola_window))
        self.definition_widget.cpsd_computation_window_selector.setCurrentIndex(self.definition_widget.cpsd_computation_window_selector.findText(group.cpsd_window))
        # Other data
        try:
            self.response_transformation_matrix = group.variables['response_transformation_matrix'][...].data
        except KeyError:
            self.response_transformation_matrix = None
        try:
            self.output_transformation_matrix = group.variables['output_transformation_matrix'][...].data
        except KeyError:
            self.output_transformation_matrix = None
        self.define_transformation_matrices(None,dialog=False)
        self.specification_frequency_lines = group.variables['specification_frequency_lines'][...].data
        self.specification_cpsd_matrix = group.variables['specification_cpsd_matrix_real'][...].data + 1j*group.variables['specification_cpsd_matrix_imag'][...].data
        self.select_python_module(None,group.control_python_script)
        self.definition_widget.control_function_input.setCurrentIndex(self.definition_widget.control_function_input.findText(group.control_python_function))
        self.definition_widget.control_parameters_text_input.setText(group.control_python_function_parameters)
        self.show_specification()
    
    def collect_environment_definition_parameters(self):
        """Collect the parameters from the user interface defining the environment

        Returns
        -------
        RandomVibrationParameters
            A metadata or parameters object containing the parameters defining
            the corresponding environment.
        """
        return RandomVibrationParameters.from_ui(self)
        
    def complete_ui(self):
        """Helper Function to continue setting up the user interface"""
        # Initialize enabled or disabled widgets
        self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        
        # Button Group IDs
        self.run_widget.test_time_button_group.setId(self.run_widget.continuous_test_radiobutton,0)
        self.run_widget.test_time_button_group.setId(self.run_widget.timed_test_radiobutton,1)
        
        # Change some sizes
        self.run_widget.total_test_time_display.setFixedWidth(45)
        self.run_widget.time_at_level_display.setFixedWidth(45)
        
        # Set common look and feel for plots
        plotWidgets = [self.definition_widget.specification_single_plot,
                       self.definition_widget.specification_sum_asds_plot,
                       self.system_id_widget.response_timehistory_plot,
                       self.system_id_widget.drive_timehistory_plot,
                       self.system_id_widget.transfer_function_phase_plot,
                       self.system_id_widget.transfer_function_amplitude_plot,
                       self.prediction_widget.excitation_display_plot,
                       self.prediction_widget.response_display_plot,
                       self.run_widget.global_test_performance_plot]
        for plotWidget in plotWidgets:
            plot_item = plotWidget.getPlotItem()
            plot_item.showGrid(True,True,0.25)
            plot_item.enableAutoRange()
            plot_item.getViewBox().enableAutoRange(enable=True)
        logscalePlotWidgets = [self.definition_widget.specification_single_plot,
                               self.definition_widget.specification_sum_asds_plot,
                               self.system_id_widget.transfer_function_amplitude_plot,
                               self.prediction_widget.excitation_display_plot,
                               self.prediction_widget.response_display_plot,
                               self.run_widget.global_test_performance_plot]
        for plotWidget in logscalePlotWidgets:
            plot_item = plotWidget.getPlotItem()
            plot_item.setLogMode(False,True)
    
    def connect_callbacks(self):
        """Helper function to connect callbacks to functions in the class"""
        # Definition
        self.definition_widget.samples_per_frame_selector.valueChanged.connect(self.update_parameters_and_clear_spec)
        self.definition_widget.system_id_overlap_percentage_selector.valueChanged.connect(self.update_parameters)
        self.definition_widget.cola_overlap_percentage_selector.valueChanged.connect(self.update_parameters)
        self.definition_widget.transformation_matrices_button.clicked.connect(self.define_transformation_matrices)
        self.definition_widget.system_id_averaging_scheme_selector.currentIndexChanged.connect(self.disable_exponential_coefficient)
        self.definition_widget.control_script_load_file_button.clicked.connect(self.select_python_module)
        self.definition_widget.control_function_input.currentIndexChanged.connect(self.update_generator_selector)
        self.definition_widget.load_spec_button.clicked.connect(self.select_spec_file)
        self.definition_widget.specification_row_selector.currentIndexChanged.connect(self.show_specification)
        self.definition_widget.specification_column_selector.currentIndexChanged.connect(self.show_specification)
        # System ID
        self.system_id_widget.preview_transfer_function_button.clicked.connect(self.preview_transfer_function)
        self.system_id_widget.acquire_transfer_function_button.clicked.connect(self.acquire_transfer_function)
        self.system_id_widget.stop_transfer_function_button.clicked.connect(self.stop_transfer_function)
        self.system_id_widget.select_transfer_function_stream_file_button.clicked.connect(self.select_transfer_function_stream_file)
        self.system_id_widget.voltage_scale_factor_selector.valueChanged.connect(self.change_transfer_function_test_level)
        self.system_id_widget.transfer_function_response_selector.currentIndexChanged.connect(self.show_frf)
        self.system_id_widget.transfer_function_reference_selector.currentIndexChanged.connect(self.show_frf)
        # Test Predictions
        self.prediction_widget.excitation_row_selector.currentIndexChanged.connect(self.show_test_predictions)
        self.prediction_widget.excitation_column_selector.currentIndexChanged.connect(self.show_test_predictions)
        self.prediction_widget.response_row_selector.currentIndexChanged.connect(self.show_test_predictions)
        self.prediction_widget.response_column_selector.currentIndexChanged.connect(self.show_test_predictions)
        self.prediction_widget.maximum_voltage_button.clicked.connect(self.show_max_voltage_prediction)
        self.prediction_widget.minimum_voltage_button.clicked.connect(self.show_min_voltage_prediction)
        self.prediction_widget.maximum_error_button.clicked.connect(self.show_max_error_prediction)
        self.prediction_widget.minimum_error_button.clicked.connect(self.show_min_error_prediction)
        # Run Test
        self.run_widget.current_test_level_selector.valueChanged.connect(self.change_control_test_level)
        self.run_widget.start_test_button.clicked.connect(self.start_control)
        self.run_widget.stop_test_button.clicked.connect(self.stop_control)
        self.run_widget.create_window_button.clicked.connect(self.create_window)
        self.run_widget.show_all_asds_button.clicked.connect(self.show_all_asds)
        self.run_widget.show_all_csds_phscoh_button.clicked.connect(self.show_all_csds_phscoh)
        self.run_widget.show_all_csds_realimag_button.clicked.connect(self.show_all_csds_realimag)
        self.run_widget.tile_windows_button.clicked.connect(self.tile_windows)
        self.run_widget.close_windows_button.clicked.connect(self.close_windows)
        self.run_timer.timeout.connect(self.update_run_time)
    
    def initialize_data_acquisition(self, data_acquisition_parameters : DataAcquisitionParameters):
        """Update the user interface with data acquisition parameters
        
        This function is called when the Data Acquisition parameters are
        initialized.  This function should set up the environment user interface
        accordingly.

        Parameters
        ----------
        data_acquisition_parameters : DataAcquisitionParameters :
            Container containing the data acquisition parameters, including
            channel table and sampling information.

        """
        self.log('Initializing Data Acquisition')
        # Initialize the plots
        # Clear plots if there is anything on them
        self.definition_widget.specification_single_plot.getPlotItem().clear()
        self.definition_widget.specification_sum_asds_plot.getPlotItem().clear()
        self.run_widget.global_test_performance_plot.getPlotItem().clear()
        
        # Now add initial lines that we can update later
        self.definition_widget.specification_single_plot.getPlotItem().addLegend()
        self.plot_data_items['specification_real'] = self.definition_widget.specification_single_plot.getPlotItem().plot(np.array([0,data_acquisition_parameters.sample_rate/2]),np.zeros(2),pen = {'color': "r", 'width': 1},name='Real Part')
        self.plot_data_items['specification_imag'] = self.definition_widget.specification_single_plot.getPlotItem().plot(np.array([0,data_acquisition_parameters.sample_rate/2]),np.zeros(2),pen = {'color': "b", 'width': 1},name='Imaginary Part')
        self.plot_data_items['specification_sum'] = self.definition_widget.specification_sum_asds_plot.getPlotItem().plot(np.array([0,data_acquisition_parameters.sample_rate/2]),np.zeros(2),pen = {'color': "r", 'width': 1})
        self.run_widget.global_test_performance_plot.getPlotItem().addLegend()
        self.plot_data_items['specification_sum_control'] = self.run_widget.global_test_performance_plot.getPlotItem().plot(np.array([0,data_acquisition_parameters.sample_rate/2]),np.zeros(2),pen = {'color': "b", 'width': 1},name='Specification')
        self.plot_data_items['sum_asds_control'] = self.run_widget.global_test_performance_plot.getPlotItem().plot(np.array([0,data_acquisition_parameters.sample_rate/2]),np.zeros(2),pen = {'color': "r", 'width': 1},name='Response')
        
        # Initialize channels
        channels = data_acquisition_parameters.channel_list
        num_control = len([channel for channel in channels if channel.control])
        num_output = len([channel for channel in channels if not channel.feedback_device is None])
        self.definition_widget.response_transformation_matrix = None#np.eye(num_control)
        self.definition_widget.output_transformation_matrix = None#np.eye(num_output)
        self.definition_widget.input_channels_display.setValue(len(channels))
        self.definition_widget.control_channels_display.setValue(num_control)
        self.definition_widget.output_channels_display.setValue(num_output)
        self.definition_widget.transform_channels_display.setValue(num_control)
        self.definition_widget.transform_outputs_display.setValue(num_output)
        self.physical_control_names = ['{:} {:} {:}'.format('' if channel.channel_type is None else channel.channel_type,channel.node_number,channel.node_direction)[:maximum_name_length]
            for channel in channels if channel.control]
        self.physical_output_names = ['{:} {:} {:}'.format('' if channel.channel_type is None else channel.channel_type,channel.node_number,channel.node_direction)[:maximum_name_length]
            for channel in channels if channel.feedback_device]
        self.control_selector_widgets = [
                self.definition_widget.specification_row_selector,
                self.definition_widget.specification_column_selector,
                self.system_id_widget.transfer_function_response_selector,
                self.prediction_widget.response_row_selector,
                self.prediction_widget.response_column_selector,
                self.run_widget.control_channel_1_selector,
                self.run_widget.control_channel_2_selector]
        self.output_selector_widgets = [
                self.system_id_widget.transfer_function_reference_selector,
                self.prediction_widget.excitation_row_selector,
                self.prediction_widget.excitation_column_selector,]
        for widget in self.control_selector_widgets:
            widget.blockSignals(True)
            widget.clear()
        for widget in self.output_selector_widgets:
            widget.blockSignals(True)
            widget.clear()
        for i,control_name in enumerate(self.physical_control_names):
            for widget in self.control_selector_widgets:
                widget.addItem('{:}: {:}'.format(i+1,control_name))
        for i,drive_name in enumerate(self.physical_output_names):
            for widget in self.output_selector_widgets:
                widget.addItem('{:}: {:}'.format(i+1,drive_name))
        for widget in self.control_selector_widgets:
            widget.blockSignals(False)
        for widget in self.output_selector_widgets:
            widget.blockSignals(False)
        self.definition_widget.sample_rate_display.setValue(data_acquisition_parameters.sample_rate)
        self.definition_widget.samples_per_frame_selector.setValue(data_acquisition_parameters.sample_rate)
        # Store for later
        self.data_acquisition_parameters = data_acquisition_parameters

    
    ### Definition Callbacks
        
    def update_parameters(self):
        """Recompute derived parameters from updated sampling parameters"""
        data = self.collect_environment_definition_parameters()
        self.definition_widget.samples_per_acquire_display.setValue(data.samples_per_acquire)
        self.definition_widget.frame_time_display.setValue(data.frame_time)
        self.definition_widget.nyquist_frequency_display.setValue(data.nyquist_frequency)
        self.definition_widget.fft_lines_display.setValue(data.fft_lines)
        self.definition_widget.frequency_spacing_display.setValue(data.frequency_spacing)
        self.definition_widget.samples_per_write_display.setValue(data.samples_per_output)
        
    def update_parameters_and_clear_spec(self):
        """Clears the specification data and updates parameters"""
        samples_per_frame = self.definition_widget.samples_per_frame_selector.value()
        if samples_per_frame % 2 != 0:
            self.definition_widget.samples_per_frame_selector.blockSignals(True)
            self.definition_widget.samples_per_frame_selector.setValue(samples_per_frame+1)
            self.definition_widget.samples_per_frame_selector.blockSignals(False)
        self.specification_frequency_lines = None
        self.specification_cpsd_matrix = None
        self.definition_widget.specification_file_name_display.setText('')
        self.show_specification()
        self.update_parameters()
        
    def define_transformation_matrices(self,clicked,dialog = True):
        """Defines the transformation matrices using the dialog box"""
        if dialog:
            (response_transformation,output_transformation,result
             ) = TransformationMatrixWindow.define_transformation_matrices(
                    self.response_transformation_matrix,
                    self.definition_widget.control_channels_display.value(),
                    self.output_transformation_matrix,
                    self.definition_widget.output_channels_display.value(),
                    self.definition_widget)
        else:
            response_transformation = self.response_transformation_matrix
            output_transformation = self.output_transformation_matrix
            result = True
        if result:
            # Update the control names
            for widget in self.control_selector_widgets:
                widget.blockSignals(True)
                widget.clear()
            if response_transformation is None:
                for i,control_name in enumerate(self.physical_control_names):
                    for widget in self.control_selector_widgets:
                        widget.addItem('{:}: {:}'.format(i+1,control_name))
                self.definition_widget.transform_channels_display.setValue(len(self.physical_control_names))
            else:
                for i in range(response_transformation.shape[0]):
                    for widget in self.control_selector_widgets:
                        widget.addItem('{:}: {:}'.format(i+1,'Virtual Response'))
                self.definition_widget.transform_channels_display.setValue(response_transformation.shape[0])
            for widget in self.control_selector_widgets:
                widget.blockSignals(False)
            # Update the output names
            for widget in self.output_selector_widgets:
                widget.blockSignals(True)
                widget.clear()
            if output_transformation is None:
                for i,drive_name in enumerate(self.physical_output_names):
                    for widget in self.output_selector_widgets:
                        widget.addItem('{:}: {:}'.format(i+1,drive_name))
                self.definition_widget.transform_outputs_display.setValue(len(self.physical_output_names))
            else:
                for i in range(output_transformation.shape[0]):
                    for widget in self.output_selector_widgets:
                        widget.addItem('{:}: {:}'.format(i+1,'Virtual Drive'))
                self.definition_widget.transform_outputs_display.setValue(output_transformation.shape[0])
            for widget in self.output_selector_widgets:
                widget.blockSignals(False)
                
            self.response_transformation_matrix = response_transformation
            self.output_transformation_matrix = output_transformation
            self.update_parameters_and_clear_spec()
    
    def disable_exponential_coefficient(self):
        """Disables the exponential averaging coefficient when averaging is linear"""
        if self.definition_widget.system_id_averaging_scheme_selector.currentIndex() == 0:
            self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        else:
            self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(True)
    
    def select_python_module(self,clicked,filename=None):
        """Loads a Python module using a dialog or the specified filename

        Parameters
        ----------
        clicked :
            The clicked event that triggered the callback.
        filename :
            File name defining the Python module for bypassing the callback when
            loading from a file (Default value = None).

        """
        if filename is None:
            filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self.definition_widget,'Select Python Module',filter='Python Modules (*.py)')
            if filename == '':
                return
        self.python_control_module = load_python_module(filename)
        functions = [function for function in inspect.getmembers(self.python_control_module)
                     if (inspect.isfunction(function[1]) and len(inspect.signature(function[1]).parameters)>=6)
                     or inspect.isgeneratorfunction(function[1])
                     or (inspect.isclass(function[1]) and all([method in function[1].__dict__ for method in ['system_id_update','control']]))]
        self.log('Loaded module {:} with functions {:}'.format(self.python_control_module.__name__,[function[0] for function in functions]))
        self.definition_widget.control_function_input.clear()
        self.definition_widget.control_script_file_path_input.setText(filename)
        for function in functions:
            self.definition_widget.control_function_input.addItem(function[0])
            
    def update_generator_selector(self):
        """Updates the function/generator selector based on the function selected"""
        if self.python_control_module is None:
            return
        function = getattr(self.python_control_module,self.definition_widget.control_function_input.itemText(self.definition_widget.control_function_input.currentIndex()))
        if inspect.isgeneratorfunction(function):
            self.definition_widget.control_function_generator_selector.setCurrentIndex(1)
        elif inspect.isclass(function):
            self.definition_widget.control_function_generator_selector.setCurrentIndex(2)
        else:
            self.definition_widget.control_function_generator_selector.setCurrentIndex(0)
    
    def select_spec_file(self,clicked,filename=None):
        """Loads a specification using a dialog or the specified filename

        Parameters
        ----------
        clicked :
            The clicked event that triggered the callback.
        filename :
            File name defining the specification for bypassing the callback when
            loading from a file (Default value = None).

        """
        if filename is None:
            filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self.definition_widget,'Select Specification File',filter='Numpyz or Mat (*.npz *.mat)')
            if filename == '':
                return
        self.definition_widget.specification_file_name_display.setText(filename)
        self.specification_frequency_lines,self.specification_cpsd_matrix = load_specification(filename,self.definition_widget.fft_lines_display.value(),self.definition_widget.frequency_spacing_display.value())
        self.show_specification()
    
    def show_specification(self):
        """Show the specification on the GUI"""
        if self.specification_cpsd_matrix is None:
            self.plot_data_items['specification_real'].setData(np.array([0,self.definition_widget.sample_rate_display.value()/2]),np.zeros(2))
            self.plot_data_items['specification_imag'].setData(np.array([0,self.definition_widget.sample_rate_display.value()/2]),np.zeros(2))
            self.plot_data_items['specification_sum'].setData(np.array([0,self.definition_widget.sample_rate_display.value()/2]),np.zeros(2))
            # enabled_state = self.run_widget.isEnabled()
            # self.run_widget.setEnabled(True)
            self.plot_data_items['specification_sum_control'].setData(np.array([0,self.definition_widget.sample_rate_display.value()/2]),np.zeros(2))
            # self.run_widget.setEnabled(enabled_state)
        else:
            row = self.definition_widget.specification_row_selector.currentIndex()
            column = self.definition_widget.specification_column_selector.currentIndex()
            spec_real = abs(self.specification_cpsd_matrix[:,row,column].real)
            spec_imag = abs(self.specification_cpsd_matrix[:,row,column].imag)
            spec_sum = abs(np.einsum('ijj',self.specification_cpsd_matrix))
            self.plot_data_items['specification_real'].setData(self.specification_frequency_lines[spec_real > 0.0],spec_real[spec_real > 0.0])
            self.plot_data_items['specification_imag'].setData(self.specification_frequency_lines[spec_imag > 0.0],spec_imag[spec_imag > 0.0])
            self.plot_data_items['specification_sum'].setData(self.specification_frequency_lines[spec_sum > 0.0],spec_sum[spec_sum > 0.0])
            # enabled_state = self.run_widget.isEnabled()
            # self.run_widget.setEnabled(True)
            self.plot_data_items['specification_sum_control'].setData(self.specification_frequency_lines[spec_sum > 0.0],spec_sum[spec_sum > 0.0])
            # self.run_widget.setEnabled(enabled_state)
    
    def initialize_environment(self):
        """Update the user interface with environment parameters
        
        This function is called when the Environment parameters are initialized.
        This function should set up the user interface accordingly.  It must
        return the parameters class of the environment that inherits from
        AbstractMetadata.

        Returns
        -------
        environment_parameters : RandomVibrationParameters
            A RandomVibrationParameters object that contains the parameters
            defining the environment.
        """
        self.log('Initializing Environment Parameters')
        data = self.collect_environment_definition_parameters()
        # Make sure everything is defined
        if data.specification_cpsd_matrix is None:
            raise ValueError('Specification has not been defined for {:}'.format(self.environment_name))
        if data.control_python_script is None:
            raise ValueError('Control function has not been loaded for {:}'.format(self.environment_name))
        # Set the value in the system_id_num_averages_display
        self.system_id_widget.system_id_num_averages_display.setValue(data.system_id_averages)
        self.system_id_widget.voltage_display.setValue(self.definition_widget.system_id_initial_drive_voltage_selector.value())
        # Initialize the rest of the plots
        self.system_id_widget.response_timehistory_plot.getPlotItem().clear()
        self.system_id_widget.drive_timehistory_plot.getPlotItem().clear()
        self.system_id_widget.transfer_function_phase_plot.getPlotItem().clear()
        self.system_id_widget.transfer_function_amplitude_plot.getPlotItem().clear()
        self.prediction_widget.excitation_display_plot.getPlotItem().clear()
        self.prediction_widget.response_display_plot.getPlotItem().clear()
        
        self.plot_data_items['control_responses'] = multiline_plotter(
                np.arange(data.samples_per_acquire*ACQUISITION_FRAMES_TO_DISPLAY)/data.sample_rate,
                np.zeros((self.definition_widget.transform_channels_display.value(),data.samples_per_acquire*ACQUISITION_FRAMES_TO_DISPLAY)),
                widget=self.system_id_widget.response_timehistory_plot,
                other_pen_options={'width':1},
                names = [self.system_id_widget.transfer_function_response_selector.itemText(i) for i in range(self.system_id_widget.transfer_function_response_selector.count())])
        self.plot_data_items['drive_outputs'] = multiline_plotter(
                np.arange(data.samples_per_acquire*ACQUISITION_FRAMES_TO_DISPLAY)/data.sample_rate,
                np.zeros((self.definition_widget.transform_outputs_display.value(),data.samples_per_acquire*ACQUISITION_FRAMES_TO_DISPLAY)),
                widget=self.system_id_widget.drive_timehistory_plot,
                other_pen_options={'width':1},
                names = [self.system_id_widget.transfer_function_reference_selector.itemText(i) for i in range(self.system_id_widget.transfer_function_reference_selector.count())])
        self.plot_data_items['transfer_function_phase'] = self.system_id_widget.transfer_function_phase_plot.getPlotItem().plot(np.arange(data.fft_lines)*data.frequency_spacing,np.zeros(data.fft_lines),pen = {'color': "r", 'width': 1},name='Phase')
        self.plot_data_items['transfer_function_amplitude'] = self.system_id_widget.transfer_function_amplitude_plot.getPlotItem().plot(np.arange(data.fft_lines)*data.frequency_spacing,np.zeros(data.fft_lines),pen = {'color': "r", 'width': 1},name='Amplitude')
        self.prediction_widget.excitation_display_plot.getPlotItem().addLegend()
        self.prediction_widget.response_display_plot.getPlotItem().addLegend()
        self.plot_data_items['response_prediction'] = multiline_plotter(
                np.arange(data.fft_lines)*data.frequency_spacing,
                np.zeros((4,data.fft_lines)),
                widget = self.prediction_widget.response_display_plot,
                other_pen_options={'width':1},
                names = ['Real Prediction','Real Spec','Imag Prediction','Imag Spec']
                )
        self.plot_data_items['excitation_prediction'] = multiline_plotter(
                np.arange(data.fft_lines)*data.frequency_spacing,
                np.zeros((2,data.fft_lines)),
                widget = self.prediction_widget.excitation_display_plot,
                other_pen_options={'width':1},
                names = ['Real Prediction','Imag Prediction']
                )
        # Store for later
        self.environment_parameters = data
        
        return data
    
    ### System Identification Callbacks
    def preview_transfer_function(self):
        """Starts the transfer function in preview mode"""
        self.log('Starting Transfer Function Preview')
        self.controller_communication_queue.put(self.log_name,(GlobalCommands.RUN_HARDWARE,None))
        self.controller_communication_queue.put(self.log_name,(GlobalCommands.START_ENVIRONMENT,self.environment_name))
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.START_TRANSFER_FUNCTION,(False,db2scale(self.system_id_widget.voltage_scale_factor_selector.value()))))
    
    def acquire_transfer_function(self):
        """Starts the transfer function in acquire mode"""
        self.log('Starting Transfer Function Acquire')
        if self.system_id_widget.stream_transfer_function_data_checkbox.isChecked():
            self.controller_communication_queue.put(self.log_name,(GlobalCommands.INITIALIZE_STREAMING,self.system_id_widget.transfer_function_stream_file_display.text()))
            self.controller_communication_queue.put(self.log_name,(GlobalCommands.START_STREAMING,None))
        self.controller_communication_queue.put(self.log_name,(GlobalCommands.RUN_HARDWARE,None))
        self.controller_communication_queue.put(self.log_name,(GlobalCommands.START_ENVIRONMENT,self.environment_name))
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.START_TRANSFER_FUNCTION,(True,db2scale(self.system_id_widget.voltage_scale_factor_selector.value()))))
        
    def stop_transfer_function(self):
        """Stops the transfer function acquisition"""
        self.log('Stopping Transfer Function')
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.STOP_TRANSFER_FUNCTION,None))
    
    def select_transfer_function_stream_file(self):
        """Select a file to save transfer function data to"""
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self.system_id_widget,'Select NetCDF File to Save Transfer Function Data',filter='NetCDF File (*.nc4)')
        if filename == '':
            return
        self.system_id_widget.transfer_function_stream_file_display.setText(filename)
        self.system_id_widget.stream_transfer_function_data_checkbox.setChecked(True)
    
    def change_transfer_function_test_level(self):
        """Updates the test level for the transfer function"""
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.ADJUST_TEST_LEVEL,db2scale(self.system_id_widget.voltage_scale_factor_selector.value())))
    
    def show_frf(self):
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.SHOW_FRF,None))
    
    ### Prediction Callbacks
    
    def show_test_predictions(self):
        """Shows test predictions"""
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.SHOW_TEST_PREDICTION,None))
    
    def show_max_voltage_prediction(self):
        widget = self.prediction_widget.excitation_voltage_list
        index = np.argmax([float(widget.item(v).text())
                           for v in range(widget.count())])
        self.prediction_widget.excitation_row_selector.setCurrentIndex(index)
        self.prediction_widget.excitation_column_selector.setCurrentIndex(index)
    def show_min_voltage_prediction(self):
        widget = self.prediction_widget.excitation_voltage_list
        index = np.argmin([float(widget.item(v).text())
                           for v in range(widget.count())])
        self.prediction_widget.excitation_row_selector.setCurrentIndex(index)
        self.prediction_widget.excitation_column_selector.setCurrentIndex(index)
    def show_max_error_prediction(self):
        widget = self.prediction_widget.response_error_list
        index = np.argmax([float(widget.item(v).text())
                           for v in range(widget.count())])
        self.prediction_widget.response_row_selector.setCurrentIndex(index)
        self.prediction_widget.response_column_selector.setCurrentIndex(index)
    def show_min_error_prediction(self):
        widget = self.prediction_widget.response_error_list
        index = np.argmin([float(widget.item(v).text())
                           for v in range(widget.count())])
        self.prediction_widget.response_row_selector.setCurrentIndex(index)
        self.prediction_widget.response_column_selector.setCurrentIndex(index)
    
    ### Control Callbacks
    def get_test_time_button_index(self):
        """Get the button index that determines the timing of the run.
        
        Debug function, not actually used."""
        print(self.run_widget.test_time_button_group.checkedId())
        
    def change_control_test_level(self):
        """Updates the test level of the control."""
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.ADJUST_TEST_LEVEL,db2scale(self.run_widget.current_test_level_selector.value())))
        self.run_level_start_time = time.time()
        # Check and see if we need to start streaming data
        if self.run_widget.current_test_level_selector.value() == self.run_widget.target_test_level_selector.value():
            self.controller_communication_queue.put(self.log_name,(GlobalCommands.AT_TARGET_LEVEL,self.environment_name))
        
    def start_control(self):
        """Starts running the control."""
        self.run_widget.stop_test_button.setEnabled(True)
        self.run_widget.start_test_button.setEnabled(False)
        self.run_widget.target_test_level_selector.setEnabled(False)
        self.run_widget.continuous_test_radiobutton.setEnabled(False)
        self.run_widget.timed_test_radiobutton.setEnabled(False)
        self.run_widget.time_test_at_target_level_checkbox.setEnabled(False)
        self.run_widget.test_time_selector.setEnabled(False)
        self.controller_communication_queue.put(self.log_name,(GlobalCommands.START_ENVIRONMENT,self.environment_name))
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.START_CONTROL,db2scale(self.run_widget.current_test_level_selector.value())))
        self.run_timer.start(250)
        self.run_start_time = time.time()
        self.run_level_start_time = self.run_start_time
        self.run_widget.test_progress_bar.setValue(0)

    def stop_control(self):
        """Stops running the control"""
        # I might have to move these into one of the tasks to ensure that we
        # have shut down successfully prior to re-enabling.
        self.run_widget.stop_test_button.setEnabled(False)
        self.run_widget.start_test_button.setEnabled(True)
        self.run_widget.target_test_level_selector.setEnabled(True)
        self.run_widget.continuous_test_radiobutton.setEnabled(True)
        self.run_widget.timed_test_radiobutton.setEnabled(True)
        self.run_widget.time_test_at_target_level_checkbox.setEnabled(True)
        self.run_widget.test_time_selector.setEnabled(True)
        self.environment_command_queue.put(self.log_name,(RandomVibrationCommands.STOP_CONTROL,None))
        self.run_timer.stop()

    def update_run_time(self):
        """Updates the time that the control has been running on the GUI"""
        # Update the total run time
        current_time = time.time()
        time_elapsed = current_time-self.run_start_time
        time_at_level_elapsed = current_time-self.run_level_start_time
        self.run_widget.total_test_time_display.setText(str(datetime.timedelta(seconds=time_elapsed)).split('.')[0])
        self.run_widget.time_at_level_display.setText(str(datetime.timedelta(seconds=time_at_level_elapsed)).split('.')[0])
        # Check if we need to stop the test due to timeout
        if self.run_widget.timed_test_radiobutton.isChecked():
            check_time = self.run_widget.test_time_selector.time()
            check_time_seconds = check_time.hour()*3600 + check_time.minute()*60 + check_time.second()
            if self.run_widget.time_test_at_target_level_checkbox.isChecked():
                if self.run_widget.current_test_level_selector.value() == self.run_widget.target_test_level_selector.value():
                    self.run_widget.test_progress_bar.setValue(time_at_level_elapsed/check_time_seconds*100)
                    if time_at_level_elapsed > check_time_seconds:
                        self.stop_control()
                else:
                    self.run_widget.test_progress_bar.setValue(0)
            else:
                self.run_widget.test_progress_bar.setValue(time_elapsed/check_time_seconds*100)
                if time_elapsed > check_time_seconds:
                    self.stop_control()
        
    def change_test_level_from_profile(self,test_level):
        """Sets the test level from a profile instruction

        Parameters
        ----------
        test_level :
            Value to set the test level to.
        """
        self.run_widget.current_test_level_selector.setValue(int(test_level))
    
    def create_window(self,event,row_index = None,column_index=None,datatype_index = None):
        """Creates a subwindow to show a specific channel information

        Parameters
        ----------
        event :
            
        row_index :
            Row index in the CPSD matrix to display (Default value = None)
        column_index :
            Column index in the CPSD matrix to display (Default value = None)
        datatype_index :
            Data type to display (real,imag,mag,phase,etc) (Default value = None)

        """
        if row_index is None:
            row_index = self.run_widget.control_channel_1_selector.currentIndex()
        if column_index is None:
            column_index = self.run_widget.control_channel_2_selector.currentIndex()
        if datatype_index is None:
            datatype_index = self.run_widget.data_type_selector.currentIndex()
        self.plot_windows.append(
                PlotWindow(None,row_index, column_index,datatype_index,
                          (self.specification_frequency_lines,self.specification_cpsd_matrix),
                          self.run_widget.control_channel_1_selector.itemText(row_index),
                          self.run_widget.control_channel_2_selector.itemText(column_index),
                          self.run_widget.data_type_selector.itemText(datatype_index)
                          ))
    
    def show_all_asds(self):
        """Creates a subwindow for each ASD in the CPSD matrix"""
        for i in range(self.specification_cpsd_matrix.shape[-1]):
            self.create_window(None,i,i,0)
        self.tile_windows()
    
    def show_all_csds_phscoh(self):
        """Creates a subwindow for each entry in the CPSD matrix showing phase and coherence"""
        for i in range(self.specification_cpsd_matrix.shape[-1]):
            for j in range(self.specification_cpsd_matrix.shape[-1]):
                if i == j:
                    datatype_index = 0
                elif i < j:
                    datatype_index = 1
                elif i > j:
                    datatype_index = 2
                self.create_window(None,i,j,datatype_index)
        self.tile_windows()
    
    def show_all_csds_realimag(self):
        """Creates a subwindow for each entry in the CPSD matrix showing real and imaginary"""
        for i in range(self.specification_cpsd_matrix.shape[-1]):
            for j in range(self.specification_cpsd_matrix.shape[-1]):
                if i == j:
                    datatype_index = 0
                elif i < j:
                    datatype_index = 3
                elif i > j:
                    datatype_index = 4
                self.create_window(None,i,j,datatype_index)
        self.tile_windows()
    
    def tile_windows(self):
        """Tile subwindow equally across the screen"""
        screen_rect = QtWidgets.QApplication.desktop().screenGeometry()
        # Go through and remove any closed windows
        self.plot_windows = [window for window in self.plot_windows if window.isVisible()]
        num_windows = len(self.plot_windows)
        ncols = int(np.ceil(np.sqrt(num_windows)))
        nrows = int(np.ceil(num_windows/ncols))
        window_width = screen_rect.width()/ncols
        window_height = screen_rect.height()/nrows
        for index,window in enumerate(self.plot_windows):
            window.resize(window_width,window_height)
            row_ind = index // ncols
            col_ind = index % ncols
            window.move(col_ind*window_width,row_ind*window_height)
            
    def close_windows(self):
        """Close all subwindows"""
        for window in self.plot_windows:
            window.close()
    
    def update_gui(self,queue_data):
        """Update the graphical interface for the environment

        Parameters
        ----------
        queue_data :
            A 2-tuple consisting of ``(message,data)`` pairs where the message
            denotes what to change and the data contains the information needed
            to be displayed.  
        """
        message,data = queue_data
        if message == 'FRF':
            # Display the data
            self.plot_data_items['transfer_function_phase'].setData(data[0],
                                np.angle(data[1][:,self.system_id_widget.transfer_function_response_selector.currentIndex(),
                                                   self.system_id_widget.transfer_function_reference_selector.currentIndex()]))
            self.plot_data_items['transfer_function_amplitude'].setData(data[0],
                                np.abs(data[1][:,self.system_id_widget.transfer_function_response_selector.currentIndex(),
                                                 self.system_id_widget.transfer_function_reference_selector.currentIndex()]))
        elif message == 'control_predictions':
            frequencies,excitation_prediction,response_prediction,spec = data
            row_index = self.prediction_widget.excitation_row_selector.currentIndex()
            column_index = self.prediction_widget.excitation_column_selector.currentIndex()
            self.plot_data_items['excitation_prediction'][0].setData(frequencies,
                                np.abs(np.real(excitation_prediction[:,row_index,column_index])))
            imag_part = np.abs(np.imag(excitation_prediction[:,row_index,column_index]))
            imag_part[imag_part < 1e-15] = 0
            self.plot_data_items['excitation_prediction'][1].setData(frequencies,
                                imag_part)
            row_index = self.prediction_widget.response_row_selector.currentIndex()
            column_index = self.prediction_widget.response_column_selector.currentIndex()
            self.plot_data_items['response_prediction'][0].setData(frequencies,
                                np.abs(np.real(response_prediction[:,row_index,column_index])))
            imag_part = np.abs(np.imag(response_prediction[:,row_index,column_index]))
            imag_part[imag_part < 1e-15] = 0
            self.plot_data_items['response_prediction'][2].setData(frequencies,
                                imag_part)
            self.plot_data_items['response_prediction'][1].setData(frequencies,
                                np.abs(np.real(spec[:,row_index,column_index])))
            imag_part = np.abs(np.imag(spec[:,row_index,column_index]))
            imag_part[imag_part < 1e-15] = 0
            self.plot_data_items['response_prediction'][3].setData(frequencies,
                                imag_part)
            
        elif message == 'time_data':
            # Display the data
            control_data,output_data = data
            for curve,this_data in zip(self.plot_data_items['control_responses'],control_data):
                x,y = curve.getData()
                y = np.concatenate((y[this_data.size:],this_data[-x.size:]),axis=0)
                curve.setData(x,y)
            # Display the data
            for curve,this_output in zip(self.plot_data_items['drive_outputs'],output_data):
                x,y = curve.getData()
                y = np.concatenate((y[this_output.size:],this_output[-x.size:]),axis=0)
                curve.setData(x,y)
        elif message == 'update_control_response':
            frequencies,cpsd = data
            self.plot_data_items['sum_asds_control'].setData(frequencies,np.einsum('ijj',cpsd).real)
            # Go through and remove any closed windows
            self.plot_windows = [window for window in self.plot_windows if window.isVisible()]
            for window in self.plot_windows:
                window.update_plot(cpsd)
        elif message == 'enable':
            widget = None
            for parent in [self.definition_widget,self.system_id_widget,self.prediction_widget,self.run_widget]:
                try:
                    widget = getattr(parent,data)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Enable Widget {:}: not found in UI'.format(data))
            widget.setEnabled(True)
        elif message == 'disable':
            widget = None
            for parent in [self.definition_widget,self.system_id_widget,self.prediction_widget,self.run_widget]:
                try:
                    widget = getattr(parent,data)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Disable Widget {:}: not found in UI'.format(data))
            widget.setEnabled(False)
        else:
            widget = None
            for parent in [self.definition_widget,self.system_id_widget,self.prediction_widget,self.run_widget]:
                try:
                    widget = getattr(parent,message)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Update Widget {:}: not found in UI'.format(message))
            if type(widget) is QtWidgets.QDoubleSpinBox:
                widget.setValue(data)
            elif type(widget) is QtWidgets.QSpinBox:
                widget.setValue(data)
            elif type(widget) is QtWidgets.QLineEdit:
                widget.setText(data)
            elif type(widget) is QtWidgets.QListWidget:
                widget.clear()
                widget.addItems(['{:.3f}'.format(d) for d in data])
                
    @staticmethod
    def create_environment_template(environment_name: str, workbook : openpyxl.workbook.workbook.Workbook):
        """Creates a template worksheet in an Excel workbook defining the
        environment.
        
        This function creates a template worksheet in an Excel workbook that
        when filled out could be read by the controller to re-create the 
        environment.
        
        This function is the "write" counterpart to the 
        ``set_parameters_from_template`` function in the ``RandomVibrationUI`` class,
        which reads the values from the template file to populate the user
        interface.

        Parameters
        ----------
        environment_name : str :
            The name of the environment that will specify the worksheet's name
        workbook : openpyxl.workbook.workbook.Workbook :
            A reference to an ``openpyxl`` workbook.

        """
        worksheet = workbook.create_sheet(environment_name)
        worksheet.cell(1,1,'Control Type')
        worksheet.cell(1,2,'Random')
        worksheet.cell(2,1,'Samples Per Frame:')
        worksheet.cell(3,1,'Test Level Ramp Time:')
        worksheet.cell(4,1,'Averaging Type:')
        worksheet.cell(5,1,'System ID Averages:')
        worksheet.cell(6,1,'Averaging Coefficient:')
        worksheet.cell(7,1,'FRF Technique:')
        worksheet.cell(8,1,'FRF Window:')
        worksheet.cell(9,1,'Overlap Percentage:')
        worksheet.cell(10,1,'System ID RMS Voltage:')
        worksheet.cell(11,1,'Update TF During Control:')
        worksheet.cell(12,1,'COLA Window:')
        worksheet.cell(13,1,'COLA Overlap:')
        worksheet.cell(14,1,'Window Exponent:')
        worksheet.cell(15,1,'Frames in CPSD:')
        worksheet.cell(16,1,'CPSD Window:')
        worksheet.cell(17,1,'Control Python Script:')
        worksheet.cell(18,1,'Control Python Function:')
        worksheet.cell(19,1,'Control Parameters:')
        worksheet.cell(20,1,'Specification File:')
        worksheet.cell(21,1,'Response Transformation Matrix:')
        worksheet.cell(22,1,'Output Transformation Matrix:')
        worksheet.cell(2,2,'# Number of Samples per Measurement Frame')
        worksheet.cell(3,2,'# Time taken to Ramp between test levels')
        worksheet.cell(4,2,'# Averaging Type')
        worksheet.cell(5,2,'# Number of Averages used when computing the FRF')
        worksheet.cell(6,2,'# Averaging Coefficient for Exponential Averaging')
        worksheet.cell(7,2,'# FRF Technique')
        worksheet.cell(8,2,'# Window used to compute FRF')
        worksheet.cell(9,2,'# Overlap percentage for CPSD and FRF calculations')
        worksheet.cell(10,2,'# RMS Value of Flat Voltage Spectrum used for System Identification')
        worksheet.cell(11,2,'# Continue updating transfer function while the controller is controlling (Y/N)')
        worksheet.cell(12,2,'# Window used for Constant Overlap and Add process')
        worksheet.cell(13,2,'# Overlap used in Constant Overlap and Add process')
        worksheet.cell(14,2,"# Exponent Applied to the COLA Window (use 0.5 unless you are sure you don't want to!)")
        worksheet.cell(15,2,'# Frames used to compute the CPSD matrix')
        worksheet.cell(16,2,'# Window used to compute the CPSD matrix')
        worksheet.cell(17,2,'# Path to the Python script containing the control law')
        worksheet.cell(18,2,'# Function name within the Python Script that will serve as the control law')
        worksheet.cell(19,2,'# Extra parameters used in the control law')
        worksheet.cell(20,2,'# Path to the file containing the Specification')
        worksheet.cell(21,2,'# Transformation matrix to apply to the response channels.  Type None if there is none.  Otherwise, make this a 2D array in the spreadsheet and move the Output Transformation Matrix line down so it will fit.  The number of columns should be the number of physical control channels.')
        worksheet.cell(22,2,'# Transformation matrix to apply to the outputs.  Type None if there is none.  Otherwise, make this a 2D array in the spreadsheet.  The number of columns should be the number of physical output channels in the environment.')
    
    def set_parameters_from_template(self, worksheet : openpyxl.worksheet.worksheet.Worksheet):
        """
        Collects parameters for the user interface from the Excel template file
        
        This function reads a filled out template worksheet to create an
        environment.  Cells on this worksheet contain parameters needed to
        specify the environment, so this function should read those cells and
        update the UI widgets with those parameters.
        
        This function is the "read" counterpart to the 
        ``create_environment_template`` function in the ``RandomVibrationUI`` class,
        which writes a template file that can be filled out by a user.
        

        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            An openpyxl worksheet that contains the environment template.
            Cells on this worksheet should contain the parameters needed for the
            user interface.

        """
        self.definition_widget.samples_per_frame_selector.setValue(int(worksheet.cell(2,2).value))
        self.definition_widget.ramp_time_spinbox.setValue(float(worksheet.cell(3,2).value))
        self.definition_widget.system_id_averaging_scheme_selector.setCurrentIndex(self.definition_widget.system_id_averaging_scheme_selector.findText(worksheet.cell(4,2).value))
        self.definition_widget.system_id_frames_to_average_selector.setValue(int(worksheet.cell(5,2).value))
        self.definition_widget.system_id_averaging_coefficient_selector.setValue(float(worksheet.cell(6,2).value))
        self.definition_widget.system_id_frf_technique_selector.setCurrentIndex(self.definition_widget.system_id_frf_technique_selector.findText(worksheet.cell(7,2).value))
        self.definition_widget.system_id_transfer_function_computation_window_selector.setCurrentIndex(self.definition_widget.system_id_transfer_function_computation_window_selector.findText(worksheet.cell(8,2).value))
        self.definition_widget.system_id_overlap_percentage_selector.setValue(float(worksheet.cell(9,2).value))
        self.definition_widget.system_id_initial_drive_voltage_selector.setValue(float(worksheet.cell(10,2).value))
        self.definition_widget.update_transfer_function_during_control_selector.setChecked(worksheet.cell(11,2).value.upper() == 'Y')
        self.definition_widget.cola_window_selector.setCurrentIndex(self.definition_widget.cola_window_selector.findText(worksheet.cell(12,2).value))
        self.definition_widget.cola_overlap_percentage_selector.setValue(float(worksheet.cell(13,2).value))
        self.definition_widget.cola_exponent_selector.setValue(float(worksheet.cell(14,2).value))
        self.definition_widget.cpsd_frames_selector.setValue(int(worksheet.cell(15,2).value))
        self.definition_widget.cpsd_computation_window_selector.setCurrentIndex(self.definition_widget.cpsd_computation_window_selector.findText(worksheet.cell(16,2).value))
        self.select_python_module(None,worksheet.cell(17,2).value)
        self.definition_widget.control_function_input.setCurrentIndex(self.definition_widget.control_function_input.findText(worksheet.cell(18,2).value))
        self.definition_widget.control_parameters_text_input.setText(str(worksheet.cell(19,2).value))
        # Now we need to find the transformation matrices' sizes
        response_channels = self.definition_widget.control_channels_display.value()
        output_channels = self.definition_widget.output_channels_display.value()
        output_transform_row = 22
        if isinstance(worksheet.cell(21,2).value,str) and worksheet.cell(21,2).value.lower() == 'none':
            self.response_transformation_matrix = None
        else:
            while True:
                if worksheet.cell(output_transform_row,1).value == 'Output Transformation Matrix:':
                    break
                output_transform_row += 1
            response_size = output_transform_row-21
            response_transformation = []
            for i in range(response_size):
                response_transformation.append([])
                for j in range(response_channels):
                    response_transformation[-1].append(float(worksheet.cell(21+i,2+j).value))
            self.response_transformation_matrix = np.array(response_transformation)
        if isinstance(worksheet.cell(output_transform_row,2).value,str) and worksheet.cell(output_transform_row,2).value.lower() == 'none':
            self.output_transformation_matrix = None
        else:
            output_transformation = []
            i = 0
            while True:
                if worksheet.cell(output_transform_row+i,2).value is None or (isinstance(worksheet.cell(output_transform_row+i,2).value,str) and worksheet.cell(output_transform_row+i,2).value.strip() == ''):
                    break
                output_transformation.append([])
                for j in range(output_channels):
                    output_transformation[-1].append(float(worksheet.cell(output_transform_row+i,2+j).value))
                i += 1
            self.output_transformation_matrix = np.array(output_transformation)
        self.define_transformation_matrices(None,dialog=False)
        self.select_spec_file(None,worksheet.cell(20,2).value)

class RandomVibrationEnvironment(AbstractEnvironment):
    """Environment defined by a random vibration control strategy"""
    def __init__(self,
                 environment_name : str,
                 queue_container : RandomEnvironmentQueues):
        """
        Random Vibration Environment Constructor that fills out the ``command_map``

        Parameters
        ----------
        environment_name : str
            Name of the environment.
        queue_container : RandomEnvironmentQueues
            Container of queues used by the Random Vibration Environment.

        """
        super().__init__(
                environment_name,
                queue_container.environment_command_queue,
                queue_container.gui_update_queue,
                queue_container.controller_communication_queue,
                queue_container.log_file_queue,
                queue_container.data_in_queue,
                queue_container.data_out_queue)
        self.queue_container = queue_container
        # Define command map
        self.command_map[RandomVibrationCommands.START_TRANSFER_FUNCTION] = self.start_transfer_function
        self.command_map[RandomVibrationCommands.STOP_TRANSFER_FUNCTION] = self.stop_environment
        self.command_map[RandomVibrationCommands.ADJUST_TEST_LEVEL] = self.adjust_test_level
        self.command_map[RandomVibrationCommands.STOP_ACQUISITION_AND_ANALYSIS] = self.stop_acquisition_and_analysis
        self.command_map[RandomVibrationCommands.SHOW_TEST_PREDICTION] = self.show_test_predictions
        self.command_map[RandomVibrationCommands.START_CONTROL] = self.start_control
        self.command_map[RandomVibrationCommands.STOP_CONTROL] = self.stop_environment
        self.command_map[RandomVibrationCommands.SHOW_FRF] = self.show_frf

    def initialize_data_acquisition_parameters(self,data_acquisition_parameters : DataAcquisitionParameters):
        """Initialize the data acquisition parameters in the environment.
        
        The environment will receive the global data acquisition parameters from
        the controller, and must set itself up accordingly.

        Parameters
        ----------
        data_acquisition_parameters : DataAcquisitionParameters :
            A container containing data acquisition parameters, including
            channels active in the environment as well as sampling parameters.
        """
        self.log('Initializing Data Acquisition Parameters')
        self.queue_container.cpsd_command_queue.put(self.environment_name,(CPSDMessages.INITIALIZE_DATA_ACQUISITION,data_acquisition_parameters))
        self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.INITIALIZE_DATA_ACQUISITION,data_acquisition_parameters))
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.INITIALIZE_DATA_ACQUISITION,data_acquisition_parameters))
        self.queue_container.data_analysis_command_queue.put(self.environment_name,(RandomDataAnalysisMessages.INITIALIZE_DATA_ACQUISITION,data_acquisition_parameters))
        self.queue_container.collector_command_queue.put(self.environment_name,(RandomDataCollectorMessages.INITIALIZE_DATA_ACQUISITION,data_acquisition_parameters))
        
    
    def initialize_environment_test_parameters(self,environment_parameters : RandomVibrationParameters):
        """
        Initialize the environment parameters specific to this environment
        
        The environment will recieve parameters defining itself from the
        user interface and must set itself up accordingly.

        Parameters
        ----------
        environment_parameters : RandomVibrationParameters
            A container containing the parameters defining the environment

        """
        self.log('Initializing Environment Parameters')
        self.queue_container.cpsd_command_queue.put(self.environment_name,(CPSDMessages.INITIALIZE_TEST_PARAMETERS,environment_parameters))
        self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.INITIALIZE_TEST_PARAMETERS,environment_parameters))
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.INITIALIZE_TEST_PARAMETERS,environment_parameters))
        self.queue_container.data_analysis_command_queue.put(self.environment_name,(RandomDataAnalysisMessages.INITIALIZE_TEST_PARAMETERS,environment_parameters))
        self.queue_container.collector_command_queue.put(self.environment_name,(RandomDataCollectorMessages.INITIALIZE_TEST_PARAMETERS,environment_parameters))
        
    def start_transfer_function(self,data):
        """Starts the transfer function running

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        """
        acquire,test_level = data
        self.log('Starting Transfer Function in {:} Mode at {:}x'.format('"Acquire"' if acquire else '"Preview"',test_level))
        self.queue_container.collector_command_queue.put(self.environment_name,(RandomDataCollectorMessages.SET_TEST_LEVEL,(100,test_level)))
        # I'm not sure why, but somehow the collector message just sent gets behind the message coming from the Signal Generator subtask...
        # I'm putting this sleep here to try to get it to not do that.
        time.sleep(0.01)
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.MUTE,None))
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.ADJUST_TEST_LEVEL,test_level))
        self.queue_container.collector_command_queue.put(self.environment_name,(RandomDataCollectorMessages.ACQUIRE,None))
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.RUN_TRANSFER_FUNCTION,None))
        self.queue_container.data_analysis_command_queue.put(self.environment_name,(RandomDataAnalysisMessages.RUN_TRANSFER_FUNCTION,acquire))
        self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.CLEAR_FRF,None))
        self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.RUN_FRF,None))
        self.queue_container.cpsd_command_queue.put(self.environment_name,(CPSDMessages.CLEAR_CPSD,None))
        self.queue_container.cpsd_command_queue.put(self.environment_name,(CPSDMessages.RUN_CPSD,None))
    
    def stop_environment(self, data):
        """Stop the environment gracefully
        
        This function defines the operations to shut down the environment
        gracefully so there is no hard stop that might damage test equipment
        or parts.

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        """
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.START_SHUTDOWN,None))
    
    def adjust_test_level(self,data):
        """Adjusts the test level of the environment

        Parameters
        ----------
        data :
            New test level
        """
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.ADJUST_TEST_LEVEL,data))
    
    def stop_acquisition_and_analysis(self,data):
        """Stops the computational tasks after the signal generation has ended

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        """
        self.queue_container.data_analysis_command_queue.put(self.environment_name,(RandomDataAnalysisMessages.STOP,None))
        self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.STOP_FRF,None))
        self.queue_container.cpsd_command_queue.put(self.environment_name,(CPSDMessages.STOP_CPSD,None))
        
    def show_test_predictions(self,data):
        """Sends a message to show the test prediction information

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        """
        self.queue_container.data_analysis_command_queue.put(self.environment_name,(RandomDataAnalysisMessages.SHOW_TEST_PREDICTION,data))
        
    def show_frf(self,data):
        """
        Sends a message to show the FRF data

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        """
        self.queue_container.data_analysis_command_queue.put(self.environment_name,(RandomDataAnalysisMessages.SHOW_FRF,data))
    
    def start_control(self,data):
        """Starts up the environment

        Parameters
        ----------
        data :
            Test level that the environment should start at

        """
        self.log('Starting Control')
        self.queue_container.collector_command_queue.put(self.environment_name,(RandomDataCollectorMessages.SET_TEST_LEVEL,(100,data)))
        # I'm not sure why, but somehow the collector message just sent gets behind the message coming from the Signal Generator subtask...
        # I'm putting this sleep here to try to get it to not do that.
        time.sleep(0.01)
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.MUTE,None))
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.ADJUST_TEST_LEVEL,data))
        self.queue_container.data_analysis_command_queue.put(self.environment_name,(RandomDataAnalysisMessages.RUN_CONTROL,None))
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(RandomSignalGenerationMessages.RUN_CONTROL,None))
        self.queue_container.collector_command_queue.put(self.environment_name,(RandomDataCollectorMessages.ACQUIRE,None))
        self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.RUN_FRF,None))
        self.queue_container.cpsd_command_queue.put(self.environment_name,(CPSDMessages.RUN_CPSD,None))
    
    def quit(self,data):
        """Stops the process because the program is ending.  This needs to 
        tell all the subprocesses to quit as well.

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``
        
        Returns
        -------
        True : bool
            Returns true to stop the run loop.
        """
        for queue in [self.queue_container.cpsd_command_queue,
                      self.queue_container.frf_command_queue,
                      self.queue_container.data_analysis_command_queue,
                      self.queue_container.signal_generation_command_queue,
                      self.queue_container.collector_command_queue]:
            queue.put(self.environment_name,(GlobalCommands.QUIT,None))
        # Return true to stop the task
        return True
        

def random_vibration_process(environment_name : str,
                 input_queue : VerboseMessageQueue,
                 gui_update_queue : Queue,
                 controller_communication_queue : VerboseMessageQueue,
                 log_file_queue : Queue,
                 data_in_queue : Queue,
                 data_out_queue : Queue):
    """Random vibration environment process function called by multiprocessing
    
    This function defines the Random Vibration Environment process that
    gets run by the multiprocessing module when it creates a new process.  It
    creates a RandomVibrationEnvironment object and runs it.

    Parameters
    ----------
    environment_name : str :
        Name of the environment
    input_queue : VerboseMessageQueue :
        Queue containing instructions for the environment
    gui_update_queue : Queue :
        Queue where GUI updates are put
    controller_communication_queue : Queue :
        Queue for global communications with the controller
    log_file_queue : Queue :
        Queue for writing log file messages
    data_in_queue : Queue :
        Queue from which data will be read by the environment
    data_out_queue : Queue :
        Queue to which data will be written that will be output by the hardware.

    """
    
    
    # Create vibration queues
    queue_container = RandomEnvironmentQueues(environment_name,
                                              input_queue,
                                              gui_update_queue,
                                              controller_communication_queue,
                                              data_in_queue,
                                              data_out_queue,
                                              log_file_queue)
    
    
    frf_proc = mp.Process(target=frf_computation_process,args=(environment_name,
                                                               queue_container.frf_command_queue,
                                                               queue_container.data_for_frf_queue,
                                                               queue_container.updated_frf_queue,
                                                               queue_container.gui_update_queue,
                                                               queue_container.log_file_queue))
    frf_proc.start()
    cpsd_proc = mp.Process(target=cpsd_computation_process,args=(environment_name,queue_container))
    cpsd_proc.start()
    analysis_proc = mp.Process(target=random_data_analysis_process,args=(environment_name,queue_container))
    analysis_proc.start()
    siggen_proc = mp.Process(target=random_signal_generation_process,args=(environment_name,queue_container))
    siggen_proc.start()
    collection_proc = mp.Process(target=random_data_collector_process,args=(environment_name,queue_container))
    collection_proc.start()
    
    process_class = RandomVibrationEnvironment(
            environment_name,
            queue_container)
    process_class.run()
    
    # Rejoin all the processes
    process_class.log('Joining Subprocesses')
    process_class.log('Joining FRF Computation')
    frf_proc.join()
    process_class.log('Joining CPSD Computation')
    cpsd_proc.join()
    process_class.log('Joining Data Analysis')
    analysis_proc.join()
    process_class.log('Joining Signal Generation')
    siggen_proc.join()
    process_class.log('Joining Data Collection')
    collection_proc.join()
    
    