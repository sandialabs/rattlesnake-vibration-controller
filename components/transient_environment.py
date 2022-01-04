# -*- coding: utf-8 -*-
"""
This file defines a Transient environment where a signal can be
loaded and played directly to the output devices.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from PyQt5 import QtWidgets,uic,QtCore
from .abstract_environment import AbstractEnvironment,AbstractMetadata,AbstractUI
from .utilities import DataAcquisitionParameters,VerboseMessageQueue,GlobalCommands,rms_time,db2scale,flush_queue,trac,align_signals,shift_signal,OverlapBuffer
from .ui_utilities import TransformationMatrixWindow,multiline_plotter,load_time_history,load_python_module,PlotTimeWindow
from .environments import (ControlTypes,environment_definition_ui_paths,
                           environment_run_ui_paths,
                           system_identification_ui_path,
                           environment_prediction_ui_paths,
                           )
import netCDF4 as nc4
from multiprocessing.queues import Queue
import numpy as np
import scipy.signal as sig
import multiprocessing as mp
import copy
import openpyxl
import inspect
from enum import Enum
import time
import os
import importlib

control_type = ControlTypes.TRANSIENT
test_level_threshold = 1.01
# max_responses_to_plot = 20
# max_samples_to_plot = 10000
maximum_name_length = 50
ACQUISITION_FRAMES_TO_DISPLAY = 1

from scipy.fft import rfft

class TransientCommands(Enum):
    """Commands accepted by the Transient Environment"""
    START_TRANSFER_FUNCTION = 0
    STOP_TRANSFER_FUNCTION = 1
    SHOW_TEST_PREDICTION = 2
    ADJUST_TEST_LEVEL = 3
    PERFORM_CONTROL_PREDICTION = 4
    SHOW_FRF = 5

class TransientQueues:
    """A set of queues used by the Transient environment"""
    def __init__(self,
                 environment_name: str,
                 environment_command_queue : VerboseMessageQueue,
                 gui_update_queue : mp.queues.Queue,
                 controller_communication_queue : VerboseMessageQueue,
                 data_in_queue : mp.queues.Queue,
                 data_out_queue : mp.queues.Queue,
                 log_file_queue : VerboseMessageQueue
                 ):
        """
        Creates a namespace to store all the queues used by the Transient Environment

        Parameters
        ----------
        environment_name : str
            Name of the environment
        environment_command_queue : VerboseMessageQueue
            Queue from which the environment will receive instructions.
        gui_update_queue : mp.queues.Queue
            Queue to which the environment will put GUI updates.
        controller_communication_queue : VerboseMessageQueue
            Queue to which the environment will put global contorller instructions.
        data_in_queue : mp.queues.Queue
            Queue from which the environment will receive data from acquisition.
        data_out_queue : mp.queues.Queue
            Queue to which the environment will write data for output.
        log_file_queue : VerboseMessageQueue
            Queue to which the environment will write log file messages.
        """
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.log_file_queue = log_file_queue
        # Queues for FRFs
        self.frf_command_queue = VerboseMessageQueue(log_file_queue,environment_name
                                                     + ' FRF Computation Command Queue')
        self.data_for_frf_queue = mp.Queue()
        self.updated_frf_queue = mp.Queue()
        # Queues for System ID
        self.signal_generation_command_queue = VerboseMessageQueue(log_file_queue,environment_name
                                                                   + ' Signal Generation Command Queue')

class TransientParameters(AbstractMetadata):
    """Storage container for parameters used by the Transient Environment
    """
    def __init__(self,sample_rate,control_signal,ramp_time,
                 averaging_type,system_id_averages,averaging_coefficient,
                 frf_technique,frf_window,overlap_percentage,frf_voltage,
                 response_transformation_matrix,output_transformation_matrix,
                 control_python_script,control_python_function,control_python_function_type,
                 control_python_function_parameters):
        """
        Container to hold signal processing parameters for the Time environment

        Parameters
        ----------
        sample_rate : int
            Number of samples per second that the controller runs.
        control_signal : np.ndarray
            Signal that will be generated by the controller, 2D array (n_control x
            n_samples)
        ramp_time : float
            Time used to start the signal from zero when the environment is started
            or stop the signal to zero when the environment is finished.
            Prevents "hard starts" or "hard stops" from damaging equipment.
        averaging_type : TYPE
            Averaging type used to compute FRFs
        system_id_averages : int
            Number of averages to compute the FRFs
        averaging_coefficient : float
            Exponential averaging coefficient
        frf_technique : TYPE
            Technique to compute FRFs
        frf_window : TYPE
            Window function used to compute FRFs
        overlap_percentage : float
            Overlap percentage in FRF computation. 0-100
        frf_voltage : float
            RMS voltage of the signal used for system identification.
        response_transformation_matrix : np.ndarray
            Response transformation.  None if Identity is used.
        output_transformation_matrix : TYPE
            Output transformation.  None if Identity is used.
        control_python_script : str
            Path to the Python file that contains the control script.
        control_python_function : str
            Function name of the control strategy within the script.
        control_python_function_type : TYPE
            Type of the Python data used for control Function/Generator/Etc.
        control_python_function_parameters : str
            Extra parameters passed to the Python function.
        """
        self.sample_rate = sample_rate
        self.control_signal = control_signal
        self.ramp_time = ramp_time
        self.averaging_type = averaging_type
        self.system_id_averages = system_id_averages
        self.averaging_coefficient = averaging_coefficient
        self.frf_technique = frf_technique
        self.frf_window = frf_window
        self.frf_voltage = frf_voltage
        self.overlap = overlap_percentage/100
        self.response_transformation_matrix = response_transformation_matrix
        self.output_transformation_matrix = output_transformation_matrix
        self.control_python_script = control_python_script
        self.control_python_function = control_python_function
        self.control_python_function_type = control_python_function_type
        self.control_python_function_parameters = control_python_function_parameters
        
    @property
    def signal_samples(self):
        """The number of samples in the signal"""
        return self.control_signal.shape[-1]
    
    @property
    def control_channels(self):
        """The number of output channels in the signal"""
        return self.control_signal.shape[0]
    
    @property
    def signal_time(self):
        """The length of the signal in seconds"""
        return self.signal_samples/self.sample_rate
    
    @property
    def fft_lines(self):
        """Property returning the frequency lines given the sampling parameters"""
        return self.signal_samples//2 + 1
    
    @property
    def frequency_spacing(self):
        """Property returning frequency line spacing given the sampling parameters"""
        return self.sample_rate/self.signal_samples
    
    @property
    def samples_per_frame(self):
        return self.control_signal.shape[-1]
    
    @property
    def ramp_samples(self):
        """The number of samples required to ramp up the signal when started"""
        return int(self.ramp_time*self.sample_rate)

    def store_to_netcdf(self,netcdf_group_handle : nc4._netCDF4.Group):
        """
        Stores parameters to a netCDF group so they can be recovered.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            Reference to the netCDF4 group in which the environment data should
            be stored.

        """
        netcdf_group_handle.ramp_time = self.ramp_time
        netcdf_group_handle.averaging_type = self.averaging_type
        netcdf_group_handle.system_id_averages = self.system_id_averages
        netcdf_group_handle.averaging_coefficient = self.averaging_coefficient
        netcdf_group_handle.frf_technique = self.frf_technique
        netcdf_group_handle.frf_window = self.frf_window
        netcdf_group_handle.overlap = self.overlap
        netcdf_group_handle.frf_voltage = self.frf_voltage
        netcdf_group_handle.control_python_script = self.control_python_script
        netcdf_group_handle.control_python_function = self.control_python_function
        netcdf_group_handle.control_python_function_type = self.control_python_function_type
        netcdf_group_handle.control_python_function_parameters = self.control_python_function_parameters
        # Save the output signal
        netcdf_group_handle.createDimension('control_channels',self.control_channels)
        netcdf_group_handle.createDimension('signal_samples',self.signal_samples)
        var = netcdf_group_handle.createVariable('control_signal','f8',('control_channels','signal_samples'))
        var[...] = self.control_signal
        if not self.response_transformation_matrix is None:
            netcdf_group_handle.createDimension('response_transformation_rows',self.response_transformation_matrix.shape[0])
            netcdf_group_handle.createDimension('response_transformation_cols',self.response_transformation_matrix.shape[1])
            var = netcdf_group_handle.createVariable('response_transformation_matrix','f8',('response_transformation_rows','response_transformation_cols'))
            var[...] = self.response_transformation_matrix
        if not self.output_transformation_matrix is None:
            netcdf_group_handle.createDimension('output_transformation_rows',self.output_transformation_matrix.shape[0])
            netcdf_group_handle.createDimension('output_transformation_cols',self.output_transformation_matrix.shape[1])
            var = netcdf_group_handle.createVariable('output_transformation_matrix','f8',('output_transformation_rows','output_transformation_cols'))
            var[...] = self.output_transformation_matrix
    
    @classmethod
    def from_ui(cls,ui):
        """Creates a TransientParameters object from the user interface

        Parameters
        ----------
        ui : TransientUI
            A Transient User Interface

        Returns
        -------
        test_parameters : TransientParameters
            Parameters corresponding to the data in the user interface.

        """
        if ui.python_control_module is None:
            control_module = None
            control_function = None
            control_function_type = None
            control_function_parameters = None
        else:
            control_module = ui.definition_widget.control_script_file_path_input.text()
            control_function = ui.definition_widget.control_function_input.itemText(ui.definition_widget.control_function_input.currentIndex())
            control_function_type = ui.definition_widget.control_function_generator_selector.currentIndex()
            control_function_parameters = ui.definition_widget.control_parameters_text_input.toPlainText()
        return cls(sample_rate=ui.definition_widget.sample_rate_display.value(),
                    control_signal=ui.signal,
                    ramp_time=ui.definition_widget.ramp_selector.value(),
                    averaging_type=ui.definition_widget.system_id_averaging_scheme_selector.itemText(ui.definition_widget.system_id_averaging_scheme_selector.currentIndex()),
                    system_id_averages=ui.definition_widget.system_id_frames_to_average_selector.value(),
                    averaging_coefficient=ui.definition_widget.system_id_averaging_coefficient_selector.value(),
                    frf_technique=ui.definition_widget.system_id_frf_technique_selector.itemText(ui.definition_widget.system_id_frf_technique_selector.currentIndex()),
                    frf_window=ui.definition_widget.system_id_transfer_function_computation_window_selector.itemText(ui.definition_widget.system_id_transfer_function_computation_window_selector.currentIndex()),
                    overlap_percentage=ui.definition_widget.system_id_overlap_percentage_selector.value(),
                    frf_voltage=ui.definition_widget.system_id_initial_drive_voltage_selector.value(),
                    response_transformation_matrix=ui.response_transformation_matrix,
                    output_transformation_matrix=ui.output_transformation_matrix,
                    control_python_script=control_module,
                    control_python_function=control_function,
                    control_python_function_type=control_function_type,
                    control_python_function_parameters=control_function_parameters)

from .frf_computation import frf_computation_process,FRFMessages
from .system_id_signal_generation import system_id_signal_generation_process,SysIdSignalGenerationMessages

class TransientUI(AbstractUI):
    """Class defining the user interface for a Transient environment.
    
    This class will contain four main UIs, the environment definition, system identification,
    test predictions, and run.
    The widgets corresponding to these interfaces are stored in TabWidgets in
    the main UI.
    
    This class defines all the call backs and user interface operations required
    for the Transient environment."""
    def __init__(self,
                 environment_name : str,
                 definition_tabwidget : QtWidgets.QTabWidget,
                 system_id_tabwidget : QtWidgets.QTabWidget,
                 test_predictions_tabwidget : QtWidgets.QTabWidget,
                 run_tabwidget : QtWidgets.QTabWidget,
                 environment_command_queue : VerboseMessageQueue,
                 controller_communication_queue : VerboseMessageQueue,
                 log_file_queue : Queue):
        """
        Constructs a Transient User Interface
        
        Given the tab widgets from the main interface as well as communication
        queues, this class assembles the user interface components specific to
        the Transient Environment

        Parameters
        ----------
        definition_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Control
            Definition main tab
        system_id_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the System
            Identification main tab.  The Transient Environment has no system 
            identification step, so this is not used.
        test_predictions_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Test Predictions
            main tab.    The Transient Environment has no system identification
            step, so this is not used.
        run_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Run
            main tab.
        environment_command_queue : VerboseMessageQueue
            Queue for sending commands to the Random Vibration Environment
        controller_communication_queue : VerboseMessageQueue
            Queue for sending global commands to the controller
        log_file_queue : Queue
            Queue where log file messages can be written.
    
        """
        super().__init__(environment_name,
             environment_command_queue,controller_communication_queue,log_file_queue)
        # Add the page to the control definition tabwidget
        self.definition_widget = QtWidgets.QWidget()
        uic.loadUi(environment_definition_ui_paths[control_type],self.definition_widget)
        definition_tabwidget.addTab(self.definition_widget,self.environment_name)
        # Add the page to the control prediction tabwidget
        self.prediction_widget = QtWidgets.QWidget()
        uic.loadUi(environment_prediction_ui_paths[control_type],self.prediction_widget)
        test_predictions_tabwidget.addTab(self.prediction_widget,self.environment_name)
        # Add the page to the system id tabwidget
        self.system_id_widget = QtWidgets.QWidget()
        uic.loadUi(system_identification_ui_path,self.system_id_widget)
        system_id_tabwidget.addTab(self.system_id_widget,self.environment_name)
        # Add the page to the run tabwidget
        self.run_widget = QtWidgets.QWidget()
        uic.loadUi(environment_run_ui_paths[control_type],self.run_widget)
        run_tabwidget.addTab(self.run_widget,self.environment_name)
        
        # Set up some persistent data
        self.data_acquisition_parameters = None
        self.environment_parameters = None
        self.signal = None
        self.show_signal_checkboxes = None
        self.plot_data_items = {}
        self.plot_windows = []
        self.response_transformation_matrix = None
        self.output_transformation_matrix = None
        self.python_control_module = None
        self.physical_control_names = None
        self.physical_output_names = None
        self.control_selector_widgets = None
        self.output_selector_widgets = None
                
        self.complete_ui()
        self.connect_callbacks()
        
        # Complete the profile commands
        self.command_map['Set Test Level'] = self.change_test_level_from_profile
        self.command_map['Set Repeat'] = self.set_repeat_from_profile
        self.command_map['Set No Repeat'] = self.set_norepeat_from_profile
        
        
    def collect_environment_definition_parameters(self) -> TransientParameters:
        """Collect the parameters from the user interface defining the environment

        Returns
        -------
        TransientParameters
            A metadata or parameters object containing the parameters defining
            the corresponding environment.
        """
        return TransientParameters.from_ui(self)
    
    def complete_ui(self):
        """Helper Function to continue setting up the user interface"""
        # Initialize enabled or disabled widgets
        self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        
        # Set common look and feel for plots
        plotWidgets = [self.definition_widget.signal_display_plot,
                       self.run_widget.output_signal_plot,
                       self.run_widget.response_signal_plot,
                       self.prediction_widget.excitation_display_plot,
                       self.prediction_widget.response_display_plot,
                       self.system_id_widget.response_timehistory_plot,
                       self.system_id_widget.drive_timehistory_plot,
                       self.system_id_widget.transfer_function_phase_plot,
                       self.system_id_widget.transfer_function_amplitude_plot]
        for plotWidget in plotWidgets:
            plot_item = plotWidget.getPlotItem()
            plot_item.showGrid(True,True,0.25)
            plot_item.enableAutoRange()
            plot_item.getViewBox().enableAutoRange(enable=True)
        logscalePlotWidgets = [self.system_id_widget.transfer_function_amplitude_plot,]
        for plotWidget in logscalePlotWidgets:
            plot_item = plotWidget.getPlotItem()
            plot_item.setLogMode(False,True)
            
    def connect_callbacks(self):
        """Helper function to connect callbacks to functions in the class"""
        # Definition
        self.definition_widget.load_signal_button.clicked.connect(self.load_signal)
        self.definition_widget.transformation_matrices_button.clicked.connect(self.define_transformation_matrices)
        self.definition_widget.system_id_averaging_scheme_selector.currentIndexChanged.connect(self.disable_exponential_coefficient)
        self.definition_widget.control_script_load_file_button.clicked.connect(self.select_python_module)
        self.definition_widget.control_function_input.currentIndexChanged.connect(self.update_generator_selector)
        # System ID
        self.system_id_widget.preview_transfer_function_button.clicked.connect(self.preview_transfer_function)
        self.system_id_widget.acquire_transfer_function_button.clicked.connect(self.acquire_transfer_function)
        self.system_id_widget.stop_transfer_function_button.clicked.connect(self.stop_transfer_function)
        self.system_id_widget.select_transfer_function_stream_file_button.clicked.connect(self.select_transfer_function_stream_file)
        self.system_id_widget.voltage_scale_factor_selector.valueChanged.connect(self.change_transfer_function_test_level)
        self.system_id_widget.transfer_function_response_selector.currentIndexChanged.connect(self.show_frf)
        self.system_id_widget.transfer_function_reference_selector.currentIndexChanged.connect(self.show_frf)
        # Test Predictions
        self.prediction_widget.excitation_selector.currentIndexChanged.connect(self.show_test_predictions)
        self.prediction_widget.response_selector.currentIndexChanged.connect(self.show_test_predictions)
        self.prediction_widget.maximum_voltage_button.clicked.connect(self.show_max_voltage_prediction)
        self.prediction_widget.minimum_voltage_button.clicked.connect(self.show_min_voltage_prediction)
        self.prediction_widget.maximum_error_button.clicked.connect(self.show_max_error_prediction)
        self.prediction_widget.minimum_error_button.clicked.connect(self.show_min_error_prediction)
        # Run Test 
        self.run_widget.start_test_button.clicked.connect(self.start_control)
        self.run_widget.stop_test_button.clicked.connect(self.stop_control)
        self.run_widget.create_window_button.clicked.connect(self.create_window)
        self.run_widget.show_all_channels_button.clicked.connect(self.show_all_channels)
        self.run_widget.tile_windows_button.clicked.connect(self.tile_windows)
        self.run_widget.close_windows_button.clicked.connect(self.close_windows)
    
    def initialize_data_acquisition(self,data_acquisition_parameters : DataAcquisitionParameters):
        """Update the user interface with data acquisition parameters
        
        This function is called when the Data Acquisition parameters are
        initialized.  This function should set up the environment user interface
        accordingly.

        Parameters
        ----------
        data_acquisition_parameters : DataAcquisitionParameters :
            Container containing the data acquisition parameters, including
            channel table and sampling information.

        """
        self.log('Initializing Data Acquisition')
        self.signal = None
        # Get channel information
        channels = data_acquisition_parameters.channel_list
        num_control = len([channel for channel in channels if channel.control])
        num_output = len([channel for channel in channels if not channel.feedback_device is None])
        # Initilize transformation matrices to none (identity)
        self.definition_widget.response_transformation_matrix = None#np.eye(num_control)
        self.definition_widget.output_transformation_matrix = None#np.eye(num_output)
        self.definition_widget.input_channels_display.setValue(len(channels))
        self.definition_widget.control_channels_display.setValue(num_control)
        self.definition_widget.output_channels_display.setValue(num_output)
        self.definition_widget.transform_channels_display.setValue(num_control)
        self.definition_widget.transform_outputs_display.setValue(num_output)
        self.physical_control_names = ['{:} {:} {:}'.format('' if channel.channel_type is None else channel.channel_type,channel.node_number,channel.node_direction)[:maximum_name_length]
            for channel in channels if channel.control]
        self.physical_output_names = ['{:} {:} {:}'.format('' if channel.channel_type is None else channel.channel_type,channel.node_number,channel.node_direction)[:maximum_name_length]
            for channel in channels if channel.feedback_device]
        self.control_selector_widgets = [
                self.system_id_widget.transfer_function_response_selector,
                self.prediction_widget.response_selector,
                self.run_widget.control_channel_selector]
        self.output_selector_widgets = [
                self.system_id_widget.transfer_function_reference_selector,
                self.prediction_widget.excitation_selector,
                ]
        for widget in self.control_selector_widgets:
            widget.blockSignals(True)
            widget.clear()
        for widget in self.output_selector_widgets:
            widget.blockSignals(True)
            widget.clear()
        for i,control_name in enumerate(self.physical_control_names):
            for widget in self.control_selector_widgets:
                widget.addItem('{:}: {:}'.format(i+1,control_name))
        for i,drive_name in enumerate(self.physical_output_names):
            for widget in self.output_selector_widgets:
                widget.addItem('{:}: {:}'.format(i+1,drive_name))
        for widget in self.control_selector_widgets:
            widget.blockSignals(False)
        for widget in self.output_selector_widgets:
            widget.blockSignals(False)
        # Add rows to the signal table
        self.definition_widget.signal_information_table.setRowCount(num_control)
        self.show_signal_checkboxes = []
        for i,name in enumerate(self.physical_control_names):
            item = QtWidgets.QTableWidgetItem()
            item.setText(name)
            item.setFlags(item.flags() ^ QtCore.Qt.ItemIsEditable)
            self.definition_widget.signal_information_table.setItem(i,1,item)
            checkbox = QtWidgets.QCheckBox()
            checkbox.setChecked(True)
            checkbox.stateChanged.connect(self.show_signal)
            self.show_signal_checkboxes.append(checkbox)
            self.definition_widget.signal_information_table.setCellWidget(i,0,checkbox)
            item = QtWidgets.QTableWidgetItem()
            item.setText('0.0')
            item.setFlags(item.flags() ^ QtCore.Qt.ItemIsEditable)
            self.definition_widget.signal_information_table.setItem(i,2,item)
            item = QtWidgets.QTableWidgetItem()
            item.setText('0.0')
            item.setFlags(item.flags() ^ QtCore.Qt.ItemIsEditable)
            self.definition_widget.signal_information_table.setItem(i,3,item)
        # Fill in the info at the bottom
        self.definition_widget.sample_rate_display.setValue(data_acquisition_parameters.sample_rate)
            
        # Clear the signal plots
        self.definition_widget.signal_display_plot.getPlotItem().clear()
        self.run_widget.output_signal_plot.getPlotItem().clear()
        self.run_widget.response_signal_plot.getPlotItem().clear()
        
        # Set initial lines
        self.plot_data_items['control_signal_definition'] = multiline_plotter(
                np.arange(2),
                np.zeros((num_control,2)),
                widget=self.definition_widget.signal_display_plot,
                other_pen_options={'width':1},
                names = self.physical_control_names)
        self.plot_data_items['output_signal_measurement'] = multiline_plotter(
                np.arange(2),
                np.zeros((num_output,2)),
                widget=self.run_widget.output_signal_plot,
                other_pen_options={'width':1},
                names = self.physical_output_names)
        self.plot_data_items['signal_range'] = self.run_widget.response_signal_plot.getPlotItem().plot(np.zeros(5),np.zeros(5),pen = {'color': "k", 'width': 1},name='Signal Lower Bound')
        self.plot_data_items['control_signal_measurement'] = multiline_plotter(
                np.arange(2),
                np.zeros((num_control,2)),
                widget=self.run_widget.response_signal_plot,
                other_pen_options={'width':1},
                names = self.physical_control_names)
        
        self.data_acquisition_parameters = data_acquisition_parameters
    
    def load_signal(self,clicked,filename=None):
        """Loads a time signal using a dialog or the specified filename

        Parameters
        ----------
        clicked :
            The clicked event that triggered the callback.
        filename :
            File name defining the specification for bypassing the callback when
            loading from a file (Default value = None).

        """
        if filename is None:
            filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self.definition_widget,'Select Signal File',filter='Numpy or Mat (*.npy *.npz *.mat)')
            if filename == '':
                return
        self.definition_widget.signal_file_name_display.setText(filename)
        self.signal = load_time_history(filename,self.definition_widget.sample_rate_display.value())
        self.definition_widget.signal_samples_display.setValue(self.signal.shape[-1])
        self.definition_widget.signal_time_display.setValue(self.signal.shape[-1]/self.definition_widget.sample_rate_display.value())
        maxs = np.max(np.abs(self.signal),axis=-1)
        rmss = rms_time(self.signal,axis=-1)
        for i,(mx,rms) in enumerate(zip(maxs,rmss)):
            self.definition_widget.signal_information_table.item(i,2).setText('{:0.2f}'.format(mx))
            self.definition_widget.signal_information_table.item(i,3).setText('{:0.2f}'.format(rms))
        self.show_signal()
    
    def show_signal(self):
        """Shows the signal on the user interface"""
        for curve,signal,check_box in zip(self.plot_data_items['control_signal_definition'],self.signal,self.show_signal_checkboxes):
            if check_box.isChecked():
                x = np.arange(signal.shape[-1])/self.definition_widget.sample_rate_display.value()
                curve.setData(x,signal)
            else:
                curve.setData((0,0),(0,0))
    
    def define_transformation_matrices(self,clicked,dialog = True):
        """Defines the transformation matrices using the dialog box"""
        if dialog:
            (response_transformation,output_transformation,result
             ) = TransformationMatrixWindow.define_transformation_matrices(
                    self.response_transformation_matrix,
                    self.definition_widget.control_channels_display.value(),
                    self.output_transformation_matrix,
                    self.definition_widget.output_channels_display.value(),
                    self.definition_widget)
        else:
            response_transformation = self.response_transformation_matrix
            output_transformation = self.output_transformation_matrix
            result = True
        if result:
            # Update the control names
            for widget in self.control_selector_widgets:
                widget.blockSignals(True)
                widget.clear()
            if response_transformation is None:
                for i,control_name in enumerate(self.physical_control_names):
                    for widget in self.control_selector_widgets:
                        widget.addItem('{:}: {:}'.format(i+1,control_name))
                self.definition_widget.transform_channels_display.setValue(len(self.physical_control_names))
            else:
                for i in range(response_transformation.shape[0]):
                    for widget in self.control_selector_widgets:
                        widget.addItem('{:}: {:}'.format(i+1,'Virtual Response'))
                self.definition_widget.transform_channels_display.setValue(response_transformation.shape[0])
            for widget in self.control_selector_widgets:
                widget.blockSignals(False)
            # Update the output names
            for widget in self.output_selector_widgets:
                widget.blockSignals(True)
                widget.clear()
            if output_transformation is None:
                for i,drive_name in enumerate(self.physical_output_names):
                    for widget in self.output_selector_widgets:
                        widget.addItem('{:}: {:}'.format(i+1,drive_name))
                self.definition_widget.transform_outputs_display.setValue(len(self.physical_output_names))
            else:
                for i in range(output_transformation.shape[0]):
                    for widget in self.output_selector_widgets:
                        widget.addItem('{:}: {:}'.format(i+1,'Virtual Drive'))
                self.definition_widget.transform_outputs_display.setValue(output_transformation.shape[0])
            for widget in self.output_selector_widgets:
                widget.blockSignals(False)
                
            self.response_transformation_matrix = response_transformation
            self.output_transformation_matrix = output_transformation
    
    def disable_exponential_coefficient(self):
        """Disables the exponential averaging coefficient when averaging is linear"""
        if self.definition_widget.system_id_averaging_scheme_selector.currentIndex() == 0:
            self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        else:
            self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(True)
    
    def select_python_module(self,clicked,filename=None):
        """Loads a Python module using a dialog or the specified filename

        Parameters
        ----------
        clicked :
            The clicked event that triggered the callback.
        filename :
            File name defining the Python module for bypassing the callback when
            loading from a file (Default value = None).

        """
        if filename is None:
            filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self.definition_widget,'Select Python Module',filter='Python Modules (*.py)')
            if filename == '':
                return
        self.python_control_module = load_python_module(filename)
        functions = [function for function in inspect.getmembers(self.python_control_module)
                     if (inspect.isfunction(function[1]) and len(inspect.signature(function[1]).parameters)>=6)
                     or inspect.isgeneratorfunction(function[1])
                     or (inspect.isclass(function[1]) and all([method in function[1].__dict__ for method in ['system_id_update','control']]))]
        self.log('Loaded module {:} with functions {:}'.format(self.python_control_module.__name__,[function[0] for function in functions]))
        self.definition_widget.control_function_input.clear()
        self.definition_widget.control_script_file_path_input.setText(filename)
        for function in functions:
            self.definition_widget.control_function_input.addItem(function[0])
            
    def update_generator_selector(self):
        """Updates the function/generator selector based on the function selected"""
        if self.python_control_module is None:
            return
        function = getattr(self.python_control_module,self.definition_widget.control_function_input.itemText(self.definition_widget.control_function_input.currentIndex()))
        if inspect.isgeneratorfunction(function):
            self.definition_widget.control_function_generator_selector.setCurrentIndex(1)
        elif inspect.isclass(function):
            self.definition_widget.control_function_generator_selector.setCurrentIndex(2)
        else:
            self.definition_widget.control_function_generator_selector.setCurrentIndex(0)
    
    
    def initialize_environment(self) -> AbstractMetadata:
        """Update the user interface with environment parameters
        
        This function is called when the Environment parameters are initialized.
        This function should set up the user interface accordingly.  It must
        return the parameters class of the environment that inherits from
        AbstractMetadata.

        Returns
        -------
        environment_parameters : TransientParameters
            A TransientParameters object that contains the parameters
            defining the environment.
        """
        self.log('Initializing Environment Parameters')
        data = self.collect_environment_definition_parameters()
        # Make sure everything is defined 
        if data.control_signal is None:
            raise ValueError('Control Signal is not defined for {:}!'.format(self.environment_name))
        if data.control_python_script is None:
            raise ValueError('Control function has not been loaded for {:}'.format(self.environment_name))
        # Set the value in the system_id_num_averages_display
        self.system_id_widget.system_id_num_averages_display.setValue(data.system_id_averages)
        self.system_id_widget.voltage_display.setValue(self.definition_widget.system_id_initial_drive_voltage_selector.value())
        # Initialize the rest of the plots
        self.system_id_widget.response_timehistory_plot.getPlotItem().clear()
        self.system_id_widget.drive_timehistory_plot.getPlotItem().clear()
        self.system_id_widget.transfer_function_phase_plot.getPlotItem().clear()
        self.system_id_widget.transfer_function_amplitude_plot.getPlotItem().clear()
        self.prediction_widget.excitation_display_plot.getPlotItem().clear()
        self.prediction_widget.response_display_plot.getPlotItem().clear()
        # Initialize the correct sizes of the arrays
        self.plot_data_items['control_responses'] = multiline_plotter(
                np.arange(data.control_signal.shape[-1]*ACQUISITION_FRAMES_TO_DISPLAY)/data.sample_rate,
                np.zeros((self.definition_widget.transform_channels_display.value(),data.control_signal.shape[-1]*ACQUISITION_FRAMES_TO_DISPLAY)),
                widget=self.system_id_widget.response_timehistory_plot,
                other_pen_options={'width':1},
                names = [self.system_id_widget.transfer_function_response_selector.itemText(i) for i in range(self.system_id_widget.transfer_function_response_selector.count())])
        self.plot_data_items['drive_outputs'] = multiline_plotter(
                np.arange(data.control_signal.shape[-1]*ACQUISITION_FRAMES_TO_DISPLAY)/data.sample_rate,
                np.zeros((self.definition_widget.transform_outputs_display.value(),data.control_signal.shape[-1]*ACQUISITION_FRAMES_TO_DISPLAY)),
                widget=self.system_id_widget.drive_timehistory_plot,
                other_pen_options={'width':1},
                names = [self.system_id_widget.transfer_function_reference_selector.itemText(i) for i in range(self.system_id_widget.transfer_function_reference_selector.count())])
        self.plot_data_items['transfer_function_phase'] = self.system_id_widget.transfer_function_phase_plot.getPlotItem().plot(
            np.arange(data.fft_lines)*data.frequency_spacing,np.zeros(data.fft_lines),pen = {'color': "r", 'width': 1},name='Phase')
        self.plot_data_items['transfer_function_amplitude'] = self.system_id_widget.transfer_function_amplitude_plot.getPlotItem().plot(
            np.arange(data.fft_lines)*data.frequency_spacing,np.zeros(data.fft_lines),pen = {'color': "r", 'width': 1},name='Amplitude')
        for plot_items in [self.plot_data_items['output_signal_measurement'],self.plot_data_items['control_signal_measurement']]:
            for curve in plot_items:
                curve.setData(
                    np.arange(data.control_signal.shape[-1]*2)/self.data_acquisition_parameters.sample_rate,
                              np.zeros((data.control_signal.shape[-1]*2))
                    )
        self.plot_data_items['response_prediction'] = multiline_plotter(
                np.arange(data.fft_lines)*data.frequency_spacing,
                np.zeros((2,data.fft_lines)),
                widget = self.prediction_widget.response_display_plot,
                other_pen_options={'width':1},
                names = ['Prediction','Spec']
                )
        self.plot_data_items['excitation_prediction'] = multiline_plotter(
                np.arange(data.fft_lines)*data.frequency_spacing,
                np.zeros((1,data.fft_lines)),
                widget = self.prediction_widget.excitation_display_plot,
                other_pen_options={'width':1},
                names = ['Prediction']
                )
        self.environment_parameters = data
        return data
    
    ### System Identification Callbacks
    def preview_transfer_function(self):
        """Starts the transfer function in preview mode"""
        self.log('Starting Transfer Function Preview')
        self.system_id_widget.preview_transfer_function_button.setEnabled(False)
        self.system_id_widget.acquire_transfer_function_button.setEnabled(False)
        self.system_id_widget.stop_transfer_function_button.setEnabled(True)
        self.environment_command_queue.put(self.log_name,(TransientCommands.START_TRANSFER_FUNCTION,
                                                          (False,self.environment_parameters.frf_voltage,
                                                           db2scale(self.system_id_widget.voltage_scale_factor_selector.value()))))
    
    def acquire_transfer_function(self):
        """Starts the transfer function in acquire mode"""
        self.log('Starting Transfer Function Acquire')
        self.system_id_widget.preview_transfer_function_button.setEnabled(False)
        self.system_id_widget.acquire_transfer_function_button.setEnabled(False)
        self.system_id_widget.stop_transfer_function_button.setEnabled(True)
        if self.system_id_widget.stream_transfer_function_data_checkbox.isChecked():
            self.controller_communication_queue.put(self.log_name,(GlobalCommands.INITIALIZE_STREAMING,self.system_id_widget.transfer_function_stream_file_display.text()))
            self.controller_communication_queue.put(self.log_name,(GlobalCommands.START_STREAMING,None))
        self.environment_command_queue.put(self.log_name,(TransientCommands.START_TRANSFER_FUNCTION,
                                                          (True,self.environment_parameters.frf_voltage,
                                                           db2scale(self.system_id_widget.voltage_scale_factor_selector.value()))))
        
    def stop_transfer_function(self):
        """Stops the transfer function acquisition"""
        self.log('Stopping Transfer Function')
        self.environment_command_queue.put(self.log_name,(TransientCommands.STOP_TRANSFER_FUNCTION,None))
    
    def select_transfer_function_stream_file(self):
        """Select a file to save transfer function data to"""
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self.system_id_widget,'Select NetCDF File to Save Transfer Function Data',filter='NetCDF File (*.nc4)')
        if filename == '':
            return
        self.system_id_widget.transfer_function_stream_file_display.setText(filename)
        self.system_id_widget.stream_transfer_function_data_checkbox.setChecked(True)
    
    def change_transfer_function_test_level(self):
        """Updates the test level for the transfer function"""
        self.environment_command_queue.put(self.log_name,(TransientCommands.ADJUST_TEST_LEVEL,db2scale(self.system_id_widget.voltage_scale_factor_selector.value())))
    
    def show_frf(self):
        """Tells the environment to show the FRF"""
        self.environment_command_queue.put(self.log_name,(TransientCommands.SHOW_FRF,None))
    
    ### Prediction Callbacks
    def show_test_predictions(self):
        """Shows test predictions"""
        self.environment_command_queue.put(self.log_name,(TransientCommands.SHOW_TEST_PREDICTION,False))
        
    def show_max_voltage_prediction(self):
        widget = self.prediction_widget.excitation_voltage_list
        index = np.argmax([float(widget.item(v).text())
                           for v in range(widget.count())])
        self.prediction_widget.excitation_selector.setCurrentIndex(index)
    def show_min_voltage_prediction(self):
        widget = self.prediction_widget.excitation_voltage_list
        index = np.argmin([float(widget.item(v).text())
                           for v in range(widget.count())])
        self.prediction_widget.excitation_selector.setCurrentIndex(index)
    def show_max_error_prediction(self):
        widget = self.prediction_widget.response_error_list
        index = np.argmax([float(widget.item(v).text())
                           for v in range(widget.count())])
        self.prediction_widget.response_selector.setCurrentIndex(index)
    def show_min_error_prediction(self):
        widget = self.prediction_widget.response_error_list
        index = np.argmin([float(widget.item(v).text())
                           for v in range(widget.count())])
        self.prediction_widget.response_selector.setCurrentIndex(index)
        
    ### Control Callbacks
    
    def start_control(self):
        """Starts running the environment"""
        self.run_widget.stop_test_button.setEnabled(True)
        self.run_widget.start_test_button.setEnabled(False)
        self.run_widget.test_level_selector.setEnabled(False)
        self.run_widget.repeat_signal_checkbox.setEnabled(False)
        self.controller_communication_queue.put(self.log_name,(GlobalCommands.START_ENVIRONMENT,self.environment_name))
        self.environment_command_queue.put(self.log_name,(GlobalCommands.START_ENVIRONMENT,
                                                          (db2scale(self.run_widget.test_level_selector.value()),
                                                           self.run_widget.repeat_signal_checkbox.isChecked())))
        self.controller_communication_queue.put(self.log_name,(GlobalCommands.AT_TARGET_LEVEL,self.environment_name))
    
    def stop_control(self):
        """Stops running the environment"""
        self.environment_command_queue.put(self.log_name,(GlobalCommands.STOP_ENVIRONMENT,
                                                          None))
    
    def create_window(self,event,control_index = None):
        """Creates a subwindow to show a specific channel information

        Parameters
        ----------
        event :
            
        control_index :
            Row index in the specification matrix to display (Default value = None)

        """
        if control_index is None:
            control_index = self.run_widget.control_channel_selector.currentIndex()
        self.plot_windows.append(
                PlotTimeWindow(None,control_index,
                          self.environment_parameters.control_signal,
                          self.data_acquisition_parameters.sample_rate,
                          self.run_widget.control_channel_selector.itemText(control_index),
                          ))
    
    def show_all_channels(self):
        """Creates a subwindow for each ASD in the CPSD matrix"""
        for i in range(self.environment_parameters.control_signal.shape[0]):
            self.create_window(None,i)
        self.tile_windows()
    
    def tile_windows(self):
        """Tile subwindow equally across the screen"""
        screen_rect = QtWidgets.QApplication.desktop().screenGeometry()
        # Go through and remove any closed windows
        self.plot_windows = [window for window in self.plot_windows if window.isVisible()]
        num_windows = len(self.plot_windows)
        ncols = int(np.ceil(np.sqrt(num_windows)))
        nrows = int(np.ceil(num_windows/ncols))
        window_width = screen_rect.width()/ncols
        window_height = screen_rect.height()/nrows
        for index,window in enumerate(self.plot_windows):
            window.resize(window_width,window_height)
            row_ind = index // ncols
            col_ind = index % ncols
            window.move(col_ind*window_width,row_ind*window_height)
            
    def close_windows(self):
        """Close all subwindows"""
        for window in self.plot_windows:
            window.close()
    
    def retrieve_metadata(self,netcdf_handle : nc4._netCDF4.Dataset):
        """Collects environment parameters from a netCDF dataset.

        This function retrieves parameters from a netCDF dataset that was written
        by the controller during streaming.  It must populate the widgets
        in the user interface with the proper information.

        This function is the "read" counterpart to the store_to_netcdf 
        function in the TransientParameters class, which will write 
        parameters to the netCDF file to document the metadata.
        
        Note that the entire dataset is passed to this function, so the function
        should collect parameters pertaining to the environment from a Group
        in the dataset sharing the environment's name, e.g.
        
        ``group = netcdf_handle.groups[self.environment_name]``
        ``self.definition_widget.parameter_selector.setValue(group.parameter)``
        
        Parameters
        ----------
        netcdf_handle : nc4._netCDF4.Dataset :
            The netCDF dataset from which the data will be read.  It should have
            a group name with the enviroment's name.
        """
        group = netcdf_handle.groups[self.environment_name]
        self.signal = group.variables['control_signal'][...].data
        self.definition_widget.ramp_selector.setValue(group.ramp_time)
        self.definition_widget.system_id_averaging_scheme_selector.setCurrentIndex(
            self.definition_widget.system_id_averaging_scheme_selector.findText(group.averaging_type))
        self.definition_widget.system_id_frames_to_average_selector.setValue(group.system_id_averages)
        self.definition_widget.system_id_averaging_coefficient_selector.setValue(group.averaging_coefficient)
        self.definition_widget.system_id_transfer_function_computation_window_selector.setCurrentIndex(
            self.definition_widget.system_id_transfer_function_computation_window_selector.findText(group.frf_window))
        self.definition_widget.system_id_overlap_percentage_selector.setValue(group.overlap*100)
        self.definition_widget.system_id_initial_drive_voltage_selector.setValue(group.frf_voltage)
        self.definition_widget.system_id_frf_technique_selector.setCurrentIndex(
            self.definition_widget.system_id_frf_technique_selector.findText(group.frf_technique))
        try:
            self.response_transformation_matrix = group.variables['response_transformation_matrix'][...].data
        except KeyError:
            self.response_transformation_matrix = None
        try:
            self.output_transformation_matrix = group.variables['output_transformation_matrix'][...].data
        except KeyError:
            self.output_transformation_matrix = None
        self.select_python_module(None,group.control_python_script)
        self.definition_widget.control_function_input.setCurrentIndex(
            self.definition_widget.control_function_input.findText(group.control_python_function))
        self.definition_widget.control_parameters_text_input.setText(group.control_python_function_parameters)
        maxs = np.max(np.abs(self.signal),axis=-1)
        rmss = rms_time(self.signal,axis=-1)
        for i,(mx,rms) in enumerate(zip(maxs,rmss)):
            self.definition_widget.signal_information_table.item(i,2).setText('{:0.2f}'.format(mx))
            self.definition_widget.signal_information_table.item(i,3).setText('{:0.2f}'.format(rms))
        self.show_signal()
    
    def change_test_level_from_profile(self,test_level):
        """Sets the test level from a profile instruction

        Parameters
        ----------
        test_level :
            Value to set the test level to.
        """
        self.run_widget.test_level_selector.setValue(int(test_level))
        
    
    def set_repeat_from_profile(self,data):
        """Sets the the signal to repeat from a profile instruction

        Parameters
        ----------
        data : Ignored
            Parameter is ignored but required by the ``command_map``

        """
        self.run_widget.repeat_signal_checkbox.setChecked(True)
    
    def set_norepeat_from_profile(self,data):
        """Sets the the signal to not repeat from a profile instruction

        Parameters
        ----------
        data : Ignored
            Parameter is ignored but required by the ``command_map``

        """
        self.run_widget.repeat_signal_checkbox.setChecked(False)
        
    def update_gui(self,queue_data):
        """Update the graphical interface for the environment

        Parameters
        ----------
        queue_data :
            A 2-tuple consisting of ``(message,data)`` pairs where the message
            denotes what to change and the data contains the information needed
            to be displayed.  
        """
        message,data = queue_data
        if message == 'time_data':
            response_data,output_data,signal_delay = data
            max_y = -1e15
            min_y = 1e15
            for curve,this_data in zip(self.plot_data_items['control_signal_measurement'],response_data):
                x,y = curve.getData()
                if np.max(y) > max_y:
                    max_y = np.max(y)
                if np.min(y) < min_y:
                    min_y = np.min(y)
                y = np.concatenate((y[this_data.size:],this_data[-x.size:]),axis=0)
                curve.setData(x,y)
            # Display the data
            for curve,this_output in zip(self.plot_data_items['output_signal_measurement'],output_data):
                x,y = curve.getData()
                y = np.concatenate((y[this_output.size:],this_output[-x.size:]),axis=0)
                curve.setData(x,y)
            if signal_delay is None:
                self.plot_data_items['signal_range'].setData(
                    np.zeros(5),np.zeros(5)
                    )
            else:
                self.plot_data_items['signal_range'].setData(
                    np.array((x[signal_delay],x[signal_delay],
                              x[signal_delay+self.environment_parameters.signal_samples-1],
                              x[signal_delay+self.environment_parameters.signal_samples-1],
                              x[signal_delay])),1.05*np.array((min_y,max_y,max_y,min_y,min_y))
                    )
        elif message == 'control_data':
            control_data,output_data = data
            # Go through and remove any closed windows
            self.plot_windows = [window for window in self.plot_windows if window.isVisible()]
            for window in self.plot_windows:
                window.update_plot(control_data)
        elif message == 'sysid_time_data':
            # Display the data
            control_data,output_data = data
            for curve,this_data in zip(self.plot_data_items['control_responses'],control_data):
                x,y = curve.getData()
                y = np.concatenate((y[this_data.size:],this_data[-x.size:]),axis=0)
                curve.setData(x,y)
            # Display the data
            for curve,this_output in zip(self.plot_data_items['drive_outputs'],output_data):
                x,y = curve.getData()
                y = np.concatenate((y[this_output.size:],this_output[-x.size:]),axis=0)
                curve.setData(x,y)
        elif message == 'FRF':
            # Display the data
            self.plot_data_items['transfer_function_phase'].setData(data[0],
                                np.angle(data[1][:,self.system_id_widget.transfer_function_response_selector.currentIndex(),
                                                   self.system_id_widget.transfer_function_reference_selector.currentIndex()]))
            self.plot_data_items['transfer_function_amplitude'].setData(data[0],
                                np.abs(data[1][:,self.system_id_widget.transfer_function_response_selector.currentIndex(),
                                                 self.system_id_widget.transfer_function_reference_selector.currentIndex()]))
        elif message == 'control_predictions':
            times,excitation_prediction,response_prediction,spec_signal = data
            index = self.prediction_widget.excitation_selector.currentIndex()
            self.plot_data_items['excitation_prediction'][0].setData(times,
                                excitation_prediction[index])
            index = self.prediction_widget.response_selector.currentIndex()
            self.plot_data_items['response_prediction'][0].setData(times,
                                response_prediction[index])
            self.plot_data_items['response_prediction'][1].setData(times,
                                spec_signal[index])
        elif message == 'enable':
            widget = None
            for parent in [self.definition_widget,self.run_widget,self.system_id_widget,self.prediction_widget]:
                try:
                    widget = getattr(parent,data)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Enable Widget {:}: not found in UI'.format(data))
            widget.setEnabled(True)
        elif message == 'disable':
            widget = None
            for parent in [self.definition_widget,self.run_widget,self.system_id_widget,self.prediction_widget]:
                try:
                    widget = getattr(parent,data)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Disable Widget {:}: not found in UI'.format(data))
            widget.setEnabled(False)
        else:
            widget = None
            for parent in [self.definition_widget,self.run_widget,self.system_id_widget,self.prediction_widget]:
                try:
                    widget = getattr(parent,message)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Update Widget {:}: not found in UI'.format(message))
            if type(widget) is QtWidgets.QDoubleSpinBox:
                widget.setValue(data)
            elif type(widget) is QtWidgets.QSpinBox:
                widget.setValue(data)
            elif type(widget) is QtWidgets.QLineEdit:
                widget.setText(data)
            elif type(widget) is QtWidgets.QListWidget:
                widget.clear()
                widget.addItems(['{:.3f}'.format(d) for d in data])
                
    @staticmethod
    def create_environment_template(environment_name: str, workbook : openpyxl.workbook.workbook.Workbook):
        """Creates a template worksheet in an Excel workbook defining the
        environment.
        
        This function creates a template worksheet in an Excel workbook that
        when filled out could be read by the controller to re-create the 
        environment.
        
        This function is the "write" counterpart to the 
        ``set_parameters_from_template`` function in the ``TransientUI`` class,
        which reads the values from the template file to populate the user
        interface.

        Parameters
        ----------
        environment_name : str :
            The name of the environment that will specify the worksheet's name
        workbook : openpyxl.workbook.workbook.Workbook :
            A reference to an ``openpyxl`` workbook.

        """
        worksheet = workbook.create_sheet(environment_name)
        worksheet.cell(1,1,'Control Type')
        worksheet.cell(1,2,'Transient')
        worksheet.cell(1,4,'Note: Replace cells with hash marks (#) to provide the requested parameters.')
        worksheet.cell(2,1,'Signal File')
        worksheet.cell(2,2,'# Path to the file that contains the time signal that will be output')
        worksheet.cell(3,1,'Ramp Time')
        worksheet.cell(3,2,'# Time for the environment to ramp between levels or from start or to stop.')
        worksheet.cell(4,1,'Averaging Type:')
        worksheet.cell(4,2,'# Averaging Type')
        worksheet.cell(5,1,'System ID Averages:')
        worksheet.cell(5,2,'# Number of Averages used when computing the FRF')
        worksheet.cell(6,1,'Averaging Coefficient:')
        worksheet.cell(6,2,'# Averaging Coefficient for Exponential Averaging')
        worksheet.cell(7,1,'FRF Technique:')
        worksheet.cell(7,2,'# FRF Technique')
        worksheet.cell(8,1,'FRF Window:')
        worksheet.cell(8,2,'# Window used to compute FRF')
        worksheet.cell(9,1,'Overlap Percentage:')
        worksheet.cell(9,2,'# Overlap percentage for CPSD and FRF calculations')
        worksheet.cell(10,1,'System ID RMS Voltage:')
        worksheet.cell(10,2,'# RMS Value of Flat Voltage Spectrum used for System Identification')
        worksheet.cell(11,1,'Control Python Script:')
        worksheet.cell(11,2,'# Path to the Python script containing the control law')
        worksheet.cell(12,1,'Control Python Function:')
        worksheet.cell(12,2,'# Function name within the Python Script that will serve as the control law')
        worksheet.cell(13,1,'Control Parameters:')
        worksheet.cell(13,2,'# Extra parameters used in the control law')
        worksheet.cell(14,1,'Response Transformation Matrix:')
        worksheet.cell(14,2,'# Transformation matrix to apply to the response channels.  Type None if there is none.  Otherwise, make this a 2D array in the spreadsheet and move the Output Transformation Matrix line down so it will fit.  The number of columns should be the number of physical control channels.')
        worksheet.cell(15,1,'Output Transformation Matrix:')
        worksheet.cell(15,2,'# Transformation matrix to apply to the outputs.  Type None if there is none.  Otherwise, make this a 2D array in the spreadsheet.  The number of columns should be the number of physical output channels in the environment.')
    
    def set_parameters_from_template(self, worksheet : openpyxl.worksheet.worksheet.Worksheet):
        """
        Collects parameters for the user interface from the Excel template file
        
        This function reads a filled out template worksheet to create an
        environment.  Cells on this worksheet contain parameters needed to
        specify the environment, so this function should read those cells and
        update the UI widgets with those parameters.
        
        This function is the "read" counterpart to the 
        ``create_environment_template`` function in the ``TransientUI`` class,
        which writes a template file that can be filled out by a user.
        

        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            An openpyxl worksheet that contains the environment template.
            Cells on this worksheet should contain the parameters needed for the
            user interface.

        """
        self.load_signal(None,worksheet.cell(2,2).value)
        self.definition_widget.ramp_selector.setValue(float(worksheet.cell(3,2).value))
        self.definition_widget.system_id_averaging_scheme_selector.setCurrentIndex(self.definition_widget.system_id_averaging_scheme_selector.findText(worksheet.cell(4,2).value))
        self.definition_widget.system_id_frames_to_average_selector.setValue(int(worksheet.cell(5,2).value))
        self.definition_widget.system_id_averaging_coefficient_selector.setValue(float(worksheet.cell(6,2).value))
        self.definition_widget.system_id_frf_technique_selector.setCurrentIndex(self.definition_widget.system_id_frf_technique_selector.findText(worksheet.cell(7,2).value))
        self.definition_widget.system_id_transfer_function_computation_window_selector.setCurrentIndex(self.definition_widget.system_id_transfer_function_computation_window_selector.findText(worksheet.cell(8,2).value))
        self.definition_widget.system_id_overlap_percentage_selector.setValue(float(worksheet.cell(9,2).value))
        self.definition_widget.system_id_initial_drive_voltage_selector.setValue(float(worksheet.cell(10,2).value))
        self.select_python_module(None,worksheet.cell(11,2).value)
        self.definition_widget.control_function_input.setCurrentIndex(self.definition_widget.control_function_input.findText(worksheet.cell(12,2).value))
        self.definition_widget.control_parameters_text_input.setText(str(worksheet.cell(13,2).value))
        # Now we need to find the transformation matrices' sizes
        response_channels = self.definition_widget.control_channels_display.value()
        output_channels = self.definition_widget.output_channels_display.value()
        output_transform_row = 15
        if isinstance(worksheet.cell(14,2).value,str) and worksheet.cell(14,2).value.lower() == 'none':
            self.response_transformation_matrix = None
        else:
            while True:
                if worksheet.cell(output_transform_row,1).value == 'Output Transformation Matrix:':
                    break
                output_transform_row += 1
            response_size = output_transform_row-14
            response_transformation = []
            for i in range(response_size):
                response_transformation.append([])
                for j in range(response_channels):
                    response_transformation[-1].append(float(worksheet.cell(14+i,2+j).value))
            self.response_transformation_matrix = np.array(response_transformation)
        if isinstance(worksheet.cell(output_transform_row,2).value,str) and worksheet.cell(output_transform_row,2).value.lower() == 'none':
            self.output_transformation_matrix = None
        else:
            output_transformation = []
            i = 0
            while True:
                if worksheet.cell(output_transform_row+i,2).value is None or (isinstance(worksheet.cell(output_transform_row+i,2).value,str) and worksheet.cell(output_transform_row+i,2).value.strip() == ''):
                    break
                output_transformation.append([])
                for j in range(output_channels):
                    output_transformation[-1].append(float(worksheet.cell(output_transform_row+i,2+j).value))
                i += 1
            self.output_transformation_matrix = np.array(output_transformation)
        self.define_transformation_matrices(None,dialog=False)
        
class TransientEnvironment(AbstractEnvironment):
    """Environment that attempts to match responses to a specified signal"""
        
    def __init__(self,
                 environment_name : str,
                 queue_container : TransientQueues):
        """
        Transient Environment Constructor 
    
        This function fills out the command map and initializes parameters to
        zero or null.

        Parameters
        ----------
        environment_name : str
            Name of the environment.
        queue_container : TransientQueues
            Container of queues used by the Transient Environment.

        """
        super().__init__(
                environment_name,
                queue_container.environment_command_queue,
                queue_container.gui_update_queue,
                queue_container.controller_communication_queue,
                queue_container.log_file_queue,
                queue_container.data_in_queue,
                queue_container.data_out_queue)
        self.queue_container = queue_container
        # Define command map
        self.command_map[GlobalCommands.START_ENVIRONMENT] = self.run_environment
        self.command_map[TransientCommands.START_TRANSFER_FUNCTION] = self.start_transfer_function
        self.command_map[TransientCommands.STOP_TRANSFER_FUNCTION] = self.stop_transfer_function
        self.command_map[TransientCommands.ADJUST_TEST_LEVEL] = self.adjust_test_level
        self.command_map[TransientCommands.PERFORM_CONTROL_PREDICTION] = self.perform_control_prediction
        self.command_map[TransientCommands.SHOW_TEST_PREDICTION] = self.show_test_prediction
        self.command_map[TransientCommands.SHOW_FRF] = self.show_frf
        # Persistent data
        self.data_acquisition_parameters = None
        self.environment_parameters = None
        self.startup = True
        self.shutdown_flag = False
        self.current_test_level = 0.0
        self.target_test_level = 0.0
        self.test_level_change = 0.0
        self.repeat = False
        self.output_channels = None
        self.control_channels = None
        self.frf_proc = None
        self.siggen_proc = None
        self.skip_frames = 0
        self.test_level = None
        self.control_sysid_buffer = None
        self.output_sysid_buffer = None
        self.buffer_position = 0
        self.frf_window = None
        self.skip_frames_for_test_level_change = None
        self.frequencies = None
        self.frf = None
        self.next_response_prediction = None
        self.next_drive = None
        self.control_function = None
        self.control_function_type = None
        self.extra_control_parameters = None
        self.control_measurement_buffer = None
        self.output_measurement_buffer = None
        self.aligned_response = None
        self.aligned_output = None
    
    def initialize_data_acquisition_parameters(self,data_acquisition_parameters : DataAcquisitionParameters):
        """Initialize the data acquisition parameters in the environment.
        
        The environment will receive the global data acquisition parameters from
        the controller, and must set itself up accordingly.

        Parameters
        ----------
        data_acquisition_parameters : DataAcquisitionParameters :
            A container containing data acquisition parameters, including
            channels active in the environment as well as sampling parameters.
        """
        self.log('Initializing Data Acquisition Parameters')
        self.data_acquisition_parameters = data_acquisition_parameters
        self.control_channels = [index for index,channel in enumerate(self.data_acquisition_parameters.channel_list) if channel.control]
        self.output_channels = [index for index,channel in enumerate(self.data_acquisition_parameters.channel_list) if not channel.feedback_device is None]
    
    def initialize_environment_test_parameters(self,environment_parameters : TransientParameters):
        """
        Initialize the environment parameters specific to this environment
        
        The environment will recieve parameters defining itself from the
        user interface and must set itself up accordingly.

        Parameters
        ----------
        environment_parameters : TransientParameters
            A container containing the parameters defining the environment

        """
        self.log('Initializing Environment Parameters')
        self.environment_parameters = environment_parameters
        self.skip_frames_for_test_level_change = int(np.ceil(
            self.environment_parameters.ramp_samples/self.data_acquisition_parameters.samples_per_read)+2)
        # Set up the buffers
        # Only clear the buffer if the data processing parameters have changed, otherwise we will zero out some data.
        control_buffer_shape = (len(self.control_channels) if self.environment_parameters.response_transformation_matrix is None
                                else self.environment_parameters.response_transformation_matrix.shape[0],
                                int(self.environment_parameters.signal_samples*(2-self.environment_parameters.overlap)))
        output_buffer_shape = (len(self.output_channels) if self.environment_parameters.output_transformation_matrix is None
                               else self.environment_parameters.output_transformation_matrix.shape[0],
                                int(self.environment_parameters.signal_samples*(2-self.environment_parameters.overlap)))
        self.frf_window = sig.windows.get_window(self.environment_parameters.frf_window.lower(),self.environment_parameters.signal_samples,fftbins=True)
        if (self.control_sysid_buffer is None or self.output_sysid_buffer is None or
            self.control_sysid_buffer.shape != control_buffer_shape or
            self.output_sysid_buffer.shape != output_buffer_shape):
            self.control_sysid_buffer = OverlapBuffer(control_buffer_shape)
            self.output_sysid_buffer = OverlapBuffer(output_buffer_shape)
        # Only set to zeros if they are none or the wrong shape
        if (self.frf is None
            or self.frf.shape != (self.environment_parameters.fft_lines,len(self.control_channels),len(self.output_channels))
            ):
            self.frf = np.zeros((self.environment_parameters.fft_lines,len(self.control_channels),len(self.output_channels)),dtype='complex128')
            self.response_prediction = np.zeros((len(self.control_channels),self.environment_parameters.signal_samples))
            self.drive_prediction = np.zeros((len(self.output_channels),self.environment_parameters.signal_samples))
        self.frequencies = self.environment_parameters.frequency_spacing * np.arange(self.environment_parameters.fft_lines)
        # Load in the control law
        path,file = os.path.split(environment_parameters.control_python_script)
        file,ext = os.path.splitext(file)
        spec = importlib.util.spec_from_file_location(file, environment_parameters.control_python_script)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        self.control_function_type = environment_parameters.control_python_function_type
        self.extra_control_parameters = environment_parameters.control_python_function_parameters
        if self.control_function_type == 1: # Generator
            # Get the generator function
            generator_function = getattr(module,environment_parameters.control_python_function)()
            # Get us to the first yield statement
            next(generator_function)
            # Define the control function as the generator's send function
            self.control_function = generator_function.send
        elif self.control_function_type == 2: # Class
            self.control_function = getattr(module,environment_parameters.control_python_function)(
                self.environment_parameters.control_signal,self.extra_control_parameters, # Required parameters
                self.data_acquisition_parameters.output_oversample,self.frf,
                self.aligned_output,self.aligned_response) # Optional parameters
        else: # Function
            self.control_function = getattr(module,environment_parameters.control_python_function)
        
    
    def start_transfer_function(self,data):
        """Starts the transfer function running

        Parameters
        ----------
        data : tuple
            A 3-tuple containing (acquire,rms_voltage,test_level) where
            acquire is True if in acquisition mode and False if in Preview mode,
            rms_voltage is the base RMS level that the signal will be generated
            ad, and test_level is the linear scaling on that base rms level

        """
        if self.startup:
            acquire,rms_voltage,self.test_level = data
            self.log('Starting Transfer Function in {:} Mode with {:} V RMS at {:}x'.format('"Acquire"' if acquire else '"Preview"',rms_voltage,self.test_level))
            # Create the processes
            self.frf_proc = mp.Process(target=frf_computation_process,args=(self.environment_name,
                                                                            self.queue_container.frf_command_queue,
                                                                            self.queue_container.data_for_frf_queue,
                                                                            self.queue_container.updated_frf_queue,
                                                                            self.queue_container.gui_update_queue,
                                                                            self.queue_container.log_file_queue))
            self.frf_proc.start()
            self.siggen_proc = mp.Process(target=system_id_signal_generation_process,args=(self.environment_name,
                                                                                      self.queue_container.signal_generation_command_queue,
                                                                                      self.queue_container.log_file_queue,
                                                                                      self.queue_container.gui_update_queue,
                                                                                      self.queue_container.data_out_queue,
                                                                                      self.queue_container.controller_communication_queue))
            self.siggen_proc.start()
            # Initialize the data
            self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.INITIALIZE_DATA_ACQUISITION,self.data_acquisition_parameters))
            self.queue_container.signal_generation_command_queue.put(self.environment_name,(SysIdSignalGenerationMessages.INITIALIZE_DATA_ACQUISITION,self.data_acquisition_parameters))
            self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.INITIALIZE_TEST_PARAMETERS,self.environment_parameters))
            self.queue_container.signal_generation_command_queue.put(self.environment_name,(SysIdSignalGenerationMessages.INITIALIZE_TEST_PARAMETERS,(self.environment_parameters.signal_samples,
                                                                                                                                                      self.environment_parameters.frequency_spacing,
                                                                                                                                                      self.environment_parameters.ramp_samples,
                                                                                                                                                      self.environment_parameters.output_transformation_matrix)))
            self.skip_frames = self.skip_frames_for_test_level_change
            time.sleep(0.01)
            self.queue_container.signal_generation_command_queue.put(self.environment_name,(SysIdSignalGenerationMessages.MUTE,None))
            self.queue_container.signal_generation_command_queue.put(self.environment_name,(SysIdSignalGenerationMessages.ADJUST_TEST_LEVEL,self.test_level))
            self.queue_container.signal_generation_command_queue.put(self.environment_name,(SysIdSignalGenerationMessages.RUN_TRANSFER_FUNCTION,rms_voltage))
            self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.CLEAR_FRF,None))
            self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.RUN_FRF,None))
            self.startup = False
        # See if any data has come in
        try:
            # Get data
            acquisition_data,last_acquisition = self.queue_container.data_in_queue.get_nowait()
            self.log('Acquired Data')
            # Parse data into control and output channels
            control_data = acquisition_data[self.control_channels]
            if not self.environment_parameters.response_transformation_matrix is None:
                control_data = self.environment_parameters.response_transformation_matrix@control_data
            output_data = acquisition_data[self.output_channels]
            if not self.environment_parameters.output_transformation_matrix is None:
                output_data = self.environment_parameters.output_transformation_matrix@output_data
            self.log('Parsed Channels')
            # Send the data up to the GUI
            self.queue_container.gui_update_queue.put((self.environment_name,('sysid_time_data',(control_data,output_data))))
            self.log('Sent Data to GUI')
            if self.skip_frames > 0:
                self.skip_frames -= 1
                self.log('Skipped Frame, {:} left'.format(self.skip_frames))
                # print('Skipped Frame, {:} left'.format(self.skip_frames))
            elif self.test_level != 0.0:
                # Add data to the buffer
                self.control_sysid_buffer.add_data(control_data)
                self.output_sysid_buffer.add_data(output_data)
                if self.control_sysid_buffer.buffer_position >= self.environment_parameters.samples_per_frame:
                    self.log('Sending Data')
                    buffer_shift = int(self.environment_parameters.signal_samples*(1-self.environment_parameters.overlap))
                    control_data = self.control_sysid_buffer.get_data(
                        self.environment_parameters.samples_per_frame,
                        -buffer_shift)
                    output_data = self.output_sysid_buffer.get_data(
                        self.environment_parameters.samples_per_frame,
                        -buffer_shift)
                    self.log('Extracted Data, Computing FFTs')
                    # print('Extracted Data, Computing FFTs')
                    control_fft = rfft(control_data*self.frf_window/self.test_level,axis=-1)
                    output_fft = rfft(output_data*self.frf_window/self.test_level,axis=-1)
                    self.queue_container.data_for_frf_queue.put(copy.deepcopy(
                            (control_fft,
                             output_fft)))
                    self.log('Sent Data')
                    # print('Sent Data')
        except mp.queues.Empty:
            pass
        # Get any frf data that exists
        frf_data = flush_queue(self.queue_container.updated_frf_queue,timeout=0.01)
        if len(frf_data) > 0:
            self.frf[:],frf_frames = frf_data[-1]
            self.queue_container.gui_update_queue.put((self.environment_name,('system_id_current_average_display',frf_frames)))
            # print('FRF Frames {:} > {:}, Stopping = {:}'.format(frf_frames,self.environment_parameters.system_id_averages,frf_frames >= self.environment_parameters.system_id_averages))
        else:
            frf_frames = 0
        if frf_frames >= self.environment_parameters.system_id_averages:
            self.stop_transfer_function(None)
            self.queue_container.environment_command_queue.put(self.environment_name,(TransientCommands.PERFORM_CONTROL_PREDICTION,None))
            self.queue_container.controller_communication_queue.put(self.environment_name,(GlobalCommands.COMPLETED_SYSTEM_ID,self.environment_name))
        else:
            self.queue_container.environment_command_queue.put(self.environment_name,(TransientCommands.START_TRANSFER_FUNCTION,None))
    
    def stop_transfer_function(self,data):
        self.log('Stopping Data Collection')
        self.queue_container.controller_communication_queue.put(self.environment_name,(GlobalCommands.STOP_HARDWARE,None))
        flush_queue(self.queue_container.data_for_frf_queue)
        self.environment_command_queue.flush(self.environment_name)
        self.control_sysid_buffer.set_buffer_position()
        self.output_sysid_buffer.set_buffer_position()
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(SysIdSignalGenerationMessages.QUIT_EVENTUALLY,None))
        self.queue_container.frf_command_queue.put(self.environment_name,(FRFMessages.STOP_FRF,None))
        self.queue_container.signal_generation_command_queue.put(self.environment_name,(SysIdSignalGenerationMessages.START_SHUTDOWN,None))
        self.queue_container.frf_command_queue.put(self.environment_name,(GlobalCommands.QUIT,None))
        last_acquisition = False
        while not last_acquisition:
            self.log('Acquiring Remaining Data')
            # Get data
            acquisition_data,last_acquisition = self.queue_container.data_in_queue.get()
            self.log('Acquired Data')
            # Parse data into control and output channels
            control_data = acquisition_data[self.control_channels]
            if not self.environment_parameters.response_transformation_matrix is None:
                control_data = self.environment_parameters.response_transformation_matrix@control_data
            output_data = acquisition_data[self.output_channels]
            if not self.environment_parameters.output_transformation_matrix is None:
                output_data = self.environment_parameters.output_transformation_matrix@output_data
            self.log('Parsed Channels')
            # Send the data up to the GUI
            self.queue_container.gui_update_queue.put((self.environment_name,('sysid_time_data',(control_data,output_data))))
            self.log('Sent Data to GUI')
        # Rejoin the processes
        print('Joining FRF Process')
        self.frf_proc.join()
        print('Joining Signal Generation Process')
        self.siggen_proc.join()
        print('Joined Processes')
        self.startup = True
        self.queue_container.gui_update_queue.put((self.environment_name,('enable','preview_transfer_function_button')))
        self.queue_container.gui_update_queue.put((self.environment_name,('enable','acquire_transfer_function_button')))
        self.queue_container.gui_update_queue.put((self.environment_name,('disable','stop_transfer_function_button')))
    
    def perform_control_prediction(self,data):
        """Performs control predictions based off the control law and transfer function
        
        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``
        """
        if self.control_function_type == 1: # Generator
            output_time_history = self.control_function((
                self.frf,
                self.environment_parameters.control_signal,
                self.data_acquisition_parameters.output_oversample,
                self.extra_control_parameters,
                None,None))
        elif self.control_function_type == 2: # Class
            self.control_function.system_id_update(self.frf)
            output_time_history = self.control_function.control(None,None)
        else: # Function
            output_time_history = self.control_function(
                self.frf,
                self.environment_parameters.control_signal,
                self.data_acquisition_parameters.output_oversample,
                self.extra_control_parameters,
                None,None
                    )
        self.next_drive = output_time_history
        self.queue_container.environment_command_queue.put(self.environment_name,(TransientCommands.SHOW_TEST_PREDICTION,True))
        
    
    def show_test_prediction(self,recompute):
        if recompute:
            # Predict response
            output_time_history_fft = np.fft.rfft(self.next_drive,axis=-1).T[...,np.newaxis]
            # Un-oversample
            nfreq = (output_time_history_fft.shape[0]-1)//self.data_acquisition_parameters.output_oversample+1
            response_time_history_fft = self.frf@output_time_history_fft[:nfreq]
            response_prediction = np.fft.irfft(response_time_history_fft.squeeze().T,axis=-1)/self.data_acquisition_parameters.output_oversample
            time_trac = trac(response_prediction,self.environment_parameters.control_signal)
            peak_voltages = np.max(np.abs(self.next_drive),axis=-1)
            self.next_response_prediction = response_prediction
            self.queue_container.gui_update_queue.put((self.environment_name,('excitation_voltage_list',peak_voltages)))
            self.queue_container.gui_update_queue.put((self.environment_name,('response_error_list',time_trac)))
        self.queue_container.gui_update_queue.put((self.environment_name,('control_predictions',(np.arange(self.environment_parameters.signal_samples)/self.data_acquisition_parameters.sample_rate,
                                                                       self.next_drive[...,::self.data_acquisition_parameters.output_oversample],
                                                                       self.next_response_prediction,
                                                                       self.environment_parameters.control_signal))))
    
    def show_frf(self,data):
        self.gui_update_queue.put((self.environment_name,('FRF',(np.fft.rfftfreq(self.environment_parameters.signal_samples,
                                                                                 1/self.data_acquisition_parameters.sample_rate),self.frf))))
        
    def run_environment(self,data):
        """Runs the Transient history environment.
        
        This function handles start up, running, and shutting down the environment

        Parameters
        ----------
        data : Tuple
            A tuple containing the test level to run the environment at and
            a boolean specifying whether or not to repeat the signal.

        """
        if self.startup:
            self.log('Starting Environment')
            self.mute(None)
            test_level,self.repeat = data
            self.adjust_test_level(test_level)
            self.log('Test Level set to {:}'.format(self.current_test_level))
            n_control_channels = len(self.control_channels) if self.environment_parameters.response_transformation_matrix is None else self.environment_parameters.response_transformation_matrix.shape[0]
            n_output_channels = len(self.output_channels) if self.environment_parameters.output_transformation_matrix is None else self.environment_parameters.output_transformation_matrix.shape[0]
            self.control_measurement_buffer = OverlapBuffer((n_control_channels,2*self.environment_parameters.signal_samples))
            self.output_measurement_buffer = OverlapBuffer((n_output_channels,2*self.environment_parameters.signal_samples))
            self.output(self.next_drive,not self.repeat) # If not repeat, then it is the last signal
            self.startup = False
        # See if any data has come in
        try:
            acquisition_data,last_acquisition = self.queue_container.data_in_queue.get_nowait()
            self.log('Acquired Data')
            scale_factor = 0.0 if self.current_test_level < 1e-10 else 1/self.current_test_level
            control_data = acquisition_data[self.control_channels]*scale_factor
            if not self.environment_parameters.response_transformation_matrix is None:
                control_data = self.environment_parameters.response_transformation_matrix@control_data
            output_data = acquisition_data[self.output_channels]*scale_factor
            if not self.environment_parameters.output_transformation_matrix is None:
                output_data = self.environment_parameters.output_transformation_matrix@output_data
            # Add the data to the buffers
            self.control_measurement_buffer.add_data(control_data)
            self.output_measurement_buffer.add_data(output_data)
            # Find alignment with the specification via output
            self.log('Aligning signal with specification')
            self.aligned_output,sample_delay,phase_change = align_signals(self.output_measurement_buffer[:],self.next_drive[:,::self.data_acquisition_parameters.output_oversample])
            self.queue_container.gui_update_queue.put((self.environment_name,('time_data',(control_data,output_data,sample_delay)))) # Sample_delay will be None if the alignment is not found
            if not self.aligned_output is None:
                self.log('Alignment Found at {:} samples'.format(sample_delay))
                self.aligned_response = shift_signal(self.control_measurement_buffer[:],self.environment_parameters.control_signal.shape[-1],sample_delay,phase_change)
                self.queue_container.gui_update_queue.put((self.environment_name,('control_data',(self.aligned_response,self.aligned_output)))) # Sample_delay will be None if the alignment is not found
                # Do the next control
                self.log('Computing next signal via control law')
                if self.control_function_type == 1: # Generator
                    output_time_history = self.control_function((
                        self.frf,
                        self.environment_parameters.control_signal,
                        self.data_acquisition_parameters.output_oversample,
                        self.extra_control_parameters,
                        self.aligned_output,self.aligned_response))
                elif self.control_function_type == 2: # Class
                    self.control_function.system_id_update(self.frf)
                    output_time_history = self.control_function.control(self.aligned_output,self.aligned_response)
                else: # Function
                    output_time_history = self.control_function(
                        self.frf,
                        self.environment_parameters.control_signal,
                        self.data_acquisition_parameters.output_oversample,
                        self.extra_control_parameters,
                        self.aligned_output,self.aligned_response)
                self.next_drive = output_time_history
                self.queue_container.environment_command_queue.put(self.environment_name,(TransientCommands.SHOW_TEST_PREDICTION,True))
        except mp.queues.Empty:
            last_acquisition = False
        last_signal = False
        # See if we need to output data
        if self.queue_container.data_out_queue.empty() and self.repeat:
            # Check to see if we're at the last test level
            if self.current_test_level == 0.0:
                last_signal = True
            self.log('Sending {:} signal'.format('last' if last_signal else 'next'))
            self.output(self.next_drive, last_signal)
        if last_signal or not self.repeat:
            self.log('Last signal is output')
            # Wait until we get the last signal from the acquisition
            while not last_acquisition:
                self.log('Waiting for Last Acquisition')
                scale_factor = 0.0 if self.current_test_level < 1e-10 else 1/self.current_test_level 
                acquisition_data,last_acquisition = self.queue_container.data_in_queue.get()
                control_data = acquisition_data[self.control_channels]*scale_factor
                if not self.environment_parameters.response_transformation_matrix is None:
                    control_data = self.environment_parameters.response_transformation_matrix@control_data
                output_data = acquisition_data[self.output_channels]*scale_factor
                if not self.environment_parameters.output_transformation_matrix is None:
                    output_data = self.environment_parameters.output_transformation_matrix@output_data
                # Add the data to the buffers
                self.control_measurement_buffer.add_data(control_data)
                self.output_measurement_buffer.add_data(output_data)
                # Find alignment with the specification via output
                self.log('Aligning signal with specification')
                self.aligned_output,sample_delay,phase_change = align_signals(self.output_measurement_buffer[:],self.next_drive[:,::self.data_acquisition_parameters.output_oversample])
                self.queue_container.gui_update_queue.put((self.environment_name,('time_data',(control_data,output_data,sample_delay)))) # Sample_delay will be None if the alignment is not found
                if not self.aligned_output is None:
                    self.log('Alignment Found at {:} samples'.format(sample_delay))
                    self.aligned_response = shift_signal(self.control_measurement_buffer[:],self.environment_parameters.control_signal.shape[-1],sample_delay,phase_change)
                    self.queue_container.gui_update_queue.put((self.environment_name,('control_data',(self.aligned_response,self.aligned_output)))) # Sample_delay will be None if the alignment is not found
                    # Do the next control
                    self.log('Computing next signal via control law')
                    if self.control_function_type == 1: # Generator
                        output_time_history = self.control_function((
                            self.frf,
                            self.environment_parameters.control_signal,
                            self.data_acquisition_parameters.output_oversample,
                            self.extra_control_parameters,
                            self.aligned_output,self.aligned_response))
                    elif self.control_function_type == 2: # Class
                        self.control_function.system_id_update(self.frf)
                        output_time_history = self.control_function.control(self.aligned_output,self.aligned_response)
                    else: # Function
                        output_time_history = self.control_function(
                            self.frf,
                            self.environment_parameters.control_signal,
                            self.data_acquisition_parameters.output_oversample,
                            self.extra_control_parameters,
                            self.aligned_output,self.aligned_response)
                    self.next_drive = output_time_history
                    self.queue_container.environment_command_queue.put(self.environment_name,(TransientCommands.SHOW_TEST_PREDICTION,True))
            self.shutdown()
            return
        self.queue_container.environment_command_queue.put(self.environment_name,(GlobalCommands.START_ENVIRONMENT,None))

    def output(self,write_data,last_signal=False):
        """Puts data to the data_out_queue and handles test level changes
        
        This function keeps track of the environment test level and scales the
        output signals accordingly prior to placing them into the data_out_queue.
        This function also handles the ramping between two test levels.

        Parameters
        ----------
        write_data : np.ndarray
            A numpy array containing the signals to be written.
            
        last_signal :
            Specifies if the signal being written is the last signal that will
            be generated due to the signal generation shutting down.  This is
            passed to the output task to tell it that there will be no more
            signals from this environment until it is restarted. (Default value
            = False)
        """
#        np.savez('test_data/signal_generation_initial_output_data_check.npz',write_data = write_data)
        # Perform the output transformation if necessary
        if not self.environment_parameters.output_transformation_matrix is None:
            self.log('Applying Transformation')
            write_data = self.output_transformation_matrix@write_data
        # Compute the test_level scaling for this dataset
        if self.test_level_change == 0.0:
            test_level = self.current_test_level
            self.log('Test Level at {:}'.format(test_level))
        else:
            test_level = self.current_test_level + (np.arange(self.environment_parameters.signal_samples*self.data_acquisition_parameters.output_oversample)+1)*self.test_level_change
            # Compute distance in steps from the target test_level and find where it is near the target
            full_level_index = np.nonzero(abs(test_level - self.test_level_target)/abs(self.test_level_change) < test_level_threshold)[0]
            # Check if any are
            if len(full_level_index) > 0:
                # If so, set all test_levels after that one to the target test_level
                test_level[full_level_index[0]+1:] = self.test_level_target
                # And update that our current test_level is now the target test_level
                self.current_test_level = self.test_level_target
                self.test_level_change = 0.0
            else:
                # Otherwise, our current test_level is the last entry in the test_level scaling
                self.current_test_level = test_level[-1]
            self.log('Test level from {:} to {:}'.format(test_level[0],test_level[-1]))
        # Write the test level-scaled data to the task
        self.log('Sending data to data_out queue')
        # np.savez('test_data/signal_generation_output_data_check.npz',write_data = write_data,test_level = test_level)
        self.queue_container.data_out_queue.put((copy.deepcopy(write_data*test_level),last_signal))

    def stop_environment(self,data):
        """Stops the environment by setting the test level to zero.

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but is required for the
            ``command_map`` calling signature.

        """
        self.adjust_test_level(0.0)

    def mute(self,data):
        """Immediately mute the signal generation task
        
        This function should primarily only be called at the beginning of an
        analysis to ensure the system ramps up from zero excitation.  Muting
        the signal generation during excitation will shock load the exciters and
        test article and may damage those hardware.

        Parameters
        ----------
        data : None
            Unused argument required due to the expectation that functions called
            by the RandomSignalGenerationProcess.Run function will have one argument
            accepting any data passed along with the instruction.

        """
        self.current_test_level = 0.0
        self.test_level_target = 0.0
        self.test_level_change = 0.0
        
    def adjust_test_level(self,data):
        """Sets a new target test level and computes the test level change per sample

        Parameters
        ----------
        data : float:
            The new test level target scale factor.

        """
        self.test_level_target = data
        self.test_level_change = (self.test_level_target - self.current_test_level)/self.environment_parameters.ramp_samples
        if self.test_level_change != 0.0:
            self.log('Changed test level to {:} from {:}, {:} change per sample'.format(self.test_level_target,self.current_test_level,self.test_level_change))

    def shutdown(self):
        """Performs final cleanup operations when the system has shut down
        
        This function is called when the environment has been instructed
        to shut down and the last acquisition data has been received.  The signal generation
        is the first process in the Random Vibration environment to stop when
        shutdown is called, so it notifies the environment process to stop the
        acquisition and analysis tasks because it is no longer generating signals
        
        """
        self.log('Shutting Down Transient Control')
        self.queue_container.environment_command_queue.flush(self.environment_name)
        # Enable the volume controls
        self.queue_container.gui_update_queue.put((self.environment_name,('enable','test_level_selector')))
        self.queue_container.gui_update_queue.put((self.environment_name,('enable','repeat_signal_checkbox')))
        self.queue_container.gui_update_queue.put((self.environment_name,('enable','start_test_button')))
        self.queue_container.gui_update_queue.put((self.environment_name,('disable','stop_test_button')))
        self.startup = True

def transient_process(environment_name : str,
                 input_queue : VerboseMessageQueue,
                 gui_update_queue : Queue,
                 controller_communication_queue : VerboseMessageQueue,
                 log_file_queue : Queue,
                 data_in_queue : Queue,
                 data_out_queue : Queue):
    """Transient environment process function called by multiprocessing
    
    This function defines the environment process that
    gets run by the multiprocessing module when it creates a new process.  It
    creates a TransientEnviornment object and runs it.

    Parameters
    ----------
    environment_name : str :
        Name of the environment
    input_queue : VerboseMessageQueue :
        Queue containing instructions for the environment
    gui_update_queue : Queue :
        Queue where GUI updates are put
    controller_communication_queue : Queue :
        Queue for global communications with the controller
    log_file_queue : Queue :
        Queue for writing log file messages
    data_in_queue : Queue :
        Queue from which data will be read by the environment
    data_out_queue : Queue :
        Queue to which data will be written that will be output by the hardware.

    """
    
    
    # Create vibration queues
    queue_container = TransientQueues(environment_name,
                                      input_queue,
                                      gui_update_queue,
                                      controller_communication_queue,
                                      data_in_queue,
                                      data_out_queue,
                                      log_file_queue)

    process_class = TransientEnvironment(
            environment_name,
            queue_container)
    process_class.run()
    
    # # Rejoin all the processes
    # process_class.log('Joining Subprocesses')
    # process_class.log('Joining FRF Computation')
    # frf_proc.join()
    # process_class.log('Joining Signal Generation')
    # siggen_proc.join()