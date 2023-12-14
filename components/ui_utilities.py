# -*- coding: utf-8 -*-
"""
User interface-specific utilities that might be used in multiple environments

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import os
from qtpy import QtWidgets, uic, QtGui
from qtpy.QtCore import Qt,QTimer
import numpy as np
import pyqtgraph
from scipy.io import loadmat
from scipy.interpolate import interp1d
import openpyxl
from typing import List

from .utilities import (coherence,error_message_qt,save_csv_matrix,
                        load_csv_matrix,trac,Channel,DataAcquisitionParameters)
from .environments import (ControlTypes,environment_long_names,
                           combined_environments_capable,control_select_ui_path,
                           environment_select_ui_path,environment_UIs,
                           transformation_matrices_ui_path,
                           modal_mdi_ui_path
                           )

ACQUISITION_FRAMES_TO_DISPLAY = 4

class ProfileTimer(QTimer):
    """A timer class that allows storage of controller instruction information"""
    def __init__(self,environment : str,operation : str,data : str):
        """
        A timer class that allows storage of controller instruction information
        
        When the timer times out, the environment, operation, and any data can
        be collected by the callback by accessing the self.sender().environment,
        .operation, or .data attributes.

        Parameters
        ----------
        environment : str
            The name of the environment (or 'Global') that the instruction will
            be sent to
        operation : str
            The operation that the environment will be instructed to perform
        data : str
            Any data corresponding to that operation that is required


        """
        super().__init__()
        self.environment = environment
        self.operation = operation
        self.data = data

def get_table_strings(tablewidget : QtWidgets.QTableWidget):
    """Collect a table of strings from a QTableWidget

    Parameters
    ----------
    tablewidget : QtWidgets.QTableWidget
        A table widget to pull the strings from

    Returns
    -------
    string_array : list[list[str]]
        A nested list of strings from the table items

    """
    string_array = []
    for row_idx in range(tablewidget.rowCount()):
        string_array.append([])
        for col_idx in range(tablewidget.columnCount()):
            value = tablewidget.item(row_idx,col_idx).text()
            string_array[-1].append(value)
    return string_array

def get_table_bools(tablewidget : QtWidgets.QTableWidget):
    """Collect a table of booleans from a QTableWidget full of QCheckBoxes

    Parameters
    ----------
    tablewidget : QtWidgets.QTableWidget
        A table widget to pull the strings from

    Returns
    -------
    bool_array : list[list[bool]]
        A nested list of booleans from the table widgets

    """
    bool_array = []
    for row_idx in range(tablewidget.rowCount()):
        bool_array.append([])
        for col_idx in range(tablewidget.columnCount()):
            value = tablewidget.cellWidget(row_idx,col_idx).isChecked()
            bool_array[-1].append(value)
    return bool_array

def load_time_history(signal_path,sample_rate):
    """Loads a time history from a given file
    
    The signal can be loaded from numpy files (.npz, .npy) or matlab files (.mat).
    For .mat and .npz files, the time data can be included in the file in the 
    't' field, or it can be excluded and the sample_rate input argument will
    be used.  If time data is specified, it will be linearly interpolated to the
    sample rate of the controller.
    For these file types, the signal should be stored in the 'signal'
    field.  For .npy files, only one array is stored, so it is treated as the
    signal, and the sample_rate input argument is used to construct the time
    data.

    Parameters
    ----------
    signal_path : str:
        Path to the file from which to load the time history
        
    sample_rate : str:
        The sample rate of the loaded signal.
        
    Returns
    -------
    signal : np.ndarray:
        A signal loaded from the file

    """
    file_base,extension = os.path.splitext(signal_path)
    if extension.lower() == '.npy':
        signal = np.load(signal_path)
    elif extension.lower() == '.npz':
        data = np.load(signal_path)
        signal = data['signal']
        try:
            times = data['t'].squeeze()
            fn = interp1d(times,signal)
            abscissa = np.arange(0,max(times)+1/sample_rate-1e-10,1/sample_rate)
            abscissa = abscissa[abscissa <= max(times)]
            signal = fn(abscissa)
        except KeyError:
            pass
    elif extension.lower() == '.mat':
        data = loadmat(signal_path)
        signal = data['signal']
        try:
            times = data['t'].squeeze()
            fn = interp1d(times,signal)
            abscissa = np.arange(0,max(times)+1/sample_rate-1e-10,1/sample_rate)
            abscissa = abscissa[abscissa <= max(times)]
            signal = fn(abscissa)
        except KeyError:
            pass
    else:
        raise ValueError("Could Not Determine the file type from the filename {:}: {:}".format(
            signal_path,extension))
    if signal.shape[-1] % 2 == 1:
        signal = signal[...,:-1]
    return signal

colororder = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
def multiline_plotter(x,y,widget = None,curve_list = None,names=None,other_pen_options = {'width':1},legend=False):
    """Helper function for PyQtGraph to deal with plots with multiple curves

    Parameters
    ----------
    x : np.ndarray
        Abscissa for the data that will be plotted, 1D array with shape n_samples
    y : np.ndarray
        Ordinates for the data that will be plotted.  2D array with shape
        n_curves x n_samples
    widget :
        The plot widget on which the curves will be drawn. (Default value = None)
    curve_list :
        Alternatively to specifying the widget, a curve list can be specified
        directly.  (Default value = None)
    names :
        Names of the curves that will appear in the legend. (Default value = None)
    other_pen_options : dict
        Additional options besides color that will be applied to the curves.
        (Default value = {'width':1})
    legend :
        Whether or not to draw a legend (Default value = False)

    Returns
    -------

    """
    if not widget is None:
        plot_item = widget.getPlotItem()
        if legend:
            plot_item.addLegend(colCount = len(y)//10)
        handles = []
        for i,this_y in enumerate(y):
            pen = {'color':colororder[i%len(colororder)]}
            pen.update(other_pen_options)
            handles.append(plot_item.plot(x,this_y,pen=pen,name=None if names is None else names[i]))
        return handles
    elif not curve_list is None:
        for this_y,curve in zip(y,curve_list):
            curve.setData(x,y)
        return curve_list
    else:
        raise ValueError('Either Widget or list of curves must be specified')

def save_combined_environments_profile_template(filename,environment_data):
    # Create the header
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Channel Table"
    hardware_worksheet = workbook.create_sheet('Hardware')
    # Create the header
    worksheet.cell(row=1,column=2,value='Test Article Definition')
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    worksheet.cell(row=1,column=5,value='Instrument Definition')
    worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=11)
    worksheet.cell(row=1,column=12,value='Channel Definition')
    worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=19)
    worksheet.cell(row=1,column=20,value='Output Feedback')
    worksheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=21)
    worksheet.cell(row=1,column=22,value='Limits')
    worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=23)
    for col_idx,val in enumerate(['Channel Index',
                                  'Node Number',
                                  'Node Direction',
                                  'Comment',
                                  'Serial Number',
                                  'Triax DoF',
                                  'Sensitivity  (mV/EU)',
                                  'Engineering Unit',
                                  'Make',
                                  'Model',
                                  'Calibration Exp Date',
                                  'Physical Device',
                                  'Physical Channel',
                                  'Type',
                                  'Minimum Value (V)',
                                  'Maximum Value (V)',
                                  'Coupling',
                                  'Current Excitation Source',
                                  'Current Excitation Value',
                                  'Physical Device',
                                  'Physical Channel',
                                  'Warning Level (EU)',
                                  'Abort Level (EU)']):
        worksheet.cell(row=2,column=1+col_idx,value=val)
    # Fill out the hardware worksheet
    hardware_worksheet.cell(1,1,'Hardware Type')
    hardware_worksheet.cell(1,2,'# Enter hardware index here')
    hardware_worksheet.cell(1,3,'Hardware Indices: 0 - NI DAQmx; 1 - LAN XI; 2 - Data Physics Quattro; 3 - Data Physics 900 Series; 4 - Exodus Modal Solution; 5 - State Space Integration; 6 - SDynPy System Integration')
    hardware_worksheet.cell(2,1,'Hardware File')
    hardware_worksheet.cell(2,2,'# Path to Hardware File (Depending on Hardware Device: 0 - Not Used; 1 - Not Used; 2 - Path to DpQuattro.dll library file; 3 - Not Used; 4 - Path to Exodus Eigensolution; 5 - Path to State Space File; 6 - Path to SDynPy system file)')
    hardware_worksheet.cell(3,1,'Sample Rate')
    hardware_worksheet.cell(3,2,'# Sample Rate of Data Acquisition System')
    hardware_worksheet.cell(4,1,'Time Per Read')
    hardware_worksheet.cell(4,2,'# Number of seconds per Read from the Data Acquisition System')
    hardware_worksheet.cell(5,1,'Time Per Write')
    hardware_worksheet.cell(5,2,'# Number of seconds per Write to the Data Acquisition System')
    hardware_worksheet.cell(6,1,'Acquisition Processes')
    hardware_worksheet.cell(6,2,'# Maximum Number of Acquisition Processes to start to pull data from hardware')
    hardware_worksheet.cell(7,1,'Integration Oversampling')
    hardware_worksheet.cell(7,2,'# For virual control, an integration oversampling can be specified')
    # Now do the environment
    worksheet.cell(row=1,column=24,value='Environments')
    for row,(value,name) in enumerate(environment_data):
        environment_UIs[value].create_environment_template(name, workbook)
        worksheet.cell(row=2,column=24+row,value=name)
    # Now create a profile page
    profile_sheet = workbook.create_sheet('Test Profile')
    profile_sheet.cell(1,1,'Time (s)')
    profile_sheet.cell(1,2,'Environment')
    profile_sheet.cell(1,3,'Operation')
    profile_sheet.cell(1,4,'Data')
    
    workbook.save(filename)

class EnvironmentSelect(QtWidgets.QDialog):
    """QDialog for selecting the environments in a combined environments run"""
    def __init__(self,parent=None):
        """
        Constructor for the EnvironmentSelect dialog box.

        Parameters
        ----------
        parent : QWidget, optional
            Parent widget to the dialog. The default is None.

        """
        super(QtWidgets.QDialog,self).__init__(parent)
        uic.loadUi(environment_select_ui_path, self)
        self.setWindowIcon(QtGui.QIcon('logo/Rattlesnake_Icon.png'))
        
        self.add_environment_button.clicked.connect(self.add_environment)
        self.remove_environment_button.clicked.connect(self.remove_environment)
        self.load_profile_button.clicked.connect(self.load_profile)
        self.save_profile_template_button.clicked.connect(self.save_profile_template)
        self.loaded_profile = None
        
    def add_environment(self):
        """Adds a row to the environment table"""
        selected_row = self.environment_display_table.rowCount()
        self.environment_display_table.insertRow(selected_row)
        combobox = QtWidgets.QComboBox()
        for control_type in combined_environments_capable:
            combobox.addItem(control_type.name.title(),control_type.value)
        self.environment_display_table.setCellWidget(selected_row,0,combobox)
    
    def remove_environment(self):
        """Removes a row from the environment table"""
        selected_row = self.environment_display_table.currentRow()
        if selected_row >= 0:
            self.environment_display_table.removeRow(selected_row)
    
    def save_profile_template(self):
        """Saves a template for the given environments table
        
        This template can be filled out by a user and then loaded."""
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self,'Save Combined Environments Template',filter='Excel File (*.xlsx)')
        if filename == '':
            return
        file_base,file_type = os.path.splitext(filename)
        # Now do the environments
        environment_data = []
        for row in range(self.environment_display_table.rowCount()):
            combobox = self.environment_display_table.cellWidget(row,0)
            value = ControlTypes(combobox.currentData())
            name = self.environment_display_table.item(row,1).text()
            environment_data.append((value,name))
        save_combined_environments_profile_template(filename, environment_data)
        
    def load_profile(self):
        """Loads a profile from an excel spreadsheet."""
        filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load Combined Environments Profile',filter='Excel File (*.xlsx)')
        if filename == '':
            return
        else:
            self.loaded_profile = filename
            self.accept()
    
    @staticmethod
    def select_environment(parent=None):
        """Creates the dialog box and then parses the output.
        
        Note that there are variable numbers of outputs for this function

        Parameters
        ----------
        parent : QWidget
            Parent to the dialog box (Default value = None)

        Returns
        -------
        result : int
            A flag specifying the outcome of the dialog box.  Will be 1 if the
            dialog was accepted, zero if cancelled, and -1 if a profile was
            loaded instead.
        environment_table : list of lists
            A list of environment type, environment name pairs that will be
            used to specify the environments in a test.
        loaded_profile : str
            File name to the profile file that needs to be loaded.  Only
            output if result == -1
        """
        dialog = EnvironmentSelect(parent)
        result = 1 if (dialog.exec_()==QtWidgets.QDialog.Accepted) else 0
        if dialog.loaded_profile is None:
            environment_table = []
            if result:
                for row in range(dialog.environment_display_table.rowCount()):
                    combobox = dialog.environment_display_table.cellWidget(row,0)
                    value = ControlTypes(combobox.currentData())
                    name = dialog.environment_display_table.item(row,1).text()
                    environment_table.append([value,name])
            # print(environment_table)
            return result,environment_table
        else:
            result = -1
            workbook = openpyxl.load_workbook(dialog.loaded_profile)
            environment_sheets = [sheet for sheet in workbook if (not sheet.title in ['Channel Table','Hardware','Test Profile']) and sheet.cell(1,1).value == 'Control Type']
            environment_table = [(ControlTypes[sheet.cell(1,2).value.upper()],sheet.title) for sheet in environment_sheets]
            workbook.close()
            return result,environment_table,dialog.loaded_profile
            
        
class ControlSelect(QtWidgets.QDialog):
    """Environment selector dialog box to select the control type for the test"""
    def __init__(self,parent=None):
        """
        Selects the environment type that gets used for the test.
        
        This function reads from the environment control types to populate the
        radiobuttons on the dialog.

        Parameters
        ----------
        parent : QWidget, optional
            Parent of the dialog box. The default is None.

        """
        super(QtWidgets.QDialog,self).__init__(parent)
        uic.loadUi(control_select_ui_path, self)
        self.setWindowIcon(QtGui.QIcon('logo/Rattlesnake_Icon.png'))
    
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        self.control_select_buttongroup = QtWidgets.QButtonGroup()
        
        # Go through and create radiobuttons for each control type
        control_types_sorted = sorted([(control_type.value,control_type) for control_type in ControlTypes])
        
        for value,control_type in control_types_sorted[1:]+control_types_sorted[:1]:
            radiobutton = QtWidgets.QRadioButton(environment_long_names[control_type])
            self.control_select_buttongroup.addButton(radiobutton,value)
            if value == ControlTypes.RANDOM.value:
                radiobutton.setChecked(True)
            self.environment_radiobutton_layout.addWidget(radiobutton)
    
    @staticmethod
    def select_control(parent=None):
        """Create the dialog box and parse the output

        Parameters
        ----------
        parent : QWidget
            Parent of the dialog box (Default value = None)

        Returns
        -------
        button_id : int
            The index of the button that was pressed
        result : bool
            True if dialog was accepted, otherwise false if cancelled.
        """
        dialog = ControlSelect(parent)
        result = dialog.exec_()==QtWidgets.QDialog.Accepted
        index = dialog.control_select_buttongroup.checkedId()
        button_id = ControlTypes(index)
        # print(button_id)
        return (button_id,result)


class PlotWindow(QtWidgets.QDialog):
    """Class defining a subwindow that displays specific channel information"""
    def __init__(self,parent,row,column,datatype,specification,row_name,column_name,datatype_name,
                 warning_matrix = None, abort_matrix = None):
        """
        Creates a window showing CPSD matrix information for a single channel.

        Parameters
        ----------
        parent : QWidget
            Parent of the window.
        row : int
            Row of the CPSD matrix to plot.
        column : int
            Column of the CPSD matrix to plot.
        datatype : int
            Type of data to plot: 0 - Magnitude, 1 - Coherence, 2 - Phase, 3 -
            Real, 4 - Imaginary.
        specification : np.ndarray
            The specification against which data will be compared.
        row_name : str
            Channel name for the row.
        column_name : str
            Channel name for the column.
        datatype_name : str
            Name for the datatype.


        """
        super(QtWidgets.QDialog,self).__init__(parent)
        self.setWindowFlags(self.windowFlags() & Qt.Tool)
        self.row = row
        self.column = column
        self.datatype = datatype
        self.frequencies = specification[0]
        self.spec_data = self.reduce_matrix(specification[1])
        self.data = np.zeros(self.spec_data.shape)
        # Now plot the data
        layout = QtWidgets.QVBoxLayout()
        plotwidget = pyqtgraph.PlotWidget()
        layout.addWidget(plotwidget)
        self.setLayout(layout)
        plot_item = plotwidget.getPlotItem()
        plot_item.showGrid(True,True,0.25)
        plot_item.enableAutoRange()
        plot_item.getViewBox().enableAutoRange(enable=True)
        if self.datatype==0:
            plot_item.setLogMode(False,True)
        plot_item.plot(self.frequencies,self.spec_data,pen = {'color': "b", 'width': 1})
        if not warning_matrix is None:
            plot_item.plot(self.frequencies,warning_matrix[0,:,row],pen = {'color': (255, 204, 0), 'width': 1, 'style':Qt.DashLine})
            plot_item.plot(self.frequencies,warning_matrix[1,:,row],pen = {'color': (255, 204, 0), 'width': 1, 'style':Qt.DashLine})
        if not abort_matrix is None:
            plot_item.plot(self.frequencies,abort_matrix[0,:,row],pen = {'color': (153, 0, 0), 'width': 1, 'style':Qt.DashLine})
            plot_item.plot(self.frequencies,abort_matrix[1,:,row],pen = {'color': (153, 0, 0), 'width': 1, 'style':Qt.DashLine})
        self.curve = plot_item.plot(self.frequencies,self.data,pen = {'color': "r", 'width': 1})
        self.setWindowTitle('{:} {:} / {:}'.format(datatype_name,row_name,column_name))
        self.show()
    
    def reduce_matrix(self,matrix):
        """Collects the data specific to the row and column and datatype

        Parameters
        ----------
        matrix : np.ndarray
            The 3D CPSD data that will be reduced

        Returns
        -------
        plot_data : np.ndarray
            The data that will be plotted

        """
        if self.datatype == 0: # Magnitude
            return np.abs(matrix[...,self.row,self.column])
        elif self.datatype == 1: # Coherence
            return coherence(matrix,(self.row,self.column))
        elif self.datatype == 2: # Phase
            return np.angle(matrix[...,self.row,self.column])    
        elif self.datatype == 3: # Real
            return np.real(matrix[...,self.row,self.column])     
        elif self.datatype == 4: # Imag
            return np.imag(matrix[...,self.row,self.column])
        else:
            raise ValueError('{:} is not a valid datatype!'.format(self.datatype))
        
    def update_plot(self,cpsd_matrix):
        """Updates the plot with the given CPSD matrix data

        Parameters
        ----------
        cpsd_matrix : np.ndarray
            3D CPSD matrix that will be reduced for plotting

        """
        self.curve.setData(self.frequencies,self.reduce_matrix(cpsd_matrix))
        
        
class PlotTimeWindow(QtWidgets.QDialog):
    """Class defining a subwindow that displays specific channel information"""
    def __init__(self,parent,index,specification,sample_rate,index_name):
        """
        Creates a window showing time history information for a single channel.

        Parameters
        ----------
        parent : QWidget
            Parent of the window.
        index : int
            Row of the time history matrix to plot
        specification : np.ndarray
            The specification against which data will be compared.
        sample_rate : int
            The sample rate of the time signal
        index_name : str
            Channel name for the row.
        """
        super(QtWidgets.QDialog,self).__init__(parent)
        self.setWindowFlags(self.windowFlags() & Qt.Tool)
        self.index = index
        self.times = np.arange(specification.shape[-1])/sample_rate
        self.spec_data = self.reduce_matrix(specification)
        self.data = np.zeros(self.spec_data.shape)
        # Now plot the data
        layout = QtWidgets.QVBoxLayout()
        plotwidget = pyqtgraph.PlotWidget()
        layout.addWidget(plotwidget)
        self.setLayout(layout)
        plot_item = plotwidget.getPlotItem()
        plot_item.showGrid(True,True,0.25)
        plot_item.enableAutoRange()
        plot_item.getViewBox().enableAutoRange(enable=True)
        plot_item.plot(self.times,self.spec_data,pen = {'color': "b", 'width': 1})
        plot_item.setLabel('left','TRAC: 0.0')
        self.plot_item = plot_item
        self.curve = plot_item.plot(self.times,self.data,pen = {'color': "r", 'width': 1})
        self.setWindowTitle('{:}'.format(index_name))
        self.show()
    
    def reduce_matrix(self,matrix):
        """Collects the data specific to the row and column and datatype

        Parameters
        ----------
        matrix : np.ndarray
            The 3D CPSD data that will be reduced

        Returns
        -------
        plot_data : np.ndarray
            The data that will be plotted

        """
        return matrix[self.index]
        
    def update_plot(self,data):
        """Updates the plot with the given CPSD matrix data

        Parameters
        ----------
        cpsd_matrix : np.ndarray
            3D CPSD matrix that will be reduced for plotting

        """
        data = self.reduce_matrix(data)
        self.curve.setData(self.times,data)
        self.plot_item.setLabel('left','TRAC: {:0.2f}'.format(trac(data,self.spec_data).squeeze()))
        
        
class TransformationMatrixWindow(QtWidgets.QDialog):
    """Dialog box for specifying transformation matrices"""
    def __init__(self,parent,current_response_transformation_matrix,num_responses,
                 current_output_transformation_matrix,num_outputs):
        """
        Creates a dialog box for specifying response and output transformations

        Parameters
        ----------
        parent : QWidget
            Parent to the dialog box.
        current_response_transformation_matrix : np.ndarray
            The current value of the transformation matrix that will be used to
            populate the entries in the table.
        num_responses : int
            Number of physical responses in the transformation.
        current_output_transformation_matrix : np.ndarray
            The current value of the transformation matrix that will be used to
            populate the entries in the table.
        num_outputs : int
            Number of physical outputs in the transformation.

        """
        super().__init__(parent)
        uic.loadUi(transformation_matrices_ui_path, self)
        
        self.response_transformation_matrix.setColumnCount(num_responses)
        self.output_transformation_matrix.setColumnCount(num_outputs)
        
        if current_response_transformation_matrix is None:
            self.set_response_transformation_identity()
        else:
            self.response_transformation_matrix.setRowCount(current_response_transformation_matrix.shape[0])
            for row_idx,row in enumerate(current_response_transformation_matrix):
                for col_idx,col in enumerate(row):
                    try:
                        self.response_transformation_matrix.item(row_idx,col_idx).setText(str(col))
                    except AttributeError:
                        item = QtWidgets.QTableWidgetItem(str(col))
                        self.response_transformation_matrix.setItem(row_idx,col_idx,item)
        if current_output_transformation_matrix is None:
            self.set_output_transformation_identity()
        else:
            self.output_transformation_matrix.setRowCount(current_output_transformation_matrix.shape[0])
            for row_idx,row in enumerate(current_output_transformation_matrix):
                for col_idx,col in enumerate(row):
                    try:
                        self.output_transformation_matrix.item(row_idx,col_idx).setText(str(col))
                    except AttributeError:
                        item = QtWidgets.QTableWidgetItem(str(col))
                        self.output_transformation_matrix.setItem(row_idx,col_idx,item)
            
        # Callbacks
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        
        self.response_transformation_add_row_button.clicked.connect(self.response_transformation_add_row)
        self.response_transformation_remove_row_button.clicked.connect(self.response_transformation_remove_row)
        self.response_transformation_save_matrix_button.clicked.connect(self.save_response_transformation_matrix)
        self.response_transformation_load_matrix_button.clicked.connect(self.load_response_transformation_matrix)
        self.response_transformation_identity_button.clicked.connect(self.set_response_transformation_identity)
        self.response_transformation_6dof_kinematic_button.clicked.connect(self.set_response_transformation_6dof)
        self.response_transformation_reversed_6dof_kinematic_button.clicked.connect(self.set_response_transformation_6dof_reversed)
        
        self.output_transformation_add_row_button.clicked.connect(self.output_transformation_add_row)
        self.output_transformation_remove_row_button.clicked.connect(self.output_transformation_remove_row)
        self.output_transformation_save_matrix_button.clicked.connect(self.save_output_transformation_matrix)
        self.output_transformation_load_matrix_button.clicked.connect(self.load_output_transformation_matrix)
        self.output_transformation_identity_button.clicked.connect(self.set_output_transformation_identity)
        self.output_transformation_6dof_kinematic_button.clicked.connect(self.set_output_transformation_6dof)
        self.output_transformation_reversed_6dof_kinematic_button.clicked.connect(self.set_output_transformation_6dof_reversed)
        
    @staticmethod
    def define_transformation_matrices(current_response_transformation_matrix,num_responses,
                 current_output_transformation_matrix,num_outputs,parent=None):
        """
        Shows the dialog and returns the transformation matrices
        
        Parameters
        ----------
        current_response_transformation_matrix : np.ndarray
            The current value of the transformation matrix that will be used to
            populate the entries in the table.
        num_responses : int
            Number of physical responses in the transformation.
        current_output_transformation_matrix : np.ndarray
            The current value of the transformation matrix that will be used to
            populate the entries in the table.
        num_outputs : int
            Number of physical outputs in the transformation.
        parent : QWidget
            Parent to the dialog box. (Default value = None)

        Returns
        -------
        response_transformation : np.ndarray
            Response transformation (or None if Identity)
        output_transformation : np.ndarray
            Output transformation (or None if Identity)
        result : bool
            True if dialog was accepted, false if cancelled.
        """
        dialog = TransformationMatrixWindow(parent,
                 current_response_transformation_matrix,num_responses,
                 current_output_transformation_matrix,num_outputs)
        result = dialog.exec_()==QtWidgets.QDialog.Accepted
        response_transformation = np.array([[float(val) for val in row] 
            for row in get_table_strings(dialog.response_transformation_matrix)])
        if all(val == response_transformation.shape[0] for val in response_transformation.shape) and np.allclose(response_transformation,np.eye(response_transformation.shape[0])):
            response_transformation = None
        output_transformation = np.array([[float(val) for val in row]
            for row in get_table_strings(dialog.output_transformation_matrix)])
        if all(val == output_transformation.shape[0] for val in output_transformation.shape) and np.allclose(output_transformation,np.eye(output_transformation.shape[0])):
            output_transformation = None
        return (response_transformation,output_transformation,result)

    def response_transformation_add_row(self):
        """Adds a row to the response transformation"""
        num_rows = self.response_transformation_matrix.rowCount()
        self.response_transformation_matrix.insertRow(num_rows)
        for col_idx in range(self.response_transformation_matrix.columnCount()):
            item = QtWidgets.QTableWidgetItem('0.0')
            self.response_transformation_matrix.setItem(num_rows,col_idx,item)
    def response_transformation_remove_row(self):
        """Removes a row from the response transformation"""
        num_rows = self.response_transformation_matrix.rowCount()
        self.response_transformation_matrix.removeRow(num_rows-1)
    def save_response_transformation_matrix(self):
        """Saves the response transformation matrix to a csv file"""
        string_array = self.get_table_strings(self.response_transformation_matrix)
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self,'Save Response Transformation',filter='Comma-separated Values (*.csv)')
        if filename == '':
            return
        save_csv_matrix(string_array,filename)
    def load_response_transformation_matrix(self):
        """Loads the response transformation from a csv file"""
        filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load Response Transformation',filter='Comma-separated values (*.csv *.txt);;Numpy Files (*.npy *.npz);;Matlab Files (*.mat)')
        if filename == '':
            return
        file_base,extension = os.path.splitext(filename)
        string_array = None
        if extension.lower() == '.npy':
            string_array = np.load(filename).astype('U')
        elif extension.lower() == '.npz':
            data = np.load(filename)
            for key,array in data.items():
                string_array = array.astype('U')
                break
        elif extension.lower() == '.mat':
            data = loadmat(filename)
            for key,array in data.items():
                if '__' in key:
                    continue
                string_array = array.astype('U')
                break
        else:
            string_array = load_csv_matrix(filename)
        if string_array is None:
            return
        # Set the number of rows
        self.response_transformation_matrix.setRowCount(len(string_array))
        num_rows = self.response_transformation_matrix.rowCount()
        num_cols = self.response_transformation_matrix.columnCount()
        for row_idx,row in enumerate(string_array):
            if row_idx == num_rows:
                break
            for col_idx,value in enumerate(row):
                if col_idx == num_cols:
                    break
                try:
                    self.response_transformation_matrix.item(row_idx,col_idx).setText(value)
                except AttributeError:
                    item = QtWidgets.QTableWidgetItem(value)
                    self.response_transformation_matrix.setItem(row_idx,col_idx,item)
    def set_response_transformation_identity(self):
        """Sets the response transformation to identity matrix (no transform)"""
        num_columns = self.response_transformation_matrix.columnCount()
        self.response_transformation_matrix.setRowCount(num_columns)
        for row_idx in range(num_columns):
            for col_idx in range(num_columns):
                if row_idx == col_idx:
                    value = 1.0
                else:
                    value = 0.0
                try:
                    self.response_transformation_matrix.item(row_idx,col_idx).setText(str(value))
                except AttributeError:
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.response_transformation_matrix.setItem(row_idx,col_idx,item)
    def set_response_transformation_6dof(self):
        """Sets the response transformation matrix to the 6DoF table"""
        num_columns = self.response_transformation_matrix.columnCount()
        if num_columns != 12:
            error_message_qt('Invalid Number of Control Channels.','Invalid Number of Control Channels.  6DoF Transform assumes 12 control accelerometer channels.')
            return
        self.response_transformation_matrix.setRowCount(6)
        matrix = [[0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0],
                   [ 0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0],
                   [ 0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25],
                   [ 0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,-0.25,0.0,0.0,-0.25],
                   [ 0.0,0.0,-0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,-0.25],
                   [ -0.125,0.125,0.0,-0.125,-0.125,0.0,0.125,-0.125,0.0,0.125,0.125,0.0]]
        for row_idx,row in enumerate(matrix):
            for col_idx,value in enumerate(row):
                try:
                    self.response_transformation_matrix.item(row_idx,col_idx).setText(str(value))
                except AttributeError:
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.response_transformation_matrix.setItem(row_idx,col_idx,item)           
    def set_response_transformation_6dof_reversed(self):
        """Sets the response transformation matrix to the 6DoF table"""
        num_columns = self.response_transformation_matrix.columnCount()
        if num_columns != 12:
            error_message_qt('Invalid Number of Control Channels.','Invalid Number of Control Channels.  6DoF Transform assumes 12 control accelerometer channels.')
            return  
        self.response_transformation_matrix.setRowCount(6)
        matrix = [[0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0],
                   [ 0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0],
                   [ 0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25],
                   [ 0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,-0.25,0.0,0.0,-0.25],
                   [ 0.0,0.0,-0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,-0.25],
                   [ -0.125,0.125,0.0,-0.125,-0.125,0.0,0.125,-0.125,0.0,0.125,0.125,0.0]]
        for row_idx,row in enumerate(matrix):
            for col_idx,value in enumerate(row):
                try:
                    self.response_transformation_matrix.item(row_idx,col_idx).setText(str(value))
                except AttributeError:
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.response_transformation_matrix.setItem(row_idx,col_idx,item)
        
    def output_transformation_add_row(self):
        """Adds a row to the output transformation"""
        num_rows = self.output_transformation_matrix.rowCount()
        self.output_transformation_matrix.insertRow(num_rows)
        for col_idx in range(self.output_transformation_matrix.columnCount()):
            item = QtWidgets.QTableWidgetItem('0.0')
            self.output_transformation_matrix.setItem(num_rows,col_idx,item)
    def output_transformation_remove_row(self):
        """Removes a row from the output tranformation"""
        num_rows = self.output_transformation_matrix.rowCount()
        self.output_transformation_matrix.removeRow(num_rows-1)
    def save_output_transformation_matrix(self):
        """Saves output transformation matrix to a CSV file"""
        string_array = self.get_table_strings(self.output_transformation_matrix)
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self,'Save Output Transformation',filter='Comma-separated Values (*.csv)')
        if filename == '':
            return
        save_csv_matrix(string_array,filename)
    def load_output_transformation_matrix(self):
        """Loads the output transformation from a CSV file"""
        filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load Output Transformation',filter='Comma-separated values (*.csv *.txt);;Numpy Files (*.npy *.npz);;Matlab Files (*.mat)')
        if filename == '':
            return
        file_base,extension = os.path.splitext(filename)
        string_array = None
        if extension.lower() == '.npy':
            string_array = np.load(filename).astype('U')
        elif extension.lower() == '.npz':
            data = np.load(filename)
            for key,array in data.items():
                string_array = array.astype('U')
                break
        elif extension.lower() == '.mat':
            data = loadmat(filename)
            for key,array in data.items():
                if '__' in key:
                    continue
                string_array = array.astype('U')
                break
        else:
            string_array = load_csv_matrix(filename)
        if string_array is None:
            return
        # Set the number of rows
        self.output_transformation_matrix.setRowCount(len(string_array))
        num_rows = self.output_transformation_matrix.rowCount()
        num_cols = self.output_transformation_matrix.columnCount()
        for row_idx,row in enumerate(string_array):
            if row_idx == num_rows:
                break
            for col_idx,value in enumerate(row):
                if col_idx == num_cols:
                    break
                try:
                    self.output_transformation_matrix.item(row_idx,col_idx).setText(value)
                except AttributeError:
                    item = QtWidgets.QTableWidgetItem(value)
                    self.output_transformation_matrix.setItem(row_idx,col_idx,item)
    def set_output_transformation_identity(self):
        """Sets the output transformation to identity (no transform)"""
        num_columns = self.output_transformation_matrix.columnCount()
        self.output_transformation_matrix.setRowCount(num_columns)
        for row_idx in range(num_columns):
            for col_idx in range(num_columns):
                if row_idx == col_idx:
                    value = 1.0
                else:
                    value = 0.0
                try:
                    self.output_transformation_matrix.item(row_idx,col_idx).setText(str(value))
                except AttributeError:
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.output_transformation_matrix.setItem(row_idx,col_idx,item)
    def set_output_transformation_6dof(self):
        """Sets the output transformation matrix to the 6DoF table"""
        num_columns = self.output_transformation_matrix.columnCount()
        if num_columns != 12:
            error_message_qt('Invalid Number of Output Signals.','Invalid Number of Output Signals.  6DoF Transform assumes 12 drive channels.')
            return  
        self.output_transformation_matrix.setRowCount(6)
        matrix = [[0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0],
                   [ 0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0],
                   [ 0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25],
                   [ 0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,-0.25,0.0,0.0,-0.25],
                   [ 0.0,0.0,-0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,-0.25],
                   [ -0.125,0.125,0.0,-0.125,-0.125,0.0,0.125,-0.125,0.0,0.125,0.125,0.0]]
        for row_idx,row in enumerate(matrix):
            for col_idx,value in enumerate(row):
                try:
                    self.output_transformation_matrix.item(row_idx,col_idx).setText(str(value))
                except AttributeError:
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.output_transformation_matrix.setItem(row_idx,col_idx,item)
                    
    def set_output_transformation_6dof_reversed(self):
        """Sets the output transformation matrix to the 6DoF table"""
        num_columns = self.output_transformation_matrix.columnCount()
        if num_columns != 12:
            error_message_qt('Invalid Number of Output Signals.','Invalid Number of Output Signals.  6DoF Transform assumes 12 drive channels.')
            return
        self.output_transformation_matrix.setRowCount(6)
        matrix = [[0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0],
                   [ 0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0],
                   [ 0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,0.25],
                   [ 0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,-0.25,0.0,0.0,-0.25],
                   [ 0.0,0.0,-0.25,0.0,0.0,0.25,0.0,0.0,0.25,0.0,0.0,-0.25],
                   [ -0.125,0.125,0.0,-0.125,-0.125,0.0,0.125,-0.125,0.0,0.125,0.125,0.0]]
        for row_idx,row in enumerate(matrix):
            for col_idx,value in enumerate(row):
                try:
                    self.output_transformation_matrix.item(row_idx,col_idx).setText(str(value))
                except AttributeError:
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.output_transformation_matrix.setItem(row_idx,col_idx,item)
                    
class ModalMDISubWindow(QtWidgets.QWidget):
    def __init__(self,parent):
        super().__init__(parent)
        uic.loadUi(modal_mdi_ui_path, self)
        
        self.parent = parent
        self.channel_names = self.parent.channel_names
        self.reference_names = np.array([self.parent.channel_names[i] for i in self.parent.reference_channel_indices])
        self.response_names = np.array([self.parent.channel_names[i] for i in self.parent.response_channel_indices])
        self.reciprocal_responses = self.parent.reciprocal_responses
        
        self.signal_selector.currentIndexChanged.connect(self.update_ui)
        self.data_type_selector.currentIndexChanged.connect(self.update_ui_no_clear)
        self.response_coordinate_selector.currentIndexChanged.connect(self.update_data)
        self.reference_coordinate_selector.currentIndexChanged.connect(self.update_data)
        
        self.primary_plotitem = self.primary_plot.getPlotItem()
        self.secondary_plotitem = self.secondary_plot.getPlotItem()
        self.primary_viewbox = self.primary_plotitem.getViewBox()
        self.secondary_viewbox = self.secondary_plotitem.getViewBox()
        self.primary_axis = self.primary_plotitem.getAxis('left')
        self.secondary_axis = self.secondary_plotitem.getAxis('left')
        
        self.secondary_plotitem.setXLink(self.primary_plotitem)
        
        self.primary_plotdataitem = pyqtgraph.PlotDataItem(
            np.arange(2),np.zeros(2),pen = {'color': "r", 'width': 1})
        self.secondary_plotdataitem = pyqtgraph.PlotDataItem(
            np.arange(2),np.zeros(2),pen = {'color': "r", 'width': 1})
        
        self.primary_viewbox.addItem(self.primary_plotdataitem)
        self.secondary_viewbox.addItem(self.secondary_plotdataitem)
        
        self.twinx_viewbox = None
        self.twinx_axis = None
        self.twinx_original_plotitem = None
        self.twinx_plotdataitem = None
        
        self.is_comparing = False
        self.primary_plotdataitem_compare = pyqtgraph.PlotDataItem(
            np.arange(2),np.zeros(2),pen = {'color': "b", 'width': 1})
        self.secondary_plotdataitem_compare = pyqtgraph.PlotDataItem(
            np.arange(2),np.zeros(2),pen = {'color': "b", 'width': 1})
        
        self.update_ui()
        
    def remove_twinx(self):
        if self.twinx_viewbox is None:
            return
        self.twinx_original_plotitem.layout.removeItem(self.twinx_axis)
        self.twinx_original_plotitem.scene().removeItem(self.twinx_viewbox)
        self.twinx_original_plotitem.scene().removeItem(self.twinx_axis)
        self.twinx_viewbox = None
        self.twinx_axis = None
        self.twinx_original_plotitem = None
        self.twinx_plot_item = None
    
    def add_twinx(self, existing_plot_item : pyqtgraph.PlotItem):
        # Create a viewbox
        self.twinx_original_plotitem = existing_plot_item
        self.twinx_viewbox = pyqtgraph.ViewBox()
        self.twinx_original_plotitem.scene().addItem(self.twinx_viewbox)
        self.twinx_axis = pyqtgraph.AxisItem('right')
        self.twinx_axis.setLogMode(False)
        self.twinx_axis.linkToView(self.twinx_viewbox)
        self.twinx_original_plotitem.layout.addItem(self.twinx_axis,2,3)
        self.updateTwinXViews()
        self.twinx_viewbox.setXLink(self.twinx_original_plotitem)
        self.twinx_original_plotitem.vb.sigResized.connect(self.updateTwinXViews)
        self.twinx_plotdataitem = pyqtgraph.PlotDataItem(
            np.arange(2),np.zeros(2),pen = {'color': "b", 'width': 1})
        self.twinx_viewbox.addItem(self.twinx_plotdataitem)
    
    def add_compare(self):
        self.is_comparing = True
        self.primary_viewbox.addItem(self.primary_plotdataitem_compare)
        self.secondary_viewbox.addItem(self.secondary_plotdataitem_compare)
        
    def remove_compare(self):
        if self.is_comparing:
            self.primary_viewbox.removeItem(self.primary_plotdataitem_compare)
            self.secondary_viewbox.removeItem(self.secondary_plotdataitem_compare)
            self.is_comparing = False
    
    def updateTwinXViews(self):
        if self.twinx_viewbox is None:
            return
        self.twinx_viewbox.setGeometry(self.twinx_original_plotitem.vb.sceneBoundingRect())
        # self.twinx_viewbox.linkedViewChanged(self.twinx_original_plotitem.vb, self.twinx_viewbox.XAxis)
        
    def update_ui_no_clear(self):
        self.update_ui(False)
        
    def update_ui(self,clear_channels=True):
        self.response_coordinate_selector.blockSignals(True)
        self.reference_coordinate_selector.blockSignals(True)
        self.remove_twinx()
        self.remove_compare()
        if self.signal_selector.currentIndex() in [0,1,2,3]: # Time or Windowed Time or Spectrum or Autospectrum
            self.reference_coordinate_selector.hide()
            self.data_type_selector.hide()
            self.secondary_plot.hide()
            if clear_channels:
                self.response_coordinate_selector.clear()
                self.reference_coordinate_selector.clear()
                for channel_name in self.channel_names:
                    self.response_coordinate_selector.addItem(channel_name)
            if self.signal_selector.currentIndex() in [0,1]:
                self.primary_axis.setLogMode(False)
                self.primary_plotdataitem.setLogMode(False,False)
            else:
                self.primary_axis.setLogMode(True)
                self.primary_plotdataitem.setLogMode(False,True)
        elif self.signal_selector.currentIndex() in [4,6,7]: # FRF or FRF Coherence or Reciprocity
            self.reference_coordinate_selector.show()
            self.data_type_selector.show()
            if self.data_type_selector.currentIndex() in [1,4]:
                self.secondary_plot.show()
                if self.signal_selector.currentIndex() == 6:
                    self.add_twinx(self.secondary_plotitem)
            else:
                self.secondary_plot.hide()
                if self.signal_selector.currentIndex() == 6:
                    self.add_twinx(self.primary_plotitem)
            if self.signal_selector.currentIndex() == 7:
                if any([val is None for val in self.reciprocal_responses]):
                    error_message_qt('Invalid Reciprocal Channels', 'Could not deterimine reciprocal channels for this test')
                    self.signal_selector.setCurrentIndex(4)
                    return
                self.add_compare()
            if clear_channels:
                self.response_coordinate_selector.clear()
                self.reference_coordinate_selector.clear()
                if self.signal_selector.currentIndex() == 7:
                    for channel_name in self.response_names[self.reciprocal_responses]:
                        self.response_coordinate_selector.addItem(channel_name)
                else:
                    for channel_name in self.response_names:
                        self.response_coordinate_selector.addItem(channel_name)
                for channel_name in self.reference_names:
                    self.reference_coordinate_selector.addItem(channel_name)
            if self.data_type_selector.currentIndex() == 0:
                self.primary_axis.setLogMode(True)
                self.primary_plotdataitem.setLogMode(False,True)
                self.primary_plotdataitem_compare.setLogMode(False,True)
            elif self.data_type_selector.currentIndex() == 1:
                self.primary_axis.setLogMode(False)
                self.primary_plotdataitem.setLogMode(False,False)
                self.primary_plotdataitem_compare.setLogMode(False,False)
                self.secondary_axis.setLogMode(True)
                self.secondary_plotdataitem.setLogMode(False,True)
                self.secondary_plotdataitem_compare.setLogMode(False,True)
            elif self.data_type_selector.currentIndex() in [2,3]:
                self.primary_axis.setLogMode(False)
                self.primary_plotdataitem.setLogMode(False,False)
                self.primary_plotdataitem_compare.setLogMode(False,False)
            elif self.data_type_selector.currentIndex() == 4:
                self.primary_axis.setLogMode(False)
                self.primary_plotdataitem.setLogMode(False,False)
                self.primary_plotdataitem_compare.setLogMode(False,False)
                self.secondary_axis.setLogMode(False)
                self.secondary_plotdataitem.setLogMode(False,False)
                self.secondary_plotdataitem_compare.setLogMode(False,False)
            if self.signal_selector.currentIndex() == 6:
                self.twinx_axis.setLogMode(False)
                self.twinx_plotdataitem.setLogMode(False,False)
        elif self.signal_selector.currentIndex() in [5]: # Coherence
            self.reference_coordinate_selector.hide()
            self.data_type_selector.hide()
            self.secondary_plot.hide()
            if clear_channels:
                self.response_coordinate_selector.clear()
                self.reference_coordinate_selector.clear()
                for channel_name in self.response_names:
                    self.response_coordinate_selector.addItem(channel_name)
            self.primary_axis.setLogMode(False)
            self.primary_plotdataitem.setLogMode(False,False)
        self.update_data()
        self.response_coordinate_selector.blockSignals(False)
        self.reference_coordinate_selector.blockSignals(False)
            
    def set_window_title(self):
        self.setWindowTitle('{:} {:} {:}'.format(
            self.signal_selector.itemText(self.signal_selector.currentIndex()),
            self.response_coordinate_selector.itemText(self.response_coordinate_selector.currentIndex()),
            self.reference_coordinate_selector.itemText(self.reference_coordinate_selector.currentIndex()) if self.signal_selector.currentIndex() == 4 else ''
            ))
            
    def update_data(self):
        self.set_window_title()
        current_index = self.signal_selector.currentIndex()
        if current_index in [0,1]: # Time history
            if self.parent.last_frame is None:
                return
            data = self.parent.last_frame[self.response_coordinate_selector.currentIndex()]
            if current_index == 1:
                data = data * self.parent.window_function
            self.primary_plotdataitem.setData(self.parent.time_abscissa,data)
        elif current_index == 2: # Spectrum
            if self.parent.last_spectrum is None:
                return
            data = self.parent.last_spectrum[self.response_coordinate_selector.currentIndex()]
            self.primary_plotdataitem.setData(self.parent.frequency_abscissa,data)
        elif current_index == 3: # Autospectrum
            if self.parent.last_autospectrum is None:
                return
            data = self.parent.last_autospectrum[self.response_coordinate_selector.currentIndex()]
            self.primary_plotdataitem.setData(self.parent.frequency_abscissa,data)
        elif current_index == 4 or current_index == 6: # FRF or FRF Coherence
            if self.parent.last_frf is None:
                return
            data = self.parent.last_frf[:,self.response_coordinate_selector.currentIndex(),self.reference_coordinate_selector.currentIndex()]
            if self.data_type_selector.currentIndex() == 0: # Magnitude
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.abs(data))
            elif self.data_type_selector.currentIndex() == 1: # Magnitude/Phase
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.angle(data))
                self.secondary_plotdataitem.setData(self.parent.frequency_abscissa,np.abs(data))
            elif self.data_type_selector.currentIndex() == 2: # Real
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.real(data))
            elif self.data_type_selector.currentIndex() == 3: # Imag
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.imag(data))
            elif self.data_type_selector.currentIndex() == 4: # Real/Imag
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.real(data))
                self.secondary_plotdataitem.setData(self.parent.frequency_abscissa,np.imag(data))
            if current_index == 6:
                data = self.parent.last_coh[self.response_coordinate_selector.currentIndex()]
                self.twinx_plotdataitem.setData(self.parent.frequency_abscissa,data)
        elif current_index == 5: # Coherence
            if self.parent.last_coh is None:
                return
            data = self.parent.last_coh[self.response_coordinate_selector.currentIndex()]
            self.primary_plotdataitem.setData(self.parent.frequency_abscissa,data)
        elif current_index == 7: # FRF or FRF Coherence
            if self.parent.last_frf is None:
                return
            resp_ind = self.response_coordinate_selector.currentIndex()
            ref_ind = self.reference_coordinate_selector.currentIndex()
            data = self.parent.last_frf[:,self.reciprocal_responses[resp_ind],ref_ind]
            compare_data = self.parent.last_frf[:,self.reciprocal_responses[ref_ind],resp_ind]
            if self.data_type_selector.currentIndex() == 0: # Magnitude
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.abs(data))
                self.primary_plotdataitem_compare.setData(self.parent.frequency_abscissa,np.abs(compare_data))
            elif self.data_type_selector.currentIndex() == 1: # Magnitude/Phase
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.angle(data))
                self.secondary_plotdataitem.setData(self.parent.frequency_abscissa,np.abs(data))
                self.primary_plotdataitem_compare.setData(self.parent.frequency_abscissa,np.angle(compare_data))
                self.secondary_plotdataitem_compare.setData(self.parent.frequency_abscissa,np.abs(compare_data))
            elif self.data_type_selector.currentIndex() == 2: # Real
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.real(data))
                self.primary_plotdataitem_compare.setData(self.parent.frequency_abscissa,np.real(compare_data))
            elif self.data_type_selector.currentIndex() == 3: # Imag
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.imag(data))
                self.primary_plotdataitem_compare.setData(self.parent.frequency_abscissa,np.imag(compare_data))
            elif self.data_type_selector.currentIndex() == 4: # Real/Imag
                self.primary_plotdataitem.setData(self.parent.frequency_abscissa,np.real(data))
                self.secondary_plotdataitem.setData(self.parent.frequency_abscissa,np.imag(data))
                self.primary_plotdataitem_compare.setData(self.parent.frequency_abscissa,np.real(compare_data))
                self.secondary_plotdataitem_compare.setData(self.parent.frequency_abscissa,np.imag(compare_data))
    
    def increment_channel(self,increment=1):
        if not self.lock_response_checkbox.isChecked():
            num_channels = self.response_coordinate_selector.count()
            current_index = self.response_coordinate_selector.currentIndex()
            new_index = (current_index + increment) % num_channels
            self.response_coordinate_selector.setCurrentIndex(new_index)
            
class ChannelMonitor(QtWidgets.QDialog):
    """Class defining a subwindow that displays specific channel information"""
    
    def __init__(self,parent,daq_settings : DataAcquisitionParameters):
        """
        Creates a window showing CPSD matrix information for a single channel.

        Parameters
        ----------
        parent : QWidget
            Parent of the window.
        """
        super(QtWidgets.QDialog,self).__init__(parent)
        self.setWindowFlags(self.windowFlags() & Qt.Tool)
        self.channels = daq_settings.channel_list
        # Set up the window
        self.graphics_layout_widget = pyqtgraph.GraphicsLayoutWidget(self)
        self.push_button = QtWidgets.QPushButton('Clear Alerts',self)
        self.channels_per_row_label = QtWidgets.QLabel('Channels per Row: ',self)
        self.channels_per_row_selector = QtWidgets.QSpinBox(self)
        self.channels_per_row_selector.setMinimum(2)
        self.channels_per_row_selector.setMaximum(100)
        self.channels_per_row_selector.setValue(20)
        self.channels_per_row_selector.setKeyboardTracking(False)
        layout = QtWidgets.QVBoxLayout()
        control_layout = QtWidgets.QHBoxLayout()
        layout.addWidget(self.graphics_layout_widget)
        control_layout.addWidget(self.channels_per_row_label)
        control_layout.addWidget(self.channels_per_row_selector)
        control_layout.addStretch()
        control_layout.addWidget(self.push_button)
        layout.addLayout(control_layout)
        self.setLayout(layout)
        # Set up defaults for the channel ranges
        self.channel_ranges = None
        self.channel_warning_limits = None
        self.channel_abort_limits = None
        self.background_bars = None
        self.history_bars = None
        self.level_bars = None
        self.history_last_update = None
        self.history_hold_frames = int(np.ceil(10*daq_settings.sample_rate/daq_settings.samples_per_read))
        self.aborted_channels = None
        # Set up defaults for the plot
        self.plots = None
        self.bar_channel_indices = None
        self.pen = pyqtgraph.mkPen(color=(0,0,0,255),width=1)
        self.background_brush = pyqtgraph.mkBrush((255,255,255))
        self.history_brush = pyqtgraph.mkBrush((124,124,255))
        self.current_brush = pyqtgraph.mkBrush((34,139,34))
        self.limit_brush = pyqtgraph.mkBrush((145,197,17))
        self.abort_brush = pyqtgraph.mkBrush((145,70,17))
        self.limit_background_brush = pyqtgraph.mkBrush((255,255,0,))
        self.abort_background_brush = pyqtgraph.mkBrush((255,0,0))
        self.limit_history_brush = pyqtgraph.mkBrush((190,190,128))
        self.abort_history_brush = pyqtgraph.mkBrush((190,62,128))
        # Connect everything and do final builds
        self.connect_callbacks()
        self.build_plot()
        self.setWindowTitle('Channel Monitor')
        self.resize(400,300)
        self.show()

    def connect_callbacks(self):
        self.channels_per_row_selector.valueChanged.connect(self.build_plot)
        self.push_button.clicked.connect(self.clear_alerts)
    
    def update_channel_list(self,daq_settings):
        self.channels = daq_settings.channel_list
        self.history_hold_frames = int(np.ceil(10*daq_settings.sample_rate/daq_settings.samples_per_read))
        self.build_plots()
    
    def clear_alerts(self):
        self.aborted_channels = [False for val in self.aborted_channels]
        for current_bar in self.level_bars:
            current_bar.setOpts(brushes=[self.current_brush])
        for history_bar in self.history_bars:
            history_bar.setOpts(brushes=[self.history_brush])
        for background_bar in self.background_bars:
            background_bar.setOpts(brushes=[self.background_brush])
    
    def build_plot(self):
        #TODO Need to get the values from the bars before deleting them so we
        # can maintain the levels from before the value was changed
        self.graphics_layout_widget.clear()
        num_channels = len(self.channels)
        num_bars = int(np.ceil(num_channels/self.channels_per_row_selector.value()))
        # Compute number of channels per bar
        channels_per_bar = [0 for i in range(num_bars)]
        for i in range(num_channels):
            channels_per_bar[i%num_bars] += 1
        
        # print('Channels per Bar {:}'.format(channels_per_bar))
        # Now let's actually make the plots
        self.plots = [self.graphics_layout_widget.addPlot(i,0) for i in range(num_bars)]
        
        # Now parse the channel ranges
        self.channel_ranges = []
        self.channel_warning_limits = []
        self.channel_abort_limits = []
        for channel in self.channels:
            try:
                max_abs_volt = np.min(np.abs([float(channel.maximum_value),float(channel.minimum_value)]))
            except (ValueError,TypeError):
                max_abs_volt = 10 # Assume 10 V range on DAQ
            try:
                sensitivity = float(channel.sensitivity)/1000 #mV -> V
            except (ValueError,TypeError):
                sensitivity = 0.01 # Assume 10 mV/EU
            max_abs_eu = max_abs_volt/sensitivity
            try:
                warning_limit = float(channel.warning_level)
            except (ValueError,TypeError):
                warning_limit = max_abs_eu*0.9 # Put out warning at 90% the max range
            try:
                abort_limit = float(channel.abort_level)
            except (ValueError,TypeError):
                abort_limit = max_abs_eu # Never abort on this channel if not specified
            self.channel_ranges.append(max_abs_eu)
            self.channel_warning_limits.append(warning_limit)
            self.channel_abort_limits.append(abort_limit)
        self.channel_ranges = np.array(self.channel_ranges)
        self.channel_warning_limits = np.array(self.channel_warning_limits)
        self.channel_abort_limits = np.array(self.channel_abort_limits)
        # Display abort limit as range rather than channel if it is lower
        abort_lower = self.channel_ranges > self.channel_abort_limits
        self.channel_ranges[abort_lower] = self.channel_abort_limits[abort_lower]
        
        # Now build the plots
        self.bar_channel_indices = []
        for i,num_channels in enumerate(channels_per_bar):
            try:
                next_starting_index = self.bar_channel_indices[-1][-1]+1
            except IndexError:
                next_starting_index = 0
            self.bar_channel_indices.append(next_starting_index+np.arange(num_channels))
        # print(self.bar_channel_indices)
        self.background_bars = []
        self.history_bars = []
        self.level_bars = []
        self.history_last_update = []
        self.aborted_channels = []
        for indices,plot in zip(self.bar_channel_indices,self.plots):
            plot.hideAxis('left')
            for j,index in enumerate(indices):
                background_bar = pyqtgraph.BarGraphItem(x=[index+1],height=1.0,width=0.9,pen=self.pen,brush=self.background_brush)
                plot.addItem(background_bar)
                self.background_bars.append(background_bar)
                history_bar = pyqtgraph.BarGraphItem(x=[index+1],height=0,width=0.9,pen=self.pen,brush=self.history_brush)
                plot.addItem(history_bar)
                self.history_bars.append(history_bar)
                current_bar = pyqtgraph.BarGraphItem(x=[index+1],height=0,width=0.9,pen=self.pen,brush=self.current_brush)
                plot.addItem(current_bar)
                self.level_bars.append(current_bar)
                self.history_last_update.append(0)
                self.aborted_channels.append(False)
                
    def update(self,channel_levels):
        # print('Data {:}'.format(channel_levels.shape))
        # print(channel_levels)
        for index,(level,current_bar,history_bar,background_bar,history_last_update,warning,abort,range,aborted) in enumerate(zip(
                channel_levels,self.level_bars,self.history_bars,self.background_bars,self.history_last_update,
                self.channel_warning_limits,self.channel_abort_limits,self.channel_ranges,self.aborted_channels)):
            # Set the current bar height
            current_height = level/range
            current_bar.setOpts(height=current_height if current_height < 1 else 1)
            # Now look at the history bar
            last_history_height = history_bar.opts.get('height')
            # print(last_history_height)
            if history_last_update > self.history_hold_frames:
                desired_history_height = last_history_height - 1/self.history_hold_frames
            else:
                desired_history_height = last_history_height
            if desired_history_height < current_height:
                desired_history_height = current_height
                self.history_last_update[index] = 0
            else:
                self.history_last_update[index] += 1
            history_bar.setOpts(height=1 if desired_history_height > 1 else desired_history_height)
            # Now look at the pen color
            if level > abort or aborted:
                current_bar.setOpts(brushes=[self.abort_brush])
                background_bar.setOpts(brushes=[self.abort_background_brush])
                history_bar.setOpts(brushes=[self.abort_history_brush])
                self.aborted_channels[index] = True
            elif level > warning:
                current_bar.setOpts(brushes=[self.limit_brush])
                background_bar.setOpts(brushes=[self.limit_background_brush])
                history_bar.setOpts(brushes=[self.limit_history_brush])