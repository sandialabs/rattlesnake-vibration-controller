import multiprocessing as mp
import queue as thqueue
from unittest import mock

import netCDF4 as nc4
import numpy as np
import openpyxl
import pytest

from rattlesnake.environment.abstract_environment import (
    Environment,
    EnvironmentInstructions,
    EnvironmentMetadata,
)
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.environment.time_environment import (
    CONTROL_TYPE,
    TEST_LEVEL_THRESHOLD,
    TimeCommands,
    TimeEnvironment,
    TimeInstructions,
    TimeMetadata,
    TimeQueues,
    TimeUICommands,
    time_process,
)
from rattlesnake.testing.mock_utilities import (
    mock_channel_list_bools,
    mock_event_container,
    mock_queue_container,
    skeleton_hardware_metadata,
)
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.utilities import GlobalCommands, RattlesnakeError, VerboseMessageQueue


# region Fixtures
@pytest.fixture
def hardware_metadata():
    return skeleton_hardware_metadata()


@pytest.fixture
def time_metadata():
    metadata = TimeMetadata(
        environment_name="Time Environment",
        channel_list_bools=mock_channel_list_bools(),
        sample_rate=1000,
        output_oversample=10,
        output_signal=np.ones((1, 2000)),
        cancel_rampdown_time=0.1,
    )
    metadata.queue_name = "Environment 0"
    return metadata


@pytest.fixture(params=[True, False], ids=["threaded", "non_threaded"])
def time_environment(request):
    use_thread = request.param
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)

    time_queues = TimeQueues(
        queue_container.environment_command_queues["Environment 0"],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues["Environment 0"],
        queue_container.environment_data_out_queues["Environment 0"],
        queue_container.log_file_queue,
    )

    return TimeEnvironment(
        "Time Environment",
        "Environment 0",
        time_queues,
        event_container.acquisition_active_event,
        event_container.output_active_event,
        event_container.environment_active_events["Environment 0"],
        event_container.environment_ready_events["Environment 0"],
    )


# endregion


# region Time Commands
def test_time_commands_have_expected_values():
    """
    Verifies that time environment commands have expected integer values.
    """
    assert TimeCommands.SET_TEST_LEVEL.value == 0
    assert TimeCommands.SET_REPEAT.value == 1
    assert TimeCommands.SET_NO_REPEAT.value == 2


def test_time_commands_valid_profile_commands():
    """
    Verifies that all time commands are valid profile commands.
    """
    assert TimeCommands.valid_profile_commands() == (
        TimeCommands.SET_TEST_LEVEL,
        TimeCommands.SET_REPEAT,
        TimeCommands.SET_NO_REPEAT,
    )


def test_time_commands_valid_data():
    """
    Verifies that time command data requirements are defined correctly.
    """
    assert TimeCommands.valid_data() == {
        TimeCommands.SET_TEST_LEVEL: int,
        TimeCommands.SET_REPEAT: type(None),
        TimeCommands.SET_NO_REPEAT: type(None),
    }


def test_time_commands_label():
    """
    Verifies that command labels are user-friendly title-case strings.
    """
    assert TimeCommands.SET_TEST_LEVEL.label == "Set Test Level"


def test_time_ui_commands():
    """
    Verifies that UI command values are unique integers.
    """
    values = [command.value for command in TimeUICommands]

    assert all(isinstance(value, int) for value in values)
    assert len(values) == len(set(values))
    assert TimeUICommands.TIME_DATA.value == 0


# endregion


# region Time Metadata
def test_time_metadata_init():
    """
    Verifies that ``TimeMetadata`` initializes and is an
    ``EnvironmentMetadata``.
    """
    metadata = TimeMetadata("Time Environment")

    assert isinstance(metadata, TimeMetadata)
    assert isinstance(metadata, EnvironmentMetadata)
    assert metadata.environment_type == CONTROL_TYPE
    assert metadata.environment_name == "Time Environment"
    assert metadata.output_oversample is None
    assert metadata.output_signal is None
    assert metadata.cancel_rampdown_time is None
    assert metadata.signal_file is None


def test_time_metadata_properties(time_metadata):
    """
    Verifies derived time metadata properties.
    """
    assert time_metadata.signal_samples == 2000
    assert time_metadata.output_channels == 1
    assert time_metadata.signal_time == 0.2
    assert time_metadata.cancel_rampdown_samples == 1000


def test_time_metadata_set_file(time_metadata):
    """
    Verifies that the signal file path can be stored.
    """
    time_metadata.set_file("signal.csv")

    assert time_metadata.signal_file == "signal.csv"


@pytest.mark.parametrize(
    "channel_list_bools, sample_rate, cancel_rampdown_time, output_signal, expected",
    [
        (mock_channel_list_bools(), 1000, 0.5, np.zeros((1, 2000)), True),
        (mock_channel_list_bools(), -10, 0.5, np.zeros((1, 2000)), RattlesnakeError),
        (mock_channel_list_bools(), 1000, None, np.zeros((1, 2000)), RattlesnakeError),
        (mock_channel_list_bools(), None, 0.5, np.zeros((1, 2000)), RattlesnakeError),
        (mock_channel_list_bools(), 1000, 0.5, None, RattlesnakeError),
        (
            mock_channel_list_bools(),
            1000,
            0.5,
            np.zeros((1, 2000, 3)),
            RattlesnakeError,
        ),
        (mock_channel_list_bools(), 1000, 0.5, np.zeros((0, 2000)), RattlesnakeError),
        ([True], 1000, 0.5, np.zeros((1, 2000)), RattlesnakeError),
    ],
)
def test_time_metadata_validate(
    channel_list_bools,
    sample_rate,
    cancel_rampdown_time,
    output_signal,
    expected,
    time_metadata,
    hardware_metadata,
):
    """
    Verifies valid metadata passes validation and invalid metadata raises
    ``RattlesnakeError``.
    """
    time_metadata.channel_list_bools = channel_list_bools
    time_metadata.sample_rate = sample_rate
    time_metadata.cancel_rampdown_time = cancel_rampdown_time
    time_metadata.output_signal = output_signal

    if expected is RattlesnakeError:
        with pytest.raises(RattlesnakeError):
            time_metadata.validate(hardware_metadata)
    else:
        assert time_metadata.validate(hardware_metadata) is True


def test_time_metadata_save_and_load_netcdf(time_metadata, hardware_metadata):
    """
    Verifies that time metadata can be saved to and loaded from a netCDF group.
    """
    dataset = nc4.Dataset("temp.nc", mode="w", diskless=True, persist=False)
    group = dataset.createGroup("Time Environment")

    try:
        time_metadata.save_metadata_to_netcdf(group)

        assert group.cancel_rampdown_time == time_metadata.cancel_rampdown_time
        assert "output_channels" in group.dimensions
        assert "signal_samples" in group.dimensions
        assert "output_signal" in group.variables
        np.testing.assert_array_equal(
            group.variables["output_signal"][...],
            time_metadata.output_signal,
        )

        loaded = TimeMetadata.load_metadata_from_netcdf(
            group,
            "Time Environment",
            mock_channel_list_bools(),
            hardware_metadata,
        )

        assert isinstance(loaded, TimeMetadata)
        assert loaded.environment_name == "Time Environment"
        assert loaded.channel_list_bools == mock_channel_list_bools()
        assert loaded.sample_rate == hardware_metadata.sample_rate
        assert loaded.output_oversample == hardware_metadata.output_oversample
        assert loaded.cancel_rampdown_time == time_metadata.cancel_rampdown_time
        np.testing.assert_array_equal(loaded.output_signal, time_metadata.output_signal)
    finally:
        dataset.close()


def test_time_metadata_create_blank_worksheet_template():
    """
    Verifies that the blank worksheet template contains time-environment fields.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    TimeMetadata.create_blank_worksheet_template(worksheet)

    assert worksheet.cell(1, 1).value == "Control Type"
    assert worksheet.cell(1, 2).value == "Time"
    assert worksheet.cell(1, 3).value == "v4.0"
    assert worksheet.cell(2, 1).value == "Signal File"
    assert worksheet.cell(3, 1).value == "Cancel Rampdown Time"


def test_time_metadata_save_metadata_to_worksheet(time_metadata):
    """
    Verifies that signal file and cancel rampdown time are saved to a worksheet.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    time_metadata.set_file("signal.csv")
    time_metadata.save_metadata_to_worksheet(worksheet)

    assert worksheet.cell(1, 2).value == "Time"
    assert worksheet.cell(2, 2).value == "signal.csv"
    assert worksheet.cell(3, 2).value == str(time_metadata.cancel_rampdown_time)


@mock.patch("rattlesnake.environment.time_environment.load_time_history")
def test_time_metadata_load_metadata_from_worksheet(
    mock_load_time_history,
    hardware_metadata,
):
    """
    Verifies that time metadata can be loaded from a worksheet.
    """
    output_signal = np.ones((1, 100))
    mock_load_time_history.return_value = output_signal

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    TimeMetadata.create_blank_worksheet_template(worksheet)
    worksheet.cell(2, 2, "signal.csv")
    worksheet.cell(3, 2, "0.25")

    loaded = TimeMetadata.load_metadata_from_worksheet(
        worksheet,
        "Time Environment",
        mock_channel_list_bools(),
        hardware_metadata,
    )

    mock_load_time_history.assert_called_once_with(
        "signal.csv",
        hardware_metadata.sample_rate,
    )
    assert loaded.environment_name == "Time Environment"
    assert loaded.signal_file == "signal.csv"
    assert loaded.cancel_rampdown_time == 0.25
    np.testing.assert_array_equal(loaded.output_signal, output_signal)


@mock.patch("rattlesnake.environment.time_environment.load_time_history")
def test_time_metadata_load_metadata_from_worksheet_load_failure(
    mock_load_time_history,
    hardware_metadata,
):
    """
    Verifies that worksheet loading uses a placeholder signal if signal loading
    fails.
    """
    mock_load_time_history.side_effect = RuntimeError("failed")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    TimeMetadata.create_blank_worksheet_template(worksheet)
    worksheet.cell(2, 2, "signal.csv")
    worksheet.cell(3, 2, "0.25")

    loaded = TimeMetadata.load_metadata_from_worksheet(
        worksheet,
        "Time Environment",
        mock_channel_list_bools(),
        hardware_metadata,
    )

    np.testing.assert_array_equal(loaded.output_signal, np.zeros((1, 1)))
    assert loaded.signal_file == "signal.csv"


def test_time_metadata_load_metadata_from_worksheet_invalid_field(hardware_metadata):
    """
    Verifies that unknown worksheet fields raise ``RattlesnakeError``.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 1, "Control Type")
    worksheet.cell(1, 2, "Time")
    worksheet.cell(2, 1, "Bad Field")
    worksheet.cell(2, 2, "value")

    with pytest.raises(RattlesnakeError):
        TimeMetadata.load_metadata_from_worksheet(
            worksheet,
            "Time Environment",
            mock_channel_list_bools(),
            hardware_metadata,
        )


# endregion


# region Time Instructions
def test_time_instructions_init():
    """
    Verifies that time instructions initialize required attributes.
    """
    instructions = TimeInstructions(
        environment_name="Time Environment",
        current_test_level=1,
        repeat=False,
    )

    assert isinstance(instructions, TimeInstructions)
    assert isinstance(instructions, EnvironmentInstructions)
    assert instructions.environment_type == EnvironmentType.TIME
    assert instructions.environment_name == "Time Environment"
    assert instructions.current_test_level == 1
    assert instructions.repeat is False


def test_time_instructions_validate():
    """
    Verifies that time instructions validate without error.
    """
    instructions = TimeInstructions(
        environment_name="Time Environment",
        current_test_level=1,
        repeat=True,
    )

    instructions.validate()


# endregion


# region Time Queues
@pytest.mark.parametrize("use_thread", [True, False])
def test_time_queues_init(use_thread):
    """
    Verifies that ``TimeQueues`` stores supplied queue references.
    """
    queue_container = mock_queue_container(use_thread)

    time_queues = TimeQueues(
        queue_container.environment_command_queues["Environment 0"],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues["Environment 0"],
        queue_container.environment_data_out_queues["Environment 0"],
        queue_container.log_file_queue,
    )

    assert isinstance(time_queues, TimeQueues)
    assert (
        time_queues.environment_command_queue
        is queue_container.environment_command_queues["Environment 0"]
    )
    assert time_queues.gui_update_queue is queue_container.gui_update_queue
    assert (
        time_queues.controller_communication_queue
        is queue_container.controller_command_queue
    )
    assert (
        time_queues.data_in_queue
        is queue_container.environment_data_in_queues["Environment 0"]
    )
    assert (
        time_queues.data_out_queue
        is queue_container.environment_data_out_queues["Environment 0"]
    )
    assert time_queues.log_file_queue is queue_container.log_file_queue


# endregion


# region Time Environment
@pytest.mark.parametrize("use_thread", [True, False])
def test_time_environment_init(use_thread):
    """
    Verifies that ``TimeEnvironment`` initializes and maps expected commands.
    """
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)

    time_queues = TimeQueues(
        queue_container.environment_command_queues["Environment 0"],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues["Environment 0"],
        queue_container.environment_data_out_queues["Environment 0"],
        queue_container.log_file_queue,
    )

    environment = TimeEnvironment(
        "Time Environment",
        "Environment 0",
        time_queues,
        event_container.acquisition_active_event,
        event_container.output_active_event,
        event_container.environment_active_events["Environment 0"],
        event_container.environment_ready_events["Environment 0"],
    )

    assert isinstance(environment, TimeEnvironment)
    assert isinstance(environment, Environment)
    assert environment.queue_container is time_queues
    assert environment.ready is True
    assert environment.active is False

    assert environment.command_map[GlobalCommands.START_ENVIRONMENT] == (
        environment.run_environment
    )
    assert environment.command_map[TimeCommands.SET_TEST_LEVEL] == (
        environment.set_test_level
    )
    assert environment.command_map[TimeCommands.SET_NO_REPEAT] == (
        environment.set_no_repeat
    )
    assert environment.command_map[TimeCommands.SET_REPEAT] == environment.set_repeat

    assert environment.shutdown_flag is False
    assert environment.current_test_level == 0.0
    assert environment.target_test_level == 0.0
    assert environment.test_level_change == 0.0
    assert environment.repeat is False
    assert environment.signal_remainder is None
    assert environment.output_channels is None
    assert environment.measurement_channels is None


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.log")
def test_time_environment_initialize_hardware(
    mock_log,
    time_environment,
    hardware_metadata,
):
    """
    Verifies that hardware metadata is stored and measurement/output channel
    indices are computed.
    """
    time_environment.clear_ready()

    time_environment.initialize_hardware(hardware_metadata)

    mock_log.assert_called_with("Initializing Data Acquisition Parameters")
    assert time_environment.hardware_metadata is hardware_metadata
    assert time_environment.measurement_channels == [0]
    assert time_environment.output_channels == [1]
    assert time_environment.ready is True


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.log")
def test_time_environment_initialize_environment(
    mock_log,
    time_environment,
    time_metadata,
):
    """
    Verifies that time environment metadata is stored.
    """
    time_environment.clear_ready()

    time_environment.initialize_environment(time_metadata)

    mock_log.assert_called_with("Initializing Environment Parameters")
    assert time_environment.environment_metadata is time_metadata
    assert time_environment.environment_name == time_metadata.environment_name
    assert time_environment.ready is True


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.shutdown")
@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.output")
@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.log")
def test_time_environment_run_environment(
    mock_log,
    mock_output,
    mock_shutdown,
    time_environment,
    time_metadata,
    hardware_metadata,
):
    """
    Verifies startup behavior, GUI updates, acquisition data forwarding,
    output chunk selection, and command-loop requeueing.
    """
    mock_gui_queue = mock.MagicMock()
    mock_data_in_queue = mock.MagicMock()
    mock_data_out_queue = mock.MagicMock()
    mock_command_queue = mock.MagicMock()

    time_environment.queue_container.gui_update_queue = mock_gui_queue
    time_environment.queue_container.data_in_queue = mock_data_in_queue
    time_environment.queue_container.data_out_queue = mock_data_out_queue
    time_environment.queue_container.environment_command_queue = mock_command_queue

    mock_data_in_queue.get_nowait.return_value = (np.ones((2, 2000)), False, 0.0)
    mock_data_out_queue.empty.return_value = True

    time_environment.initialize_hardware(hardware_metadata)
    time_environment.environment_metadata = time_metadata

    instructions = TimeInstructions(
        environment_name="Time Environment",
        current_test_level=1,
        repeat=True,
    )

    time_environment.run_environment(instructions)

    assert time_environment.active is True
    assert time_environment.repeat is True
    assert time_environment.current_test_level != 0.0

    mock_gui_queue.put.assert_any_call(
        (
            "Time Environment",
            (UICommands.SET_ENVIRONMENT_INSTRUCTIONS, instructions),
        )
    )
    mock_gui_queue.put.assert_any_call(
        ("Time Environment", (UICommands.ENVIRONMENT_STARTED, None))
    )

    time_data_call = None
    for call in mock_gui_queue.put.call_args_list:
        payload = call.args[0]
        if (
            payload[0] == "Time Environment"
            and payload[1][0] == TimeUICommands.TIME_DATA
        ):
            time_data_call = payload
            break

    assert time_data_call is not None
    measurement_data, output_data = time_data_call[1][1]
    np.testing.assert_array_equal(measurement_data, np.ones((1, 2000)))
    np.testing.assert_array_equal(output_data, np.ones((1, 2000)))

    np.testing.assert_array_equal(
        mock_output.call_args.args[0],
        time_metadata.output_signal[:, : hardware_metadata.samples_per_write],
    )
    assert mock_output.call_args.args[1] is False

    mock_command_queue.put.assert_called_once_with(
        "Time Environment",
        (GlobalCommands.START_ENVIRONMENT, None),
    )
    mock_shutdown.assert_not_called()


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.shutdown")
@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.output")
def test_time_environment_run_environment_final_signal_shutdown(
    mock_output,
    mock_shutdown,
    time_environment,
    time_metadata,
    hardware_metadata,
):
    """
    Verifies that a final signal waits for final acquisition data and then
    shuts down.
    """
    mock_data_in_queue = mock.MagicMock()
    mock_data_out_queue = mock.MagicMock()
    mock_gui_queue = mock.MagicMock()
    mock_command_queue = mock.MagicMock()

    time_environment.queue_container.data_in_queue = mock_data_in_queue
    time_environment.queue_container.data_out_queue = mock_data_out_queue
    time_environment.queue_container.gui_update_queue = mock_gui_queue
    time_environment.queue_container.environment_command_queue = mock_command_queue

    mock_data_in_queue.get_nowait.side_effect = thqueue.Empty
    mock_data_in_queue.get.return_value = (np.ones((2, 100)), True, 0.0)
    mock_data_out_queue.empty.return_value = True

    time_environment.initialize_hardware(hardware_metadata)
    time_environment.environment_metadata = time_metadata
    time_environment.signal_remainder = time_metadata.output_signal.copy()
    time_environment.set_active()
    time_environment.current_test_level = 0.0
    time_environment.repeat = False

    time_environment.run_environment(None)

    assert mock_output.call_args.args[1] is True
    mock_shutdown.assert_called_once_with()
    mock_command_queue.put.assert_not_called()


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.log")
@pytest.mark.parametrize("test_level_change", [0.0, -0.001])
def test_time_environment_output(
    mock_log,
    test_level_change,
    time_environment,
):
    """
    Verifies constant and ramped test-level output data.
    """
    mock_data_out_queue = mock.MagicMock()
    time_environment.queue_container.data_out_queue = mock_data_out_queue

    time_environment.test_level_change = test_level_change
    time_environment.current_test_level = 1.0
    time_environment.test_level_target = 0.8

    write_data = np.ones((1, 1000))

    time_environment.output(write_data, False)

    if test_level_change == 0.0:
        mock_log.assert_has_calls(
            [
                mock.call("Test Level at 1.0"),
                mock.call("Sending data to data_out queue"),
            ]
        )
        expected = np.ones((1, 1000))
    else:
        mock_log.assert_has_calls(
            [
                mock.call("Test level from 0.999 to 0.8"),
                mock.call("Sending data to data_out queue"),
            ]
        )
        expected = 1.0 + (np.arange(1000) + 1) * test_level_change
        full_level_index = np.nonzero(
            abs(expected - 0.8) / abs(test_level_change) < TEST_LEVEL_THRESHOLD
        )[0]
        expected[full_level_index[0] + 1 :] = 0.8
        expected = expected.reshape(1, -1)

    queued_data, last_signal, _ = mock_data_out_queue.put.call_args.args[0]
    np.testing.assert_array_almost_equal(queued_data, expected)
    assert last_signal is False


@mock.patch(
    "rattlesnake.environment.time_environment.TimeEnvironment.adjust_test_level"
)
def test_time_environment_set_test_level(
    mock_adjust_test_level,
    time_environment,
):
    """
    Verifies that setting test level adjusts level and sends a GUI update.
    """
    mock_gui_queue = mock.MagicMock()
    time_environment.queue_container.gui_update_queue = mock_gui_queue

    time_environment.set_test_level(6)

    mock_adjust_test_level.assert_called_once()
    mock_gui_queue.put.assert_called_once_with(
        ("Time Environment", (TimeCommands.SET_TEST_LEVEL, 6))
    )


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.log")
def test_time_environment_set_no_repeat(mock_log, time_environment):
    """
    Verifies that repeat mode can be disabled.
    """
    mock_gui_queue = mock.MagicMock()
    time_environment.queue_container.gui_update_queue = mock_gui_queue
    time_environment.repeat = True

    time_environment.set_no_repeat(None)

    assert time_environment.repeat is False
    mock_log.assert_called_once_with("Repeat turned off")
    mock_gui_queue.put.assert_called_once_with(
        ("Time Environment", (TimeCommands.SET_NO_REPEAT, None))
    )


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.log")
def test_time_environment_set_repeat(mock_log, time_environment):
    """
    Verifies that repeat mode can be enabled.
    """
    mock_gui_queue = mock.MagicMock()
    time_environment.queue_container.gui_update_queue = mock_gui_queue
    time_environment.repeat = False

    time_environment.set_repeat(None)

    assert time_environment.repeat is True
    mock_log.assert_called_once_with("Repeat turned on")
    mock_gui_queue.put.assert_called_once_with(
        ("Time Environment", (TimeCommands.SET_REPEAT, None))
    )


@mock.patch(
    "rattlesnake.environment.time_environment.TimeEnvironment.adjust_test_level"
)
def test_time_environment_stop_environment(
    mock_adjust_test_level,
    time_environment,
):
    """
    Verifies that stopping the environment ramps the test level to zero.
    """
    time_environment.stop_environment(None)

    mock_adjust_test_level.assert_called_once_with(0.0)


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.log")
def test_time_environment_adjust_test_level(
    mock_log,
    time_environment,
    time_metadata,
):
    """
    Verifies that test-level target and per-sample change are updated.
    """
    time_environment.current_test_level = 1.0
    time_environment.environment_metadata = time_metadata

    time_environment.adjust_test_level(0.8)

    assert time_environment.test_level_target == 0.8
    assert time_environment.test_level_change == pytest.approx(
        (0.8 - 1.0) / time_metadata.cancel_rampdown_samples
    )
    mock_log.assert_called_once_with(
        "Changed test level to 0.8 from 1.0, "
        f"{time_environment.test_level_change} change per sample"
    )


def test_time_environment_adjust_test_level_no_change(time_environment, time_metadata):
    """
    Verifies that no log is written when the target level equals the current
    level.
    """
    time_environment.current_test_level = 1.0
    time_environment.environment_metadata = time_metadata

    with mock.patch.object(time_environment, "log") as mock_log:
        time_environment.adjust_test_level(1.0)

    assert time_environment.test_level_target == 1.0
    assert time_environment.test_level_change == 0.0
    mock_log.assert_not_called()


@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment.log")
def test_time_environment_shutdown(mock_log, time_environment):
    """
    Verifies that shutdown flushes the command queue, clears active state, and
    sends an environment-ended GUI update.
    """
    mock_gui_queue = mock.MagicMock()
    mock_command_queue = mock.MagicMock()

    time_environment.queue_container.gui_update_queue = mock_gui_queue
    time_environment.queue_container.environment_command_queue = mock_command_queue
    time_environment.set_active()

    time_environment.shutdown()

    mock_log.assert_called_once_with("Shutting Down Time History Generation")
    mock_command_queue.flush.assert_called_once_with("Time Environment")
    assert time_environment.active is False
    mock_gui_queue.put.assert_called_once_with(
        ("Time Environment", (UICommands.ENVIRONMENT_ENDED, None))
    )


# endregion


# region Time Process
@pytest.mark.parametrize("use_thread", [True, False])
@mock.patch("rattlesnake.environment.time_environment.TimeEnvironment")
def test_time_process(mock_time_environment_class, use_thread):
    """
    Verifies that ``time_process`` constructs a ``TimeEnvironment`` and calls
    its ``run`` method.
    """
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)

    time_process(
        "Environment Name",
        "Environment 0",
        queue_container.environment_command_queues["Environment 0"],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.log_file_queue,
        queue_container.environment_data_in_queues["Environment 0"],
        queue_container.environment_data_out_queues["Environment 0"],
        event_container.acquisition_active_event,
        event_container.output_active_event,
        event_container.environment_active_events["Environment 0"],
        event_container.environment_ready_events["Environment 0"],
        event_container.environment_close_events["Environment 0"],
        event_container.environment_sysid_active_events["Environment 0"],
        event_container.environment_sysid_stored_events["Environment 0"],
        event_container.ping_alive_event,
        use_thread,
    )

    mock_time_environment_class.assert_called_once()
    mock_instance = mock_time_environment_class.return_value
    mock_instance.run.assert_called_once_with(
        event_container.environment_close_events["Environment 0"]
    )


# endregion
