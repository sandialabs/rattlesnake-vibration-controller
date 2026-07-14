from __future__ import annotations

import netCDF4 as nc4
import numpy as np
import openpyxl
import pytest
from typing import List

from rattlesnake.utilities import RattlesnakeError
from rattlesnake.hardware.abstract_hardware import (
    HardwareMetadata,
    HardwareAcquisition,
    HardwareOutput,
)
from rattlesnake.hardware.hardware_utilities import Channel, HardwareType
from rattlesnake.hardware.hardware_registry import (
    HARDWARE_METADATA,
    HARDWARE_ACQUISITION,
    HARDWARE_OUTPUT,
)
from rattlesnake.hardware.skeleton_hardware import (
    SkeletonHardwareMetadata,
    SkeletonHardwareAcquisition,
    SkeletonHardwareOutput,
)
from rattlesnake.user_interface.ui_utilities import HardwareAssistModules
from rattlesnake.testing.mock_utilities import (
    IMPLEMENTED_HARDWARE,
    instantiate_with_mocks,
    mock_channel_list,
    skeleton_hardware_metadata,
)


# region Fixtures
@pytest.fixture
def channel_list() -> list[Channel]:
    return mock_channel_list()


@pytest.fixture
def hardware_metadata() -> SkeletonHardwareMetadata:
    return skeleton_hardware_metadata()


@pytest.fixture
def hardware_acquisition() -> SkeletonHardwareAcquisition:
    return SkeletonHardwareAcquisition()


@pytest.fixture
def hardware_output() -> SkeletonHardwareOutput:
    return SkeletonHardwareOutput()


# endregion


# region Hardware Metadata
@pytest.mark.parametrize("hardware_type", IMPLEMENTED_HARDWARE)
def test_hardware_metadata(hardware_type, channel_list: list[Channel]):
    """
    Verifies that metadata subclasses from the hardware registry initialize
    required metadata attributes and preserve the supplied channel list,
    sample rate, read time, and write time.
    """
    metadata_class = HARDWARE_METADATA[hardware_type]

    metadata = instantiate_with_mocks(
        metadata_class,
        channel_list=channel_list,
        sample_rate=1024,
        time_per_read=0.25,
        time_per_write=0.125,
    )

    assert isinstance(metadata, HardwareMetadata)
    assert metadata.hardware_type == hardware_type
    assert metadata.channel_list is channel_list
    assert metadata.sample_rate == 1024
    assert metadata.time_per_read == 0.25
    assert metadata.time_per_write == 0.125


def test_hardware_metadata_init(hardware_metadata: SkeletonHardwareMetadata):
    """
    Verifies that skeleton hardware metadata initializes required attributes
    with the expected default values.
    """
    assert isinstance(hardware_metadata, HardwareMetadata)

    assert hasattr(hardware_metadata, "hardware_type")
    assert hasattr(hardware_metadata, "channel_list")
    assert hasattr(hardware_metadata, "sample_rate")
    assert hasattr(hardware_metadata, "time_per_read")
    assert hasattr(hardware_metadata, "time_per_write")
    assert hasattr(hardware_metadata, "output_oversample")

    assert hardware_metadata.hardware_type == HardwareType.SKELETON
    assert hardware_metadata.sample_rate == 1024
    assert hardware_metadata.time_per_read == 0.25
    assert hardware_metadata.time_per_write == 0.125
    assert hardware_metadata.output_oversample == 1


def test_hardware_metadata_channel_list_property(channel_list: list[Channel]):
    """
    Verifies that the channel list property returns the configured hardware
    channel list and can be updated.
    """
    metadata = skeleton_hardware_metadata(channel_list=[])
    assert metadata.channel_list == []

    metadata.channel_list = channel_list
    assert metadata.channel_list is channel_list


def test_hardware_metadata_samples_per_read():
    """
    Verifies that ``samples_per_read`` returns the expected number of samples
    per acquisition frame.
    """
    metadata = skeleton_hardware_metadata(sample_rate=1000, time_per_read=0.25)

    assert metadata.samples_per_read == 250


def test_hardware_metadata_samples_per_write():
    """
    Verifies that ``samples_per_write`` includes the output oversampling
    factor.
    """
    metadata = skeleton_hardware_metadata(
        sample_rate=1000,
        time_per_write=0.3,
        output_oversample=10,
    )

    assert metadata.samples_per_write == 3000


def test_hardware_metadata_nyquist_frequency():
    """
    Verifies that the Nyquist frequency is half the acquisition sample rate.
    """
    metadata = skeleton_hardware_metadata(sample_rate=1000)

    assert metadata.nyquist_frequency == 500


def test_hardware_metadata_output_sample_rate():
    """
    Verifies that output sample rate is the acquisition sample rate multiplied
    by the output oversampling factor.
    """
    metadata = skeleton_hardware_metadata(sample_rate=1000, output_oversample=10)

    assert metadata.output_sample_rate == 10000


def test_hardware_metadata_validate_truth(
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Verifies that a valid skeleton hardware metadata object passes validation.
    """
    hardware_metadata.validate()


def test_hardware_metadata_validate_duplicate_channels(
    channel_list: list[Channel],
):
    """
    Verifies that duplicate channels raise ``RattlesnakeError``.
    """
    metadata = skeleton_hardware_metadata(channel_list=channel_list + [channel_list[0]])

    with pytest.raises(RattlesnakeError):
        metadata.validate()


def test_hardware_metadata_valid_channel_dict(
    hardware_metadata: SkeletonHardwareMetadata,
    channel_list: list[Channel],
):
    """
    Verifies that the valid-channel dictionary contains all channel attributes
    and maps each attribute to a list of valid values.
    """
    valid_channel_dict = hardware_metadata.valid_channel_dict(channel_list[0])

    assert set(valid_channel_dict) == set(Channel().channel_attr_list)
    assert all(isinstance(value, list) for value in valid_channel_dict.values())


def test_hardware_metadata_assist_mode_modules(
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Verifies that the assist-mode module dictionary contains all channel
    attributes and defaults each attribute to ``HardwareAssistModules.NONE``.
    """
    assist_mode_modules = hardware_metadata.assist_mode_modules

    assert set(assist_mode_modules) == set(Channel().channel_attr_list)
    assert all(
        value == HardwareAssistModules.NONE for value in assist_mode_modules.values()
    )


def test_hardware_metadata_save_channel_table_to_workbook(
    channel_list: list[Channel],
):
    """
    Saves a hardware channel table to an Excel workbook and verifies that the
    worksheet contains the expected title, section headers, column labels, and
    channel indices.
    """
    workbook = openpyxl.Workbook()

    SkeletonHardwareMetadata.save_channel_table_to_workbook(channel_list, workbook)

    worksheet = workbook.active

    assert worksheet.title == "Channel Table"
    assert worksheet.cell(row=1, column=2).value == "Test Article Definition"
    assert worksheet.cell(row=1, column=5).value == "Instrument Definition"
    assert worksheet.cell(row=1, column=12).value == "Channel Definition"
    assert worksheet.cell(row=1, column=20).value == "Output Feedback"
    assert worksheet.cell(row=1, column=22).value == "Limits"

    assert worksheet.cell(row=2, column=1).value == "Channel Index"
    assert worksheet.cell(row=2, column=2).value == "Node Number"
    assert worksheet.cell(row=3, column=1).value == 0
    assert worksheet.cell(row=4, column=1).value == 1


def test_hardware_metadata_load_channel_table_from_workbook(
    channel_list: list[Channel],
):
    """
    Saves a hardware channel table to an Excel workbook and then loads it back
    into a channel list. Verifies that each loaded channel preserves the
    expected channel attribute values.
    """
    workbook = openpyxl.Workbook()
    SkeletonHardwareMetadata.save_channel_table_to_workbook(channel_list, workbook)

    loaded_channel_list = SkeletonHardwareMetadata.load_channel_table_from_workbook(
        workbook
    )

    assert len(loaded_channel_list) == len(channel_list)

    for loaded_channel, expected_channel in zip(loaded_channel_list, channel_list):
        for attr in Channel().channel_attr_list:
            expected_value = getattr(expected_channel, attr)
            loaded_value = getattr(loaded_channel, attr)

            if expected_value is None:
                assert loaded_value in (None, "")
            else:
                assert str(loaded_value) == str(expected_value)


def test_hardware_metadata_load_channel_table_from_workbook_multiple_channel_sheets():
    """
    Verifies that loading a channel table from a workbook with multiple
    candidate channel table worksheets raises ``RattlesnakeError``.
    """
    workbook = openpyxl.Workbook()
    workbook.active.title = "Channel Table"
    workbook.create_sheet("Backup Channel Table")

    with pytest.raises(RattlesnakeError):
        SkeletonHardwareMetadata.load_channel_table_from_workbook(workbook)


def test_hardware_metadata_save_blank_hardware_to_workbook():
    """
    Saves a blank hardware worksheet to an Excel workbook and verifies that
    the expected hardware metadata labels are written.
    """
    workbook = openpyxl.Workbook()

    SkeletonHardwareMetadata.save_blank_hardware_to_workbook(workbook)

    assert "Hardware" in workbook.sheetnames

    worksheet = workbook["Hardware"]

    assert worksheet.cell(1, 1).value == "Hardware Type"
    assert worksheet.cell(2, 1).value == "Hardware File"
    assert worksheet.cell(3, 1).value == "Sample Rate"
    assert worksheet.cell(4, 1).value == "Time Per Read"
    assert worksheet.cell(5, 1).value == "Time Per Write"
    assert worksheet.cell(6, 1).value == "Maximum Acquisition Processes"
    assert worksheet.cell(7, 1).value == "Integration Oversampling"
    assert worksheet.cell(8, 1).value == "Task Trigger"
    assert worksheet.cell(9, 1).value == "Task Trigger Output Channel"
    assert worksheet.cell(10, 1).value == "Damping Ratio"


def test_hardware_metadata_save_metadata_to_workbook(
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Saves skeleton hardware metadata to an Excel workbook and verifies that the
    hardware worksheet and channel table are created with expected metadata
    field values.
    """
    workbook = openpyxl.Workbook()

    SkeletonHardwareMetadata.save_blank_hardware_to_workbook(workbook)
    hardware_metadata.save_metadata_to_workbook(workbook)

    assert "Channel Table" in workbook.sheetnames
    assert "Hardware" in workbook.sheetnames

    hardware_worksheet = workbook["Hardware"]

    assert hardware_worksheet.cell(1, 2).value == str(
        hardware_metadata.hardware_type.value
    )
    assert hardware_worksheet.cell(3, 2).value == str(hardware_metadata.sample_rate)
    assert hardware_worksheet.cell(4, 2).value == str(hardware_metadata.time_per_read)
    assert hardware_worksheet.cell(5, 2).value == str(hardware_metadata.time_per_write)


@pytest.mark.parametrize("hardware_type", IMPLEMENTED_HARDWARE)
def test_hardware_metadata_load_save_workbook(
    hardware_type, channel_list: List[Channel]
):
    """
    Saves skeleton hardware metadata to an Excel workbook and then loads it
    back into a metadata object. Verifies that the loaded metadata preserves
    the expected hardware type, timing parameters, output oversampling, and
    channel list length.
    """
    metadata_class = HARDWARE_METADATA[hardware_type]

    hardware_metadata = instantiate_with_mocks(
        metadata_class,
        channel_list=channel_list,
        sample_rate=1024,
        time_per_read=0.25,
        time_per_write=0.125,
    )

    workbook = openpyxl.Workbook()

    hardware_metadata.save_blank_hardware_to_workbook(workbook)
    hardware_metadata.save_metadata_to_workbook(workbook)

    loaded_metadata = metadata_class.load_metadata_from_workbook(workbook)

    assert loaded_metadata.sample_rate == hardware_metadata.sample_rate
    assert loaded_metadata.time_per_read == hardware_metadata.time_per_read
    assert loaded_metadata.time_per_write == hardware_metadata.time_per_write
    assert loaded_metadata.output_oversample == hardware_metadata.output_oversample
    assert len(loaded_metadata.channel_list) == len(hardware_metadata.channel_list)


def test_hardware_metadata_save_metadata_to_netcdf_contents(
    hardware_metadata: SkeletonHardwareMetadata,
    tmp_path,
):
    """
    Saves skeleton hardware metadata to a netCDF dataset and verifies that the
    expected dimensions, variables, groups, and dataset attributes are created.
    """
    path = tmp_path / "hardware_metadata.nc4"

    with nc4.Dataset(path, "w", format="NETCDF4") as dataset:
        hardware_metadata.save_metadata_to_netcdf(dataset)

    with nc4.Dataset(path, "r") as dataset:
        assert "response_channels" in dataset.dimensions
        assert "output_channels" in dataset.dimensions
        assert "time_samples" in dataset.dimensions
        assert "time_data" in dataset.variables
        assert "channels" in dataset.groups

        assert dataset.file_version == "3.0.0"
        assert dataset.sample_rate == hardware_metadata.sample_rate
        assert dataset.hardware == hardware_metadata.hardware_type.value
        assert dataset.output_oversample == hardware_metadata.output_oversample

        assert dataset.dimensions["response_channels"].size == len(
            hardware_metadata.channel_list
        )


@pytest.mark.parametrize("hardware_type", IMPLEMENTED_HARDWARE)
def test_hardware_metadata_load_save_netcdf(
    hardware_type,
    channel_list: List[Channel],
    tmp_path,
):
    """
    Saves skeleton hardware metadata to a netCDF dataset and then loads it
    back into a metadata object. Verifies that the loaded metadata preserves
    the expected hardware type, sample rate, timing parameters, output
    oversampling, and channel list length.
    """
    metadata_class = HARDWARE_METADATA[hardware_type]

    hardware_metadata = instantiate_with_mocks(
        metadata_class,
        channel_list=channel_list,
        sample_rate=1024,
        time_per_read=0.25,
        time_per_write=0.125,
    )

    path = tmp_path / "hardware_metadata.nc4"

    with nc4.Dataset(path, "w", format="NETCDF4") as dataset:
        hardware_metadata.save_metadata_to_netcdf(dataset)

    with nc4.Dataset(path, "r") as dataset:
        loaded_metadata = metadata_class.load_metadata_from_netcdf(dataset)

    assert loaded_metadata.sample_rate == hardware_metadata.sample_rate
    assert loaded_metadata.time_per_read == pytest.approx(
        hardware_metadata.samples_per_read / hardware_metadata.sample_rate
    )
    assert loaded_metadata.time_per_write == pytest.approx(
        hardware_metadata.samples_per_write / hardware_metadata.output_sample_rate
    )
    assert loaded_metadata.output_oversample == hardware_metadata.output_oversample
    assert len(loaded_metadata.channel_list) == len(hardware_metadata.channel_list)


# endregion


# region Hardware Acquisition
@pytest.mark.parametrize("hardware_type", IMPLEMENTED_HARDWARE)
def test_hardware_acquisition(hardware_type: HardwareType):
    """
    Verifies that acquisition subclasses from the hardware registry can be
    instantiated and implement the hardware acquisition interface.
    """
    acquisition_class = HARDWARE_ACQUISITION[hardware_type]

    hardware_acquisition = instantiate_with_mocks(acquisition_class)

    assert isinstance(hardware_acquisition, HardwareAcquisition)


def test_hardware_acquisition_initialize_hardware(
    hardware_metadata: SkeletonHardwareMetadata,
    hardware_acquisition: SkeletonHardwareAcquisition,
):
    """
    Verifies that acquisition subclasses store the supplied hardware metadata
    during hardware initialization.
    """

    hardware_acquisition.initialize_hardware(hardware_metadata)

    assert hardware_acquisition.metadata == hardware_metadata


def test_hardware_acquisition_init(
    hardware_acquisition: SkeletonHardwareAcquisition,
):
    """
    Verifies that the skeleton acquisition class implements the hardware
    acquisition interface.
    """
    assert isinstance(hardware_acquisition, HardwareAcquisition)


def test_hardware_acquisition_start(
    hardware_acquisition: SkeletonHardwareAcquisition,
):
    """
    Verifies that skeleton acquisition hardware can be started.
    """
    hardware_acquisition.start()

    assert hardware_acquisition.started is True


def test_hardware_acquisition_read(
    hardware_acquisition: SkeletonHardwareAcquisition,
):
    """
    Verifies that reading from skeleton acquisition hardware returns a NumPy
    array with the expected shape.
    """
    data = hardware_acquisition.read()

    assert isinstance(data, np.ndarray)
    assert data.shape == (2, 10)


def test_hardware_acquisition_read_remaining(
    hardware_acquisition: SkeletonHardwareAcquisition,
):
    """
    Verifies that reading remaining acquisition data from skeleton acquisition
    hardware returns a NumPy array with the expected shape.
    """
    data = hardware_acquisition.read_remaining()

    assert isinstance(data, np.ndarray)
    assert data.shape == (2, 3)


def test_hardware_acquisition_stop(
    hardware_acquisition: SkeletonHardwareAcquisition,
):
    """
    Verifies that skeleton acquisition hardware can be stopped.
    """
    hardware_acquisition.stop()

    assert hardware_acquisition.stopped is True


def test_hardware_acquisition_close(
    hardware_acquisition: SkeletonHardwareAcquisition,
):
    """
    Verifies that skeleton acquisition hardware can be closed.
    """
    hardware_acquisition.close()

    assert hardware_acquisition.closed is True


def test_hardware_acquisition_get_acquisition_delay(
    hardware_acquisition: SkeletonHardwareAcquisition,
):
    """
    Verifies that skeleton acquisition hardware returns the expected
    acquisition delay.
    """
    acquisition_delay = hardware_acquisition.get_acquisition_delay()

    assert isinstance(acquisition_delay, int)
    assert acquisition_delay == 0


def test_hardware_acquisition_functions(
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Calls all skeleton acquisition interface methods and verifies that the
    methods store metadata, update state flags, and return expected data
    types.
    """
    hardware_acquisition = SkeletonHardwareAcquisition()

    hardware_acquisition.initialize_hardware(hardware_metadata)
    hardware_acquisition.start()
    read_data = hardware_acquisition.read()
    remaining_data = hardware_acquisition.read_remaining()
    hardware_acquisition.stop()
    hardware_acquisition.close()
    acquisition_delay = hardware_acquisition.get_acquisition_delay()

    assert hardware_acquisition.metadata is hardware_metadata
    assert hardware_acquisition.started is True
    assert hardware_acquisition.stopped is True
    assert hardware_acquisition.closed is True
    assert isinstance(read_data, np.ndarray)
    assert isinstance(remaining_data, np.ndarray)
    assert isinstance(acquisition_delay, int)


# endregion


# region Hardware Output
@pytest.mark.parametrize("hardware_type", IMPLEMENTED_HARDWARE)
def test_hardware_output(hardware_type: HardwareType):
    """
    Verifies that output subclasses from the hardware registry can be
    instantiated and implement the hardware output interface.
    """
    output_class = HARDWARE_OUTPUT[hardware_type]

    hardware_output = instantiate_with_mocks(output_class)

    assert isinstance(hardware_output, HardwareOutput)


def test_hardware_output_initialize_hardware(
    hardware_metadata: SkeletonHardwareMetadata, hardware_output: SkeletonHardwareOutput
):
    """
    Verifies that output subclasses store the supplied hardware metadata
    during hardware initialization.
    """
    hardware_output.initialize_hardware(hardware_metadata)

    assert hardware_output.metadata is hardware_metadata


def test_hardware_output_init(
    hardware_output: SkeletonHardwareOutput,
):
    """
    Verifies that the skeleton output class implements the hardware output
    interface.
    """
    assert isinstance(hardware_output, HardwareOutput)


def test_hardware_output_start(
    hardware_output: SkeletonHardwareOutput,
):
    """
    Verifies that skeleton output hardware can be started.
    """
    hardware_output.start()

    assert hardware_output.started is True


def test_hardware_output_write(
    hardware_output: SkeletonHardwareOutput,
):
    """
    Verifies that skeleton output hardware accepts and stores output data.
    """
    output_data = np.zeros((1, 100))

    hardware_output.write(output_data)

    assert hardware_output.last_write is output_data


def test_hardware_output_stop(
    hardware_output: SkeletonHardwareOutput,
):
    """
    Verifies that skeleton output hardware can be stopped.
    """
    hardware_output.stop()

    assert hardware_output.stopped is True


def test_hardware_output_close(
    hardware_output: SkeletonHardwareOutput,
):
    """
    Verifies that skeleton output hardware can be closed.
    """
    hardware_output.close()

    assert hardware_output.closed is True


def test_hardware_output_ready_for_new_output(
    hardware_output: SkeletonHardwareOutput,
):
    """
    Verifies that skeleton output hardware reports readiness for new output.
    """
    assert hardware_output.ready_for_new_output() is True


def test_hardware_output_functions(
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Calls all skeleton output interface methods and verifies that the methods
    store metadata, update state flags, store output data, and report readiness
    for additional output.
    """
    hardware_output = SkeletonHardwareOutput()
    output_data = np.zeros((1, 100))

    hardware_output.initialize_hardware(hardware_metadata)
    hardware_output.start()
    hardware_output.write(output_data)
    hardware_output.stop()
    hardware_output.close()
    ready = hardware_output.ready_for_new_output()

    assert hardware_output.metadata is hardware_metadata
    assert hardware_output.started is True
    assert hardware_output.last_write is output_data
    assert hardware_output.stopped is True
    assert hardware_output.closed is True
    assert ready is True


# endregion
