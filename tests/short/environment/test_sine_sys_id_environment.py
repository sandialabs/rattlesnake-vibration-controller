"""Tests for the sine sys-id environment.

These tests drive the metadata file flows (worksheet with external
specification .npz files, netCDF with breakpoint tables) and the environment
lifecycle the controller uses in a real sine test (initialize
hardware/environment/sysid with real specification signal synthesis ->
transfer function -> SYSTEM_ID_COMPLETE -> QUIT), using hand-fed queues with
no spawned sub-workers.

SineEnvironment.start_control is a self-re-enqueueing real-time loop that
consumes acquired frames against live signal-generation state; it requires
running sub-workers and is exercised end-to-end by
tests/long/test_qualification.py, so it is not covered here.  The control
prediction path is covered with the prediction patched out because it
requires excitation state produced by that loop.
"""

import queue as thqueue
from unittest import mock

import netCDF4 as nc4
import numpy as np
import openpyxl
import pytest

from rattlesnake.environment.abstract_environment import (
    EnvironmentInstructions,
    EnvironmentMetadata,
)
from rattlesnake.environment.abstract_sysid_environment import (
    SysIdEnvironment,
    SysIdUICommands,
    SystemIdCommands,
)
from rattlesnake.environment.sine_sys_id_environment import (
    SineCommands,
    SineEnvironment,
    SineInstructions,
    SineMetadata,
    SineQueues,
    SineUICommands,
    sine_process,
)
from rattlesnake.environment.sine_sys_id_utilities import (
    DefaultSineControlLaw,
    SineSpecification,
)
from rattlesnake.process.abstract_sysid_data_analysis import (
    SysIdDataAnalysisCommands,
    SysIdDataAnalysisUICommands,
    SysIdMetadata,
)
from rattlesnake.process.data_collector import DataCollectorCommands
from rattlesnake.process.signal_generation_process import SignalGenerationCommands
from rattlesnake.process.spectral_processing import SpectralProcessingCommands
from rattlesnake.testing.mock_utilities import (
    drain_queue_commands,
    get_queue_messages,
    sysid_measurement_data_package,
    sysid_measurement_metadata,
    numeric_hardware_metadata,
    write_sine_spec_npz,
    mock_event_container,
    mock_queue_container,
)
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.utilities import GlobalCommands, RattlesnakeError, flush_queue

ENVIRONMENT_NAME = "Sine Environment"
QUEUE_NAME = "Environment 0"


def flush_environment_queues(environment):
    """Drains every queue so no multiprocessing feeder thread is left
    blocked on a full pipe, which would hang the interpreter at exit."""
    queues = environment.queue_container
    for queue in (
        queues.gui_update_queue,
        queues.environment_command_queue,
        queues.data_analysis_command_queue,
        queues.collector_command_queue,
        queues.signal_generation_command_queue,
        queues.spectral_command_queue,
    ):
        flush_queue(queue, timeout=0.05)


# region Fixtures
def make_sine_specification(amplitude=1.0):
    # A single 10 Hz to 20 Hz linear sweep at 10 Hz/s (1 second long)
    return SineSpecification(
        name="Tone 1",
        start_time=0.0,
        num_control=1,
        frequency_breakpoints=np.array([10.0, 20.0]),
        amplitude_breakpoints=np.full((2, 1), amplitude),
        phase_breakpoints=np.zeros((2, 1)),
        sweep_type_breakpoints=np.array([0], dtype="u1"),
        sweep_rate_breakpoints=np.array([10.0]),
    )


def make_sine_metadata(amplitude=1.0):
    return SineMetadata(
        environment_name=ENVIRONMENT_NAME,
        channel_list_bools=[True, True],
        sample_rate=1000,
        samples_per_frame=250,
        number_of_channels=2,
        specifications=[make_sine_specification(amplitude)],
        ramp_time=0.05,
        buffer_blocks=2,
        control_convergence=0.5,
        update_drives_after_environment=False,
        phase_fit=False,
        allow_automatic_aborts=False,
        tracking_filter_type=0,
        tracking_filter_cutoff=0.5,
        tracking_filter_order=2,
        vk_filter_order=1,
        vk_filter_bandwidth=1.0,
        vk_filter_blocksize=100,
        vk_filter_overlap=0.25,
        control_python_script="",
        control_python_class="",
        control_python_parameters="",
        control_channel_indices=[0],
        output_channel_indices=[1],
        response_transformation_matrix=None,
        output_transformation_matrix=None,
        sysid_metadata=sysid_measurement_metadata(),
    )


@pytest.fixture
def sine_metadata():
    sine_metadata = make_sine_metadata()
    sine_metadata.queue_name = QUEUE_NAME

    return sine_metadata


@pytest.fixture(params=[True, False], ids=["threaded", "non_threaded"])
def sine_environment(request):
    use_thread = request.param
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)
    sine_queues = SineQueues(
        ENVIRONMENT_NAME,
        queue_container.environment_command_queues[QUEUE_NAME],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues[QUEUE_NAME],
        queue_container.environment_data_out_queues[QUEUE_NAME],
        queue_container.log_file_queue,
        use_thread,
    )
    sine_environment = SineEnvironment(
        ENVIRONMENT_NAME,
        QUEUE_NAME,
        sine_queues,
        event_container.acquisition_active_event,
        event_container.output_active_event,
        event_container.environment_active_events[QUEUE_NAME],
        event_container.environment_ready_events[QUEUE_NAME],
        event_container.environment_sysid_active_events[QUEUE_NAME],
        event_container.environment_sysid_stored_events[QUEUE_NAME],
    )
    yield sine_environment
    flush_environment_queues(sine_environment)


def initialized_environment(sine_environment, sine_metadata):
    """Runs the environment through the real hardware/environment init flow."""
    sine_environment.initialize_hardware(numeric_hardware_metadata())
    sine_environment.initialize_environment(sine_metadata)
    # Discard the initialization queue traffic so tests can assert on
    # the traffic of the method under test only
    for queue in (
        sine_environment.queue_container.data_analysis_command_queue,
        sine_environment.queue_container.collector_command_queue,
        sine_environment.queue_container.signal_generation_command_queue,
        sine_environment.queue_container.spectral_command_queue,
    ):
        drain_queue_commands(queue, ENVIRONMENT_NAME)
    flush_queue(sine_environment.gui_update_queue, timeout=0.1)
    return sine_environment


# region Metadata
def test_sine_metadata_init(sine_metadata):
    assert isinstance(sine_metadata, SineMetadata)
    assert isinstance(sine_metadata, EnvironmentMetadata)
    assert isinstance(sine_metadata.sysid_metadata, SysIdMetadata)


def test_sine_metadata_properties(sine_metadata):
    assert sine_metadata.ramp_samples == 50
    assert sine_metadata.response_channel_indices == [0]
    assert sine_metadata.reference_channel_indices == [1]
    assert sine_metadata.num_response_channels == 1
    assert sine_metadata.num_reference_channels == 1


def test_sine_metadata_transformed_channel_counts(sine_metadata):
    sine_metadata.response_transformation_matrix = np.ones((3, 1))
    sine_metadata.reference_transformation_matrix = np.ones((2, 1))

    assert sine_metadata.num_response_channels == 3
    assert sine_metadata.num_reference_channels == 2


@pytest.mark.parametrize(
    "channel_list_bools, environment_name, expected",
    [
        ([True, True], ENVIRONMENT_NAME, True),
        ([True], ENVIRONMENT_NAME, RattlesnakeError),
        ([True, True], 17, RattlesnakeError),
    ],
)
def test_sine_metadata_validate(
    channel_list_bools, environment_name, expected, sine_metadata
):
    hardware_metadata = numeric_hardware_metadata()
    sine_metadata.channel_list_bools = channel_list_bools
    sine_metadata.environment_name = environment_name

    if expected is RattlesnakeError:
        with pytest.raises(RattlesnakeError):
            sine_metadata.validate(hardware_metadata)
    else:
        sine_metadata.validate(hardware_metadata)


def test_sine_specification_equality():
    assert make_sine_specification() == make_sine_specification()
    assert not make_sine_specification() == make_sine_specification(amplitude=2.0)


@pytest.mark.parametrize("with_transformations", [False, True])
def test_sine_metadata_netcdf_round_trip(sine_metadata, with_transformations):
    if with_transformations:
        sine_metadata.response_transformation_matrix = np.arange(2.0).reshape(2, 1)
        sine_metadata.reference_transformation_matrix = np.arange(3.0).reshape(3, 1)
    dataset = nc4.Dataset("temp.nc", mode="w", diskless=True, persist=False)
    netcdf_group = dataset.createGroup(ENVIRONMENT_NAME)

    sine_metadata.save_metadata_to_netcdf(netcdf_group)
    loaded_metadata = SineMetadata.load_metadata_from_netcdf(
        netcdf_group,
        ENVIRONMENT_NAME,
        [True, True],
        numeric_hardware_metadata(),
    )

    assert loaded_metadata.samples_per_frame == 250
    assert loaded_metadata.ramp_time == 0.05
    assert loaded_metadata.buffer_blocks == 2
    assert loaded_metadata.control_convergence == 0.5
    assert loaded_metadata.update_drives_after_environment is False
    assert loaded_metadata.phase_fit is False
    assert loaded_metadata.allow_automatic_aborts is False
    assert loaded_metadata.tracking_filter_type == 0
    assert loaded_metadata.tracking_filter_cutoff == 0.5
    assert loaded_metadata.tracking_filter_order == 2
    assert loaded_metadata.vk_filter_order == 1
    assert loaded_metadata.vk_filter_bandwidth == 1.0
    assert loaded_metadata.vk_filter_blocksize == 100
    assert loaded_metadata.vk_filter_overlap == 0.25
    np.testing.assert_array_equal(
        np.asarray(loaded_metadata.control_channel_indices), [0]
    )
    assert loaded_metadata.output_channel_indices == [1]
    # The specification round-trips through its breakpoint table
    assert len(loaded_metadata.specifications) == 1
    loaded_specification = loaded_metadata.specifications[0]
    assert loaded_specification.name == "Tone 1"
    assert loaded_specification == sine_metadata.specifications[0]
    assert loaded_metadata.sysid_metadata == sine_metadata.sysid_metadata

    dataset.close()


def test_sine_metadata_worksheet_round_trip(tmp_path, sine_metadata):
    # This is the profile .xlsx flow: the specification comes from an
    # external .npz file referenced on the worksheet
    spec_path = str(tmp_path / "sine_spec.npz")
    spec_data = write_sine_spec_npz(spec_path)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    sine_metadata.save_metadata_to_worksheet(worksheet)
    worksheet.cell(35, 2, spec_path)

    loaded_metadata = SineMetadata.load_metadata_from_worksheet(
        worksheet,
        ENVIRONMENT_NAME,
        [True, True],
        numeric_hardware_metadata(),
    )

    assert loaded_metadata.ramp_time == 0.05
    assert loaded_metadata.control_convergence == 0.5
    assert loaded_metadata.update_drives_after_environment is False
    assert loaded_metadata.phase_fit is False
    assert loaded_metadata.allow_automatic_aborts is False
    assert loaded_metadata.buffer_blocks == 2
    # DFT/VK selection round-trips through the "DFT"/"VK" worksheet cell
    assert loaded_metadata.tracking_filter_type == 0
    assert loaded_metadata.tracking_filter_cutoff == 0.5
    assert loaded_metadata.tracking_filter_order == 2
    assert loaded_metadata.vk_filter_order == 1
    assert loaded_metadata.vk_filter_bandwidth == 1.0
    assert loaded_metadata.vk_filter_blocksize == 100
    assert loaded_metadata.vk_filter_overlap == 0.25
    assert loaded_metadata.control_python_script == ""
    assert loaded_metadata.control_channel_indices == [0]
    assert loaded_metadata.output_channel_indices == [1]
    # The samples per frame comes from the hardware read size on load
    assert loaded_metadata.samples_per_frame == 250
    assert loaded_metadata.sysid_metadata == sine_metadata.sysid_metadata
    # The specification was rebuilt from the .npz file
    assert len(loaded_metadata.specifications) == 1
    loaded_specification = loaded_metadata.specifications[0]
    assert loaded_specification.name == "Tone 1"
    np.testing.assert_allclose(
        loaded_specification.breakpoint_table["frequency"],
        spec_data["frequency"],
    )
    np.testing.assert_allclose(
        loaded_specification.breakpoint_table["amplitude"],
        spec_data["amplitude"],
    )
    np.testing.assert_allclose(
        loaded_specification.breakpoint_table["sweep_rate"][:-1],
        spec_data["sweep_rate"],
    )


def test_sine_metadata_worksheet_multiple_specification_files(tmp_path, sine_metadata):
    first_spec_path = str(tmp_path / "tone_1.npz")
    second_spec_path = str(tmp_path / "tone_2.npz")
    write_sine_spec_npz(first_spec_path, name="Tone 1")
    write_sine_spec_npz(
        second_spec_path, name="Tone 2", frequencies=(30.0, 40.0), amplitude=2.0
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    sine_metadata.save_metadata_to_worksheet(worksheet)
    worksheet.cell(35, 2, first_spec_path)
    worksheet.cell(35, 3, second_spec_path)

    loaded_metadata = SineMetadata.load_metadata_from_worksheet(
        worksheet,
        ENVIRONMENT_NAME,
        [True, True],
        numeric_hardware_metadata(),
    )

    assert len(loaded_metadata.specifications) == 2
    assert loaded_metadata.specifications[0].name == "Tone 1"
    assert loaded_metadata.specifications[1].name == "Tone 2"
    np.testing.assert_allclose(
        loaded_metadata.specifications[1].breakpoint_table["frequency"],
        [30.0, 40.0],
    )


# region Instructions
def test_sine_instructions_init():
    sine_instructions = SineInstructions(
        ENVIRONMENT_NAME,
        control_test_level=0.0,
        control_tones=None,
        control_start_time=None,
        control_end_time=None,
    )

    assert isinstance(sine_instructions, SineInstructions)
    assert isinstance(sine_instructions, EnvironmentInstructions)
    assert sine_instructions.control_test_level == 0.0
    sine_instructions.validate()


# region Queues
@pytest.mark.parametrize("use_thread", [True, False])
def test_sine_queues_init(use_thread):
    queue_container = mock_queue_container(use_thread)
    sine_queues = SineQueues(
        ENVIRONMENT_NAME,
        queue_container.environment_command_queues[QUEUE_NAME],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues[QUEUE_NAME],
        queue_container.environment_data_out_queues[QUEUE_NAME],
        queue_container.log_file_queue,
        use_thread,
    )

    assert isinstance(sine_queues, SineQueues)
    for queue_name in (
        "data_analysis_command_queue",
        "signal_generation_command_queue",
        "spectral_command_queue",
        "collector_command_queue",
        "time_history_to_generate_queue",
    ):
        assert hasattr(sine_queues, queue_name)


# region Environment
def test_sine_environment_init(sine_environment):
    assert isinstance(sine_environment, SineEnvironment)
    assert isinstance(sine_environment, SysIdEnvironment)
    assert sine_environment.ready
    assert sine_environment.control_startup
    for command in (
        GlobalCommands.QUIT,
        GlobalCommands.INITIALIZE_HARDWARE,
        GlobalCommands.INITIALIZE_ENVIRONMENT,
        GlobalCommands.INITIALIZE_SYSTEM_ID,
        GlobalCommands.START_SYSTEM_ID_TRANSFER,
        GlobalCommands.START_ENVIRONMENT,
        SineCommands.START_CONTROL,
        SineCommands.STOP_CONTROL,
        SineCommands.SAVE_CONTROL_DATA,
        SineCommands.SET_TEST_LEVEL,
        SysIdDataAnalysisCommands.SYSTEM_ID_COMPLETE,
    ):
        assert command in sine_environment.command_map


def test_sine_environment_initialize_hardware(sine_environment):
    hardware_metadata = numeric_hardware_metadata()
    sine_environment.initialize_hardware(hardware_metadata)

    assert sine_environment.hardware_metadata == hardware_metadata
    assert sine_environment.ready


def test_sine_environment_initialize_environment(sine_environment, sine_metadata):
    """The real initialization use case: the default control law is built
    and the specification sweep signals are synthesized."""
    sine_environment.initialize_hardware(numeric_hardware_metadata())
    sine_environment.initialize_environment(sine_metadata)

    assert sine_environment.environment_metadata is sine_metadata
    assert isinstance(sine_environment.control_class, DefaultSineControlLaw)
    assert sine_environment.ramp_samples == 50

    # The specification signals were synthesized: a 1 second sweep at
    # 1000 Hz sample rate plus start and end ramps of 50 samples each
    combined_signals = sine_environment.specification_signals_combined
    assert combined_signals.shape[0] == 1
    assert combined_signals.shape[-1] >= 1000
    assert np.max(np.abs(combined_signals)) == pytest.approx(1.0, rel=0.01)

    (analysis_message,) = get_queue_messages(
        sine_environment.queue_container.data_analysis_command_queue,
        ENVIRONMENT_NAME,
        1,
    )
    assert analysis_message == (
        GlobalCommands.INITIALIZE_ENVIRONMENT,
        ENVIRONMENT_NAME,
    )

    gui_message = sine_environment.gui_update_queue.get(timeout=10)
    assert gui_message[0] == ENVIRONMENT_NAME
    assert gui_message[1][0] == SineUICommands.SPECIFICATION_FOR_PLOTTING
    assert len(gui_message[1][1]) == 7

    assert sine_environment.ready


def test_sine_environment_initialize_environment_same_specification_keeps_state(
    sine_environment, sine_metadata
):
    initialized_environment(sine_environment, sine_metadata)
    excitation_sentinel = np.ones((1, 10))
    sine_environment.excitation_signals_combined = excitation_sentinel

    sine_environment.initialize_environment(sine_metadata)

    assert sine_environment.excitation_signals_combined is excitation_sentinel


def test_sine_environment_initialize_environment_new_specification_resets_state(
    sine_environment, sine_metadata
):
    initialized_environment(sine_environment, sine_metadata)
    sine_environment.excitation_signals_combined = np.ones((1, 10))

    new_metadata = make_sine_metadata(amplitude=2.0)
    new_metadata.queue_name = QUEUE_NAME
    sine_environment.initialize_environment(new_metadata)

    assert sine_environment.excitation_signals_combined is None
    assert sine_environment.sysid_data.sysid_frf is None


def test_sine_environment_initialize_sysid(sine_environment, sine_metadata):
    initialized_environment(sine_environment, sine_metadata)
    new_sysid_metadata = sysid_measurement_metadata(frame_size=100)

    sine_environment.initialize_sysid(new_sysid_metadata)

    assert sine_environment.environment_metadata.sysid_metadata == new_sysid_metadata
    (analysis_message,) = get_queue_messages(
        sine_environment.queue_container.data_analysis_command_queue,
        ENVIRONMENT_NAME,
        1,
    )
    assert analysis_message == (
        SysIdDataAnalysisCommands.INITIALIZE_PARAMETERS,
        new_sysid_metadata,
    )
    assert sine_environment.ready


def test_sine_environment_start_transfer_function(sine_environment, sine_metadata):
    initialized_environment(sine_environment, sine_metadata)

    sine_environment.start_transfer_function(None)

    queues = sine_environment.queue_container
    collector_messages = get_queue_messages(
        queues.collector_command_queue, ENVIRONMENT_NAME, 4
    )
    assert [message for message, _ in collector_messages] == [
        DataCollectorCommands.FORCE_INITIALIZE_COLLECTOR,
        DataCollectorCommands.SET_TEST_LEVEL,
        DataCollectorCommands.ACQUIRE,
        DataCollectorCommands.CLEAR_KURTOSIS_BUFFER,
    ]
    siggen_messages = get_queue_messages(
        queues.signal_generation_command_queue, ENVIRONMENT_NAME, 5
    )
    assert [message for message, _ in siggen_messages] == [
        SignalGenerationCommands.INITIALIZE_PARAMETERS,
        SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
        SignalGenerationCommands.MUTE,
        SignalGenerationCommands.ADJUST_TEST_LEVEL,
        SignalGenerationCommands.GENERATE_SIGNALS,
    ]
    (analysis_message,) = get_queue_messages(
        queues.data_analysis_command_queue, ENVIRONMENT_NAME, 1
    )
    assert analysis_message[0] == SysIdDataAnalysisCommands.RUN_TRANSFER_FUNCTION
    spectral_messages = get_queue_messages(
        queues.spectral_command_queue, ENVIRONMENT_NAME, 3
    )
    assert [message for message, _ in spectral_messages] == [
        SpectralProcessingCommands.INITIALIZE_PARAMETERS,
        SpectralProcessingCommands.CLEAR_SPECTRAL_PROCESSING,
        SpectralProcessingCommands.RUN_SPECTRAL_PROCESSING,
    ]
    assert sine_environment.sysid_active
    gui_message = queues.gui_update_queue.get(timeout=10)
    assert gui_message == (ENVIRONMENT_NAME, (SysIdUICommands.SYSID_STARTED, None))


def test_sine_environment_system_id_complete(sine_environment, sine_metadata):
    initialized_environment(sine_environment, sine_metadata)
    data_package = sysid_measurement_data_package(sine_metadata.sysid_metadata)

    with mock.patch.object(
        SineEnvironment, "perform_control_prediction"
    ) as mock_prediction:
        sine_environment.system_id_complete(
            (sine_metadata.sysid_metadata, data_package)
        )

    assert sine_environment.sysid_data == data_package
    assert sine_environment.sysid_stored
    mock_prediction.assert_called_with(True)

    gui_message = sine_environment.gui_update_queue.get(timeout=10)
    assert gui_message[1][0] == SysIdDataAnalysisUICommands.TRANSFER_COMPLETED
    gui_message = sine_environment.gui_update_queue.get(timeout=10)
    assert gui_message[0] == UICommands.COMPLETED_SYSTEM_ID


def test_sine_environment_set_test_level(sine_environment, sine_metadata):
    initialized_environment(sine_environment, sine_metadata)

    # While inactive, the test level change is forwarded to the UI
    sine_environment.set_test_level(-3.0)
    gui_message = sine_environment.gui_update_queue.get(timeout=10)
    assert gui_message == (
        ENVIRONMENT_NAME,
        (SineCommands.SET_TEST_LEVEL, -3.0),
    )

    # While the environment is running, the level cannot be changed
    sine_environment.set_active()
    sine_environment.set_test_level(-6.0)
    with pytest.raises(thqueue.Empty):
        sine_environment.gui_update_queue.get(timeout=0.5)


def test_sine_environment_save_control_data(tmp_path, sine_environment, sine_metadata):
    initialized_environment(sine_environment, sine_metadata)
    output_path = str(tmp_path / "control_data.npz")
    sine_environment.control_response_signals_combined = [np.ones((1, 100))]
    sine_environment.control_response_amplitudes = [np.ones((1, 100))]
    sine_environment.control_response_phases = [np.zeros((1, 100))]
    sine_environment.control_drive_modifications = [np.ones((1, 2))]
    sine_environment.control_response_frequencies = np.linspace(10, 20, 100)
    sine_environment.control_response_arguments = np.linspace(0, 100, 100)
    sine_environment.control_target_phases = np.zeros((1, 100))
    sine_environment.control_target_amplitudes = np.ones((1, 100))

    sine_environment.save_control_data(output_path)

    with np.load(output_path) as saved_data:
        assert saved_data["sample_rate"] == 1024
        assert saved_data["output_oversample"] == 1
        np.testing.assert_array_equal(saved_data["names"], ["Tone 1"])
        np.testing.assert_allclose(
            saved_data["control_response_signals_combined_0"], np.ones((1, 100))
        )
        np.testing.assert_allclose(
            saved_data["control_response_frequencies"],
            np.linspace(10, 20, 100),
        )


def test_sine_environment_stop_environment(sine_environment, sine_metadata):
    initialized_environment(sine_environment, sine_metadata)

    sine_environment.stop_environment(None)

    (siggen_message,) = get_queue_messages(
        sine_environment.queue_container.signal_generation_command_queue,
        ENVIRONMENT_NAME,
        1,
    )
    assert siggen_message == (SignalGenerationCommands.START_SHUTDOWN, None)


def test_sine_environment_shutdown(sine_environment, sine_metadata):
    initialized_environment(sine_environment, sine_metadata)
    sine_environment.set_active()
    sine_environment.control_startup = False

    sine_environment.shutdown()

    assert not sine_environment.active
    assert sine_environment.control_startup
    gui_message = sine_environment.gui_update_queue.get(timeout=10)
    assert gui_message == (
        ENVIRONMENT_NAME,
        (UICommands.ENVIRONMENT_ENDED, None),
    )


def test_sine_environment_quit(sine_environment):
    halt_flag = sine_environment.quit(None)

    assert halt_flag is True
    queues = sine_environment.queue_container
    for queue in (
        queues.data_analysis_command_queue,
        queues.collector_command_queue,
        queues.signal_generation_command_queue,
        queues.spectral_command_queue,
    ):
        (message,) = get_queue_messages(queue, ENVIRONMENT_NAME, 1)
        assert message == (GlobalCommands.QUIT, None)


# region Run loop
@mock.patch(
    "rattlesnake.environment.abstract_environment.Environment.log",
    new=mock.MagicMock(),
)
@mock.patch.object(SineEnvironment, "perform_control_prediction")
def test_sine_environment_run_full_lifecycle(
    mock_prediction, sine_environment, sine_metadata
):
    """Replays the controller's real message sequence through run().

    The control prediction is patched out because it requires excitation
    state produced by the live control loop.  VerboseMessageQueue.get is
    mocked with a finite side_effect so exhaustion raises instead of
    hanging if the loop misbehaves.
    """
    hardware_metadata = numeric_hardware_metadata()
    sysid_metadata = sine_metadata.sysid_metadata
    data_package = sysid_measurement_data_package(sysid_metadata)

    controller_sequence = [
        (GlobalCommands.INITIALIZE_HARDWARE, hardware_metadata),
        (GlobalCommands.INITIALIZE_ENVIRONMENT, sine_metadata),
        (GlobalCommands.INITIALIZE_SYSTEM_ID, sysid_metadata),
        (GlobalCommands.START_SYSTEM_ID_TRANSFER, None),
        (
            SysIdDataAnalysisCommands.SYSTEM_ID_COMPLETE,
            (sysid_metadata, data_package),
        ),
        (GlobalCommands.STOP_SYSTEM_ID, False),
        (SignalGenerationCommands.SHUTDOWN_ACHIEVED, None),
        (DataCollectorCommands.SHUTDOWN_ACHIEVED, None),
        (SpectralProcessingCommands.SHUTDOWN_ACHIEVED, None),
        (SysIdDataAnalysisCommands.SHUTDOWN_ACHIEVED, None),
        (SystemIdCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None),
        (SineCommands.STOP_CONTROL, None),
        (GlobalCommands.QUIT, None),
    ]

    mock_shutdown = mock.MagicMock()
    mock_shutdown.is_set.return_value = False
    with mock.patch(
        "rattlesnake.utilities.VerboseMessageQueue.get",
        side_effect=controller_sequence,
    ):
        sine_environment.run(mock_shutdown)

    assert sine_environment.sysid_data == data_package
    assert sine_environment.sysid_stored
    assert not sine_environment.sysid_active
    mock_prediction.assert_called_with(True)

    queues = sine_environment.queue_container
    for queue in (
        queues.data_analysis_command_queue,
        queues.collector_command_queue,
        queues.signal_generation_command_queue,
        queues.spectral_command_queue,
    ):
        commands = drain_queue_commands(queue, ENVIRONMENT_NAME)
        assert commands[-1] == GlobalCommands.QUIT


# region sine_process
@pytest.mark.parametrize("use_thread", [True, False])
@mock.patch("rattlesnake.environment.sine_sys_id_environment.mp.Process")
@mock.patch("rattlesnake.environment.sine_sys_id_environment.threading.Thread")
@mock.patch("rattlesnake.environment.sine_sys_id_environment.SineEnvironment")
def test_sine_process(mock_environment_class, mock_thread, mock_process, use_thread):
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)

    sine_process(
        ENVIRONMENT_NAME,
        QUEUE_NAME,
        queue_container.environment_command_queues[QUEUE_NAME],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.log_file_queue,
        queue_container.environment_data_in_queues[QUEUE_NAME],
        queue_container.environment_data_out_queues[QUEUE_NAME],
        event_container.acquisition_active_event,
        event_container.output_active_event,
        event_container.environment_active_events[QUEUE_NAME],
        event_container.environment_ready_events[QUEUE_NAME],
        event_container.environment_close_events[QUEUE_NAME],
        event_container.environment_sysid_active_events[QUEUE_NAME],
        event_container.environment_sysid_stored_events[QUEUE_NAME],
        event_container.ping_alive_event,
        use_thread,
    )

    mock_environment_class.return_value.run.assert_called()
    worker_class = mock_thread if use_thread else mock_process
    assert worker_class.call_count == 4
