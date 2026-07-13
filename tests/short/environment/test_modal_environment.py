import multiprocessing as mp
import threading
from unittest import mock

import netCDF4 as nc4
import numpy as np
import openpyxl
import pytest

from rattlesnake.environment.abstract_environment import (
    Environment,
    EnvironmentInstructions,
    EnvironmentMetadata,
)
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.environment.modal_environment import (
    CONTROL_TYPE,
    ModalCommands,
    ModalEnvironment,
    ModalInstructions,
    ModalMetadata,
    ModalQueues,
    ModalUICommands,
    modal_process,
)
from rattlesnake.process.data_collector import (
    Acceptance,
    AcquisitionType,
    CollectorMetadata,
    DataCollectorCommands,
    TriggerSlope,
    Window,
)
from rattlesnake.process.signal_generation import (
    BurstRandomSignalGenerator,
    ChirpSignalGenerator,
    PseudorandomSignalGenerator,
    RandomSignalGenerator,
    SineSignalGenerator,
    SquareSignalGenerator,
)
from rattlesnake.process.signal_generation_process import (
    SignalGenerationCommands,
    SignalGenerationMetadata,
)
from rattlesnake.process.spectral_processing import (
    AveragingTypes,
    Estimator,
    SpectralProcessingCommands,
    SpectralProcessingMetadata,
)
from rattlesnake.testing.mock_utilities import (
    mock_channel_list_bools,
    mock_event_container,
    mock_queue_container,
    skeleton_hardware_metadata,
)
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.utilities import GlobalCommands, RattlesnakeError, VerboseMessageQueue


# region Fixtures
@pytest.fixture
def hardware_metadata():
    return skeleton_hardware_metadata()


@pytest.fixture
def modal_metadata(hardware_metadata):
    return ModalMetadata(
        environment_name="Modal Environment",
        channel_list_bools=mock_channel_list_bools(),
        sample_rate=hardware_metadata.sample_rate,
        samples_per_frame=1024,
        averaging_type="Linear",
        num_averages=8,
        averaging_coefficient=0.25,
        frf_technique="H1",
        frf_window="rectangle",
        overlap_percent=50.0,
        trigger_type="Free Run",
        accept_type="Accept All",
        wait_for_steady_state=0.25,
        trigger_channel=0,
        pretrigger_percent=10.0,
        trigger_slope_positive=True,
        trigger_level_percent=10.0,
        hysteresis_level_percent=5.0,
        hysteresis_frame_percent=20.0,
        signal_generator_type="none",
        signal_generator_level=1.0,
        signal_generator_min_frequency=1.0,
        signal_generator_max_frequency=100.0,
        signal_generator_on_percent=50.0,
        acceptance_function=None,
        reference_channel_indices=[1],
        response_channel_indices=[0],
        output_channel_indices=[1],
        output_oversample=hardware_metadata.output_oversample,
        exponential_window_value_at_frame_end=0.25,
    )


@pytest.fixture(params=[True, False], ids=["threaded", "non_threaded"])
def modal_queues(request):
    use_thread = request.param
    queue_container = mock_queue_container(use_thread)

    return ModalQueues(
        "Modal Environment",
        queue_container.environment_command_queues["Environment 0"],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues["Environment 0"],
        queue_container.environment_data_out_queues["Environment 0"],
        queue_container.log_file_queue,
        use_thread,
    )


@pytest.fixture(params=[True, False], ids=["threaded", "non_threaded"])
def modal_environment(request):
    use_thread = request.param
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)

    modal_queues = ModalQueues(
        "Modal Environment",
        queue_container.environment_command_queues["Environment 0"],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues["Environment 0"],
        queue_container.environment_data_out_queues["Environment 0"],
        queue_container.log_file_queue,
        use_thread,
    )

    return ModalEnvironment(
        "Modal Environment",
        "Environment 0",
        modal_queues,
        event_container.acquisition_active_event,
        event_container.output_active_event,
        event_container.environment_active_events["Environment 0"],
        event_container.environment_ready_events["Environment 0"],
    )


# endregion


# region Commands
def test_modal_commands_have_expected_values():
    """
    Verifies that modal command values are expected integers.
    """
    assert ModalCommands.ACCEPT_FRAME.value == 2
    assert ModalCommands.RUN_CONTROL.value == 3
    assert ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN.value == 4


def test_modal_commands_valid_profile_commands():
    """
    Verifies that modal commands do not define profile commands.
    """
    assert ModalCommands.valid_profile_commands() == ()


def test_modal_commands_valid_data():
    """
    Verifies that modal command data requirements are defined correctly.
    """
    assert ModalCommands.valid_data() == {
        ModalCommands.ACCEPT_FRAME: int,
        ModalCommands.RUN_CONTROL: type(None),
        ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN: type(None),
    }


def test_modal_ui_commands_have_unique_integer_values():
    """
    Verifies that modal UI command values are unique integers.
    """
    values = [command.value for command in ModalUICommands]

    assert all(isinstance(value, int) for value in values)
    assert len(values) == len(set(values))
    assert ModalUICommands.SPECTRAL_UPDATE.value == 1


# endregion


# region Metadata
def test_modal_metadata_init(modal_metadata):
    """
    Verifies that modal metadata initializes and derived properties are
    computed correctly.
    """
    assert isinstance(modal_metadata, ModalMetadata)
    assert isinstance(modal_metadata, EnvironmentMetadata)

    assert modal_metadata.environment_type == CONTROL_TYPE
    assert modal_metadata.environment_name == "Modal Environment"

    assert modal_metadata.overlap == 0.5
    assert modal_metadata.pretrigger == 0.1
    assert modal_metadata.trigger_level == 0.1
    assert modal_metadata.hysteresis_level == 0.05
    assert modal_metadata.hysteresis_length == 0.2
    assert modal_metadata.signal_generator_on_fraction == 0.5

    assert modal_metadata.samples_per_acquire == 512
    assert modal_metadata.frame_time == pytest.approx(
        modal_metadata.samples_per_frame / modal_metadata.sample_rate
    )
    assert modal_metadata.nyquist_frequency == modal_metadata.sample_rate / 2
    assert modal_metadata.fft_lines == modal_metadata.samples_per_frame // 2 + 1
    assert modal_metadata.frequency_spacing == pytest.approx(
        modal_metadata.sample_rate / modal_metadata.samples_per_frame
    )
    assert modal_metadata.skip_frames == int(
        np.ceil(
            modal_metadata.wait_for_steady_state
            * modal_metadata.sample_rate
            / (modal_metadata.samples_per_frame * (1 - modal_metadata.overlap))
        )
    )
    assert modal_metadata.disabled_signals == []
    assert modal_metadata.hysteresis_samples == int(0.2 * 1024)


@pytest.mark.parametrize(
    "signal_generator_type, expected_class",
    [
        ("none", PseudorandomSignalGenerator),
        ("random", RandomSignalGenerator),
        ("pseudorandom", PseudorandomSignalGenerator),
        ("burst", BurstRandomSignalGenerator),
        ("chirp", ChirpSignalGenerator),
        ("square", SquareSignalGenerator),
        ("sine", SineSignalGenerator),
    ],
)
def test_modal_metadata_get_signal_generator(
    modal_metadata,
    signal_generator_type,
    expected_class,
):
    """
    Verifies that supported modal signal generator types construct expected
    signal generator objects.
    """
    modal_metadata.signal_generator_type = signal_generator_type

    signal_generator = modal_metadata.get_signal_generator()

    assert isinstance(signal_generator, expected_class)


def test_modal_metadata_get_signal_generator_invalid(modal_metadata):
    """
    Verifies that invalid signal generator types return ``None``.
    """
    modal_metadata.signal_generator_type = "invalid"

    assert modal_metadata.get_signal_generator() is None


def test_modal_metadata_get_trigger_levels(modal_metadata, hardware_metadata):
    """
    Verifies conversion of trigger and hysteresis levels to volts and
    engineering units.
    """
    trigger_v, trigger_eu, hysteresis_v, hysteresis_eu = (
        modal_metadata.get_trigger_levels(hardware_metadata.channel_list)
    )

    assert trigger_v == pytest.approx(1.0)
    assert trigger_eu == pytest.approx(1.0)
    assert hysteresis_v == pytest.approx(0.5)
    assert hysteresis_eu == pytest.approx(0.5)


def test_modal_metadata_get_trigger_levels_zero_channel_values(modal_metadata):
    """
    Verifies trigger level conversion uses defaults when channel values are
    zero.
    """
    channel = mock.MagicMock()
    channel.maximum_value = 0.0
    channel.sensitivity = 0.0

    trigger_v, trigger_eu, hysteresis_v, hysteresis_eu = (
        modal_metadata.get_trigger_levels([channel])
    )

    assert trigger_v == pytest.approx(1.0)
    assert trigger_eu == pytest.approx(1.0)
    assert hysteresis_v == pytest.approx(0.5)
    assert hysteresis_eu == pytest.approx(0.5)


def test_modal_metadata_disabled_signals(modal_metadata):
    """
    Verifies disabled signal calculation.
    """
    modal_metadata.output_channel_indices = [0, 1, 2]
    modal_metadata.response_channel_indices = [0]
    modal_metadata.reference_channel_indices = [2]

    assert modal_metadata.disabled_signals == [1]


@pytest.mark.parametrize("signal_generator", [None, mock.MagicMock()])
def test_modal_metadata_generate_signal(modal_metadata, signal_generator):
    """
    Verifies generated signal behavior with and without a configured signal
    generator.
    """
    modal_metadata.signal_generator = signal_generator

    signal = modal_metadata.generate_signal()

    if signal_generator is None:
        np.testing.assert_array_equal(
            signal,
            np.zeros(
                (
                    len(modal_metadata.output_channel_indices),
                    modal_metadata.samples_per_frame * modal_metadata.output_oversample,
                )
            ),
        )
    else:
        signal_generator.generate_frame.assert_called_once_with()
        assert signal is signal_generator.generate_frame.return_value[0]


def test_modal_metadata_validate(modal_metadata, hardware_metadata):
    """
    Verifies that modal metadata delegates common validation successfully.
    """
    assert modal_metadata.validate(hardware_metadata) is None


def test_modal_metadata_save_and_load_netcdf(modal_metadata, hardware_metadata):
    """
    Verifies that modal metadata can be saved to and loaded from netCDF.
    """
    dataset = nc4.Dataset("modal.nc", mode="w", diskless=True, persist=False)
    group = dataset.createGroup("Modal Environment")

    try:
        modal_metadata.save_metadata_to_netcdf(group)

        assert group.samples_per_frame == modal_metadata.samples_per_frame
        assert group.averaging_type == modal_metadata.averaging_type
        assert group.num_averages == modal_metadata.num_averages
        assert group.frf_technique == modal_metadata.frf_technique
        assert group.frf_window == modal_metadata.frf_window
        assert "reference_channel_indices" in group.variables
        assert "response_channel_indices" in group.variables

        loaded = ModalMetadata.load_metadata_from_netcdf(
            group,
            "Modal Environment",
            mock_channel_list_bools(),
            hardware_metadata,
        )

        assert isinstance(loaded, ModalMetadata)
        assert loaded.environment_name == "Modal Environment"
        assert loaded.channel_list_bools == mock_channel_list_bools()
        assert loaded.sample_rate == hardware_metadata.sample_rate
        assert loaded.samples_per_frame == modal_metadata.samples_per_frame
        assert loaded.averaging_type == modal_metadata.averaging_type
        assert loaded.num_averages == modal_metadata.num_averages
        assert loaded.frf_technique == modal_metadata.frf_technique
        assert loaded.frf_window == modal_metadata.frf_window
        np.testing.assert_array_equal(
            loaded.reference_channel_indices,
            modal_metadata.reference_channel_indices,
        )
        np.testing.assert_array_equal(
            loaded.response_channel_indices,
            modal_metadata.response_channel_indices,
        )
    finally:
        dataset.close()


def test_modal_metadata_create_blank_worksheet_template():
    """
    Verifies that the blank worksheet template contains modal fields.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    ModalMetadata.create_blank_worksheet_template(worksheet)

    assert worksheet.cell(1, 1).value == "Control Type"
    assert worksheet.cell(1, 2).value == "Modal"
    assert worksheet.cell(2, 1).value == "Samples Per Frame:"
    assert worksheet.cell(6, 1).value == "FRF Technique:"
    assert worksheet.cell(18, 1).value == "Signal Generator Type"
    assert worksheet.cell(26, 1).value == "Reference Channels"
    assert worksheet.cell(27, 1).value == "Disabled Channels"


def test_modal_metadata_save_metadata_to_worksheet(modal_metadata):
    """
    Verifies that modal metadata values are written to a worksheet.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    modal_metadata.save_metadata_to_worksheet(worksheet)

    assert worksheet.cell(1, 2).value == "Modal"
    assert worksheet.cell(2, 2).value == modal_metadata.samples_per_frame
    assert worksheet.cell(3, 2).value == modal_metadata.averaging_type
    assert worksheet.cell(4, 2).value == modal_metadata.num_averages
    assert worksheet.cell(6, 2).value == modal_metadata.frf_technique
    assert worksheet.cell(7, 2).value == modal_metadata.frf_window
    assert worksheet.cell(10, 2).value == modal_metadata.trigger_type
    assert worksheet.cell(11, 2).value == modal_metadata.accept_type
    assert worksheet.cell(12, 2).value == modal_metadata.trigger_channel + 1
    assert worksheet.cell(14, 2).value == "Positive"
    assert worksheet.cell(18, 2).value == modal_metadata.signal_generator_type
    assert worksheet.cell(26, 2).value == 2


def test_modal_metadata_load_metadata_from_worksheet(
    modal_metadata,
    hardware_metadata,
):
    """
    Verifies that modal metadata can be reconstructed from a worksheet.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    modal_metadata.save_metadata_to_worksheet(worksheet)

    loaded = ModalMetadata.load_metadata_from_worksheet(
        worksheet,
        "Modal Environment",
        mock_channel_list_bools(),
        hardware_metadata,
    )

    assert isinstance(loaded, ModalMetadata)
    assert loaded.environment_name == "Modal Environment"
    assert loaded.sample_rate == hardware_metadata.sample_rate
    assert loaded.samples_per_frame == modal_metadata.samples_per_frame
    assert loaded.averaging_type == modal_metadata.averaging_type
    assert loaded.num_averages == modal_metadata.num_averages
    assert loaded.frf_technique == modal_metadata.frf_technique
    assert loaded.frf_window == modal_metadata.frf_window
    assert loaded.overlap == pytest.approx(modal_metadata.overlap)
    assert loaded.trigger_channel == modal_metadata.trigger_channel
    assert loaded.trigger_slope_positive == modal_metadata.trigger_slope_positive
    assert loaded.reference_channel_indices == modal_metadata.reference_channel_indices
    assert loaded.output_channel_indices == modal_metadata.output_channel_indices


# endregion


# region Instructions
def test_modal_instructions_init():
    """
    Verifies that modal instructions initialize as environment instructions.
    """
    instructions = ModalInstructions("Modal Environment")

    assert isinstance(instructions, ModalInstructions)
    assert isinstance(instructions, EnvironmentInstructions)
    assert instructions.environment_type == EnvironmentType.MODAL
    assert instructions.environment_name == "Modal Environment"


def test_modal_instructions_validate():
    """
    Verifies that modal instructions validate without error.
    """
    instructions = ModalInstructions("Modal Environment")

    instructions.validate()


# endregion


# region Queues
@pytest.mark.parametrize("use_thread", [True, False])
def test_modal_queues_init(use_thread):
    """
    Verifies that modal queue containers initialize all expected queues.
    """
    queue_container = mock_queue_container(use_thread)

    modal_queues = ModalQueues(
        "Modal Environment",
        queue_container.environment_command_queues["Environment 0"],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues["Environment 0"],
        queue_container.environment_data_out_queues["Environment 0"],
        queue_container.log_file_queue,
        use_thread,
    )

    assert isinstance(modal_queues, ModalQueues)
    assert (
        modal_queues.environment_command_queue
        is queue_container.environment_command_queues["Environment 0"]
    )
    assert modal_queues.gui_update_queue is queue_container.gui_update_queue
    assert (
        modal_queues.controller_communication_queue
        is queue_container.controller_command_queue
    )
    assert (
        modal_queues.data_in_queue
        is queue_container.environment_data_in_queues["Environment 0"]
    )
    assert (
        modal_queues.data_out_queue
        is queue_container.environment_data_out_queues["Environment 0"]
    )
    assert modal_queues.log_file_queue is queue_container.log_file_queue

    assert hasattr(modal_queues, "data_for_spectral_computation_queue")
    assert hasattr(modal_queues, "updated_spectral_quantities_queue")
    assert hasattr(modal_queues, "signal_generation_update_queue")
    assert isinstance(modal_queues.spectral_command_queue, VerboseMessageQueue)
    assert isinstance(modal_queues.collector_command_queue, VerboseMessageQueue)
    assert isinstance(
        modal_queues.signal_generation_command_queue,
        VerboseMessageQueue,
    )


# endregion


# region Environment
def test_modal_environment_init(modal_environment):
    """
    Verifies that modal environment initializes successfully and maps commands.
    """
    assert isinstance(modal_environment, ModalEnvironment)
    assert isinstance(modal_environment, Environment)

    assert modal_environment.ready is True
    assert modal_environment.active is False
    assert modal_environment.frame_number == 0
    assert modal_environment.siggen_shutdown_achieved is False
    assert modal_environment.collector_shutdown_achieved is False
    assert modal_environment.spectral_shutdown_achieved is False

    assert modal_environment.command_map[ModalCommands.ACCEPT_FRAME] == (
        modal_environment.accept_frame
    )
    assert modal_environment.command_map[GlobalCommands.START_ENVIRONMENT] == (
        modal_environment.start_environment
    )
    assert modal_environment.command_map[ModalCommands.RUN_CONTROL] == (
        modal_environment.run_control
    )
    assert modal_environment.command_map[ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN] == (
        modal_environment.check_for_shutdown
    )


def test_modal_environment_initialize_hardware(
    modal_environment,
    hardware_metadata,
):
    """
    Verifies that hardware metadata is stored and ready event is set.
    """
    modal_environment.clear_ready()

    modal_environment.initialize_hardware(hardware_metadata)

    assert modal_environment.hardware_metadata is hardware_metadata
    assert modal_environment.ready is True


@mock.patch(
    "rattlesnake.environment.modal_environment.ModalEnvironment.get_spectral_processing_metadata"
)
@mock.patch(
    "rattlesnake.environment.modal_environment.ModalEnvironment.get_signal_generation_metadata"
)
@mock.patch(
    "rattlesnake.environment.modal_environment.ModalEnvironment.get_data_collector_metadata"
)
def test_modal_environment_initialize_environment(
    mock_collector_metadata,
    mock_signal_metadata,
    mock_spectral_metadata,
    modal_environment,
    modal_metadata,
):
    """
    Verifies that modal environment metadata is stored and subprocess
    initialization commands are sent.
    """
    mock_collector_metadata.return_value = "collector metadata"
    mock_signal_metadata.return_value = "signal metadata"
    mock_spectral_metadata.return_value = "spectral metadata"

    modal_environment.queue_container.collector_command_queue = mock.MagicMock()
    modal_environment.queue_container.signal_generation_command_queue = mock.MagicMock()
    modal_environment.queue_container.spectral_command_queue = mock.MagicMock()

    modal_environment.clear_ready()

    modal_environment.initialize_environment(modal_metadata)

    assert modal_environment.environment_name == modal_metadata.environment_name
    assert modal_environment.environment_metadata is modal_metadata
    assert modal_environment.ready is True

    modal_environment.queue_container.collector_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (DataCollectorCommands.INITIALIZE_COLLECTOR, "collector metadata"),
    )
    modal_environment.queue_container.signal_generation_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (SignalGenerationCommands.INITIALIZE_PARAMETERS, "signal metadata"),
    )
    modal_environment.queue_container.spectral_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (SpectralProcessingCommands.INITIALIZE_PARAMETERS, "spectral metadata"),
    )


@pytest.mark.parametrize(
    "trigger_type, expected_acquisition_type",
    [
        ("Free Run", AcquisitionType.FREE_RUN),
        ("First Frame", AcquisitionType.TRIGGER_FIRST_FRAME),
        ("Every Frame", AcquisitionType.TRIGGER_EVERY_FRAME),
    ],
)
def test_modal_environment_get_data_collector_metadata_trigger_types(
    modal_environment,
    modal_metadata,
    hardware_metadata,
    trigger_type,
    expected_acquisition_type,
):
    """
    Verifies trigger type conversion in collector metadata.
    """
    modal_environment.hardware_metadata = hardware_metadata
    modal_environment.environment_metadata = modal_metadata
    modal_metadata.trigger_type = trigger_type

    metadata = modal_environment.get_data_collector_metadata()

    assert isinstance(metadata, CollectorMetadata)
    assert metadata.acquisition_type == expected_acquisition_type
    assert metadata.num_channels == len(hardware_metadata.channel_list)
    assert metadata.response_channel_indices == modal_metadata.response_channel_indices
    assert (
        metadata.reference_channel_indices == modal_metadata.reference_channel_indices
    )


@pytest.mark.parametrize(
    "accept_type, expected_acceptance, expected_function",
    [
        ("Accept All", Acceptance.AUTOMATIC, None),
        ("Manual", Acceptance.MANUAL, None),
        ("Autoreject...", Acceptance.AUTOMATIC, ("module.py", "accept")),
    ],
)
def test_modal_environment_get_data_collector_metadata_acceptance_types(
    modal_environment,
    modal_metadata,
    hardware_metadata,
    accept_type,
    expected_acceptance,
    expected_function,
):
    """
    Verifies acceptance type conversion in collector metadata.
    """
    modal_environment.hardware_metadata = hardware_metadata
    modal_environment.environment_metadata = modal_metadata
    modal_metadata.accept_type = accept_type
    modal_metadata.acceptance_function = expected_function

    metadata = modal_environment.get_data_collector_metadata()

    assert metadata.acceptance == expected_acceptance
    assert metadata.acceptance_function == expected_function


@pytest.mark.parametrize(
    "frf_window, expected_window",
    [
        ("hann", Window.HANN),
        ("rectangle", Window.RECTANGLE),
        ("exponential", Window.EXPONENTIAL),
    ],
)
def test_modal_environment_get_data_collector_metadata_windows(
    modal_environment,
    modal_metadata,
    hardware_metadata,
    frf_window,
    expected_window,
):
    """
    Verifies FRF window conversion in collector metadata.
    """
    modal_environment.hardware_metadata = hardware_metadata
    modal_environment.environment_metadata = modal_metadata
    modal_metadata.frf_window = frf_window

    metadata = modal_environment.get_data_collector_metadata()

    assert metadata.window == expected_window


@pytest.mark.parametrize(
    "field_name, invalid_value",
    [
        ("trigger_type", "Invalid Trigger"),
        ("accept_type", "Invalid Acceptance"),
        ("frf_window", "invalid_window"),
    ],
)
def test_modal_environment_get_data_collector_metadata_invalid_values(
    modal_environment,
    modal_metadata,
    hardware_metadata,
    field_name,
    invalid_value,
):
    """
    Verifies invalid collector metadata options raise ``ValueError``.
    """
    modal_environment.hardware_metadata = hardware_metadata
    modal_environment.environment_metadata = modal_metadata
    setattr(modal_metadata, field_name, invalid_value)

    with pytest.raises(ValueError):
        modal_environment.get_data_collector_metadata()


@pytest.mark.parametrize(
    "averaging_type, expected_averaging_type",
    [
        ("Linear", AveragingTypes.LINEAR),
        ("Exponential", AveragingTypes.EXPONENTIAL),
    ],
)
@pytest.mark.parametrize(
    "frf_technique, expected_estimator",
    [
        ("H1", Estimator.H1),
        ("H2", Estimator.H2),
        ("H3", Estimator.H3),
        ("Hv", Estimator.HV),
    ],
)
def test_modal_environment_get_spectral_processing_metadata(
    modal_environment,
    modal_metadata,
    averaging_type,
    expected_averaging_type,
    frf_technique,
    expected_estimator,
):
    """
    Verifies spectral processing metadata conversion.
    """
    modal_environment.environment_metadata = modal_metadata
    modal_metadata.averaging_type = averaging_type
    modal_metadata.frf_technique = frf_technique

    metadata = modal_environment.get_spectral_processing_metadata()

    assert isinstance(metadata, SpectralProcessingMetadata)
    assert metadata.averaging_type == expected_averaging_type
    assert metadata.frf_estimator == expected_estimator
    assert metadata.averages == modal_metadata.num_averages
    assert metadata.frequency_spacing == modal_metadata.frequency_spacing
    assert metadata.sample_rate == modal_metadata.sample_rate
    assert metadata.num_frequency_lines == modal_metadata.fft_lines


def test_modal_environment_get_spectral_processing_metadata_invalid_estimator(
    modal_environment,
    modal_metadata,
):
    """
    Verifies invalid FRF estimator raises ``ValueError``.
    """
    modal_environment.environment_metadata = modal_metadata
    modal_metadata.frf_technique = "Invalid"

    with pytest.raises(ValueError):
        modal_environment.get_spectral_processing_metadata()


def test_modal_environment_get_signal_generation_metadata(
    modal_environment,
    modal_metadata,
    hardware_metadata,
):
    """
    Verifies signal generation metadata creation.
    """
    modal_environment.hardware_metadata = hardware_metadata
    modal_environment.environment_metadata = modal_metadata

    metadata = modal_environment.get_signal_generation_metadata()

    assert isinstance(metadata, SignalGenerationMetadata)
    assert metadata.samples_per_write == hardware_metadata.samples_per_write


def test_modal_environment_get_signal_generator(modal_environment):
    """
    Verifies delegation to modal metadata signal generator.
    """
    metadata = mock.MagicMock()
    metadata.get_signal_generator.return_value = "signal generator"
    modal_environment.environment_metadata = metadata

    assert modal_environment.get_signal_generator() == "signal generator"
    metadata.get_signal_generator.assert_called_once_with()


@mock.patch("rattlesnake.environment.modal_environment.time.sleep")
@mock.patch(
    "rattlesnake.environment.modal_environment.ModalEnvironment.get_spectral_processing_metadata"
)
@mock.patch(
    "rattlesnake.environment.modal_environment.ModalEnvironment.get_signal_generator"
)
@mock.patch(
    "rattlesnake.environment.modal_environment.ModalEnvironment.get_signal_generation_metadata"
)
@mock.patch(
    "rattlesnake.environment.modal_environment.ModalEnvironment.get_data_collector_metadata"
)
@mock.patch("rattlesnake.environment.modal_environment.ModalEnvironment.log")
def test_modal_environment_start_environment(
    mock_log,
    mock_collector_metadata,
    mock_signal_metadata,
    mock_signal_generator,
    mock_spectral_metadata,
    mock_sleep,
    modal_environment,
    modal_metadata,
):
    """
    Verifies modal startup commands and GUI update behavior.
    """
    modal_environment.environment_metadata = modal_metadata

    modal_environment.queue_container.collector_command_queue = mock.MagicMock()
    modal_environment.queue_container.signal_generation_command_queue = mock.MagicMock()
    modal_environment.queue_container.spectral_command_queue = mock.MagicMock()
    modal_environment.queue_container.environment_command_queue = mock.MagicMock()
    modal_environment.queue_container.gui_update_queue = mock.MagicMock()

    mock_collector_metadata.return_value = "collector metadata"
    mock_signal_metadata.return_value = "signal metadata"
    mock_signal_generator.return_value = "signal generator"
    mock_spectral_metadata.return_value = "spectral metadata"

    modal_environment.start_environment(None)

    mock_log.assert_called_with("Starting Modal")
    assert modal_environment.siggen_shutdown_achieved is False
    assert modal_environment.collector_shutdown_achieved is False
    assert modal_environment.spectral_shutdown_achieved is False
    assert modal_environment.active is True

    collector_puts = modal_environment.queue_container.collector_command_queue.put
    collector_puts.assert_has_calls(
        [
            mock.call(
                "Modal Environment",
                (
                    DataCollectorCommands.FORCE_INITIALIZE_COLLECTOR,
                    "collector metadata",
                ),
            ),
            mock.call(
                "Modal Environment",
                (
                    DataCollectorCommands.SET_TEST_LEVEL,
                    (modal_metadata.skip_frames, 1),
                ),
            ),
            mock.call(
                "Modal Environment",
                (DataCollectorCommands.ACQUIRE, None),
            ),
        ]
    )

    signal_puts = modal_environment.queue_container.signal_generation_command_queue.put
    signal_puts.assert_has_calls(
        [
            mock.call(
                "Modal Environment",
                (
                    SignalGenerationCommands.INITIALIZE_PARAMETERS,
                    "signal metadata",
                ),
            ),
            mock.call(
                "Modal Environment",
                (
                    SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
                    "signal generator",
                ),
            ),
            mock.call("Modal Environment", (SignalGenerationCommands.MUTE, None)),
            mock.call(
                "Modal Environment",
                (SignalGenerationCommands.ADJUST_TEST_LEVEL, 1.0),
            ),
            mock.call(
                "Modal Environment",
                (SignalGenerationCommands.GENERATE_SIGNALS, None),
            ),
        ]
    )

    spectral_puts = modal_environment.queue_container.spectral_command_queue.put
    spectral_puts.assert_has_calls(
        [
            mock.call(
                "Modal Environment",
                (
                    SpectralProcessingCommands.INITIALIZE_PARAMETERS,
                    "spectral metadata",
                ),
            ),
            mock.call(
                "Modal Environment",
                (SpectralProcessingCommands.CLEAR_SPECTRAL_PROCESSING, None),
            ),
            mock.call(
                "Modal Environment",
                (SpectralProcessingCommands.RUN_SPECTRAL_PROCESSING, None),
            ),
        ]
    )

    modal_environment.queue_container.environment_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (ModalCommands.RUN_CONTROL, None),
    )
    modal_environment.queue_container.gui_update_queue.put.assert_called_once_with(
        ("Modal Environment", (UICommands.ENVIRONMENT_STARTED, None))
    )


@mock.patch("rattlesnake.environment.modal_environment.flush_queue")
@mock.patch("rattlesnake.environment.modal_environment.ModalEnvironment.log")
def test_modal_environment_run_control_with_spectral_data(
    mock_log,
    mock_flush_queue,
    modal_environment,
    modal_metadata,
):
    """
    Verifies spectral data forwarding to the GUI and control-loop requeueing.
    """
    modal_environment.environment_metadata = modal_metadata
    modal_environment._gui_update_queue = mock.MagicMock()
    modal_environment.queue_container.environment_command_queue = mock.MagicMock()

    spectral_data = (
        "frames",
        "frequencies",
        "frf",
        "coherence",
        "response_cpsd",
        "reference_cpsd",
        "condition",
    )
    mock_flush_queue.return_value = [spectral_data]

    modal_environment.run_control(None)

    mock_log.assert_called_with("Received Data")
    modal_environment.gui_update_queue.put.assert_called_once_with(
        (
            "Modal Environment",
            (
                ModalUICommands.SPECTRAL_UPDATE,
                (
                    "frames",
                    modal_metadata.num_averages,
                    "frequencies",
                    "frf",
                    "coherence",
                    "response_cpsd",
                    "reference_cpsd",
                    "condition",
                ),
            ),
        )
    )
    modal_environment.queue_container.environment_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (ModalCommands.RUN_CONTROL, None),
    )


@mock.patch("rattlesnake.environment.modal_environment.time.sleep")
@mock.patch("rattlesnake.environment.modal_environment.flush_queue")
def test_modal_environment_run_control_without_spectral_data(
    mock_flush_queue,
    mock_sleep,
    modal_environment,
):
    """
    Verifies that the control loop waits and requeues when no spectral data are
    available.
    """
    modal_environment.queue_container.environment_command_queue = mock.MagicMock()
    mock_flush_queue.return_value = []

    modal_environment.run_control(None)

    mock_sleep.assert_called_once()
    modal_environment.queue_container.environment_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (ModalCommands.RUN_CONTROL, None),
    )


def test_modal_environment_siggen_shutdown_achieved_fn(modal_environment):
    """
    Verifies that signal generation shutdown flag is set.
    """
    modal_environment.siggen_shutdown_achieved_fn(None)

    assert modal_environment.siggen_shutdown_achieved is True


def test_modal_environment_collector_shutdown_achieved_fn(modal_environment):
    """
    Verifies that collector shutdown flag is set.
    """
    modal_environment.collector_shutdown_achieved_fn(None)

    assert modal_environment.collector_shutdown_achieved is True


def test_modal_environment_spectral_shutdown_achieved_fn(modal_environment):
    """
    Verifies that spectral shutdown flag is set.
    """
    modal_environment.spectral_shutdown_achieved_fn(None)

    assert modal_environment.spectral_shutdown_achieved is True


@mock.patch("rattlesnake.environment.modal_environment.ModalEnvironment.log")
def test_modal_environment_check_for_shutdown_complete(
    mock_log,
    modal_environment,
):
    """
    Verifies complete-shutdown behavior.
    """
    modal_environment.siggen_shutdown_achieved = True
    modal_environment.collector_shutdown_achieved = True
    modal_environment.spectral_shutdown_achieved = True
    modal_environment.set_active()
    modal_environment.queue_container.gui_update_queue = mock.MagicMock()

    modal_environment.check_for_shutdown(None)

    mock_log.assert_called_with("Shutdown Achieved")
    assert modal_environment.active is False
    modal_environment.queue_container.gui_update_queue.put.assert_called_once_with(
        ("Modal Environment", (UICommands.ENVIRONMENT_ENDED, None))
    )


@mock.patch("rattlesnake.environment.modal_environment.time.sleep")
def test_modal_environment_check_for_shutdown_incomplete(
    mock_sleep,
    modal_environment,
):
    """
    Verifies incomplete-shutdown requeue behavior.
    """
    modal_environment.siggen_shutdown_achieved = False
    modal_environment.collector_shutdown_achieved = True
    modal_environment.spectral_shutdown_achieved = True
    modal_environment._command_queue = mock.MagicMock()

    modal_environment.check_for_shutdown(None)

    mock_sleep.assert_called_once_with(1)
    modal_environment.environment_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None),
    )


def test_modal_environment_accept_frame(modal_environment):
    """
    Verifies that frame acceptance is forwarded to collector.
    """
    modal_environment.queue_container.collector_command_queue = mock.MagicMock()

    modal_environment.accept_frame(True)

    modal_environment.queue_container.collector_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (DataCollectorCommands.ACCEPT, True),
    )


@mock.patch("rattlesnake.environment.modal_environment.flush_queue")
@mock.patch("rattlesnake.environment.modal_environment.ModalEnvironment.log")
def test_modal_environment_stop_environment(
    mock_log,
    mock_flush_queue,
    modal_environment,
):
    """
    Verifies modal graceful shutdown commands.
    """
    modal_environment.queue_container.environment_command_queue = mock.MagicMock()
    modal_environment.queue_container.collector_command_queue = mock.MagicMock()
    modal_environment.queue_container.signal_generation_command_queue = mock.MagicMock()
    modal_environment.queue_container.spectral_command_queue = mock.MagicMock()

    modal_environment.stop_environment(None)

    mock_log.assert_called_with("Stopping Control")
    mock_flush_queue.assert_called_once_with(
        modal_environment.queue_container.environment_command_queue
    )

    modal_environment.queue_container.collector_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (DataCollectorCommands.SET_TEST_LEVEL, (1000, 1)),
    )
    modal_environment.queue_container.signal_generation_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (SignalGenerationCommands.START_SHUTDOWN, None),
    )
    modal_environment.queue_container.spectral_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (SpectralProcessingCommands.STOP_SPECTRAL_PROCESSING, None),
    )
    modal_environment.queue_container.environment_command_queue.put.assert_called_once_with(
        "Modal Environment",
        (ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None),
    )


def test_modal_environment_quit(modal_environment):
    """
    Verifies that quit commands are sent to modal subprocesses.
    """
    modal_environment.queue_container.spectral_command_queue = mock.MagicMock()
    modal_environment.queue_container.signal_generation_command_queue = mock.MagicMock()
    modal_environment.queue_container.collector_command_queue = mock.MagicMock()

    result = modal_environment.quit(None)

    assert result is True
    for command_queue in [
        modal_environment.queue_container.spectral_command_queue,
        modal_environment.queue_container.signal_generation_command_queue,
        modal_environment.queue_container.collector_command_queue,
    ]:
        command_queue.put.assert_called_once_with(
            "Modal Environment",
            (GlobalCommands.QUIT, None),
        )


# endregion


# region Process
@pytest.mark.parametrize("use_thread", [True, False])
@mock.patch("rattlesnake.environment.modal_environment.ModalEnvironment")
@mock.patch("rattlesnake.environment.modal_environment.data_collector_process")
@mock.patch("rattlesnake.environment.modal_environment.signal_generation_process")
@mock.patch("rattlesnake.environment.modal_environment.spectral_processing_process")
def test_modal_process(
    mock_spectral_processing_process,
    mock_signal_generation_process,
    mock_data_collector_process,
    mock_modal_environment_class,
    use_thread,
):
    """
    Verifies that modal subprocesses are started, the modal environment is run,
    and subprocesses are joined.
    """
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)

    mock_process = mock.MagicMock()
    mock_process_class = mock.MagicMock(return_value=mock_process)

    patch_target = (
        "rattlesnake.environment.modal_environment.threading.Thread"
        if use_thread
        else "rattlesnake.environment.modal_environment.mp.Process"
    )

    with mock.patch(patch_target, mock_process_class):
        modal_process(
            "Modal Environment",
            "Environment 0",
            queue_container.environment_command_queues["Environment 0"],
            queue_container.gui_update_queue,
            queue_container.controller_command_queue,
            queue_container.log_file_queue,
            queue_container.environment_data_in_queues["Environment 0"],
            queue_container.environment_data_out_queues["Environment 0"],
            event_container.acquisition_active_event,
            event_container.output_active_event,
            event_container.environment_active_events["Environment 0"],
            event_container.environment_ready_events["Environment 0"],
            event_container.environment_close_events["Environment 0"],
            event_container.environment_sysid_active_events["Environment 0"],
            event_container.environment_sysid_stored_events["Environment 0"],
            event_container.ping_alive_event,
            use_thread,
        )

    assert mock_process_class.call_count == 3
    assert mock_process.start.call_count == 3
    assert mock_process.join.call_count == 3

    mock_modal_environment_class.assert_called_once()
    mock_modal_environment_class.return_value.run.assert_called_once_with(
        event_container.environment_close_events["Environment 0"]
    )
    mock_modal_environment_class.return_value.log.assert_has_calls(
        [
            mock.call("Joining Subprocesses"),
            mock.call("Joining Spectral Computation"),
            mock.call("Joining Signal Generation"),
            mock.call("Joining Data Collection"),
        ]
    )


# endregion
