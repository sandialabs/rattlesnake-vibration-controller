"""Tests for the random vibration sys-id environment.

These tests drive the metadata file flows (worksheet, netCDF, specification
.npz) and the environment lifecycle the controller uses in a real test
(initialize hardware/environment/sysid -> transfer function ->
SYSTEM_ID_COMPLETE -> start/stop control -> shutdown handshake -> QUIT),
using hand-fed queues with no spawned sub-workers.

The closed-loop control math itself lives in
rattlesnake.process.random_vibration_sys_id_data_analysis and is exercised
end-to-end by tests/long/test_qualification.py, so it is not covered here.
"""

from unittest import mock

import netCDF4 as nc4
import numpy as np
import openpyxl
import pytest

from rattlesnake.environment.abstract_environment import (
    EnvironmentInstructions,
    EnvironmentMetadata,
)
from rattlesnake.environment.abstract_sysid_environment import (
    SysIdEnvironment,
    SysIdUICommands,
    SystemIdCommands,
)
from rattlesnake.environment.random_vibration_sys_id_environment import (
    RandomVibrationCommands,
    RandomVibrationEnvironment,
    RandomVibrationInstructions,
    RandomVibrationMetadata,
    RandomVibrationQueues,
    RandomVibrationUICommands,
    random_vibration_process,
)
from rattlesnake.process.abstract_sysid_data_analysis import (
    SysIdDataAnalysisCommands,
    SysIdDataAnalysisUICommands,
    SysIdMetadata,
)
from rattlesnake.process.data_collector import DataCollectorCommands
from rattlesnake.process.random_vibration_sys_id_data_analysis import (
    RandomVibrationDataAnalysisCommands,
)
from rattlesnake.process.signal_generation import (
    CPSDSignalGenerator,
    RandomSignalGenerator,
)
from rattlesnake.process.signal_generation_process import SignalGenerationCommands
from rattlesnake.process.spectral_processing import SpectralProcessingCommands
from rattlesnake.testing.mock_utilities import (
    drain_queue_commands,
    get_queue_messages,
    sysid_measurement_data_package,
    sysid_measurement_metadata,
    numeric_channel_list,
    numeric_hardware_metadata,
    outer_coordinate,
    write_cpsd_spec_npz,
    mock_event_container,
    mock_queue_container,
)
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.utilities import (
    GlobalCommands,
    RattlesnakeError,
    db2scale,
    flush_queue,
)

ENVIRONMENT_NAME = "Random Environment"
QUEUE_NAME = "Environment 0"


def flush_environment_queues(environment):
    """Drains every queue so no multiprocessing feeder thread is left
    blocked on a full pipe, which would hang the interpreter at exit."""
    queues = environment.queue_container
    for queue in (
        queues.gui_update_queue,
        queues.environment_command_queue,
        queues.data_analysis_command_queue,
        queues.collector_command_queue,
        queues.signal_generation_command_queue,
        queues.spectral_command_queue,
    ):
        flush_queue(queue, timeout=0.05)


# region Fixtures
def make_random_metadata():
    fft_lines = 101
    return RandomVibrationMetadata(
        environment_name=ENVIRONMENT_NAME,
        channel_list_bools=[True, True],
        sample_rate=1000,
        number_of_channels=2,
        samples_per_frame=200,
        test_level_ramp_time=0.1,
        cola_window="hann",
        cola_overlap=0.5,
        cola_window_exponent=0.5,
        sigma_clip=5.0,
        update_tf_during_control=False,
        frames_in_cpsd=20,
        cpsd_window="Hann",
        cpsd_overlap=0.5,
        percent_lines_out=0.1,
        allow_automatic_aborts=False,
        control_python_script="",
        control_python_function="",
        control_python_function_type=0,
        control_python_function_parameters="",
        control_channel_indices=[0],
        output_channel_indices=[1],
        specification_frequency_lines=5.0 * np.arange(fft_lines),
        specification_cpsd_matrix=np.ones((fft_lines, 1, 1), dtype="complex128"),
        specification_warning_matrix=np.full((2, fft_lines, 1), np.nan),
        specification_abort_matrix=np.full((2, fft_lines, 1), np.nan),
        response_transformation_matrix=None,
        output_transformation_matrix=None,
        sysid_metadata=sysid_measurement_metadata(),
    )


@pytest.fixture
def random_metadata():
    random_metadata = make_random_metadata()
    random_metadata.queue_name = QUEUE_NAME

    return random_metadata


@pytest.fixture(params=[True, False], ids=["threaded", "non_threaded"])
def random_environment(request):
    use_thread = request.param
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)
    random_queues = RandomVibrationQueues(
        ENVIRONMENT_NAME,
        queue_container.environment_command_queues[QUEUE_NAME],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues[QUEUE_NAME],
        queue_container.environment_data_out_queues[QUEUE_NAME],
        queue_container.log_file_queue,
        use_thread,
    )
    random_environment = RandomVibrationEnvironment(
        ENVIRONMENT_NAME,
        QUEUE_NAME,
        random_queues,
        event_container.acquisition_active_event,
        event_container.output_active_event,
        event_container.environment_active_events[QUEUE_NAME],
        event_container.environment_ready_events[QUEUE_NAME],
        event_container.environment_sysid_active_events[QUEUE_NAME],
        event_container.environment_sysid_stored_events[QUEUE_NAME],
    )
    yield random_environment
    flush_environment_queues(random_environment)


def initialized_environment(random_environment, random_metadata):
    """Runs the environment through the real hardware/environment init flow."""
    random_environment.initialize_hardware(numeric_hardware_metadata())
    random_environment.initialize_environment(random_metadata)
    # Discard the initialization queue traffic so tests can assert on
    # the traffic of the method under test only
    for queue in (
        random_environment.queue_container.data_analysis_command_queue,
        random_environment.queue_container.collector_command_queue,
        random_environment.queue_container.signal_generation_command_queue,
        random_environment.queue_container.spectral_command_queue,
    ):
        drain_queue_commands(queue, ENVIRONMENT_NAME)
    return random_environment


# region Metadata
def test_random_metadata_init(random_metadata):
    assert isinstance(random_metadata, RandomVibrationMetadata)
    assert isinstance(random_metadata, EnvironmentMetadata)
    assert isinstance(random_metadata.sysid_metadata, SysIdMetadata)


def test_random_metadata_default_sysid():
    random_metadata = make_random_metadata()
    random_metadata.sysid_metadata = None
    rebuilt = RandomVibrationMetadata(
        environment_name=ENVIRONMENT_NAME,
        channel_list_bools=[True, True],
        sample_rate=1000,
        number_of_channels=2,
        samples_per_frame=200,
        test_level_ramp_time=0.1,
        cola_window="hann",
        cola_overlap=0.5,
        cola_window_exponent=0.5,
        sigma_clip=5.0,
        update_tf_during_control=False,
        frames_in_cpsd=20,
        cpsd_window="Hann",
        cpsd_overlap=0.5,
        percent_lines_out=0.1,
        allow_automatic_aborts=False,
        control_python_script="",
        control_python_function="",
        control_python_function_type=None,
        control_python_function_parameters="",
        control_channel_indices=[0],
        output_channel_indices=[1],
        specification_frequency_lines=None,
        specification_cpsd_matrix=None,
        specification_warning_matrix=None,
        specification_abort_matrix=None,
        response_transformation_matrix=None,
        output_transformation_matrix=None,
        sysid_metadata=None,
    )
    assert isinstance(rebuilt.sysid_metadata, SysIdMetadata)
    assert rebuilt.sysid_metadata.sample_rate == 1000


def test_random_metadata_properties(random_metadata):
    assert random_metadata.samples_per_acquire == 100
    assert random_metadata.frame_time == 0.2
    assert random_metadata.nyquist_frequency == 500
    assert random_metadata.fft_lines == 101
    assert random_metadata.frequency_spacing == 5.0
    assert random_metadata.samples_per_output == 100
    assert random_metadata.overlapped_output_samples == 100
    assert random_metadata.skip_frames == 1
    assert random_metadata.response_channel_indices == [0]
    assert random_metadata.reference_channel_indices == [1]
    assert random_metadata.num_response_channels == 1
    assert random_metadata.num_reference_channels == 1


def test_random_metadata_transformed_channel_counts(random_metadata):
    random_metadata.response_transformation_matrix = np.ones((3, 1))
    random_metadata.reference_transformation_matrix = np.ones((2, 1))

    assert random_metadata.num_response_channels == 3
    assert random_metadata.num_reference_channels == 2


@pytest.mark.parametrize(
    "channel_list_bools, environment_name, expected",
    [
        ([True, True], ENVIRONMENT_NAME, True),
        ([True], ENVIRONMENT_NAME, RattlesnakeError),
        ([True, True], 17, RattlesnakeError),
    ],
)
def test_random_metadata_validate(
    channel_list_bools, environment_name, expected, random_metadata
):
    hardware_metadata = numeric_hardware_metadata()
    random_metadata.channel_list_bools = channel_list_bools
    random_metadata.environment_name = environment_name

    if expected is RattlesnakeError:
        with pytest.raises(RattlesnakeError):
            random_metadata.validate(hardware_metadata)
    else:
        random_metadata.validate(hardware_metadata)


# region Specification loading
def test_random_metadata_load_specification(tmp_path, random_metadata):
    spec_path = str(tmp_path / "spec.npz")
    spec_data = write_cpsd_spec_npz(spec_path, 5.0 * np.arange(101), 1)

    random_metadata.load_specification(numeric_channel_list(), spec_path)

    np.testing.assert_allclose(
        random_metadata.specification_frequency_lines, 5.0 * np.arange(101)
    )
    np.testing.assert_allclose(
        random_metadata.specification_cpsd_matrix, spec_data["cpsd"]
    )
    assert random_metadata.specification_warning_matrix.shape == (2, 101, 1)
    np.testing.assert_allclose(
        random_metadata.specification_warning_matrix[0, :, 0],
        spec_data["warning_lower"][:, 0],
    )
    np.testing.assert_allclose(
        random_metadata.specification_warning_matrix[1, :, 0],
        spec_data["warning_upper"][:, 0],
    )
    np.testing.assert_allclose(
        random_metadata.specification_abort_matrix[0, :, 0],
        spec_data["abort_lower"][:, 0],
    )
    np.testing.assert_allclose(
        random_metadata.specification_abort_matrix[1, :, 0],
        spec_data["abort_upper"][:, 0],
    )


def test_random_metadata_load_specification_no_limits(tmp_path, random_metadata):
    spec_path = str(tmp_path / "spec.npz")
    write_cpsd_spec_npz(spec_path, 5.0 * np.arange(101), 1, warning=False, abort=False)

    random_metadata.load_specification(numeric_channel_list(), spec_path)

    assert np.all(np.isnan(random_metadata.specification_warning_matrix))
    assert np.all(np.isnan(random_metadata.specification_abort_matrix))


def test_random_metadata_load_specification_partial_band(tmp_path, random_metadata):
    # A specification narrower than the analysis bandwidth should only fill
    # the matching frequency lines
    spec_path = str(tmp_path / "spec.npz")
    spec_frequencies = 5.0 * np.arange(10, 21)
    write_cpsd_spec_npz(spec_path, spec_frequencies, 1)

    random_metadata.load_specification(numeric_channel_list(), spec_path)

    cpsd = random_metadata.specification_cpsd_matrix
    assert cpsd.shape == (101, 1, 1)
    np.testing.assert_allclose(cpsd[10:21, 0, 0], 1.0)
    np.testing.assert_allclose(cpsd[:10], 0.0)
    np.testing.assert_allclose(cpsd[21:], 0.0)
    assert np.all(np.isnan(random_metadata.specification_warning_matrix[:, :10]))


def test_random_metadata_load_specification_off_grid_lines(tmp_path, random_metadata):
    # Spec frequencies that don't land on analysis frequency lines are skipped
    spec_path = str(tmp_path / "spec.npz")
    write_cpsd_spec_npz(spec_path, 5.0 * np.arange(101) + 0.5, 1)

    random_metadata.load_specification(numeric_channel_list(), spec_path)

    np.testing.assert_allclose(random_metadata.specification_cpsd_matrix, 0.0)


def test_random_metadata_load_specification_coordinate_reduction(
    tmp_path, random_metadata
):
    # A 2-channel specification with coordinates reduced onto a single
    # control channel: node 2 in the spec is the environment's node 1 channel
    spec_path = str(tmp_path / "spec.npz")
    coordinate = outer_coordinate([2, 1], [1, 1])
    write_cpsd_spec_npz(spec_path, 5.0 * np.arange(101), 2, coordinate=coordinate)

    # Control channel is node 1, which is the second spec channel (psd = 2)
    random_metadata.load_specification(numeric_channel_list(), spec_path)

    cpsd = random_metadata.specification_cpsd_matrix
    assert cpsd.shape == (101, 1, 1)
    np.testing.assert_allclose(cpsd[:, 0, 0], 2.0)
    np.testing.assert_allclose(
        random_metadata.specification_warning_matrix[1, :, 0], 4.0
    )


# region netCDF round trip
@pytest.mark.parametrize("with_transformations", [False, True])
def test_random_metadata_netcdf_round_trip(random_metadata, with_transformations):
    if with_transformations:
        random_metadata.response_transformation_matrix = np.arange(2.0).reshape(2, 1)
        random_metadata.reference_transformation_matrix = np.arange(3.0).reshape(3, 1)
    dataset = nc4.Dataset("temp.nc", mode="w", diskless=True, persist=False)
    netcdf_group = dataset.createGroup(ENVIRONMENT_NAME)

    random_metadata.save_metadata_to_netcdf(netcdf_group)
    loaded_metadata = RandomVibrationMetadata.load_metadata_from_netcdf(
        netcdf_group,
        ENVIRONMENT_NAME,
        [True, True],
        numeric_hardware_metadata(),
    )

    assert loaded_metadata.samples_per_frame == 200
    assert loaded_metadata.test_level_ramp_time == 0.1
    assert loaded_metadata.cola_window == "hann"
    assert loaded_metadata.cpsd_window == "Hann"
    assert loaded_metadata.update_tf_during_control is False
    assert loaded_metadata.allow_automatic_aborts is False
    assert loaded_metadata.frames_in_cpsd == 20
    assert loaded_metadata.percent_lines_out == 0.1
    assert loaded_metadata.control_python_function_type == 0
    np.testing.assert_array_equal(
        np.asarray(loaded_metadata.control_channel_indices), [0]
    )
    assert loaded_metadata.output_channel_indices == [1]
    np.testing.assert_allclose(
        np.asarray(loaded_metadata.specification_cpsd_matrix),
        random_metadata.specification_cpsd_matrix,
    )
    np.testing.assert_allclose(
        np.asarray(loaded_metadata.specification_frequency_lines),
        random_metadata.specification_frequency_lines,
    )
    if with_transformations:
        np.testing.assert_allclose(
            np.asarray(loaded_metadata.response_transformation_matrix),
            random_metadata.response_transformation_matrix,
        )
        np.testing.assert_allclose(
            np.asarray(loaded_metadata.reference_transformation_matrix),
            random_metadata.reference_transformation_matrix,
        )
    else:
        assert loaded_metadata.response_transformation_matrix is None
        assert loaded_metadata.reference_transformation_matrix is None
    assert loaded_metadata.sysid_metadata == random_metadata.sysid_metadata

    dataset.close()


# region Worksheet round trip
def test_random_metadata_worksheet_round_trip(tmp_path, random_metadata):
    # This is the profile .xlsx flow: metadata is saved to a worksheet that
    # references an external specification file, then loaded back.  The
    # specification .npz is written without a coordinate key because the
    # worksheet loader builds its reduction coordinate from the output
    # channels rather than the control channels (suspected pre-existing bug).
    spec_path = str(tmp_path / "spec.npz")
    spec_data = write_cpsd_spec_npz(spec_path, 5.0 * np.arange(101), 1)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    random_metadata.save_metadata_to_worksheet(worksheet)
    worksheet.cell(34, 2, spec_path)

    loaded_metadata = RandomVibrationMetadata.load_metadata_from_worksheet(
        worksheet,
        ENVIRONMENT_NAME,
        [True, True],
        numeric_hardware_metadata(),
    )

    assert loaded_metadata.samples_per_frame == 200
    assert loaded_metadata.test_level_ramp_time == 0.1
    assert loaded_metadata.cola_window == "hann"
    assert loaded_metadata.cola_overlap == 0.5
    assert loaded_metadata.cola_window_exponent == 0.5
    assert loaded_metadata.update_tf_during_control is False
    assert loaded_metadata.frames_in_cpsd == 20
    assert loaded_metadata.cpsd_window == "Hann"
    assert loaded_metadata.cpsd_overlap == 0.5
    assert loaded_metadata.percent_lines_out == 0.1
    assert loaded_metadata.allow_automatic_aborts is False
    assert loaded_metadata.sigma_clip == 5.0
    assert loaded_metadata.control_python_script == ""
    assert loaded_metadata.control_python_function_type is None
    assert loaded_metadata.control_channel_indices == [0]
    assert loaded_metadata.output_channel_indices == [1]
    assert loaded_metadata.response_transformation_matrix is None
    assert loaded_metadata.reference_transformation_matrix is None
    assert loaded_metadata.sysid_metadata == random_metadata.sysid_metadata
    np.testing.assert_allclose(
        loaded_metadata.specification_cpsd_matrix, spec_data["cpsd"]
    )
    np.testing.assert_allclose(
        loaded_metadata.specification_warning_matrix[1, :, 0],
        spec_data["warning_upper"][:, 0],
    )


def test_random_metadata_worksheet_round_trip_no_specification(random_metadata):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    random_metadata.save_metadata_to_worksheet(worksheet)

    loaded_metadata = RandomVibrationMetadata.load_metadata_from_worksheet(
        worksheet,
        ENVIRONMENT_NAME,
        [True, True],
        numeric_hardware_metadata(),
    )

    assert loaded_metadata.specification_cpsd_matrix is None
    assert loaded_metadata.specification_frequency_lines is None


# region Instructions
def test_random_instructions_init():
    random_instructions = RandomVibrationInstructions(ENVIRONMENT_NAME, -6.0)

    assert isinstance(random_instructions, RandomVibrationInstructions)
    assert isinstance(random_instructions, EnvironmentInstructions)
    assert random_instructions.control_test_level == -6.0
    random_instructions.validate()


# region Queues
@pytest.mark.parametrize("use_thread", [True, False])
def test_random_queues_init(use_thread):
    queue_container = mock_queue_container(use_thread)
    random_queues = RandomVibrationQueues(
        ENVIRONMENT_NAME,
        queue_container.environment_command_queues[QUEUE_NAME],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.environment_data_in_queues[QUEUE_NAME],
        queue_container.environment_data_out_queues[QUEUE_NAME],
        queue_container.log_file_queue,
        use_thread,
    )

    assert isinstance(random_queues, RandomVibrationQueues)
    for queue_name in (
        "data_analysis_command_queue",
        "signal_generation_command_queue",
        "spectral_command_queue",
        "collector_command_queue",
    ):
        assert hasattr(random_queues, queue_name)


# region Environment
def test_random_environment_init(random_environment):
    assert isinstance(random_environment, RandomVibrationEnvironment)
    assert isinstance(random_environment, SysIdEnvironment)
    assert random_environment.ready
    for command in (
        GlobalCommands.QUIT,
        GlobalCommands.INITIALIZE_HARDWARE,
        GlobalCommands.INITIALIZE_ENVIRONMENT,
        GlobalCommands.INITIALIZE_SYSTEM_ID,
        GlobalCommands.START_SYSTEM_ID_TRANSFER,
        GlobalCommands.START_ENVIRONMENT,
        GlobalCommands.STOP_ENVIRONMENT,
        SysIdDataAnalysisCommands.SYSTEM_ID_COMPLETE,
        RandomVibrationCommands.ADJUST_TEST_LEVEL,
        RandomVibrationCommands.CHANGE_SPECIFICATION,
        RandomVibrationCommands.CHECK_FOR_COMPLETE_SHUTDOWN,
    ):
        assert command in random_environment.command_map


def test_random_environment_initialize_hardware(random_environment):
    hardware_metadata = numeric_hardware_metadata()
    random_environment.initialize_hardware(hardware_metadata)

    assert random_environment.hardware_metadata == hardware_metadata
    assert random_environment.ready


def test_random_environment_initialize_environment(random_environment, random_metadata):
    random_environment.initialize_hardware(numeric_hardware_metadata())
    random_environment.initialize_environment(random_metadata)

    # Identity check: the metadata __eq__ is defeated by NaN entries in the
    # warning/abort matrices, so equality cannot be used here
    assert random_environment.environment_metadata is random_metadata

    queues = random_environment.queue_container
    (analysis_message,) = get_queue_messages(
        queues.data_analysis_command_queue, ENVIRONMENT_NAME, 1
    )
    # The metadata payload is pickled through the queue in non-threaded
    # mode, so assert on its contents rather than its identity
    assert (
        analysis_message[0]
        == RandomVibrationDataAnalysisCommands.INITIALIZE_ENVIRONMENT
    )
    assert analysis_message[1].samples_per_frame == 200
    assert analysis_message[1].control_channel_indices == [0]

    (collector_message,) = get_queue_messages(
        queues.collector_command_queue, ENVIRONMENT_NAME, 1
    )
    assert collector_message[0] == DataCollectorCommands.INITIALIZE_COLLECTOR
    collector_metadata = collector_message[1]
    assert collector_metadata.frame_size == 200
    assert collector_metadata.overlap_fraction == 0.5

    (siggen_message,) = get_queue_messages(
        queues.signal_generation_command_queue, ENVIRONMENT_NAME, 1
    )
    assert siggen_message[0] == SignalGenerationCommands.INITIALIZE_PARAMETERS
    # ramp time * sample rate * output oversample
    assert siggen_message[1].ramp_samples == 0.1 * 1000 * 1

    (spectral_message,) = get_queue_messages(
        queues.spectral_command_queue, ENVIRONMENT_NAME, 1
    )
    assert spectral_message[0] == SpectralProcessingCommands.INITIALIZE_PARAMETERS
    assert spectral_message[1].num_frequency_lines == 101

    assert random_environment.ready


def test_random_environment_initialize_sysid(random_environment, random_metadata):
    initialized_environment(random_environment, random_metadata)
    new_sysid_metadata = sysid_measurement_metadata(frame_size=100)

    random_environment.initialize_sysid(new_sysid_metadata)

    assert random_environment.environment_metadata.sysid_metadata == new_sysid_metadata
    (analysis_message,) = get_queue_messages(
        random_environment.queue_container.data_analysis_command_queue,
        ENVIRONMENT_NAME,
        1,
    )
    assert analysis_message == (
        SysIdDataAnalysisCommands.INITIALIZE_PARAMETERS,
        new_sysid_metadata,
    )
    assert random_environment.ready


def test_random_environment_start_transfer_function(
    random_environment, random_metadata
):
    initialized_environment(random_environment, random_metadata)

    random_environment.start_transfer_function(None)

    queues = random_environment.queue_container
    collector_messages = get_queue_messages(
        queues.collector_command_queue, ENVIRONMENT_NAME, 4
    )
    assert [message for message, _ in collector_messages] == [
        DataCollectorCommands.FORCE_INITIALIZE_COLLECTOR,
        DataCollectorCommands.SET_TEST_LEVEL,
        DataCollectorCommands.ACQUIRE,
        DataCollectorCommands.CLEAR_KURTOSIS_BUFFER,
    ]
    # skip frames from the sysid ramp time, full level for system id
    assert collector_messages[1][1] == (
        random_metadata.sysid_metadata.sysid_skip_frames,
        1,
    )

    siggen_messages = get_queue_messages(
        queues.signal_generation_command_queue, ENVIRONMENT_NAME, 5
    )
    assert [message for message, _ in siggen_messages] == [
        SignalGenerationCommands.INITIALIZE_PARAMETERS,
        SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
        SignalGenerationCommands.MUTE,
        SignalGenerationCommands.ADJUST_TEST_LEVEL,
        SignalGenerationCommands.GENERATE_SIGNALS,
    ]
    assert isinstance(siggen_messages[1][1], RandomSignalGenerator)
    assert siggen_messages[3][1] == 1.0

    (analysis_message,) = get_queue_messages(
        queues.data_analysis_command_queue, ENVIRONMENT_NAME, 1
    )
    assert analysis_message[0] == SysIdDataAnalysisCommands.RUN_TRANSFER_FUNCTION

    spectral_messages = get_queue_messages(
        queues.spectral_command_queue, ENVIRONMENT_NAME, 3
    )
    assert [message for message, _ in spectral_messages] == [
        SpectralProcessingCommands.INITIALIZE_PARAMETERS,
        SpectralProcessingCommands.CLEAR_SPECTRAL_PROCESSING,
        SpectralProcessingCommands.RUN_SPECTRAL_PROCESSING,
    ]

    assert random_environment.sysid_active
    gui_message = queues.gui_update_queue.get(timeout=10)
    assert gui_message == (ENVIRONMENT_NAME, (SysIdUICommands.SYSID_STARTED, None))


def test_random_environment_adjust_test_level(random_environment, random_metadata):
    initialized_environment(random_environment, random_metadata)

    random_environment.adjust_test_level(-6.0)

    queues = random_environment.queue_container
    (siggen_message,) = get_queue_messages(
        queues.signal_generation_command_queue, ENVIRONMENT_NAME, 1
    )
    assert siggen_message == (
        SignalGenerationCommands.ADJUST_TEST_LEVEL,
        db2scale(-6.0),
    )
    (collector_message,) = get_queue_messages(
        queues.collector_command_queue, ENVIRONMENT_NAME, 1
    )
    assert collector_message == (
        DataCollectorCommands.SET_TEST_LEVEL,
        (random_metadata.skip_frames, db2scale(-6.0)),
    )
    gui_message = queues.gui_update_queue.get(timeout=10)
    assert gui_message == (
        ENVIRONMENT_NAME,
        (RandomVibrationUICommands.ADJUST_TEST_LEVEL, -6.0),
    )


def test_random_environment_system_id_complete(random_environment, random_metadata):
    initialized_environment(random_environment, random_metadata)
    data_package = sysid_measurement_data_package(random_metadata.sysid_metadata)

    random_environment.system_id_complete(
        (random_metadata.sysid_metadata, data_package)
    )

    assert random_environment.sysid_data == data_package
    assert random_environment.sysid_stored

    queues = random_environment.queue_container
    gui_message = queues.gui_update_queue.get(timeout=10)
    assert gui_message[0] == ENVIRONMENT_NAME
    assert gui_message[1][0] == SysIdDataAnalysisUICommands.TRANSFER_COMPLETED
    gui_message = queues.gui_update_queue.get(timeout=10)
    assert gui_message[0] == UICommands.COMPLETED_SYSTEM_ID

    (analysis_message,) = get_queue_messages(
        queues.data_analysis_command_queue, ENVIRONMENT_NAME, 1
    )
    assert analysis_message == (
        RandomVibrationDataAnalysisCommands.PERFORM_CONTROL_PREDICTION,
        None,
    )


def test_random_environment_start_control(random_environment, random_metadata):
    initialized_environment(random_environment, random_metadata)
    random_instructions = RandomVibrationInstructions(ENVIRONMENT_NAME, -3.0)

    random_environment.start_control(random_instructions)

    assert random_environment.active
    assert not random_environment.siggen_shutdown_achieved
    assert not random_environment.collector_shutdown_achieved
    assert not random_environment.spectral_shutdown_achieved
    assert not random_environment.analysis_shutdown_achieved

    queues = random_environment.queue_container
    gui_message = queues.gui_update_queue.get(timeout=10)
    assert gui_message[0] == ENVIRONMENT_NAME
    assert gui_message[1][0] == UICommands.SET_ENVIRONMENT_INSTRUCTIONS
    assert gui_message[1][1].control_test_level == -3.0

    collector_messages = get_queue_messages(
        queues.collector_command_queue, ENVIRONMENT_NAME, 3
    )
    assert [message for message, _ in collector_messages] == [
        DataCollectorCommands.INITIALIZE_COLLECTOR,
        DataCollectorCommands.SET_TEST_LEVEL,
        DataCollectorCommands.ACQUIRE,
    ]
    assert collector_messages[1][1] == (
        random_metadata.skip_frames,
        db2scale(-3.0),
    )

    siggen_messages = get_queue_messages(
        queues.signal_generation_command_queue, ENVIRONMENT_NAME, 5
    )
    assert [message for message, _ in siggen_messages] == [
        SignalGenerationCommands.INITIALIZE_PARAMETERS,
        SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
        SignalGenerationCommands.MUTE,
        SignalGenerationCommands.ADJUST_TEST_LEVEL,
        SignalGenerationCommands.GENERATE_SIGNALS,
    ]
    assert isinstance(siggen_messages[1][1], CPSDSignalGenerator)
    assert siggen_messages[3][1] == db2scale(-3.0)

    (analysis_message,) = get_queue_messages(
        queues.data_analysis_command_queue, ENVIRONMENT_NAME, 1
    )
    assert analysis_message == (
        RandomVibrationDataAnalysisCommands.RUN_CONTROL,
        None,
    )

    spectral_messages = get_queue_messages(
        queues.spectral_command_queue, ENVIRONMENT_NAME, 3
    )
    assert [message for message, _ in spectral_messages] == [
        SpectralProcessingCommands.INITIALIZE_PARAMETERS,
        SpectralProcessingCommands.CLEAR_SPECTRAL_PROCESSING,
        SpectralProcessingCommands.RUN_SPECTRAL_PROCESSING,
    ]


def test_random_environment_stop_environment(random_environment, random_metadata):
    initialized_environment(random_environment, random_metadata)

    random_environment.stop_environment(None)

    queues = random_environment.queue_container
    (collector_message,) = get_queue_messages(
        queues.collector_command_queue, ENVIRONMENT_NAME, 1
    )
    assert collector_message == (
        DataCollectorCommands.SET_TEST_LEVEL,
        (random_metadata.skip_frames * 10, 1),
    )
    (siggen_message,) = get_queue_messages(
        queues.signal_generation_command_queue, ENVIRONMENT_NAME, 1
    )
    assert siggen_message == (SignalGenerationCommands.START_SHUTDOWN, None)
    (spectral_message,) = get_queue_messages(
        queues.spectral_command_queue, ENVIRONMENT_NAME, 1
    )
    assert spectral_message == (
        SpectralProcessingCommands.STOP_SPECTRAL_PROCESSING,
        None,
    )
    (analysis_message,) = get_queue_messages(
        queues.data_analysis_command_queue, ENVIRONMENT_NAME, 1
    )
    assert analysis_message == (
        RandomVibrationDataAnalysisCommands.STOP_CONTROL,
        None,
    )
    # The environment re-enqueues the shutdown check for itself
    (environment_message,) = get_queue_messages(
        random_environment.environment_command_queue, ENVIRONMENT_NAME, 1
    )
    assert environment_message == (
        RandomVibrationCommands.CHECK_FOR_COMPLETE_SHUTDOWN,
        None,
    )


def test_random_environment_check_for_control_shutdown(
    random_environment, random_metadata
):
    initialized_environment(random_environment, random_metadata)
    random_environment.set_active()
    random_environment.siggen_shutdown_achieved = True
    random_environment.collector_shutdown_achieved = True
    random_environment.spectral_shutdown_achieved = True
    random_environment.analysis_shutdown_achieved = True

    random_environment.check_for_control_shutdown(None)

    assert not random_environment.active
    gui_message = random_environment.gui_update_queue.get(timeout=10)
    assert gui_message == (
        ENVIRONMENT_NAME,
        (RandomVibrationUICommands.ENABLE_CONTROL, None),
    )


def test_random_environment_save_spectral_data(
    tmp_path, random_environment, random_metadata
):
    initialized_environment(random_environment, random_metadata)
    output_path = str(tmp_path / "control_data.nc4")

    random_environment.save_spectral_data(output_path)

    with nc4.Dataset(output_path, "r") as dataset:
        assert ENVIRONMENT_NAME in dataset.groups
        assert dataset.groups[ENVIRONMENT_NAME].samples_per_frame == 200
    (analysis_message,) = get_queue_messages(
        random_environment.queue_container.data_analysis_command_queue,
        ENVIRONMENT_NAME,
        1,
    )
    assert analysis_message == (
        RandomVibrationDataAnalysisCommands.SAVE_CONTROL_DATA,
        output_path,
    )


def test_random_environment_change_specification(
    tmp_path, random_environment, random_metadata
):
    initialized_environment(random_environment, random_metadata)
    spec_path = str(tmp_path / "new_spec.npz")
    new_cpsd_value = 7.0
    spec_frequencies = 5.0 * np.arange(101)
    num_lines = spec_frequencies.size
    cpsd = np.full((num_lines, 1, 1), new_cpsd_value, dtype="complex128")
    np.savez(spec_path, f=spec_frequencies, cpsd=cpsd)

    random_environment.change_specification(spec_path)

    np.testing.assert_allclose(
        random_environment.environment_metadata.specification_cpsd_matrix[:, 0, 0],
        new_cpsd_value,
    )
    # initialize_environment re-ran, so the sub-workers were re-initialized
    (analysis_message,) = get_queue_messages(
        random_environment.queue_container.data_analysis_command_queue,
        ENVIRONMENT_NAME,
        1,
    )
    assert (
        analysis_message[0]
        == RandomVibrationDataAnalysisCommands.INITIALIZE_ENVIRONMENT
    )
    gui_messages = []
    while True:
        gui_messages.append(random_environment.gui_update_queue.get(timeout=0.5))
        if gui_messages[-1][1][0] == RandomVibrationUICommands.CHANGE_SPECIFICATION:
            break
    assert gui_messages[-1][1][1][0] == spec_path


# region Run loop
@mock.patch(
    "rattlesnake.environment.abstract_environment.Environment.log",
    new=mock.MagicMock(),
)
def test_random_environment_run_full_lifecycle(random_environment, random_metadata):
    """Replays the controller's real message sequence through run().

    This is the message order RattlesnakeController sends over a test:
    initialization, system identification, control at level, stop, shutdown
    handshake, and quit.  VerboseMessageQueue.get is mocked with a finite
    side_effect so exhaustion raises instead of hanging if the loop
    misbehaves.
    """
    hardware_metadata = numeric_hardware_metadata()
    sysid_metadata = random_metadata.sysid_metadata
    data_package = sysid_measurement_data_package(sysid_metadata)
    random_instructions = RandomVibrationInstructions(ENVIRONMENT_NAME, 0.0)

    controller_sequence = [
        (GlobalCommands.INITIALIZE_HARDWARE, hardware_metadata),
        (GlobalCommands.INITIALIZE_ENVIRONMENT, random_metadata),
        (GlobalCommands.INITIALIZE_SYSTEM_ID, sysid_metadata),
        (GlobalCommands.START_SYSTEM_ID_TRANSFER, None),
        (
            SysIdDataAnalysisCommands.SYSTEM_ID_COMPLETE,
            (sysid_metadata, data_package),
        ),
        (GlobalCommands.STOP_SYSTEM_ID, False),
        (SignalGenerationCommands.SHUTDOWN_ACHIEVED, None),
        (DataCollectorCommands.SHUTDOWN_ACHIEVED, None),
        (SpectralProcessingCommands.SHUTDOWN_ACHIEVED, None),
        (SysIdDataAnalysisCommands.SHUTDOWN_ACHIEVED, None),
        (SystemIdCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None),
        (GlobalCommands.START_ENVIRONMENT, random_instructions),
        (GlobalCommands.STOP_ENVIRONMENT, None),
        (SignalGenerationCommands.SHUTDOWN_ACHIEVED, None),
        (DataCollectorCommands.SHUTDOWN_ACHIEVED, None),
        (SpectralProcessingCommands.SHUTDOWN_ACHIEVED, None),
        (SysIdDataAnalysisCommands.SHUTDOWN_ACHIEVED, None),
        (RandomVibrationCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None),
        (GlobalCommands.QUIT, None),
    ]

    mock_shutdown = mock.MagicMock()
    mock_shutdown.is_set.return_value = False
    with mock.patch(
        "rattlesnake.utilities.VerboseMessageQueue.get",
        side_effect=controller_sequence,
    ):
        random_environment.run(mock_shutdown)

    # The full lifecycle completed: system id data stored, environment
    # stopped, and QUIT fanned out to all four sub-workers
    assert random_environment.sysid_data == data_package
    assert random_environment.sysid_stored
    assert not random_environment.sysid_active
    assert not random_environment.active

    queues = random_environment.queue_container
    for queue in (
        queues.data_analysis_command_queue,
        queues.collector_command_queue,
        queues.signal_generation_command_queue,
        queues.spectral_command_queue,
    ):
        commands = drain_queue_commands(queue, ENVIRONMENT_NAME)
        assert commands[-1] == GlobalCommands.QUIT


# region random_vibration_process
@pytest.mark.parametrize("use_thread", [True, False])
@mock.patch("rattlesnake.environment.random_vibration_sys_id_environment.mp.Process")
@mock.patch(
    "rattlesnake.environment.random_vibration_sys_id_environment" ".threading.Thread"
)
@mock.patch(
    "rattlesnake.environment.random_vibration_sys_id_environment"
    ".RandomVibrationEnvironment"
)
def test_random_vibration_process(
    mock_environment_class, mock_thread, mock_process, use_thread
):
    queue_container = mock_queue_container(use_thread)
    event_container = mock_event_container(use_thread)

    random_vibration_process(
        ENVIRONMENT_NAME,
        QUEUE_NAME,
        queue_container.environment_command_queues[QUEUE_NAME],
        queue_container.gui_update_queue,
        queue_container.controller_command_queue,
        queue_container.log_file_queue,
        queue_container.environment_data_in_queues[QUEUE_NAME],
        queue_container.environment_data_out_queues[QUEUE_NAME],
        event_container.acquisition_active_event,
        event_container.output_active_event,
        event_container.environment_active_events[QUEUE_NAME],
        event_container.environment_ready_events[QUEUE_NAME],
        event_container.environment_close_events[QUEUE_NAME],
        event_container.environment_sysid_active_events[QUEUE_NAME],
        event_container.environment_sysid_stored_events[QUEUE_NAME],
        event_container.ping_alive_event,
        use_thread,
    )

    mock_environment_class.return_value.run.assert_called()
    worker_class = mock_thread if use_thread else mock_process
    assert worker_class.call_count == 4
    worker_targets = {
        call.kwargs["target"].__name__ for call in worker_class.call_args_list
    }
    assert worker_targets == {
        "spectral_processing_process",
        "random_data_analysis_process",
        "signal_generation_process",
        "data_collector_process",
    }
