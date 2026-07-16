import queue as thqueue
import threading
import inspect
import time

import netCDF4 as nc4
import numpy as np
import openpyxl
import pytest
import multiprocessing as mp

from rattlesnake.utilities import RattlesnakeError, GlobalCommands, VerboseMessageQueue
from rattlesnake.hardware.skeleton_hardware import SkeletonHardwareMetadata
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.environment.skeleton_environment import (
    SkeletonMetadata,
    SkeletonInstructions,
    SkeletonQueues,
    SkeletonEnvironment,
)
from rattlesnake.environment.environment_registry import (
    ENVIRONMENT_COMMANDS,
    ENVIRONMENT_METADATA,
    ENVIRONMENT_INSTRUCTION,
    ENVIRONMENT_CLASS,
    ENVIRONMENT_PROCESS,
)
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.examples.example_registry import (
    ENVIRONMENT_DICT,
    EXAMPLE_NETCDF,
    EXAMPLE_WORKSHEET,
)
from rattlesnake.examples.hardware.sdynpy_system.sdynpy_system_metadata import (
    manual_sdynpy_system_metadata,
)
from rattlesnake.testing.mock_utilities import (
    IMPLEMENTED_ENVIRONMENT,
    instantiate_with_mocks,
    mock_channel_list_bools,
    skeleton_hardware_metadata,
    skeleton_environment_metadata,
    skeleton_environment_instructions,
    skeleton_queues,
    skeleton_environment,
)
from rattlesnake.environment.skeleton_environment import SkeletonCommands

# Environment types that ship an example netCDF/worksheet file (used by the
# headless example) and are therefore usable for load/save round-trip checks.
EXAMPLE_ENVIRONMENT_TYPES = list(EXAMPLE_NETCDF)

# Attributes/fields that are intentionally overwritten with an absolute path
# after loading (e.g. control law scripts), so their value in the example file
# will never match what gets saved back out.
_NETCDF_SKIP_ATTRIBUTES = {"control_python_script"}


def _diff_netcdf_groups(original_group, saved_group, prefix=""):
    """Recursively compares attributes, variables, and subgroups between two
    netCDF4 groups, returning a list of human-readable differences.

    Attributes that only appear in ``saved_group`` are ignored, since example
    files may predate fields that were added later (the loaders fall back to
    a default in that case). Attributes only in ``original_group`` indicate
    real data loss and are reported.
    """
    differences = []

    original_attrs = set(original_group.ncattrs()) - _NETCDF_SKIP_ATTRIBUTES
    saved_attrs = set(saved_group.ncattrs()) - _NETCDF_SKIP_ATTRIBUTES

    for name in sorted(original_attrs - saved_attrs):
        differences.append(f"{prefix}{name}: present in example, missing after save")

    for name in sorted(original_attrs & saved_attrs):
        original_value = getattr(original_group, name)
        saved_value = getattr(saved_group, name)
        if isinstance(original_value, np.ndarray) or isinstance(
            saved_value, np.ndarray
        ):
            equal = np.array_equal(original_value, saved_value, equal_nan=True)
        else:
            equal = original_value == saved_value
        if not equal:
            differences.append(f"{prefix}{name}: {original_value!r} != {saved_value!r}")

    original_vars = set(original_group.variables)
    saved_vars = set(saved_group.variables)
    for name in sorted(original_vars ^ saved_vars):
        differences.append(f"{prefix}variables/{name}: only present in one file")

    for name in sorted(original_vars & saved_vars):
        original_data = np.ma.filled(original_group.variables[name][...], np.nan)
        saved_data = np.ma.filled(saved_group.variables[name][...], np.nan)
        if not np.array_equal(original_data, saved_data, equal_nan=True):
            differences.append(f"{prefix}variables/{name}: values differ")

    original_groups = set(original_group.groups)
    saved_groups = set(saved_group.groups)
    for name in sorted(original_groups & saved_groups):
        differences.extend(
            _diff_netcdf_groups(
                original_group.groups[name],
                saved_group.groups[name],
                prefix=f"{prefix}{name}/",
            )
        )
    for name in sorted(original_groups ^ saved_groups):
        differences.append(f"{prefix}groups/{name}: only present in one file")

    return differences


def _normalize_cell_value(value):
    """Treats a blank cell and an empty string as equivalent."""
    return "" if value is None else value


def _is_comment_cell(value):
    """Identifies human-readable documentation cells (e.g. '# ...') that are
    regenerated fresh by ``save_metadata_to_worksheet`` and aren't expected to
    match older example files that predate a given field."""
    return isinstance(value, str) and value.strip().startswith("#")


def _diff_worksheets(original_worksheet, saved_worksheet):
    """Compares two worksheets cell by cell, returning a list of
    ``(row, col, original_value, saved_value)`` tuples for values that differ.

    Documentation/comment cells are ignored, as are rows whose label mentions
    a "script", since those are always overwritten with an absolute path after
    loading rather than sourced from the worksheet's saved value.
    """
    differences = []
    max_row = max(original_worksheet.max_row, saved_worksheet.max_row)
    max_col = max(original_worksheet.max_column, saved_worksheet.max_column)

    for row in range(1, max_row + 1):
        row_label = str(
            original_worksheet.cell(row, 1).value
            or saved_worksheet.cell(row, 1).value
            or ""
        ).lower()
        if "script" in row_label:
            continue
        for col in range(1, max_col + 1):
            original_value = _normalize_cell_value(
                original_worksheet.cell(row, col).value
            )
            saved_value = _normalize_cell_value(saved_worksheet.cell(row, col).value)
            if _is_comment_cell(original_value) or _is_comment_cell(saved_value):
                continue
            if original_value != saved_value:
                differences.append((row, col, original_value, saved_value))

    return differences


# region Fixtures
@pytest.fixture
def hardware_metadata():
    return skeleton_hardware_metadata()


@pytest.fixture()
def environment_metadata():
    return skeleton_environment_metadata()


@pytest.fixture()
def environment_instructions():
    return skeleton_environment_instructions()


@pytest.fixture()
def queue_container():
    return skeleton_queues()


@pytest.fixture
def environment():
    return skeleton_environment()


# endregion


# region Environment Commands
@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment_commands_have_unique_integer_values(environment_type):
    """
    Iterates through each enum member to confirm unique integer values.
    Verifies that ``VALID_PROFILE_COMMANDS`` is a tuple of ints and
    ``VALID_DATA`` is a dict mapping ints to types.
    """
    command_class = ENVIRONMENT_COMMANDS[environment_type]

    command_members = [
        member
        for name, member in command_class.__members__.items()
        if name not in {"VALID_PROFILE_COMMANDS", "VALID_DATA"}
    ]

    values = [member.value for member in command_members]

    assert all(isinstance(value, int) for value in values)
    assert len(values) == len(set(values))

    valid_profile_commands = command_class.VALID_PROFILE_COMMANDS.value
    valid_data = command_class.VALID_DATA.value

    assert isinstance(valid_profile_commands, tuple)
    assert all(isinstance(command, int) for command in valid_profile_commands)

    assert isinstance(valid_data, dict)
    assert all(isinstance(key, int) for key in valid_data.keys())
    assert all(isinstance(value, type) for value in valid_data.values())


def test_environment_commands_label():
    """
    Checks that labels replace underscores with spaces and use title
    case.
    """
    assert SkeletonCommands.EXAMPLE_SET_TEST_LEVEL.label == "Example Set Test Level"


def test_environment_commands_valid_profile_commands():
    """
    Ensures this method returns a tuple of EnvironmentCommands
    members corresponding to VALID_PROFILE_COMMANDS.
    """
    assert SkeletonCommands.valid_profile_commands() == (
        SkeletonCommands.EXAMPLE_SET_TEST_LEVEL,
        SkeletonCommands.EXAMPLE_FLOAT_COMMAND,
    )


def test_environment_commands_valid_data():
    """
    Verifies this method returns a dict mapping enum members to their
    predefined types.
    """
    assert SkeletonCommands.valid_data() == {
        SkeletonCommands.EXAMPLE_SET_TEST_LEVEL: type(None),
        SkeletonCommands.EXAMPLE_FLOAT_COMMAND: float,
    }


# endregion


# region Environment Metadata
@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment_metadata(environment_type):
    """
    Verifies that subclasses within ``ENVIRONMENT_METADATA`` in the
    registry initialize required metadata attributes and preserve the
    supplied environment name, channel mask, and sample rate.
    """
    metadata_class = ENVIRONMENT_METADATA[environment_type]
    channel_list_bools = mock_channel_list_bools()
    metadata = instantiate_with_mocks(
        metadata_class,
        environment_name="test_environment",
        channel_list_bools=channel_list_bools,
        sample_rate=2048,
    )

    assert metadata.environment_name == "test_environment"
    assert metadata.channel_list_bools == channel_list_bools
    assert metadata.sample_rate == 2048


def test_environment_metadata_init():
    """
    Confirms that initialization stores the environment type,
    environment name, sample rate, channel list bools, and
    initializes ``queue_name`` to ``None``.
    """
    metadata = skeleton_environment_metadata(
        environment_name="Env A",
        channel_list_bools=[True, False, True],
        sample_rate=2048,
    )

    assert metadata.environment_type == EnvironmentType.SKELETON
    assert metadata.environment_name == "Env A"
    assert metadata.channel_list_bools == [True, False, True]
    assert metadata.sample_rate == 2048
    assert metadata.queue_name is None


@pytest.mark.parametrize(
    "channel_list_bools, expected",
    [
        ([True, True], [0, 1]),
        ([True, False], [0]),
    ],
)
def test_environment_metadata_channel_indices(channel_list_bools, expected):
    """
    Verifies that selected channel indices correspond to true
    entries in ``channel_list_bools``.
    """
    metadata = skeleton_environment_metadata(channel_list_bools=channel_list_bools)

    assert metadata.channel_indices == expected


def test_environment_metadata_environment_channel_list():
    """
    Confirms that the returned channel list contains only channels
    selected by ``channel_list_bools`` and preserves their original
    order.
    """
    metadata = skeleton_environment_metadata(channel_list_bools=[True, False, True])
    channel_list = ["ch0", "ch1", "ch2"]

    assert metadata.environment_channel_list(channel_list) == ["ch0", "ch2"]


def test_environment_metadata_validate_truth(
    hardware_metadata: SkeletonHardwareMetadata,
    environment_metadata: SkeletonMetadata,
):
    """
    Verifies that valid skeleton metadata class passes the validation check.
    """

    environment_metadata.validate(hardware_metadata)


def test_environment_metadata_validate_invalid_environment_type(
    hardware_metadata: SkeletonHardwareMetadata,
    environment_metadata: SkeletonMetadata,
):
    """
    Verifies that an error is thrown with an invalid environment type object.
    """
    environment_metadata.environment_type = object()

    with pytest.raises(RattlesnakeError):
        environment_metadata.validate(hardware_metadata)


def test_environment_metadata_validate_invalid_environment_name(
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Verifies that an error is thrown when the environment name is not a string.
    """
    metadata = skeleton_environment_metadata(environment_name=123)

    with pytest.raises(RattlesnakeError):
        metadata.validate(hardware_metadata)


def test_environment_metadata_validate_invalid_channel_list(
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Verifies that an error is thrown when an invalid channel list is given to the metadata.
    """
    metadata = skeleton_environment_metadata(channel_list_bools=[True, False, True])

    with pytest.raises(RattlesnakeError):
        metadata.validate(hardware_metadata)


@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment_metadata_save_load_netcdf(
    environment_type, tmp_path, hardware_metadata: SkeletonHardwareMetadata
):
    """
    Saves a valid metadata subclass to a netcdf file and then loads
    it back into a metadata object. Verifies that the metadata object
    is valid and that the netcdf handle contains environment_name and
    environment_type.
    """
    metadata_class = ENVIRONMENT_METADATA[environment_type]
    metadata = ENVIRONMENT_DICT[environment_type]["manual"](hardware_metadata)
    metadata.environment_name = "Environment Name"

    path = tmp_path / "metadata.nc"

    with nc4.Dataset(path, "w") as dataset:
        group = dataset.createGroup(metadata.environment_name)
        metadata.save_metadata_to_netcdf(group)

    with nc4.Dataset(path, "r") as dataset:
        load_group = dataset.groups["Environment Name"]
        loaded = metadata_class.load_metadata_from_netcdf(
            load_group,
            environment_name="Environment Name",
            channel_list_bools=mock_channel_list_bools(),
            hardware_metadata=hardware_metadata,
        )

    assert loaded.environment_name == "Environment Name"
    assert loaded.channel_list_bools == [True, True]
    assert loaded.sample_rate == 1024
    loaded.validate(hardware_metadata)


@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment_metadata_save_load_worksheet(
    environment_type, hardware_metadata: SkeletonHardwareMetadata
):
    """
    Saves a valid metadata subclass to an Excel worksheet and then loads
    it back into a metadata object. Verifies that the worksheet contains
    the expected metadata header values and that the loaded metadata
    preserves the expected environment name, channel selection, and sample
    rate before passing validation.
    """
    metadata_class = ENVIRONMENT_METADATA[environment_type]
    metadata = ENVIRONMENT_DICT[environment_type]["manual"](hardware_metadata)
    metadata.environment_name = "Environment Name"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    metadata.save_metadata_to_worksheet(worksheet)

    assert worksheet.cell(1, 1).value == "Control Type"
    assert worksheet.cell(1, 3).value == "v4.0"

    loaded = metadata_class.load_metadata_from_worksheet(
        worksheet=worksheet,
        environment_name="Environment Name",
        channel_list_bools=mock_channel_list_bools(),
        hardware_metadata=hardware_metadata,
    )

    assert loaded.environment_name == "Environment Name"
    assert loaded.channel_list_bools == [True, True]
    assert loaded.sample_rate == 1024
    loaded.validate(hardware_metadata)


@pytest.mark.parametrize("environment_type", EXAMPLE_ENVIRONMENT_TYPES)
def test_environment_metadata_load_save_netcdf(environment_type, tmp_path):
    """
    Loads environment metadata from the example netCDF file used by the
    headless example, saves it back out to a new netCDF file, and confirms
    the saved file reproduces the example aside from fields that are always
    overwritten at load time (e.g. absolute paths to control law scripts)
    or fields the example predates (which fall back to defaults).
    """
    hardware_metadata = manual_sdynpy_system_metadata()
    metadata = ENVIRONMENT_DICT[environment_type]["netcdf"](hardware_metadata)

    path = tmp_path / "metadata.nc4"
    with nc4.Dataset(path, "w", format="NETCDF4") as dataset:
        group = dataset.createGroup(metadata.environment_name)
        metadata.save_metadata_to_netcdf(group)
        dataset.close()

    with (
        nc4.Dataset(EXAMPLE_NETCDF[environment_type], "r") as original_dataset,
        nc4.Dataset(path, "r") as saved_dataset,
    ):
        original_group = original_dataset.groups[metadata.environment_name]
        saved_group = saved_dataset.groups[metadata.environment_name]
        differences = _diff_netcdf_groups(original_group, saved_group)

        original_dataset.close()
        saved_dataset.close()

    assert differences == []


@pytest.mark.parametrize("environment_type", EXAMPLE_ENVIRONMENT_TYPES)
def test_environment_metadata_load_save_worksheet(environment_type):
    """
    Loads environment metadata from the example worksheet used by the
    headless example, saves it back out to a new worksheet, and confirms
    the saved worksheet reproduces the example aside from fields that are
    always overwritten at load time (e.g. absolute paths to control law
    scripts) or documentation comments regenerated from the current template.
    """
    hardware_metadata = manual_sdynpy_system_metadata()
    metadata = ENVIRONMENT_DICT[environment_type]["worksheet"](hardware_metadata)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = metadata.environment_name
    metadata.save_metadata_to_worksheet(worksheet)

    original_workbook = openpyxl.load_workbook(
        EXAMPLE_WORKSHEET[environment_type], read_only=True
    )
    original_worksheet = original_workbook[metadata.environment_name]

    differences = _diff_worksheets(original_worksheet, worksheet)
    original_workbook.close()

    assert differences == []


# endregion


# region Environment Instructions
@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment_instructions(environment_type):
    """
    Verifies that instruction subclasses initialize the required
    environment type and environment name attributes.
    """
    instruction_class = ENVIRONMENT_INSTRUCTION[environment_type]

    instructions = instantiate_with_mocks(
        instruction_class, environment_name="test_environment"
    )

    assert instructions.environment_name == "test_environment"


def test_environment_instructions_init():
    """
    Confirms that initialization stores the environment type and
    environment name.
    """
    instructions = skeleton_environment_instructions(environment_name="Env A")

    assert instructions.environment_type == EnvironmentType.SKELETON
    assert instructions.environment_name == "Env A"


def test_environment_instructions_validate_truth(
    environment_instructions: SkeletonInstructions,
):
    """
    Verifies that a valid skeleton instruction subclass passes the validation
    check.
    """
    environment_instructions.validate()


# endregion


# region Environment
@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment(environment_type):
    """
    Verifies that instruction subclasses initialize the required
    environment name attribute.
    """
    environment_class = ENVIRONMENT_CLASS[environment_type]

    environment = instantiate_with_mocks(
        environment_class, environment_name="test_environment"
    )

    assert environment.environment_name == "test_environment"


def test_environment_init():
    """
    Confirms that initialization stores all queues and events, initializes
    metadata attributes to ``None``, and maps the default global commands.
    """
    environment = skeleton_environment(
        environment_name="Environment Name", queue_name="Queue Name"
    )
    assert environment.environment_name == "Environment Name"
    assert environment.queue_name == "Queue Name"
    assert environment.hardware_metadata is None
    assert environment.environment_metadata is None

    assert GlobalCommands.QUIT in environment.command_map
    assert GlobalCommands.INITIALIZE_HARDWARE in environment.command_map
    assert GlobalCommands.INITIALIZE_ENVIRONMENT in environment.command_map
    assert GlobalCommands.STOP_ENVIRONMENT in environment.command_map


def test_environment_command_map(environment: SkeletonEnvironment):
    """
    Verifies that the default command map contains expected global
    commands and maps them to callable methods.
    """
    assert environment.command_map[GlobalCommands.QUIT] == environment.quit
    assert environment.command_map[GlobalCommands.INITIALIZE_HARDWARE] == (
        environment.initialize_hardware
    )
    assert environment.command_map[GlobalCommands.INITIALIZE_ENVIRONMENT] == (
        environment.initialize_environment
    )
    assert environment.command_map[GlobalCommands.STOP_ENVIRONMENT] == (
        environment.stop_environment
    )


def test_environment_map_command(environment: SkeletonEnvironment):
    """
    Confirms that a new command can be added to the command map and maps
    to the provided callable.
    """
    command = object()

    def handler(data):
        return data

    environment.map_command(command, handler)

    assert environment.command_map[command] == handler


def test_environment_set_ready(environment: SkeletonEnvironment):
    """
    Verifies that calling this method sets the ready event.
    """
    environment.set_ready()

    assert environment.ready is True


def test_environment_clear_ready(environment: SkeletonEnvironment):
    """
    Verifies that calling this method clears the ready event.
    """
    environment._ready_event.set()
    environment.clear_ready()

    assert environment.ready is False


def test_environment_set_active(environment: SkeletonEnvironment):
    """
    Verifies that calling this method sets the active event.
    """
    environment.set_active()

    assert environment.active is True


def test_environment_clear_active(environment: SkeletonEnvironment):
    """
    Verifies that calling this method clears the active event.
    """
    environment._active_event.set()
    environment.clear_active()

    assert environment.active is False


def test_environment_acquisition_active(environment: SkeletonEnvironment):
    """
    Verifies that this property reflects the state of the acquisition
    active event.
    """
    assert environment.acquisition_active is False
    environment._acquisition_active_event.set()
    assert environment.acquisition_active is True


def test_environment_output_active(environment: SkeletonEnvironment):
    """
    Verifies that this property reflects the state of the output active
    event.
    """
    assert environment.output_active is False
    environment._output_active_event.set()
    assert environment.output_active is True


@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment_initialize_hardware(
    environment_type,
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Verifies that each implemented environment stores the supplied hardware
    metadata during hardware initialization and sets its ready event when
    initialization completes.
    """
    environment_class = ENVIRONMENT_CLASS[environment_type]
    environment = instantiate_with_mocks(
        environment_class,
        environment_name="test_environment",
        ready_event=mp.Event(),
    )
    environment.initialize_hardware(hardware_metadata)

    assert environment.hardware_metadata is hardware_metadata
    assert environment.ready is True


@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment_initialize_environment(
    environment_type,
    hardware_metadata: SkeletonHardwareMetadata,
):
    """
    Verifies that each implemented environment stores the supplied environment
    metadata during environment initialization, updates its environment name
    from the metadata, and sets its ready event when initialization completes.
    """
    environment_metadata_class = ENVIRONMENT_METADATA[environment_type]
    environment_metadata = instantiate_with_mocks(
        environment_metadata_class,
        environment_name="Environment A",
    )
    environment_class = ENVIRONMENT_CLASS[environment_type]
    environment = instantiate_with_mocks(
        environment_class,
        environment_name="test_environment",
        ready_event=mp.Event(),
    )
    environment.hardware_metadata = hardware_metadata
    environment.initialize_environment(environment_metadata)

    assert environment.environment_metadata is environment_metadata
    assert environment.environment_name == environment_metadata.environment_name
    assert environment.ready is True


def test_environment_queue_name(environment: SkeletonEnvironment):
    """
    Verifies that this property returns the queue name supplied during
    initialization.
    """
    assert environment.queue_name == "Queue Name"


def test_environment_environment_command_queue(queue_container: SkeletonQueues):
    """
    Verifies that this property returns the command queue supplied during
    initialization.
    """
    environment = skeleton_environment(queue_container=queue_container)
    assert (
        environment.environment_command_queue
        is queue_container.environment_command_queue
    )


def test_environment_data_in_queue(queue_container: SkeletonQueues):
    """
    Verifies that this property returns the data input queue supplied
    during initialization.
    """
    environment = skeleton_environment(queue_container=queue_container)
    assert environment.data_in_queue is queue_container.data_in_queue


def test_environment_data_out_queue(queue_container: SkeletonQueues):
    """
    Verifies that this property returns the data output queue supplied
    during initialization.
    """
    environment = skeleton_environment(queue_container=queue_container)
    assert environment.data_out_queue is queue_container.data_out_queue


def test_environment_gui_update_queue(queue_container: SkeletonQueues):
    """
    Verifies that this property returns the GUI update queue supplied
    during initialization.
    """
    environment = skeleton_environment(queue_container=queue_container)
    assert environment.gui_update_queue is queue_container.gui_update_queue


def test_environment_controller_command_queue(queue_container: SkeletonQueues):
    """
    Verifies that this property returns the controller command queue
    supplied during initialization.
    """
    environment = skeleton_environment(queue_container=queue_container)
    assert (
        environment.controller_command_queue
        is queue_container.controller_communication_queue
    )


def test_environment_log_file_queue(queue_container: SkeletonQueues):
    """
    Verifies that this property returns the log file queue supplied during
    initialization.
    """
    environment = skeleton_environment(queue_container=queue_container)
    assert environment.log_file_queue is queue_container.log_file_queue


def test_environment_log(environment: SkeletonEnvironment):
    """
    Verifies that calling this method places a formatted log message on
    the log file queue.
    """
    environment.log("hello world")

    time.sleep(1)
    log_message = environment.log_file_queue.get_nowait()

    assert "Environment Name -- hello world" in log_message


def test_environment_run_quit(queue_container: SkeletonQueues):
    """
    Verifies that the command loop exits when a mapped command returns a
    truthy halt flag.
    """
    environment = skeleton_environment(queue_container=queue_container)
    shutdown_event = threading.Event()

    queue_container.environment_command_queue.put(
        "Testing", (GlobalCommands.QUIT, None)
    )
    time.sleep(1)
    environment.run(shutdown_event)
    time.sleep(1)

    logs = []
    while not environment.log_file_queue.empty():
        logs.append(environment.log_file_queue.get_nowait())

    assert any("Starting Process" in message for message in logs)
    assert any("Stopping Process" in message for message in logs)


def test_environment_run_undefined_command(queue_container: SkeletonQueues):
    """
    Verifies that an undefined command is logged and does not halt the
    environment.
    """
    environment = skeleton_environment(queue_container=queue_container)
    shutdown_event = threading.Event()

    queue_container.environment_command_queue.put(
        "Testing", (SkeletonCommands.EXAMPLE_UNDEFINED_COMMAND, None)
    )
    time.sleep(1)
    queue_container.environment_command_queue.put(
        "Testing", (GlobalCommands.QUIT, None)
    )
    time.sleep(1)
    environment.run(shutdown_event)

    logs = []
    while not environment.log_file_queue.empty():
        logs.append(environment.log_file_queue.get_nowait())

    assert any("Undefined Message" in message for message in logs)
    assert any("Stopping Process" in message for message in logs)


def test_environment_run_command_exception():
    """
    Verifies that an exception raised by a mapped command is logged and
    sent to the GUI update queue.
    """
    queue_container = skeleton_queues()
    environment = skeleton_environment(queue_container=queue_container)
    shutdown_event = threading.Event()

    def boom(data):
        raise RuntimeError("BOOM")

    environment.map_command(SkeletonCommands.EXAMPLE_UNDEFINED_COMMAND, boom)

    queue_container.environment_command_queue.put(
        "Testing", (SkeletonCommands.EXAMPLE_UNDEFINED_COMMAND, None)
    )
    time.sleep(1)
    queue_container.environment_command_queue.put(
        "Testing", (GlobalCommands.QUIT, None)
    )
    time.sleep(1)
    environment.run(shutdown_event)

    gui_message, gui_data = environment.gui_update_queue.get_nowait()

    assert gui_message == UICommands.ERROR
    assert gui_data[0] == "Environment Name Error"
    assert "RuntimeError: BOOM" in gui_data[1]


def test_environment_stop_environment(environment: SkeletonEnvironment):
    """
    Verifies that a skeleton environment subclass sets the shutdown flag so
    the control loop can perform a graceful shutdown on its next iteration.
    """
    environment.set_active()
    environment.stop_environment(None)
    assert environment.shutdown_flag is True


@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
def test_environment_quit(environment_type):
    """
    Verifies that this method returns ``True``.
    """
    environment_class = ENVIRONMENT_CLASS[environment_type]

    environment = instantiate_with_mocks(
        environment_class, environment_name="test_environment"
    )
    assert environment.quit(None) is True


@pytest.mark.parametrize("environment_type", IMPLEMENTED_ENVIRONMENT)
@pytest.mark.parametrize("use_thread", [True, False])
def test_processes(use_thread, environment_type):
    """
    Verifies that valid environment process functions receive correct parameters and
    shutdown properly when shutdown event is set.
    """

    if use_thread:
        new_process = threading.Thread
        new_event = threading.Event
        new_queue = thqueue.Queue
    else:
        new_process = mp.Process
        new_event = mp.Event
        new_queue = mp.Queue

    process_function = ENVIRONMENT_PROCESS[environment_type]

    shutdown_event = new_event()
    shutdown_event.set()

    environment_process = new_process(
        target=process_function,
        args=(
            "Skeleton Environment",
            "skeleton_environment_queue",
            VerboseMessageQueue(mp.Queue(), new_queue(), "Command Queue"),
            new_queue(),
            VerboseMessageQueue(mp.Queue(), new_queue(), "Controller Queue"),
            mp.Queue(),
            new_queue(),
            new_queue(),
            new_event(),
            new_event(),
            new_event(),
            new_event(),
            shutdown_event,
            new_event(),
            new_event(),
            new_event(),
            use_thread,
        ),
    )

    environment_process.start()
    environment_process.join(timeout=10)

    if environment_process.is_alive():
        if use_thread:
            pytest.fail(
                f"Thread for {environment_type!r} did not shut down within 10 seconds"
            )
        else:
            environment_process.terminate()
            environment_process.join(timeout=2)

            if environment_process.is_alive():
                environment_process.kill()
                environment_process.join(timeout=2)

            pytest.fail(
                f"Process for {environment_type!r} did not shut down within 10 seconds"
            )
