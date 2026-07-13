import multiprocessing as mp
import queue as thqueue
import threading
import time
from unittest import mock

import netCDF4 as nc4
import numpy as np
import openpyxl
import pytest

from rattlesnake.environment.abstract_sysid_environment import (
    SystemIdCommands,
    SysIdUICommands,
    SysIdEnvironmentMetadata,
    SysIdEnvironment,
)
from rattlesnake.environment.environment_registry import (
    ENVIRONMENT_CLASS,
    ENVIRONMENT_METADATA,
    UNIMPLEMENTED_ENVIRONMENT,
    SYSID_ENVIRONMENTS,
)
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.examples.example_registry import ENVIRONMENT_DICT, SYSID_DICT
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.process.abstract_sysid_data_analysis import (
    SysIdDataAnalysisCommands,
    SysIdDataAnalysisUICommands,
    SysIdDataPackage,
    SysIdMetadata,
)
from rattlesnake.process.data_collector import (
    AcquisitionType,
    CollectorMetadata,
    DataCollectorCommands,
)
from rattlesnake.process.signal_generation import (
    BurstRandomSignalGenerator,
    ChirpSignalGenerator,
    PseudorandomSignalGenerator,
    RandomSignalGenerator,
)
from rattlesnake.process.signal_generation_process import (
    SignalGenerationCommands,
    SignalGenerationMetadata,
)
from rattlesnake.process.spectral_processing import (
    SpectralProcessingCommands,
    SpectralProcessingMetadata,
)
from rattlesnake.testing.mock_utilities import (
    instantiate_with_mocks,
    mock_channel_list_bools,
    skeleton_hardware_metadata,
    skeleton_sysid_environment_metadata,
    skeleton_sysid_environment,
    skeleton_sysid_queues,
)
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.utilities import GlobalCommands, VerboseMessageQueue

# region Fixtures
IMPLEMENTED_SYSID_ENVIRONMENT = [
    environment
    for environment in EnvironmentType
    if environment not in UNIMPLEMENTED_ENVIRONMENT
    and environment in SYSID_ENVIRONMENTS
]


@pytest.fixture
def hardware_metadata():
    metadata = skeleton_hardware_metadata()
    return metadata


@pytest.fixture
def sysid_metadata():
    metadata = SYSID_DICT["manual"](skeleton_hardware_metadata())
    return metadata


@pytest.fixture
def environment_metadata(sysid_metadata):
    return skeleton_sysid_environment_metadata(sysid_metadata=sysid_metadata)


@pytest.fixture
def environment(hardware_metadata, environment_metadata):
    environment = skeleton_sysid_environment()
    environment.hardware_metadata = hardware_metadata
    environment.environment_metadata = environment_metadata
    return environment


# endregion


# region System Id Commands
def test_sysid_commands():
    """
    Iterates through each enum member to confirm unique integer values.
    """
    values = [member.value for member in SystemIdCommands]

    assert all(isinstance(value, int) for value in values)
    assert len(values) == len(set(values))


def test_sysid_ui_commands():
    """
    Iterates through each enum member to confirm unique integer values.
    """
    values = [member.value for member in SysIdUICommands]

    assert all(isinstance(value, int) for value in values)
    assert len(values) == len(set(values))


# endregion


# region System Id Environment Metadata
@pytest.mark.parametrize("environment_type", IMPLEMENTED_SYSID_ENVIRONMENT)
def test_sysid_environment_metadata(environment_type, hardware_metadata):
    """
    Verifies that registered system identification metadata classes initialize
    required base metadata attributes and include ``sysid_metadata``.
    """
    metadata = ENVIRONMENT_DICT[environment_type]["manual"](hardware_metadata)

    assert isinstance(metadata, SysIdEnvironmentMetadata)
    assert metadata.environment_type == environment_type
    assert isinstance(metadata.environment_name, str)
    assert isinstance(metadata.channel_list_bools, list)
    assert isinstance(metadata.sample_rate, int)
    assert isinstance(metadata.sysid_metadata, SysIdMetadata)


def test_sysid_environment_metadata_init_with_supplied_sysid_metadata(sysid_metadata):
    """
    Verifies that supplied system identification metadata is stored.
    """
    metadata = skeleton_sysid_environment_metadata(sysid_metadata=sysid_metadata)

    assert metadata.sysid_metadata is sysid_metadata


def test_sysid_environment_metadata_init_with_default_sysid_metadata():
    """
    Verifies that default system identification metadata is created when none
    is supplied.
    """
    metadata = skeleton_sysid_environment_metadata(sysid_metadata=None)

    assert isinstance(metadata.sysid_metadata, SysIdMetadata)


def test_sysid_environment_metadata_number_of_channels(environment_metadata):
    """
    Verifies that subclasses return the correct number of environment channels.
    """
    assert environment_metadata.number_of_channels == 2


def test_sysid_environment_metadata_response_channel_indices(environment_metadata):
    """
    Verifies that subclasses return valid response channel indices.
    """
    assert environment_metadata.response_channel_indices == [0]


def test_sysid_environment_metadata_reference_channel_indices(environment_metadata):
    """
    Verifies that subclasses return valid reference channel indices.
    """
    assert environment_metadata.reference_channel_indices == [1]


def test_sysid_environment_metadata_num_response_channels_without_transformation():
    """
    Verifies that response channel count uses physical response channels when
    no transformation matrix is supplied.
    """
    metadata = skeleton_sysid_environment_metadata(response_transformation_matrix=None)

    assert metadata.num_response_channels == 1


def test_sysid_environment_metadata_num_response_channels_with_transformation():
    """
    Verifies that response channel count uses the number of rows in the
    response transformation matrix when one is supplied.
    """
    response_transformation_matrix = np.ones((3, 1))
    metadata = skeleton_sysid_environment_metadata(
        response_transformation_matrix=response_transformation_matrix
    )

    assert metadata.num_response_channels == 3


def test_sysid_environment_metadata_num_reference_channels_without_transformation():
    """
    Verifies that reference channel count uses physical reference channels when
    no transformation matrix is supplied.
    """
    metadata = skeleton_sysid_environment_metadata(reference_transformation_matrix=None)

    assert metadata.num_reference_channels == 1


def test_sysid_environment_metadata_num_reference_channels_with_transformation():
    """
    Verifies that reference channel count uses the number of rows in the
    reference transformation matrix when one is supplied.
    """
    reference_transformation_matrix = np.ones((2, 1))
    metadata = skeleton_sysid_environment_metadata(
        reference_transformation_matrix=reference_transformation_matrix
    )

    assert metadata.num_reference_channels == 2


def test_sysid_environment_metadata_response_transformation_matrix():
    """
    Verifies that subclasses return the configured response transformation
    matrix.
    """
    response_transformation_matrix = np.array([[1.0], [2.0]])
    metadata = skeleton_sysid_environment_metadata(
        response_transformation_matrix=response_transformation_matrix
    )

    np.testing.assert_array_equal(
        metadata.response_transformation_matrix,
        response_transformation_matrix,
    )


def test_sysid_environment_metadata_reference_transformation_matrix():
    """
    Verifies that subclasses return the configured reference transformation
    matrix.
    """
    reference_transformation_matrix = np.array([[1.0], [2.0]])
    metadata = skeleton_sysid_environment_metadata(
        reference_transformation_matrix=reference_transformation_matrix
    )

    np.testing.assert_array_equal(
        metadata.reference_transformation_matrix,
        reference_transformation_matrix,
    )


def test_sysid_environment_metadata_validate_truth(
    environment_metadata, hardware_metadata
):
    """
    Verifies that valid system identification metadata passes validation.
    """
    environment_metadata.validate(hardware_metadata)


def test_sysid_environment_metadata_eq_truth(sysid_metadata):
    """
    Verifies that equivalent metadata objects compare equal.
    """
    metadata_1 = skeleton_sysid_environment_metadata(sysid_metadata=sysid_metadata)
    metadata_2 = skeleton_sysid_environment_metadata(sysid_metadata=sysid_metadata)

    assert metadata_1 == metadata_2


def test_sysid_environment_metadata_eq_false(sysid_metadata):
    """
    Verifies that mismatched metadata objects compare unequal.
    """
    metadata_1 = skeleton_sysid_environment_metadata(
        environment_name="Environment A",
        sysid_metadata=sysid_metadata,
    )
    metadata_2 = skeleton_sysid_environment_metadata(
        environment_name="Environment B",
        sysid_metadata=sysid_metadata,
    )

    assert metadata_1 != metadata_2


@pytest.mark.parametrize("environment_type", IMPLEMENTED_SYSID_ENVIRONMENT)
def test_sysid_environment_metadata_load_save_netcdf(
    environment_type, tmp_path, hardware_metadata
):
    """
    Saves a registered system identification metadata subclass to a netCDF file
    and loads it back into a metadata object.
    """
    metadata_class = ENVIRONMENT_METADATA[environment_type]
    metadata = ENVIRONMENT_DICT[environment_type]["manual"](hardware_metadata)
    metadata.environment_name = "Environment Name"

    path = tmp_path / "sysid_metadata.nc4"

    with nc4.Dataset(path, "w") as dataset:
        group = dataset.createGroup(metadata.environment_name)
        metadata.save_metadata_to_netcdf(group)

    with nc4.Dataset(path, "r") as dataset:
        load_group = dataset.groups["Environment Name"]
        loaded = metadata_class.load_metadata_from_netcdf(
            load_group,
            environment_name="Environment Name",
            channel_list_bools=mock_channel_list_bools(),
            hardware_metadata=hardware_metadata,
        )

    assert isinstance(loaded, SysIdEnvironmentMetadata)
    assert loaded.environment_name == "Environment Name"
    assert loaded.channel_list_bools == [True, True]
    assert loaded.sample_rate == hardware_metadata.sample_rate
    assert isinstance(loaded.sysid_metadata, SysIdMetadata)
    loaded.validate(hardware_metadata)


@pytest.mark.parametrize("environment_type", IMPLEMENTED_SYSID_ENVIRONMENT)
def test_sysid_environment_metadata_load_save_worksheet(
    environment_type, hardware_metadata
):
    """
    Saves a registered system identification metadata subclass to an Excel
    worksheet and loads it back into a metadata object.
    """
    metadata_class = ENVIRONMENT_METADATA[environment_type]
    metadata = ENVIRONMENT_DICT[environment_type]["manual"](hardware_metadata)
    metadata.environment_name = "Environment Name"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    metadata.save_metadata_to_worksheet(worksheet)

    assert worksheet.cell(1, 1).value == "Control Type"
    assert worksheet.cell(1, 3).value == "v4.0"

    loaded = metadata_class.load_metadata_from_worksheet(
        worksheet=worksheet,
        environment_name="Environment Name",
        channel_list_bools=mock_channel_list_bools(),
        hardware_metadata=hardware_metadata,
    )

    assert isinstance(loaded, SysIdEnvironmentMetadata)
    assert loaded.environment_name == "Environment Name"
    assert loaded.channel_list_bools == [True, True]
    assert loaded.sample_rate == hardware_metadata.sample_rate
    assert isinstance(loaded.sysid_metadata, SysIdMetadata)
    loaded.validate(hardware_metadata)


def test_sysid_environment_metadata_save_sysid_matrix_to_worksheet():
    """
    Verifies that response and output transformation matrices are written to
    the worksheet.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    response_matrix = np.array([[1.0, 2.0], [3.0, 4.0]])
    output_matrix = np.array([[5.0, 6.0]])

    SysIdEnvironmentMetadata.save_sysid_matrix_to_worksheet(
        worksheet,
        response_matrix,
        output_matrix,
        start_row=5,
    )

    assert worksheet.cell(5, 2).value == 1.0
    assert worksheet.cell(5, 3).value == 2.0
    assert worksheet.cell(6, 2).value == 3.0
    assert worksheet.cell(6, 3).value == 4.0
    assert worksheet.cell(7, 1).value == "Output Transformation Matrix:"
    assert worksheet.cell(7, 2).value == 5.0
    assert worksheet.cell(7, 3).value == 6.0


def test_sysid_environment_metadata_save_sysid_matrix_to_worksheet_none():
    """
    Verifies that ``None`` transformation matrices are written as ``"None"``.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    SysIdEnvironmentMetadata.save_sysid_matrix_to_worksheet(
        worksheet,
        response_matrix=None,
        output_matrix=None,
        start_row=5,
    )

    assert worksheet.cell(5, 2).value == "None"
    assert worksheet.cell(6, 2).value == "None"


def test_sysid_environment_metadata_load_sysid_matrix_from_worksheet():
    """
    Verifies that response and output transformation matrices are loaded from
    worksheet values.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(5, 1, "Response Transformation Matrix:")
    worksheet.cell(5, 2, 1.0)
    worksheet.cell(5, 3, 2.0)
    worksheet.cell(6, 2, 3.0)
    worksheet.cell(6, 3, 4.0)

    worksheet.cell(7, 1, "Output Transformation Matrix:")
    worksheet.cell(7, 2, 5.0)
    worksheet.cell(7, 3, 6.0)

    response_matrix, output_matrix = (
        SysIdEnvironmentMetadata.load_sysid_matrix_from_worksheet(
            worksheet,
            start_row=5,
        )
    )

    np.testing.assert_array_equal(
        response_matrix,
        np.array([[1.0, 2.0], [3.0, 4.0]]),
    )
    np.testing.assert_array_equal(output_matrix, np.array([[5.0, 6.0]]))


def test_sysid_environment_metadata_load_sysid_matrix_from_worksheet_none():
    """
    Verifies that worksheet ``"None"`` values are loaded as ``None`` matrices.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(5, 2, "None")
    worksheet.cell(6, 2, "None")

    response_matrix, output_matrix = (
        SysIdEnvironmentMetadata.load_sysid_matrix_from_worksheet(
            worksheet,
            start_row=5,
        )
    )

    assert response_matrix is None
    assert output_matrix is None


# endregion


# region System Id Environment
@pytest.mark.parametrize("environment_type", IMPLEMENTED_SYSID_ENVIRONMENT)
def test_sysid_environment(environment_type):
    """
    Verifies that registered system identification environment subclasses
    initialize required environment attributes.
    """
    environment_class = ENVIRONMENT_CLASS[environment_type]

    environment = instantiate_with_mocks(
        environment_class,
        environment_name="test_environment",
    )

    assert isinstance(environment, SysIdEnvironment)
    assert environment.environment_name == "test_environment"


def test_sysid_environment_init(environment):
    """
    Confirms that initialization stores queues and events, initializes system
    identification state, and maps system identification commands.
    """
    assert environment.environment_name == "Skeleton SysId Environment"
    assert environment.queue_name == "Mock SysId Queue"

    assert isinstance(environment.sysid_data, SysIdDataPackage)

    assert environment.collector_shutdown_achieved is True
    assert environment.spectral_shutdown_achieved is True
    assert environment.siggen_shutdown_achieved is True
    assert environment.analysis_shutdown_achieved is True

    assert GlobalCommands.INITIALIZE_SYSTEM_ID in environment.command_map
    assert GlobalCommands.START_SYSTEM_ID_NOISE in environment.command_map
    assert GlobalCommands.START_SYSTEM_ID_TRANSFER in environment.command_map
    assert GlobalCommands.STOP_SYSTEM_ID in environment.command_map
    assert GlobalCommands.SAVE_SYSTEM_ID in environment.command_map
    assert GlobalCommands.LOAD_SYSTEM_ID in environment.command_map
    assert SystemIdCommands.CHECK_FOR_COMPLETE_SHUTDOWN in environment.command_map


def test_sysid_environment_sysid_active(environment):
    """
    Verifies that the ``sysid_active`` property reflects the event state.
    """
    assert environment.sysid_active is False

    environment.set_sysid_active()

    assert environment.sysid_active is True


def test_sysid_environment_set_sysid_active(environment):
    """
    Verifies that calling this method sets the system identification active
    event.
    """
    environment.set_sysid_active()

    assert environment.sysid_active is True


def test_sysid_environment_clear_sysid_active(environment):
    """
    Verifies that calling this method clears the system identification active
    event.
    """
    environment.set_sysid_active()
    environment.clear_sysid_active()

    assert environment.sysid_active is False


def test_sysid_environment_sysid_stored(environment):
    """
    Verifies that the ``sysid_stored`` property reflects the event state.
    """
    assert environment.sysid_stored is False

    environment.set_sysid_stored()

    assert environment.sysid_stored is True


def test_sysid_environment_set_sysid_stored(environment):
    """
    Verifies that calling this method sets the system identification stored
    event.
    """
    environment.set_sysid_stored()

    assert environment.sysid_stored is True


def test_sysid_environment_clear_sysid_stored(environment):
    """
    Verifies that calling this method clears the system identification stored
    event.
    """
    environment.set_sysid_stored()
    environment.clear_sysid_stored()

    assert environment.sysid_stored is False


def test_sysid_environment_initialize_hardware(environment, hardware_metadata):
    """
    Verifies that hardware metadata is stored and that the environment is
    marked ready.
    """
    environment.clear_ready()

    environment.initialize_hardware(hardware_metadata)

    assert environment.hardware_metadata is hardware_metadata
    assert environment.ready is True


def test_sysid_environment_initialize_environment(environment, environment_metadata):
    """
    Verifies that environment metadata is stored and the data analysis process
    receives environment initialization information.
    """
    environment.clear_ready()

    environment.initialize_environment(environment_metadata)

    assert environment.environment_metadata is environment_metadata
    assert environment.environment_name == environment_metadata.environment_name
    assert environment.ready is True

    message, data = environment.data_analysis_command_queue.get(
        environment_metadata.environment_name
    )

    assert message == GlobalCommands.INITIALIZE_ENVIRONMENT
    assert data == "Skeleton SysId Environment"


def test_sysid_environment_initialize_sysid(environment, sysid_metadata):
    """
    Verifies that system identification metadata is stored and forwarded to the
    data analysis process.
    """
    environment.clear_ready()

    environment.initialize_sysid(sysid_metadata)

    assert environment.environment_metadata.sysid_metadata is sysid_metadata
    assert environment.ready is True

    message, data = environment.data_analysis_command_queue.get(
        environment.environment_name
    )

    assert message == SysIdDataAnalysisCommands.INITIALIZE_PARAMETERS
    assert data == sysid_metadata


def test_sysid_environment_get_sysid_data_collector_metadata(environment):
    """
    Verifies that collector metadata is populated from system identification
    metadata.
    """
    collector_metadata = environment.get_sysid_data_collector_metadata()

    assert isinstance(collector_metadata, CollectorMetadata)
    assert collector_metadata.num_channels == 2
    assert collector_metadata.response_channel_indices == [0]
    assert collector_metadata.reference_channel_indices == [1]
    assert collector_metadata.acquisition_type == AcquisitionType.FREE_RUN


def test_sysid_environment_get_sysid_spectral_processing_metadata_noise(environment):
    """
    Verifies that spectral processing metadata is populated for noise
    measurements.
    """
    spectral_metadata = environment.get_sysid_spectral_processing_metadata(
        is_noise=True
    )

    assert isinstance(spectral_metadata, SpectralProcessingMetadata)


def test_sysid_environment_get_sysid_spectral_processing_metadata_transfer(
    environment,
):
    """
    Verifies that spectral processing metadata is populated for transfer
    function measurements.
    """
    spectral_metadata = environment.get_sysid_spectral_processing_metadata(
        is_noise=False
    )

    assert isinstance(spectral_metadata, SpectralProcessingMetadata)


def test_sysid_environment_get_sysid_spectral_processing_metadata_invalid_estimator(
    environment,
):
    """
    Verifies that an invalid FRF estimator raises a ``ValueError``.
    """
    original_estimator = environment.environment_metadata.sysid_metadata.sysid_estimator
    environment.environment_metadata.sysid_metadata.sysid_estimator = "Invalid"

    with pytest.raises(ValueError):
        environment.get_sysid_spectral_processing_metadata()

    environment.environment_metadata.sysid_metadata.sysid_estimator = original_estimator


def test_sysid_environment_get_sysid_signal_generation_metadata(environment):
    """
    Verifies that signal generation metadata is populated correctly.
    """
    signal_generation_metadata = environment.get_sysid_signal_generation_metadata()

    assert isinstance(signal_generation_metadata, SignalGenerationMetadata)
    assert signal_generation_metadata.samples_per_write == 128


@pytest.mark.parametrize(
    "signal_type,expected_class",
    [
        ("Random", RandomSignalGenerator),
        ("Pseudorandom", PseudorandomSignalGenerator),
        ("Burst Random", BurstRandomSignalGenerator),
        ("Chirp", ChirpSignalGenerator),
    ],
)
def test_sysid_environment_get_sysid_signal_generator(
    environment,
    signal_type,
    expected_class,
):
    """
    Verifies that the correct signal generator type is created for each
    supported system identification signal type.
    """
    environment.environment_metadata.sysid_metadata.sysid_signal_type = signal_type

    signal_generator = environment.get_sysid_signal_generator()

    assert isinstance(signal_generator, expected_class)


def test_sysid_environment_load_noise(environment):
    """
    Verifies that noise data is forwarded to the data analysis command queue.
    """
    data = object()

    environment.load_noise(data)

    message, queued_data = environment.data_analysis_command_queue.get(
        environment.environment_name
    )

    assert message == SysIdDataAnalysisCommands.LOAD_NOISE
    assert isinstance(queued_data, object)


def test_sysid_environment_load_transfer_function(environment):
    """
    Verifies that transfer-function data is forwarded to the data analysis
    command queue.
    """
    data = object()

    environment.load_transfer_function(data)

    message, queued_data = environment.data_analysis_command_queue.get(
        environment.environment_name
    )

    assert message == SysIdDataAnalysisCommands.LOAD_TRANSFER_FUNCTION
    assert isinstance(queued_data, object)


def test_sysid_environment_save_system_id_to_file_npz(
    environment,
    tmp_path,
):
    """
    Verifies that system identification data can be saved to a NumPy ``.npz``
    file and that the environment is marked ready afterward.
    """
    path = tmp_path / "sysid_data.npz"

    environment.clear_ready()
    environment.save_system_id_to_file(path)

    assert path.exists()
    assert environment.ready is True


def test_sysid_environment_load_system_id_from_package():
    """
    Verifies that loaded system identification data is forwarded to data
    analysis and GUI update queues.
    """
    sysid_data = SysIdDataPackage()
    sysid_data.sysid_frames = 1
    sysid_data.frequencies = np.array([1.0])
    sysid_data.sysid_frf = np.array([1.0])
    sysid_data.sysid_coherence = np.array([1.0])
    sysid_data.sysid_response_cpsd = np.array([1.0])
    sysid_data.sysid_reference_cpsd = np.array([1.0])
    sysid_data.sysid_condition = np.array([1.0])

    queue_container = skeleton_sysid_queues()
    environment = skeleton_sysid_environment(queue_container=queue_container)
    environment.environment_metadata = skeleton_sysid_environment_metadata()
    environment.load_system_id_from_package(sysid_data)

    time.sleep(1)

    message, data = queue_container.data_analysis_command_queue.get(
        environment.environment_name, timeout=0.1
    )

    assert message == SysIdDataAnalysisCommands.LOAD_SYSTEM_ID
    assert isinstance(data, SysIdDataPackage)

    gui_environment_name, gui_payload = queue_container.gui_update_queue.get_nowait()

    assert gui_environment_name == environment.environment_name
    assert gui_payload[0] == SysIdDataAnalysisUICommands.SYSID_UPDATE

    message, data = environment.environment_command_queue.get(
        environment.environment_name
    )

    assert message == SysIdDataAnalysisCommands.SYSTEM_ID_COMPLETE
    assert isinstance(data[1], SysIdDataPackage)


def test_sysid_environment_start_noise(environment):
    """
    Verifies that all required subprocesses are commanded when starting a noise
    measurement.
    """
    environment.start_noise(None)

    assert environment.sysid_active is True

    message, _ = environment.collector_command_queue.get(environment.environment_name)
    assert message == DataCollectorCommands.FORCE_INITIALIZE_COLLECTOR

    message, _ = environment.signal_generator_command_queue.get(
        environment.environment_name
    )
    assert message == SignalGenerationCommands.INITIALIZE_PARAMETERS

    message, _ = environment.data_analysis_command_queue.get(
        environment.environment_name
    )
    assert message == SysIdDataAnalysisCommands.RUN_NOISE

    message, _ = environment.spectral_processing_command_queue.get(
        environment.environment_name
    )
    assert message == SpectralProcessingCommands.INITIALIZE_PARAMETERS

    gui_environment_name, gui_payload = environment.gui_update_queue.get_nowait()

    assert gui_environment_name == environment.environment_name
    assert gui_payload == (SysIdUICommands.SYSID_STARTED, None)


def test_sysid_environment_start_transfer_function(environment):
    """
    Verifies that all required subprocesses are commanded when starting a
    transfer-function measurement.
    """
    environment.start_transfer_function(None)
    time.sleep(1)

    assert environment.sysid_active is True

    message, _ = environment.collector_command_queue.get(environment.environment_name)
    assert message == DataCollectorCommands.FORCE_INITIALIZE_COLLECTOR

    message, _ = environment.signal_generator_command_queue.get(
        environment.environment_name
    )
    assert message == SignalGenerationCommands.INITIALIZE_PARAMETERS

    message, _ = environment.data_analysis_command_queue.get(
        environment.environment_name
    )
    assert message == SysIdDataAnalysisCommands.RUN_TRANSFER_FUNCTION

    message, _ = environment.spectral_processing_command_queue.get(
        environment.environment_name
    )
    assert message == SpectralProcessingCommands.INITIALIZE_PARAMETERS

    gui_environment_name, gui_payload = environment.gui_update_queue.get_nowait()

    assert gui_environment_name == environment.environment_name
    assert gui_payload == (SysIdUICommands.SYSID_STARTED, None)


def test_sysid_environment_stop_system_id(environment):
    """
    Verifies that shutdown commands are sent to system identification
    subprocesses and that a shutdown check is queued.
    """
    environment.stop_system_id(True)
    time.sleep(1)

    message, _ = environment.collector_command_queue.get(environment.environment_name)
    assert message == DataCollectorCommands.SET_TEST_LEVEL

    message, _ = environment.signal_generator_command_queue.get(
        environment.environment_name
    )
    assert message == SignalGenerationCommands.START_SHUTDOWN

    message, _ = environment.spectral_processing_command_queue.get(
        environment.environment_name
    )
    assert message == SpectralProcessingCommands.STOP_SPECTRAL_PROCESSING

    message, _ = environment.data_analysis_command_queue.get(
        environment.environment_name
    )
    assert message == SysIdDataAnalysisCommands.STOP_SYSTEM_ID

    message, _ = environment.environment_command_queue.get(environment.environment_name)
    assert message == SystemIdCommands.CHECK_FOR_COMPLETE_SHUTDOWN


def test_sysid_environment_siggen_shutdown_achieved_fn(environment):
    """
    Verifies that the signal generation shutdown flag is set.
    """
    environment.siggen_shutdown_achieved = False

    environment.siggen_shutdown_achieved_fn(None)

    assert environment.siggen_shutdown_achieved is True


def test_sysid_environment_collector_shutdown_achieved_fn(environment):
    """
    Verifies that the collector shutdown flag is set.
    """
    environment.collector_shutdown_achieved = False

    environment.collector_shutdown_achieved_fn(None)

    assert environment.collector_shutdown_achieved is True


def test_sysid_environment_spectral_shutdown_achieved_fn(environment):
    """
    Verifies that the spectral processing shutdown flag is set.
    """
    environment.spectral_shutdown_achieved = False

    environment.spectral_shutdown_achieved_fn(None)

    assert environment.spectral_shutdown_achieved is True


def test_sysid_environment_analysis_shutdown_achieved_fn(environment):
    """
    Verifies that the data analysis shutdown flag is set.
    """
    environment.analysis_shutdown_achieved = False

    environment.analysis_shutdown_achieved_fn(None)

    assert environment.analysis_shutdown_achieved is True


def test_sysid_environment_check_for_sysid_shutdown_complete(environment):
    """
    Verifies that system identification is marked inactive when all shutdown
    flags are complete.
    """
    environment.set_sysid_active()
    environment.siggen_shutdown_achieved = True
    environment.collector_shutdown_achieved = True
    environment.spectral_shutdown_achieved = True
    environment.analysis_shutdown_achieved = True

    environment.check_for_sysid_shutdown(None)
    time.sleep(1)

    assert environment.sysid_active is False

    gui_environment_name, gui_payload = environment.gui_update_queue.get_nowait()

    assert gui_environment_name == environment.environment_name
    assert gui_payload == (SysIdUICommands.SYSID_ENDED, None)


def test_sysid_environment_check_for_sysid_shutdown_incomplete(environment):
    """
    Verifies that incomplete shutdown state is logged and another shutdown
    check is queued.
    """
    environment.siggen_shutdown_achieved = False
    environment.collector_shutdown_achieved = True
    environment.spectral_shutdown_achieved = True
    environment.analysis_shutdown_achieved = True

    with mock.patch("rattlesnake.environment.abstract_sysid_environment.time.sleep"):
        environment.check_for_sysid_shutdown(None)

    message, _ = environment.environment_command_queue.get(environment.environment_name)

    assert message == SystemIdCommands.CHECK_FOR_COMPLETE_SHUTDOWN


def test_sysid_environment_system_id_noise_complete(environment):
    """
    Verifies that noise completion is logged and forwarded to the GUI.
    """
    data = object()

    environment.system_id_noise_complete(data)
    time.sleep(1)

    gui_environment_name, gui_payload = environment.gui_update_queue.get_nowait()

    assert gui_environment_name == environment.environment_name
    assert gui_payload[0] == SysIdDataAnalysisUICommands.NOISE_COMPLETED
    assert isinstance(gui_payload[1], object)


def test_sysid_environment_system_id_complete(environment, sysid_metadata):
    """
    Verifies that completed system identification data is stored and GUI
    completion messages are sent.
    """
    sysid_data = SysIdDataPackage()
    data = (sysid_metadata, sysid_data)

    environment.system_id_complete(data)
    time.sleep(1)

    assert environment.sysid_data is sysid_data
    assert environment.sysid_stored is True

    gui_environment_name, gui_payload = environment.gui_update_queue.get_nowait()

    assert gui_environment_name == environment.environment_name
    assert gui_payload[0] == SysIdDataAnalysisUICommands.TRANSFER_COMPLETED
    assert isinstance(gui_payload[1][0], SysIdMetadata)
    assert isinstance(gui_payload[1][1], SysIdDataPackage)

    gui_message, gui_data = environment.gui_update_queue.get_nowait()

    assert gui_message == UICommands.COMPLETED_SYSTEM_ID
    assert gui_data[0] == environment.environment_name
    assert isinstance(gui_data[1][0], SysIdMetadata)
    assert isinstance(gui_data[1][1], SysIdDataPackage)


def test_sysid_environment_stop_environment(environment):
    """
    Verifies that a system identification environment subclass performs
    graceful shutdown behavior.
    """
    environment.set_active()

    environment.stop_environment(None)

    assert environment.active is False


def test_sysid_environment_quit(environment):
    """
    Verifies that quit commands are sent to all system identification
    subprocesses and that this method returns ``True``.
    """
    assert environment.quit(None) is True

    for command_queue in [
        environment.collector_command_queue,
        environment.signal_generator_command_queue,
        environment.spectral_processing_command_queue,
        environment.data_analysis_command_queue,
    ]:
        message, data = command_queue.get(environment.environment_name)

        assert message == GlobalCommands.QUIT
        assert data is None


@pytest.mark.parametrize("environment_type", IMPLEMENTED_SYSID_ENVIRONMENT)
@pytest.mark.parametrize("use_thread", [True, False])
def test_sysid_processes(use_thread, environment_type):
    """
    Verifies that registered system identification environment process
    functions receive correct parameters and shut down properly when the
    shutdown event is set.
    """
    from rattlesnake.environment.environment_registry import ENVIRONMENT_PROCESS

    if use_thread:
        new_process = threading.Thread
        new_event = threading.Event
        new_queue = thqueue.Queue
    else:
        new_process = mp.Process
        new_event = mp.Event
        new_queue = mp.Queue

    process_function = ENVIRONMENT_PROCESS[environment_type]

    log_file_queue = mp.Queue()
    shutdown_event = new_event()
    shutdown_event.set()

    environment_process = new_process(
        target=process_function,
        args=(
            "Skeleton SysId Environment",
            "mock_sysid_environment_queue",
            VerboseMessageQueue(log_file_queue, new_queue(), "Command Queue"),
            new_queue(),
            VerboseMessageQueue(log_file_queue, new_queue(), "Controller Queue"),
            log_file_queue,
            new_queue(),
            new_queue(),
            new_event(),
            new_event(),
            new_event(),
            new_event(),
            shutdown_event,
            new_event(),
            new_event(),
            new_event(),
            use_thread,
        ),
    )

    environment_process.start()
    environment_process.join(timeout=10)

    if environment_process.is_alive():
        if use_thread:
            pytest.fail(
                f"Thread for {environment_type!r} did not shut down within 10 seconds"
            )

        environment_process.terminate()
        environment_process.join(timeout=2)

        if environment_process.is_alive():
            environment_process.kill()
            environment_process.join(timeout=2)

        pytest.fail(
            f"Process for {environment_type!r} did not shut down within 10 seconds"
        )


# endregion
