from types import SimpleNamespace
from unittest import mock

import netCDF4 as nc4
import numpy as np
import openpyxl
import pytest

from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.environment.time_environment import TimeCommands
from rattlesnake.hardware.hardware_utilities import HardwareType
from rattlesnake.load_utilities import (
    discover_environment_type_in_old_netcdf,
    load_metadata_from_netcdf,
    load_metadata_from_workbook,
    load_profile_from_workbook,
    save_profile_to_workbook,
    save_rattlesnake_to_workbook,
)
from rattlesnake.profile_manager import ProfileEvent
from rattlesnake.hardware.skeleton_hardware import SkeletonHardwareMetadata
from rattlesnake.environment.skeleton_environment import (
    SkeletonMetadata,
    SkeletonCommands,
)
from rattlesnake.testing.mock_utilities import (
    skeleton_hardware_metadata,
    skeleton_environment_metadata,
)
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.utilities import GlobalCommands, RattlesnakeError


# region Fixtures
@pytest.fixture
def hardware_type():
    return HardwareType.SKELETON


@pytest.fixture
def environment_type():
    return EnvironmentType.SKELETON


@pytest.fixture
def mock_hardware_metadata():
    return skeleton_hardware_metadata()


@pytest.fixture
def mock_environment_metadata():
    metadata = skeleton_environment_metadata()
    metadata.channel_list_bools = [True, False]
    return metadata


# endregion


# region netCDF Metadata Loading and Saving
def test_load_metadata_from_netcdf(hardware_type, environment_type):
    """
    Verifies that hardware metadata and environment metadata are loaded from a
    netCDF dataset using the registered metadata classes.
    """
    dataset = mock.MagicMock()
    dataset.hardware = hardware_type.value
    dataset.variables = {
        "environment_names": mock.MagicMock(),
        "environment_types": mock.MagicMock(),
        "environment_active_channels": mock.MagicMock(),
    }
    dataset.variables["environment_names"].__getitem__.return_value = np.array(
        ["Env A"]
    )
    dataset.variables["environment_types"].__getitem__.return_value = (
        environment_type.value
    )
    dataset.variables["environment_active_channels"].__getitem__.return_value = (
        np.array([True, False])
    )

    environment_group = mock.MagicMock()
    dataset.groups = {"Env A": environment_group}

    with (
        mock.patch.dict(
            "rattlesnake.load_utilities.HARDWARE_METADATA",
            {hardware_type: SkeletonHardwareMetadata},
            clear=False,
        ),
        mock.patch.dict(
            "rattlesnake.load_utilities.ENVIRONMENT_METADATA",
            {environment_type: SkeletonMetadata},
            clear=False,
        ),
    ):
        hardware_metadata, environment_metadata_list = load_metadata_from_netcdf(
            dataset
        )

    assert isinstance(hardware_metadata, SkeletonHardwareMetadata)
    assert hardware_metadata.hardware_type == hardware_type

    assert len(environment_metadata_list) == 1
    environment_metadata = environment_metadata_list[0]
    assert isinstance(environment_metadata, SkeletonMetadata)
    assert environment_metadata.environment_name == "Env A"
    assert (environment_metadata.channel_list_bools == [True, False]).all()


def test_load_metadata_from_netcdf_discovers_legacy_environment_type(
    hardware_type,
    environment_type,
):
    """
    Verifies that old netCDF files without explicit environment type variables
    use legacy environment type discovery.
    """
    dataset = mock.MagicMock()
    dataset.hardware = hardware_type.value
    dataset.variables = {
        "environment_names": mock.MagicMock(),
        "environment_types": mock.MagicMock(),
        "environment_active_channels": mock.MagicMock(),
    }
    dataset.variables["environment_names"].__getitem__.return_value = np.array(
        ["Env A"]
    )
    dataset.variables["environment_types"].__getitem__.side_effect = KeyError("missing")
    dataset.variables["environment_active_channels"].__getitem__.return_value = (
        np.array([True, False])
    )

    environment_group = SimpleNamespace(
        cancel_rampdown_time=0.1, example_window_size="Example Window Size"
    )
    dataset.groups = {"Env A": environment_group}

    with (
        mock.patch.dict(
            "rattlesnake.load_utilities.HARDWARE_METADATA",
            {hardware_type: SkeletonHardwareMetadata},
            clear=False,
        ),
        mock.patch.dict(
            "rattlesnake.load_utilities.ENVIRONMENT_METADATA",
            {EnvironmentType.TIME: SkeletonMetadata},
            clear=False,
        ),
    ):
        _, environment_metadata_list = load_metadata_from_netcdf(dataset)

    assert len(environment_metadata_list) == 1
    assert environment_metadata_list[0].environment_name == "Env A"


@pytest.mark.parametrize(
    "attribute_name, expected_environment_type",
    [
        ("tracking_filter_type", EnvironmentType.SINE),
        ("update_tf_during_control", EnvironmentType.RANDOM),
        ("num_averages", EnvironmentType.MODAL),
        ("test_level_ramp_time", EnvironmentType.TRANSIENT),
        ("cancel_rampdown_time", EnvironmentType.TIME),
    ],
)
def test_discover_environment_type_in_old_netcdf(
    attribute_name,
    expected_environment_type,
):
    """
    Verifies that known legacy netCDF attributes map to expected environment
    types.
    """
    group = SimpleNamespace()
    setattr(group, attribute_name, True)

    assert discover_environment_type_in_old_netcdf(group) == expected_environment_type


def test_discover_environment_type_in_old_netcdf_invalid():
    """
    Verifies that unrecognized legacy netCDF groups raise ``RattlesnakeError``.
    """
    group = SimpleNamespace()

    with pytest.raises(RattlesnakeError):
        discover_environment_type_in_old_netcdf(group)


# endregion


# region Workbook Metadata Loading and Saving
def build_metadata_workbook(hardware_type, environment_type):
    workbook = openpyxl.Workbook()
    channel_sheet = workbook.active
    channel_sheet.title = "Channel Table"

    hardware_sheet = workbook.create_sheet("Hardware")
    hardware_sheet.cell(1, 1, "Hardware Type")
    hardware_sheet.cell(1, 2, str(hardware_type.value))

    # Environment membership columns start at column 24.
    channel_sheet.cell(row=2, column=24, value="Env A")
    channel_sheet.cell(row=3, column=24, value="x")
    channel_sheet.cell(row=4, column=24, value="")

    environment_sheet = workbook.create_sheet("Env A")
    environment_sheet.cell(1, 1, "Control Type")
    environment_sheet.cell(1, 2, environment_type.name)
    environment_sheet.cell(2, 1, "Example Window Size")
    environment_sheet.cell(2, 2, "Example Window Size")

    profile_sheet = workbook.create_sheet("Test Profile")
    profile_sheet.cell(1, 1, "Time (s)")
    profile_sheet.cell(1, 2, "Environment")
    profile_sheet.cell(1, 3, "Operation")
    profile_sheet.cell(1, 4, "Data")
    profile_sheet.cell(2, 1, "0")
    profile_sheet.cell(2, 2, "Global")
    profile_sheet.cell(2, 3, "Start Streaming")
    profile_sheet.cell(2, 4, "")

    return workbook


def test_load_metadata_from_workbook(hardware_type, environment_type):
    """
    Verifies that hardware metadata, environment metadata, and profile events
    are loaded from a workbook.
    """
    workbook = build_metadata_workbook(hardware_type, environment_type)

    with (
        mock.patch.dict(
            "rattlesnake.load_utilities.HARDWARE_METADATA",
            {hardware_type: SkeletonHardwareMetadata},
            clear=False,
        ),
        mock.patch.dict(
            "rattlesnake.load_utilities.ENVIRONMENT_METADATA",
            {environment_type: SkeletonMetadata},
            clear=False,
        ),
    ):
        hardware_metadata, environment_metadata_list, profile_event_list = (
            load_metadata_from_workbook(workbook)
        )

    assert isinstance(hardware_metadata, SkeletonHardwareMetadata)
    assert hardware_metadata.hardware_type == hardware_type

    assert len(environment_metadata_list) == 1
    environment_metadata = environment_metadata_list[0]
    assert environment_metadata.environment_name == "Env A"
    assert environment_metadata.environment_type == environment_type
    assert (
        environment_metadata.channel_list_bools == []
    )  # This is since channel table is blank

    assert len(profile_event_list) == 1
    assert profile_event_list[0].timestamp == 0.0
    assert profile_event_list[0].environment_name == "Global"
    assert profile_event_list[0].command == GlobalCommands.START_STREAMING
    assert profile_event_list[0].data is None


def test_save_rattlesnake_to_workbook(
    mock_hardware_metadata,
    mock_environment_metadata,
):
    """
    Verifies that hardware metadata, environment channel markers, environment
    worksheets, and profile events are written to a workbook.
    """
    workbook = openpyxl.Workbook()
    profile_event = ProfileEvent(
        timestamp=1.5,
        environment_name="Global",
        command=GlobalCommands.START_STREAMING,
        data=None,
    )

    save_rattlesnake_to_workbook(
        workbook,
        hardware_metadata=mock_hardware_metadata,
        environment_metadata_dict={"Env A": mock_environment_metadata},
        profile_event_list=[profile_event],
    )

    assert "Hardware" in workbook.sheetnames
    assert "Env A" in workbook.sheetnames
    assert "Test Profile" in workbook.sheetnames

    channel_sheet = workbook["Channel Table"]
    assert channel_sheet.cell(row=1, column=24).value == "Environments"
    assert channel_sheet.cell(row=2, column=24).value == "Env A"
    assert channel_sheet.cell(row=3, column=24).value == "x"
    assert channel_sheet.cell(row=4, column=24).value is None

    environment_sheet = workbook["Env A"]
    assert environment_sheet.cell(1, 1).value == "Control Type"
    assert environment_sheet.cell(1, 2).value == "Skeleton"

    profile_sheet = workbook["Test Profile"]
    assert profile_sheet.cell(1, 1).value == "Time (s)"
    assert profile_sheet.cell(2, 1).value == str(profile_event.timestamp)
    assert profile_sheet.cell(2, 2).value == "Global"
    assert profile_sheet.cell(2, 3).value == profile_event.command.label


def test_save_rattlesnake_to_workbook_environment_type_template(
    mock_hardware_metadata,
    environment_type,
):
    """
    Verifies that environment type values create blank environment worksheet
    templates.
    """
    workbook = openpyxl.Workbook()

    with mock.patch.dict(
        "rattlesnake.load_utilities.ENVIRONMENT_METADATA",
        {environment_type: SkeletonMetadata},
        clear=False,
    ):
        save_rattlesnake_to_workbook(
            workbook,
            hardware_metadata=mock_hardware_metadata,
            environment_metadata_dict={"Env Template": environment_type},
            profile_event_list=[],
        )

    assert "Env Template" in workbook.sheetnames
    worksheet = workbook["Env Template"]
    assert worksheet.cell(1, 1).value == "Control Type"
    assert worksheet.cell(1, 2).value == "Skeleton"


def test_save_rattlesnake_to_workbook_no_environments(mock_hardware_metadata):
    """
    Verifies that saving with no environments still creates the profile sheet
    and environment column header.
    """
    workbook = openpyxl.Workbook()

    save_rattlesnake_to_workbook(
        workbook,
        hardware_metadata=mock_hardware_metadata,
        environment_metadata_dict={},
        profile_event_list=[],
    )

    assert "Test Profile" in workbook.sheetnames
    assert workbook["Channel Table"].cell(row=1, column=24).value == "Environments"


# endregion


# region Profile Workbook Loading and Saving
def build_profile_workbook():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Test Profile"
    worksheet.cell(1, 1, "Time (s)")
    worksheet.cell(1, 2, "Environment")
    worksheet.cell(1, 3, "Operation")
    worksheet.cell(1, 4, "Data")
    return workbook, worksheet


def test_load_profile_from_workbook_global_command():
    """
    Verifies that global profile commands are loaded correctly from a
    worksheet.
    """
    workbook, worksheet = build_profile_workbook()
    worksheet.cell(2, 1, 0.0)
    worksheet.cell(2, 2, "Global")
    worksheet.cell(2, 3, "Start Streaming")
    worksheet.cell(2, 4, "")

    profile_event_list = load_profile_from_workbook(
        workbook,
        {"Global": "Global"},
    )

    assert len(profile_event_list) == 1
    event = profile_event_list[0]
    assert event.timestamp == 0.0
    assert event.environment_name == "Global"
    assert event.command == GlobalCommands.START_STREAMING
    assert event.data is None


def test_load_profile_from_workbook_environment_command(environment_type):
    """
    Verifies that environment-specific commands are loaded correctly from a
    worksheet.
    """
    workbook, worksheet = build_profile_workbook()
    worksheet.cell(2, 1, 2.5)
    worksheet.cell(2, 2, "Env A")
    worksheet.cell(2, 3, "Example Float Command")
    worksheet.cell(2, 4, 1.25)

    with mock.patch.dict(
        "rattlesnake.load_utilities.ENVIRONMENT_COMMANDS",
        {environment_type: SkeletonCommands},
        clear=False,
    ):
        profile_event_list = load_profile_from_workbook(
            workbook,
            {"Global": "Global", "Env A": environment_type},
        )

    assert len(profile_event_list) == 1
    event = profile_event_list[0]
    assert event.timestamp == 2.5
    assert event.environment_name == "Env A"
    assert event.command == SkeletonCommands.EXAMPLE_FLOAT_COMMAND
    assert event.data == 1.25


def test_load_profile_from_workbook_skips_set_environment_instructions(
    environment_type,
):
    """
    Verifies that ``SET_ENVIRONMENT_INSTRUCTIONS`` rows are skipped.
    """
    workbook, worksheet = build_profile_workbook()
    worksheet.cell(2, 1, 1.0)
    worksheet.cell(2, 2, "Env A")
    worksheet.cell(2, 3, "Set Environment Instructions")
    worksheet.cell(2, 4, "ignored")

    with mock.patch.dict(
        "rattlesnake.load_utilities.ENVIRONMENT_COMMANDS",
        {environment_type: SkeletonCommands},
        clear=False,
    ):
        profile_event_list = load_profile_from_workbook(
            workbook,
            {"Global": "Global", "Env A": environment_type},
        )

    assert profile_event_list == []


def test_load_profile_from_workbook_invalid_command(environment_type):
    """
    Verifies that invalid command strings raise ``RattlesnakeError``.
    """
    workbook, worksheet = build_profile_workbook()
    worksheet.cell(2, 1, 1.0)
    worksheet.cell(2, 2, "Env A")
    worksheet.cell(2, 3, "Not A Command")
    worksheet.cell(2, 4, "")

    with (
        mock.patch.dict(
            "rattlesnake.load_utilities.ENVIRONMENT_COMMANDS",
            {environment_type: SkeletonCommands},
            clear=False,
        ),
        pytest.raises(RattlesnakeError),
    ):
        load_profile_from_workbook(
            workbook,
            {"Global": "Global", "Env A": environment_type},
        )


def test_load_profile_from_workbook_stops_at_blank_row():
    """
    Verifies that profile loading stops at the first blank timestamp row.
    """
    workbook, worksheet = build_profile_workbook()
    worksheet.cell(2, 1, None)
    worksheet.cell(2, 2, "Global")
    worksheet.cell(2, 3, "Start Streaming")

    profile_event_list = load_profile_from_workbook(
        workbook,
        {"Global": "Global"},
    )

    assert profile_event_list == []


def test_save_profile_to_workbook():
    """
    Verifies that profile headers and event rows are written.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    profile_event = ProfileEvent(
        timestamp=3.0,
        environment_name="Global",
        command=GlobalCommands.STOP_STREAMING,
        data=None,
    )

    save_profile_to_workbook(worksheet, [profile_event])

    assert worksheet.cell(1, 1).value == "Time (s)"
    assert worksheet.cell(1, 2).value == "Environment"
    assert worksheet.cell(1, 3).value == "Operation"
    assert worksheet.cell(1, 4).value == "Data"

    assert worksheet.cell(2, 1).value == "3.0"
    assert worksheet.cell(2, 2).value == "Global"
    assert worksheet.cell(2, 3).value == "Stop Streaming"
    assert worksheet.cell(2, 4).value == "None"


def test_save_profile_to_workbook_skips_environment_instruction_event():
    """
    Verifies that instruction-update events are skipped when saving profiles.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    skipped_event = ProfileEvent(
        timestamp=1.0,
        environment_name="Env A",
        command=UICommands.SET_ENVIRONMENT_INSTRUCTIONS,
        data="instructions",
    )
    saved_event = ProfileEvent(
        timestamp=2.0,
        environment_name="Global",
        command=GlobalCommands.START_STREAMING,
        data=None,
    )

    save_profile_to_workbook(worksheet, [skipped_event, saved_event])

    assert worksheet.cell(2, 1).value == "2.0"
    assert worksheet.cell(2, 2).value == "Global"
    assert worksheet.cell(2, 3).value == "Start Streaming"


def test_save_profile_to_workbook_empty_profile():
    """
    Verifies that saving an empty profile writes only headers.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    save_profile_to_workbook(worksheet, [])

    assert worksheet.cell(1, 1).value == "Time (s)"
    assert worksheet.cell(2, 1).value is None


# endregion
