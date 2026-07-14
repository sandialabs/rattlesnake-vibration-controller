import multiprocessing as mp
import queue as thqueue
import random
import socket
import string
import threading
from unittest import mock

import numpy as np
import openpyxl
import pytest
from scipy.io import savemat

from rattlesnake.testing.mock_utilities import (
    clear_log_queue,
    clear_verbose_queue,
    fake_time,
    mock_queue_container,
    mock_event_container,
)
from rattlesnake.utilities import (
    EventContainer,
    GlobalCommands,
    IPAddress,
    OverlapBuffer,
    QueueContainer,
    RattlesnakeError,
    VerboseMessageQueue,
    align_signals,
    autofill_single_ip_address,
    coherence,
    corr_norm_signal_spec,
    corr_norm_spec2,
    correlation_norm_signal_spec_ratio,
    correlation_norm_spec_ratio,
    cpsd_to_time_history,
    db2scale,
    find_lanxi_devices,
    flush_queue,
    load_csv_matrix,
    load_python_module,
    load_time_history,
    log_file_task,
    moving_sum,
    norm_ratio,
    power2db,
    read_transformation_matrix_from_worksheet,
    reduce_array_by_coordinate,
    rms_csd,
    rms_time,
    save_csv_matrix,
    scale2db,
    search_for_lanxi_devices,
    shift_signal,
    check_lanxi_candidate,
    trac,
    wrap,
)


# region Fixtures
@pytest.fixture
def log_file_queue():
    return mp.Queue()


@pytest.fixture
def name_manager():
    manager = mp.Manager()
    yield manager
    manager.shutdown()


@pytest.fixture(params=[True, False], ids=["threaded", "non_threaded"])
def verbose_fixture(request, log_file_queue, name_manager):
    use_thread = request.param

    if use_thread:
        queue = thqueue.Queue()
    else:
        queue = mp.Queue()

    return (
        VerboseMessageQueue(log_file_queue, queue, "VerboseQueue", name_manager),
        use_thread,
    )


@pytest.fixture
def queue_container():
    return mock_queue_container()


@pytest.fixture
def event_container():
    return mock_event_container()


# endregion


# region log_file_task
@mock.patch("builtins.open", new_callable=mock.mock_open)
def test_log_file_process(mock_file, log_file_queue: mp.Queue):
    """
    Verifies that queued log messages are written, quit messages terminate the
    log task, and the file is flushed.
    """
    message = "This is a test\n"
    log_file_queue.put(message)
    log_file_queue.put(GlobalCommands.QUIT)

    shutdown_event = mock.MagicMock()
    shutdown_event.is_set.return_value = False

    log_file_task(log_file_queue, shutdown_event)

    mock_file().write.assert_any_call(message)
    mock_file().write.assert_called_with("Program quitting, logging terminated.")
    mock_file().flush.assert_called()


@mock.patch("builtins.open", new_callable=mock.mock_open)
def test_log_file_process_replaces_extra_newlines(mock_file, log_file_queue: mp.Queue):
    """
    Verifies that multi-line log messages replace all but the final newline
    with slash delimiters.
    """
    log_file_queue.put("line 1\nline 2\nline 3\n")
    log_file_queue.put(GlobalCommands.QUIT)

    shutdown_event = mock.MagicMock()
    shutdown_event.is_set.return_value = False

    log_file_task(log_file_queue, shutdown_event)

    mock_file().write.assert_any_call("line 1////line 2////line 3\n")


@mock.patch("builtins.open", new_callable=mock.mock_open)
def test_log_file_process_shutdown_event_exits_without_read(
    mock_file, log_file_queue: mp.Queue
):
    """
    Verifies that the log task exits immediately if the shutdown event is set.
    """
    shutdown_event = mock.MagicMock()
    shutdown_event.is_set.return_value = True

    log_file_task(log_file_queue, shutdown_event)

    mock_file().write.assert_not_called()


# endregion


# region RattlesnakeError and GlobalCommands
def test_rattlesnake_error():
    """
    Verifies that ``RattlesnakeError`` is an exception type.
    """
    with pytest.raises(RattlesnakeError):
        raise RattlesnakeError("test error")


def test_global_commands_have_unique_integer_values():
    """
    Verifies that all global command values are unique integers.
    """
    values = [command.value for command in GlobalCommands]

    assert all(isinstance(value, int) for value in values)
    assert len(values) == len(set(values))


def test_global_command_label():
    """
    Verifies that command labels are title-case human-readable strings.
    """
    assert GlobalCommands.INITIALIZE_HARDWARE.label == "Initialize Hardware"
    assert GlobalCommands.START_SYSTEM_ID_NOISE.label == "Start System Id Noise"


# endregion


# region VerboseMessageQueue
def test_verbose_queue_init(log_file_queue: mp.Queue, name_manager):
    """
    Verifies that the verbose queue initializes successfully.
    """
    verbose_queue = VerboseMessageQueue(
        log_file_queue,
        mp.Queue(),
        "VerboseQueue",
        name_manager,
    )

    assert isinstance(verbose_queue, VerboseMessageQueue)
    assert verbose_queue.base_name == "VerboseQueue"
    assert verbose_queue.log_queue is log_file_queue
    assert verbose_queue.last_put_message is None
    assert verbose_queue.last_get_message is None
    assert verbose_queue.time_threshold == 1.0


def test_verbose_queue_name(verbose_fixture: VerboseMessageQueue):
    """
    Verifies that the log name includes the assigned environment name.
    """
    verbose_queue, _ = verbose_fixture

    assert verbose_queue.log_name == "VerboseQueue"

    verbose_queue.assign_environment("Environment 0")

    assert verbose_queue.log_name == "VerboseQueue | Environment 0"


def test_verbose_queue_name_without_manager(log_file_queue: VerboseMessageQueue):
    """
    Verifies that queues without a name manager use only the base name.
    """
    verbose_queue = VerboseMessageQueue(
        log_file_queue,
        thqueue.Queue(),
        "VerboseQueue",
    )

    assert verbose_queue.log_name == "VerboseQueue"


def test_verbose_message_id(verbose_fixture: VerboseMessageQueue, random_seed=42):
    """
    Verifies deterministic message ID generation when the random seed is fixed.
    """
    verbose_queue, _ = verbose_fixture

    random.seed(random_seed)
    message_id = verbose_queue.generate_message_id()

    random.seed(random_seed)
    expected_id = "".join(
        random.choice(string.ascii_letters + string.digits) for _ in range(6)
    )

    assert message_id == expected_id


@mock.patch("rattlesnake.utilities.VerboseMessageQueue.generate_message_id")
def test_verbose_message_queue_put(
    mock_generate_message_id, verbose_fixture: VerboseMessageQueue
):
    """
    Verifies that put stores the generated message ID and message payload on
    the underlying queue.
    """
    verbose_queue, _ = verbose_fixture
    mock_base_queue = mock.MagicMock()
    verbose_queue.base_queue = mock_base_queue

    mock_generate_message_id.return_value = "message-id"
    task_name = "Test verbose queue"
    message_data_tuple = (GlobalCommands.QUIT, "Information")

    verbose_queue.put(task_name, message_data_tuple)

    mock_generate_message_id.assert_called_once_with(8)
    mock_base_queue.put.assert_called_once_with(("message-id", message_data_tuple))


@mock.patch("rattlesnake.utilities.time.time")
@mock.patch("rattlesnake.utilities.VerboseMessageQueue.generate_message_id")
def test_verbose_message_queue_put_suppresses_repeated_logs(
    mock_generate_message_id,
    mock_time,
    verbose_fixture: VerboseMessageQueue,
):
    """
    Verifies that repeated messages within the time threshold use an empty
    message ID and are not logged again.
    """
    verbose_queue, _ = verbose_fixture
    mock_base_queue = mock.MagicMock()
    verbose_queue.base_queue = mock_base_queue

    mock_time.side_effect = [0.0, 0.5]
    mock_generate_message_id.return_value = "message-id"

    message_data_tuple = (GlobalCommands.QUIT, None)

    verbose_queue.put("Task", message_data_tuple)
    verbose_queue.put("Task", message_data_tuple)

    assert mock_base_queue.put.call_args_list == [
        mock.call(("message-id", message_data_tuple)),
        mock.call(("", message_data_tuple)),
    ]


def test_verbose_message_queue_get(verbose_fixture: VerboseMessageQueue):
    """
    Verifies that get returns the message payload from the underlying queue.
    """
    verbose_queue, _ = verbose_fixture
    mock_base_queue = mock.MagicMock()
    verbose_queue.base_queue = mock_base_queue

    message_data_tuple = (GlobalCommands.QUIT, "Information")
    mock_base_queue.get.return_value = ("message-id", message_data_tuple)

    data = verbose_queue.get("Test verbose queue")

    assert data == message_data_tuple


def test_verbose_message_queue_get_empty_id_does_not_log(
    verbose_fixture: VerboseMessageQueue,
):
    """
    Verifies that messages with an empty ID are returned without a get log.
    """
    verbose_queue, _ = verbose_fixture
    mock_log_queue = mock.MagicMock()
    mock_base_queue = mock.MagicMock()
    verbose_queue.log_queue = mock_log_queue
    verbose_queue.base_queue = mock_base_queue

    message_data_tuple = (GlobalCommands.QUIT, "Information")
    mock_base_queue.get.return_value = ("", message_data_tuple)

    assert verbose_queue.get("Task") == message_data_tuple
    mock_log_queue.put.assert_not_called()


@mock.patch("rattlesnake.utilities.VerboseMessageQueue.generate_message_id")
@mock.patch("rattlesnake.utilities.datetime")
def test_verbose_queue_log(
    mock_datetime,
    mock_generate_message_id,
    log_file_queue: mp.Queue,
    verbose_fixture: VerboseMessageQueue,
):
    """
    Exercises put and get logging through a helper process or thread.
    """
    verbose_queue, use_thread = verbose_fixture
    new_process = threading.Thread if use_thread else mp.Process

    verbose_array = mp.Array("i", 1)
    verbose_value = 10
    log_string = mp.Array("c", 500)
    log_string.value = b""

    task_name = "Test verbose queue"
    message_data_tuple = (GlobalCommands.QUIT, verbose_value)

    mock_datetime.now = fake_time
    mock_generate_message_id.return_value = "1"

    verbose_queue.put(task_name, message_data_tuple)

    verbose_process = new_process(
        target=clear_verbose_queue,
        args=(verbose_queue, "Get Queue", verbose_array),
    )
    verbose_process.start()
    verbose_process.join()

    log_file_process = new_process(
        target=clear_log_queue,
        args=(log_file_queue, log_string),
    )
    log_file_process.start()
    log_file_process.join()

    assert verbose_array[0] == verbose_value
    assert b"put QUIT" in log_string.value
    assert b"got QUIT" in log_string.value


@mock.patch("rattlesnake.utilities.VerboseMessageQueue.generate_message_id")
@mock.patch("rattlesnake.utilities.datetime")
def test_verbose_message_queue_flush(
    mock_datetime,
    mock_generate_message_id,
    log_file_queue: mp.Queue,
    verbose_fixture: VerboseMessageQueue,
):
    """
    Verifies that flush returns queued message payloads.
    """
    verbose_queue, use_thread = verbose_fixture
    new_process = threading.Thread if use_thread else mp.Process

    log_string = mp.Array("c", 500)
    log_string.value = b""

    task_name = "Test verbose flush"
    message_data_tuple = (GlobalCommands.QUIT, "This should have data")

    mock_datetime.now = fake_time
    mock_generate_message_id.return_value = "1"

    verbose_queue.put(task_name, message_data_tuple)
    data = verbose_queue.flush(task_name)

    log_file_process = new_process(
        target=clear_log_queue,
        args=(log_file_queue, log_string),
    )
    log_file_process.start()
    log_file_process.join()

    assert data == [message_data_tuple]
    assert b"flushed VerboseQueue" in log_string.value


def test_verbose_queue_empty_close_join_thread(verbose_fixture: VerboseMessageQueue):
    """
    Verifies that empty, close, and join_thread are available and do not raise.
    """
    verbose_queue, _ = verbose_fixture

    verbose_queue.empty()
    verbose_queue.close()
    verbose_queue.join_thread()

    assert True


# endregion


# region QueueContainer, EventContainer, flush_queue
def test_queue_container_init(queue_container: QueueContainer):
    """
    Verifies that ``QueueContainer`` stores supplied queues.
    """
    assert isinstance(queue_container, QueueContainer)
    assert isinstance(queue_container.controller_command_queue, VerboseMessageQueue)
    assert isinstance(queue_container.acquisition_command_queue, VerboseMessageQueue)
    assert isinstance(queue_container.output_command_queue, VerboseMessageQueue)
    assert isinstance(queue_container.streaming_command_queue, VerboseMessageQueue)
    assert "Environment 0" in queue_container.environment_command_queues
    assert "Environment 0" in queue_container.environment_data_in_queues
    assert "Environment 0" in queue_container.environment_data_out_queues


def test_event_container_init(event_container: EventContainer):
    """
    Verifies that ``EventContainer`` stores supplied events.
    """
    assert isinstance(event_container, EventContainer)
    assert isinstance(event_container.controller_ready_event, threading.Event)
    assert isinstance(event_container.acquisition_ready_event, threading.Event)
    assert isinstance(event_container.output_ready_event, threading.Event)
    assert isinstance(event_container.streaming_ready_event, threading.Event)
    assert "Environment 0" in event_container.environment_ready_events
    assert "Environment 0" in event_container.environment_active_events
    assert "Environment 0" in event_container.environment_sysid_active_events
    assert "Environment 0" in event_container.environment_sysid_stored_events


@pytest.mark.parametrize("use_thread", [True, False])
def test_flush_queue_standard_queue(use_thread):
    """
    Verifies that ``flush_queue`` removes all items from a standard queue.
    """
    if use_thread:
        new_queue = thqueue.Queue
    else:
        new_queue = mp.Queue
    q = new_queue()
    q.put("a")
    q.put("b")

    assert flush_queue(q) == ["a", "b"]
    assert q.empty()


def test_flush_queue_verbose_queue(log_file_queue, name_manager):
    """
    Verifies that ``flush_queue`` works with ``VerboseMessageQueue``.
    """
    verbose_queue = VerboseMessageQueue(
        log_file_queue,
        thqueue.Queue(),
        "VerboseQueue",
        name_manager,
    )
    verbose_queue.put("Task", (GlobalCommands.QUIT, None))

    assert flush_queue(verbose_queue) == [(GlobalCommands.QUIT, None)]


@pytest.mark.parametrize("use_thread", [True, False])
def test_flush_queue_with_timeout_on_empty_queue(use_thread):
    """
    Verifies that flushing an empty queue with a timeout returns an empty list.
    """
    if use_thread:
        new_queue = thqueue.Queue
    else:
        new_queue = mp.Queue
    q = new_queue()

    assert flush_queue(q, timeout=0.01) == []


# endregion


# region LAN-XI Network Utilities
def test_autofill_single_ip_address_valid_ip_returns_unchanged():
    """
    Verifies that valid IP records are returned unchanged.
    """
    ip_address = IPAddress(host_name="host", ipv4_address="1.2.3.4", valid_ip=True)

    assert autofill_single_ip_address(ip_address) is ip_address


def test_autofill_single_ip_address_host_name_lookup():
    """
    Verifies that host-name-only records attempt IP lookup.
    """
    ip_address = IPAddress(host_name="host")
    ip_address.get_ip_from_host_name = mock.MagicMock()

    result = autofill_single_ip_address(ip_address)

    assert result is ip_address
    ip_address.get_ip_from_host_name.assert_called()


def test_autofill_single_ip_address_ipv4_lookup():
    """
    Verifies that IPv4 records attempt host-name lookup and then IP lookup.
    """
    ip_address = IPAddress(ipv4_address="1.2.3.4")
    ip_address.get_host_name_from_ip = mock.MagicMock()
    ip_address.get_ip_from_host_name = mock.MagicMock()

    result = autofill_single_ip_address(ip_address)

    assert result is ip_address
    ip_address.get_host_name_from_ip.assert_called_once()
    ip_address.get_ip_from_host_name.assert_called()


@mock.patch("rattlesnake.utilities.find_lanxi_devices")
@mock.patch("rattlesnake.utilities.time.perf_counter")
def test_search_for_lanxi_devices(mock_perf_counter, mock_find_lanxi_devices):
    """
    Verifies that LAN-XI search returns unique devices by IPv4 address.
    """
    device_1 = IPAddress(ipv4_address="169.254.1.2")
    device_2 = IPAddress(ipv4_address="169.254.1.2")
    mock_find_lanxi_devices.side_effect = [[device_1], [device_2]]
    mock_perf_counter.side_effect = [0.0, 0.1, 0.2, 1.1]

    devices = search_for_lanxi_devices(timeout=1.0)

    assert devices == [device_2]


@mock.patch("rattlesnake.utilities.IPAddress.get_ip_from_host_name")
@mock.patch("rattlesnake.utilities.check_lanxi_candidate")
@mock.patch("rattlesnake.utilities.subprocess.run")
def test_find_lanxi_devices(
    mock_subprocess_run,
    mock_check_lanxi_candidate,
    mock_get_ip_from_host_name,
):
    """
    Verifies that ARP output is parsed and valid LAN-XI candidates are returned.
    """
    mock_subprocess_run.return_value.stdout = """
    ? (169.254.1.2) at aa:bb:cc:dd:ee:ff
    ? (169.254.999.2) at invalid
    """
    info = {"module": {"type": {"number": "3050"}, "serial": "123456"}}
    sync = {"syncmode": "standalone"}
    mock_check_lanxi_candidate.return_value = (
        "BK3050-123456",
        "169.254.1.2",
        info,
        sync,
        True,
    )

    devices = find_lanxi_devices()

    assert len(devices) == 1
    assert devices[0].host_name == "BK3050-123456"
    assert devices[0].ipv4_address == "169.254.1.2"
    assert devices[0].module_info == info
    assert devices[0].sync_type == sync
    mock_get_ip_from_host_name.assert_called_once()


@mock.patch("rattlesnake.utilities.requests.get")
def test_check_lanxi_candidate_valid(mock_requests_get):
    """
    Verifies that a valid LAN-XI candidate returns host metadata.
    """
    info_response = mock.MagicMock()
    info_response.json.return_value = {
        "module": {"type": {"number": "3050"}, "serial": "123456"}
    }
    info_response.raise_for_status.return_value = None

    sync_response = mock.MagicMock()
    sync_response.json.return_value = {"syncmode": "standalone"}

    mock_requests_get.side_effect = [info_response, sync_response]

    host_name, ipv4, info, sync, valid = check_lanxi_candidate("169.254.1.2")

    assert host_name == "BK3050-123456"
    assert ipv4 == "169.254.1.2"
    assert info == info_response.json.return_value
    assert sync == sync_response.json.return_value
    assert valid is True


@mock.patch("rattlesnake.utilities.requests.get")
def test_check_lanxi_candidate_invalid(mock_requests_get):
    """
    Verifies that failed LAN-XI candidate requests return invalid status.
    """
    mock_requests_get.side_effect = RuntimeError("network failure")

    host_name, ipv4, info, sync, valid = check_lanxi_candidate("169.254.1.2")

    assert host_name is None
    assert ipv4 == "169.254.1.2"
    assert info is None
    assert sync is None
    assert valid is False


def test_ip_address_init():
    """
    Verifies that ``IPAddress`` stores default and supplied attributes.
    """
    ip_address = IPAddress("host", "1.2.3.4", "[fe80::1]", True)

    assert ip_address.host_name == "host"
    assert ip_address.ipv4_address == "1.2.3.4"
    assert ip_address.ipv6_address == "[fe80::1]"
    assert ip_address.valid_ip is True
    assert ip_address.module_info is None
    assert ip_address.sync_type is None
    assert ip_address.validation_timeout == 5


@mock.patch("rattlesnake.utilities.socket.getaddrinfo")
def test_ip_address_get_ip_from_host_name(mock_getaddrinfo):
    """
    Verifies that IP addresses are resolved from a host name.
    """
    mock_getaddrinfo.return_value = [
        (socket.AF_INET, None, None, None, ("1.2.3.4", 0)),
        (socket.AF_INET6, None, None, None, ("fe80::1", 0, 0, 2)),
    ]

    ip_address = IPAddress(host_name="host")

    ip_address.get_ip_from_host_name()

    assert ip_address.ipv4_address == "1.2.3.4"
    assert ip_address.ipv6_address == "[fe80::1%2]"


@mock.patch("rattlesnake.utilities.socket.getaddrinfo")
def test_ip_address_get_ip_from_host_name_failure(mock_getaddrinfo):
    """
    Verifies that failed host lookup marks the record invalid.
    """
    mock_getaddrinfo.side_effect = RuntimeError("lookup failure")
    ip_address = IPAddress(host_name="host", valid_ip=True)

    ip_address.get_ip_from_host_name()

    assert ip_address.valid_ip is False


def test_ip_address_get_ip_from_host_name_no_host():
    """
    Verifies that missing host name marks the record invalid.
    """
    ip_address = IPAddress(valid_ip=True)

    ip_address.get_ip_from_host_name()

    assert ip_address.valid_ip is False


@mock.patch("rattlesnake.utilities.requests.get")
def test_ip_address_get_host_name_from_ip(mock_requests_get):
    """
    Verifies that LAN-XI module information can be resolved from an IP address.
    """
    info_response = mock.MagicMock()
    info_response.json.return_value = {
        "module": {"type": {"number": "3050"}, "serial": "123456"}
    }
    sync_response = mock.MagicMock()
    sync_response.json.return_value = {"syncmode": "standalone"}
    mock_requests_get.side_effect = [info_response, sync_response]

    ip_address = IPAddress(ipv4_address="1.2.3.4")
    ip_address.get_host_name_from_ip()

    assert ip_address.host_name == "BK3050-123456"
    assert ip_address.module_info == info_response.json.return_value
    assert ip_address.sync_type == sync_response.json.return_value
    assert ip_address.valid_ip is True


@mock.patch("rattlesnake.utilities.requests.get")
def test_ip_address_get_host_name_from_ip_failure(mock_requests_get):
    """
    Verifies that failed LAN-XI lookup marks the record invalid.
    """
    mock_requests_get.side_effect = RuntimeError("network failure")
    ip_address = IPAddress(ipv4_address="1.2.3.4", valid_ip=True)

    ip_address.get_host_name_from_ip()

    assert ip_address.valid_ip is False


@mock.patch("rattlesnake.utilities.requests.get")
def test_ip_address_validate_ipv6(mock_requests_get):
    """
    Verifies validation using an IPv6 address.
    """
    info_response = mock.MagicMock()
    info_response.json.return_value = {"module": "info"}
    sync_response = mock.MagicMock()
    sync_response.json.return_value = {"sync": "type"}
    mock_requests_get.side_effect = [info_response, sync_response]

    ip_address = IPAddress(ipv6_address="[fe80::1]")

    ip_address.validate()

    assert ip_address.module_info == {"module": "info"}
    assert ip_address.sync_type == {"sync": "type"}
    assert ip_address.valid_ip is True


# endregion


# region Loading Utilities
def test_load_time_history_npy(tmp_path):
    """
    Verifies loading a NumPy ``.npy`` time-history file and truncating odd
    sample counts.
    """
    path = tmp_path / "signal.npy"
    np.save(path, np.arange(5.0))

    signal = load_time_history(path, sample_rate=1000)

    np.testing.assert_array_equal(signal, np.arange(4.0))


def test_load_time_history_npz_with_signal_only(tmp_path):
    """
    Verifies loading a NumPy ``.npz`` time-history file without time data.
    """
    path = tmp_path / "signal.npz"
    np.savez(path, signal=np.array([[1.0, 2.0, 3.0, 4.0]]))

    signal = load_time_history(path, sample_rate=1000)

    np.testing.assert_array_equal(signal, np.array([[1.0, 2.0, 3.0, 4.0]]))


def test_load_time_history_npz_with_time_interpolation(tmp_path):
    """
    Verifies that ``.npz`` files with time data are interpolated to the
    requested sample rate.
    """
    path = tmp_path / "signal.npz"
    np.savez(path, signal=np.array([0.0, 1.0, 2.0]), t=np.array([0.0, 0.5, 1.0]))

    signal = load_time_history(path, sample_rate=2)

    np.testing.assert_array_equal(signal, np.array([0.0, 1.0]))


def test_load_time_history_mat(tmp_path):
    """
    Verifies loading a MATLAB ``.mat`` time-history file.
    """
    path = tmp_path / "signal.mat"
    savemat(path, {"signal": np.array([[1.0, 2.0, 3.0, 4.0]])})

    signal = load_time_history(path, sample_rate=1000)

    np.testing.assert_array_equal(signal, np.array([[1.0, 2.0, 3.0, 4.0]]))


def test_load_time_history_invalid_extension(tmp_path):
    """
    Verifies that unknown time-history file extensions raise ``ValueError``.
    """
    path = tmp_path / "signal.txt"
    path.write_text("1,2,3", encoding="utf-8")

    with pytest.raises(ValueError):
        load_time_history(path, sample_rate=1000)


def test_load_and_save_csv_matrix(tmp_path):
    """
    Verifies that CSV matrices can be saved and loaded.
    """
    path = tmp_path / "matrix.csv"
    data = [["1", "2", "3"], ["4", "5", "6"]]

    save_csv_matrix(data, path)
    loaded = load_csv_matrix(path)

    assert loaded == data


def test_load_python_module(tmp_path):
    """
    Verifies that a Python module can be loaded from a file path.
    """
    path = tmp_path / "example_module.py"
    path.write_text("VALUE = 123\n", encoding="utf-8")

    module = load_python_module(path)

    assert module.VALUE == 123


def test_read_transformation_matrix_from_worksheet():
    """
    Verifies reading a numeric transformation matrix from a worksheet.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 1, 1)
    worksheet.cell(1, 2, 2)
    worksheet.cell(2, 1, 3)
    worksheet.cell(2, 2, 4)

    matrix = read_transformation_matrix_from_worksheet(
        worksheet,
        start_row=1,
        num_rows=2,
        start_col=1,
    )

    np.testing.assert_array_equal(matrix, np.array([[1.0, 2.0], [3.0, 4.0]]))


@pytest.mark.parametrize("first_cell", [None, "None", " none "])
def test_read_transformation_matrix_from_worksheet_none(first_cell):
    """
    Verifies that blank or ``None`` matrix entries return ``None``.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 1, first_cell)

    matrix = read_transformation_matrix_from_worksheet(
        worksheet,
        start_row=1,
        num_rows=1,
        start_col=1,
    )

    assert matrix is None


def test_read_transformation_matrix_from_worksheet_stops_at_comment():
    """
    Verifies matrix row reading stops at comment cells.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 1, 1)
    worksheet.cell(1, 2, "# comment")

    matrix = read_transformation_matrix_from_worksheet(
        worksheet,
        start_row=1,
        num_rows=1,
        start_col=1,
    )

    np.testing.assert_array_equal(matrix, np.array([[1.0]]))


# endregion


# region Math Operations
def test_coherence_full_matrix():
    """
    Verifies coherence computation for a CPSD matrix.
    """
    cpsd = np.array(
        [
            [[4.0 + 0j, 2.0 + 0j], [2.0 + 0j, 4.0 + 0j]],
            [[9.0 + 0j, 3.0 + 0j], [3.0 + 0j, 9.0 + 0j]],
        ]
    )

    coh = coherence(cpsd)

    assert coh.shape == cpsd.shape
    np.testing.assert_allclose(coh[:, 0, 0], 1.0)
    np.testing.assert_allclose(coh[:, 1, 1], 1.0)
    np.testing.assert_allclose(coh[0, 0, 1], 0.25)
    np.testing.assert_allclose(coh[1, 0, 1], 1.0 / 9.0)


def test_coherence_row_column():
    """
    Verifies coherence computation for a single row-column pair.
    """
    cpsd = np.array([[[4.0 + 0j, 2.0 + 0j], [2.0 + 0j, 4.0 + 0j]]])

    coh = coherence(cpsd, row_column=(0, 1))

    np.testing.assert_allclose(coh, np.array([0.25]))


def test_cpsd_to_time_history_shape():
    """
    Verifies that CPSD synthesis returns a channel-by-samples time history.
    """
    np.random.seed(1)
    cpsd = np.zeros((5, 2, 2), dtype=complex)
    cpsd[:, 0, 0] = 1.0
    cpsd[:, 1, 1] = 1.0

    output = cpsd_to_time_history(cpsd, sample_rate=100.0, df=1.0)

    assert output.shape == (2, 8)


def test_reduce_array_by_coordinate():
    """
    Verifies coordinate-based array reduction for a simple two-dimensional
    array.
    """
    dtype = np.dtype([("node", "U1"), ("direction", "i4")])
    coord_a = np.array(("A", 1), dtype=dtype)[()]
    coord_b = np.array(("B", -2), dtype=dtype)[()]

    array = np.array([[1.0, 2.0], [3.0, 4.0]])
    coordinate = np.array([[coord_a, coord_a], [coord_b, coord_b]], dtype=dtype)
    control_coordinate = np.array([coord_a, coord_b], dtype=dtype)

    reduced = reduce_array_by_coordinate(array, coordinate, control_coordinate)

    np.testing.assert_array_equal(reduced, array)


def test_reduce_array_by_coordinate_missing_coordinate():
    """
    Verifies that requesting a missing coordinate raises ``ValueError``.
    """
    dtype = np.dtype([("node", "U1"), ("direction", "i4")])
    coord_a = np.array(("A", 1), dtype=dtype)[()]
    coord_b = np.array(("B", 2), dtype=dtype)[()]

    array = np.array([[1.0]])
    coordinate = np.array([[coord_a, coord_a]], dtype=dtype)
    control_coordinate = np.array([coord_b], dtype=dtype)

    with pytest.raises(ValueError):
        reduce_array_by_coordinate(array, coordinate, control_coordinate)


def test_db_scale_power_conversions():
    """
    Verifies decibel, power, and scale conversions.
    """
    assert db2scale(20) == pytest.approx(10.0)
    assert scale2db(10) == pytest.approx(20.0)
    assert power2db(100) == pytest.approx(20.0)


def test_rms_time():
    """
    Verifies RMS over a time signal.
    """
    signal = np.array([3.0, 4.0])

    assert rms_time(signal) == pytest.approx(np.sqrt(12.5))


def test_rms_time_axis_keepdims():
    """
    Verifies RMS over an axis with kept dimensions.
    """
    signal = np.array([[3.0, 4.0], [0.0, 0.0]])

    result = rms_time(signal, axis=-1, keepdims=True)

    np.testing.assert_allclose(result, np.array([[np.sqrt(12.5)], [0.0]]))


def test_rms_csd():
    """
    Verifies RMS computation from a CSD matrix.
    """
    csd = np.zeros((2, 2, 2), dtype=complex)
    csd[:, 0, 0] = [1.0, 3.0]
    csd[:, 1, 1] = [4.0, 5.0]

    result = rms_csd(csd, df=0.5)

    np.testing.assert_allclose(result, np.sqrt(np.array([2.0, 4.5])))


def test_trac_self():
    """
    Verifies that TRAC of a signal against itself is one.
    """
    signal = np.array([[1.0, 2.0, 3.0]])

    result = trac(signal)

    np.testing.assert_allclose(result, np.array([1.0]))


def test_trac_pair():
    """
    Verifies TRAC between two signal arrays.
    """
    signal_1 = np.array([[1.0, 0.0]])
    signal_2 = np.array([[0.0, 1.0]])

    result = trac(signal_1, signal_2)

    np.testing.assert_allclose(result, np.array([0.0]))


def test_moving_sum():
    """
    Verifies moving sum along the final axis.
    """
    signal = np.array([1.0, 2.0, 3.0, 4.0])

    result = moving_sum(signal, 2)

    np.testing.assert_array_equal(result, np.array([3.0, 5.0, 7.0]))


# endregion


# region Correlation and Alignment
def test_correlation_metrics():
    """
    Verifies correlation metric functions return arrays of expected length.
    """
    signal = np.array([[0.0, 1.0, 2.0, 3.0, 0.0]])
    specification = np.array([[1.0, 2.0, 3.0]])

    metrics = [
        corr_norm_signal_spec,
        corr_norm_spec2,
        norm_ratio,
        correlation_norm_spec_ratio,
        correlation_norm_signal_spec_ratio,
    ]

    for metric in metrics:
        result = metric(signal, specification)
        assert result.shape == (3,)


def test_align_signals_without_subsample():
    """
    Verifies integer-delay signal alignment.
    """
    measurement_buffer = np.array([[0.0, 0.0, 1.0, 2.0, 3.0, 0.0]])
    specification = np.array([[1.0, 2.0, 3.0]])

    aligned, delay, phase_slope, found_correlation = align_signals(
        measurement_buffer,
        specification,
        correlation_threshold=0.5,
        perform_subsample=False,
    )

    np.testing.assert_array_equal(aligned, specification)
    assert delay == 2
    assert phase_slope is None
    assert found_correlation == pytest.approx(1.0)


def test_align_signals_below_threshold():
    """
    Verifies that alignment returns ``None`` values below threshold.
    """
    measurement_buffer = np.zeros((1, 5))
    specification = np.ones((1, 3))

    aligned, delay, phase_slope, found_correlation = align_signals(
        measurement_buffer,
        specification,
        correlation_threshold=0.5,
        perform_subsample=False,
    )

    assert aligned is None
    assert delay is None
    assert phase_slope is None
    assert found_correlation is None


def test_align_signals_with_custom_metric():
    """
    Verifies that a custom correlation metric can be supplied.
    """
    measurement_buffer = np.array([[0.0, 1.0, 2.0, 3.0]])
    specification = np.array([[1.0, 2.0]])

    def metric(signal, spec):
        return np.array([0.1, 1.0, 0.2])

    aligned, delay, _, found_correlation = align_signals(
        measurement_buffer,
        specification,
        correlation_threshold=0.5,
        perform_subsample=False,
        correlation_metric=metric,
    )

    np.testing.assert_array_equal(aligned, np.array([[1.0, 2.0]]))
    assert delay == 1
    assert found_correlation == pytest.approx(1.0)


def test_shift_signal():
    """
    Verifies sample-based signal shifting with zero phase slope.
    """
    signal = np.array([[0.0, 1.0, 2.0, 3.0]])

    shifted = shift_signal(signal, samples_to_keep=2, sample_delay=1, phase_slope=0.0)

    np.testing.assert_allclose(shifted, np.array([[1.0, 2.0]]))


def test_wrap():
    """
    Verifies angle wrapping.
    """
    data = np.array([-3 * np.pi, 0.0, 3 * np.pi])

    wrapped = wrap(data)

    np.testing.assert_allclose(wrapped, np.array([-np.pi, 0.0, -np.pi]))


# endregion


# region OverlapBuffer
def test_overlap_buffer_init():
    """
    Verifies overlap buffer initialization.
    """
    buffer = OverlapBuffer((2, 5), buffer_axis=-1, starting_value=1.0)

    assert buffer.buffer_position == 0
    assert buffer.buffer_axis == 1
    assert buffer.shape == (2, 5)
    np.testing.assert_array_equal(buffer.buffer_data, np.ones((2, 5)))


def test_overlap_buffer_add_data_noshift():
    """
    Verifies adding data without changing buffer position.
    """
    buffer = OverlapBuffer((1, 5))
    buffer.add_data_noshift(np.array([[1.0, 2.0]]))

    np.testing.assert_array_equal(
        buffer.buffer_data, np.array([[0.0, 0.0, 0.0, 1.0, 2.0]])
    )
    assert buffer.buffer_position == 0


def test_overlap_buffer_add_data():
    """
    Verifies adding data and updating buffer position.
    """
    buffer = OverlapBuffer((1, 5))
    buffer.add_data(np.array([[1.0, 2.0, 3.0]]))

    np.testing.assert_array_equal(
        buffer.buffer_data, np.array([[0.0, 0.0, 1.0, 2.0, 3.0]])
    )
    assert buffer.buffer_position == 3


def test_overlap_buffer_add_data_clamps_to_buffer_size():
    """
    Verifies that adding more data than the buffer size retains the newest
    samples and clamps the buffer position.
    """
    buffer = OverlapBuffer((1, 3))
    buffer.add_data(np.array([[1.0, 2.0, 3.0, 4.0, 5.0]]))

    np.testing.assert_array_equal(buffer.buffer_data, np.array([[3.0, 4.0, 5.0]]))
    assert buffer.buffer_position == 3


def test_overlap_buffer_get_data_noshift():
    """
    Verifies reading data without shifting buffer position.
    """
    buffer = OverlapBuffer((1, 5))
    buffer.add_data(np.array([[1.0, 2.0, 3.0]]))

    data = buffer.get_data_noshift(2)

    np.testing.assert_array_equal(data, np.array([[1.0, 2.0]]))
    assert buffer.buffer_position == 3


def test_overlap_buffer_get_data_noshift_too_many_samples():
    """
    Verifies that reading more samples than available raises ``ValueError``.
    """
    buffer = OverlapBuffer((1, 5))
    buffer.add_data(np.array([[1.0, 2.0]]))

    with pytest.raises(ValueError):
        buffer.get_data_noshift(3)


def test_overlap_buffer_get_data_default_shift():
    """
    Verifies reading data with default buffer position shift.
    """
    buffer = OverlapBuffer((1, 5))
    buffer.add_data(np.array([[1.0, 2.0, 3.0]]))

    data = buffer.get_data(2)

    np.testing.assert_array_equal(data, np.array([[1.0, 2.0]]))
    assert buffer.buffer_position == 1


def test_overlap_buffer_get_data_custom_shift():
    """
    Verifies reading data with an explicit buffer position shift.
    """
    buffer = OverlapBuffer((1, 5))
    buffer.add_data(np.array([[1.0, 2.0, 3.0]]))

    data = buffer.get_data(2, buffer_shift=-1)

    np.testing.assert_array_equal(data, np.array([[1.0, 2.0]]))
    assert buffer.buffer_position == 2


def test_overlap_buffer_shift_buffer_position_clamps():
    """
    Verifies that buffer position shifting is clamped to valid bounds.
    """
    buffer = OverlapBuffer((1, 5))

    buffer.shift_buffer_position(10)
    assert buffer.buffer_position == 5

    buffer.shift_buffer_position(-10)
    assert buffer.buffer_position == 0


def test_overlap_buffer_set_buffer_position_clamps():
    """
    Verifies that setting buffer position is clamped to valid bounds.
    """
    buffer = OverlapBuffer((1, 5))

    buffer.set_buffer_position(10)
    assert buffer.buffer_position == 5

    buffer.set_buffer_position(-10)
    assert buffer.buffer_position == 0


def test_overlap_buffer_getitem():
    """
    Verifies indexing into the underlying buffer.
    """
    buffer = OverlapBuffer((1, 3))
    buffer.add_data(np.array([[1.0, 2.0, 3.0]]))

    np.testing.assert_array_equal(buffer[0], np.array([1.0, 2.0, 3.0]))


def test_overlap_buffer_non_last_axis():
    """
    Verifies buffering along an axis other than the final axis.
    """
    buffer = OverlapBuffer((5, 1), buffer_axis=0)
    buffer.add_data(np.array([[1.0], [2.0]]))

    np.testing.assert_array_equal(
        buffer.buffer_data,
        np.array([[0.0], [0.0], [0.0], [1.0], [2.0]]),
    )
    assert buffer.buffer_position == 2


# endregion
