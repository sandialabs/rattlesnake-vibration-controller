# -*- coding: utf-8 -*-
"""
Abstract hardware definition that can be used to implement new hardware devices

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from abc import ABC, abstractmethod
from typing import List, Tuple

import numpy as np
import netCDF4 as nc4
import openpyxl

from rattlesnake.utilities import RattlesnakeError
from rattlesnake.hardware.hardware_utilities import Channel, HardwareType
from rattlesnake.user_interface.ui_utilities import HardwareAssistModules


# region Metadata
class HardwareMetadata:
    """
    Abstract class that contains values to fully define how the hardware is setup.

    This class contains attributes required to run acquisition, output, and streaming
    processes. The class should also contain extra attributes required for the
    HardwareAcquisition and HardwareOutput class specific to that HardwareType.
    """

    def __init__(
        self,
        hardware_type: HardwareType,
        channel_list: List[Channel],
        sample_rate: int,
        time_per_read: float,
        time_per_write: float,
        *,
        output_oversample: int = 1,
    ):
        self.hardware_type = hardware_type
        self.channel_list = channel_list
        self.sample_rate = sample_rate
        self.time_per_read = time_per_read
        self.time_per_write = time_per_write
        # Used for virtual hardware but still required for normal hardware
        self.output_oversample = output_oversample

    @property
    def samples_per_read(self):
        """Property returning the number of samples per read frame."""
        return round(self.sample_rate * self.time_per_read)

    @property
    def samples_per_write(self):
        """Property returning the number of samples per write frame."""
        return round(self.sample_rate * self.time_per_write * self.output_oversample)

    @property
    def nyquist_frequency(self):
        """Property returning the Nyquist frequency of the data acquisition."""
        return self.sample_rate / 2

    @property
    def output_sample_rate(self):
        """Property returning the output sample rate."""
        return self.sample_rate * self.output_oversample

    # endregion

    # region Validation
    @abstractmethod
    def validate(self):
        """
        Method to check if the metadata object fully defines the hardware and is valid
        for that machine.

        If possible should check which devices are connected to the machine at a given
        time and make sure that they are valid inputs to the initialize_hardware function
        of the HardwareAcquisition and HardwareOutput classes.

        Throw detailed errors while validating, these errors will show up in log files for
        debugging and will not stop the main process from running.
        """
        if len(self.channel_list) != len(set(self.channel_list)):
            raise RattlesnakeError("Duplicate channels found in channel_list")

    @abstractmethod
    def valid_channel_dict(self, channel: Channel):
        valid_dict = {}
        for attr in Channel().channel_attr_list:
            valid_dict[attr] = []
        return valid_dict

    @property
    @abstractmethod
    def assist_mode_modules(self):
        assist_modules = {}
        for attr in Channel().channel_attr_list:
            assist_modules[attr] = HardwareAssistModules.NONE
        return assist_modules

    # endregion

    # region Loading
    @classmethod
    def load_channel_table_from_netcdf(
        cls, netcdf_dataset: nc4.Dataset
    ) -> List[Channel]:
        channel_table = netcdf_dataset["channels"]

        channel_list = []
        num_channels = netcdf_dataset.dimensions["response_channels"].size
        channel_attr_list = Channel().channel_attr_list
        for row_idx in range(num_channels):
            channel = Channel()
            for attr in channel_attr_list:
                value = channel_table[attr][row_idx]
                value = None if isinstance(value, str) and not value.strip() else value
                setattr(channel, attr, value)

            if not channel.is_empty:  # optional safety check
                channel_list.append(channel)

        return channel_list

    @classmethod
    def save_channel_table_to_workbook(
        self, channel_list: List[Channel], workbook: openpyxl.workbook.workbook.Workbook
    ):
        worksheet = workbook.active
        worksheet.title = "Channel Table"
        worksheet.cell(row=1, column=2, value="Test Article Definition")
        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
        worksheet.cell(row=1, column=5, value="Instrument Definition")
        worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=11)
        worksheet.cell(row=1, column=12, value="Channel Definition")
        worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=19)
        worksheet.cell(row=1, column=20, value="Output Feedback")
        worksheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=21)
        worksheet.cell(row=1, column=22, value="Limits")
        worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=23)
        for col_idx, val in enumerate(
            [
                "Channel Index",
                "Node Number",
                "Node Direction",
                "Comment",
                "Serial Number",
                "Triax DoF",
                "Sensitivity (mV/EU)",
                "Engineering Unit",
                "Make",
                "Model",
                "Calibration Exp Date",
                "Physical Device",
                "Physical Channel",
                "Type",
                "Minimum Value (V)",
                "Maximum Value (V)",
                "Coupling",
                "Current Excitation Source",
                "Current Excitation Value",
                "Physical Device",
                "Physical Channel",
                "Warning Level (EU)",
                "Abort Level (EU)",
            ]
        ):
            worksheet.cell(row=2, column=1 + col_idx, value=val)
        # Fill out values
        channel_attr_list = Channel().channel_attr_list
        for row, channel in enumerate(channel_list):
            row_idx = row + 3
            worksheet.cell(row=row_idx, column=1, value=row)
            for col, attr in enumerate(channel_attr_list):
                col_idx = col + 2
                val = getattr(channel, attr)
                val = str(val) if val is not None else ""
                worksheet.cell(row=row_idx, column=col_idx, value=val)

    @classmethod
    def load_channel_table_from_workbook(
        cls, workbook: openpyxl.workbook.workbook.Workbook
    ) -> List[Channel]:
        sheets = workbook.sheetnames

        if len(sheets) > 1:
            sheets = [sheet for sheet in sheets if "channel" in sheet.lower()]
        if len(sheets) > 1:
            raise RattlesnakeError(
                "Multiple channel table sheets located in Excel Spreadsheet"
            )
        if len(sheets) == 0:
            raise RattlesnakeError(
                "Excel Spreadsheet does not contain a channel table sheet"
            )

        worksheet = workbook[sheets[0]]

        channel_list = []
        channel_attr_list = Channel().channel_attr_list
        for row in worksheet.iter_rows(min_row=3, min_col=2, max_col=23):
            channel = Channel()
            for col, cell in enumerate(row):
                value = cell.value
                value = None if isinstance(value, str) and not value.strip() else value
                setattr(channel, channel_attr_list[col], cell.value)
            if channel.is_empty:
                break
            channel_list.append(channel)

        return channel_list

    @classmethod
    def save_blank_hardware_to_workbook(
        cls, workbook: openpyxl.workbook.workbook.Workbook
    ):
        if "Hardware" in workbook.sheetnames:
            hardware_worksheet = workbook["Hardware"]
        else:
            hardware_worksheet = workbook.create_sheet("Hardware")
        hardware_worksheet.cell(1, 1, "Hardware Type")
        hardware_worksheet.cell(1, 2, "# Enter hardware index here")
        hardware_worksheet.cell(
            1,
            3,
            "Hardware Indices: 0 - NI DAQmx; 1 - LAN XI; 2 - Data Physics Quattro; "
            "3 - Data Physics 900 Series; 4 - Exodus Modal Solution; 5 - State Space Integration; "
            "6 - SDynPy System Integration",
        )
        hardware_worksheet.cell(2, 1, "Hardware File")
        hardware_worksheet.cell(
            2,
            3,
            "# Path to Hardware File (Depending on Hardware Device: 0 - Not Used; 1 - Not Used; "
            "2 - Path to DpQuattro.dll library file; 3 - Not Used; 4 - Path to Exodus Eigensolution; "
            "5 - Path to State Space File; 6 - Path to SDynPy system file)",
        )
        hardware_worksheet.cell(3, 1, "Sample Rate")
        hardware_worksheet.cell(3, 3, "# Sample Rate of Data Acquisition System")
        hardware_worksheet.cell(4, 1, "Time Per Read")
        hardware_worksheet.cell(
            4, 3, "# Number of seconds per Read from the Data Acquisition System"
        )
        hardware_worksheet.cell(5, 1, "Time Per Write")
        hardware_worksheet.cell(
            5, 3, "# Number of seconds per Write to the Data Acquisition System"
        )
        hardware_worksheet.cell(6, 1, "Maximum Acquisition Processes")
        hardware_worksheet.cell(
            6,
            3,
            "# Maximum Number of Acquisition Processes to start to pull data from hardware",
        )
        hardware_worksheet.cell(
            6,
            4,
            "Only Used by LAN-XI Hardware.  This row can be deleted if LAN-XI is not used",
        )
        hardware_worksheet.cell(7, 1, "Integration Oversampling")
        hardware_worksheet.cell(
            7, 3, "# For virtual control, an integration oversampling can be specified"
        )
        hardware_worksheet.cell(
            7,
            3,
            "Only used for virtual control (Exodus, State Space, or SDynPy).  "
            "This row can be deleted if these are not used.",
        )
        hardware_worksheet.cell(8, 1, "Task Trigger")
        hardware_worksheet.cell(8, 3, "# Start trigger type")
        hardware_worksheet.cell(
            8,
            3,
            "Task Triggers: 0 - Internal, 1 - PFI0 with external trigger, 2 - PFI0 with Analog Output "
            "trigger.  Only used for NI hardware.  This row can be deleted if NI is not used.",
        )
        hardware_worksheet.cell(9, 1, "Task Trigger Output Channel")
        hardware_worksheet.cell(
            9, 3, "# Physical device and channel that generates a trigger signal"
        )
        hardware_worksheet.cell(
            9,
            4,
            "Only used if Task Triggers is 2.  Only used for NI hardware.  "
            "This row can be deleted if it is not used.",
        )
    
    @classmethod
    @abstractmethod
    def load_metadata_from_workbook(cls, workbook: openpyxl.workbook.workbook.Workbook):
        channel_list = cls.load_channel_table_from_workbook(workbook)

        # This is a holdover from previous worksheet logic
        # Need default values
        hardware_type = HardwareType.NONE
        sample_rate = 1000
        time_per_read = 0.1
        time_per_write = 0.1
        output_oversample = 1

        # Hardware
        hardware_sheet = workbook["Hardware"]
        for row in hardware_sheet.rows:
            name = str(row[0].value).lower().strip().replace(" ", "_")
            value = row[1].value
            if value is None or value == "":
                continue
            match name:
                case "hardware_type":
                    hardware_type_int = int(value)
                    hardware_type = HardwareType(hardware_type_int)
                case "sample_rate":
                    sample_rate = int(value)
                case "time_per_read":
                    time_per_read = float(value)
                case "time_per_write":
                    time_per_write = float(value)
                case "integration_oversampling":
                    output_oversample = int(value)
                case _:
                    continue

        return (
            hardware_type,
            channel_list,
            sample_rate,
            time_per_read,
            time_per_write,
            output_oversample,
        )

    @abstractmethod
    def save_metadata_to_netcdf(self, netcdf_dataset: nc4.Dataset):
        stream_variable = "time_data"
        stream_dimension = "time_samples"
        # Create dimensions
        netcdf_dataset.createDimension("response_channels", len(self.channel_list))
        netcdf_dataset.createDimension(
            "output_channels",
            len(
                [
                    channel
                    for channel in self.channel_list
                    if channel.feedback_device is not None
                ]
            ),
        )
        netcdf_dataset.createDimension(stream_dimension, None)
        # Create attributes
        netcdf_dataset.file_version = "3.0.0"
        netcdf_dataset.sample_rate = self.sample_rate
        netcdf_dataset.time_per_write = self.samples_per_write / self.output_sample_rate
        netcdf_dataset.time_per_read = self.samples_per_read / self.sample_rate
        netcdf_dataset.hardware = self.hardware_type.value
        netcdf_dataset.output_oversample = self.output_oversample
        # Create Variables
        netcdf_dataset.createVariable(
            stream_variable, "f8", ("response_channels", stream_dimension)
        )

        # Create channel table variables
        labels = [
            ["node_number", str],
            ["node_direction", str],
            ["comment", str],
            ["serial_number", str],
            ["triax_dof", str],
            ["sensitivity", str],
            ["unit", str],
            ["make", str],
            ["model", str],
            ["expiration", str],
            ["physical_device", str],
            ["physical_channel", str],
            ["channel_type", str],
            ["minimum_value", str],
            ["maximum_value", str],
            ["coupling", str],
            ["excitation_source", str],
            ["excitation", str],
            ["feedback_device", str],
            ["feedback_channel", str],
            ["warning_level", str],
            ["abort_level", str],
        ]
        for label, netcdf_datatype in labels:
            var = netcdf_dataset.createVariable(
                "/channels/" + label, netcdf_datatype, ("response_channels",)
            )
            channel_data = [getattr(channel, label) for channel in self.channel_list]
            if netcdf_datatype == "i1":
                channel_data = np.array([1 if val else 0 for val in channel_data])
            else:
                channel_data = ["" if val is None else val for val in channel_data]
            for i, cd in enumerate(channel_data):
                var[i] = str(cd)

    @classmethod
    @abstractmethod
    def load_metadata_from_netcdf(cls, netcdf_dataset: nc4.Dataset):
        channel_list = cls.load_channel_table_from_netcdf(netcdf_dataset)

        # Hardware
        hardware_type = HardwareType(netcdf_dataset.hardware)
        sample_rate = int(netcdf_dataset.sample_rate)
        time_per_read = float(netcdf_dataset.time_per_read)
        time_per_write = float(netcdf_dataset.time_per_write)
        output_oversample = int(netcdf_dataset.output_oversample)

        return (
            hardware_type,
            channel_list,
            sample_rate,
            time_per_read,
            time_per_write,
            output_oversample,
        )

    @abstractmethod
    def save_metadata_to_workbook(self, workbook: openpyxl.workbook.workbook.Workbook):
        self.save_channel_table_to_workbook(self.channel_list, workbook)

        if "Hardware" in workbook.sheetnames:
            hardware_worksheet = workbook["Hardware"]
        else:
            hardware_worksheet = workbook.create_sheet("Hardware")

        # Fill out values
        hardware_type = self.hardware_type
        hardware_worksheet.cell(1, 2, str(hardware_type.value))
        hardware_worksheet.cell(3, 2, str(self.sample_rate))
        hardware_worksheet.cell(4, 2, str(self.time_per_read))
        hardware_worksheet.cell(5, 2, str(self.time_per_write))

    # endregion


# region Acquisition
class HardwareAcquisition(ABC):
    """
    Abstract class defining the interface between the controller and acquisition.

    This class defines the interfaces between the controller and the
    data acquisition portion of the hardware.  It is run by the Acquisition
    process, and must define how to get data from the test hardware into the
    controller.
    """

    @abstractmethod
    def initialize_hardware(self, metadata: HardwareMetadata) -> None:
        """
        Initialize the hardware and set up channels and sampling properties.

        The function must create channels on the hardware corresponding to
        the channels in the test.  It must also set the sampling rates.

        Parameters
        ----------
        metadata : HardwareMetadata
            Hardware specific metadata class containing the sampling properties
            and channel list to store to the HardwareAcquisition.
        """

    @abstractmethod
    def start(self) -> None:
        """Method to start acquiring data from the hardware."""

    @abstractmethod
    def read(self) -> np.ndarray:
        """Method to read a frame of data from the hardware that returns
        an appropriately sized np.ndarray."""

    @abstractmethod
    def read_remaining(self) -> np.ndarray:
        """Method to read the rest of the data on the acquisition from the hardware
        that returns an appropriately sized np.ndarray."""

    @abstractmethod
    def stop(self) -> None:
        """Method to stop the acquisition."""

    @abstractmethod
    def close(self) -> None:
        """Method to close down the hardware."""

    @abstractmethod
    def get_acquisition_delay(self) -> int:
        """Get the number of samples between output and acquisition.

        This function is designed to handle buffering done in the output
        hardware, ensuring that all data written to the output is read by the
        acquisition.  If a output hardware has a buffer, there may be a non-
        negligable delay between when output is written to the device and
        actually played out from the device."""


# endregion


# region Output
class HardwareOutput(ABC):
    """Abstract class defining the interface between the controller and output

    This class defines the interfaces between the controller and the
    output or source portion of the hardware.  It is run by the Output
    process, and must define how to get write data to the hardware from the
    control system"""

    @abstractmethod
    def initialize_hardware(self, metadata: HardwareMetadata) -> None:
        """
        Initialize the hardware and set up sources and sampling properties

        The function must create channels on the hardware corresponding to
        the sources in the test.  It must also set the sampling rates.

        Parameters
        ----------
        metadata : HardwareMetadata :
            Hardware specific metdata class that defines the sampling properties
            and channel list for a given hardware.
        """
        pass

    @abstractmethod
    def start(self) -> None:
        """Method to start outputting data to the hardware"""
        pass

    @abstractmethod
    def write(self, data) -> None:
        """
        Method to write a np.ndarray with a frame of data to the hardware

        Parameters
        ----------
        data : np.ndarray :
        num_channels x buffer_size array to write to the output hardware
        """
        pass

    @abstractmethod
    def stop(self) -> None:
        """Method to stop the output"""
        pass

    @abstractmethod
    def close(self) -> None:
        """Method to close down the hardware"""
        pass

    @abstractmethod
    def ready_for_new_output(self) -> bool:
        """Method that returns true if the hardware should accept a new signal"""
        pass


# endregion
