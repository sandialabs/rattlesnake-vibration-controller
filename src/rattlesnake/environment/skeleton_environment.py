import copy
from enum import Enum
from typing import List
import multiprocessing as mp
import multiprocessing.queues as mpqueue
import queue as thqueue

import netCDF4 as nc4
import openpyxl
import numpy as np

from rattlesnake.utilities import GlobalCommands, RattlesnakeError, VerboseMessageQueue
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.environment.abstract_environment import (
    Environment,
    EnvironmentCommands,
    EnvironmentInstructions,
    EnvironmentMetadata,
)
from rattlesnake.user_interface.ui_utilities import UICommands

ENVIRONMENT_TYPE = EnvironmentType.SKELETON


# region Commands
class SkeletonCommands(EnvironmentCommands):
    """
    Enumeration of commands that the controller can send to the environment.

    This enum defines command values intended for use in profile events,
    allowing the controller to issue environment-specific instructions at
    designated times.

    Attributes
    ----------
    EXAMPLE_COMMAND : int
        Example command identifier.

    VALID_PROFILE_COMMANDS : tuple of int
        Tuple of commands that are permitted for use as profile events.

    VALID_DATA : dict of int to type
        Mapping from each command to its valid associated data type. This is
        used to validate a profile event list before it is provided to the
        environment.
    """

    EXAMPLE_RUN_ENVIRONMENT = 0
    EXAMPLE_SET_TEST_LEVEL = 1
    EXAMPLE_FLOAT_COMMAND = 2
    EXAMPLE_UNDEFINED_COMMAND = 3

    VALID_PROFILE_COMMANDS = (EXAMPLE_SET_TEST_LEVEL, EXAMPLE_FLOAT_COMMAND)

    VALID_DATA = {
        EXAMPLE_SET_TEST_LEVEL: type(None),
        EXAMPLE_FLOAT_COMMAND: float,
    }


class SkeletonUICommands(Enum):
    """
    Enumeration of commands that the environment can send to the user interface.

    This enum defines commands used to tell the user interface to do something in response to
    the environment. These are usually put directly into the environment_command_queue within
    the control logic loop of the environment.

    Attributes:
    EXAMPLE_UI_COMMAND : int
        Example command identifier.
    """

    EXAMPLE_UI_SHOW_DATA = 0
    EXAMPLE_UI_SET_TEST_LEVEL = 1


# endregion


# region Metadata
class SkeletonMetadata(EnvironmentMetadata):
    """
    Metadata required to define the skeleton environment.

    This class stores the parameters needed to configure an environment after
    hardware initialization. It is used to construct the default metadata
    object when loading Rattlesnake from headless execution, worksheets, or
    netCDF files.

    Parameters
    ----------
    environment_name : str
        Name of the environment, used for logging and identification.
    channel_list_bools : list of bool
        Boolean mask mapping the global channel table to the channels enabled
        for this environment.
    sample_rate : float
        Hardware sample rate in samples per second.
    example_window_size : str
        Example environment-specific parameter.

    Attributes
    ----------
    example_window_size : str
        Example environment-specific parameter.
    """

    def __init__(
        self,
        environment_name: str,
        channel_list_bools: List[bool],
        sample_rate: float,
        example_window_size: str,
    ):
        """
        Initialize the skeleton environment metadata.

        Parameters
        ----------
        environment_name : str
            Name of the environment, used for logging and identification.
        channel_list_bools : list of bool
            Boolean mask mapping the global channel table to the channels
            enabled for this environment.
        sample_rate : float
            Hardware sample rate in samples per second.
        example_window_size : str
            Example environment-specific parameter.
        """
        super().__init__(
            ENVIRONMENT_TYPE,
            environment_name,
            channel_list_bools,
            sample_rate,
        )
        self.example_window_size = example_window_size

    # endregion

    # region Validation
    def validate(self, hardware_metadata: HardwareMetadata):
        """
        Validate the metadata for use with the current hardware configuration.

        This method should raise an error if the metadata is invalid.
        Returning a ``RattlesnakeError`` causes the user interface to display
        the contained message. Other exception types will cause the user
        interface to display a traceback.

        Parameters
        ----------
        hardware_metadata : HardwareMetadata
            Hardware metadata associated with the current Rattlesnake object.
            This is primarily used to validate ``channel_list_bools`` against
            the available hardware channels.

        Raises
        ------
        RattlesnakeError
            If the metadata is invalid for the current hardware or environment
            configuration.
        """
        super().validate(hardware_metadata)

    # endregion

    # region Loading
    def save_metadata_to_netcdf(
        self,
        netcdf_group_handle: nc4._netCDF4.Group,
    ):
        """
        Save environment metadata to a netCDF group.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            netCDF group in which the environment metadata should be stored.

        Notes
        -----
        This method should not close ``netcdf_group_handle``.
        """
        netcdf_group_handle.example_window_size = self.example_window_size

    @classmethod
    def load_metadata_from_netcdf(
        cls,
        netcdf_group_handle: nc4._netCDF4.Group,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        """
        Load environment metadata from a netCDF group.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            netCDF group containing the stored environment metadata.
        environment_name : str
            Name of the environment, used for logging and identification.
        channel_list_bools : list of bool
            Boolean mask mapping the global channel table to the channels
            enabled for this environment.
        hardware_metadata : HardwareMetadata
            Hardware metadata used to supply hardware-dependent values such as
            the sample rate.

        Returns
        -------
        SkeletonMetadata
            Metadata instance reconstructed from the netCDF group.
        """
        example_window_size = netcdf_group_handle.example_window_size

        return cls(
            environment_name,
            channel_list_bools,
            hardware_metadata.sample_rate,
            example_window_size,
        )

    @classmethod
    def create_blank_worksheet_template(cls, worksheet):
        """
        Create a blank worksheet template for skeleton environment metadata.

        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            Worksheet to populate with the template layout and field labels.

        Notes
        ----------
        Worksheet cell 1, 2 must be set to a string of the environment type.
        """
        super().create_blank_worksheet_template(worksheet)

        worksheet.cell(1, 2, ENVIRONMENT_TYPE.name.title())
        worksheet.cell(2, 1, "Example Window Size")
        worksheet.cell(
            2,
            3,
            "# None type object used to show how to store example data to worksheets",
        )

    def save_metadata_to_worksheet(
        self,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
    ):
        """
        Save environment metadata to a worksheet.

        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            Worksheet in which the environment metadata should be stored.
        """
        super().save_metadata_to_worksheet(worksheet)

        if self.example_window_size:
            worksheet.cell(2, 2, self.example_window_size)

    @classmethod
    def load_metadata_from_worksheet(
        cls,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        """
        Load environment metadata from a worksheet.

        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            Worksheet containing the stored environment metadata.
        environment_name : str
            Name of the environment, used for logging and identification.
        channel_list_bools : list of bool
            Boolean mask mapping the global channel table to the channels
            enabled for this environment.
        hardware_metadata : HardwareMetadata
            Hardware metadata used to supply hardware-dependent values such as
            the sample rate.

        Returns
        -------
        SkeletonMetadata
            Metadata instance reconstructed from the worksheet.

        Raises
        ------
        RattlesnakeError
            If the worksheet contains an unexpected parameter name for this
            environment.
        """
        for row in worksheet.rows:
            name = str(row[0].value).lower().strip().replace(" ", "_")
            value = row[1].value
            match name:
                case "control_type":
                    continue
                case "example_window_size":
                    example_window_size = float(value)
                case "":
                    continue
                case _:
                    raise RattlesnakeError(
                        f"{name} does not go with {ENVIRONMENT_TYPE} environment"
                    )

        return cls(
            environment_name,
            channel_list_bools,
            hardware_metadata.sample_rate,
            example_window_size,
        )

    # endregion


# region Instructions
class SkeletonInstructions(EnvironmentInstructions):
    """
    Instructions used to initialize the skeleton environment for a test.

    This object is created and passed to the environment when
    ``GlobalCommands.START_ENVIRONMENT`` is issued. It stores parameters that
    may change frequently between tests and should be initialized at the start
    of each run.

    Parameters
    ----------
    environment_name : str
        Name of the environment, used for logging and identification.
    example_test_level : float
        Example test-level setting for the environment.

    Attributes
    ----------
    example_test_level : float
        Example test-level setting for the environment.
    """

    def __init__(self, environment_name: str, example_test_level: float):
        """
        Initialize the skeleton environment instructions.

        Parameters
        ----------
        environment_name : str
            Name of the environment, used for logging and identification.
        example_test_level : float
            Example test-level setting for the environment.
        """
        super().__init__(ENVIRONMENT_TYPE, environment_name)

        self.example_test_level = example_test_level

    def validate(self):
        return super().validate()


# endregion


# region Queues
class SkeletonQueues:
    """Container for queues passed to the environment on process startup"""

    def __init__(
        self,
        environment_command_queue: VerboseMessageQueue,
        gui_update_queue: mp.queues.Queue,
        controller_communication_queue: VerboseMessageQueue,
        data_in_queue: mp.queues.Queue,
        data_out_queue: mp.queues.Queue,
        log_file_queue: VerboseMessageQueue,
    ):
        """
        Creates a namespace to store all the queues used by the Time Environment

        Parameters
        ----------
        environment_command_queue : VerboseMessageQueue
            Queue from which the environment will receive instructions.
        gui_update_queue : mp.queues.Queue
            Queue to which the environment will put GUI updates.
        controller_communication_queue : VerboseMessageQueue
            Queue to which the environment will put global contorller instructions.
        data_in_queue : mp.queues.Queue
            Queue from which the environment will receive data from acquisition.
        data_out_queue : mp.queues.Queue
            Queue to which the environment will write data for output.
        log_file_queue : VerboseMessageQueue
            Queue to which the environment will write log file messages.
        """
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.log_file_queue = log_file_queue


# endregion


# region Environment
class SkeletonEnvironment(Environment):

    def __init__(
        self,
        environment_name: str,
        queue_name: str,
        queue_container: SkeletonQueues,
        acquisition_active_event: mp.synchronize.Event,
        output_active_event: mp.synchronize.Event,
        active_event: mp.synchronize.Event,
        ready_event: mp.synchronize.Event,
    ):
        """

        Parameters
        ----------
        environment_name : str
            Name of the environment.
        queue_container : SkeletonQueues
            Container of queues used by the Skeleton Environment.
        acqusition_active_event: mp.Event
            Event that is set when the acqusition process is actively reading from the hardware
        output_active_event: mp.Event
            Event that is set when the output process is actively sending data to the hardware
        active_event: mp.Event
            Event that needs to be set when the environment is processing data from acqusition and sending data to output
        ready_event: mp.Event
            Event that is checked by the controller to make sure that the environment process recieved information correctly
            without any errors.
        """
        super().__init__(
            environment_name,
            queue_name,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.controller_communication_queue,
            queue_container.log_file_queue,
            queue_container.data_in_queue,
            queue_container.data_out_queue,
            acquisition_active_event,
            output_active_event,
            active_event,
            ready_event,
        )

        # Define command map
        self.command_map[GlobalCommands.START_ENVIRONMENT] = self.start_environment
        self.command_map[SkeletonCommands.EXAMPLE_RUN_ENVIRONMENT] = self.run_control
        self.command_map[SkeletonCommands.EXAMPLE_SET_TEST_LEVEL] = self.set_test_level

        # Persistent data
        self.test_level = 0
        self.shutdown_flag = True
        self.last_acqusition = False
        self.control_channels = []
        self.output_signal = []

        # Tell controller that initialization was successful
        self.set_ready()

    # endregion

    # region State Sync
    def initialize_hardware(self, hardware_metadata: HardwareMetadata):
        super().initialize_hardware(hardware_metadata)

        self.control_channels = [
            index
            for index, channel in enumerate(hardware_metadata.channel_list)
            if channel.feedback_device is not None
        ]
        self.output_signal = np.zeros(
            (len(self.control_channels), self.hardware_metadata.samples_per_write)
        )
        self.set_ready()

    def initialize_environment(self, environment_metadata: SkeletonMetadata):
        super().initialize_environment(environment_metadata)
        self.set_ready()

    # endregion

    # region Commands
    def start_environment(self, data: SkeletonInstructions):
        if not self.active:
            # Store instructions
            if data is not None:
                test_level = data.example_test_level
                self.test_level = test_level
                self.gui_update_queue.put(
                    (
                        self.environment_name,
                        (SkeletonUICommands.EXAMPLE_UI_SET_TEST_LEVEL, test_level),
                    )
                )

            # Set startup flags
            self.set_active()
            self.shutdown_flag = False
            self.last_acqusition = False
            self.gui_update_queue.put(
                (self.environment_name, (UICommands.ENVIRONMENT_STARTED, None))
            )

            # Start Run Environment loop
            self.environment_command_queue.put(
                self.environment_name, (SkeletonCommands.EXAMPLE_RUN_ENVIRONMENT, None)
            )

    def run_control(self, data: None):
        # Get data from data in queue and send it to user interface
        try:
            acqusition_data, self.last_acqusition = self.data_in_queue.get_nowait()
            self.gui_update_queue.put(
                (
                    self.environment_name,
                    (SkeletonUICommands.EXAMPLE_UI_SHOW_DATA, acqusition_data),
                ),
            )
        except (thqueue.Empty, mpqueue.Empty):
            pass

        # If required, put data to data out queue
        if self.data_out_queue.empty():
            self.data_out_queue.put(
                (copy.deepcopy(self.output_signal), self.shutdown_flag)
            )
            if self.shutdown_flag:
                self.shutdown_flag = False

        # If environment is shutting down, flush queue and update UI
        if self.last_acqusition:
            self.environment_command_queue.flush(self.environment_name)
            self.gui_update_queue.put(
                (self.environment_name, (UICommands.ENVIRONMENT_ENDED, None))
            )
            self.clear_active()
        # Run control if environment is not shutting down
        else:
            self.environment_command_queue.put(
                self.environment_name, (SkeletonCommands.EXAMPLE_RUN_ENVIRONMENT, None)
            )

    def stop_environment(self, data):
        # Set shutdown flag so the run_control knows to stop control loop
        self.shutdown_flag = True

    def set_test_level(self, data):
        self.test_level = data
        print(f"Setting test level {self.test_level}")


def skeleton_process(
    environment_name: str,
    queue_name: str,
    input_queue: VerboseMessageQueue,
    gui_update_queue: mp.Queue,
    controller_command_queue: VerboseMessageQueue,
    log_file_queue: mp.Queue,
    data_in_queue: mp.Queue,
    data_out_queue: mp.Queue,
    acquisition_active_event: mp.synchronize.Event,
    output_active_event: mp.synchronize.Event,
    active_event: mp.synchronize.Event,
    ready_event: mp.synchronize.Event,
    shutdown_event: mp.synchronize.Event,
    sysid_active_event: mp.synchronize.Event,
    sysid_stored_event: mp.synchronize.Event,
    ping_alive_event: mp.synchronize.Event,
    threaded: bool,
):
    queue_container = SkeletonQueues(
        input_queue,
        gui_update_queue,
        controller_command_queue,
        data_in_queue,
        data_out_queue,
        log_file_queue,
    )

    process_class = SkeletonEnvironment(
        environment_name,
        queue_name,
        queue_container,
        acquisition_active_event,
        output_active_event,
        active_event,
        ready_event,
    )
    process_class.run(shutdown_event)
