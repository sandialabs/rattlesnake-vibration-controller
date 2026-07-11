# -*- coding: utf-8 -*-
"""
This file defines a transient environment that utilizes system
identification.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import importlib
import multiprocessing as mp
import multiprocessing.sharedctypes  # pylint: disable=unused-import
import os
import threading
import traceback
import inspect
from enum import Enum
from multiprocessing.queues import Queue
from typing import List

import openpyxl
import netCDF4 as nc4
import numpy as np
import scipy.signal as sig

from rattlesnake.utilities import (
    GlobalCommands,
    VerboseMessageQueue,
    db2scale,
    align_signals,
    shift_signal,
    load_python_module,
    load_time_history,
    trac,
)
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.environment.abstract_environment import (
    EnvironmentInstructions,
    EnvironmentCommands,
)
from rattlesnake.environment.abstract_interactive_control_law import (
    AbstractControlLawComputation,
)
from rattlesnake.environment.abstract_sysid_environment import (
    SysIdEnvironment,
    SysIdEnvironmentMetadata,
)
from rattlesnake.environment.abstract_interactive_control_law import ControlLawCommands

from rattlesnake.process.abstract_sysid_data_analysis import (
    SysIdMetadata,
    SysIdDataPackage,
    sysid_data_analysis_process,
)
from rattlesnake.process.data_collector import (
    FrameBuffer,
    data_collector_process,
)
from rattlesnake.process.signal_generation import (
    TransientSignalGenerator,
)
from rattlesnake.process.signal_generation_process import (
    SignalGenerationCommands,
    SignalGenerationMetadata,
    signal_generation_process,
)
from rattlesnake.process.spectral_processing import (
    spectral_processing_process,
)
from rattlesnake.user_interface.ui_utilities import UICommands

# %% Global Variables
CONTROL_TYPE = EnvironmentType.TRANSIENT
BUFFER_SIZE_SAMPLES_PER_READ_MULTIPLIER = 2


# region Commands
class TransientCommands(EnvironmentCommands):
    """Valid commands for the transient environment"""

    START_CONTROL = 0
    STOP_CONTROL = 1
    PERFORM_CONTROL_PREDICTION = 3
    SET_TEST_LEVEL = 5
    SET_REPEAT = 6
    SET_NO_REPEAT = 7

    VALID_PROFILE_COMMANDS = {
        SET_TEST_LEVEL,
        SET_REPEAT,
        SET_NO_REPEAT,
    }
    VALID_DATA = {
        SET_TEST_LEVEL: int,
        SET_REPEAT: type(None),
        SET_NO_REPEAT: type(None),
    }


class TransientUICommands(Enum):
    INTERACTIVE_CONTROL_SYSID_UPDATE = 0
    CONTROL_PREDICTIONS = 1
    TIME_DATA = 2
    CONTROL_DATA = 3


# endregion


class TransientMetadata(SysIdEnvironmentMetadata):
    """Metadata required to define a transient control law in rattlesnake."""

    # region Metadata
    def __init__(
        self,
        environment_name,
        channel_list_bools,
        sample_rate,
        number_of_channels,
        control_signal,
        ramp_time,
        control_python_script,
        control_python_function,
        control_python_function_type,
        control_python_function_parameters,
        control_channel_indices,
        output_channel_indices,
        response_transformation_matrix,
        output_transformation_matrix,
        sysid_metadata=None,
    ):
        super().__init__(
            CONTROL_TYPE,
            environment_name,
            channel_list_bools,
            sample_rate,
            sysid_metadata,
        )
        self.number_of_channels = number_of_channels
        self.sample_rate = sample_rate
        self.control_signal = control_signal
        self.test_level_ramp_time = ramp_time
        self.control_python_script = control_python_script
        self.control_python_function = control_python_function
        self.control_python_function_type = control_python_function_type
        self.control_python_function_parameters = control_python_function_parameters
        self.control_channel_indices = control_channel_indices
        self.output_channel_indices = output_channel_indices
        self.response_transformation_matrix = response_transformation_matrix
        self.reference_transformation_matrix = output_transformation_matrix

    @property
    def ramp_samples(self):
        """Number of samples to ramp down to zero when aborting a test"""
        return int(self.test_level_ramp_time * self.sample_rate)

    @property
    def number_of_channels(self):
        """Total number of channels in the environment"""
        return self._number_of_channels

    @number_of_channels.setter
    def number_of_channels(self, value):
        """Sets the total number of channels in the environment"""
        self._number_of_channels = value

    @property
    def response_channel_indices(self):
        """Indices identifying which channels are control channels"""
        return self.control_channel_indices

    @property
    def reference_channel_indices(self):
        """Indices identifying which channels are reference or excitation channels"""
        return self.output_channel_indices

    @property
    def response_transformation_matrix(self):
        """Transformation matrix applied to the control channels"""
        return self._response_transformation_matrix

    @response_transformation_matrix.setter
    def response_transformation_matrix(self, value):
        """Sets the transformation matrix for the control channels"""
        self._response_transformation_matrix = value

    @property
    def reference_transformation_matrix(self):
        """Transformation matrix applied to the excitation channels"""
        return self._reference_transformation_matrix

    @reference_transformation_matrix.setter
    def reference_transformation_matrix(self, value):
        """Sets the transformation matrix applied to the excitation channels"""
        self._reference_transformation_matrix = value

    @property
    def sample_rate(self):
        """Gets the sample rate of the data acquisition system"""
        return self._sample_rate

    @sample_rate.setter
    def sample_rate(self, value):
        """Sets the sample rate of the data acquisition system"""
        self._sample_rate = value

    @property
    def signal_samples(self):
        """Gets the number of samples in the signal that is being controlled to"""
        return self.control_signal.shape[-1]

    # endregion

    # region Validation
    def validate(self, hardware_metadata):
        return super().validate(hardware_metadata)

    # endregion

    # region Loading
    def save_metadata_to_netcdf(
        self,
        netcdf_group_handle: nc4._netCDF4.Group,  # pylint: disable=c-extension-no-member
    ):
        """Stores the metadata in a netcdf group

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            A group in a NetCDF4 group defining the environment's medatadata
        """
        super().save_metadata_to_netcdf(netcdf_group_handle)
        netcdf_group_handle.test_level_ramp_time = self.test_level_ramp_time
        netcdf_group_handle.control_python_script = self.control_python_script
        netcdf_group_handle.control_python_function = self.control_python_function
        # netCDF attributes cannot be None; -1 marks "no control script"
        netcdf_group_handle.control_python_function_type = (
            -1
            if self.control_python_function_type is None
            else self.control_python_function_type
        )
        netcdf_group_handle.control_python_function_parameters = (
            self.control_python_function_parameters
        )
        # Save the output signal
        netcdf_group_handle.createDimension(
            "control_channels", len(self.control_channel_indices)
        )
        netcdf_group_handle.createDimension(
            "specification_channels", self.control_signal.shape[0]
        )
        netcdf_group_handle.createDimension("signal_samples", self.signal_samples)
        var = netcdf_group_handle.createVariable(
            "control_signal", "f8", ("specification_channels", "signal_samples")
        )
        var[...] = self.control_signal
        # Control Channels
        var = netcdf_group_handle.createVariable(
            "control_channel_indices", "i4", ("control_channels")
        )
        var[...] = self.control_channel_indices
        # Transformation Matrix
        if self.response_transformation_matrix is not None:
            netcdf_group_handle.createDimension(
                "response_transformation_rows",
                self.response_transformation_matrix.shape[0],
            )
            netcdf_group_handle.createDimension(
                "response_transformation_cols",
                self.response_transformation_matrix.shape[1],
            )
            var = netcdf_group_handle.createVariable(
                "response_transformation_matrix",
                "f8",
                ("response_transformation_rows", "response_transformation_cols"),
            )
            var[...] = self.response_transformation_matrix
        if self.reference_transformation_matrix is not None:
            netcdf_group_handle.createDimension(
                "reference_transformation_rows",
                self.reference_transformation_matrix.shape[0],
            )
            netcdf_group_handle.createDimension(
                "reference_transformation_cols",
                self.reference_transformation_matrix.shape[1],
            )
            var = netcdf_group_handle.createVariable(
                "reference_transformation_matrix",
                "f8",
                ("reference_transformation_rows", "reference_transformation_cols"),
            )
            var[...] = self.reference_transformation_matrix

    @classmethod
    def load_metadata_from_netcdf(
        cls,
        netcdf_group_handle: nc4._netCDF4.Group,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        sysid_metadata = SysIdMetadata.load_metadata_from_netcdf(
            netcdf_group_handle, hardware_metadata
        )
        sample_rate = hardware_metadata.sample_rate

        test_level_ramp_time = netcdf_group_handle.test_level_ramp_time
        control_python_script = netcdf_group_handle.control_python_script
        control_python_function = netcdf_group_handle.control_python_function
        control_python_function_type = netcdf_group_handle.control_python_function_type
        if control_python_function_type == -1:
            control_python_function_type = None
        control_python_function_parameters = (
            netcdf_group_handle.control_python_function_parameters
        )
        # Load Variables
        control_signal = netcdf_group_handle.variables["control_signal"][:]
        control_channel_indices = netcdf_group_handle.variables[
            "control_channel_indices"
        ][:]

        # Extract number of channels from group or hardware
        number_of_channels = netcdf_group_handle.dimensions[
            "specification_channels"
        ].size

        # Handle derived channel lists (matching your example pattern)
        environment_channel_list = [
            channel
            for channel, channel_bool in zip(
                hardware_metadata.channel_list, channel_list_bools
            )
            if channel_bool
        ]

        output_channel_indices = [
            index
            for index, channel in enumerate(environment_channel_list)
            if channel.feedback_device is not None
        ]

        # Optional Transformation Matrices
        response_transformation_matrix = None
        if "response_transformation_matrix" in netcdf_group_handle.variables:
            response_transformation_matrix = netcdf_group_handle.variables[
                "response_transformation_matrix"
            ][:]

        reference_transformation_matrix = None
        if "reference_transformation_matrix" in netcdf_group_handle.variables:
            reference_transformation_matrix = netcdf_group_handle.variables[
                "reference_transformation_matrix"
            ][:]

        return cls(
            environment_name=environment_name,
            channel_list_bools=channel_list_bools,
            sample_rate=sample_rate,
            number_of_channels=number_of_channels,
            control_signal=control_signal,
            ramp_time=test_level_ramp_time,
            control_python_script=control_python_script,
            control_python_function=control_python_function,
            control_python_function_type=control_python_function_type,
            control_python_function_parameters=control_python_function_parameters,
            control_channel_indices=control_channel_indices,
            output_channel_indices=output_channel_indices,
            response_transformation_matrix=response_transformation_matrix,
            output_transformation_matrix=reference_transformation_matrix,
            sysid_metadata=sysid_metadata,
        )

    @classmethod
    def create_blank_worksheet_template(cls, worksheet):
        super().create_blank_worksheet_template(worksheet)
        worksheet.cell(1, 2, "Transient")
        worksheet.cell(2, 1, "Signal File")
        worksheet.cell(
            2, 3, "# Path to the file that contains the time signal that will be output"
        )
        worksheet.cell(3, 1, "Ramp Time")
        worksheet.cell(
            3,
            3,
            "# Time for the environment to ramp between levels or from start or to stop.",
        )
        worksheet.cell(4, 1, "Control Python Script:")
        worksheet.cell(4, 3, "# Path to the Python script containing the control law")
        worksheet.cell(5, 1, "Control Python Function:")
        worksheet.cell(
            5,
            3,
            "# Function name within the Python Script that will serve as the control law",
        )
        worksheet.cell(6, 1, "Control Parameters:")
        worksheet.cell(6, 3, "# Extra parameters used in the control law")
        worksheet.cell(7, 1, "Control Channels (1-based):")
        worksheet.cell(7, 3, "# List of channels, one per cell on this row")
        SysIdMetadata.create_blank_worksheet_template(worksheet, start_row=8)
        worksheet.cell(24, 1, "Response Transformation Matrix:")
        worksheet.cell(
            24,
            2,
            "# Transformation matrix to apply to the response channels.  Type None if there "
            "is none.  Otherwise, make this a 2D array in the spreadsheet and move the Output "
            "Transformation Matrix line down so it will fit.  The number of columns should be "
            "the number of physical control channels.",
        )
        worksheet.cell(25, 1, "Output Transformation Matrix:")
        worksheet.cell(
            25,
            2,
            "# Transformation matrix to apply to the outputs.  Type None if there is none.  "
            "Otherwise, make this a 2D array in the spreadsheet.  The number of columns should "
            "be the number of physical output channels in the environment.",
        )

    def save_metadata_to_worksheet(
        self, worksheet: openpyxl.worksheet.worksheet.Worksheet
    ):
        super().save_metadata_to_worksheet(worksheet)
        if self.test_level_ramp_time is not None:
            worksheet.cell(3, 2, self.test_level_ramp_time)
        if self.control_python_script is not None:
            worksheet.cell(4, 2, self.control_python_script)
        if self.control_python_function is not None:
            worksheet.cell(5, 2, self.control_python_function)
        if self.control_python_function_parameters is not None:
            worksheet.cell(6, 2, self.control_python_function_parameters)
        if self.control_channel_indices is not None:
            for idx, channel_ind in enumerate(self.control_channel_indices):
                col_idx = idx + 2
                worksheet.cell(7, col_idx, channel_ind + 1)
        self.sysid_metadata.save_metadata_to_worksheet(worksheet, start_row=8)
        self.save_sysid_matrix_to_worksheet(
            worksheet,
            self.response_transformation_matrix,
            self.reference_transformation_matrix,
            start_row=24,
        )

    @classmethod
    def load_metadata_from_worksheet(
        cls,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        sample_rate = hardware_metadata.sample_rate
        number_of_channels = sum(channel_list_bools)
        environment_channel_list = [
            channel
            for channel, channel_bool in zip(
                hardware_metadata.channel_list, channel_list_bools
            )
            if channel_bool
        ]

        ramp_time = float(worksheet.cell(3, 2).value)

        control_python_script = (
            worksheet.cell(4, 2).value if worksheet.cell(4, 2).value is not None else ""
        )
        control_python_function = (
            worksheet.cell(5, 2).value if worksheet.cell(5, 2).value is not None else ""
        )
        control_python_function_parameters = (
            worksheet.cell(6, 2).value if worksheet.cell(6, 2).value is not None else ""
        )
        control_channel_indices = []
        column_index = 2
        while True:
            channel_ind = worksheet.cell(7, column_index).value
            if channel_ind is None or (
                isinstance(channel_ind, str)
                and (channel_ind.startswith("#") or channel_ind.strip() == "")
            ):
                break
            try:
                control_channel_indices.append(int(channel_ind) - 1)
            except:
                break
            column_index += 1
        output_channel_indices = [
            index
            for index, channel in enumerate(environment_channel_list)
            if channel.feedback_device is not None
        ]

        # Find python module type
        if control_python_script:
            python_control_module = load_python_module(control_python_script)
            function = getattr(python_control_module, control_python_function)
            control_python_function_type = None
            if inspect.isgeneratorfunction(function):
                control_python_function_type = 1
            elif inspect.isclass(function) and issubclass(
                function, AbstractControlLawComputation
            ):
                control_python_function_type = 2
            elif inspect.isclass(function):
                control_python_function_type = 3
            else:
                control_python_function_type = 0
        else:
            control_python_function_type = None

        response_transformation_matrix, output_transformation_matrix = (
            cls.load_sysid_matrix_from_worksheet(worksheet, start_row=24)
        )

        sysid_metadata = SysIdMetadata.load_metadata_from_worksheet(
            worksheet, hardware_metadata, 8
        )

        metadata = cls(
            environment_name=environment_name,
            channel_list_bools=channel_list_bools,
            sample_rate=sample_rate,
            number_of_channels=number_of_channels,
            control_signal=None,
            ramp_time=ramp_time,
            control_python_script=control_python_script,
            control_python_function=control_python_function,
            control_python_function_type=control_python_function_type,
            control_python_function_parameters=control_python_function_parameters,
            control_channel_indices=control_channel_indices,
            output_channel_indices=output_channel_indices,
            response_transformation_matrix=response_transformation_matrix,
            output_transformation_matrix=output_transformation_matrix,
            sysid_metadata=sysid_metadata,
        )

        spec_filename = worksheet.cell(2, 2).value
        if spec_filename is not None:
            metadata.control_signal = load_time_history(
                spec_filename, hardware_metadata.sample_rate
            )

        return metadata


# endregion


# region Instructions
class TransientInstructions(EnvironmentInstructions):
    def __init__(self, environment_name, test_level, repeat):
        super().__init__(CONTROL_TYPE, environment_name)
        self.test_level = test_level
        self.repeat = repeat

    def validate(self):
        return super().validate()


# endregion


# region Queues
class TransientQueues:
    """A container class for the queues that this environment will manage."""

    def __init__(
        self,
        environment_name: str,
        environment_command_queue: VerboseMessageQueue,
        gui_update_queue: Queue,
        controller_communication_queue: VerboseMessageQueue,
        data_in_queue: Queue,
        data_out_queue: Queue,
        log_file_queue: VerboseMessageQueue,
    ):
        """A container class for the queues that transient will manage.

        The environment uses many queues to pass data between the various pieces.
        This class organizes those queues into one common namespace.


        Parameters
        ----------
        environment_name : str
            Name of the environment
        environment_command_queue : VerboseMessageQueue
            Queue that is read by the environment for environment commands
        gui_update_queue : mp.queues.Queue
            Queue where various subtasks put instructions for updating the
            widgets in the user interface
        controller_communication_queue : VerboseMessageQueue
            Queue that is read by the controller for global controller commands
        data_in_queue : mp.queues.Queue
            Multiprocessing queue that connects the acquisition subtask to the
            environment subtask.  Each environment will retrieve acquired data
            from this queue.
        data_out_queue : mp.queues.Queue
            Multiprocessing queue that connects the output subtask to the
            environment subtask.  Each environment will put data that it wants
            the controller to generate in this queue.
        log_file_queue : VerboseMessageQueue
            Queue for putting logging messages that will be read by the logging
            subtask and written to a file.
        """
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.data_analysis_command_queue = VerboseMessageQueue(
            log_file_queue,
            mp.Queue(),
            environment_name + " Data Analysis Command Queue",
        )
        self.signal_generation_command_queue = VerboseMessageQueue(
            log_file_queue,
            mp.Queue(),
            environment_name + " Signal Generation Command Queue",
        )
        self.spectral_command_queue = VerboseMessageQueue(
            log_file_queue,
            mp.Queue(),
            environment_name + " Spectral Computation Command Queue",
        )
        self.collector_command_queue = VerboseMessageQueue(
            log_file_queue,
            mp.Queue(),
            environment_name + " Data Collector Command Queue",
        )
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.data_for_spectral_computation_queue = mp.Queue()
        self.updated_spectral_quantities_queue = mp.Queue()
        self.time_history_to_generate_queue = mp.Queue()
        self.log_file_queue = log_file_queue


class TransientEnvironment(SysIdEnvironment):
    """Class defining calculations for the transient environment"""

    # region Environment
    def __init__(
        self,
        environment_name: str,
        queue_name: str,
        queue_container: TransientQueues,
        acquisition_active_event: mp.synchronize.Event,
        output_active_event: mp.synchronize.Event,
        active_event: mp.synchronize.Event,
        ready_event: mp.synchronize.Event,
        sysid_active_event: mp.synchronize.Event,
        sysid_stored_event: mp.synchronize.Event,
    ):
        super().__init__(
            environment_name,
            queue_name,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.controller_communication_queue,
            queue_container.log_file_queue,
            queue_container.collector_command_queue,
            queue_container.signal_generation_command_queue,
            queue_container.spectral_command_queue,
            queue_container.data_analysis_command_queue,
            queue_container.data_in_queue,
            queue_container.data_out_queue,
            acquisition_active_event,
            output_active_event,
            active_event,
            ready_event,
            sysid_active_event,
            sysid_stored_event,
        )
        self.map_command(
            TransientCommands.PERFORM_CONTROL_PREDICTION,
            self.perform_control_prediction,
        )
        self.map_command(GlobalCommands.START_ENVIRONMENT, self.start_control)
        self.map_command(TransientCommands.START_CONTROL, self.start_control)
        self.map_command(TransientCommands.STOP_CONTROL, self.stop_environment)
        self.map_command(
            ControlLawCommands.UPDATE_INTERACTIVE_CONTROL_PARAMETERS,
            self.update_interactive_control_parameters,
        )
        self.map_command(
            ControlLawCommands.SEND_INTERACTIVE_COMMAND, self.send_interactive_command
        )
        # Persistent data
        self.hardware_metadata = None
        self.environment_metadata = None
        self.queue_container = queue_container
        self.control_function_type = None
        self.extra_control_parameters = None
        self.control_function = None
        self.aligned_output = None
        self.aligned_response = None
        self.next_drive = None
        self.predicted_response = None
        self.startup = True
        self.shutdown_flag = False
        self.repeat = False
        self.test_level = 0
        self.control_buffer = None
        self.output_buffer = None
        self.last_signal_found = None
        self.has_sent_interactive_control_transfer_function_results = False
        self.last_interactive_parameters = None

        self.set_ready()

    # endregion

    # region State Sync
    def initialize_hardware(self, hardware_metadata):
        super().initialize_hardware(hardware_metadata)
        self.set_ready()

    def initialize_environment(self, environment_metadata: TransientMetadata):
        if (
            self.environment_metadata is None
            or self.environment_metadata.control_signal.shape
            != environment_metadata.control_signal.shape
        ):
            self.sysid_data = SysIdDataPackage()
            self.control_function_type = None
            self.extra_control_parameters = None
            self.control_function = None
            self.aligned_output = None
            self.aligned_response = None
            self.next_drive = None
            self.predicted_response = None
        super().initialize_environment(environment_metadata)
        # Load in the control law
        _, file = os.path.split(environment_metadata.control_python_script)
        file, _ = os.path.splitext(file)
        spec = importlib.util.spec_from_file_location(
            file, environment_metadata.control_python_script
        )
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        self.control_function_type = environment_metadata.control_python_function_type
        self.extra_control_parameters = (
            environment_metadata.control_python_function_parameters
        )
        if self.control_function_type == 1:  # Generator
            # Get the generator function
            generator_function = getattr(
                module, environment_metadata.control_python_function
            )()
            # Get us to the first yield statement
            next(generator_function)
            # Define the control function as the generator's send function
            self.control_function = generator_function.send
        elif self.control_function_type == 2:  # Class
            self.control_function = getattr(
                module, environment_metadata.control_python_function
            )(
                self.hardware_metadata.sample_rate,
                self.environment_metadata.control_signal,
                self.hardware_metadata.output_oversample,
                self.extra_control_parameters,  # Required parameters
                self.environment_metadata.sysid_metadata.sysid_frequency_spacing,  # Frequency Spacing
                self.sysid_data.sysid_frf,  # Transfer Functions
                self.sysid_data.sysid_response_noise,  # Noise levels and correlation
                self.sysid_data.sysid_reference_noise,  # from the system identification
                self.sysid_data.sysid_response_cpsd,  # Response levels and correlation
                self.sysid_data.sysid_reference_cpsd,  # from the system identification
                self.sysid_data.sysid_coherence,  # Coherence from the system identification
                self.sysid_data.sysid_frames,  # Number of frames in the CPSD and FRF matrices
                self.environment_metadata.sysid_metadata.sysid_averages,  # Total frames that
                # could be in the CPSD and FRF matrices
                self.aligned_output,  # Last excitation signal for drive-based control
                self.aligned_response,
            )  # Last response signal for error-based correction
        elif self.control_function_type == 3:  # Interactive Class
            control_class = getattr(
                module, environment_metadata.control_python_function
            )
            self.control_function = control_class(
                self.environment_name,
                self.gui_update_queue,
                self.hardware_metadata.sample_rate,
                self.environment_metadata.control_signal,
                self.hardware_metadata.output_oversample,
                self.extra_control_parameters,  # Required parameters
                self.environment_metadata.sysid_metadata.sysid_frequency_spacing,  # Frequency Spacing
                self.sysid_data.sysid_frf,  # Transfer Functions
                self.sysid_data.sysid_response_noise,  # Noise levels and correlation
                self.sysid_data.sysid_reference_noise,  # from the system identification
                self.sysid_data.sysid_response_cpsd,  # Response levels and correlation
                self.sysid_data.sysid_reference_cpsd,  # from the system identification
                self.sysid_data.sysid_coherence,  # Coherence from the system identification
                self.sysid_data.sysid_frames,  # Number of frames in the CPSD and FRF matrices
                self.environment_metadata.sysid_metadata.sysid_averages,  # Total frames tha
                # could be in the CPSD and FRF matrices
                self.aligned_output,  # Last excitation signal for drive-based control
                self.aligned_response,
            )  # Last response signal for error-based correction
            self.last_interactive_parameters = None
            self.has_sent_interactive_control_transfer_function_results = False
        else:  # Function
            self.control_function = getattr(
                module, environment_metadata.control_python_function
            )

        self.set_ready()

    def initialize_sysid(self, sysid_metadata):
        super().initialize_sysid(sysid_metadata)

        self.set_ready()

    def update_interactive_control_parameters(self, interactive_control_parameters):
        """Updates the interactive control law based on received parameters"""
        if self.environment_metadata.control_python_function_type == 3:  # Interactive
            self.control_function.update_parameters(interactive_control_parameters)
            self.last_interactive_parameters = interactive_control_parameters
        else:
            raise ValueError(
                "Received an UPDATE_INTERACTIVE_CONTROL_PARAMETERS signal without an "
                "interactive control law.  How did this happen?"
            )

    def get_signal_generation_metadata(self):
        """Collects the metadata required to define the signal generation process"""
        return SignalGenerationMetadata(
            samples_per_write=self.hardware_metadata.samples_per_write,
            level_ramp_samples=self.environment_metadata.test_level_ramp_time
            * self.environment_metadata.sample_rate
            * self.hardware_metadata.output_oversample,
            output_transformation_matrix=self.environment_metadata.reference_transformation_matrix,
        )

    # endregion

    # region Commands
    def send_interactive_command(self, command):
        """General method that can be used by an interactive UI object to pass commands
        and data to its corresponding computation object"""
        if self.environment_metadata.control_python_function_type == 3:  # Interactive
            self.control_function.send_command(command)
        else:
            raise ValueError(
                "Received an SEND_INTERACTIVE_COMMAND signal without an interactive "
                "control law.  How did this happen?"
            )

    def system_id_complete(self, data):
        """Sends the message that system identification is complete and control calculations
        should be performed"""
        super().system_id_complete(data)

        # Perform the control prediction
        self.perform_control_prediction(True)
        self.set_sysid_stored()

    def perform_control_prediction(self, sysid_update):
        """Performs the control prediction based on system identification information"""
        if self.sysid_data.sysid_frf is None:
            self.gui_update_queue.put(
                (
                    UICommands.ERROR,
                    (
                        "Perform System Identification",
                        "Perform System ID before performing test predictions",
                    ),
                )
            )
            return
        if self.control_function_type == 1:  # Generator
            output_time_history = self.control_function(
                (
                    self.hardware_metadata.sample_rate,
                    self.environment_metadata.control_signal,
                    self.environment_metadata.sysid_metadata.sysid_frequency_spacing,
                    self.sysid_data.sysid_frf,  # Transfer Functions
                    self.sysid_data.sysid_response_noise,  # Noise levels and correlation
                    self.sysid_data.sysid_reference_noise,  # from the system identification
                    self.sysid_data.sysid_response_cpsd,  # Response levels and correlation
                    self.sysid_data.sysid_reference_cpsd,  # from the system identification
                    self.sysid_data.sysid_coherence,  # Coherence from the system identification
                    self.sysid_data.sysid_frames,  # Number of frames in the CPSD and FRF matrices
                    self.environment_metadata.sysid_metadata.sysid_averages,  # Total frames that could be in
                    #  the CPSD and FRF matrices
                    self.hardware_metadata.output_oversample,
                    self.extra_control_parameters,  # Required parameters
                    self.next_drive,  # Last excitation signal for drive-based control
                    self.predicted_response,  # Last response signal for error correction
                )
            )
        elif self.control_function_type in [2, 3]:  # Class or Interactive Class
            if (
                self.environment_metadata.control_python_function == 2
                or not self.has_sent_interactive_control_transfer_function_results
            ):
                if sysid_update:
                    self.control_function.system_id_update(
                        self.environment_metadata.sysid_metadata.sysid_frequency_spacing,
                        self.sysid_data.sysid_frf,  # Transfer Functions
                        self.sysid_data.sysid_response_noise,  # Noise levels and correlation
                        self.sysid_data.sysid_reference_noise,  # from the system identification
                        self.sysid_data.sysid_response_cpsd,  # Response levels and correlation
                        self.sysid_data.sysid_reference_cpsd,  # from the system identification
                        self.sysid_data.sysid_coherence,  # Coherence from the system identification
                        self.sysid_data.sysid_frames,  # Number of frames in the CPSD and FRF matrices
                        self.environment_metadata.sysid_metadata.sysid_averages,  # Total frames that
                        # could be in the CPSD and FRF matrices
                    )

                if self.environment_metadata.control_python_function_type == 3:
                    self.gui_update_queue.put(
                        (
                            self.environment_name,
                            (
                                TransientUICommands.INTERACTIVE_CONTROL_SYSID_UPDATE,
                                (
                                    self.sysid_data.sysid_frf,
                                    self.sysid_data.sysid_response_noise,
                                    self.sysid_data.sysid_reference_noise,
                                    self.sysid_data.sysid_response_cpsd,
                                    self.sysid_data.sysid_reference_cpsd,
                                    self.sysid_data.sysid_coherence,
                                ),
                            ),
                        )
                    )
                    self.has_sent_interactive_control_transfer_function_results = True
            if (
                self.environment_metadata.control_python_function_type == 2
                or self.last_interactive_parameters is not None
            ):
                output_time_history = self.control_function.control(
                    self.next_drive, self.predicted_response
                )
            else:
                self.log(
                    "Have not yet received control parameters from interactive control law!"
                )
                output_time_history = None
                return
        else:  # Function
            output_time_history = self.control_function(
                self.hardware_metadata.sample_rate,
                self.environment_metadata.control_signal,
                self.environment_metadata.sysid_metadata.sysid_frequency_spacing,
                self.sysid_data.sysid_frf,  # Transfer Functions
                self.sysid_data.sysid_response_noise,  # Noise levels and correlation
                self.sysid_data.sysid_reference_noise,  # from the system identification
                self.sysid_data.sysid_response_cpsd,  # Response levels and correlation
                self.sysid_data.sysid_reference_cpsd,  # from the system identification
                self.sysid_data.sysid_coherence,  # Coherence from the system identification
                self.sysid_data.sysid_frames,  # Number of frames in the CPSD and FRF matrices
                self.environment_metadata.sysid_metadata.sysid_averages,  # Total frames that could
                # be in the CPSD and FRF matrices
                self.hardware_metadata.output_oversample,
                self.extra_control_parameters,  # Required parameters
                self.next_drive,  # Last excitation signal for drive-based control
                self.predicted_response,  # Last response signal for error correction
            )
        self.next_drive = output_time_history
        self.show_test_prediction()

    def show_test_prediction(self):
        """Sends the test predictions to the UI"""
        # print('Drive Signals {:}'.format(self.next_drive.shape))
        drive_signals = self.next_drive[:, :: self.hardware_metadata.output_oversample]
        impulse_responses = np.moveaxis(
            np.fft.irfft(self.sysid_data.sysid_frf, axis=0), 0, -1
        )

        self.predicted_response = np.zeros(
            (impulse_responses.shape[0], drive_signals.shape[-1])
        )

        for i, impulse_response_row in enumerate(impulse_responses):
            for _, (impulse, drive) in enumerate(
                zip(impulse_response_row, drive_signals)
            ):
                # print('Convolving {:},{:}'.format(i,j))
                self.predicted_response[i, :] += sig.convolve(drive, impulse, "full")[
                    : drive_signals.shape[-1]
                ]

        # print('Response Prediction {:}'.format(self.predicted_response.shape))
        # print('Control Signal {:}'.format(self.environment_metadata.control_signal.shape))
        time_trac = trac(
            self.predicted_response, self.environment_metadata.control_signal
        )
        peak_voltages = np.max(np.abs(self.next_drive), axis=-1)
        self.gui_update_queue.put(
            (
                self.environment_name,
                (UICommands.SET_ATTR, ("excitation_voltage_list", peak_voltages)),
            )
        )
        self.gui_update_queue.put(
            (
                self.environment_name,
                (UICommands.SET_ATTR, ("response_error_list", time_trac)),
            )
        )
        self.gui_update_queue.put(
            (
                self.environment_name,
                (
                    TransientUICommands.CONTROL_PREDICTIONS,
                    (
                        np.arange(self.environment_metadata.control_signal.shape[-1])
                        / self.hardware_metadata.sample_rate,
                        drive_signals,
                        self.predicted_response,
                        self.environment_metadata.control_signal,
                    ),
                ),
            )
        )

    def start_control(self, data: TransientInstructions):
        """Starts up the control to generate the signal"""
        if self.startup:
            self.test_level = db2scale(data.test_level)
            self.repeat = data.repeat
            self.gui_update_queue.put(
                (
                    self.environment_name,
                    (UICommands.SET_ENVIRONMENT_INSTRUCTIONS, data),
                )
            )
            self.log("Starting Environment")
            self.siggen_shutdown_achieved = False
            # Set up the signal generation
            self.queue_container.signal_generation_command_queue.put(
                self.environment_name,
                (
                    SignalGenerationCommands.INITIALIZE_PARAMETERS,
                    self.get_signal_generation_metadata(),
                ),
            )
            self.queue_container.signal_generation_command_queue.put(
                self.environment_name,
                (
                    SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
                    TransientSignalGenerator(self.next_drive, self.repeat),
                ),
            )
            self.queue_container.signal_generation_command_queue.put(
                self.environment_name,
                (SignalGenerationCommands.SET_TEST_LEVEL, self.test_level),
            )
            # Tell the signal generation to start generating signals
            self.queue_container.signal_generation_command_queue.put(
                self.environment_name, (SignalGenerationCommands.GENERATE_SIGNALS, None)
            )
            # Set up the measurement buffers
            n_control_channels = (
                len(self.environment_metadata.control_channel_indices)
                if self.environment_metadata.response_transformation_matrix is None
                else self.environment_metadata.response_transformation_matrix.shape[0]
            )
            n_output_channels = (
                len(self.environment_metadata.output_channel_indices)
                if self.environment_metadata.reference_transformation_matrix is None
                else self.environment_metadata.reference_transformation_matrix.shape[0]
            )
            self.control_buffer = FrameBuffer(
                n_control_channels,
                0,
                0,
                False,
                0,
                0,
                0,
                self.environment_metadata.control_signal.shape[-1],
                0,
                False,
                False,
                False,
                0,
                buffer_size_frame_multiplier=1
                + (
                    self.hardware_metadata.samples_per_read
                    * BUFFER_SIZE_SAMPLES_PER_READ_MULTIPLIER
                    / self.environment_metadata.control_signal.shape[-1]
                ),
                starting_value=0.0,
            )
            self.output_buffer = FrameBuffer(
                n_output_channels,
                0,
                0,
                False,
                0,
                0,
                0,
                self.environment_metadata.control_signal.shape[-1],
                0,
                False,
                False,
                False,
                0,
                buffer_size_frame_multiplier=1
                + (
                    self.hardware_metadata.samples_per_read
                    * BUFFER_SIZE_SAMPLES_PER_READ_MULTIPLIER
                    / self.environment_metadata.control_signal.shape[-1]
                ),
                starting_value=0.0,
            )
            self.startup = False
            self.set_active()
            self.gui_update_queue.put(
                (self.environment_name, (UICommands.ENVIRONMENT_STARTED, None))
            )
        # See if any data has come in
        try:
            acquisition_data, last_acquisition = (
                self.queue_container.data_in_queue.get_nowait()
            )
            if self.last_signal_found is not None:
                self.last_signal_found -= self.hardware_metadata.samples_per_read
            if last_acquisition:
                self.log(
                    f"Acquired Last Data, Signal Generation "
                    f"Shutdown Achieved: {self.siggen_shutdown_achieved}"
                )
            else:
                self.log("Acquired Data")
            scale_factor = 0.0 if self.test_level < 1e-10 else 1 / self.test_level
            control_data = (
                acquisition_data[self.environment_metadata.control_channel_indices]
                * scale_factor
            )
            if self.environment_metadata.response_transformation_matrix is not None:
                control_data = (
                    self.environment_metadata.response_transformation_matrix
                    @ control_data
                )
            output_data = (
                acquisition_data[self.environment_metadata.output_channel_indices]
                * scale_factor
            )
            if self.environment_metadata.reference_transformation_matrix is not None:
                output_data = (
                    self.environment_metadata.reference_transformation_matrix
                    @ output_data
                )
            # Add the data to the buffers
            self.control_buffer.add_data(control_data)
            self.output_buffer.add_data(output_data)
            if last_acquisition:
                # Find alignment with the specification via output
                self.log("Aligning signal with specification")
                (
                    self.aligned_output,
                    sample_delay,
                    phase_change,
                    _,
                ) = align_signals(
                    self.output_buffer[:],
                    self.next_drive[:, :: self.hardware_metadata.output_oversample],
                    correlation_threshold=0.5,
                )
            else:
                (
                    self.aligned_output,
                    sample_delay,
                    phase_change,
                    _,
                ) = (None, None, None, None)
            self.queue_container.gui_update_queue.put(
                (
                    self.environment_name,
                    (
                        TransientUICommands.TIME_DATA,
                        (control_data, output_data, sample_delay),
                    ),
                )
            )  # Sample_delay will be None if the alignment is not found
            if self.aligned_output is not None:
                self.log(f"Alignment Found at {sample_delay} samples")
                self.aligned_response = shift_signal(
                    self.control_buffer[:],
                    self.environment_metadata.control_signal.shape[-1],
                    sample_delay,
                    phase_change,
                )
                time_trac = trac(
                    self.aligned_response, self.environment_metadata.control_signal
                )
                self.gui_update_queue.put(
                    (
                        self.environment_name,
                        (
                            UICommands.SET_ATTR,
                            ("control_response_error_list", time_trac),
                        ),
                    )
                )
                self.queue_container.gui_update_queue.put(
                    (
                        self.environment_name,
                        (
                            TransientUICommands.CONTROL_DATA,
                            (self.aligned_response, self.aligned_output),
                        ),
                    )
                )
                # Do the next control
                self.log(
                    f"Last Signal Found: {self.last_signal_found}, "
                    f"Current Signal Found: {sample_delay}"
                )
                # We don't want to keep a signal if it starts during the last signal.
                # Multiply by 0.8 to give a little wiggle room in case the
                # last signal wasn't found exactly at the right place.
                if (
                    self.last_signal_found is None
                    or (
                        self.last_signal_found
                        + self.environment_metadata.control_signal.shape[-1] * 0.8
                    )
                    < sample_delay
                ):
                    self.next_drive = self.aligned_output
                    self.predicted_response = self.aligned_response
                    self.log("Computing next signal via control law")
                    self.perform_control_prediction(False)
                    self.last_signal_found = sample_delay
                else:
                    self.log("Signal was found previously, not controlling")
        except mp.queues.Empty:
            last_acquisition = False
        # See if we need to keep going
        if self.siggen_shutdown_achieved and last_acquisition:
            self.shutdown()
        else:
            self.queue_container.environment_command_queue.put(
                self.environment_name, (TransientCommands.START_CONTROL, None)
            )

    def stop_environment(self, data):
        """Starts the shutdown sequence based on commands from the UI"""
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name, (SignalGenerationCommands.START_SHUTDOWN, None)
        )

    # endregion

    # region Shutdown
    def shutdown(self):
        """Let the UI know that this environment has completely shut down"""
        self.log("Environment Shut Down")
        self.clear_active()
        self.gui_update_queue.put(
            (self.environment_name, (UICommands.ENVIRONMENT_ENDED, None))
        )
        self.startup = True

    # endregion


# region Process
def transient_process(
    environment_name: str,
    queue_name: str,
    input_queue: VerboseMessageQueue,
    gui_update_queue: mp.Queue,
    controller_command_queue: VerboseMessageQueue,
    log_file_queue: mp.Queue,
    data_in_queue: mp.Queue,
    data_out_queue: mp.Queue,
    acquisition_active_event: mp.synchronize.Event,
    output_active_event: mp.synchronize.Event,
    active_event: mp.synchronize.Event,
    ready_event: mp.synchronize.Event,
    shutdown_event: mp.synchronize.Event,
    sysid_active_event: mp.synchronize.Event,
    sysid_stored_event: mp.synchronize.Event,
    ping_alive_event: mp.synchronize.Event,
    threaded: bool,
):
    """
    Transient vibration environment process function called by multiprocessing

    This function defines the Transient Vibration Environment process that
    gets run by the multiprocessing module when it creates a new process.  It
    creates a TransientEnvironment object and runs it.

    Parameters
    ----------
    environment_name : str :
        Name of the environment
    input_queue : VerboseMessageQueue :
        Queue containing instructions for the environment
    gui_update_queue : Queue :
        Queue where GUI updates are put
    controller_communication_queue : Queue :
        Queue for global communications with the controller
    log_file_queue : Queue :
        Queue for writing log file messages
    data_in_queue : Queue :
        Queue from which data will be read by the environment
    data_out_queue : Queue :
        Queue to which data will be written that will be output by the hardware.
    acquisition_active : mp.sharedctypes.Synchronized
        A synchronized value that indicates when the acquisition is active
    output_active : mp.sharedctypes.Synchronized
        A synchronized value that indicates when the output is active
    """
    try:
        if threaded:
            new_process = threading.Thread  # worker threads
        else:
            new_process = mp.Process  # worker processes

        # Create vibration queues
        queue_container = TransientQueues(
            environment_name,
            input_queue,
            gui_update_queue,
            controller_command_queue,
            data_in_queue,
            data_out_queue,
            log_file_queue,
        )

        spectral_proc = new_process(
            target=spectral_processing_process,
            args=(
                environment_name,
                queue_container.spectral_command_queue,
                queue_container.data_for_spectral_computation_queue,
                queue_container.updated_spectral_quantities_queue,
                queue_container.environment_command_queue,
                queue_container.gui_update_queue,
                queue_container.log_file_queue,
            ),
        )
        spectral_proc.start()
        analysis_proc = new_process(
            target=sysid_data_analysis_process,
            args=(
                environment_name,
                queue_container.data_analysis_command_queue,
                queue_container.updated_spectral_quantities_queue,
                queue_container.time_history_to_generate_queue,
                queue_container.environment_command_queue,
                queue_container.gui_update_queue,
                queue_container.log_file_queue,
                ping_alive_event,
            ),
        )
        analysis_proc.start()
        siggen_proc = new_process(
            target=signal_generation_process,
            args=(
                environment_name,
                queue_container.signal_generation_command_queue,
                queue_container.time_history_to_generate_queue,
                queue_container.data_out_queue,
                queue_container.environment_command_queue,
                queue_container.log_file_queue,
                queue_container.gui_update_queue,
            ),
        )
        siggen_proc.start()
        collection_proc = new_process(
            target=data_collector_process,
            args=(
                environment_name,
                queue_container.collector_command_queue,
                queue_container.data_in_queue,
                [queue_container.data_for_spectral_computation_queue],
                queue_container.environment_command_queue,
                queue_container.log_file_queue,
                queue_container.gui_update_queue,
            ),
        )
        collection_proc.start()

        process_class = TransientEnvironment(
            environment_name,
            queue_name,
            queue_container,
            acquisition_active_event,
            output_active_event,
            active_event,
            ready_event,
            sysid_active_event,
            sysid_stored_event,
        )
        process_class.run(shutdown_event)

        # Rejoin all the processes
        process_class.log("Joining Subprocesses")
        process_class.log("Joining Spectral Computation")
        spectral_proc.join()
        process_class.log("Joining Data Analysis")
        analysis_proc.join()
        process_class.log("Joining Signal Generation")
        siggen_proc.join()
        process_class.log("Joining Data Collection")
        collection_proc.join()
    except Exception:  # pylint: disable = broad-exception-caught
        print(traceback.format_exc())


# endregion
