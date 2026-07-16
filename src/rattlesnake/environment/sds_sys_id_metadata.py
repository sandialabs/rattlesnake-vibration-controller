"""
This file defines the metadata that describes a MIMO Shock environment

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from enum import Enum

import netCDF4 as nc4
import numpy as np
from typing import List
import openpyxl

from rattlesnake.environment.abstract_sysid_environment import SysIdEnvironmentMetadata
from rattlesnake.environment.sds_sys_id_utilities import octspace, DecayedSineTable
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.process.abstract_sysid_data_analysis import SysIdMetadata


class ToneStrategy(Enum):
    """Enumeration specifying different ways of specifying sine tones"""

    FROM_SPEC = 0
    OCTAVE = 1
    MANUAL = 2


class ToneParameters:
    """Class to contain tone information"""

    def __init__(self, tone_strategy: ToneStrategy, tone_data: None | np.ndarray):
        """Initialize tone data information for storing with metadata

        Parameters
        ----------
        tone_strategy : ToneStrategy
            A tone strategy describing how to interpret the `tone_data` argument.
        tone_data : None | np.ndarray
            If `tone_strategy` is FROM_SPEC, this should be None and will be ignored.
            If `tone_stragegy` is OCTAVE, then this should be a size 3 NumPy array with values
            [min_frequency, max_frequency, tones_per_octave].
            If `tone_strategy` is MANUAL, then this should be a 1D NumPy array consisting of
            the individual sine tones, not including the compensation pulse if used.
        """
        if not isinstance(tone_strategy, ToneStrategy):
            raise ValueError("`tone_strategy` must be one of the `ToneStrategy` enumeration values")
        self.tone_strategy = tone_strategy
        if self.tone_strategy == ToneStrategy.FROM_SPEC:
            self.tone_data = None
        elif self.tone_strategy == ToneStrategy.OCTAVE:
            tone_data = np.array(tone_data).flatten()
            if tone_data.size != 3:
                raise ValueError("`tone_data` must be length-3 for `OCTAVE` tone strategy.")
            self.tone_data = tone_data
        else:
            self.tone_data = np.array(tone_data).flatten()


class CompPulseParameters:
    """Class to contain compenstation pulse data"""

    def __init__(
        self,
        use_compensation_pulse: bool,
        compensation_frequency: None | float = None,
        compensation_decay: None | float = None,
    ):
        """Initialize compensation pulse information for storing with metadata

        Parameters
        ----------
        use_compensation_pulse : bool
            True if a compensation pulse is used.
        compensation_frequency : None | float
            The frequency at which the compensation tone is defined.
            If None, it will be selected automatically.   Will be ignored if not used.
        compensation_decay : None | float
            The compensation decay as a fraction (0.95) rather than a percentage (95%).  Can be
            None if a compensation pulse is not used.  Will be ignored if not used.
        """
        self.use_compensation_pulse = use_compensation_pulse
        if not self.use_compensation_pulse:
            compensation_frequency = None
            compensation_decay = None
        self.compensation_frequency = compensation_frequency
        if compensation_decay is None and use_compensation_pulse:
            raise ValueError(
                "`compensation_decay` must be specified if a compensation pulse is used."
            )
        self.compensation_decay = compensation_decay


class DecayStrategy(Enum):
    """Enumeration containing different ways to define decay in a sum of decayed sines table"""

    DAMPING = 0
    TIME_CONSTANT = 1
    NUM_TIME_CONSTANTS = 2


class DecayParameters:
    """A class to store data defining how the sine terms decay"""

    def __init__(
        self, decay_strategy: DecayStrategy, common_decay: bool, decay_data: float | np.ndarray
    ):
        """Initializes the decay data

        Parameters
        ----------
        decay_strategy : DecayStrategy
            The strategy used to define the decay parameters in the sum of decayed sine table
        common_decay : bool
            If True, a single decay parameter is used for all tones.  If False, one must be
            specified for each sine tone
        decay_data : float | np.ndarray
            If `common_decay` is True, this should be a single floating point number defining the
            decay value.  If `common_decay` is False, this should be a np.ndarray with the same
            number of values as sine tone frequencies in the environment.
        """
        self.decay_strategy = decay_strategy
        self.common_decay = common_decay
        self.decay_data = np.array(decay_data).flatten()


def convert_decay_strategy(
    old_values: np.ndarray,
    frequencies: np.ndarray,
    block_length: float,
    old_strategy: DecayStrategy,
    new_strategy: DecayStrategy,
):
    """Convert between different decay strategies

    Parameters
    ----------
    old_values : np.ndarray
        Damping values defined in the stratgy given in `old_strategy`
    frequencies : np.ndarray
        Frequencies (in Hz) corresponding to the damping values in `old_values`
    block_length : float
        Length of the time block (in seconds) over which the number of time constants will be
        evaluated
    old_strategy : DecayStrategy
        The decay strategy in which the the `old_values` are defined.
    new_strategy : DecayStrategy
        The new strategy in which the decay values will be returned.

    Returns
    -------
    new_values : np.ndarray
        Decay values defined in the form specified by `new_strategy`

    Raises
    ------
    ValueError
        If invalid decay strategies are provided
    """
    if old_strategy == new_strategy:
        # If the strategies are the same, return the old values directly
        return old_values

    # Convert frequencies to angular frequencies (omega)
    omega = 2 * np.pi * frequencies

    # Conversion logic
    if old_strategy == DecayStrategy.TIME_CONSTANT:
        tau = old_values
        if new_strategy == DecayStrategy.DAMPING:
            # Convert tau to zeta
            zeta = 1 / (tau * omega)
            return zeta
        if new_strategy == DecayStrategy.NUM_TIME_CONSTANTS:
            # Convert tau to number of time constants per block
            num_time_constants = block_length / tau
            return num_time_constants

    elif old_strategy == DecayStrategy.DAMPING:
        zeta = old_values
        tau = 1 / (zeta * omega)
        if new_strategy == DecayStrategy.TIME_CONSTANT:
            # Convert zeta to tau
            return tau
        if new_strategy == DecayStrategy.NUM_TIME_CONSTANTS:
            # Convert zeta to number of time constants per block
            num_time_constants = block_length / tau
            return num_time_constants

    elif old_strategy == DecayStrategy.NUM_TIME_CONSTANTS:
        num_time_constants = old_values
        tau = block_length / num_time_constants
        if new_strategy == DecayStrategy.TIME_CONSTANT:
            # Convert number of time constants per block to tau
            return tau
        if new_strategy == DecayStrategy.DAMPING:
            # Convert number of time constants per block to zeta
            zeta = 1 / (tau * omega)
            return zeta

    # If no valid conversion is found, raise an error
    raise ValueError("Invalid conversion between damping strategies.")


class SRSType(Enum):
    """Enumeration containing different ways to compute an SRS"""

    PRIMARY_POS = 1
    PRIMARY_NEG = 2
    PRIMARY_ABSMAX = 3
    RESIDUAL_POS = 4
    RESIDUAL_NEG = 5
    RESIDUAL_ABSMAX = 6
    MAXIMUM_POS = 7
    MAXIMUM_NEG = 8
    MAXIMUM_ABSMAX = 9


class SRSDisplacementType(Enum):
    """Enumeration containing different types of displacement to compute SRSs from"""

    ABSOLUTE = 1
    RELATIVE = -1


class SRSParameters:
    """Class defining parameters for how SRSs are computed."""

    def __init__(
        self, srs_type: SRSType, srs_displacement: SRSDisplacementType, srs_damping: float
    ):
        """Initializes an object to store SRS computation parameters

        Parameters
        ----------
        srs_type : SRSType
            The method for computing maximum response for the SRS
        srs_displacement : SRSDisplacementType
            The response type to compute maximums from
        srs_damping : float
            The damping value used for the single-degree-of-freedom oscilators in the SRS.
        """
        self.srs_type = srs_type
        self.srs_displacement = srs_displacement
        self.srs_damping = srs_damping


class SDSParameters:
    """Class defining how sum-of-decayed-sine tables are computed"""

    def __init__(
        self, iterations: int, convergence: float, scale_factor: float, error_tolerance: float
    ):
        self.iterations = iterations
        self.convergence = convergence
        self.scale_factor = scale_factor
        self.error_tolerance = error_tolerance


class SpecParameters:
    """A class for storing SRS values that define the environment specification"""

    def __init__(
        self,
        frequencies: np.ndarray,
        srs_spec: np.ndarray,
        srs_lower_limit: np.ndarray,
        srs_upper_limit: np.ndarray,
        num_hits: int,
    ):
        """Initializes specification parameters

        Parameters
        ----------
        frequencies : np.ndarray
            The frequency tones defined in the specification.  This should be a 1D numpy array
            with each entry representing a different frequency value.
        srs_spec : np.ndarray
            The values of the SRS at each frequency defined in the `frequencies` array for each
            of the control degrees of freedom (physical or virtual) in the environment.  This should
            be 2-dimensional, with the rows equal to the number of frequencies and the columns equal
            to the number of control channels.  If control is not required for specific sine tones
            or channels, a NaN can be placed at that row and column.
        srs_lower_limit : np.ndarray
            The values of the lower limit SRS at each frequency defined in the `frequencies` array
            for each
            of the control degrees of freedom (physical or virtual) in the environment.  This should
            be 2-dimensional, with the rows equal to the number of frequencies and the columns equal
            to the number of control channels.  If a limit is not required for specific sine tones
            or channels, a NaN can be placed at that row and column.
        srs_upper_limit : np.ndarray
            The values of the upper limit SRS at each frequency defined in the `frequencies` array
            for each
            of the control degrees of freedom (physical or virtual) in the environment.  This should
            be 2-dimensional, with the rows equal to the number of frequencies and the columns equal
            to the number of control channels.  If a limit is not required for specific sine tones
            or channels, a NaN can be placed at that row and column.
        num_hits : int
            The number of hits to apply to the test article in the environment.
        """
        self.frequencies = frequencies
        self.srs_spec = srs_spec
        self.srs_lower_limit = srs_lower_limit
        self.srs_upper_limit = srs_upper_limit
        self.num_hits = num_hits


class ControlLawType(Enum):
    """Enumeration containing acceptable types of objects to use for a control law"""

    FUNCTION = 0
    CLASS = 1
    INTERACTIVE_CLASS = 2


class ControlParameters:
    """Class to store control law data"""

    def __init__(
        self,
        control_script: str | None,
        control_object: str | None,
        control_type: ControlLawType | None,
        control_parameters: dict | None,
    ):
        """Initializes an object to store information about the custom control law

        Parameters
        ----------
        control_script : str
            The path to the python script containing the control law
        control_object : str
            The name of the item (function, generator, class) containing the control law
        control_type : ControlLawType
            The type of item defining the control law
        control_parameters : dict
            Any extra keyword arguments to pass to the control law.
        """
        self.control_script = control_script
        self.control_object = control_object
        self.control_type = control_type
        self.control_parameters = control_parameters


def _parse_scalar_string(value):
    if isinstance(value, (int, float, bool)):
        return value
    if value is None:
        return None
    if not isinstance(value, str):
        return value

    text = value.strip()
    if text == "":
        return ""

    if text.lower() in ("true", "y", "yes"):
        return True
    if text.lower() in ("false", "n", "no"):
        return False

    try:
        return int(text)
    except ValueError:
        pass

    try:
        return float(text)
    except ValueError:
        pass

    return text


class SDSMetadata(SysIdEnvironmentMetadata):
    """Metadata required to define a Shock control law in rattlesnake."""

    def __init__(
        self,
        *,
        environment_name: str,
        channel_list_bools: list,
        sample_rate: int,
        num_channels: int,
        block_size: int,
        tone_data: ToneParameters,
        compensation_pulse_data: CompPulseParameters,
        decay_data: DecayParameters,
        srs_data: SRSParameters,
        sds_data: SDSParameters,
        control_script_data: ControlParameters,
        control_channel_indices: np.ndarray,
        output_channel_indices: np.ndarray,
        response_transformation_matrix: None | np.ndarray,
        excitation_transformation_matrix: None | np.ndarray,
        specification_data: SpecParameters,
        sysid_metadata=None,
    ):
        super().__init__(
            EnvironmentType.SDS,
            environment_name,
            channel_list_bools,
            sample_rate,
            sysid_metadata,
        )
        self.block_size = block_size
        self.number_of_channels = num_channels
        self.compensation_pulse_data = compensation_pulse_data
        self.decay_data = decay_data
        self.srs_data = srs_data
        self.sds_data = sds_data
        self.tone_data = tone_data
        self.sample_rate = sample_rate
        self.control_script_data = control_script_data
        self.control_channel_indices = control_channel_indices
        self.output_channel_indices = output_channel_indices
        self.response_transformation_matrix = response_transformation_matrix
        self.reference_transformation_matrix = excitation_transformation_matrix
        self.specification_data = specification_data

    @property
    def number_of_channels(self):
        """Total number of channels in the environment"""
        return self._number_of_channels

    @number_of_channels.setter
    def number_of_channels(self, value):
        """Sets the total number of channels in the environment"""
        self._number_of_channels = value

    @property
    def response_channel_indices(self):
        """Indices identifying which channels are control channels"""
        return self.control_channel_indices

    @property
    def reference_channel_indices(self):
        """Indices identifying which channels are reference or excitation channels"""
        return self.output_channel_indices

    @property
    def response_transformation_matrix(self):
        """Transformation matrix applied to the control channels"""
        return self._response_transformation_matrix

    @response_transformation_matrix.setter
    def response_transformation_matrix(self, value):
        """Sets the transformation matrix for the control channels"""
        self._response_transformation_matrix = value

    @property
    def reference_transformation_matrix(self):
        """Transformation matrix applied to the excitation channels"""
        return self._reference_transformation_matrix

    @reference_transformation_matrix.setter
    def reference_transformation_matrix(self, value):
        """Sets the transformation matrix applied to the excitation channels"""
        self._reference_transformation_matrix = value

    @property
    def sample_rate(self):
        """Gets the sample rate of the data acquisition system"""
        return self._sample_rate

    @sample_rate.setter
    def sample_rate(self, value):
        """Sets the sample rate of the data acquisition system"""
        self._sample_rate = value

    def get_compensation_pulse_frequency(self):
        if self.compensation_pulse_data.compensation_frequency is None:
            return min(self.get_sds_frequencies()) / 3
        else:
            return self.compensation_pulse_data.compensation_frequency

    def get_sds_frequencies(self):
        if self.tone_data.tone_strategy == ToneStrategy.FROM_SPEC:
            return self.specification_data.frequencies
        if self.tone_data.tone_strategy == ToneStrategy.OCTAVE:
            return octspace(*self.tone_data.tone_data)
        if self.tone_data.tone_strategy == ToneStrategy.MANUAL:
            return self.tone_data.tone_data

    def get_sds_decays(self):
        frequencies = self.get_sds_frequencies()
        if self.decay_data.common_decay:
            decay_values = self.decay_data.decay_data[0] * np.ones(len(frequencies))
        else:
            decay_values = self.decay_data.decay_data
        return convert_decay_strategy(
            decay_values,
            frequencies,
            self.block_size / self.sample_rate,
            self.decay_data.decay_strategy,
            DecayStrategy.DAMPING,
        )

    def get_sds_frequencies_w_compensation_pulse(self):
        if self.tone_data.tone_strategy == ToneStrategy.FROM_SPEC:
            frequencies = self.specification_data.frequencies
        if self.tone_data.tone_strategy == ToneStrategy.OCTAVE:
            frequencies = octspace(*self.tone_data.tone_data)
        if self.tone_data.tone_strategy == ToneStrategy.MANUAL:
            frequencies = self.tone_data.tone_data
        if self.compensation_pulse_data.use_compensation_pulse:
            frequencies = np.concatenate((frequencies, [self.get_compensation_pulse_frequency()]))
        return frequencies

    def get_sds_decays_w_compensation_pulse(self):
        frequencies = self.get_sds_frequencies()
        if self.decay_data.common_decay:
            decay_values = self.decay_data.decay_data[0] * np.ones(len(frequencies))
        else:
            decay_values = self.decay_data.decay_data
        decay_values = convert_decay_strategy(
            decay_values,
            frequencies,
            self.block_size / self.sample_rate,
            self.decay_data.decay_strategy,
            DecayStrategy.DAMPING,
        )
        if self.compensation_pulse_data.use_compensation_pulse:
            decay_values = np.concatenate(
                (decay_values, [self.compensation_pulse_data.compensation_decay])
            )
        return decay_values

    @classmethod
    def create_blank_worksheet_template(cls, worksheet):
        super().create_blank_worksheet_template(worksheet)

        worksheet.cell(1, 2, "SDS")

        worksheet.cell(2, 1, "Block Size")
        worksheet.cell(2, 3, "# Number of samples in one SDS hit block")

        worksheet.cell(3, 1, "Tone Strategy")
        worksheet.cell(3, 3, '# One of "From Spec", "Octave", or "Manual"')

        worksheet.cell(4, 1, "Tone Data")
        worksheet.cell(
            4,
            3,
            '# If "From Spec", leave blank. If "Octave", fill columns as min_freq, max_freq, tones/octave. '
            'If "Manual", fill columns with tone frequencies.',
        )

        worksheet.cell(5, 1, "Use Compensation Pulse")
        worksheet.cell(5, 3, "# Y or N")

        worksheet.cell(6, 1, "Compensation Frequency")
        worksheet.cell(6, 3, '# Frequency in Hz or "Auto"')

        worksheet.cell(7, 1, "Compensation Decay")
        worksheet.cell(7, 3, "# Compensation decay as fraction (e.g. 0.95)")

        worksheet.cell(8, 1, "Decay Strategy")
        worksheet.cell(
            8,
            3,
            '# One of "Damping", "Time Constant", or "Num Time Constants"',
        )

        worksheet.cell(9, 1, "Common Decay")
        worksheet.cell(9, 3, "# Y or N")

        worksheet.cell(10, 1, "Decay Data")
        worksheet.cell(
            10,
            3,
            "# If common decay, put one value in column 2. Otherwise fill one value per tone across columns.",
        )

        worksheet.cell(11, 1, "SRS Type")
        worksheet.cell(11, 3, "# Name of SRS type enumeration")

        worksheet.cell(12, 1, "SRS Displacement")
        worksheet.cell(12, 3, '# One of "Absolute" or "Relative"')

        worksheet.cell(13, 1, "SRS Damping")
        worksheet.cell(13, 3, "# Fractional damping, e.g. 0.03")

        worksheet.cell(14, 1, "SDS Iterations")
        worksheet.cell(14, 3, "# Number of SDS synthesis iterations")

        worksheet.cell(15, 1, "SDS Convergence")
        worksheet.cell(15, 3, "# Fractional convergence, e.g. 0.8")

        worksheet.cell(16, 1, "SDS Scale Factor")
        worksheet.cell(16, 3, "# Scale factor, e.g. 1.02")

        worksheet.cell(17, 1, "SDS Error Tolerance")
        worksheet.cell(17, 3, "# Fractional error tolerance, e.g. 0.05")

        worksheet.cell(18, 1, "Specification File")
        worksheet.cell(18, 3, "# Path to .npz specification file to load")

        worksheet.cell(19, 1, "Control Python Script")
        worksheet.cell(19, 3, "# Path to control law Python script")

        worksheet.cell(20, 1, "Control Python Function/Object")
        worksheet.cell(20, 3, "# Function or class name in script")

        worksheet.cell(21, 1, "Control Python Type")
        worksheet.cell(
            21,
            3,
            '# One of "Function", "Class", or "Interactive"',
        )

        worksheet.cell(22, 1, "Control Channels (1-based)")
        worksheet.cell(22, 3, "# List of channels, one per cell on this row")

        worksheet.cell(23, 1, "Control Parameters")
        worksheet.cell(
            23,
            3,
            '# key=value pairs, one per cell, e.g. "rcond=1e-10"',
        )

        worksheet.cell(24, 1, "Specification Num Hits")
        worksheet.cell(24, 3, "# Integer number of target hits")

        SysIdMetadata.create_blank_worksheet_template(worksheet, start_row=26)

        worksheet.cell(42, 1, "Response Transformation Matrix:")
        worksheet.cell(
            42,
            2,
            "# Type None if not used, otherwise put matrix values starting in column 2",
        )

        worksheet.cell(43, 1, "Output Transformation Matrix:")
        worksheet.cell(
            43,
            2,
            "# Type None if not used, otherwise put matrix values starting in column 2",
        )

    @classmethod
    def load_metadata_from_worksheet(
        cls,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        sample_rate = hardware_metadata.sample_rate
        num_channels = sum(channel_list_bools)

        environment_channel_list = [
            channel
            for channel, channel_bool in zip(hardware_metadata.channel_list, channel_list_bools)
            if channel_bool
        ]

        output_channel_indices = [
            index
            for index, channel in enumerate(environment_channel_list)
            if channel.feedback_device is not None
        ]

        block_size = int(worksheet.cell(2, 2).value)

        # Tone strategy
        tone_strategy_text = str(worksheet.cell(3, 2).value).strip().lower()
        if tone_strategy_text in ["from spec", "from specification", "from_spec"]:
            tone_data = ToneParameters(ToneStrategy.FROM_SPEC, None)
        elif tone_strategy_text in ["octave", "oct"]:
            octave_values = []
            col = 2
            while True:
                value = worksheet.cell(4, col).value
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    break
                octave_values.append(float(value))
                col += 1
            tone_data = ToneParameters(ToneStrategy.OCTAVE, np.array(octave_values))
        elif tone_strategy_text in ["manual"]:
            manual_values = []
            col = 2
            while True:
                value = worksheet.cell(4, col).value
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    break
                manual_values.append(float(value))
                col += 1
            tone_data = ToneParameters(ToneStrategy.MANUAL, np.array(manual_values))
        else:
            raise ValueError(f"Unknown Tone Strategy {worksheet.cell(3, 2).value}")

        # Compensation pulse
        use_comp = str(worksheet.cell(5, 2).value).strip().upper() == "Y"
        comp_freq_raw = worksheet.cell(6, 2).value
        if isinstance(comp_freq_raw, str) and comp_freq_raw.strip().lower() == "auto":
            compensation_frequency = None
        elif comp_freq_raw is None or comp_freq_raw == "":
            compensation_frequency = None
        else:
            compensation_frequency = float(comp_freq_raw)

        comp_decay_raw = worksheet.cell(7, 2).value
        compensation_decay = None if comp_decay_raw in (None, "") else float(comp_decay_raw)

        compensation_pulse_data = CompPulseParameters(
            use_compensation_pulse=use_comp,
            compensation_frequency=compensation_frequency,
            compensation_decay=compensation_decay,
        )

        # Decay
        decay_strategy_text = str(worksheet.cell(8, 2).value).strip().lower().replace("_", " ")
        if decay_strategy_text in ["damping", "zeta"]:
            decay_strategy = DecayStrategy.DAMPING
        elif decay_strategy_text in ["time constant", "tau", "time const"]:
            decay_strategy = DecayStrategy.TIME_CONSTANT
        elif decay_strategy_text in ["num time constants", "number of time constants", "ntc"]:
            decay_strategy = DecayStrategy.NUM_TIME_CONSTANTS
        else:
            raise ValueError(f"Unknown Decay Strategy {worksheet.cell(8, 2).value}")

        common_decay = str(worksheet.cell(9, 2).value).strip().upper() == "Y"

        decay_values = []
        col = 2
        while True:
            value = worksheet.cell(10, col).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                break
            decay_values.append(float(value))
            col += 1
        decay_data = np.array(decay_values)

        decay_parameters = DecayParameters(decay_strategy, common_decay, decay_data)

        # SRS parameters
        srs_type = SRSType[str(worksheet.cell(11, 2).value).strip().upper()]
        srs_disp_text = str(worksheet.cell(12, 2).value).strip().lower()
        if srs_disp_text == "absolute":
            srs_displacement = SRSDisplacementType.ABSOLUTE
        elif srs_disp_text == "relative":
            srs_displacement = SRSDisplacementType.RELATIVE
        else:
            raise ValueError(f"Unknown SRS displacement {worksheet.cell(12, 2).value}")

        srs_damping = float(worksheet.cell(13, 2).value)
        srs_data = SRSParameters(srs_type, srs_displacement, srs_damping)

        # SDS synthesis parameters
        sds_data = SDSParameters(
            iterations=int(worksheet.cell(14, 2).value),
            convergence=float(worksheet.cell(15, 2).value),
            scale_factor=float(worksheet.cell(16, 2).value),
            error_tolerance=float(worksheet.cell(17, 2).value),
        )

        # Specification file
        spec_filename = worksheet.cell(18, 2).value
        if spec_filename is None or (
            isinstance(spec_filename, str) and spec_filename.strip() == ""
        ):
            raise ValueError("SDS worksheet requires a specification file path in row 18 column 2.")

        spec_data = np.load(spec_filename)
        specification_data = SpecParameters(
            frequencies=spec_data["f"],
            srs_spec=spec_data["srs"],
            srs_lower_limit=spec_data["lower_limit"],
            srs_upper_limit=spec_data["upper_limit"],
            num_hits=int(spec_data["num_hits"]),
        )

        # Control data
        control_script = worksheet.cell(19, 2).value
        control_object = worksheet.cell(20, 2).value
        control_type_text = str(worksheet.cell(21, 2).value).strip().lower()
        if control_type_text == "function":
            control_type = ControlLawType.FUNCTION
        elif control_type_text == "class":
            control_type = ControlLawType.CLASS
        elif control_type_text == "interactive":
            control_type = ControlLawType.INTERACTIVE_CLASS
        else:
            raise ValueError(f"Unknown control type {worksheet.cell(21, 2).value}")

        control_channel_indices = []
        col = 2
        while True:
            value = worksheet.cell(22, col).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                break
            control_channel_indices.append(int(value) - 1)
            col += 1
        control_channel_indices = np.array(control_channel_indices, dtype=int)

        control_parameters = {}
        col = 2
        while True:
            value = worksheet.cell(23, col).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                break
            text = str(value)
            if "=" not in text:
                raise ValueError(f'Invalid control parameter entry "{text}", expected key=value')
            key, raw_val = text.split("=", 1)
            control_parameters[key.strip()] = _parse_scalar_string(raw_val.strip())
            col += 1

        control_script_data = ControlParameters(
            control_script=control_script,
            control_object=control_object,
            control_type=control_type,
            control_parameters=control_parameters,
        )

        # Override num_hits from worksheet if desired
        num_hits_cell = worksheet.cell(24, 2).value
        if num_hits_cell is not None and str(num_hits_cell).strip() != "":
            specification_data.num_hits = int(num_hits_cell)

        sysid_metadata = SysIdMetadata.load_metadata_from_worksheet(
            worksheet, hardware_metadata, start_row=26
        )

        response_transformation_matrix, output_transformation_matrix = (
            cls.load_sysid_matrix_from_worksheet(worksheet, start_row=42)
        )

        return cls(
            environment_name=environment_name,
            channel_list_bools=channel_list_bools,
            sample_rate=sample_rate,
            num_channels=num_channels,
            block_size=block_size,
            tone_data=tone_data,
            compensation_pulse_data=compensation_pulse_data,
            decay_data=decay_parameters,
            srs_data=srs_data,
            sds_data=sds_data,
            control_script_data=control_script_data,
            control_channel_indices=control_channel_indices,
            output_channel_indices=np.array(output_channel_indices, dtype=int),
            response_transformation_matrix=response_transformation_matrix,
            excitation_transformation_matrix=output_transformation_matrix,
            specification_data=specification_data,
            sysid_metadata=sysid_metadata,
        )

    def save_metadata_to_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        super().save_metadata_to_worksheet(worksheet)

        worksheet.cell(2, 2, self.block_size)

        tone_strategy_names = {
            ToneStrategy.FROM_SPEC: "From Spec",
            ToneStrategy.OCTAVE: "Octave",
            ToneStrategy.MANUAL: "Manual",
        }
        worksheet.cell(3, 2, tone_strategy_names[self.tone_data.tone_strategy])

        if self.tone_data.tone_strategy == ToneStrategy.OCTAVE:
            for i, value in enumerate(self.tone_data.tone_data):
                worksheet.cell(4, 2 + i, float(value))
        elif self.tone_data.tone_strategy == ToneStrategy.MANUAL:
            for i, value in enumerate(self.tone_data.tone_data):
                worksheet.cell(4, 2 + i, float(value))

        worksheet.cell(5, 2, "Y" if self.compensation_pulse_data.use_compensation_pulse else "N")
        worksheet.cell(
            6,
            2,
            (
                "Auto"
                if self.compensation_pulse_data.compensation_frequency is None
                else self.compensation_pulse_data.compensation_frequency
            ),
        )
        if self.compensation_pulse_data.compensation_decay is not None:
            worksheet.cell(7, 2, self.compensation_pulse_data.compensation_decay)

        decay_strategy_names = {
            DecayStrategy.DAMPING: "Damping",
            DecayStrategy.TIME_CONSTANT: "Time Constant",
            DecayStrategy.NUM_TIME_CONSTANTS: "Num Time Constants",
        }
        worksheet.cell(8, 2, decay_strategy_names[self.decay_data.decay_strategy])
        worksheet.cell(9, 2, "Y" if self.decay_data.common_decay else "N")

        if self.decay_data.common_decay:
            worksheet.cell(10, 2, float(self.decay_data.decay_data[0]))
        else:
            for i, value in enumerate(self.decay_data.decay_data):
                worksheet.cell(10, 2 + i, float(value))

        worksheet.cell(11, 2, self.srs_data.srs_type.name)
        worksheet.cell(12, 2, self.srs_data.srs_displacement.name.title())
        worksheet.cell(13, 2, self.srs_data.srs_damping)

        worksheet.cell(14, 2, self.sds_data.iterations)
        worksheet.cell(15, 2, self.sds_data.convergence)
        worksheet.cell(16, 2, self.sds_data.scale_factor)
        worksheet.cell(17, 2, self.sds_data.error_tolerance)

        # Intentionally leave spec file blank when saving template
        worksheet.cell(18, 2, "")
        worksheet.cell(18, 3, "# Required for loading: path to .npz specification file")

        worksheet.cell(19, 2, self.control_script_data.control_script)
        worksheet.cell(20, 2, self.control_script_data.control_object)
        worksheet.cell(21, 2, self.control_script_data.control_type.name.title().replace("_", " "))

        for idx, channel_ind in enumerate(self.control_channel_indices):
            worksheet.cell(22, 2 + idx, int(channel_ind) + 1)

        if self.control_script_data.control_parameters is not None:
            for idx, (key, value) in enumerate(self.control_script_data.control_parameters.items()):
                worksheet.cell(23, 2 + idx, f"{key}={value}")

        worksheet.cell(24, 2, int(self.specification_data.num_hits))

        self.sysid_metadata.save_metadata_to_worksheet(worksheet, start_row=26)

        self.save_sysid_matrix_to_worksheet(
            worksheet,
            self.response_transformation_matrix,
            self.reference_transformation_matrix,
            start_row=42,
        )

    @classmethod
    def load_metadata_from_netcdf(
        cls,
        netcdf_group_handle: nc4._netCDF4.Group,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        sample_rate = hardware_metadata.sample_rate
        num_channels = sum(channel_list_bools)
        block_size = int(netcdf_group_handle.block_size)

        environment_channel_list = [
            channel
            for channel, channel_bool in zip(hardware_metadata.channel_list, channel_list_bools)
            if channel_bool
        ]

        output_channel_indices = [
            index
            for index, channel in enumerate(environment_channel_list)
            if channel.feedback_device is not None
        ]

        # ----------------------------
        # Tone parameters
        # ----------------------------
        tone_grp = netcdf_group_handle.groups["tone_parameters"]
        tone_strategy = ToneStrategy(tone_grp.strategy)
        if "tone_data" in tone_grp.variables:
            tone_values = np.array(tone_grp.variables["tone_data"][...]).flatten()
        else:
            tone_values = None
        tone_data = ToneParameters(tone_strategy, tone_values)

        # ----------------------------
        # Compensation pulse parameters
        # ----------------------------
        comp_grp = netcdf_group_handle.groups["compensation_pulse_parameters"]
        use_compensation_pulse = bool(comp_grp.use_compensation_pulse)
        compensation_frequency = (
            comp_grp.compensation_frequency if hasattr(comp_grp, "compensation_frequency") else None
        )
        compensation_decay = (
            comp_grp.compensation_decay if hasattr(comp_grp, "compensation_decay") else None
        )
        compensation_pulse_data = CompPulseParameters(
            use_compensation_pulse=use_compensation_pulse,
            compensation_frequency=compensation_frequency,
            compensation_decay=compensation_decay,
        )

        # ----------------------------
        # Decay parameters
        # ----------------------------
        decay_grp = netcdf_group_handle.groups["decay_parameters"]
        decay_strategy = DecayStrategy(decay_grp.decay_strategy)
        common_decay = bool(decay_grp.common_decay)
        if "decay_data" in decay_grp.variables:
            decay_values = np.array(decay_grp.variables["decay_data"][...]).flatten()
        else:
            decay_values = np.array([decay_grp.decay_data], dtype=float)
        decay_data = DecayParameters(decay_strategy, common_decay, decay_values)

        # ----------------------------
        # SRS parameters
        # ----------------------------
        srs_grp = netcdf_group_handle.groups["srs_parameters"]
        srs_data = SRSParameters(
            SRSType(srs_grp.srs_type),
            SRSDisplacementType(srs_grp.srs_displacement),
            float(srs_grp.srs_damping),
        )

        # ----------------------------
        # SDS parameters
        # ----------------------------
        sds_grp = netcdf_group_handle.groups["sds_parameters"]
        sds_data = SDSParameters(
            iterations=int(sds_grp.iterations),
            convergence=float(sds_grp.convergence),
            scale_factor=float(sds_grp.scale_factor),
            error_tolerance=float(sds_grp.error_tolerance),
        )

        # ----------------------------
        # Specification parameters
        # ----------------------------
        spec_grp = netcdf_group_handle.groups["specification_parameters"]
        specification_data = SpecParameters(
            frequencies=np.array(spec_grp.variables["frequencies"][...]),
            srs_spec=np.array(spec_grp.variables["srs_spec"][...]),
            srs_lower_limit=np.array(spec_grp.variables["srs_lower_limit"][...]),
            srs_upper_limit=np.array(spec_grp.variables["srs_upper_limit"][...]),
            num_hits=int(spec_grp.num_hits),
        )

        # ----------------------------
        # Control parameters
        # ----------------------------
        control_grp = netcdf_group_handle.groups["control_parameters"]
        control_type = ControlLawType(control_grp.control_type)
        control_script = control_grp.control_script
        control_object = control_grp.control_object

        control_parameters = {}
        if "control_extra_parameters" in control_grp.groups:
            control_grp_params = control_grp.groups["control_extra_parameters"]
            for key in control_grp_params.ncattrs():
                control_parameters[key] = getattr(control_grp_params, key)

        control_script_data = ControlParameters(
            control_script=control_script,
            control_object=control_object,
            control_type=control_type,
            control_parameters=control_parameters,
        )

        # ----------------------------
        # Control channel indices
        # ----------------------------
        control_channel_indices = np.array(
            netcdf_group_handle.variables["control_channel_indices"][...],
            dtype=int,
        )

        # ----------------------------
        # Transformation matrices
        # ----------------------------
        response_transformation_matrix = None
        if "response_transformation_matrix" in netcdf_group_handle.variables:
            response_transformation_matrix = np.array(
                netcdf_group_handle.variables["response_transformation_matrix"][...]
            )

        excitation_transformation_matrix = None
        if "reference_transformation_matrix" in netcdf_group_handle.variables:
            excitation_transformation_matrix = np.array(
                netcdf_group_handle.variables["reference_transformation_matrix"][...]
            )

        # ----------------------------
        # SysID metadata
        # ----------------------------
        sysid_metadata = SysIdMetadata.load_metadata_from_netcdf(
            netcdf_group_handle,
            hardware_metadata,
        )

        return cls(
            environment_name=environment_name,
            channel_list_bools=channel_list_bools,
            sample_rate=sample_rate,
            num_channels=num_channels,
            block_size=block_size,
            tone_data=tone_data,
            compensation_pulse_data=compensation_pulse_data,
            decay_data=decay_data,
            srs_data=srs_data,
            sds_data=sds_data,
            control_script_data=control_script_data,
            control_channel_indices=control_channel_indices,
            output_channel_indices=np.array(output_channel_indices, dtype=int),
            response_transformation_matrix=response_transformation_matrix,
            excitation_transformation_matrix=excitation_transformation_matrix,
            specification_data=specification_data,
            sysid_metadata=sysid_metadata,
        )

    def save_metadata_to_netcdf(
        self,
        netcdf_group_handle: nc4._netCDF4.Group,  # pylint: disable=c-extension-no-member
    ):
        super().save_metadata_to_netcdf(netcdf_group_handle)
        netcdf_group_handle.block_size = self.block_size
        # Create groups for different portions of the metadata
        tone_grp = netcdf_group_handle.createGroup("tone_parameters")
        decay_grp = netcdf_group_handle.createGroup("decay_parameters")
        srs_grp = netcdf_group_handle.createGroup("srs_parameters")
        sds_grp = netcdf_group_handle.createGroup("sds_parameters")
        comp_grp = netcdf_group_handle.createGroup("compensation_pulse_parameters")
        control_grp = netcdf_group_handle.createGroup("control_parameters")
        spec_grp = netcdf_group_handle.createGroup("specification_parameters")
        # Tone group
        tone_grp.strategy = self.tone_data.tone_strategy.value
        if self.tone_data.tone_data is not None:
            tone_grp.createDimension("tone_data_size", self.tone_data.tone_data.size)
            var = tone_grp.createVariable("tone_data", "f8", ("tone_data_size",))
            var[...] = self.tone_data.tone_data
        # Compensation pulse
        comp_grp.use_compensation_pulse = (
            1 if self.compensation_pulse_data.use_compensation_pulse else 0
        )
        if self.compensation_pulse_data.compensation_frequency is not None:
            comp_grp.compensation_frequency = self.compensation_pulse_data.compensation_frequency
        if self.compensation_pulse_data.compensation_decay is not None:
            comp_grp.compensation_decay = self.compensation_pulse_data.compensation_decay
        # Decay parameters
        decay_grp.decay_strategy = self.decay_data.decay_strategy.value
        decay_grp.common_decay = 1 if self.decay_data.common_decay else 0
        if self.decay_data.common_decay:
            decay_grp.decay_data = self.decay_data.decay_data[0]
        else:
            decay_grp.createDimension("num_decays", self.decay_data.decay_data.size)
            var = decay_grp.createVariable("decay_data", "f8", ("num_decays",))
            var[...] = self.decay_data.decay_data
        # SRS Group
        srs_grp.srs_type = self.srs_data.srs_type.value
        srs_grp.srs_displacement = self.srs_data.srs_displacement.value
        srs_grp.srs_damping = self.srs_data.srs_damping
        # SDS Group
        sds_grp.iterations = self.sds_data.iterations
        sds_grp.convergence = self.sds_data.convergence
        sds_grp.scale_factor = self.sds_data.scale_factor
        sds_grp.error_tolerance = self.sds_data.error_tolerance
        # Specification
        spec_grp.num_hits = self.specification_data.num_hits
        spec_grp.createDimension("num_frequencies", self.specification_data.frequencies.size)
        spec_grp.createDimension("num_spec_signals", self.specification_data.srs_spec.shape[1])
        var = spec_grp.createVariable("frequencies", "f8", ("num_frequencies"))
        var[...] = self.specification_data.frequencies
        var = spec_grp.createVariable("srs_spec", "f8", ("num_frequencies", "num_spec_signals"))
        var[...] = self.specification_data.srs_spec
        var = spec_grp.createVariable(
            "srs_lower_limit", "f8", ("num_frequencies", "num_spec_signals")
        )
        var[...] = self.specification_data.srs_lower_limit
        var = spec_grp.createVariable(
            "srs_upper_limit", "f8", ("num_frequencies", "num_spec_signals")
        )
        var[...] = self.specification_data.srs_upper_limit
        # Control group
        control_grp.control_type = self.control_script_data.control_type.value
        control_grp.control_script = self.control_script_data.control_script
        control_grp.control_object = self.control_script_data.control_object
        control_grp_params = control_grp.createGroup("control_extra_parameters")
        for key, value in self.control_script_data.control_parameters.items():
            setattr(control_grp_params, key, value)
        netcdf_group_handle.createDimension("control_channels", len(self.control_channel_indices))
        if self.response_transformation_matrix is None:
            netcdf_group_handle.createDimension(
                "specification_channels", len(self.control_channel_indices)
            )
        else:
            netcdf_group_handle.createDimension(
                "specification_channels", self.response_transformation_matrix.shape[0]
            )
        # Control Channels
        var = netcdf_group_handle.createVariable(
            "control_channel_indices", "i4", ("control_channels")
        )
        var[...] = self.control_channel_indices
        # Transformation Matrix
        if self.response_transformation_matrix is not None:
            var = netcdf_group_handle.createVariable(
                "response_transformation_matrix",
                "f8",
                ("specification_channels", "control_channels"),
            )
            var[...] = self.response_transformation_matrix
        if self.reference_transformation_matrix is not None:
            netcdf_group_handle.createDimension(
                "reference_transformation_rows",
                self.reference_transformation_matrix.shape[0],
            )
            netcdf_group_handle.createDimension(
                "reference_transformation_cols",
                self.reference_transformation_matrix.shape[1],
            )
            var = netcdf_group_handle.createVariable(
                "reference_transformation_matrix",
                "f8",
                ("reference_transformation_rows", "reference_transformation_cols"),
            )
            var[...] = self.reference_transformation_matrix

    def validate(self, hardware_metadata):
        return super().validate(hardware_metadata)
