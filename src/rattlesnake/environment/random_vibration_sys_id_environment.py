# -*- coding: utf-8 -*-
"""
This file defines a Random Vibration Environment where a specification is
defined and the controller solves for excitations that will cause the test
article to match the specified response.

This environment has a number of subprocesses, including CPSD and FRF
computation, data analysis, and signal generation.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import inspect
import threading
import multiprocessing as mp
import multiprocessing.sharedctypes  # pylint: disable=unused-import
import time
from enum import Enum
from multiprocessing.queues import Queue
from typing import List

import netCDF4 as nc4
import numpy as np
import openpyxl

from rattlesnake.environment.random_vibration_sys_id_utilities import (
    load_specification,
)
from rattlesnake.environment.abstract_interactive_control_law import (
    AbstractControlLawComputation,
)
from rattlesnake.environment.environment_utilities import (
    EnvironmentType,
)
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.environment.abstract_environment import (
    EnvironmentInstructions,
    EnvironmentCommands,
)
from rattlesnake.environment.abstract_sysid_environment import (
    SysIdEnvironment,
    SysIdEnvironmentMetadata,
)
from rattlesnake.process.abstract_sysid_data_analysis import SysIdMetadata
from rattlesnake.utilities import (
    GlobalCommands,
    VerboseMessageQueue,
    db2scale,
    load_python_module,
    _direction_map,
)
from rattlesnake.environment.abstract_interactive_control_law import ControlLawCommands
from rattlesnake.process.data_collector import (
    Acceptance,
    AcquisitionType,
    CollectorMetadata,
    DataCollectorCommands,
    TriggerSlope,
    Window,
    data_collector_process,
)
from rattlesnake.process.signal_generation import (
    CPSDSignalGenerator,
)
from rattlesnake.process.signal_generation_process import (
    SignalGenerationCommands,
    SignalGenerationMetadata,
    signal_generation_process,
)
from rattlesnake.process.spectral_processing import (
    AveragingTypes,
    Estimator,
    SpectralProcessingCommands,
    SpectralProcessingMetadata,
    spectral_processing_process,
)
from rattlesnake.user_interface.ui_utilities import UICommands

CONTROL_TYPE = EnvironmentType.RANDOM


# region Commands
class RandomVibrationCommands(EnvironmentCommands):
    """Valid random vibration commands"""

    ADJUST_TEST_LEVEL = 0
    CHECK_FOR_COMPLETE_SHUTDOWN = 1
    RECOMPUTE_PREDICTION = 2
    CHANGE_SPECIFICATION = 3
    SAVE_CONTROL_DATA = 4

    VALID_PROFILE_COMMANDS = (
        ADJUST_TEST_LEVEL,
        CHANGE_SPECIFICATION,
        SAVE_CONTROL_DATA,
    )
    VALID_DATA = {
        ADJUST_TEST_LEVEL: float,
        CHANGE_SPECIFICATION: str,
        SAVE_CONTROL_DATA: str,
    }


class RandomVibrationUICommands(Enum):
    ENABLE_CONTROL = 0
    CHANGE_SPECIFICATION = 1
    ADJUST_TEST_LEVEL = 2


# endregion


# region Metadata
class RandomVibrationMetadata(SysIdEnvironmentMetadata):
    """Container to hold the signal processing parameters of the environment"""

    def __init__(
        self,
        *,
        environment_name: str,
        channel_list_bools: list,
        sample_rate: int,
        number_of_channels,
        samples_per_frame,
        test_level_ramp_time,
        cola_window,
        cola_overlap,
        cola_window_exponent,
        sigma_clip,
        update_tf_during_control,
        frames_in_cpsd,
        cpsd_window,
        cpsd_overlap,
        percent_lines_out,
        allow_automatic_aborts,
        control_python_script,
        control_python_function,
        control_python_function_type,
        control_python_function_parameters,
        control_channel_indices,
        output_channel_indices,
        specification_frequency_lines,
        specification_cpsd_matrix,
        specification_warning_matrix,
        specification_abort_matrix,
        response_transformation_matrix,
        output_transformation_matrix,
        sysid_metadata=None,
    ):
        super().__init__(
            CONTROL_TYPE,
            environment_name,
            channel_list_bools,
            sample_rate,
            sysid_metadata,
        )
        self.number_of_channels = number_of_channels
        self.sample_rate = sample_rate
        self.samples_per_frame = samples_per_frame
        self.test_level_ramp_time = test_level_ramp_time
        self.cpsd_overlap = cpsd_overlap
        self.update_tf_during_control = update_tf_during_control
        self.cola_window = cola_window
        self.cola_overlap = cola_overlap
        self.cola_window_exponent = cola_window_exponent
        self.sigma_clip = sigma_clip
        self.frames_in_cpsd = frames_in_cpsd
        self.cpsd_window = cpsd_window
        self.response_transformation_matrix = response_transformation_matrix
        self.reference_transformation_matrix = output_transformation_matrix
        self.control_python_script = control_python_script
        self.control_python_function = control_python_function
        self.control_python_function_type = control_python_function_type
        self.control_python_function_parameters = control_python_function_parameters
        self.control_channel_indices = control_channel_indices
        self.output_channel_indices = output_channel_indices
        self.specification_frequency_lines = specification_frequency_lines
        self.specification_cpsd_matrix = specification_cpsd_matrix
        self.specification_warning_matrix = specification_warning_matrix
        self.specification_abort_matrix = specification_abort_matrix
        self.percent_lines_out = percent_lines_out
        self.allow_automatic_aborts = allow_automatic_aborts

    @property
    def sample_rate(self):
        return self._sample_rate

    @sample_rate.setter
    def sample_rate(self, value):
        self._sample_rate = value

    @property
    def number_of_channels(self):
        return self._number_of_channels

    @number_of_channels.setter
    def number_of_channels(self, value):
        self._number_of_channels = value

    @property
    def reference_channel_indices(self):
        return self.output_channel_indices

    @property
    def response_channel_indices(self):
        return self.control_channel_indices

    @property
    def response_transformation_matrix(self):
        return self._response_transformation_matrix

    @response_transformation_matrix.setter
    def response_transformation_matrix(self, value):
        self._response_transformation_matrix = value

    @property
    def reference_transformation_matrix(self):
        return self._reference_transformation_matrix

    @reference_transformation_matrix.setter
    def reference_transformation_matrix(self, value):
        self._reference_transformation_matrix = value

    @property
    def samples_per_acquire(self):
        """Property returning the samples per acquisition step given the overlap"""
        return int(self.samples_per_frame * (1 - self.cpsd_overlap))

    @property
    def frame_time(self):
        """Property returning the time per measurement frame"""
        return self.samples_per_frame / self.sample_rate

    @property
    def nyquist_frequency(self):
        """Property returning half the sample rate"""
        return self.sample_rate / 2

    @property
    def fft_lines(self):
        """Property returning the frequency lines given the sampling parameters"""
        return self.samples_per_frame // 2 + 1

    @property
    def frequency_spacing(self):
        """Property returning frequency line spacing given the sampling parameters"""
        return self.sample_rate / self.samples_per_frame

    @property
    def samples_per_output(self):
        """Property returning the samples per output given the COLA overlap"""
        return int(self.samples_per_frame * (1 - self.cola_overlap))

    @property
    def overlapped_output_samples(self):
        """Property returning the number of output samples that are overlapped."""
        return self.samples_per_frame - self.samples_per_output

    @property
    def skip_frames(self):
        """Property returning the number of frames to skip when changing levels"""
        return int(
            np.ceil(
                self.test_level_ramp_time
                * self.sample_rate
                / (self.samples_per_frame * (1 - self.cpsd_overlap))
            )
        )

    # endregion

    # region Validation
    def validate(self, hardware_metadata):
        return super().validate(hardware_metadata)

    # endregion

    # region Loading
    def load_specification(self, environment_channel_list, filename):
        coord_dtype = np.dtype([("node", "<u8"), ("direction", "i1")])
        if self.response_transformation_matrix is not None:
            control_coordinate = None
        else:
            control_coordinate = np.array(
                [
                    (
                        environment_channel_list[i].node_number,
                        _direction_map[environment_channel_list[i].node_direction],
                    )
                    for i in self.control_channel_indices
                ],
                dtype=coord_dtype,
            )

        (
            self.specification_frequency_lines,
            self.specification_cpsd_matrix,
            self.specification_warning_matrix,
            self.specification_abort_matrix,
        ) = load_specification(
            filename,
            self.fft_lines,
            self.frequency_spacing,
            control_coordinate,
        )

    def save_metadata_to_netcdf(
        self,
        netcdf_group_handle: nc4._netCDF4.Group,  # pylint: disable=c-extension-no-member
    ):
        """Store parameters to a group in a netCDF streaming file.

        This function stores parameters from the environment into the netCDF
        file in a group with the environment's name as its name.  The function
        will receive a reference to the group within the dataset and should
        store the environment's parameters into that group in the form of
        attributes, dimensions, or variables.

        This function is the "write" counterpart to the retrieve_metadata
        function in the RandomVibrationUI class, which will read parameters from
        the netCDF file to populate the parameters in the user interface.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            A reference to the Group within the netCDF dataset where the
            environment's metadata is stored.

        """
        super().save_metadata_to_netcdf(netcdf_group_handle)
        netcdf_group_handle.samples_per_frame = self.samples_per_frame
        netcdf_group_handle.test_level_ramp_time = self.test_level_ramp_time
        netcdf_group_handle.cpsd_overlap = self.cpsd_overlap
        netcdf_group_handle.update_tf_during_control = (
            1 if self.update_tf_during_control else 0
        )
        netcdf_group_handle.cola_window = self.cola_window
        netcdf_group_handle.cola_overlap = self.cola_overlap
        netcdf_group_handle.cola_window_exponent = self.cola_window_exponent
        netcdf_group_handle.frames_in_cpsd = self.frames_in_cpsd
        netcdf_group_handle.cpsd_window = self.cpsd_window
        netcdf_group_handle.control_python_script = self.control_python_script
        netcdf_group_handle.control_python_function = self.control_python_function
        netcdf_group_handle.control_python_function_type = (
            self.control_python_function_type
        )
        netcdf_group_handle.control_python_function_parameters = (
            self.control_python_function_parameters
        )
        netcdf_group_handle.allow_automatic_aborts = (
            1 if self.allow_automatic_aborts else 0
        )
        # Specifications
        netcdf_group_handle.createDimension("fft_lines", self.fft_lines)
        netcdf_group_handle.createDimension("two", 2)
        netcdf_group_handle.createDimension(
            "specification_channels", self.specification_cpsd_matrix.shape[-1]
        )
        var = netcdf_group_handle.createVariable(
            "specification_frequency_lines", "f8", ("fft_lines",)
        )
        var[...] = self.specification_frequency_lines
        var = netcdf_group_handle.createVariable(
            "specification_cpsd_matrix_real",
            "f8",
            ("fft_lines", "specification_channels", "specification_channels"),
        )
        var[...] = self.specification_cpsd_matrix.real
        var = netcdf_group_handle.createVariable(
            "specification_cpsd_matrix_imag",
            "f8",
            ("fft_lines", "specification_channels", "specification_channels"),
        )
        var[...] = self.specification_cpsd_matrix.imag
        var = netcdf_group_handle.createVariable(
            "specification_warning_matrix",
            "f8",
            ("two", "fft_lines", "specification_channels"),
        )
        var[...] = self.specification_warning_matrix.real
        var = netcdf_group_handle.createVariable(
            "specification_abort_matrix",
            "f8",
            ("two", "fft_lines", "specification_channels"),
        )
        var[...] = self.specification_abort_matrix.real
        # Transformation matrices
        if self.response_transformation_matrix is not None:
            netcdf_group_handle.createDimension(
                "response_transformation_rows",
                self.response_transformation_matrix.shape[0],
            )
            netcdf_group_handle.createDimension(
                "response_transformation_cols",
                self.response_transformation_matrix.shape[1],
            )
            var = netcdf_group_handle.createVariable(
                "response_transformation_matrix",
                "f8",
                ("response_transformation_rows", "response_transformation_cols"),
            )
            var[...] = self.response_transformation_matrix
        if self.reference_transformation_matrix is not None:
            netcdf_group_handle.createDimension(
                "reference_transformation_rows",
                self.reference_transformation_matrix.shape[0],
            )
            netcdf_group_handle.createDimension(
                "reference_transformation_cols",
                self.reference_transformation_matrix.shape[1],
            )
            var = netcdf_group_handle.createVariable(
                "reference_transformation_matrix",
                "f8",
                ("reference_transformation_rows", "reference_transformation_cols"),
            )
            var[...] = self.reference_transformation_matrix
        # Control channels
        netcdf_group_handle.createDimension(
            "control_channels", len(self.control_channel_indices)
        )
        var = netcdf_group_handle.createVariable(
            "control_channel_indices", "i4", ("control_channels")
        )
        var[...] = self.control_channel_indices

    @classmethod
    def load_metadata_from_netcdf(
        cls,
        netcdf_group_handle: nc4._netCDF4.Group,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        """Collect environment parameters from a netCDF group."""

        sample_rate = hardware_metadata.sample_rate
        number_of_channels = sum(channel_list_bools)

        environment_channel_list = [
            channel
            for channel, channel_bool in zip(
                hardware_metadata.channel_list, channel_list_bools
            )
            if channel_bool
        ]

        output_channel_indices = [
            index
            for index, channel in enumerate(environment_channel_list)
            if channel.feedback_device is not None
        ]

        samples_per_frame = netcdf_group_handle.samples_per_frame
        test_level_ramp_time = netcdf_group_handle.test_level_ramp_time
        cpsd_overlap = netcdf_group_handle.cpsd_overlap
        update_tf_during_control = bool(netcdf_group_handle.update_tf_during_control)
        cola_window = netcdf_group_handle.cola_window
        cola_overlap = netcdf_group_handle.cola_overlap
        cola_window_exponent = netcdf_group_handle.cola_window_exponent
        frames_in_cpsd = netcdf_group_handle.frames_in_cpsd
        cpsd_window = netcdf_group_handle.cpsd_window
        control_python_script = netcdf_group_handle.control_python_script
        control_python_function = netcdf_group_handle.control_python_function
        control_python_function_type = netcdf_group_handle.control_python_function_type
        control_python_function_parameters = (
            netcdf_group_handle.control_python_function_parameters
        )
        allow_automatic_aborts = bool(netcdf_group_handle.allow_automatic_aborts)

        control_channel_indices = netcdf_group_handle.variables[
            "control_channel_indices"
        ][...]

        specification_frequency_lines = netcdf_group_handle.variables[
            "specification_frequency_lines"
        ][...]
        specification_cpsd_matrix = (
            netcdf_group_handle.variables["specification_cpsd_matrix_real"][...]
            + 1j * netcdf_group_handle.variables["specification_cpsd_matrix_imag"][...]
        )
        specification_warning_matrix = netcdf_group_handle.variables[
            "specification_warning_matrix"
        ][...]
        specification_abort_matrix = netcdf_group_handle.variables[
            "specification_abort_matrix"
        ][...]

        response_transformation_matrix = None
        if "response_transformation_matrix" in netcdf_group_handle.variables:
            response_transformation_matrix = netcdf_group_handle.variables[
                "response_transformation_matrix"
            ][...]

        reference_transformation_matrix = None
        if "reference_transformation_matrix" in netcdf_group_handle.variables:
            reference_transformation_matrix = netcdf_group_handle.variables[
                "reference_transformation_matrix"
            ][...]

        sysid_metadata = SysIdMetadata.load_metadata_from_netcdf(
            netcdf_group_handle, hardware_metadata
        )

        return cls(
            environment_name=environment_name,
            channel_list_bools=channel_list_bools,
            sample_rate=sample_rate,
            number_of_channels=number_of_channels,
            samples_per_frame=samples_per_frame,
            test_level_ramp_time=test_level_ramp_time,
            cola_window=cola_window,
            cola_overlap=cola_overlap,
            cola_window_exponent=cola_window_exponent,
            sigma_clip=5,
            update_tf_during_control=update_tf_during_control,
            frames_in_cpsd=frames_in_cpsd,
            cpsd_window=cpsd_window,
            cpsd_overlap=cpsd_overlap,
            percent_lines_out=0.1,  # TODO This is wrong
            allow_automatic_aborts=allow_automatic_aborts,
            control_python_script=control_python_script,
            control_python_function=control_python_function,
            control_python_function_type=control_python_function_type,
            control_python_function_parameters=control_python_function_parameters,
            control_channel_indices=control_channel_indices,
            output_channel_indices=output_channel_indices,
            specification_frequency_lines=specification_frequency_lines,
            specification_cpsd_matrix=specification_cpsd_matrix,
            specification_warning_matrix=specification_warning_matrix,
            specification_abort_matrix=specification_abort_matrix,
            response_transformation_matrix=response_transformation_matrix,
            output_transformation_matrix=reference_transformation_matrix,
            sysid_metadata=sysid_metadata,
        )

    @classmethod
    def create_blank_worksheet_template(cls, worksheet):
        super().create_blank_worksheet_template(worksheet)
        worksheet.cell(1, 2, "Random")
        worksheet.cell(2, 1, "Samples Per Frame:")
        worksheet.cell(2, 3, "# Number of Samples per Measurement Frame")
        worksheet.cell(3, 1, "Test Level Ramp Time:")
        worksheet.cell(3, 3, "# Time taken to Ramp between test levels")
        worksheet.cell(4, 1, "COLA Window:")
        worksheet.cell(4, 3, "# Window used for Constant Overlap and Add process")
        worksheet.cell(5, 1, "COLA Overlap %:")
        worksheet.cell(5, 3, "# Overlap used in Constant Overlap and Add process")
        worksheet.cell(6, 1, "COLA Window Exponent:")
        worksheet.cell(
            6,
            3,
            "# Exponent Applied to the COLA Window (use 0.5 unless you "
            "are sure you don't want to!)",
        )
        worksheet.cell(7, 1, "Update System ID During Control:")
        worksheet.cell(
            7,
            3,
            "# Continue updating transfer function while the controller is controlling (Y/N)",
        )
        worksheet.cell(8, 1, "Frames in CPSD:")
        worksheet.cell(8, 3, "# Frames used to compute the CPSD matrix")
        worksheet.cell(9, 1, "CPSD Window:")
        worksheet.cell(9, 3, "# Window used to compute the CPSD matrix")
        worksheet.cell(10, 1, "CPSD Overlap %:")
        worksheet.cell(10, 3, "# Overlap percentage for CPSD calculations")
        worksheet.cell(11, 1, "Allow Automatic Aborts")
        worksheet.cell(
            11,
            3,
            "# Shut down the test automatically if an abort level is reached (Y/N)",
        )
        worksheet.cell(12, 1, "Control Python Script:")
        worksheet.cell(12, 3, "# Path to the Python script containing the control law")
        worksheet.cell(13, 1, "Control Python Function:")
        worksheet.cell(
            13,
            3,
            "# Function or class name within the Python Script that will serve as the control law",
        )
        worksheet.cell(14, 1, "Control Parameters:")
        worksheet.cell(14, 3, "# Extra parameters used in the control law")
        worksheet.cell(15, 1, "Control Channels (1-based):")
        worksheet.cell(16, 1, "Sigma Clipping")
        worksheet.cell(
            16, 3, "# Standard-deviation threshold used to reject outlier data."
        )
        SysIdMetadata.create_blank_worksheet_template(worksheet, start_row=17)
        worksheet.cell(33, 1, "Specification File:")
        worksheet.cell(33, 3, "# Path to the file containing the Specification")
        worksheet.cell(34, 1, "Response Transformation Matrix:")
        worksheet.cell(
            34,
            2,
            "# Transformation matrix to apply to the response channels.  Type None if there "
            "is none.  Otherwise, make this a 2D array in the spreadsheet and move the Output "
            "Transformation Matrix line down so it will fit.  The number of columns should be the "
            "number of physical control channels.",
        )
        worksheet.cell(35, 1, "Output Transformation Matrix:")
        worksheet.cell(
            35,
            2,
            "# Transformation matrix to apply to the outputs.  Type None if there is none.  "
            "Otherwise, make this a 2D array in the spreadsheet.  The number of columns should be "
            "the number of physical output channels in the environment.",
        )

    def save_metadata_to_worksheet(
        self, worksheet: openpyxl.worksheet.worksheet.Worksheet
    ):
        super().save_metadata_to_worksheet(worksheet)

        if self.samples_per_frame is not None:
            worksheet.cell(2, 2, self.samples_per_frame)
        if self.test_level_ramp_time is not None:
            worksheet.cell(3, 2, self.test_level_ramp_time)
        if self.cola_window is not None:
            worksheet.cell(4, 2, self.cola_window)
        if self.cola_overlap is not None:
            worksheet.cell(5, 2, self.cola_overlap)
        if self.cola_window_exponent is not None:
            worksheet.cell(6, 2, self.cola_window_exponent)
        if self.update_tf_during_control is not None:
            worksheet.cell(7, 2, "Y" if self.update_tf_during_control else "N")
        if self.frames_in_cpsd is not None:
            worksheet.cell(8, 2, self.frames_in_cpsd)
        if self.cpsd_window is not None:
            worksheet.cell(9, 2, self.cpsd_window)
        if self.cpsd_overlap is not None:
            worksheet.cell(10, 2, self.cpsd_overlap)
        if self.allow_automatic_aborts is not None:
            worksheet.cell(11, 2, "Y" if self.allow_automatic_aborts else "N")
        if self.control_python_script is not None:
            worksheet.cell(12, 2, self.control_python_script)
        if self.control_python_function is not None:
            worksheet.cell(13, 2, self.control_python_function)
        if self.control_python_function_parameters is not None:
            worksheet.cell(14, 2, self.control_python_function_parameters)
        if self.control_channel_indices is not None:
            for idx, channel_ind in enumerate(self.control_channel_indices):
                col_idx = idx + 2
                worksheet.cell(15, col_idx, channel_ind + 1)
        if self.sigma_clip is not None:
            worksheet.cell(16, 2, self.sigma_clip)
        self.sysid_metadata.save_metadata_to_worksheet(worksheet, start_row=17)
        self.save_sysid_matrix_to_worksheet(
            worksheet,
            self.response_transformation_matrix,
            self.reference_transformation_matrix,
            start_row=34,
        )

    @classmethod
    def load_metadata_from_worksheet(
        cls,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        sample_rate = hardware_metadata.sample_rate
        number_of_channels = sum(channel_list_bools)
        environment_channel_list = [
            channel
            for channel, channel_bool in zip(
                hardware_metadata.channel_list, channel_list_bools
            )
            if channel_bool
        ]

        output_channel_indices = [
            index
            for index, channel in enumerate(environment_channel_list)
            if channel.feedback_device is not None
        ]

        samples_per_frame = int(worksheet.cell(2, 2).value)
        test_level_ramp_time = float(worksheet.cell(3, 2).value)
        cola_window = worksheet.cell(4, 2).value
        cola_overlap = float(worksheet.cell(5, 2).value)
        cola_window_exponent = float(worksheet.cell(6, 2).value)
        update_tf_during_control = worksheet.cell(7, 2).value.upper() == "Y"
        frames_in_cpsd = int(worksheet.cell(8, 2).value)
        cpsd_window = worksheet.cell(9, 2).value
        cpsd_overlap = float(worksheet.cell(10, 2).value)
        allow_automatic_aborts = worksheet.cell(11, 2).value.upper() == "Y"

        control_python_script = (
            worksheet.cell(12, 2).value
            if worksheet.cell(12, 2).value is not None
            else ""
        )
        control_python_function = (
            worksheet.cell(13, 2).value
            if worksheet.cell(13, 2).value is not None
            else ""
        )
        control_python_function_parameters = (
            worksheet.cell(14, 2).value
            if worksheet.cell(14, 2).value is not None
            else ""
        )
        control_channel_indices = []
        column_index = 2
        while True:
            channel_ind = worksheet.cell(15, column_index).value
            if channel_ind is None or (
                isinstance(channel_ind, str) and channel_ind.strip() == ""
            ):
                break
            try:
                control_channel_indices.append(int(channel_ind) - 1)
            except:
                break
            column_index += 1
        sigma_clip = float(worksheet.cell(16, 2).value)

        sysid_metadata = SysIdMetadata.load_metadata_from_worksheet(
            worksheet, hardware_metadata, 17
        )

        response_transformation_matrix, output_transformation_matrix = (
            cls.load_sysid_matrix_from_worksheet(worksheet, start_row=34)
        )

        # Find python module type
        if control_python_script:
            python_control_module = load_python_module(control_python_script)
            function = getattr(python_control_module, control_python_function)
            control_python_function_type = None
            if inspect.isgeneratorfunction(function):
                control_python_function_type = 1
            elif inspect.isclass(function) and issubclass(
                function, AbstractControlLawComputation
            ):
                control_python_function_type = 2
            elif inspect.isclass(function):
                control_python_function_type = 3
            else:
                control_python_function_type = 0
        else:
            control_python_function_type = None

        coord_dtype = np.dtype([("node", "<u8"), ("direction", "i1")])
        if response_transformation_matrix is not None:
            control_coordinate = None
        else:
            control_coordinate = np.array(
                [
                    (
                        hardware_metadata.channel_list[i].node_number,
                        _direction_map[
                            hardware_metadata.channel_list[i].node_direction
                        ],
                    )
                    for i in output_channel_indices
                ],
                dtype=coord_dtype,
            )

        metadata = cls(
            environment_name=environment_name,
            channel_list_bools=channel_list_bools,
            sample_rate=sample_rate,
            number_of_channels=number_of_channels,
            samples_per_frame=samples_per_frame,
            test_level_ramp_time=test_level_ramp_time,
            cola_window=cola_window,
            cola_overlap=cola_overlap,
            cola_window_exponent=cola_window_exponent,
            sigma_clip=sigma_clip,
            update_tf_during_control=update_tf_during_control,
            frames_in_cpsd=frames_in_cpsd,
            cpsd_window=cpsd_window,
            cpsd_overlap=cpsd_overlap,
            percent_lines_out=0.1,  # TODO This is wrong
            allow_automatic_aborts=allow_automatic_aborts,
            control_python_script=control_python_script,
            control_python_function=control_python_function,
            control_python_function_type=control_python_function_type,
            control_python_function_parameters=control_python_function_parameters,
            control_channel_indices=control_channel_indices,
            output_channel_indices=output_channel_indices,
            specification_frequency_lines=None,
            specification_cpsd_matrix=None,
            specification_warning_matrix=None,
            specification_abort_matrix=None,
            response_transformation_matrix=response_transformation_matrix,
            output_transformation_matrix=output_transformation_matrix,
            sysid_metadata=sysid_metadata,
        )

        # Load specification
        specification_file = worksheet.cell(33, 2).value
        if specification_file is not None:
            (
                metadata.specification_frequency_lines,
                metadata.specification_cpsd_matrix,
                metadata.specification_warning_matrix,
                metadata.specification_abort_matrix,
            ) = load_specification(
                specification_file,
                metadata.fft_lines,
                metadata.frequency_spacing,
                control_coordinate,
            )

        return metadata


# region Instructions
class RandomVibrationInstructions(EnvironmentInstructions):
    def __init__(self, environment_name, control_test_level):
        super().__init__(CONTROL_TYPE, environment_name)
        self.control_test_level = control_test_level

    def validate(self):
        return super().validate()


# endregion


# region Queues
class RandomVibrationQueues:
    """A container class for the queues that random vibration will manage."""

    def __init__(
        self,
        environment_name: str,
        environment_command_queue: VerboseMessageQueue,
        gui_update_queue: mp.queues.Queue,
        controller_communication_queue: VerboseMessageQueue,
        data_in_queue: mp.queues.Queue,
        data_out_queue: mp.queues.Queue,
        log_file_queue: VerboseMessageQueue,
    ):
        """A container class for the queues that random vibration will manage.

        The environment uses many queues to pass data between the various pieces.
        This class organizes those queues into one common namespace.


        Parameters
        ----------
        environment_name : str
            Name of the environment
        environment_command_queue : VerboseMessageQueue
            Queue that is read by the environment for environment commands
        gui_update_queue : mp.queues.Queue
            Queue where various subtasks put instructions for updating the
            widgets in the user interface
        controller_communication_queue : VerboseMessageQueue
            Queue that is read by the controller for global controller commands
        data_in_queue : mp.queues.Queue
            Multiprocessing queue that connects the acquisition subtask to the
            environment subtask.  Each environment will retrieve acquired data
            from this queue.
        data_out_queue : mp.queues.Queue
            Multiprocessing queue that connects the output subtask to the
            environment subtask.  Each environment will put data that it wants
            the controller to generate in this queue.
        log_file_queue : VerboseMessageQueue
            Queue for putting logging messages that will be read by the logging
            subtask and written to a file.
        """
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.data_analysis_command_queue = VerboseMessageQueue(
            log_file_queue,
            mp.Queue(),
            environment_name + " Data Analysis Command Queue",
        )
        self.signal_generation_command_queue = VerboseMessageQueue(
            log_file_queue,
            mp.Queue(),
            environment_name + " Signal Generation Command Queue",
        )
        self.spectral_command_queue = VerboseMessageQueue(
            log_file_queue,
            mp.Queue(),
            environment_name + " Spectral Computation Command Queue",
        )
        self.collector_command_queue = VerboseMessageQueue(
            log_file_queue,
            mp.Queue(),
            environment_name + " Data Collector Command Queue",
        )
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.data_for_spectral_computation_queue = mp.Queue()
        self.updated_spectral_quantities_queue = mp.Queue()
        self.cpsd_to_generate_queue = mp.Queue()
        self.log_file_queue = log_file_queue


# endregion

from rattlesnake.process.random_vibration_sys_id_data_analysis import (  # noqa: E402 pylint: disable=wrong-import-position
    RandomVibrationDataAnalysisCommands,
    random_data_analysis_process,
)


class RandomVibrationEnvironment(SysIdEnvironment):
    """Random Environment class defining the interface with the controller"""

    # region Environment
    def __init__(
        self,
        environment_name: str,
        queue_name: str,
        queue_container: RandomVibrationQueues,
        acquisition_active_event: mp.synchronize.Event,
        output_active_event: mp.synchronize.Event,
        active_event: mp.synchronize.Event,
        ready_event: mp.synchronize.Event,
        sysid_active_event: mp.synchronize.Event,
        sysid_stored_event: mp.synchronize.Event,
    ):
        """
        Random Vibration Environment Constructor that fills out the ``command_map``

        Parameters
        ----------
        environment_name : str
            Name of the environment.
        queue_container : RandomVibrationQueues
            Container of queues used by the Random Vibration Environment.

        """
        super().__init__(
            environment_name,
            queue_name,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.controller_communication_queue,
            queue_container.log_file_queue,
            queue_container.collector_command_queue,
            queue_container.signal_generation_command_queue,
            queue_container.spectral_command_queue,
            queue_container.data_analysis_command_queue,
            queue_container.data_in_queue,
            queue_container.data_out_queue,
            acquisition_active_event,
            output_active_event,
            active_event,
            ready_event,
            sysid_active_event,
            sysid_stored_event,
        )
        self.map_command(GlobalCommands.START_ENVIRONMENT, self.start_control)
        self.map_command(
            RandomVibrationDataAnalysisCommands.STOP_CONTROL, self.stop_environment
        )
        self.map_command(
            RandomVibrationCommands.ADJUST_TEST_LEVEL, self.adjust_test_level
        )
        self.map_command(
            RandomVibrationCommands.SAVE_CONTROL_DATA, self.save_spectral_data
        )
        self.map_command(
            RandomVibrationCommands.CHANGE_SPECIFICATION, self.change_specification
        )
        self.map_command(
            RandomVibrationCommands.CHECK_FOR_COMPLETE_SHUTDOWN,
            self.check_for_control_shutdown,
        )
        self.map_command(
            RandomVibrationCommands.RECOMPUTE_PREDICTION, self.recompute_prediction
        )
        self.map_command(
            ControlLawCommands.UPDATE_INTERACTIVE_CONTROL_PARAMETERS,
            self.update_interactive_control_parameters,
        )
        self.map_command(
            ControlLawCommands.SEND_INTERACTIVE_COMMAND, self.send_interactive_command
        )
        self.queue_container = queue_container

        self.set_ready()

    # endregion

    # region StateSync
    def initialize_hardware(self, hardware_metadata):
        super().initialize_hardware(hardware_metadata)

        self.set_ready()

    def initialize_environment(self, environment_metadata: RandomVibrationMetadata):
        self.environment_name = environment_metadata.environment_name
        self.environment_metadata = environment_metadata

        # Set up the data analysis
        self.queue_container.data_analysis_command_queue.put(
            self.environment_name,
            (
                RandomVibrationDataAnalysisCommands.INITIALIZE_ENVIRONMENT,
                self.environment_metadata,
            ),
        )

        # Set up the collector
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (
                DataCollectorCommands.INITIALIZE_COLLECTOR,
                self.get_data_collector_metadata(),
            ),
        )
        # Set up the signal generation
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (
                SignalGenerationCommands.INITIALIZE_PARAMETERS,
                self.get_signal_generation_metadata(),
            ),
        )
        # Set up the spectral processing
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (
                SpectralProcessingCommands.INITIALIZE_PARAMETERS,
                self.get_spectral_processing_metadata(),
            ),
        )

        self.set_ready()

    def initialize_sysid(self, sysid_metadata):
        super().initialize_sysid(sysid_metadata)

        self.set_ready()

    def update_interactive_control_parameters(self, parameters):
        """Sends updated parameters to the interactive control law on the data analysis process"""
        self.queue_container.data_analysis_command_queue.put(
            self.environment_name,
            (ControlLawCommands.UPDATE_INTERACTIVE_CONTROL_PARAMETERS, parameters),
        )

    def get_data_collector_metadata(self):
        """Gets relevant metadata for the data collector process"""
        num_channels = self.environment_metadata.number_of_channels
        response_channel_indices = self.environment_metadata.response_channel_indices
        reference_channel_indices = self.environment_metadata.reference_channel_indices
        acquisition_type = AcquisitionType.FREE_RUN
        acceptance = Acceptance.AUTOMATIC
        acceptance_function = None
        overlap_fraction = self.environment_metadata.cpsd_overlap
        trigger_channel_index = 0
        trigger_slope = TriggerSlope.POSITIVE
        trigger_level = 0
        trigger_hysteresis = 0
        trigger_hysteresis_samples = 0
        pretrigger_fraction = 0
        frame_size = self.environment_metadata.samples_per_frame
        window = (
            Window.HANN if self.environment_metadata.cpsd_window == "Hann" else None
        )
        # use number of sysid averages as kurtosis buffer size
        # (could maybe make this match the test duration if user is using the "Time at Level"
        # function, would need to pass info from the RandomVibrationUI object)
        kurtosis_buffer_length = self.environment_metadata.sysid_metadata.sysid_averages

        return CollectorMetadata(
            num_channels,
            response_channel_indices,
            reference_channel_indices,
            acquisition_type,
            acceptance,
            acceptance_function,
            overlap_fraction,
            trigger_channel_index,
            trigger_slope,
            trigger_level,
            trigger_hysteresis,
            trigger_hysteresis_samples,
            pretrigger_fraction,
            frame_size,
            window,
            kurtosis_buffer_length=kurtosis_buffer_length,
            response_transformation_matrix=self.environment_metadata.response_transformation_matrix,
            reference_transformation_matrix=self.environment_metadata.reference_transformation_matrix,
        )

    def get_signal_generation_metadata(self):
        """Gets relevant metadata for the signal generation process"""
        return SignalGenerationMetadata(
            samples_per_write=self.hardware_metadata.samples_per_write,
            level_ramp_samples=self.environment_metadata.test_level_ramp_time
            * self.environment_metadata.sample_rate
            * self.hardware_metadata.output_oversample,
            output_transformation_matrix=self.environment_metadata.reference_transformation_matrix,
        )

    def get_signal_generator(self):
        """Gets the signal generator object that will generate signals for the environment"""
        return CPSDSignalGenerator(
            self.environment_metadata.sample_rate,
            self.environment_metadata.samples_per_frame,
            self.environment_metadata.num_reference_channels,
            None,
            self.environment_metadata.cola_overlap,
            self.environment_metadata.cola_window,
            self.environment_metadata.cola_window_exponent,
            self.environment_metadata.sigma_clip,
            self.hardware_metadata.output_oversample,
        )

    def get_spectral_processing_metadata(self):
        """Gets the required metadata for the spectral processing process"""
        averaging_type = AveragingTypes.LINEAR
        averages = self.environment_metadata.frames_in_cpsd
        exponential_averaging_coefficient = 0
        if self.environment_metadata.sysid_metadata.sysid_estimator == "H1":
            frf_estimator = Estimator.H1
        elif self.environment_metadata.sysid_metadata.sysid_estimator == "H2":
            frf_estimator = Estimator.H2
        elif self.environment_metadata.sysid_metadata.sysid_estimator == "H3":
            frf_estimator = Estimator.H3
        elif self.environment_metadata.sysid_metadata.sysid_estimator == "Hv":
            frf_estimator = Estimator.HV
        else:
            raise ValueError(
                f"Invalid FRF Estimator {self.environment_metadata.sysid_metadata.sysid_estimator}"
            )
        num_response_channels = self.environment_metadata.num_response_channels
        num_reference_channels = self.environment_metadata.num_reference_channels
        frequency_spacing = self.environment_metadata.frequency_spacing
        sample_rate = self.environment_metadata.sample_rate
        num_frequency_lines = self.environment_metadata.fft_lines
        return SpectralProcessingMetadata(
            averaging_type,
            averages,
            exponential_averaging_coefficient,
            frf_estimator,
            num_response_channels,
            num_reference_channels,
            frequency_spacing,
            sample_rate,
            num_frequency_lines,
        )

    # endregion

    # region Commands
    def adjust_test_level(self, data):
        """Adjusts the test level of the environment to the specified level"""
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.ADJUST_TEST_LEVEL, db2scale(data)),
        )
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (
                DataCollectorCommands.SET_TEST_LEVEL,
                (self.environment_metadata.skip_frames, db2scale(data)),
            ),
        )
        self.gui_update_queue.put(
            (
                self.environment_name,
                (RandomVibrationUICommands.ADJUST_TEST_LEVEL, data),
            )
        )

    def send_interactive_command(self, command):
        """General method that can be used by an interactive UI object to pass commands and data to
        its corresponding computation object"""
        if self.environment_metadata.control_python_function_type == 3:  # Interactive
            self.queue_container.data_analysis_command_queue.put(
                self.environment_name,
                (ControlLawCommands.SEND_INTERACTIVE_COMMAND, command),
            )
        else:
            raise ValueError(
                "Received an SEND_INTERACTIVE_COMMAND signal without an interactive control law.  "
                "How did this happen?"
            )

    def system_id_complete(self, data):
        """Triggered when system identification has been completed, starting control predictions"""
        super().system_id_complete(data)
        self.queue_container.data_analysis_command_queue.put(
            self.environment_name,
            (RandomVibrationDataAnalysisCommands.PERFORM_CONTROL_PREDICTION, None),
        )
        self.set_sysid_stored()

    def recompute_prediction(self, data):  # pylint: disable=unused-argument
        """Sends a signal to the data analysis process to recompute test predictions"""
        self.queue_container.data_analysis_command_queue.put(
            self.environment_name,
            (RandomVibrationDataAnalysisCommands.PERFORM_CONTROL_PREDICTION, None),
        )

    def start_control(self, data: RandomVibrationInstructions):
        """Starts the environment at the specified test level"""
        self.log("Starting Control")
        test_level = db2scale(data.control_test_level)
        self.gui_update_queue.put(
            (
                self.environment_name,
                (UICommands.SET_ENVIRONMENT_INSTRUCTIONS, data),
            )
        )
        self.siggen_shutdown_achieved = False
        self.collector_shutdown_achieved = False
        self.spectral_shutdown_achieved = False
        self.analysis_shutdown_achieved = False
        # Set up the collector
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (
                DataCollectorCommands.INITIALIZE_COLLECTOR,
                self.get_data_collector_metadata(),
            ),
        )

        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (
                DataCollectorCommands.SET_TEST_LEVEL,
                (self.environment_metadata.skip_frames, test_level),
            ),
        )
        time.sleep(0.01)

        # Set up the signal generation
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (
                SignalGenerationCommands.INITIALIZE_PARAMETERS,
                self.get_signal_generation_metadata(),
            ),
        )

        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (
                SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
                self.get_signal_generator(),
            ),
        )

        self.queue_container.signal_generation_command_queue.put(
            self.environment_name, (SignalGenerationCommands.MUTE, None)
        )

        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.ADJUST_TEST_LEVEL, test_level),
        )

        # Tell the collector to start acquiring data
        self.queue_container.collector_command_queue.put(
            self.environment_name, (DataCollectorCommands.ACQUIRE, None)
        )

        # Tell the signal generation to start generating signals
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name, (SignalGenerationCommands.GENERATE_SIGNALS, None)
        )

        # # Set up the data analysis
        # self.queue_container.data_analysis_command_queue.put(
        #     self.environment_name,
        #     (RandomVibrationDataAnalysisCommands.INITIALIZE_PARAMETERS,
        #      self.environment_metadata))

        # Start the data analysis running
        self.queue_container.data_analysis_command_queue.put(
            self.environment_name,
            (RandomVibrationDataAnalysisCommands.RUN_CONTROL, None),
        )

        # Set up the spectral processing
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (
                SpectralProcessingCommands.INITIALIZE_PARAMETERS,
                self.get_spectral_processing_metadata(),
            ),
        )

        # Tell the spectral analysis to clear and start acquiring
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.CLEAR_SPECTRAL_PROCESSING, None),
        )

        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.RUN_SPECTRAL_PROCESSING, None),
        )

        self.set_active()

    def stop_environment(self, data):
        """Stop the environment gracefully

        This function defines the operations to shut down the environment
        gracefully so there is no hard stop that might damage test equipment
        or parts.

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        """
        self.log("Stopping Control")
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (
                DataCollectorCommands.SET_TEST_LEVEL,
                (self.environment_metadata.skip_frames * 10, 1),
            ),
        )
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name, (SignalGenerationCommands.START_SHUTDOWN, None)
        )
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.STOP_SPECTRAL_PROCESSING, None),
        )
        self.queue_container.data_analysis_command_queue.put(
            self.environment_name,
            (RandomVibrationDataAnalysisCommands.STOP_CONTROL, None),
        )
        self.queue_container.environment_command_queue.put(
            self.environment_name,
            (RandomVibrationCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None),
        )

    def save_spectral_data(self, data):
        filename = data
        netcdf_dataset = nc4.Dataset(  # pylint: disable=no-member
            filename, "w", format="NETCDF4", clobber=True
        )
        if self.environment_name not in netcdf_dataset.groups:
            netcdf_handle = netcdf_dataset.createGroup(self.environment_name)
        else:
            netcdf_handle = netcdf_dataset.groups[self.environment_name]
        self.environment_metadata.save_metadata_to_netcdf(netcdf_handle)
        netcdf_dataset.close()

        self.data_analysis_command_queue.put(
            self.environment_name,
            (RandomVibrationDataAnalysisCommands.SAVE_CONTROL_DATA, filename),
        )

    def change_specification(self, data):
        """
        Loads in a new specification and starts controlling to it

        Parameters
        ----------
        new_specification_file : str
            File path to a new specification file

        """
        filename = data
        new_metadata = self.environment_metadata
        new_metadata.load_specification(self.hardware_metadata.channel_list, filename)
        self.initialize_environment(new_metadata)

        self.gui_update_queue.put(
            (
                self.environment_name,
                (
                    RandomVibrationUICommands.CHANGE_SPECIFICATION,
                    (filename, new_metadata),
                ),
            )
        )

    # region Shutdown
    def check_for_control_shutdown(self, data):  # pylint: disable=unused-argument
        """Checks the different processes to see if the controller has shut down gracefully"""
        if (
            self.siggen_shutdown_achieved
            and self.collector_shutdown_achieved
            and self.spectral_shutdown_achieved
            and self.analysis_shutdown_achieved
        ):
            self.log("Shutdown Achieved")
            self.gui_update_queue.put(
                (
                    self.environment_name,
                    (RandomVibrationUICommands.ENABLE_CONTROL, None),
                )
            )
            self.clear_active()
        else:
            # Recheck some time later
            time.sleep(1)
            self.environment_command_queue.put(
                self.environment_name,
                (RandomVibrationCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None),
            )

    def quit(self, data):
        """Closes down the environment permanently as the software is exiting"""
        for queue in [
            self.queue_container.spectral_command_queue,
            self.queue_container.data_analysis_command_queue,
            self.queue_container.signal_generation_command_queue,
            self.queue_container.collector_command_queue,
        ]:
            queue.put(self.environment_name, (GlobalCommands.QUIT, None))
        # Return true to stop the task
        return True

    # endregion


# region Process
def random_vibration_process(
    environment_name: str,
    queue_name: str,
    input_queue: VerboseMessageQueue,
    gui_update_queue: mp.Queue,
    controller_command_queue: VerboseMessageQueue,
    log_file_queue: mp.Queue,
    data_in_queue: mp.Queue,
    data_out_queue: mp.Queue,
    acquisition_active_event: mp.synchronize.Event,
    output_active_event: mp.synchronize.Event,
    active_event: mp.synchronize.Event,
    ready_event: mp.synchronize.Event,
    shutdown_event: mp.synchronize.Event,
    sysid_active_event: mp.synchronize.Event,
    sysid_stored_event: mp.synchronize.Event,
    ping_alive_event: mp.synchronize.Event,
    threaded: bool,
):
    """Random vibration environment process function called by multiprocessing

    This function defines the Random Vibration Environment process that
    gets run by the multiprocessing module when it creates a new process.  It
    creates a RandomVibrationEnvironment object and runs it.

    Parameters
    ----------
    environment_name : str :
        Name of the environment
    input_queue : VerboseMessageQueue :
        Queue containing instructions for the environment
    gui_update_queue : Queue :
        Queue where GUI updates are put
    controller_communication_queue : Queue :
        Queue for global communications with the controller
    log_file_queue : Queue :
        Queue for writing log file messages
    data_in_queue : Queue :
        Queue from which data will be read by the environment
    data_out_queue : Queue :
        Queue to which data will be written that will be output by the hardware.
    acquisition_active : mp.sharedctypes.Synchronized
        A synchronized value that indicates when the acquisition is active
    output_active : mp.sharedctypes.Synchronized
        A synchronized value that indicates when the output is active
    """
    # Create vibration queues
    if threaded:
        new_process = threading.Thread  # worker threads
    else:
        new_process = mp.Process  # worker processes
    queue_container = RandomVibrationQueues(
        environment_name,
        input_queue,
        gui_update_queue,
        controller_command_queue,
        data_in_queue,
        data_out_queue,
        log_file_queue,
    )

    spectral_proc = new_process(
        target=spectral_processing_process,
        args=(
            environment_name,
            queue_container.spectral_command_queue,
            queue_container.data_for_spectral_computation_queue,
            queue_container.updated_spectral_quantities_queue,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.log_file_queue,
        ),
    )
    spectral_proc.start()
    analysis_proc = new_process(
        target=random_data_analysis_process,
        args=(
            environment_name,
            queue_container.data_analysis_command_queue,
            queue_container.updated_spectral_quantities_queue,
            queue_container.cpsd_to_generate_queue,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.log_file_queue,
            ping_alive_event,
        ),
    )
    analysis_proc.start()
    siggen_proc = new_process(
        target=signal_generation_process,
        args=(
            environment_name,
            queue_container.signal_generation_command_queue,
            queue_container.cpsd_to_generate_queue,
            queue_container.data_out_queue,
            queue_container.environment_command_queue,
            queue_container.log_file_queue,
            queue_container.gui_update_queue,
        ),
    )
    siggen_proc.start()
    collection_proc = new_process(
        target=data_collector_process,
        args=(
            environment_name,
            queue_container.collector_command_queue,
            queue_container.data_in_queue,
            [queue_container.data_for_spectral_computation_queue],
            queue_container.environment_command_queue,
            queue_container.log_file_queue,
            queue_container.gui_update_queue,
        ),
    )

    collection_proc.start()
    process_class = RandomVibrationEnvironment(
        environment_name,
        queue_name,
        queue_container,
        acquisition_active_event,
        output_active_event,
        active_event,
        ready_event,
        sysid_active_event,
        sysid_stored_event,
    )
    process_class.run(shutdown_event)

    # Rejoin all the processes
    process_class.log("Joining Subprocesses")
    process_class.log("Joining Spectral Computation")
    spectral_proc.join()
    process_class.log("Joining Data Analysis")
    analysis_proc.join()
    process_class.log("Joining Signal Generation")
    siggen_proc.join()
    process_class.log("Joining Data Collection")
    collection_proc.join()


# endregion
