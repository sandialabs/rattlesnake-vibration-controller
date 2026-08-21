import sys
import os
from qtpy import QtWidgets, QtCore, uic
from rattlesnake.user_interface.sds_sys_id_prediction_table import SDSPredictionTable
from rattlesnake.engine import RattlesnakeController
from rattlesnake.environment.sds_sys_id_utilities import decayed_sine_table
from rattlesnake.utilities import DIRECTORY
import numpy as np
import csv
import openpyxl


class SDSRunTableDialog(QtWidgets.QDialog):

    def __init__(
        self,
        rattlesnake: RattlesnakeController,
        environment_name,
        prediction_mode,
        parent=None,
        other_voltage_lists=None,
        other_error_lists=None,
    ):
        super().__init__(parent)
        uic.loadUi(
            os.path.join(DIRECTORY, "user_interface", "ui_files", "srs_sds_run_table.ui"),
            self,
        )
        self.setWindowTitle("SDS Run Table")
        self.run_table = SDSPredictionTable(
            self.prediction_table_placeholder,
            rattlesnake,
            environment_name,
            prediction_mode=prediction_mode,
            other_voltage_lists=other_voltage_lists,
            other_error_lists=other_error_lists,
        )

        self.load_table_button.clicked.connect(self.load_sds_table)
        self.save_table_button.clicked.connect(self.save_sds_table)
        self.allow_manual_updates_checkbox.stateChanged.connect(self.set_table_lock)
        self.allow_automatic_updates_checkbox.stateChanged.connect(self.set_automatic_updates)

    def set_run_table(self, run_table):
        self.run_table = run_table

    def closeEvent(self, event):
        event.ignore()
        self.hide()

    @staticmethod
    def _is_header_row(row):
        if row is None:
            return False
        text = [("" if v is None else str(v).strip().lower()) for v in row[:4]]
        expected = ["frequency", "amplitude", "delay", "decay"]
        return text == expected

    @staticmethod
    def _to_float(value):
        if value is None:
            return np.nan
        if isinstance(value, (int, float, np.integer, np.floating)):
            return float(value)
        text = str(value).strip()
        if text == "":
            return np.nan
        return float(text)

    def _load_csv_table(self, filename):
        rows = []
        with open(filename, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) == 0:
                    continue
                rows.append(row)

        if len(rows) == 0:
            raise ValueError("CSV file is empty.")

        if self._is_header_row(rows[0]):
            rows = rows[1:]

        if len(rows) == 0:
            raise ValueError("CSV file contains only a header row.")

        parsed = np.array(
            [[self._to_float(cell) for cell in row[:4]] for row in rows],
            dtype=float,
        )

        if parsed.shape[1] < 4:
            raise ValueError("CSV file must contain at least 4 columns.")

        return parsed[:, 0], parsed[:, 1], parsed[:, 2], parsed[:, 3]

    def _load_xlsx_table(self, filename):
        workbook = openpyxl.load_workbook(filename, data_only=True)
        if len(workbook.worksheets) != 1:
            raise ValueError("XLSX file must contain exactly one worksheet.")

        worksheet = workbook.worksheets[0]

        rows = []
        for row in worksheet.iter_rows(values_only=True):
            if row is None:
                continue
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            rows.append(list(row))

        workbook.close()

        if len(rows) == 0:
            raise ValueError("XLSX worksheet is empty.")

        if self._is_header_row(rows[0]):
            rows = rows[1:]

        if len(rows) == 0:
            raise ValueError("XLSX worksheet contains only a header row.")

        parsed = np.array(
            [[self._to_float(cell) for cell in row[:4]] for row in rows],
            dtype=float,
        )

        if parsed.shape[1] < 4:
            raise ValueError("XLSX worksheet must contain at least 4 columns.")

        return parsed[:, 0], parsed[:, 1], parsed[:, 2], parsed[:, 3]

    def load_sds_table(self):
        filename, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Load SDS Table",
            filter="Supported Files (*.npz *.csv *.xlsx);;NumPy Files (*.npz);;CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*.*)",
        )
        if filename == "":
            return

        _, extension = os.path.splitext(filename)
        extension = extension.lower()

        try:
            if extension == ".npz":
                data = np.load(filename)

                required_keys = {"frequency", "amplitude", "delay", "decay"}
                missing = required_keys.difference(data.files)
                if missing:
                    raise ValueError(
                        "The selected file is missing required arrays: "
                        + ", ".join(sorted(missing))
                    )

                frequencies = np.array(data["frequency"]).flatten()
                amplitudes = np.array(data["amplitude"])
                delays = np.array(data["delay"])
                decays = np.array(data["decay"])

                if amplitudes.ndim == 1:
                    amplitudes = amplitudes[:, np.newaxis]
                if delays.ndim == 1:
                    delays = delays[:, np.newaxis]
                if decays.ndim == 1:
                    decays = decays[:, np.newaxis]

                num_rows = frequencies.size

                if amplitudes.shape[0] != num_rows:
                    raise ValueError(
                        f"Amplitude array first dimension ({amplitudes.shape[0]}) does not "
                        f"match number of frequencies ({num_rows})."
                    )
                if delays.shape[0] != num_rows:
                    raise ValueError(
                        f"Delay array first dimension ({delays.shape[0]}) does not "
                        f"match number of frequencies ({num_rows})."
                    )
                if decays.shape[0] != num_rows:
                    raise ValueError(
                        f"Decay array first dimension ({decays.shape[0]}) does not "
                        f"match number of frequencies ({num_rows})."
                    )

                if self.run_table.drive_names is not None:
                    expected_num_signals = len(self.run_table.drive_names)
                elif self.run_table.sds_table is not None:
                    expected_num_signals = self.run_table.sds_table["amplitude"].shape[1]
                else:
                    expected_num_signals = amplitudes.shape[1]

                if amplitudes.shape[1] != expected_num_signals:
                    raise ValueError(
                        f"Amplitude array second dimension ({amplitudes.shape[1]}) does not "
                        f"match expected number of drive channels ({expected_num_signals})."
                    )
                if delays.shape[1] != expected_num_signals:
                    raise ValueError(
                        f"Delay array second dimension ({delays.shape[1]}) does not "
                        f"match expected number of drive channels ({expected_num_signals})."
                    )
                if decays.shape[1] != expected_num_signals:
                    raise ValueError(
                        f"Decay array second dimension ({decays.shape[1]}) does not "
                        f"match expected number of drive channels ({expected_num_signals})."
                    )

                self.run_table.sds_table = decayed_sine_table(
                    frequencies,
                    amplitudes,
                    decays,
                    delays,
                )

            elif extension == ".csv":
                frequencies, amplitudes, delays, decays = self._load_csv_table(filename)
                self._apply_single_channel_table(frequencies, amplitudes, delays, decays)

            elif extension == ".xlsx":
                frequencies, amplitudes, delays, decays = self._load_xlsx_table(filename)
                self._apply_single_channel_table(frequencies, amplitudes, delays, decays)

            else:
                raise ValueError(f"Unsupported file extension: {extension}")

            self.run_table.update_table_ui()
            self.run_table.update_drive_plot_ui()
            self.run_table.update_response_plot_ui()
            self.run_table.update_all_voltages_ui()

            self.set_table_lock()

            # Treat load as a manual table modification: recompute prediction immediately
            self.run_table.perform_prediction()

        except Exception as exc:
            QtWidgets.QMessageBox.critical(
                self,
                "Invalid SDS Table File",
                str(exc),
            )
            return

    def _apply_single_channel_table(self, frequencies, amplitudes, delays, decays):
        if self.run_table.sds_table is None:
            raise ValueError("Run SDS table has not been initialized yet.")

        current_index = self.run_table.parent_widget.excitation_selector.currentIndex()
        current_freqs = np.array(self.run_table.sds_table["frequency"]).flatten()

        if frequencies.size != current_freqs.size:
            raise ValueError(
                f"Spreadsheet contains {frequencies.size} rows, but current SDS table "
                f"contains {current_freqs.size} rows. Spreadsheet imports must match the existing frequency count."
            )

        if not np.allclose(frequencies, current_freqs, rtol=1e-8, atol=1e-10):
            raise ValueError(
                "Spreadsheet frequency column does not match the current SDS table frequencies."
            )

        self.run_table.sds_table["amplitude"][:, current_index] = amplitudes
        self.run_table.sds_table["delay"][:, current_index] = delays
        self.run_table.sds_table["decay"][:, current_index] = decays

    def set_table_lock(self):
        manual_updates_enabled = self.allow_manual_updates_checkbox.isChecked()

        if manual_updates_enabled:
            self.run_table.lock_table(all_data=False, frequencies=True)
            self.load_table_button.setEnabled(True)
        else:
            self.run_table.lock_table(all_data=True)
            self.load_table_button.setEnabled(False)

    def set_automatic_updates(self):
        pass

    def _get_current_channel_table_arrays(self):
        if self.run_table.sds_table is None:
            raise ValueError("Run SDS table has not been initialized yet.")

        current_index = self.run_table.parent_widget.excitation_selector.currentIndex()

        frequencies = np.array(self.run_table.sds_table["frequency"]).flatten()
        amplitudes = np.array(self.run_table.sds_table["amplitude"][:, current_index]).flatten()
        delays = np.array(self.run_table.sds_table["delay"][:, current_index]).flatten()
        decays = np.array(self.run_table.sds_table["decay"][:, current_index]).flatten()

        return frequencies, amplitudes, delays, decays

    def _save_csv_table(self, filename):
        frequencies, amplitudes, delays, decays = self._get_current_channel_table_arrays()

        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["frequency", "amplitude", "delay", "decay"])
            for row in zip(frequencies, amplitudes, delays, decays):
                writer.writerow(row)

    def _save_xlsx_table(self, filename):
        frequencies, amplitudes, delays, decays = self._get_current_channel_table_arrays()

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet.cell(1, 1, "frequency")
        worksheet.cell(1, 2, "amplitude")
        worksheet.cell(1, 3, "delay")
        worksheet.cell(1, 4, "decay")

        for row_index, (freq, amp, delay, decay) in enumerate(
            zip(frequencies, amplitudes, delays, decays),
            start=2,
        ):
            worksheet.cell(row_index, 1, float(freq))
            worksheet.cell(row_index, 2, float(amp))
            worksheet.cell(row_index, 3, float(delay))
            worksheet.cell(row_index, 4, float(decay))

        workbook.save(filename)

    def save_sds_table(self):
        if self.run_table.sds_table is None:
            QtWidgets.QMessageBox.critical(
                self,
                "No SDS Table",
                "There is no SDS table to save.",
            )
            return

        filename, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save SDS Table",
            filter="NumPy Files (*.npz);;CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*.*)",
        )
        if filename == "":
            return

        _, extension = os.path.splitext(filename)
        extension = extension.lower()

        try:
            if extension == ".npz":
                np.savez(
                    filename,
                    frequency=np.array(self.run_table.sds_table["frequency"]).copy(),
                    amplitude=np.array(self.run_table.sds_table["amplitude"]).copy(),
                    delay=np.array(self.run_table.sds_table["delay"]).copy(),
                    decay=np.array(self.run_table.sds_table["decay"]).copy(),
                )
            elif extension == ".csv":
                self._save_csv_table(filename)
            elif extension == ".xlsx":
                self._save_xlsx_table(filename)
            else:
                raise ValueError(
                    f"Unsupported file extension: {extension}. " "Please use .npz, .csv, or .xlsx."
                )

        except Exception as exc:
            QtWidgets.QMessageBox.critical(
                self,
                "Failed to Save SDS Table",
                str(exc),
            )
            return
