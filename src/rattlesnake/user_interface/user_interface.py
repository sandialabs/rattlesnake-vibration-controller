# -*- coding: utf-8 -*-
"""
Controller subsystem to handle the user interface, including callback
assignment and displaying results.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import copy
import ctypes
import multiprocessing as mp
import os
import re
import sys
import time
import traceback
from typing import Any
from datetime import datetime

import netCDF4
import numpy as np
import openpyxl
import pyqtgraph
from qtpy import QtCore, QtGui, QtWidgets, uic

from rattlesnake.engine import RattlesnakeState, RattlesnakeController
from rattlesnake.utilities import (
    DIRECTORY,
    RattlesnakeError,
    GlobalCommands,
)
from rattlesnake.load_utilities import (
    save_rattlesnake_to_workbook,
    save_profile_to_workbook,
    load_profile_from_workbook,
)
from rattlesnake.profile_manager import VALID_COMMANDS, ProfileEvent
from rattlesnake.hardware.hardware_utilities import Channel, HardwareType
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.hardware.hardware_registry import HARDWARE_METADATA
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.environment.abstract_environment import EnvironmentInstructions
from rattlesnake.process.streaming import StreamType, StreamMetadata
from rattlesnake.user_interface.ui_utilities import (
    error_message_qt,
    UICommands,
    EventWatcherError,
    EventWatcher,
    Updater,
    ProfileTimer,
    ChannelMonitor,
    IPAddress,
    IPAddressManager,
)
from rattlesnake.user_interface.ui_registry import (
    UI_HARDWARE_OPTIONS,
    UI_ASK_FOR_FILE,
    UI_HARDWARE_WIDGETS,
    ENVIRONMENT_UIS,
    UI_ENVIRONMENT_OPTIONS,
)

# region Defaults
# pyqtgraph.setConfigOption('leftButtonPan',False)
pyqtgraph.setConfigOption("background", "w")
pyqtgraph.setConfigOption("foreground", "k")
QtCore.QDir.addSearchPath(
    "images", os.path.join(DIRECTORY, "user_interface", "themes", "images")
)
TASK_NAME = "UI"
VERSION = "3.1.1"
RATTLESNAKE_UI_PATH = os.path.join(
    DIRECTORY, "user_interface", "ui_files", "combined_environments_controller.ui"
)
BUFFER_ROWS = 10
MIN_ROWS = 30


# region User Interface
class RattlesnakeUI(QtWidgets.QMainWindow):
    """Main user interface from which the rattlesnake controller object is controlled."""

    def __init__(self, rattlesnake: RattlesnakeController):
        """
        Initializes user interface from an existing rattlesnake controller object.

        The rattlesnake controller object is created outside the UI so that the
        window can close without having to wait for the full rattlesnake.shutdown event to
        occur.

        Parameters
        ----------
        rattlesnake : RattlesnakeController
            The rattlesnake controller object that the UI is going to represent.
        """
        super(RattlesnakeUI, self).__init__()

        uic.loadUi(RATTLESNAKE_UI_PATH, self)
        self.channel_monitor_window = None

        # Communication objects
        self.rattlesnake = rattlesnake
        self.environment_uis = {}
        self.profile_table_list = []
        self.profile_timer_list = []
        self.theme = "Light"

        # Updater process
        self.event_thread = None
        self.event_watcher = None
        self.threadpool = QtCore.QThreadPool()
        self.gui_updater = Updater(self.gui_update_queue)
        self.threadpool.start(self.gui_updater)
        self.gui_updater.signals.update.connect(self.update_gui)

        # Storage properties
        self.hardware_file = None
        self.lanxi_ip_addresses = []

        # Complete UI layout
        self.connect_callbacks()
        self.complete_ui()

        # Store any presets to the UI
        self.load_ui_from_rattlesnake()

        # Show UI
        self.show()

        # Tell Rattlesnake it now has a gui
        self.gui_update_queue.put((UICommands.GUI_SETUP_FINISHED, None))

    def complete_ui(self):
        """
        Helper function to set up the default format of the user interface.
        """
        # Universal
        self.setMinimumWidth(500)
        # Disable all tabs except the first
        for i in range(1, self.rattlesnake_tabs.count() - 1):
            self.rattlesnake_tabs.setTabEnabled(i, False)
        self.rattlesnake_tabs.tabBar().setTabVisible(2, False)
        self.rattlesnake_tabs.tabBar().setTabVisible(3, False)
        self.channel_monitor_button.setVisible(False)
        # Set icons and window
        icon = QtGui.QIcon("logo/Rattlesnake_Icon.png")
        self.tray_icon = QtWidgets.QSystemTrayIcon(self)
        self.tray_icon.setIcon(icon)
        self.tray_icon.show()
        if sys.platform.startswith(
            "win"
        ):  # This fixes windows treating taskbar icon as python.exe
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                f"sandia.rattlesnake.{VERSION}"
            )
        self.setWindowIcon(icon)
        self.setWindowTitle("Rattlesnake Vibration Controller")
        self.change_color_theme(self.theme)

        # Channel Table
        self.table_layout.setStretch(0, 5)  # Channel table
        self.table_layout.setStretch(1, 1)  # Environments table
        self.channel_table.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeToContents
        )

        # Hardware
        available_hardware = UI_HARDWARE_OPTIONS.keys()
        self.hardware_widgets = {
            "sample_rate": [self.sample_rate_label, self.sample_rate_selector],
            "lanxi_ip": [self.lanxi_ip_address_button],
            "lanxi_sample_rate": [self.lanxi_sample_rate_selector],
            "buffer_size": [self.buffer_size_label, self.buffer_size_selector],
            "lanxi_processes": [
                self.lanxi_maximum_acquisition_processes_label,
                self.lanxi_maximum_acquisition_processes_selector,
            ],
            "integration_oversample": [
                self.integration_oversample_label,
                self.integration_oversample_selector,
            ],
            "damping_ratio": [self.damping_ratio_label, self.damping_ratio_selector],
            "task_trigger": [self.task_trigger_label, self.task_trigger_selector],
            "trigger_output": [self.trigger_output_label, self.trigger_output_selector],
            "select_file": [self.select_file_button],
        }
        self.hardware_selector.addItems(available_hardware)
        self.update_hardware_widget_visibility()

        # Environment
        available_environments = UI_ENVIRONMENT_OPTIONS.keys()
        self.add_environment_combobox.addItems(available_environments)
        self.environment_channel_table.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeToContents
        )
        self.environment_channel_table.horizontalHeader().setVisible(True)
        self.environment_channel_table.verticalHeader().setVisible(True)
        self.environment_channel_table.setColumnCount(0)
        self.environment_channel_table.hide()

        # Acquisition
        self.streaming_widgets = [
            self.no_streaming_radiobutton,
            self.profile_streaming_radiobutton,
            self.test_level_streaming_radiobutton,
            self.streaming_environment_select_combobox,
            self.immediate_streaming_radiobutton,
            self.select_streaming_file_button,
            self.manual_streaming_radiobutton,
            self.manual_streaming_trigger_button,
        ]
        self.manual_streaming_trigger_button.hide()

        # Profile
        # Create blank timer incase profile list is invalid and it never gets initialized
        self.profile_list_update_timer = QtCore.QTimer()
        self.profile_table.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeToContents
        )
        self.run_profile_widget.setEnabled(False)

    def connect_callbacks(self):
        """
        Helper function to connect callbacks to widgets in the user interface.
        """
        # Universal
        self.channel_monitor_button.clicked.connect(self.show_channel_monitor)
        self.color_theme_combobox.currentTextChanged.connect(self.change_color_theme)
        self.load_test_file_button.clicked.connect(self.load_ui_from_test_file)
        self.save_template_button.clicked.connect(self.save_template_from_ui)

        # Channel Table
        # self.sample_rate_selector.valueChanged.connect(self.sample_rate_update)
        self.lanxi_ip_address_button.clicked.connect(self.ip_lookup)
        self.channel_table.setContextMenuPolicy(QtCore.Qt.ActionsContextMenu)
        self.channel_table.itemChanged.connect(self.add_empty_channel_table_rows)
        channel_table_scroll = self.channel_table.verticalScrollBar()
        channel_table_scroll.valueChanged.connect(self.sync_environment_table)
        self.load_channel_table_button.clicked.connect(
            self.load_channel_table_from_file
        )
        self.save_channel_table_button.clicked.connect(self.save_channel_table_to_file)
        # self.assist_channel_table_checkbox.stateChanged.connect(
        #     self.assist_channel_table_init
        # )
        # Copy
        self.channel_table_action_copy = QtWidgets.QAction("Copy", self.channel_table)
        self.channel_table_action_copy.setShortcut("Ctrl+C")
        self.channel_table_action_copy.triggered.connect(self.copy_channel_table)
        self.channel_table.addAction(self.channel_table_action_copy)
        # Paste
        self.channel_table_action_paste = QtWidgets.QAction("Paste", self.channel_table)
        self.channel_table_action_paste.setShortcut("Ctrl+V")
        self.channel_table_action_paste.triggered.connect(self.paste_channel_table)
        self.channel_table.addAction(self.channel_table_action_paste)
        # Delete
        self.channel_table_action_delete = QtWidgets.QAction(
            "Delete", self.channel_table
        )
        self.channel_table_action_delete.setShortcut("Del")
        self.channel_table_action_delete.triggered.connect(self.delete_channel_table)
        self.channel_table.addAction(self.channel_table_action_delete)
        # Insert Row
        self.channel_table_action_insert_row = QtWidgets.QAction(
            "Insert Row", self.channel_table
        )
        self.channel_table_action_insert_row.triggered.connect(
            self.channel_table_insert_row
        )
        self.channel_table.addAction(self.channel_table_action_insert_row)
        # Delete Row
        self.channel_table_action_delete_row = QtWidgets.QAction(
            "Delete Row", self.channel_table
        )
        self.channel_table_action_delete_row.triggered.connect(
            self.channel_table_delete_row
        )
        self.channel_table.addAction(self.channel_table_action_delete_row)

        # Hardware
        self.hardware_selector.currentTextChanged.connect(self.hardware_update)
        self.initialize_hardware_button.clicked.connect(self.initialize_hardware)
        self.select_file_button.clicked.connect(self.select_hardware_file)

        # Environments
        environment_table_scroll = self.environment_channel_table.verticalScrollBar()
        environment_table_scroll.valueChanged.connect(self.sync_channel_table)
        self.add_environment_combobox.currentTextChanged.connect(self.add_environment)
        self.remove_environment_button.clicked.connect(self.remove_environment)
        self.environment_channel_table.horizontalHeader().sectionDoubleClicked.connect(
            self.rename_environment
        )
        self.initialize_environments_button.clicked.connect(
            self.initialize_environments
        )

        # Acquisition
        self.select_streaming_file_button.clicked.connect(self.select_streaming_file)
        self.arm_test_button.clicked.connect(self.start_acquisition)
        self.disarm_test_button.clicked.connect(self.stop_acquisition)
        self.manual_streaming_radiobutton.toggled.connect(
            self.show_hide_manual_streaming
        )
        self.manual_streaming_trigger_button.clicked.connect(self.start_stop_streaming)

        # Profiles
        self.add_profile_event_button.clicked.connect(self.add_profile_event)
        self.remove_profile_event_button.clicked.connect(self.remove_profile_event)
        self.save_profile_button.clicked.connect(self.save_profile_list)
        self.load_profile_button.clicked.connect(self.load_profile_list)
        self.initialize_profile_button.clicked.connect(self.initialize_profile)
        self.start_profile_button.clicked.connect(self.start_profile)
        self.stop_profile_button.clicked.connect(self.stop_profile)

    def change_color_theme(self, text: str):
        """Updates the color scheme of the UI"""
        if text == "Light":
            self.setStyleSheet("")
        elif text == "Dark":
            dark_theme_path = os.path.join(
                DIRECTORY, "user_interface", "themes", "dark_theme.txt"
            )
            with open(dark_theme_path, encoding="utf-8") as file:
                stylesheet = file.read()
            images_path = os.path.join(
                DIRECTORY, "user_interface", "themes", "images"
            ).replace("\\", "/")
            print(f"Images Path: {images_path}")
            stylesheet.replace(r"%%IMAGES_PATH%%", images_path)
            self.setStyleSheet(stylesheet)

    # endregion

    # region Process
    @property
    def gui_update_queue(self):
        return self.rattlesnake.queue_container.gui_update_queue

    @property
    def log_file_queue(self):
        return self.rattlesnake.queue_container.log_file_queue

    @property
    def timeout(self):
        return self.rattlesnake.timeout

    @property
    def has_system_id(self):
        if self.system_id_environment_tabs.count() != 0:
            return True
        return False

    @property
    def has_test_pred(self):
        if self.test_prediction_environment_tabs.count() != 0:
            return True
        return False

    def log(self, string):
        """Pass a message to the log_file_queue along with date/time and task name

        Parameters
        ----------
        string : str
            Message that will be written to the queue

        """
        self.log_file_queue.put(f"{datetime.now()}: {TASK_NAME} -- {string}\n")

    def display_error(self, error):
        tb = traceback.format_exc()
        self.log(f"ERROR\n\n {tb}")
        if isinstance(error, RattlesnakeError):
            self.gui_update_queue.put(
                (UICommands.ERROR, ("Rattlesnake Error", f"ERROR:\n\n{error}"))
            )
        elif isinstance(error, EventWatcherError):
            # Just print event watcher errors to console line to avoid double showing errors in UI
            print(error)
        elif isinstance(error, str):
            self.gui_update_queue.put(
                (UICommands.ERROR, ("Rattlesnake Error", f"ERROR:\n\n{error}"))
            )
        else:
            self.gui_update_queue.put(
                (UICommands.ERROR, ("Unknown Error", f"ERROR:\n\n{tb}"))
            )

    def create_event_watcher(
        self,
        ready_event_list,
        active_event_list,
        *,
        active_event_check: bool = None,
        timeout: float = None,
    ):
        if timeout is None:
            timeout = self.timeout

        if getattr(self, "event_thread", None) or getattr(self, "event_watcher", None):
            self.event_watcher.cancel()
            self.cleanup_event_watcher()
        self.event_thread = QtCore.QThread()
        self.event_watcher = EventWatcher(
            ready_event_list,
            active_event_list,
            active_event_check=active_event_check,
            timeout=timeout,
        )
        self.event_watcher.moveToThread(self.event_thread)
        self.event_thread.started.connect(self.event_watcher.run)

    def cleanup_event_watcher(self):
        if getattr(self, "event_thread", None):
            self.event_thread.quit()
            self.event_thread.wait()
            self.event_thread.deleteLater()
            self.event_thread = None
        if getattr(self, "event_watcher", None):
            self.event_watcher.deleteLater()
            self.event_watcher = None

    def update_gui(self, queue_data: tuple[UICommands, Any]):
        """Update the graphical interface for the main controller

        Parameters
        ----------
        queue_data : tuple[UICommands, Any]
            A 2-tuple consisting of ``(message, data)`` pairs where the message
            denotes what to change and the data contains the information needed
            to be displayed.
        """
        command, data = queue_data
        if command in self.environment_uis.keys():
            self.environment_uis[command].update_gui(data)
            return

        match command:
            case UICommands.ERROR:
                dialog_title, error_message = data
                error_message_qt(dialog_title, error_message)
            case UICommands.HARDWARE_STARTED:
                self.display_acquisition_started()
            case UICommands.HARDWARE_ENDED:
                self.display_acquisition_ended()
            case UICommands.COMPLETED_SYSTEM_ID:
                environment, _ = data
                print(f"System Id Completed for {environment}")
                self.rattlesnake_tabs.setTabEnabled(3, True)
                self.rattlesnake_tabs.setTabEnabled(4, True)
            case UICommands.MONITOR:
                if self.channel_monitor_window is not None:
                    if not self.channel_monitor_window.isVisible():
                        self.channel_monitor_window = None
                    else:
                        self.channel_monitor_window.update(data)
            case UICommands.STOP:
                self.disarm_test()
            case UICommands.ENABLE:
                widget = getattr(self, data)
                widget.setEnabled(True)
            case UICommands.DISABLE:
                widget = getattr(self, data)
                widget.setEnabled(False)
            case UICommands.ENABLE_TAB:
                self.rattlesnake_tabs.setTabEnabled(data, True)
                self.rattlesnake_tabs.setCurrentIndex(data)
            case UICommands.DISABLE_TAB:
                self.rattlesnake_tabs.setTabEnabled(data, False)
            case UICommands.GUI_SETUP_FINISHED:
                self.rattlesnake.setup_gui()
            case _:
                widget = getattr(self, command)
                if isinstance(widget, QtWidgets.QDoubleSpinBox):
                    widget.setValue(data)
                elif isinstance(widget, QtWidgets.QSpinBox):
                    widget.setValue(data)
                elif isinstance(widget, QtWidgets.QLineEdit):
                    widget.setText(data)
                elif isinstance(widget, QtWidgets.QListWidget):
                    widget.clear()
                    widget.addItems([f"{d:.3f}" for d in data])

    # endregion

    # region Loading
    def load_ui_from_test_file(self, filepath=None):
        """
        Callback to select file path, verify existance and load that file to
        the user interface.
        """
        if not filepath:
            filepath, _ = QtWidgets.QFileDialog.getOpenFileName(
                self,
                "Load Rattlesnake Template File",
                filter="Rattlesnake Files (*.nc4 *.xlsx);;NetCDF Files (*.nc4);;Excel Files (*.xlsx);;All Files (*.*)",
            )
            if filepath == "":
                return

        try:
            self.rattlesnake.load_rattlesnake_from_template(filepath)
        except Exception as e:  # pylint: disable=broad-exception-caught
            self.display_error(e)
            return

        self.load_ui_from_rattlesnake()

    def load_ui_from_rattlesnake(self):
        """
        Gets the current state of the rattlesnake object and formats
        user interface to represent that state.
        """
        # Get rattlesnake state
        state = self.rattlesnake.state
        has_profile = self.rattlesnake.has_profile
        has_streamed = self.rattlesnake.has_streamed

        # Reset UI
        for i in range(1, self.rattlesnake_tabs.count() - 1):
            self.rattlesnake_tabs.setTabEnabled(i, False)
        self.rattlesnake_tabs.tabBar().setTabVisible(2, False)
        self.rattlesnake_tabs.tabBar().setTabVisible(3, False)

        environment_names = list(self.environment_uis.keys())
        for environment_name in environment_names:
            self.remove_environment(None, environment_name)

        for event_idx in reversed(range(self.profile_table.rowCount())):
            self.remove_profile_event(None, event_idx)

        # Stores state to UI
        match state:
            case RattlesnakeState.INIT:
                return
            case RattlesnakeState.HARDWARE_STORE:
                self.load_ui_from_hardware()
            case RattlesnakeState.ENVIRONMENT_STORE:
                self.load_ui_from_hardware()
                self.load_ui_from_environments()
                # Enable next tab (sysid/profile)
                if self.has_system_id:
                    self.rattlesnake_tabs.setTabEnabled(2, True)
                    self.rattlesnake_tabs.setCurrentIndex(2)
                else:
                    self.rattlesnake_tabs.setTabEnabled(4, True)
                    self.rattlesnake_tabs.setCurrentIndex(4)
                if has_profile:
                    self.load_ui_from_profile()
                    self.initialize_profile()
                if has_streamed:
                    self.load_ui_from_stream_metadata()
            case RattlesnakeState.SYS_ID_ACTIVE:
                self.load_ui_from_hardware()
                self.load_ui_from_environments()
                # Enable sys id tab
                self.rattlesnake_tabs.setTabEnabled(2, True)
                self.rattlesnake_tabs.setCurrentIndex(2)
                if has_profile:
                    self.load_ui_from_profile()
                if has_streamed:
                    self.load_ui_from_stream_metadata()
            case RattlesnakeState.HARDWARE_ACTIVE:
                self.load_ui_from_hardware()
                self.load_ui_from_environments()
                if self.has_system_id:
                    self.rattlesnake_tabs.setTabEnabled(2, True)
                if self.has_test_pred:
                    self.rattlesnake_tabs.setTabEnabled(3, True)
                self.rattlesnake_tabs.setTabEnabled(4, True)
                if has_profile:
                    self.load_ui_from_profile()
                self.initialize_profile()
                self.load_ui_from_stream_metadata()
                self.display_acquisition_started()
                self.rattlesnake_tabs.setCurrentIndex(5)
            case RattlesnakeState.ENVIRONMENT_ACTIVE:
                self.load_ui_from_hardware()
                self.load_ui_from_environments()
                if self.has_system_id:
                    self.rattlesnake_tabs.setTabEnabled(2, True)
                if self.has_test_pred:
                    self.rattlesnake_tabs.setTabEnabled(3, True)
                self.rattlesnake_tabs.setTabEnabled(4, True)
                if has_profile:
                    self.load_ui_from_profile()
                self.initialize_profile()
                self.load_ui_from_stream_metadata()
                self.display_acquisition_started()
                self.rattlesnake_tabs.setCurrentIndex(5)
                # Display environment started for each active environment
                for (
                    queue_name,
                    active_event,
                ) in self.rattlesnake.event_container.environment_active_events.items():
                    if active_event.is_set():
                        environment_name = (
                            self.rattlesnake.environment_manager.environment_names[
                                queue_name
                            ]
                        )
                        self.environment_uis[
                            environment_name
                        ].display_environment_started()

    def load_ui_from_hardware(self):
        """
        Loads the channel table and hardware setup values from the hardware
        metadata object owned by the rattlesnake object to the user interface.
        """
        hardware_metadata = self.rattlesnake.hardware_metadata

        # Fill out channel table
        channel_list = hardware_metadata.channel_list
        self.channel_table.blockSignals(True)
        self.channel_table.setRowCount(len(channel_list))
        attr_list = Channel().channel_attr_list
        for row, channel in enumerate(channel_list):
            for col, attr_name in enumerate(attr_list):
                value = getattr(channel, attr_name)
                value = str(value) if value else None

                item = QtWidgets.QTableWidgetItem(value)
                self.channel_table.setItem(row, col, item)
        self.channel_table.blockSignals(False)
        self.add_empty_channel_table_rows()

        match hardware_metadata.hardware_type:
            case HardwareType.SDYNPY_SYSTEM:
                self.hardware_selector.blockSignals(True)
                self.hardware_selector.setCurrentText("SDynPy System Integration...")
                self.hardware_selector.blockSignals(False)
                self.update_hardware_widget_visibility()
                self.hardware_file = hardware_metadata.hardware_file
                self.sample_rate_selector.setValue(hardware_metadata.sample_rate)
                self.buffer_size_selector.setValue(hardware_metadata.time_per_read)
                self.integration_oversample_selector.setValue(
                    hardware_metadata.output_oversample
                )
            case _:
                self.display_error(
                    f"Loading from {hardware_metadata.hardware_type} is not yet implemented"
                )

    def load_ui_from_environments(self):
        """
        Loads the environment metadata list from the rattlesnake object to
        the user interface.
        """
        hardware_metadata = self.rattlesnake.hardware_metadata
        environment_metadata_dict = self.rattlesnake.environment_metadata

        for environment_idx, environment_metadata in enumerate(
            environment_metadata_dict.values()
        ):
            # Add environments
            environment_type = environment_metadata.environment_type
            self.add_environment(environment_type)

            environment_name = environment_metadata.environment_name
            if (
                environment_name not in self.environment_uis.keys()
            ):  # Dont rename if they were already using default name
                self.rename_environment(environment_idx, environment_name)

            self.environment_uis[environment_name].initialize_hardware(
                hardware_metadata
            )
            self.environment_uis[environment_name].set_environment_metadata(
                environment_metadata
            )
            self.environment_uis[environment_name].initialize_environment(
                environment_metadata
            )

        self.update_environment_tabs()
        streaming_environment_items = [""] + list(self.environment_uis.keys())
        self.streaming_environment_select_combobox.clear()
        self.streaming_environment_select_combobox.addItems(streaming_environment_items)
        self.rattlesnake_tabs.setTabEnabled(1, True)

    def load_ui_from_profile(self):
        """
        Loads the profile event list from the rattlesnake object to the
        user interface.
        """
        profile_event_list = self.rattlesnake.last_profile_event_list

        for profile_event in profile_event_list:
            timestamp = profile_event.timestamp
            environment_name = profile_event.environment_name
            command = profile_event.command
            data = profile_event.data

            # If command is START_ENVIRONMENT, add the instructions command so that
            # the user can remove those instructions if they desire
            if command is GlobalCommands.START_ENVIRONMENT and isinstance(
                data, EnvironmentInstructions
            ):
                self.add_profile_event()
                row = self.profile_table.rowCount() - 1
                timestamp_spinbox = self.profile_table.cellWidget(row, 0)
                timestamp_spinbox.setValue(timestamp)
                environment_combobox = self.profile_table.cellWidget(row, 1)
                environment_combobox.setCurrentText(environment_name)
                command_combobox = self.profile_table.cellWidget(row, 2)
                command_combobox.setCurrentText(
                    UICommands.SET_ENVIRONMENT_INSTRUCTIONS.label
                )
                data_item = QtWidgets.QTableWidgetItem("")
                data_item.setData(QtCore.Qt.ItemDataRole.UserRole, data)
                self.profile_table.setItem(row, 3, data_item)
                data = None

            data = str(data) if data is not None else ""
            data = data if data.strip() != "" else ""

            self.add_profile_event()
            row = self.profile_table.rowCount() - 1
            timestamp_spinbox = self.profile_table.cellWidget(row, 0)
            timestamp_spinbox.setValue(timestamp)
            environment_combobox = self.profile_table.cellWidget(row, 1)
            environment_combobox.setCurrentText(environment_name)
            command_combobox = self.profile_table.cellWidget(row, 2)
            command_combobox.setCurrentText(command.label)
            data_item = QtWidgets.QTableWidgetItem(data)
            self.profile_table.setItem(row, 3, data_item)

    def load_ui_from_stream_metadata(self):
        """
        Loads the stream metadata object from the rattlesnake object to
        the user interface.
        """
        stream_metadata = self.rattlesnake.last_stream_metadata

        match stream_metadata.stream_type:
            case StreamType.NO_STREAM:
                self.no_streaming_radiobutton.setChecked(True)
            case StreamType.PROFILE_INSTRUCTION:
                self.profile_streaming_radiobutton.setChecked(True)
            case StreamType.TEST_LEVEL:
                self.test_level_streaming_radiobutton.setChecked(True)
                self.streaming_environment_select_combobox.setCurrentText(
                    stream_metadata.test_level_environment_name
                )
            case StreamType.IMMEDIATELY:
                self.immediate_streaming_radiobutton.setChecked(True)
            case StreamType.MANUAL:
                self.manual_streaming_radiobutton.setChecked(True)

        self.streaming_file_display.setText(stream_metadata.stream_file)

    def save_template_from_ui(self):
        """
        Saves an excel template from the current rattlesnake object state. Inputs
        current saved values into the template.
        TODO implement blank environment template loading~
        """
        filepath, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Combined Environments Template",
            filter="Excel File (*.xlsx)",
        )
        if filepath == "":
            return

        try:
            # Hardware
            hardware_metadata = self.get_hardware_metadata_no_channels()
            channel_list = self.get_channel_list()
            hardware_metadata.channel_list = channel_list

            # Environments
            environment_metadata_list = []
            for environment_ui in self.environment_uis.values():
                metadata = environment_ui.get_environment_metadata(channel_list)
                environment_metadata_list.append(metadata)

            # Profiles
            profile_event_list = []
            num_rows = self.profile_table.rowCount()
            for row in range(num_rows):
                timestamp = self.profile_table.cellWidget(row, 0).value()
                environment_name = self.profile_table.cellWidget(row, 1).currentText()
                command = self.profile_table.cellWidget(row, 2).currentData()
                data_item = self.profile_table.item(row, 3)
                data_text = data_item.text() if data_item is not None else ""

                # Skip environment instructions
                if command == "Set Environment Instructions":
                    continue

                event = ProfileEvent(timestamp, environment_name, command, data_text)

                profile_event_list.append(event)

            workbook = openpyxl.Workbook()
            save_rattlesnake_to_workbook(
                workbook,
                hardware_metadata,
                environment_metadata_list,
                profile_event_list,
            )
            workbook.save(filepath)
        except Exception as e:  # pylint: disable=broad-exception-caught
            self.display_error(e)
            return

    # endregion

    # region Channel Table
    def get_channel(self, row):
        channel = Channel()
        channel_attr_list = channel.channel_attr_list
        for col in range(self.channel_table.columnCount()):
            attr = channel_attr_list[col]
            item = self.channel_table.item(row, col)
            # Check if item exists and has text
            if item and item.text().strip():
                setattr(channel, attr, item.text())

        return channel

    def get_channel_list(self):
        channel_list = []
        channel_attr_list = Channel().channel_attr_list
        for row in range(self.channel_table.rowCount()):
            channel = Channel()
            for col in range(self.channel_table.columnCount()):
                attr = channel_attr_list[col]
                item = self.channel_table.item(row, col)
                # Check if item exists and has text
                if item and item.text().strip():
                    setattr(channel, attr, item.text())

            if channel.is_empty:
                break

            channel_list.append(channel)

        return channel_list

    def sync_channel_table(self):
        """Callback to synchronize scrolling between channel tables"""
        self.channel_table.verticalScrollBar().setValue(
            self.environment_channel_table.verticalScrollBar().value()
        )

    def sync_environment_table(self):
        """Callback to synchronize scrolling between channel tables"""
        self.environment_channel_table.verticalScrollBar().setValue(
            self.channel_table.verticalScrollBar().value()
        )

    def add_empty_channel_table_rows(self, item=None):
        self.channel_table.blockSignals(True)
        num_rows = self.channel_table.rowCount()
        last_row = -1
        for row_idx in reversed(range(num_rows)):
            channel = self.get_channel(row_idx)

            if not channel.is_empty:
                last_row = row_idx + 1
                break

        desired_rows = max(last_row + BUFFER_ROWS, MIN_ROWS)
        if self.channel_table.rowCount() != desired_rows:
            self.channel_table.setRowCount(desired_rows)
            self.set_environment_table_row_count(desired_rows)

        if self.assist_channel_table_checkbox.isChecked():
            widget_range = range(num_rows, desired_rows)
            self.assist_channel_table_init(True, widget_range)

        self.channel_table.blockSignals(False)

    def set_environment_table_row_count(self, desired_rows):
        num_rows = self.environment_channel_table.rowCount()
        num_cols = self.environment_channel_table.columnCount()
        self.environment_channel_table.setRowCount(desired_rows)
        if desired_rows > num_rows:
            for row in range(num_rows, desired_rows):
                for col in range(num_cols):
                    checkbox = QtWidgets.QCheckBox()
                    checkbox.setChecked(False)
                    self.environment_channel_table.setCellWidget(row, col, checkbox)

    def copy_channel_table(self):
        """Function to copy text from channel table in a format that Excel recognizes"""
        if self.assist_channel_table_checkbox.isChecked():
            self.display_error("Please remove assist mode for copy functionality")
            return

        clipboard = QtWidgets.QApplication.clipboard()
        selected_ranges = self.channel_table.selectedRanges()
        if selected_ranges:
            # Get selected range
            selected_range = selected_ranges[0]
            copied_text = ""
            rows = range(selected_range.topRow(), selected_range.bottomRow() + 1)
            columns = range(
                selected_range.leftColumn(), selected_range.rightColumn() + 1
            )
            # Put tabs inbetween columns, newlines inbetween rows
            copied_text = []
            for row in rows:
                row_data = []
                for column in columns:
                    item = self.channel_table.item(row, column)
                    row_data.append(
                        item.text() if item else ""
                    )  # Empty cells should be "" not None
                copied_text.append("\t".join(row_data))  # Tab betewen columns
            copied_text = "\n".join(copied_text)  # Newline between rows
            clipboard.setText(copied_text)

    def paste_channel_table(self):
        """Function to paste clipboard starting from top left cell"""
        if self.assist_channel_table_checkbox.isChecked():
            self.display_error("Please remove assist mode for paste functionality")
            return

        self.channel_table.blockSignals(True)
        selection_range = self.channel_table.selectedRanges()
        self.channel_table.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.Fixed
        )
        if selection_range:
            # Get top left cell
            top_left_row = selection_range[0].topRow()
            top_left_column = selection_range[0].leftColumn()
            # Get clipboard text
            clipboard = QtWidgets.QApplication.clipboard()
            if clipboard.mimeData().hasText():
                clipboard_text = clipboard.text()
                # Split clipboard text with newlines between rows
                rows = clipboard_text.splitlines()
                # Split clipboard text with tabs between columns
                array_text = [row.split("\t") for row in rows]
                num_row = len(array_text)
                bottom_row = top_left_row + num_row
                if self.channel_table.rowCount() < bottom_row:
                    self.channel_table.setRowCount(bottom_row)
                # Paste the text into the table
                for i, row in enumerate(array_text):
                    for j, cell_text in enumerate(row):
                        cell_text = cell_text if cell_text is not None else ""
                        item = QtWidgets.QTableWidgetItem(cell_text)
                        self.channel_table.setItem(
                            top_left_row + i, top_left_column + j, item
                        )
        self.channel_table.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeToContents
        )
        self.channel_table.blockSignals(False)
        self.add_empty_channel_table_rows()

    def delete_channel_table(self):
        """Function to delete text from a channel table when delete is pressed"""
        if self.assist_channel_table_checkbox.isChecked():
            self.display_error("Please remove assist mode for delete functionality")
            return

        self.channel_table.blockSignals(True)
        selection_range = self.channel_table.selectedRanges()
        self.channel_table.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.Fixed
        )
        if selection_range:
            # Get the selected range
            selected_range = selection_range[0]
            rows = range(selected_range.topRow(), selected_range.bottomRow() + 1)
            columns = range(
                selected_range.leftColumn(), selected_range.rightColumn() + 1
            )
            # Clear the selected cells
            for row in rows:
                for column in columns:
                    clear_item = QtWidgets.QTableWidgetItem("")
                    self.channel_table.setItem(row, column, clear_item)
        self.channel_table.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeToContents
        )
        self.channel_table.blockSignals(False)
        self.add_empty_channel_table_rows()

    def channel_table_insert_row(self):
        """Function to insert row in right click menu on channel table"""
        if self.assist_channel_table_checkbox.isChecked():
            self.display_error("Please remove assist mode for insert functionality")
            return

        selection_range = self.channel_table.selectedRanges()

        if selection_range:
            # Find the top row and insert row above it
            top_row = selection_range[0].topRow()
            self.channel_table.insertRow(top_row)
            self.environment_channel_table.insertRow(top_row)
            num_col = self.environment_channel_table.columnCount()
            for col in range(num_col):
                checkbox = QtWidgets.QCheckBox()
                checkbox.setChecked(False)
                self.environment_channel_table.setCellWidget(top_row, col, checkbox)

            # Update vertical header for both tables
            row_count = self.channel_table.rowCount()
            indices = [str(i + 1) for i in range(row_count)]
            self.channel_table.setVerticalHeaderLabels(indices)
            self.environment_channel_table.setVerticalHeaderLabels(indices)

        self.add_empty_channel_table_rows()

    def channel_table_delete_row(self):
        """Function to delete row in right click menu on channel table"""
        if self.assist_channel_table_checkbox.isChecked():
            self.display_error("Please remove assist mode for delete functionality")
            return

        selected_ranges = self.channel_table.selectedRanges()
        # If channel table is clicked delete rows starting from highest index
        if selected_ranges:
            selected_range = selected_ranges[0]
            rows = range(selected_range.topRow(), selected_range.bottomRow() + 1)
            for row_idx in reversed(rows):
                self.channel_table.removeRow(row_idx)
                self.environment_channel_table.removeRow(row_idx)

        self.add_empty_channel_table_rows()

    def load_channel_table_from_file(self):
        """
        Loads a channel table using a file dialog or the specified filename
        """
        filepath, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Load Rattlesnake Template File",
            filter="Rattlesnake Files (*.nc4 *.xlsx);;NetCDF Files (*.nc4);;Excel Files (*.xlsx);;All Files (*.*)",
        )
        if filepath == "":
            return

        if not os.access(filepath, os.R_OK):
            self.display_error(f"You do not have permissions to open {filepath}")
            return

        filename, filetype = os.path.splitext(filepath)

        match filetype:
            case ".nc4":
                dataset = netCDF4.Dataset(filepath)
                channel_list = HardwareMetadata.load_channel_table_from_netcdf(dataset)
            case ".xlsx":
                workbook = openpyxl.load_workbook(filepath, read_only=True)
                channel_list = HardwareMetadata.load_channel_table_from_workbook(
                    workbook
                )

        self.channel_table.blockSignals(True)
        self.channel_table.setRowCount(len(channel_list))
        attr_list = Channel().channel_attr_list
        for row, channel in enumerate(channel_list):
            for col, attr_name in enumerate(attr_list):
                value = getattr(channel, attr_name)
                value = str(value) if value else None

                item = QtWidgets.QTableWidgetItem(value)
                self.channel_table.setItem(row, col, item)
        self.channel_table.blockSignals(False)
        self.add_empty_channel_table_rows()

    def save_channel_table_to_file(self):
        """Save the channel table to a file"""
        filepath, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Combined Environments Template",
            filter="Excel File (*.xlsx)",
        )
        if filepath == "":
            return

        channel_list = self.get_channel_list()
        workbook = openpyxl.Workbook()
        HardwareMetadata.save_channel_table_to_workbook(channel_list, workbook)
        workbook.save(filepath)

    # endregion

    # region Hardware

    def hardware_update(self, hardware_text):
        """Callback to provide options when hardware is selected"""
        hardware_type = UI_HARDWARE_OPTIONS[hardware_text]
        if hardware_type in UI_ASK_FOR_FILE:
            filename, file_filter = QtWidgets.QFileDialog.getOpenFileName(
                self, "Load a SDynPy System", filter="Numpy File (*.npz)"
            )
            # Check for 'cancel' dialog
            if filename == "" or filename is None:
                self.hardware_selector.blockSignals(True)
                self.hardware_selector.setCurrentText("Select Hardware")
                self.hardware_selector.blockSignals(False)
                return
            self.hardware_file = filename
        else:
            self.hardware_file = None

        if self.assist_channel_table_checkbox.isChecked():
            self.assist_channel_table_init(True)

        self.update_hardware_widget_visibility()

    def update_hardware_widget_visibility(self):
        """Helper function to update the visibility of the sampling parameters group box"""
        hardware_text = self.hardware_selector.currentText()
        hardware_type = UI_HARDWARE_OPTIONS[hardware_text]
        visible_widgets = UI_HARDWARE_WIDGETS.get(hardware_type, set())

        for name, widgets in self.hardware_widgets.items():
            for widget in widgets:
                widget.setVisible(name in visible_widgets)

    def ip_lookup(self):
        """Creates an IP Lookup window"""
        ipv4_pattern = r"^((25[0-5]|(2[0-4]|1[0-9]|[1-9]|)[0-9])(\.(?!$)|$)){4}$"
        ipv6_pattern = r"\[\s*([0-9a-fA-F]{1,4}:){0,7}(:[0-9a-fA-F]{1,4})*%?\d*\s*\]"
        stored_addresses = self.lanxi_ip_addresses

        bknum = []
        ipv4 = []
        ipv6 = []
        for ip_address in stored_addresses:
            bknum.append(ip_address.host_name)
            ipv4.append(ip_address.ipv4_address)
            ipv6.append(ip_address.ipv6_address)

        # Loop through table devices and append unique IP addresses
        for row in range(self.channel_table.rowCount()):
            table_item = self.channel_table.item(row, 10)
            if table_item is None:
                table_text = ""
            else:
                table_text = table_item.text()
            if re.search(ipv4_pattern, table_text) is not None:
                if table_text not in ipv4:
                    stored_addresses.append(IPAddress(None, table_text, None))
                    ipv4.append(table_text)
            elif re.search(ipv6_pattern, table_text) is not None:
                if table_text not in ipv6:
                    stored_addresses.append(IPAddress(None, None, table_text))
                    ipv6.append(table_text)
            elif table_text != "":
                if table_text not in bknum:
                    stored_addresses.append(IPAddress(table_text, None, None))
                    bknum.append(table_text)

        ip_manager = IPAddressManager(stored_addresses)
        ip_manager.exec()

    # def sample_rate_update(self):
    #     """Updates the sample rate selector based on valid available rates"""
    #     if self.hardware_selector.currentIndex() == 2:
    #         current_value = self.sample_rate_selector.value()
    #         valid_dp_sample_rates = np.array(
    #             [
    #                 16,
    #                 20,
    #                 25,
    #                 32,
    #                 40,
    #                 50,
    #                 64,
    #                 80,
    #                 100,
    #                 128,
    #                 160,
    #                 200,
    #                 256,
    #                 320,
    #                 400,
    #                 512,
    #                 640,
    #                 800,
    #                 1024,
    #                 1280,
    #                 1600,
    #                 2048,
    #                 2560,
    #                 3200,
    #                 4096,
    #                 5120,
    #                 6400,
    #                 8192,
    #                 10240,
    #                 12800,
    #                 20480,
    #                 25600,
    #                 40960,
    #                 51200,
    #                 102400,
    #             ]
    #         )
    #         closest_index = np.argmin(abs(valid_dp_sample_rates - current_value))
    #         closest_rate = valid_dp_sample_rates[closest_index]
    #         # Check if it is either one above or one below a previous rate
    #         if (
    #             current_value - closest_rate == 1
    #             and closest_index != len(valid_dp_sample_rates) - 1
    #         ):
    #             closest_index += 1
    #             closest_rate = valid_dp_sample_rates[closest_index]
    #         elif current_value - closest_rate == -1 and closest_index != 0:
    #             closest_index -= 1
    #             closest_rate = valid_dp_sample_rates[closest_index]
    #         self.sample_rate_selector.blockSignals(True)
    #         self.sample_rate_selector.setValue(closest_rate)
    #         self.sample_rate_selector.blockSignals(False)

    # def task_trigger_update(self):
    #     """Updates task trigger widgets based on other widget's selections"""
    #     if (
    #         self.hardware_selector.currentIndex() == 0
    #         and self.task_trigger_selector.currentIndex() == 2
    #     ):
    #         self.task_trigger_output_selector.show()
    #         self.task_trigger_output_label.show()
    #     else:
    #         self.task_trigger_output_selector.hide()
    #         self.task_trigger_output_label.hide()

    def select_hardware_file(self):
        filename, file_filter = QtWidgets.QFileDialog.getOpenFileName(
            self, "Load a SDynPy System", filter="Numpy File (*.npz)"
        )
        # Check for 'cancel' dialog
        if filename == "" or filename is None:
            return
        self.hardware_file = filename

    def initialize_hardware(self):
        self.log("Initializing Hardware")

        # Prevent user from initializing multiple times
        self.initialize_hardware_button.setEnabled(False)

        try:
            # Build hardware metadata
            hardware_metadata = self.get_hardware_metadata_no_channels()
            if hardware_metadata is None:
                return self.initialize_hardware_error("Invalid hardware type chosen.")

            # Get channel list from UI
            channel_list = self.get_channel_list()
            hardware_metadata.channel_list = channel_list

            # Send hardware metadata to rattlesnake
            self.rattlesnake.initialize_hardware(hardware_metadata)

            environment_channel_list = self.get_environment_channel_list()
            for environment_name, environment_ui in self.environment_uis.items():
                hardware_metadata.channel_list = environment_channel_list[
                    environment_name
                ]
                environment_ui.initialize_hardware(hardware_metadata)

        except Exception as e:
            self.initialize_hardware_error(e)
            return

        # Block until hardware metadata has been stored
        ready_event_list = [
            self.rattlesnake.event_container.acquisition_ready_event,
            self.rattlesnake.event_container.output_ready_event,
            *self.rattlesnake.environment_manager.ready_event_list,
        ]
        active_event_list = []
        self.create_event_watcher(ready_event_list, active_event_list)
        self.event_watcher.ready.connect(
            lambda metadata=hardware_metadata: self.initialize_hardware_ready(metadata)
        )
        self.event_watcher.error.connect(self.initialize_hardware_error)
        self.event_thread.start()

    def initialize_hardware_ready(self, metadata):
        # Clear QThread
        self.cleanup_event_watcher()

        # Update rattlesnake state
        self.rattlesnake.hardware_metadata = metadata

        # Unlock UI
        self.initialize_hardware_button.setEnabled(True)
        self.update_environment_tabs()
        num_environments = len(self.environment_uis)
        if num_environments == 0:
            self.rattlesnake_tabs.setTabEnabled(1, False)
        else:
            self.rattlesnake_tabs.setTabEnabled(1, True)
            self.rattlesnake_tabs.setCurrentIndex(1)

    def initialize_hardware_error(self, error_message):
        # Clear QThread
        self.cleanup_event_watcher()

        # Lock UI
        # If not acquiring, disable future tabs
        if self.rattlesnake.state in (
            RattlesnakeState.INIT,
            RattlesnakeState.HARDWARE_STORE,
            RattlesnakeState.ENVIRONMENT_STORE,
        ):
            for i in range(1, self.rattlesnake_tabs.count() - 1):
                self.rattlesnake_tabs.setTabEnabled(i, False)

        # Unlock UI
        self.initialize_hardware_button.setEnabled(True)
        self.display_error(error_message)

    def get_hardware_metadata_no_channels(self):

        hardware_text = self.hardware_selector.currentText()
        hardware_type = UI_HARDWARE_OPTIONS[hardware_text]
        if hardware_type == "Select":
            return None

        channel_list = []
        hardware_metadata_class = HARDWARE_METADATA[hardware_type]
        match hardware_type:
            case HardwareType.NI_DAQMX:
                sample_rate = self.sample_rate_selector.value()
                time_per_read = self.buffer_size_selector.value()
                time_per_write = self.buffer_size_selector.value()
                task_trigger = self.task_trigger_selector.currentIndex()
                output_trigger_generator = self.trigger_output_selector.text()
                return hardware_metadata_class(
                    channel_list, 
                    sample_rate,
                    time_per_read,
                    time_per_write,
                    task_trigger,
                    output_trigger_generator
                )

            case HardwareType.LAN_XI:
                sample_rate = 2 ** self.lanxi_sample_rate_selector.currentIndex() * 4096
                time_per_read = self.buffer_size_selector.value()
                time_per_write = self.buffer_size_selector.value()
                output_oversample = 16384 // sample_rate
                if output_oversample == 0:
                    output_oversample = 1
                maximum_acquisition_processes = self.lanxi_maximum_acquisition_processes_selector.value()
                return hardware_metadata_class(
                    channel_list, 
                    sample_rate,
                    time_per_read,
                    time_per_write,
                    output_oversample,
                    maximum_acquisition_processes
                )
            case HardwareType.DP_QUATTRO:
                return
            case HardwareType.DP_900:
                return
            case HardwareType.EXODUS:
                return
            case HardwareType.STATE_SPACE:
                sample_rate = self.sample_rate_selector.value()
                time_per_read = self.buffer_size_selector.value()
                time_per_write = self.buffer_size_selector.value()
                output_oversample = self.integration_oversample_selector.value()
                hardware_file = self.hardware_file
                return hardware_metadata_class(
                    channel_list,
                    sample_rate,
                    time_per_read,
                    time_per_write,
                    output_oversample,
                    hardware_file,
                )
            case HardwareType.SDYNPY_SYSTEM:
                sample_rate = self.sample_rate_selector.value()
                time_per_read = self.buffer_size_selector.value()
                time_per_write = self.buffer_size_selector.value()
                output_oversample = self.integration_oversample_selector.value()
                hardware_file = self.hardware_file
                return hardware_metadata_class(
                    channel_list,
                    sample_rate,
                    time_per_read,
                    time_per_write,
                    output_oversample,
                    hardware_file,
                )
            case HardwareType.SDYNPY_FRF:
                sample_rate = self.sample_rate_selector.value()
                time_per_read = self.buffer_size_selector.value()
                time_per_write = self.buffer_size_selector.value()
                output_oversample = self.integration_oversample_selector.value()
                hardware_file = self.hardware_file
                return hardware_metadata_class(
                    channel_list,
                    sample_rate,
                    time_per_read,
                    time_per_write,
                    output_oversample,
                    hardware_file,
                )
            case _:
                return None

    # endregion

    # region Environment
    def get_environment_channel_list(self):
        channel_list = self.get_channel_list()
        environment_channel_list = {}
        num_rows = self.environment_channel_table.rowCount()
        num_cols = self.environment_channel_table.columnCount()
        for col in range(num_cols):
            header_item = self.environment_channel_table.horizontalHeaderItem(col)
            environment_name = header_item.text()
            selected_channels = []

            for row in range(num_rows):
                checkbox = self.environment_channel_table.cellWidget(row, col)
                if checkbox is None:
                    continue
                if checkbox.isChecked():
                    if row < len(channel_list):
                        selected_channels.append(channel_list[row])
            environment_channel_list[environment_name] = selected_channels

        return environment_channel_list

    def update_environment_tabs(self):

        # Definition tabs
        self.environment_definition_environment_tabs.setCurrentIndex(-1)
        self.environment_definition_environment_tabs.clear()
        for environment_name, environment_ui in self.environment_uis.items():
            definition_widget = environment_ui.definition_widget
            if definition_widget is not None:
                self.environment_definition_environment_tabs.addTab(
                    definition_widget, environment_name
                )

        # System Identification tab
        self.rattlesnake_tabs.tabBar().setTabVisible(2, False)
        self.system_id_environment_tabs.setCurrentIndex(-1)
        self.system_id_environment_tabs.clear()
        for environment_name, environment_ui in self.environment_uis.items():
            system_id_widget = environment_ui.system_id_widget
            if system_id_widget is not None:
                self.system_id_environment_tabs.addTab(
                    system_id_widget, environment_name
                )
                self.rattlesnake_tabs.tabBar().setTabVisible(2, True)

        # Prediction tab
        self.rattlesnake_tabs.tabBar().setTabVisible(3, False)
        self.test_prediction_environment_tabs.setCurrentIndex(-1)
        self.test_prediction_environment_tabs.clear()
        for environment_name, environment_ui in self.environment_uis.items():
            prediction_widget = environment_ui.prediction_widget
            if prediction_widget is not None:
                self.test_prediction_environment_tabs.addTab(
                    prediction_widget, environment_name
                )
                self.rattlesnake_tabs.tabBar().setTabVisible(3, True)

        # Run tab
        self.run_environment_tabs.setCurrentIndex(-1)
        self.run_environment_tabs.clear()
        for environment_name, environment_ui in self.environment_uis.items():
            run_widget = environment_ui.run_widget
            if run_widget is not None:
                self.run_environment_tabs.addTab(run_widget, environment_name)

        # Disable run tabs
        for i in range(self.run_environment_tabs.count()):
            self.run_environment_tabs.widget(i).setEnabled(False)

    def add_environment(self, environment_type: str | EnvironmentType):
        """Function used to add an environment"""
        # If comming from UI, environment_type will be text in combobox
        if isinstance(environment_type, str):
            environment_type = UI_ENVIRONMENT_OPTIONS[environment_type]

        if environment_type is None:
            return

        idx = 0
        environment_name = f"{environment_type.name} {idx}"
        while environment_name in self.environment_uis.keys():
            idx += 1
            environment_name = f"{environment_type.name} {idx}"

        environment_ui_class = ENVIRONMENT_UIS[environment_type]
        environment_ui = environment_ui_class(environment_name, self.rattlesnake)

        # Update environment UIs and channel table
        self.environment_uis[environment_name] = environment_ui
        new_col = self.environment_channel_table.columnCount()
        self.environment_channel_table.insertColumn(new_col)
        self.environment_channel_table.setHorizontalHeaderItem(
            new_col, QtWidgets.QTableWidgetItem(environment_name)
        )
        self.environment_channel_table.show()

        # Set checkboxes
        channel_list = self.get_channel_list()
        num_channels = len(channel_list)
        num_rows = self.environment_channel_table.rowCount()
        for row in range(num_rows):
            checkbox = QtWidgets.QCheckBox()
            checkbox.setChecked(row < num_channels)
            self.environment_channel_table.setCellWidget(row, new_col, checkbox)

        # Reset add environment combobox
        self.add_environment_combobox.setCurrentIndex(0)

    def remove_environment(self, clicked=None, environment_name=None):
        # Find selected ranges on the environment channel table
        if environment_name:
            for col in range(self.environment_channel_table.columnCount()):
                item = self.environment_channel_table.horizontalHeaderItem(col)
                if item and item.text() == environment_name:
                    columns = [col]
        else:
            selected_ranges = self.environment_channel_table.selectedRanges()
            if not selected_ranges:
                self.display_error(
                    "Please select an environment in environment channel table to remove"
                )
                return
            # Remove selected columns from environment table and environment_uis
            selected_range = selected_ranges[0]
            columns = range(
                selected_range.leftColumn(), selected_range.rightColumn() + 1
            )

        for col in sorted(columns, reverse=True):
            header_item = self.environment_channel_table.horizontalHeaderItem(col)
            environment_name = header_item.text()
            self.environment_uis.pop(environment_name)
            self.environment_channel_table.removeColumn(col)

        # If all environments are removed, hide environment channel table
        if len(self.environment_uis) == 0:
            self.environment_channel_table.hide()

    def rename_environment(self, col_idx: int, new_name: str = None):
        """Function to rename an environment

        Parameters
        ----------
        index : int :
            The index of the environment to rename
        """

        # Pull header text from environment_channel_table
        header_item = self.environment_channel_table.horizontalHeaderItem(col_idx)
        current_name = header_item.text()

        # If name not given, ask user for a name
        if not new_name:
            # Create dialog box to get a new name
            new_name, ok_chosen = QtWidgets.QInputDialog.getText(
                self, "Rename Tab", "Enter new tab name:", text=current_name
            )
            if not ok_chosen:
                return
            new_name = new_name.strip()
            if not new_name:
                return

        # Make sure name does not already exist
        if new_name in self.environment_uis:
            self.display_error(
                "The new name already exists. Please choose a different name."
            )
            return

        # Replace old name in dict with new name while keeping order
        # This is scuffed but is very specific to this case
        ordered_dict = {}
        for environment_name, environment_ui in self.environment_uis.items():
            if environment_name == current_name:
                environment_ui.environment_name = new_name
                ordered_dict[new_name] = environment_ui
            else:
                ordered_dict[environment_name] = environment_ui
        self.environment_uis = ordered_dict
        header_item.setText(new_name)

    def initialize_environments(self):
        self.log("Initializing Environment")

        # Prevent user from initializing multiple times
        self.initialize_environments_button.setEnabled(False)

        try:
            # Build environment metadata list
            environment_metadata_list = []
            for environment_ui in self.environment_uis.values():
                metadata = environment_ui.get_environment_metadata(
                    self.rattlesnake.hardware_metadata.channel_list
                )
                environment_ui.initialize_environment(metadata)
                environment_metadata_list.append(metadata)

            # Send hardware metadata to rattlesnake
            environment_metadata = self.rattlesnake.initialize_environments(
                environment_metadata_list
            )

        except Exception as e:
            self.initialize_environments_error(e)
            return

        # Block until environment metadata has been stored
        ready_event_list = [
            self.rattlesnake.event_container.acquisition_ready_event,
            self.rattlesnake.event_container.output_ready_event,
            *self.rattlesnake.environment_manager.ready_event_list,
        ]
        active_event_list = []
        self.create_event_watcher(ready_event_list, active_event_list)
        self.event_watcher.ready.connect(
            lambda metadata=environment_metadata: self.initialize_environments_ready(
                metadata
            )
        )
        self.event_watcher.error.connect(self.initialize_environments_error)
        self.event_thread.start()

    def initialize_environments_ready(self, metadata):
        # Clear QThread
        self.cleanup_event_watcher()

        # Update rattlesnake state
        self.rattlesnake.environment_metadata = metadata

        # Unlock UI
        streaming_environment_items = [""] + list(self.environment_uis.keys())
        self.streaming_environment_select_combobox.clear()
        self.streaming_environment_select_combobox.addItems(streaming_environment_items)
        self.initialize_environments_button.setEnabled(True)

        if self.has_system_id:
            self.rattlesnake_tabs.setTabEnabled(2, True)
            self.rattlesnake_tabs.setCurrentIndex(2)
        elif self.has_test_pred:
            self.rattlesnake_tabs.setTabEnabled(3, True)
            self.rattlesnake_tabs.setCurrentIndex(3)
        else:
            self.rattlesnake_tabs.setTabEnabled(4, True)
            self.rattlesnake_tabs.setCurrentIndex(4)

    def initialize_environments_error(self, error_message):
        # Clear QThread
        self.cleanup_event_watcher()

        # Update rattlesnake state
        self.rattlesnake.environment_metadata = []

        # Lock future UI
        self.streaming_environment_select_combobox.clear()
        # If not acquiring, disable future tabs
        if self.rattlesnake.state in (
            RattlesnakeState.INIT,
            RattlesnakeState.HARDWARE_STORE,
            RattlesnakeState.ENVIRONMENT_STORE,
        ):
            for i in range(2, self.rattlesnake_tabs.count() - 1):
                self.rattlesnake_tabs.setTabEnabled(i, False)

        # Unlock UI
        self.initialize_environments_button.setEnabled(True)
        self.display_error(error_message)

    # endregion

    # region Acquisition
    def enable_ready_run_tabs(self):
        for i in range(self.run_environment_tabs.count()):
            environment_name = self.run_environment_tabs.tabText(i)
            queue_name = self.rattlesnake.environment_manager.queue_names_dict[
                environment_name
            ]
            bool_ready = (
                self.rattlesnake.environment_manager.acquisition_ready_environments[
                    queue_name
                ]
            )

            if bool_ready:
                self.run_environment_tabs.widget(i).setEnabled(True)
            else:
                self.run_environment_tabs.widget(i).setEnabled(False)

    def show_channel_monitor(self):
        """
        Shows the channel monitor window.
        """
        if (self.channel_monitor_window is None) or (
            not self.channel_monitor_window.isVisible()
        ):
            self.channel_monitor_window = ChannelMonitor(
                None, self.global_daq_parameters
            )
        else:
            pass  # TODO Need to raise the window to the front, or close and reopen

    def show_hide_manual_streaming(self):
        """Shows or hides the manual streaming button depending on which streaming type is chosen"""
        if self.manual_streaming_radiobutton.isChecked():
            self.manual_streaming_trigger_button.setVisible(True)
        else:
            self.manual_streaming_trigger_button.setVisible(False)

    def start_stop_streaming(self):
        """Starts or stops streaming manually"""
        if self.manual_streaming_trigger_button.text() == "Stop\nStreaming":
            self.manual_streaming_trigger_button.setText("Start\nStreaming")
            self.rattlesnake.stop_streaming()
        else:
            self.manual_streaming_trigger_button.setText("Stop\nStreaming")
            self.rattlesnake.start_streaming()

    def get_stream_metadata(self):
        stream_file = self.streaming_file_display.text()

        if self.no_streaming_radiobutton.isChecked():
            stream_type = StreamType.NO_STREAM
            stream_metadata = StreamMetadata(stream_type)
        elif self.profile_streaming_radiobutton.isChecked():
            stream_type = StreamType.PROFILE_INSTRUCTION
            stream_metadata = StreamMetadata(stream_type, stream_file)
        elif self.test_level_streaming_radiobutton.isChecked():
            stream_type = StreamType.TEST_LEVEL
            test_level_environment_name = (
                self.streaming_environment_select_combobox.currentText()
            )
            stream_metadata = StreamMetadata(
                stream_type, stream_file, test_level_environment_name
            )
        elif self.immediate_streaming_radiobutton.isChecked():
            stream_type = StreamType.IMMEDIATELY
            stream_metadata = StreamMetadata(stream_type, stream_file)
        elif self.manual_streaming_radiobutton.isChecked():
            stream_type = StreamType.MANUAL
            stream_metadata = StreamMetadata(stream_type, stream_file)
        else:
            stream_metadata = StreamMetadata()  # Default incase of error

        return stream_metadata

    def select_streaming_file(self):
        """Selects a file to stream data to disk"""
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Select NetCDF File to Save Control Data",
            filter="NetCDF File (*.nc4)",
        )
        if filename == "":
            return
        self.streaming_file_display.setText(filename)

    def display_acquisition_started(self):
        # Acquisition
        self.arm_test_button.setEnabled(False)
        self.disarm_test_button.setEnabled(True)
        # Environment
        self.enable_ready_run_tabs()
        # Streaming
        for widget in self.streaming_widgets:
            widget.setEnabled(False)
        self.manual_streaming_trigger_button.setEnabled(True)
        # Profile
        self.run_profile_widget.setEnabled(True)

    def display_acquisition_ended(self):
        # Acquisition
        self.arm_test_button.setEnabled(True)
        self.disarm_test_button.setEnabled(False)
        # Environment
        for i in range(self.run_environment_tabs.count()):
            self.run_environment_tabs.widget(i).setEnabled(False)
        # Streaming
        for widget in self.streaming_widgets:
            widget.setEnabled(True)
        self.manual_streaming_trigger_button.setEnabled(False)
        # Profile
        self.run_profile_widget.setEnabled(False)

    def start_acquisition(self):
        self.log("Starting hardware acquistion")
        self.arm_test_button.setEnabled(False)
        for widget in self.streaming_widgets:
            widget.setEnabled(False)

        try:
            stream_metadata = self.get_stream_metadata()

            self.rattlesnake.start_acquisition(stream_metadata)
        except Exception as e:
            self.start_acquisition_error(e)

        ready_event_list = [
            self.rattlesnake.event_container.streaming_ready_event,
        ]
        active_event_list = [
            self.rattlesnake.event_container.acquisition_active_event,
            self.rattlesnake.event_container.output_active_event,
        ]
        self.create_event_watcher(
            ready_event_list, active_event_list, active_event_check=True
        )
        self.event_watcher.ready.connect(self.start_acqusition_ready)
        self.event_watcher.error.connect(self.start_acquisition_error)
        self.event_thread.start()

    def start_acqusition_ready(self):
        self.cleanup_event_watcher()

        # Unlock UI
        self.display_acquisition_started()

    def start_acquisition_error(self, error):
        self.cleanup_event_watcher()

        # Show error
        self.display_error(error)

        # Unlock UI
        if self.rattlesnake.state is RattlesnakeState.HARDWARE_ACTIVE:
            self.display_acquisition_started()
        else:
            self.display_acquisition_ended()

    def stop_acquisition(self):
        # Stop acquisition is a special case where this needs to stop the test no matter what
        self.stop_profile()
        self.log("Stopping hardware acquisition")
        self.disarm_test_button.setEnabled(False)
        for i in range(self.run_environment_tabs.count()):
            self.run_environment_tabs.widget(i).setEnabled(False)
        self.manual_streaming_trigger_button.setEnabled(False)

        try:
            self.rattlesnake.stop_acquisition()
        except Exception as e:
            self.stop_acquisition_error(e)

        # Make sure event watcher will work since disarm daq is usually an "oh crap" moment
        ready_event_list = [
            self.rattlesnake.event_container.controller_ready_event,
        ]
        active_event_list = [
            self.rattlesnake.event_container.acquisition_active_event,
            self.rattlesnake.event_container.output_active_event,
            *self.rattlesnake.environment_manager.active_event_list,
        ]
        self.create_event_watcher(
            ready_event_list, active_event_list, active_event_check=False
        )
        self.event_watcher.ready.connect(self.stop_acquistion_ready)
        self.event_watcher.error.connect(self.stop_acquisition_error)
        self.event_thread.start()

    def stop_acquistion_ready(self):
        self.cleanup_event_watcher()

        # Unlock UI
        self.display_acquisition_ended()

    def stop_acquisition_error(self, error):
        self.cleanup_event_watcher()

        # Show error
        self.display_error(error)

        # Unlock UI
        if self.rattlesnake.state is RattlesnakeState.HARDWARE_ACTIVE:
            self.display_acquisition_started()
        else:
            self.display_acquisition_ended()

    # endregion

    # region Profile
    def load_profile_list(self, filepath=None):
        if not filepath:
            filepath, _ = QtWidgets.QFileDialog.getOpenFileName(
                self,
                "Load Profile Excel File",
                filter="Excel Files (*.xlsx);;All Files (*.*)",
            )
            if filepath == "":
                return

        if not os.access(filepath, os.R_OK):
            self.display_error(f"You do not have permissions to open {filepath}")
            return

        environment_types = {
            environment_name: self.environment_uis[environment_name].environment_type
            for environment_name in self.environment_uis.keys()
        }
        filename, filetype = os.path.splitext(filepath)
        match filetype:
            case ".nc4":
                self.display_error(f"Netcdf files do not store profile lists")
            case ".xlsx":
                workbook = openpyxl.load_workbook(filepath, read_only=True)
                profile_event_list = load_profile_from_workbook(
                    workbook, environment_types
                )
                self.rattlesnake.initialize_profile_event_list(profile_event_list)
                self.load_ui_from_profile()

    def save_profile_list(self, filepath=None):
        if not filepath:
            filepath, _ = QtWidgets.QFileDialog.getSaveFileName(
                self,
                "Save Combined Environments Template",
                filter="Excel File (*.xlsx)",
            )
        if filepath == "":
            return

        profile_event_list = []
        num_rows = self.profile_table.rowCount()
        for row in range(num_rows):
            timestamp = self.profile_table.cellWidget(row, 0).value()
            environment_name = self.profile_table.cellWidget(row, 1).currentText()
            command = self.profile_table.cellWidget(row, 2).currentData()
            data_item = self.profile_table.item(row, 3)
            data_text = data_item.text() if data_item is not None else ""

            # Skip environment instructions
            if command == "Set Environment Instructions":
                continue

            event = ProfileEvent(timestamp, environment_name, command, data_text)

            profile_event_list.append(event)

        workbook = openpyxl.Workbook()
        save_profile_to_workbook(workbook, profile_event_list)
        workbook.save(filepath)

    def add_profile_event(self, clicked=None):
        # Dont let user run profile until they intialize
        self.run_profile_widget.hide()
        # Create new row in profile table
        selected_row = self.profile_table.rowCount()
        self.profile_table.insertRow(selected_row)

        # Add spinbox to select time
        timestamp_spinbox = QtWidgets.QDoubleSpinBox()
        timestamp_spinbox.setMaximum(1e6)
        self.profile_table.setCellWidget(selected_row, 0, timestamp_spinbox)
        timestamp_spinbox.valueChanged.connect(self.update_profile_plot)

        # Add combobox with environment names
        environment_combobox = QtWidgets.QComboBox()
        environment_combobox.addItem("Global")
        for environment_name in self.environment_uis.keys():
            environment_combobox.addItem(environment_name)
        self.profile_table.setCellWidget(selected_row, 1, environment_combobox)
        environment_combobox.currentTextChanged.connect(
            lambda text, row=selected_row: self.update_profile_operations(text, row)
        )

        operation_combobox = QtWidgets.QComboBox()
        valid_commands = VALID_COMMANDS["Global"]
        valid_operations = [command.label for command in valid_commands]
        for command, operation in zip(valid_commands, valid_operations):
            operation_combobox.addItem(operation, userData=command)
        self.profile_table.setCellWidget(selected_row, 2, operation_combobox)
        operation_combobox.currentIndexChanged.connect(self.update_profile_plot)

        data_item = QtWidgets.QTableWidgetItem()
        self.profile_table.setItem(selected_row, 3, data_item)

        self.update_profile_plot()

    def remove_profile_event(self, clicked=None, selected_row=None):
        # Dont let user run profile until they intialize
        self.run_profile_widget.hide()

        # Find selected row
        if selected_row is None:
            selected_row = self.profile_table.currentRow()

        # Remove selected row
        if selected_row >= 0:
            self.profile_table.removeRow(selected_row)

        self.update_profile_plot()

    def update_profile_operations(self, environment_name, row):
        """Update profile operations given a selected environment"""
        if environment_name == "Global":
            environment_type = "Global"
        else:
            environment_type = self.environment_uis[environment_name].environment_type

        # Find valid commands for that environment type
        valid_commands = VALID_COMMANDS[environment_type]
        valid_operations = [command.label for command in valid_commands]

        # Set operation combobox to those commands
        operation_combobox = self.profile_table.cellWidget(row, 2)
        operation_combobox.blockSignals(True)
        operation_combobox.clear()
        for command, operation in zip(valid_commands, valid_operations):
            operation_combobox.addItem(operation, userData=command)
        operation_combobox.blockSignals(False)

        self.update_profile_plot()

    def update_profile_plot(self):
        """Updates the plot of profile events"""
        # Format plot
        plot_item = self.profile_timeline_plot.getPlotItem()
        plot_item.clear()
        plot_item.showGrid(True, True, 0.25)
        plot_item.disableAutoRange()

        if self.theme == "Light":
            text_color = (0, 0, 0)
        else:
            text_color = (255, 255, 255)
        max_time = 0
        for row in range(self.profile_table.rowCount()):
            timestamp = self.profile_table.cellWidget(row, 0).value()
            environment_index = self.profile_table.cellWidget(row, 1).currentIndex()
            operation = self.profile_table.cellWidget(row, 2).currentText()
            data_item = self.profile_table.item(row, 3)
            data = data_item.text() if data_item is not None else ""
            data = data if data.strip() != "" else ""

            # Add point and text to plot at correct location
            plot_item.plot(
                [timestamp], [environment_index], pen=None, symbol="o", pxMode=True
            )
            text_item = pyqtgraph.TextItem(
                f"{row + 1}: " + operation + (": " + data), color=text_color, angle=-15
            )
            plot_item.addItem(text_item)
            text_item.setPos(timestamp, environment_index)

            if timestamp > max_time:
                max_time = timestamp

        # Label axis and scale range
        environment_names = list(self.environment_uis.keys())
        axis = plot_item.getAxis("left")
        axis.setTicks(
            [[(i, name) for i, name in enumerate(["Global"] + environment_names)], []]
        )
        plot_item.setXRange(0, max_time * 1.1)
        plot_item.setYRange(-1, len(environment_names))

    def display_event_strings(self, event_string_list):
        self.upcoming_instructions_list.clear()
        self.upcoming_instructions_list.addItems(event_string_list)

    def reset_profile_ui_timers(self):
        profile_timer_list = []
        event_string_list = []
        for event in self.profile_table_list:
            timestamp = event.timestamp
            environment_name = event.environment_name
            command = event.command
            data_item = event.data
            data_text = data_item.text() if data_item is not None else ""

            timer = ProfileTimer(timestamp, environment_name, command, data_text)
            timer.setSingleShot(True)
            timer.timeout.connect(
                lambda name=environment_name: self.switch_profile_environment(name)
            )
            profile_timer_list.append(timer)

            event_string = (
                f"{timestamp:0.2f} {environment_name} {command.label} {data_text}"
            )
            event_string_list.append(event_string)

        self.upcoming_instructions_list.clear()
        self.upcoming_instructions_list.addItems(event_string_list)
        self.profile_timer_list = profile_timer_list

    def start_profile_event_timers(self):
        for timer in self.profile_timer_list:
            timer.start(int(timer.timestamp * 1000))

        self.profile_list_update_timer = QtCore.QTimer()
        self.profile_list_update_timer.timeout.connect(self.update_profile_list)
        self.profile_list_update_timer.start(250)

    def update_profile_list(self):
        """Updates the list of upcoming profile events."""
        event_string_list = []
        for timer in self.profile_timer_list:
            remaining_time = timer.remainingTime() / 1000
            environment_name = timer.environment_name
            command = timer.command
            data = timer.data

            if remaining_time > 0:
                event_string = (
                    f"{remaining_time:0.2f} {environment_name} {command.label} {data}"
                )
                event_string_list.append(event_string)

        self.upcoming_instructions_list.clear()
        self.upcoming_instructions_list.addItems(event_string_list)
        if not event_string_list:
            self.reset_profile_ui_timers()
            self.profile_list_update_timer.stop()

    def switch_profile_environment(self, environment_name):
        if self.show_profile_change_checkbox.isChecked():
            for i in range(self.run_environment_tabs.count()):
                if self.run_environment_tabs.tabText(i) == environment_name:
                    self.run_environment_tabs.setCurrentIndex(i)
                    break

    def initialize_profile(self):
        self.log("Initializing Profile Event List")
        sort_list = []
        num_rows = self.profile_table.rowCount()
        for row in range(num_rows):
            timestamp = self.profile_table.cellWidget(row, 0).value()
            environment_name = self.profile_table.cellWidget(row, 1).currentText()
            command = self.profile_table.cellWidget(row, 2).currentData()
            data_item = self.profile_table.item(row, 3)
            data_text = data_item.text() if data_item is not None else ""

            profile_event = ProfileEvent(
                timestamp, environment_name, command, data_item
            )

            sort_list.append((timestamp, row, profile_event))

        # Order profile table in ascending timestamp, then row number
        sort_list.sort(key=lambda x: (x[0], x[1]))
        profile_table_list = [row[2] for row in sort_list]
        self.profile_table_list = profile_table_list

        # Build timers and reset event list in run tab
        self.reset_profile_ui_timers()

        # Unlock UI
        if len(profile_table_list) == 0:
            self.run_profile_widget.hide()
        else:
            self.run_profile_widget.show()
        self.rattlesnake_tabs.setTabEnabled(5, True)
        self.rattlesnake_tabs.setCurrentIndex(5)

    def start_profile(self):
        """
        Build valid profile_event_list and give it to the controller.

        This function starts by walking through the initialized event list and storing information/building
        instructions so that the event list is adaptive to the user's inputs into the run tab UI.
        """
        self.start_profile_button.setEnabled(False)
        for i in range(self.run_environment_tabs.count()):
            self.run_environment_tabs.widget(i).setEnabled(False)
        try:
            # Figure out initial environment UI
            initial_instructions = {}
            for environment_name, environment_ui in self.environment_uis.items():
                environment_instructions = environment_ui.get_environment_instructions()
                initial_instructions[environment_name] = environment_instructions

            # Walk through events and store to UI/build event.data
            profile_event_list = []
            max_timestamp = 0
            for event in self.profile_table_list:
                timestamp = event.timestamp
                environment_name = event.environment_name
                command = event.command
                data_item = event.data

                # For start_environment pull instruction from current UI
                if command is GlobalCommands.START_ENVIRONMENT:
                    data = self.environment_uis[
                        environment_name
                    ].get_environment_instructions()
                elif (
                    command is UICommands.SET_ENVIRONMENT_INSTRUCTIONS
                ):  # Store data to the UI but dont add it as an event
                    data = data_item.data(QtCore.Qt.ItemDataRole.UserRole)
                    self.environment_uis[environment_name].set_environment_instructions(
                        data
                    )
                    continue
                elif isinstance(command, GlobalCommands):
                    data = None
                else:  # Convert data str to correct data type
                    data = data_item.text() if data_item is not None else ""
                    data = data if data.strip() != "" else ""
                    validator = command.valid_data()[
                        command
                    ]  # This syntax is a weird consequence of enums
                    try:
                        if validator is type(None):
                            data = None
                        else:
                            data = validator(data)
                    except:
                        raise RattlesnakeError(
                            f"{environment_name} profile event {command} requires {validator} data type"
                        )

                # Update environment ui
                if environment_name != "Global" and command not in (
                    GlobalCommands.START_ENVIRONMENT,
                    GlobalCommands.STOP_ENVIRONMENT,
                ):
                    self.environment_uis[environment_name].update_gui((command, data))

                # Add to profile_event_list
                profile_event = ProfileEvent(timestamp, environment_name, command, data)
                profile_event_list.append(profile_event)

                if timestamp > max_timestamp:
                    max_timestamp = timestamp

            # Reset UI to initial UI
            for environment_name, instruction in initial_instructions.items():
                self.environment_uis[environment_name].set_environment_instructions(
                    instruction
                )

            # Start Rattlesnake from profile_event_list
            self.rattlesnake.start_profile(profile_event_list)
            self.start_profile_event_timers()
        except Exception as e:
            self.start_profile_error(e)
            return

        ready_event_list = [self.rattlesnake.event_container.controller_ready_event]
        active_event_list = []
        self.create_event_watcher(
            ready_event_list, active_event_list, timeout=max_timestamp + self.timeout
        )
        self.event_watcher.ready.connect(self.profile_closed_out)
        self.event_watcher.error.connect(self.start_profile_error)
        self.event_thread.start()

        self.stop_profile_button.setEnabled(True)

    def profile_closed_out(self):
        self.cleanup_event_watcher()

        self.reset_profile_ui_timers()
        self.profile_list_update_timer.stop()

        self.enable_ready_run_tabs()
        self.start_profile_button.setEnabled(True)
        self.stop_profile_button.setEnabled(False)

    def start_profile_error(self, error):
        self.cleanup_event_watcher()

        self.reset_profile_ui_timers()
        self.profile_list_update_timer.stop()

        # Show error
        self.display_error(error)

        # Unlock UI
        self.start_profile_button.setEnabled(True)
        self.enable_ready_run_tabs()

    def stop_profile(self):
        self.stop_profile_button.setEnabled(False)
        try:
            self.rattlesnake.stop_profile()
        except Exception as e:
            self.stop_profile_error(e)
            return

    def stop_profile_error(self, error):
        # Show error
        self.display_error(error)

        # Unlock UI
        self.stop_profile_button.setEnabled(True)

    # endregion

    # region Shutdown
    def closeEvent(self, event: QtGui.QCloseEvent):  # pylint: disable=invalid-name
        """
        Event triggered when closing the software to gracefully shut down.

        Parameters
        ----------
        event : QtGui.QCloseEvent
            The close event, which is accepted.
        """
        self.gui_update_queue.put((GlobalCommands.QUIT, None))
        self.threadpool.waitForDone()

        event.accept()

    # endregion
