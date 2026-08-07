import os

import netCDF4 as nc4
import openpyxl

from rattlesnake.engine import RattlesnakeController
from rattlesnake.load_utilities import (
    load_metadata_from_netcdf,
    load_metadata_from_workbook,
    save_rattlesnake_to_workbook,
)
from rattlesnake.utilities import RattlesnakeError, save_rattlesnake_to_netcdf
from rattlesnake.testing.builders import initialize_rattlesnake_object


def save_hardware_metadata_to_file(rattlesnake: RattlesnakeController, filename: str):
    extension = os.path.splitext(filename)[1]
    if extension == ".nc4":
        with nc4.Dataset(filename, "w", format="NETCDF4") as dataset:
            save_rattlesnake_to_netcdf(
                dataset,
                hardware_metadata=rattlesnake.hardware_metadata,
            )
    elif extension == ".xlsx":
        workbook = openpyxl.Workbook()
        save_rattlesnake_to_workbook(
            workbook,
            hardware_metadata=rattlesnake.hardware_metadata,
        )
        workbook.save(filename)
    else:
        raise RattlesnakeError(
            f"Rattlesnake only saves .xlsx or .nc4 files, got {filename!r}"
        )


def save_environment_metadata_to_file(
    rattlesnake: RattlesnakeController, filename: str
):
    extension = os.path.splitext(filename)[1]
    if extension == ".nc4":
        with nc4.Dataset(filename, "w", format="NETCDF4") as dataset:
            save_rattlesnake_to_netcdf(
                dataset, environment_metadata_dict=rattlesnake.environment_metadata
            )
    elif extension == ".xlsx":
        workbook = openpyxl.Workbook()
        save_rattlesnake_to_workbook(
            workbook,
            environment_metadata_dict=rattlesnake.environment_metadata,
        )
        workbook.save(filename)
    else:
        raise RattlesnakeError(
            f"Rattlesnake only saves .xlsx or .nc4 files, got {filename!r}"
        )


def save_profile_event_list_to_file(rattlesnake: RattlesnakeController, filename: str):
    extension = os.path.splitext(filename)[1]
    if extension == ".xlsx":
        workbook = openpyxl.Workbook()
        save_rattlesnake_to_workbook(
            workbook,
            profile_event_list=rattlesnake.last_profile_event_list,
        )
        workbook.save(filename)
    else:
        raise RattlesnakeError(
            f"Rattlesnake only saves profiles to .xlsx files, got {filename!r}"
        )


def save_rattlesnake_state_to_file(rattlesnake: RattlesnakeController, filename: str):
    extension = os.path.splitext(filename)[1]
    if extension == ".nc4":
        with nc4.Dataset(filename, "w", format="NETCDF4") as dataset:
            save_rattlesnake_to_netcdf(
                dataset,
                hardware_metadata=rattlesnake.hardware_metadata,
                environment_metadata_dict=rattlesnake.environment_metadata,
            )
    elif extension == ".xlsx":
        workbook = openpyxl.Workbook()
        save_rattlesnake_to_workbook(
            workbook,
            hardware_metadata=rattlesnake.hardware_metadata,
            environment_metadata_dict=rattlesnake.environment_metadata,
            profile_event_list=rattlesnake.last_profile_event_list,
        )
        workbook.save(filename)
    else:
        raise RattlesnakeError(
            f"Rattlesnake only saves .xlsx or .nc4 files, got {filename!r}"
        )


def load_hardware_metadata_from_file(filename: str):
    extension = os.path.splitext(filename)[1]
    if extension == ".nc4":
        dataset = nc4.Dataset(filename)
        hardware_metadata, _ = load_metadata_from_netcdf(dataset)
        return hardware_metadata
    elif extension == ".xlsx":
        workbook = openpyxl.load_workbook(filename)
        hardware_metadata, _, _ = load_metadata_from_workbook(workbook)
        return hardware_metadata
    else:
        raise RattlesnakeError(
            f"Rattlesnake only loads .xlsx or .nc4 files, got {filename!r}"
        )


def load_environment_metadata_from_file(filename: str):
    extension = os.path.splitext(filename)[1]
    if extension == ".nc4":
        dataset = nc4.Dataset(filename)
        _, environment_metadata_list = load_metadata_from_netcdf(dataset)
        return environment_metadata_list
    elif extension == ".xlsx":
        workbook = openpyxl.load_workbook(filename)
        _, environment_metadata_list, _ = load_metadata_from_workbook(workbook)
        return environment_metadata_list
    else:
        raise RattlesnakeError(
            f"Rattlesnake only loads .xlsx or .nc4 files, got {filename!r}"
        )


def load_profile_event_list_from_file(filename: str):
    extension = os.path.splitext(filename)[1]
    if extension == ".xlsx":
        workbook = openpyxl.load_workbook(filename)
        _, _, profile_event_list = load_metadata_from_workbook(workbook)
        return profile_event_list
    else:
        raise RattlesnakeError(
            f"Rattlesnake only loads profile event lists from .xlsx, got {filename!r}"
        )


def load_rattlesnake_from_file(filename: str):
    extension = os.path.splitext(filename)[1]
    if extension == ".nc4":
        dataset = nc4.Dataset(filename)
        hardware_metadata, environment_metadata_list = load_metadata_from_netcdf(
            dataset
        )
        initialize_rattlesnake_object(hardware_metadata, environment_metadata_list)
        return environment_metadata_list
    elif extension == ".xlsx":
        workbook = openpyxl.load_workbook(filename)
        hardware_metadata, environment_metadata_list, profile_event_list = (
            load_metadata_from_workbook(workbook)
        )
        initialize_rattlesnake_object(
            hardware_metadata, environment_metadata_list, profile_event_list
        )
        return environment_metadata_list
    else:
        raise RattlesnakeError(
            f"Rattlesnake only loads .xlsx or .nc4 files, got {filename!r}"
        )
