# -*- coding: utf-8 -*-
"""
Controller subsystem to handle the user interface, including callback 
assignment and displaying results.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from qtpy import QtWidgets, uic, QtGui
from qtpy.QtCore import QThreadPool,QRunnable,QObject,Signal,Slot,QTimer
import time
import datetime
import netCDF4
import pyqtgraph
pyqtgraph.setConfigOption('background', 'w')
pyqtgraph.setConfigOption('foreground', 'k')
#pyqtgraph.setConfigOption('leftButtonPan',False)
import os
import openpyxl
import numpy as np
import multiprocessing as mp
import copy
import traceback

from .utilities import (Channel,error_message_qt,QueueContainer,
                        VerboseMessageQueue,
                        GlobalCommands,DataAcquisitionParameters)

from .environments import environment_UIs as all_environment_UIs,ui_path

from .ui_utilities import get_table_bools,ProfileTimer,ChannelMonitor

task_name = 'UI'

class UpdaterSignals(QObject):
    """Defines the signals that will be sent from the GUI Updater to the GUI
    
    Supported signals are:
    
    finished
        empty
    
    update
        `tuple` (widget_id,data)
    """
    finished = Signal()
    update = Signal(tuple)

class Updater(QRunnable):
    """Updater thread to collect results from the subsystems and reflect the
    changes in the GUI
    """
    def __init__(self,update_queue):
        """
        Initializes the updater with the queue and signals that will be emitted
        when the queue has data in it.

        Parameters
        ----------
        update_queue : mp.queues.Queue
            Queue from which events will be captured.

        """
        super(Updater, self).__init__()
        self.update_queue = update_queue
        self.signals = UpdaterSignals()
        self.verbose_queue = isinstance(self.update_queue,VerboseMessageQueue)
        
    @Slot()
    def run(self):
        """Continually capture update events from the queue"""
        while True:
            if self.verbose_queue:
                queue_data = self.update_queue.get(task_name)
            else:
                queue_data = self.update_queue.get()
            if queue_data[0] == GlobalCommands.QUIT:
                break
            self.signals.update.emit(queue_data)
        self.signals.finished.emit()
        time.sleep(1)

class Ui(QtWidgets.QMainWindow):
    """Main user interface from which the controller will be controlled."""
    def __init__(self,environments,queue_container : QueueContainer,profile_file = None):
        """
        Create the user interface with the specified parameters and queues

        Parameters
        ----------
        environments : iterable
            Iterable of control_type,control_name values to use to set up the
            environments.
        queue_container : QueueContainer
            Namespace containing the queues that are used by the controller.
        profile_file : str, optional
            File path to an optional profile file that will be loaded to set
            up the controller. The default is None.

        """
        try:
            # Store input data
            self.queue_container = queue_container
            self.environment_types = {name:control_type for control_type,name in environments}
            self.environments = [name for control_type,name in environments]
            self.environment_metadata = {name:None for name in self.environments}
            self.profile_events = None
            self.profile_timers = None
            self.profile_list_update_timer = None
            self.channel_monitor_window = None
            
            # Create the user interface
            super(Ui, self).__init__()
            uic.loadUi(ui_path, self)
    
            # Add tabs to the empty widgets based on the environments        
            self.environment_UIs = {}
            for environment_name,environment_type in self.environment_types.items():
                environment_ui = all_environment_UIs[environment_type]
                self.environment_UIs[environment_name] = environment_ui(
                        environment_name,
                        self.environment_definition_environment_tabs,
                        self.system_id_environment_tabs,
                        self.test_prediction_environment_tabs,
                        self.run_environment_tabs,
                        self.queue_container.environment_command_queues[environment_name],
                        self.queue_container.controller_communication_queue,
                        self.queue_container.log_file_queue)
                    
            # Remove the system ID and test prediction tab if not used.
            if self.system_id_environment_tabs.count() == 0:
                self.rattlesnake_tabs.removeTab(self.rattlesnake_tabs.indexOf(self.system_id_tab))
                self.has_system_id = False
                self.complete_system_ids = None
            else:
                self.has_system_id = True
                self.complete_system_ids = {self.system_id_environment_tabs.tabText(i):False for i in range(self.system_id_environment_tabs.count())}
            if self.test_prediction_environment_tabs.count() == 0:
                self.rattlesnake_tabs.removeTab(self.rattlesnake_tabs.indexOf(self.test_prediction_tab))
                self.has_test_predictions = False
            else:
                self.has_test_predictions = True
            
            # I might add this back in later, but for now we will just always show
            # this tab.
            # # If there is only one environment, remove the test profile tab
            # if len(self.environments) == 1:
            #     self.rattlesnake_tabs.removeTab(self.rattlesnake_tabs.indexOf(self.profile_tab))
            #     # Also remove profile information from the run test page
            #     self.run_profile_widget.hide()
            #     self.has_test_profile = False
            # else:
            #     self.has_test_profile = True
            
            self.streaming_environment_select_combobox.addItems(self.environments)

            self.manual_streaming_trigger_button.setVisible(False)
            
            for i in range(self.run_environment_tabs.count()):
                self.run_environment_tabs.widget(i).setEnabled(False)
                
            self.threadpool=QThreadPool()
            self.gui_updater = Updater(self.queue_container.gui_update_queue)
            # Create a side thread to collect global messages
            self.controller_instructions_collector = Updater(self.queue_container.controller_communication_queue)
    
            # Start Workers
            self.threadpool.start(self.gui_updater)
            self.threadpool.start(self.controller_instructions_collector)
            
            # Complete the remaining user interface
            self.complete_ui()
            self.connect_callbacks()
            
            # Create the command map for profile instructions
            self.command_map = {'Start Streaming':self.start_streaming,
                                'Stop Streaming':self.stop_streaming,
                                'Disarm DAQ':self.disarm_test}
            
            # Create a field to hold the loaded hardware file
            self.hardware_file = None
            self.setWindowIcon(QtGui.QIcon('logo/Rattlesnake_Icon.png'))
            self.setWindowTitle('Rattlesnake Vibration Controller')
            self.show()
        
            # If there is a loaded profile file, we need to handle it
            # print('Loading Profile')
            if not profile_file is None:
                # Channel Table
                # print('Loading Channel Table')
                self.load_channel_table(None,profile_file)
                # print('Loading Workbook')
                workbook = openpyxl.load_workbook(profile_file)
                # Hardware
                # print('Setting Hardware')
                hardware_sheet = workbook['Hardware']
                hardware_index = int(hardware_sheet.cell(1,2).value)
                self.hardware_selector.blockSignals(True)
                self.hardware_selector.setCurrentIndex(hardware_index)
                self.hardware_selector.blockSignals(False)
                self.hardware_update(select_file = False)
                self.hardware_file = hardware_sheet.cell(2,2).value
                if hardware_index == 1:
                    self.lanxi_sample_rate_selector.setCurrentIndex(round(np.log2(int(hardware_sheet.cell(3,2).value)/4096)))
                else:
                    self.sample_rate_selector.setValue(int(hardware_sheet.cell(3,2).value))
                self.time_per_read_selector.setValue(hardware_sheet.cell(4,2).value)
                self.time_per_write_selector.setValue(hardware_sheet.cell(5,2).value)
                self.lanxi_maximum_acquisition_processes_selector.setValue(hardware_sheet.cell(6,2).value)
                self.integration_oversample_selector.setValue(hardware_sheet.cell(7,2).value)
                # print('Initializing Data Acquisition')
                self.initialize_data_acquisition()
                # Now go through and do the environments
                for environment_name,environment_ui in self.environment_UIs.items():
                    # print('Setting Environment {:}'.format(environment_name))
                    environment_ui.set_parameters_from_template(workbook[environment_name])
                # print('Initializing Environments')
                self.initialize_environment_parameters()
                # Now the profile
                # print('Setting Test Profile')
                profile_sheet = workbook['Test Profile']
                index = 2
                profile_timestamps = []
                profile_environment_names = []
                profile_operation_names = []
                profile_data_names = []
                while True:
                    timestamp = profile_sheet.cell(index,1).value
                    environment = profile_sheet.cell(index,2).value
                    operation = profile_sheet.cell(index,3).value
                    data = profile_sheet.cell(index,4).value
                    if timestamp is None or (isinstance(timestamp,str) and timestamp.strip() == ''):
                        break
                    # print('Adding Profile Event {:}, {:}, {:}, {:}'.format(timestamp,environment,operation,data))
                    # self.add_profile_event(None,timestamp,environment,operation,data)
                    profile_timestamps.append(timestamp)
                    profile_environment_names.append(environment)
                    profile_operation_names.append(operation)
                    profile_data_names.append(data)
                    index += 1
                # print('Closing Workbook')
                workbook.close()
                # start_time = time.time()
                self.profile_table.setRowCount(len(profile_timestamps))
                # insert_row_time = time.time()
                # print('Time to Insert Row: {:}'.format(insert_row_time-start_time))
                for selected_row,(timestamp,environment,operation,data) in enumerate(zip(
                        profile_timestamps,profile_environment_names,
                        profile_operation_names,profile_data_names)):
                    timestamp_spinbox = QtWidgets.QDoubleSpinBox()
                    timestamp_spinbox.setMaximum(1e6)
                    timestamp_spinbox.setValue(float(timestamp))
                    self.profile_table.setCellWidget(selected_row,0,timestamp_spinbox)
                    # create_spinbox_time = time.time()
                    # print('Time to Create Spinbox: {:}'.format(create_spinbox_time-insert_row_time))
                    # Next a combobox sets the environment
                    environment_combobox = QtWidgets.QComboBox()
                    environment_combobox.addItem('Global')
                    for environment_name in self.environments:
                        environment_combobox.addItem(environment_name)
                    environment_combobox.setCurrentIndex(environment_combobox.findText(environment))
                    self.profile_table.setCellWidget(selected_row,1,environment_combobox)
                    # create_environment_combobox_time = time.time()
                    # print('Time to Create Environment Combobox: {:}'.format(create_environment_combobox_time-create_spinbox_time))
                    # Next a combobox sets the operation
                    if environment_combobox.currentIndex() == 0:
                        operations = [operation for operation in self.command_map]
                    else:
                        environment_name = self.environments[environment_combobox.currentIndex()-1]
                        operations = [op for op in self.environment_UIs[environment_name].command_map]
                    operation_combobox = QtWidgets.QComboBox()
                    for op in operations:
                        operation_combobox.addItem(op)
                    operation_combobox.setCurrentIndex(operation_combobox.findText(operation))
                    self.profile_table.setCellWidget(selected_row,2,operation_combobox)
                    # create_operation_combobox_time = time.time()
                    # print('Time to Create Operation Combobox: {:}'.format(create_operation_combobox_time-create_environment_combobox_time))
                    data_item = QtWidgets.QTableWidgetItem()
                    data_item.setText(str(data))
                    self.profile_table.setItem(selected_row,3,data_item)
                    # create_data_entry_time = time.time()
                    # print('Time to Data Entry: {:}'.format(create_data_entry_time-create_operation_combobox_time))
                    timestamp_spinbox.valueChanged.connect(self.update_profile_plot)
                    environment_combobox.currentIndexChanged.connect(self.update_operations)
                    operation_combobox.currentIndexChanged.connect(self.update_profile_plot)
                    # connect_callbacks_time = time.time()
                    # print('Time to Connect Callbacks: {:}'.format(connect_callbacks_time-create_data_entry_time))
                    # insert_row_time = connect_callbacks_time
                self.update_profile_plot()
            self.profile_table.itemChanged.connect(self.update_profile_plot)
                    
        except Exception:
            print(traceback.format_exc())
            
    def log(self,string):
        """Pass a message to the log_file_queue along with date/time and task name

        Parameters
        ----------
        string : str
            Message that will be written to the queue

        """
        self.queue_container.log_file_queue.put('{:}: {:} -- {:}\n'.format(datetime.datetime.now(),task_name,string))
        
    def complete_ui(self):
        """Helper function to complete setting up of the User Interface"""
        self.lanxi_sample_rate_selector.hide()
        self.lanxi_maximum_acquisition_processes_label.hide()
        self.lanxi_maximum_acquisition_processes_selector.hide()
        self.integration_oversample_selector.hide()
        self.integration_oversample_label.hide()
        # Fill in the channel table with empty strings
        for row_idx in range(self.channel_table.rowCount()):
            for col_idx in range(self.channel_table.columnCount()):
                item = QtWidgets.QTableWidgetItem('')
                self.channel_table.setItem(row_idx,col_idx,item)

        # Disable all tabs except the first
        for i in range(1,self.rattlesnake_tabs.count()-1):
            self.rattlesnake_tabs.setTabEnabled(i,False)
        
        # Reindex button groups
        self.streaming_button_group.setId(self.immediate_streaming_radiobutton,0)
        self.streaming_button_group.setId(self.test_level_streaming_radiobutton,1)
        self.streaming_button_group.setId(self.no_streaming_radiobutton,2)
        self.streaming_button_group.setId(self.profile_streaming_radiobutton,3)

        # Put values into the environment channel table
        self.environment_channels_table.setColumnCount(len(self.environments))
        self.environment_channels_table.setHorizontalHeaderLabels(self.environments)
        for row in range(self.environment_channels_table.rowCount()):
            for col in range(self.environment_channels_table.columnCount()):
                checkbox = QtWidgets.QCheckBox()
                if len(self.environments)==1:
                    checkbox.setChecked(True)
                self.environment_channels_table.setCellWidget(row,col,checkbox)
        if len(self.environments)==1:
            self.environment_channels_table.hide()
        max_cpus = mp.cpu_count()
        self.lanxi_maximum_acquisition_processes_selector.setMaximum(max_cpus)
        self.lanxi_maximum_acquisition_processes_selector.setValue(
            max_cpus - len(self.environments) if max_cpus > len(self.environments) else 1
            )

    def connect_callbacks(self):
        """Helper function to connect callbacks to widgets in the user interface"""
        # Stop program
        self.stop_program_button.clicked.connect(self.stop_program)
        # Channel Monitor
        self.channel_monitor_button.clicked.connect(self.show_channel_monitor)
        # Channel Table Tab
        self.load_channel_table_button.clicked.connect(self.load_channel_table)
        self.save_channel_table_button.clicked.connect(self.save_channel_table)
        self.initialize_data_acquisition_button.clicked.connect(self.initialize_data_acquisition)
        self.load_test_file_button.clicked.connect(self.load_test_file)
        self.hardware_selector.currentIndexChanged.connect(self.hardware_update)
        self.sample_rate_selector.valueChanged.connect(self.sample_rate_update)
        channel_table_scroll = self.channel_table.verticalScrollBar()
        channel_table_scroll.valueChanged.connect(self.sync_environment_table)
        environment_table_scroll = self.environment_channels_table.verticalScrollBar()
        environment_table_scroll.valueChanged.connect(self.sync_channel_table)
        
        # Control Definition Tab
        self.initialize_environments_button.clicked.connect(self.initialize_environment_parameters)
        
        # Profile Callbacks
        self.initialize_profile_button.clicked.connect(self.initialize_profile)
        self.save_profile_button.clicked.connect(self.save_profile)
        self.load_profile_button.clicked.connect(self.load_profile)
        self.add_profile_event_button.clicked.connect(self.add_profile_event)
        self.remove_profile_event_button.clicked.connect(self.remove_profile_event)
    
        # Run Test Tab
        self.select_streaming_file_button.clicked.connect(self.select_control_streaming_file)
        self.arm_test_button.clicked.connect(self.arm_test)
        self.disarm_test_button.clicked.connect(self.disarm_test)
        self.start_profile_button.clicked.connect(self.start_profile)
        self.stop_profile_button.clicked.connect(self.stop_profile)
        self.manual_streaming_radiobutton.toggled.connect(self.show_hide_manual_streaming)
        self.manual_streaming_trigger_button.clicked.connect(self.start_stop_streaming)
        
        # GUI Updater Signals
        self.gui_updater.signals.update.connect(self.update_gui)
        self.controller_instructions_collector.signals.update.connect(self.handle_controller_instructions)
    
    #%% Utility Functions
    def get_channel_table_strings(self):
        """Collect the strings in the channel table"""
        string_array = []
        for row_idx in range(self.channel_table.rowCount()):
            string_array.append([])
            for col_idx in range(self.channel_table.columnCount()):
                value = self.channel_table.item(row_idx,col_idx).text()
                string_array[-1].append(value)
        return string_array
    
    #%% Data Acquisition Callbacks

    def load_channel_table(self,clicked,filename = None):
        """Loads a channel table using a file dialog or the specified filename

        Parameters
        ----------
        clicked :
            The clicked event that triggered the callback.
        filename :
            File name defining the channel table for bypassing the callback when
            loading from a file (Default value = None).

        """
        num_environments = len(self.environments)
        if filename is None:
            filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load Channel Table',filter='Spreadsheets (*.xlsx *.csv *.txt)')
            if filename == '':
                return
        self.log('Loading Channel Table {:}'.format(filename))
        file_base,file_type = os.path.splitext(filename)
        if file_type == '.xlsx':
            workbook = openpyxl.load_workbook(filename,read_only=True)
            sheets = workbook.sheetnames
            if len(sheets) > 1:
                sheets = [sheet for sheet in sheets if 'channel' in sheet.lower()]
            if len(sheets) > 1:
                error_dialog = QtWidgets.QErrorMessage()
                error_dialog.showMessage('Could not identify channel table in Excel Spreadsheet\nIf multiple sheets exist, only 1 should have the word "channel" in it')
                return
            worksheet = workbook[sheets[0]]
            data_array = []
            environment_names = [worksheet.cell(2,24+i).value for i in range(num_environments)]
            for row in worksheet.iter_rows(min_row=3,max_col=23 + num_environments):
                data_array.append([])
                for col_idx,cell in enumerate(row):
                    data_array[-1].append(cell.value)
                if data_array[-1][0] is None:
                    data_array = data_array[:-1]
                    break
            workbook.close()
        elif file_type == '.csv' or file_type == '.txt':
            with open(filename,'r') as f:
                data_array = []
                for row_idx,line in enumerate(f):
                    if row_idx < 1:
                        continue
                    elif row_idx == 1:
                        environment_names = [val.strip() for val in line.split(',')][23:]
                        continue
                    data_array.append([val.strip() for val in line.split(',')])
            # Now split the data array off into the environment table
        channel_table_data_array = [row[:23] for row in data_array]
        environment_data_array = [row[23:] for row in data_array]
        # Now complete the table
        for row_idx,row_data in enumerate(channel_table_data_array):
            for col_idx,cell_data in enumerate(row_data):
                if col_idx == 0:
                    continue
                self.channel_table.item(row_idx,col_idx-1).setText('' if cell_data is None else str(cell_data))
        if num_environments > 1:
            for environment_index,environment_name in enumerate(environment_names):
                try:
                    environment_table_column = self.environments.index(environment_name)
                except ValueError:
                    error_message_qt('Invalid Environment Name','Invalid Environment Name {:}, Valid Environments are {:}.\n\nEnvironment channels not defined.'.format(environment_name,self.environments))
                    return
                for row_index,table_row in enumerate(environment_data_array):
                    try:
                        value = not (table_row[environment_index] == '' or table_row[environment_index] is None)
                    except IndexError:
                        value = False
                    self.environment_channels_table.cellWidget(row_index,environment_table_column).setChecked(value)

    def save_channel_table(self):
        """Save the channel table to a file"""
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self,'Save Channel Table',filter='Excel File (*.xlsx);;Comma-separated Values (*.csv)')
        if filename == '':
            return
        self.log('Saving Channel Table {:}'.format(filename))
        string_array = self.get_channel_table_strings()
        file_base,file_type = os.path.splitext(filename)
        if file_type == '.xlsx':
            # Create the header
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Channel Table"
            # Create the header
            worksheet.cell(row=1,column=2,value='Test Article Definition')
            worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
            worksheet.cell(row=1,column=5,value='Instrument Definition')
            worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=11)
            worksheet.cell(row=1,column=12,value='Channel Definition')
            worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=19)
            worksheet.cell(row=1,column=20,value='Output Feedback')
            worksheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=21)
            worksheet.cell(row=1,column=22,value='Limits')
            worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=23)
            for col_idx,val in enumerate(['Channel Index',
                                          'Node Number',
                                          'Node Direction',
                                          'Comment',
                                          'Serial Number',
                                          'Triax DoF',
                                          'Sensitivity  (mV/EU)',
                                          'Engineering Unit',
                                          'Make',
                                          'Model',
                                          'Calibration Exp Date',
                                          'Physical Device',
                                          'Physical Channel',
                                          'Type',
                                          'Minimum Value (V)',
                                          'Maximum Value (V)',
                                          'Coupling',
                                          'Current Excitation Source',
                                          'Current Excitation Value',
                                          'Physical Device',
                                          'Physical Channel',
                                          'Warning Level (EU)',
                                          'Abort Level (EU)']):
                worksheet.cell(row=2,column=1+col_idx,value=val)
            for row_idx,row in enumerate(string_array):
                worksheet.cell(row=row_idx+3,column=1,value=row_idx+1)
                for col_idx,col in enumerate(row):
                    if col == '':
                        continue
                    worksheet.cell(row=row_idx+3,column=col_idx+2,value=col)
            # Now do the environment
            if len(self.environments)>1:
                bool_array = get_table_bools(self.environment_channels_table)
                worksheet.cell(row=1,column=24,value='Environments')
                for index,name in enumerate(self.environments):
                    worksheet.cell(row=2,column=24+index,value=name)
                for row_idx,row in enumerate(bool_array):
                    for col_idx,col in enumerate(row):
                        if col:
                            worksheet.cell(row=row_idx+3,column=col_idx+24,value='X')
            workbook.save(filename)
        elif file_type == '.csv' or file_type == '.txt':
            error_message_qt('Not Implemented!','Output to CSV Not Implemented Yet!')
            
    def load_test_file(self):
        """Loads a test file using a file dialog
        """
        filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load Test NetCDF File',filter='NetCDF File (*.nc4);;All Files (*.*)')
        if filename == '':
            return
        dataset = netCDF4.Dataset(filename)
        # Channel Table
        channel_table = dataset['channels']
        # Node Number
        data = channel_table['node_number'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,0).setText(value)
        # Node Direction
        data = channel_table['node_direction'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,1).setText(value)
        # Comment
        data = channel_table['comment'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,2).setText(value)
        # SN
        data = channel_table['serial_number'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,3).setText(value)
        # Triax Dof
        data = channel_table['triax_dof'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,4).setText(value)
        # Sensitivity
        data = channel_table['sensitivity'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,5).setText(str(value))
        # Units
        data = channel_table['unit'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,6).setText(value)
        # Make
        data = channel_table['make'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,7).setText(value)
        # Model
        data = channel_table['model'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,8).setText(value)
        # Expiration Date
        data = channel_table['expiration'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,9).setText(value)
        # Read Device
        data = channel_table['physical_device'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,10).setText(value)
        # Read Channel
        data = channel_table['physical_channel'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,11).setText(value)
        # Type
        data = channel_table['channel_type'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,12).setText(value)
        # Min Volts
        data = channel_table['minimum_value'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,13).setText(str(value))
        # Max Volts
        data = channel_table['maximum_value'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,14).setText(str(value))
        # Coupling
        data = channel_table['coupling'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,15).setText(value)
        # Excitation Source
        data = channel_table['excitation_source'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,16).setText(value)
        # Excitation
        data = channel_table['excitation'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,17).setText(str(value))
        # Output Device
        data = channel_table['feedback_device'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,18).setText(value)
        # Output Channel
        data = channel_table['feedback_channel'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,19).setText(value)
        # Output Device
        data = channel_table['warning_level'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,20).setText(value)
        # Output Channel
        data = channel_table['abort_level'][...]
        for row_idx,value in enumerate(data):
            self.channel_table.item(row_idx,21).setText(value)
        # Environment Table
        for saved_environment_index,saved_environment_name in enumerate(dataset.variables['environment_names'][...]):
            environment_index = self.environments.index(saved_environment_name)
            for channel_index,bool_row in enumerate(dataset.variables['environment_active_channels'][:,saved_environment_index]):
                boolean = bool(bool_row)
                widget = self.environment_channels_table.cellWidget(channel_index,environment_index)
                widget.setChecked(boolean)
        # Hardware
        self.hardware_selector.blockSignals(True)
        try:
            self.hardware_selector.setCurrentIndex(dataset.hardware)
            self.hardware_file = None if dataset.hardware_file == 'None' else dataset.hardware_file
        except AttributeError:
            self.hardware_selector.setCurrentIndex(0)
            self.hardware_file = None
        self.hardware_selector.blockSignals(False)
        # Show the right widgets
        self.hardware_update(select_file = False)
        if self.hardware_selector.currentIndex() == 1:
            self.lanxi_sample_rate_selector.setCurrentIndex(np.log2(dataset.sample_rate//4096))
            self.lanxi_maximum_acquisition_processes_selector.setValue(dataset.maximum_acquisition_processes)
        else:
            self.sample_rate_selector.setValue(dataset.sample_rate)
        self.integration_oversample_selector.setValue(dataset.output_oversample)
        self.time_per_read_selector.setValue(dataset.time_per_read)
        self.time_per_write_selector.setValue(dataset.time_per_write)
        # Initialize files
        self.initialize_data_acquisition()
        # Set the test parameters
        for environment in self.environments:
            self.environment_UIs[environment].retrieve_metadata(dataset)
        self.initialize_environment_parameters()
    
    def hardware_update(self, current_index = None, select_file = True):
        """Callback to provide options when hardware is selected"""
        current_index = self.hardware_selector.currentIndex()
        if current_index == 0: # NIDAQmx
            self.sample_rate_selector.show()
            self.lanxi_sample_rate_selector.hide()
            self.lanxi_maximum_acquisition_processes_label.hide()
            self.lanxi_maximum_acquisition_processes_selector.hide()
            self.integration_oversample_selector.hide()
            self.integration_oversample_label.hide()
            self.hardware_file = None
        elif current_index == 1: # LAN-XI
            self.sample_rate_selector.hide()
            self.lanxi_sample_rate_selector.show()
            self.lanxi_maximum_acquisition_processes_label.show()
            self.lanxi_maximum_acquisition_processes_selector.show()
            self.integration_oversample_selector.hide()
            self.integration_oversample_label.hide()
            self.hardware_file = None
        elif current_index == 2: # DP Quattro
            # Load in the library file
            if select_file:
                filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Data Physics API',filter='Quattro API (DpQuattro.dll)')
                if filename == '':
                    self.hardware_selector.setCurrentIndex(0)
                    return
                else:
                    self.hardware_file = filename
            self.sample_rate_selector.show()
            self.lanxi_sample_rate_selector.hide()
            self.lanxi_maximum_acquisition_processes_label.hide()
            self.lanxi_maximum_acquisition_processes_selector.hide()
            self.integration_oversample_selector.hide()
            self.integration_oversample_label.hide()
            self.sample_rate_update()
        elif current_index == 3: # DP 900
            error_message_qt('Data Physics 900 Series','Data Physics 900 Series is not yet available.  Expected Quarter 1 of 2024.')
            self.hardware_selector.setCurrentIndex(0)
        elif current_index == 4: # Exodus
            # Load in an exodus file
            if select_file:
                filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load Exodus File with Eigensolution',filter='Exodus File (*.exo *.e)')
                if filename == '':
                    self.hardware_selector.setCurrentIndex(0)
                    return
                else:
                    self.hardware_file = filename
            self.sample_rate_selector.show()
            self.lanxi_sample_rate_selector.hide()
            self.lanxi_maximum_acquisition_processes_label.hide()
            self.lanxi_maximum_acquisition_processes_selector.hide()
            self.integration_oversample_selector.show()
            self.integration_oversample_label.show()
        elif current_index == 5: # State Space File
            # Load in a state space file
            if select_file:
                filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load Numpy or Matlab File with State Space Matrices A B C D',filter='Matlab or Numpy File (*.mat *.npz)')
                if filename == '':
                    self.hardware_selector.setCurrentIndex(0)
                    return
                else:
                    self.hardware_file = filename
            self.sample_rate_selector.show()
            self.lanxi_sample_rate_selector.hide()
            self.lanxi_maximum_acquisition_processes_label.hide()
            self.lanxi_maximum_acquisition_processes_selector.hide()
            self.integration_oversample_selector.show()
            self.integration_oversample_label.show()
        elif current_index == 6:
            # Load in an sdynpy system
            if select_file:
                filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load a SDynPy System',filter='Numpy File (*.npz)')
                if filename == '':
                    self.hardware_selector.setCurrentIndex(0)
                    self.hardware_file = None
                else:
                    self.hardware_file = filename
            self.sample_rate_selector.show()
            self.lanxi_sample_rate_selector.hide()
            self.lanxi_maximum_acquisition_processes_label.hide()
            self.lanxi_maximum_acquisition_processes_selector.hide()
            self.integration_oversample_selector.show()
            self.integration_oversample_label.show()
        else:
            error_message_qt('Invalid Hardware Type!','You have selected an invalid hardware type.  How did you do this?!')
    
    def sample_rate_update(self):
        if self.hardware_selector.currentIndex() == 2:
            current_value = self.sample_rate_selector.value()
            valid_dp_sample_rates = np.array([
                16, 20, 25, 32, 40, 50, 64, 80, 100, 128, 160, 200, 256, 320,
                400, 512, 640, 800, 1024, 1280, 1600, 2048, 2560, 3200, 4096, 
                5120, 6400, 8192, 10240, 12800, 20480, 25600, 40960, 51200,
                102400])
            closest_index = np.argmin(abs(valid_dp_sample_rates - current_value))
            closest_rate = valid_dp_sample_rates[closest_index]
            # Check if it is either one above or one below a previous rate
            if current_value - closest_rate == 1 and closest_index != len(valid_dp_sample_rates) - 1:
                closest_index += 1
                closest_rate = valid_dp_sample_rates[closest_index]
            elif current_value - closest_rate == -1 and closest_index != 0:
                closest_index -= 1
                closest_rate = valid_dp_sample_rates[closest_index]
            self.sample_rate_selector.blockSignals(True)
            self.sample_rate_selector.setValue(closest_rate)
            self.sample_rate_selector.blockSignals(False)
    
    def initialize_data_acquisition(self):
        """Initializes the data acquisition hardware
        
        This function collects the information from the channel table as well
        as the hardware information to create a DataAcquisitionParameters object
        that gets passed to each environment through its command queue.
        
        It also sends the data acquisition parameters to the acquisition and
        output subtasks.
        """
        self.log('Initializing Data Acquisition')
        channels = []
        environment_booleans = []
        channel_table_strings = self.get_channel_table_strings()
        environment_channels = get_table_bools(self.environment_channels_table)
#        print('User Interface {:} Channels'.format(len(channel_table_strings)))
        for index,(row,environment_bools) in enumerate(zip(channel_table_strings,environment_channels)):
            try:
                channel = Channel.from_channel_table_row(row)
            except ValueError as e:
                self.log('Bad Entry in Channel {:}, {:}'.format(index+1,e))
                error_message_qt('Channel Table Error','Bad Entry in Channel {:}\n\n{:}'.format(index+1,e))
                return
            if not channel is None:
                channels.append(channel)
                environment_booleans.append(environment_bools)
        # Go through and initialize the channel information for each environment
        environment_booleans = np.array(environment_booleans)
        environment_channel_indices = {}
        if self.hardware_selector.currentIndex() == 1:
            sample_rate = 2**self.lanxi_sample_rate_selector.currentIndex()*4096
            output_oversample = 16384//sample_rate
            if output_oversample == 0:
                output_oversample = 1
            acquisition_processes = self.lanxi_maximum_acquisition_processes_selector.value()
        elif self.hardware_selector.currentIndex() in [4,5,6]:
            sample_rate = self.sample_rate_selector.value()
            output_oversample = self.integration_oversample_selector.value()
            acquisition_processes = 1
        else:
            sample_rate = self.sample_rate_selector.value()
            output_oversample = 1
            acquisition_processes = 1
        for environment_index,environment in enumerate(self.environments):
            environment_channel_list = copy.deepcopy([channel for channel,environment_bool in zip(channels,environment_booleans) if environment_bool[environment_index]])
            environment_channel_indices[environment] = [index for index,environment_bool in enumerate(environment_booleans) if environment_bool[environment_index]]
            environment_daq_parameters = DataAcquisitionParameters(environment_channel_list,
                                                                   sample_rate,
                                                                   round(sample_rate*self.time_per_read_selector.value()),
                                                                   round(sample_rate*self.time_per_write_selector.value()*output_oversample),
                                                                   self.hardware_selector.currentIndex(),
                                                                   self.hardware_file,
                                                                   self.environments,
                                                                   environment_booleans,
                                                                   output_oversample,
                                                                   acquisition_processes)
            self.queue_container.environment_command_queues[environment].put(task_name,(GlobalCommands.INITIALIZE_DATA_ACQUISITION,environment_daq_parameters))
            self.environment_UIs[environment].initialize_data_acquisition(environment_daq_parameters)
        self.global_daq_parameters = DataAcquisitionParameters(channels,
                                                          sample_rate,
                                                          round(sample_rate*self.time_per_read_selector.value()),
                                                          round(sample_rate*self.time_per_write_selector.value()*output_oversample),
                                                          self.hardware_selector.currentIndex(),
                                                          self.hardware_file,
                                                          self.environments,
                                                          environment_booleans,
                                                          output_oversample,
                                                          acquisition_processes)
        self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.INITIALIZE_DATA_ACQUISITION,
                              (self.global_daq_parameters,
                               environment_channel_indices)))
        self.queue_container.output_command_queue.put(task_name,(GlobalCommands.INITIALIZE_DATA_ACQUISITION,
                              (self.global_daq_parameters,
                               environment_channel_indices)))
        self.channel_monitor_button.setEnabled(True)
        if not self.channel_monitor_window is None:
            self.channel_monitor_window.update_channel_list(self.global_daq_parameters)
        for i in range(2,self.rattlesnake_tabs.count()-1):
            self.rattlesnake_tabs.setTabEnabled(i,False)
        self.rattlesnake_tabs.setTabEnabled(1,True)
        self.rattlesnake_tabs.setCurrentIndex(1)
        
    #%% Test Parameters Callbacks
    
    def initialize_environment_parameters(self):
        """Initializes the environment parameters
        
        This function initializes the environment-specific parameters for each
        environment by calling the initialize_environment function of each
        environment-specific user interface."""
        for environment in self.environments:
            environment_parameters = self.environment_UIs[environment].initialize_environment()
            self.environment_metadata[environment] = environment_parameters
            self.queue_container.environment_command_queues[environment].put(task_name,(GlobalCommands.INITIALIZE_ENVIRONMENT_PARAMETERS,environment_parameters))
        
        # Enable the next section
        self.rattlesnake_tabs.setTabEnabled(2,True)
        self.rattlesnake_tabs.setCurrentIndex(2)
        
        # If there are test predictions
        if self.has_test_predictions:
            self.rattlesnake_tabs.setTabEnabled(3,True)
            
    #%% Run test callbacks
    def select_control_streaming_file(self):
        """Selects a file to stream data to disk"""
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self,'Select NetCDF File to Save Control Data',filter='NetCDF File (*.nc4)')
        if filename == '':
            return
        self.streaming_file_display.setText(filename)

    def arm_test(self):
        """Starts the data acquisition running in preparation for control"""
        if not self.no_streaming_radiobutton.isChecked() and len(self.streaming_file_display.text())==0:
            error_message_qt('No Streaming File Selected','Please select a file into which data will be streamed.')
            return
        self.log('Arming Test Hardware')
        self.queue_container.controller_communication_queue.put(task_name,(GlobalCommands.RUN_HARDWARE,None))
        self.no_streaming_radiobutton.setEnabled(False)
        self.profile_streaming_radiobutton.setEnabled(False)
        self.test_level_streaming_radiobutton.setEnabled(False)
        self.streaming_environment_select_combobox.setEnabled(False)
        self.immediate_streaming_radiobutton.setEnabled(False)
        self.select_streaming_file_button.setEnabled(False)
        self.manual_streaming_radiobutton.setEnabled(False)
        self.manual_streaming_trigger_button.setEnabled(True)
        self.arm_test_button.setEnabled(False)
        self.disarm_test_button.setEnabled(True)
        self.start_profile_button.setEnabled(True)
        self.stop_profile_button.setEnabled(True)
        for i in range(self.run_environment_tabs.count()):
            self.run_environment_tabs.widget(i).setEnabled(True)
        for environment,ui in self.environment_UIs.items():
            try:
                ui.disable_system_id_daq_armed()
            except AttributeError:
                pass
        if (self.profile_streaming_radiobutton.isChecked()
            or self.test_level_streaming_radiobutton.isChecked()
            or self.immediate_streaming_radiobutton.isChecked()
            or self.manual_streaming_radiobutton.isChecked()):
            file_path = self.streaming_file_display.text()
            self.queue_container.streaming_command_queue.put(task_name,(GlobalCommands.INITIALIZE_STREAMING,(file_path,self.global_daq_parameters,self.environment_metadata)))
        if self.immediate_streaming_radiobutton.isChecked():
            self.start_streaming()
        
    def disarm_test(self):
        """Stops the data acquisition from running and shuts down all environments"""
        self.log('Disarming Test Hardware')
        self.queue_container.controller_communication_queue.put(task_name,(GlobalCommands.STOP_HARDWARE,None))
        for environment,ui in self.environment_UIs.items():
            ui.stop_control()
        # for environment,queue in self.queue_container.environment_command_queues.items():
        #     queue.put(task_name,(GlobalCommands.STOP_ENVIRONMENT,None))
        self.no_streaming_radiobutton.setEnabled(True)
        self.profile_streaming_radiobutton.setEnabled(True)
        self.test_level_streaming_radiobutton.setEnabled(True)
        self.streaming_environment_select_combobox.setEnabled(True)
        self.immediate_streaming_radiobutton.setEnabled(True)
        self.manual_streaming_radiobutton.setEnabled(True)
        self.manual_streaming_trigger_button.setEnabled(False)
        self.manual_streaming_trigger_button.setText('Start\nStreaming')
        self.select_streaming_file_button.setEnabled(True)
        self.arm_test_button.setEnabled(True)
        self.disarm_test_button.setEnabled(False)
        self.start_profile_button.setEnabled(False)
        self.stop_profile_button.setEnabled(False)
        for i in range(self.run_environment_tabs.count()):
            self.run_environment_tabs.widget(i).setEnabled(False)
        for environment,ui in self.environment_UIs.items():
            try:
                ui.enable_system_id_daq_disarmed()
            except AttributeError:
                pass
            
    def start_profile(self):
        """Starts running the test profile"""
        self.log('Running Profile')
        # Create the QTimers
        self.profile_timers = []
        for timestamp,environment_name,operation,data in self.profile_events:
            timer = ProfileTimer(environment_name,operation,data)
            timer.setSingleShot(True)
            timer.timeout.connect(self.fire_profile_event)
            timer.start(int(timestamp*1000))
            self.profile_timers.append(timer)
        self.profile_list_update_timer = QTimer()
        self.profile_list_update_timer.timeout.connect(self.update_profile_list)
        self.profile_list_update_timer.start(250)
    
    def fire_profile_event(self):
        """Activates a given profile event """
        widget = self.sender()
        environment_name = widget.environment
        operation = widget.operation
        data = widget.data
        self.log('Profile Firing Event {:} {:} {:}'.format(environment_name,operation,data))
        if self.show_profile_change_checkbox.isChecked():
            if not environment_name == 'Global':
                environment_index = self.environments.index(environment_name)
                self.run_environment_tabs.setCurrentIndex(environment_index)
        if environment_name == 'Global':
            if operation == 'Start Streaming' and (not self.profile_streaming_radiobutton.isChecked()):
                return
            self.command_map[operation]()
        elif operation in ['Start Control','Stop Control']:
            self.environment_UIs[environment_name].command_map[operation]()
        else:
            self.environment_UIs[environment_name].command_map[operation](data)
    
    def update_profile_list(self):
        """Updates the list of upcoming profile events."""
        profile_representation = []
        for timer,profile_event in zip(self.profile_timers,self.profile_events):
            remaining_time = timer.remainingTime()/1000
            if remaining_time > 0:
                profile_representation.append([remaining_time] + profile_event[1:])
        self.upcoming_instructions_list.clear()
        self.upcoming_instructions_list.addItems(['{:0.2f} {:} {:} {:}'.format(*profile_event) for profile_event in sorted(profile_representation)])
        if len(profile_representation) == 0:
            self.stop_profile()

    def stop_profile(self):
        """Stops running the profile"""
        for timer in self.profile_timers:
            timer.stop()
        self.profile_list_update_timer.stop()

    def initialize_profile(self):
        """Initializes the profile list in the controller"""
        self.profile_events = []
        for row in range(self.profile_table.rowCount()):
            self.profile_events.append([float(self.profile_table.cellWidget(row,0).value()),
                                        self.profile_table.cellWidget(row,1).currentText(),
                                        self.profile_table.cellWidget(row,2).currentText(),
                                        self.profile_table.item(row,3).text()])
        if len(self.profile_events) == 0:
            self.run_profile_widget.hide()
        else:
            self.run_profile_widget.show()
        self.upcoming_instructions_list.clear()
        self.upcoming_instructions_list.addItems(['{:0.2f} {:} {:} {:}'.format(*profile_event) for profile_event in sorted(self.profile_events)])
        for i in range(self.rattlesnake_tabs.count()-1):
            self.rattlesnake_tabs.setTabEnabled(i,True)
                    
        self.rattlesnake_tabs.setCurrentIndex(self.rattlesnake_tabs.count()-2)

    def save_profile(self):
        """Save the profile to a spreadsheet file"""
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self,'Save Test Profile',filter='Excel File (*.xlsx)')
        if filename == '':
            return
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Test Profile"
        worksheet.cell(1,1,'Time (s)')
        worksheet.cell(1,2,'Environment')
        worksheet.cell(1,3,'Operation')
        worksheet.cell(1,4,'Data')
        for row in range(self.profile_table.rowCount()):
            worksheet.cell(row+2,1,float(self.profile_table.cellWidget(row,0).value()))
            worksheet.cell(row+2,2,self.profile_table.cellWidget(row,1).currentText())
            worksheet.cell(row+2,3,self.profile_table.cellWidget(row,2).currentText())
            worksheet.cell(row+2,4,self.profile_table.item(row,3).text())
        workbook.save(filename)
    
    def load_profile(self):
        """Load a profile from a spreadsheet file"""
        filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self,'Load Test Profile',filter='Excel File (*.xlsx)')
        if filename == '':
            return
        workbook = openpyxl.load_workbook(filename)
        profile_sheet = workbook['Test Profile']
        index = 2
        while True:
            timestamp = profile_sheet.cell(index,1).value
            environment = profile_sheet.cell(index,2).value
            operation = profile_sheet.cell(index,3).value
            data = profile_sheet.cell(index,4).value
            if timestamp is None or (isinstance(timestamp,str) and timestamp.strip() == ''):
                break
            self.add_profile_event(None,timestamp,environment,operation,data)
            index += 1
    
    def add_profile_event(self,clicked=None,timestamp = None,environment = None,operation = None,data = None):
        """Adds an event to the profile either by clicking a button or by specifying it

        Parameters
        ----------
        clicked :
            The clicked event. (Default value = None)
        timestamp :
            Optional timestamp to give to the controller (Default value = None)
        environment :
            Optional environment the profile instruction corresponds to
            (Default value = None)
        operation :
            Optional operation specified by the profile instruction
            (Default value = None)
        data :
            Optional data needed by the operation (Default value = None)

        """
        # start_time = time.time()
        # Create the row in the profile table
        selected_row = self.profile_table.rowCount()
        self.profile_table.insertRow(selected_row)
        # insert_row_time = time.time()
        # print('Time to Insert Row: {:}'.format(insert_row_time-start_time))
        # First entry is a spinbox
        timestamp_spinbox = QtWidgets.QDoubleSpinBox()
        timestamp_spinbox.setMaximum(1e6)
        self.profile_table.setCellWidget(selected_row,0,timestamp_spinbox)
        # create_spinbox_time = time.time()
        # print('Time to Create Spinbox: {:}'.format(create_spinbox_time-insert_row_time))
        # Next a combobox sets the environment
        environment_combobox = QtWidgets.QComboBox()
        environment_combobox.addItem('Global')
        for environment_name in self.environments:
            environment_combobox.addItem(environment_name)
        self.profile_table.setCellWidget(selected_row,1,environment_combobox)
        # create_environment_combobox_time = time.time()
        # print('Time to Create Environment Combobox: {:}'.format(create_environment_combobox_time-create_spinbox_time))
        # Next a combobox sets the operation
        operation_combobox = QtWidgets.QComboBox()
        for op in self.command_map:
            operation_combobox.addItem(op)
        self.profile_table.setCellWidget(selected_row,2,operation_combobox)
        # create_operation_combobox_time = time.time()
        # print('Time to Create Operation Combobox: {:}'.format(create_operation_combobox_time-create_environment_combobox_time))
        data_item = QtWidgets.QTableWidgetItem()
        self.profile_table.setItem(selected_row,3,data_item)
        # create_data_entry_time = time.time()
        # print('Time to Data Entry: {:}'.format(create_data_entry_time-create_operation_combobox_time))
        # Connect the callbacks
        timestamp_spinbox.valueChanged.connect(self.update_profile_plot)
        environment_combobox.currentIndexChanged.connect(self.update_operations)
        operation_combobox.currentIndexChanged.connect(self.update_profile_plot)
        # connect_callbacks_time = time.time()
        # print('Time to Connect Callbacks: {:}'.format(connect_callbacks_time-create_data_entry_time))
        # Initialize parameters if necessary
        if not timestamp is None:
            timestamp_spinbox.setValue(float(timestamp))
        # initialize_time_time = time.time()
        # print('Time to Initialize Timestamp: {:}'.format(initialize_time_time-connect_callbacks_time))
        if not environment is None:
            environment_combobox.setCurrentIndex(environment_combobox.findText(environment))
        # initialize_environment_time = time.time()
        # print('Time to Initialize Timestamp: {:}'.format(initialize_environment_time-initialize_time_time))
        if not operation is None:
            operation_combobox.setCurrentIndex(operation_combobox.findText(operation))
        # initialize_operation_time = time.time()
        # print('Time to Initialize Timestamp: {:}'.format(initialize_operation_time-initialize_environment_time))
        if not data is None:
            data_item.setText(str(data))
        # initialize_data_time = time.time()
        # print('Time to Initialize Data: {:}'.format(initialize_data_time-initialize_operation_time))
        # Update the plot
        self.update_profile_plot()
        # update_plot_time = time.time()
        # print('Time to Update Plot: {:}'.format(update_plot_time-initialize_data_time))

    def update_operations(self):
        """Update profile operations given a selected environment
        """
        widget = self.sender()
        if widget.currentIndex() == 0:
            operations = [operation for operation in self.command_map]
        else:
            environment_name = self.environments[widget.currentIndex()-1]
            operations = [operation for operation in self.environment_UIs[environment_name].command_map]
        for row in range(self.profile_table.rowCount()):
            if widget is self.profile_table.cellWidget(row,1):
                print('Found Widget at {:}'.format(row))
                break
        operation_combobox = self.profile_table.cellWidget(row,2)
        operation_combobox.blockSignals(True)
        operation_combobox.clear()
        for operation in operations:
            operation_combobox.addItem(operation)
        operation_combobox.blockSignals(False)
        self.update_profile_plot()
    
    def update_profile_plot(self):
        """Updates the plot of profile events"""
        plot_item = self.profile_timeline_plot.getPlotItem()
        plot_item.clear()
        plot_item.showGrid(True,True,0.25)
        plot_item.disableAutoRange()
        max_time = 0
        for row in range(self.profile_table.rowCount()):
            time = self.profile_table.cellWidget(row,0).value()
            if time > max_time:
                max_time = time
            plot_item.plot([time],[self.profile_table.cellWidget(row,1).currentIndex()],pen=None,symbol='o',pxMode=True)
            text_item = pyqtgraph.TextItem('{:}: '.format(row+1)+self.profile_table.cellWidget(row,2).currentText() + (': '+self.profile_table.item(row,3).text() if self.profile_table.item(row,3).text().strip() != '' else ''),color=(0,0,0),angle=-15)
            plot_item.addItem(text_item)
            text_item.setPos(time,self.profile_table.cellWidget(row,1).currentIndex())
        axis = plot_item.getAxis('left')
        axis.setTicks([[(i,name) for i,name in enumerate(['Global']+self.environments)],[]])
        plot_item.setXRange(0,max_time*1.1)
        plot_item.setYRange(-1,len(self.environments))
        
    def remove_profile_event(self):
        """Removes a profile event from the list of events"""
        selected_row = self.profile_table.currentRow()
        if selected_row >= 0:
            self.profile_table.removeRow(selected_row)
        self.update_profile_plot()

    def start_streaming(self):
        """Tells acquisition to start sending data to streaming"""
        self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.START_STREAMING,None))
        
    def stop_streaming(self):
        """Tells the acquisition to stop sending data to streaming"""
        self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.STOP_STREAMING,None))
    
    def show_hide_manual_streaming(self):
        if self.manual_streaming_radiobutton.isChecked():
            self.manual_streaming_trigger_button.setVisible(True)
        else:
            self.manual_streaming_trigger_button.setVisible(False)
    
    def start_stop_streaming(self):
        if self.manual_streaming_trigger_button.text() == 'Stop\nStreaming':
            self.manual_streaming_trigger_button.setText('Start\nStreaming')
            self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.STOP_STREAMING,None))
        else:
            self.manual_streaming_trigger_button.setText('Stop\nStreaming')
            self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.START_STREAMING,None))
        
    #%% Other Callbacks
    def sync_environment_table(self):
        """Callback to synchronize scrolling between channel tables"""
        self.environment_channels_table.verticalScrollBar().setValue(self.channel_table.verticalScrollBar().value())
    
    def sync_channel_table(self):
        """Callback to synchronize scrolling between channel tables"""
        self.channel_table.verticalScrollBar().setValue(self.environment_channels_table.verticalScrollBar().value())
    
    def stop_program(self):
        """Callback to stop the entire program"""
        self.close()
        
    def update_gui(self,queue_data):
        """Update the graphical interface for the main controller

        Parameters
        ----------
        queue_data :
            A 2-tuple consisting of ``(message,data)`` pairs where the message
            denotes what to change and the data contains the information needed
            to be displayed.  

        """
        message,data = queue_data
#        self.log('Updating GUI {:}'.format(message))
        if message == 'error':
            error_message_qt(data[0],data[1])
            return
        elif message in self.environments:
            self.environment_UIs[message].update_gui(data)
        elif message == 'monitor':
            if not self.channel_monitor_window is None:
                if not self.channel_monitor_window.isVisible():
                    self.channel_monitor_window = None
                else:
                    self.channel_monitor_window.update(data)
        elif message == 'update_metadata':
            environment_name,metadata = data
            self.environment_metadata[environment_name] = metadata
        elif message == 'stop':
            self.disarm_test()
        elif message == 'enable':
            widget = getattr(self,data)
            widget.setEnabled(True)
        elif message == 'disable':
            widget = getattr(self,data)
            widget.setEnabled(False)
        elif message == 'enable_tab':
            self.rattlesnake_tabs.setTabEnabled(data,True)
            self.rattlesnake_tabs.setCurrentIndex(data)
        elif message == 'disable_tab':
            self.rattlesnake_tabs.setTabEnabled(data,False)
        else:
            widget = getattr(self,message)
            if type(widget) is QtWidgets.QDoubleSpinBox:
                widget.setValue(data)
            elif type(widget) is QtWidgets.QSpinBox:
                widget.setValue(data)
            elif type(widget) is QtWidgets.QLineEdit:
                widget.setText(data)
            elif type(widget) is QtWidgets.QListWidget:
                widget.clear()
                widget.addItems(['{:.3f}'.format(d) for d in data])
#        self.log('Update took {:} seconds'.format(time.time()-start_time))
        
    def handle_controller_instructions(self,queue_data):
        """Handler function for global controller instructions

        Parameters
        ----------
        queue_data :
            A 2-tuple consisting of ``(message,data)`` pairs where the message
            denotes what to change and the data contains the information needed
            to be displayed.  

        """
        message,data = queue_data
        self.log('Received Global Instruction {:}'.format(message.name))
        if message == GlobalCommands.QUIT:
            self.stop_program()
        elif message == GlobalCommands.INITIALIZE_DATA_ACQUISITION:
            self.initialize_data_acquisition()
        elif message == GlobalCommands.INITIALIZE_ENVIRONMENT_PARAMETERS:
            self.initialize_environment_parameters()
        elif message == GlobalCommands.UPDATE_METADATA:
            environment,metadata = data
            self.environment_metadata[environment] = metadata
        elif message == GlobalCommands.RUN_HARDWARE:
            self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.RUN_HARDWARE,data))
            self.queue_container.output_command_queue.put(task_name,(GlobalCommands.RUN_HARDWARE,data))
        elif message == GlobalCommands.STOP_HARDWARE:
            self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.STOP_HARDWARE,data))
            self.queue_container.output_command_queue.put(task_name,(GlobalCommands.STOP_HARDWARE,data))
        elif message == GlobalCommands.INITIALIZE_STREAMING:
            self.queue_container.streaming_command_queue.put(task_name,(GlobalCommands.INITIALIZE_STREAMING,(data,self.global_daq_parameters,self.environment_metadata)))
        elif message == GlobalCommands.STREAMING_DATA:
            self.queue_container.streaming_command_queue.put(task_name,(GlobalCommands.STREAMING_DATA,data))
        elif message == GlobalCommands.FINALIZE_STREAMING:
            self.queue_container.streaming_command_queue.put(task_name,(GlobalCommands.FINALIZE_STREAMING,data))
        elif message == GlobalCommands.START_ENVIRONMENT:
            self.queue_container.output_command_queue.put(task_name,(GlobalCommands.START_ENVIRONMENT,data))
        elif message == GlobalCommands.STOP_ENVIRONMENT:
            self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.STOP_ENVIRONMENT,data))
        elif message == GlobalCommands.START_STREAMING:
            self.start_streaming()
        elif message == GlobalCommands.STOP_STREAMING:
            self.queue_container.acquisition_command_queue.put(task_name,(GlobalCommands.STOP_STREAMING,data))
        elif message == GlobalCommands.COMPLETED_SYSTEM_ID:
            self.complete_system_ids[data] = True
            if all([flag for environment,flag in self.complete_system_ids.items()]):
                if self.has_test_predictions:
                    self.rattlesnake_tabs.setTabEnabled(4,True)
                else:
                    self.rattlesnake_tabs.setTabEnabled(3,True)
        elif message == GlobalCommands.AT_TARGET_LEVEL:
            environment_name = data
            if self.test_level_streaming_radiobutton.isChecked() and self.streaming_environment_select_combobox.currentText() == environment_name:
                self.start_streaming()
    
    def closeEvent(self,event):
        """Event triggered when closing the software to gracefully shut down.

        Parameters
        ----------
        event :
            The close event, which is accepted.

        """
        for environment_name,command_queue in self.queue_container.environment_command_queues.items():
            command_queue.put(task_name,(GlobalCommands.QUIT,None))
            
        self.queue_container.gui_update_queue.put((GlobalCommands.QUIT,None))
        self.queue_container.controller_communication_queue.put(task_name,(GlobalCommands.QUIT,None))
        
        for command_queue in [self.queue_container.acquisition_command_queue,
                              self.queue_container.output_command_queue,
                              self.queue_container.streaming_command_queue]:
            command_queue.put(task_name,(GlobalCommands.QUIT,None))
        
        event.accept()
        
    def show_channel_monitor(self):
        """
        Shows the channel monitor window
        """
        if (self.channel_monitor_window is None) or (not self.channel_monitor_window.isVisible()):
            self.channel_monitor_window = ChannelMonitor(None,self.global_daq_parameters)
        else:
            pass # TODO Need to raise the window to the front, or close and reopen