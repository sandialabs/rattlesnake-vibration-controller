# -*- coding: utf-8 -*-
"""
This file defines a Modal Testing Environment where users can perform
hammer or shaker modal tests and export FRFs and other relevant data.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from qtpy import QtWidgets,uic
from qtpy.QtCore import Qt
from .abstract_environment import AbstractEnvironment,AbstractUI,AbstractMetadata
from .utilities import (Channel,VerboseMessageQueue,GlobalCommands,
                        DataAcquisitionParameters,error_message_qt,
                        load_python_module, flush_queue)
from .ui_utilities import multiline_plotter,ModalMDISubWindow
from .environments import (ControlTypes,environment_definition_ui_paths,
                           environment_prediction_ui_paths,
                           environment_run_ui_paths,
                           system_identification_ui_path,
                           modal_mdi_ui_path
                          )
from .signal_generation import (RandomSignalGenerator,BurstRandomSignalGenerator,
                                PseudorandomSignalGenerator,ChirpSignalGenerator,
                                SquareSignalGenerator,SineSignalGenerator)
from typing import List
from multiprocessing.queues import Queue
from datetime import datetime
import traceback
import os
import netCDF4 as nc4
import openpyxl
from enum import Enum
import multiprocessing as mp
import inspect
import numpy as np
import time
import pyqtgraph as pg
import scipy.signal as sig
import netCDF4 as nc4
from glob import glob

control_type = ControlTypes.MODAL
maximum_name_length = 50

WAIT_TIME = 0.02

class ModalCommands(Enum):
    START_CONTROL = 0
    STOP_CONTROL = 1
    ACCEPT_FRAME = 2
    RUN_CONTROL = 3
    CHECK_FOR_COMPLETE_SHUTDOWN = 4

class ModalQueues:
    """A set of queues used by the modal environment"""
    def __init__(self,environment_name : str,
                 environment_command_queue : VerboseMessageQueue,
                 gui_update_queue : mp.queues.Queue,
                 controller_communication_queue : VerboseMessageQueue,
                 data_in_queue : mp.queues.Queue,
                 data_out_queue : mp.queues.Queue,
                 log_file_queue : VerboseMessageQueue
                 ):
        """
        Creates a namespace to store all the queues used by the Modal Environment

        Parameters
        ----------
        environment_command_queue : VerboseMessageQueue
            Queue from which the environment will receive instructions.
        gui_update_queue : mp.queues.Queue
            Queue to which the environment will put GUI updates.
        controller_communication_queue : VerboseMessageQueue
            Queue to which the environment will put global contorller instructions.
        data_in_queue : mp.queues.Queue
            Queue from which the environment will receive data from acquisition.
        data_out_queue : mp.queues.Queue
            Queue to which the environment will write data for output.
        log_file_queue : VerboseMessageQueue
            Queue to which the environment will write log file messages.
        """
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.log_file_queue = log_file_queue
        self.data_for_spectral_computation_queue = mp.Queue()
        self.updated_spectral_quantities_queue = mp.Queue()
        self.signal_generation_update_queue = mp.Queue()
        self.spectral_command_queue = VerboseMessageQueue(log_file_queue,environment_name + ' Spectral Computation Command Queue')
        self.collector_command_queue = VerboseMessageQueue(log_file_queue,environment_name + ' Data Collector Command Queue')
        self.signal_generation_command_queue = VerboseMessageQueue(log_file_queue,environment_name + ' Signal Generation Command Queue')

class ModalMetadata(AbstractMetadata):
    """Class for storing metadata for an environment.
    
    This class is used as a storage container for parameters used by an
    environment.  It is returned by the environment UI's
    ``collect_environment_definition_parameters`` function as well as its 
    ``initialize_environment`` function.  Various parts of the controller and
    environment will query the class's data members for parameter values.
    """
    def __init__(self,sample_rate : float,
                 samples_per_frame : int,
                 averaging_type : str,
                 num_averages : int,
                 averaging_coefficient : float,
                 frf_technique : str,
                 frf_window : str,
                 overlap_percent : float,
                 trigger_type : str,
                 accept_type : str,
                 wait_for_steady_state : float,
                 trigger_channel : int,
                 pretrigger_percent : float,
                 trigger_slope_positive : bool,
                 trigger_level_percent : float,
                 hysteresis_level_percent : float,
                 hysteresis_frame_percent : float,
                 signal_generator_type : str,
                 signal_generator_level : float,
                 signal_generator_min_frequency : float,
                 signal_generator_max_frequency : float,
                 signal_generator_on_percent : float,
                 acceptance_function,
                 reference_channel_indices,
                 response_channel_indices,
                 output_channel_indices,
                 data_acquisition_parameters : DataAcquisitionParameters,
                 exponential_window_value_at_frame_end : float
                 ):
        self.sample_rate = sample_rate
        self.samples_per_frame = samples_per_frame
        self.averaging_type = averaging_type
        self.num_averages = num_averages
        self.averaging_coefficient = averaging_coefficient
        self.frf_technique = frf_technique
        self.frf_window = frf_window
        self.overlap = overlap_percent/100
        self.trigger_type = trigger_type
        self.accept_type = accept_type
        self.wait_for_steady_state = wait_for_steady_state
        self.trigger_channel = trigger_channel
        self.pretrigger = pretrigger_percent/100
        self.trigger_slope_positive = trigger_slope_positive
        self.trigger_level = trigger_level_percent/100
        self.hysteresis_level = hysteresis_level_percent/100
        self.hysteresis_length = hysteresis_frame_percent/100
        self.signal_generator_type = signal_generator_type
        self.signal_generator_level = signal_generator_level
        self.signal_generator_min_frequency = signal_generator_min_frequency
        self.signal_generator_max_frequency = signal_generator_max_frequency
        self.signal_generator_on_fraction = signal_generator_on_percent/100
        self.acceptance_function = acceptance_function
        self.reference_channel_indices = reference_channel_indices
        self.response_channel_indices = response_channel_indices
        self.output_channel_indices = output_channel_indices
        self.exponential_window_value_at_frame_end = exponential_window_value_at_frame_end
        # Set up signal generator
        self.output_oversample = data_acquisition_parameters.output_oversample
        self.signal_generator = self.get_signal_generator()
        
    def get_signal_generator(self):
        if self.signal_generator_type == 'none':
            signal_generator = PseudorandomSignalGenerator(
                rms = 0.0, 
                sample_rate = self.sample_rate,
                num_samples_per_frame = self.samples_per_frame,
                num_signals = len(self.output_channel_indices),
                low_frequency_cutoff = self.signal_generator_min_frequency,
                high_frequency_cutoff = self.signal_generator_max_frequency,
                output_oversample = self.output_oversample)
        elif self.signal_generator_type == 'random':
            signal_generator = RandomSignalGenerator(
                rms = self.signal_generator_level,
                sample_rate = self.sample_rate,
                num_samples_per_frame = self.samples_per_frame,
                num_signals = len(self.output_channel_indices),
                low_frequency_cutoff = self.signal_generator_min_frequency,
                high_frequency_cutoff = self.signal_generator_max_frequency,
                cola_overlap = 0.5,
                cola_window = 'hann',
                cola_exponent = 0.5,
                output_oversample = self.output_oversample)
        elif self.signal_generator_type == 'pseudorandom':
            signal_generator = PseudorandomSignalGenerator(
                rms = self.signal_generator_level, 
                sample_rate = self.sample_rate,
                num_samples_per_frame = self.samples_per_frame,
                num_signals = len(self.output_channel_indices),
                low_frequency_cutoff = self.signal_generator_min_frequency,
                high_frequency_cutoff = self.signal_generator_max_frequency,
                output_oversample = self.output_oversample)
        elif self.signal_generator_type == 'burst':
            signal_generator = BurstRandomSignalGenerator(
                rms = self.signal_generator_level, 
                sample_rate = self.sample_rate,
                num_samples_per_frame = self.samples_per_frame,
                num_signals = len(self.output_channel_indices),
                low_frequency_cutoff = self.signal_generator_min_frequency,
                high_frequency_cutoff = self.signal_generator_max_frequency,
                on_fraction = self.signal_generator_on_fraction,
                ramp_fraction = 0.05,
                output_oversample = self.output_oversample)
        elif self.signal_generator_type == 'chirp':
            signal_generator = ChirpSignalGenerator(
                level = self.signal_generator_level,
                sample_rate = self.sample_rate,
                num_samples_per_frame = self.samples_per_frame,
                num_signals = len(self.output_channel_indices),
                low_frequency_cutoff = self.signal_generator_min_frequency,
                high_frequency_cutoff = self.signal_generator_max_frequency,
                output_oversample = self.output_oversample)
        elif self.signal_generator_type == 'square':
            signal_generator = SquareSignalGenerator(
                level = self.signal_generator_level,
                sample_rate = self.sample_rate,
                num_samples_per_frame = self.samples_per_frame,
                num_signals = len(self.output_channel_indices),
                frequency = self.signal_generator_min_frequency,
                phase = 0,
                on_fraction = self.signal_generator_on_fraction,
                output_oversample = self.output_oversample)
        elif self.signal_generator_type == 'sine':
            signal_generator = SineSignalGenerator(
                level = self.signal_generator_level,
                sample_rate = self.sample_rate,
                num_samples_per_frame = self.samples_per_frame,
                num_signals = len(self.output_channel_indices),
                frequency = self.signal_generator_min_frequency,
                phase = 0,
                output_oversample = self.output_oversample)
        else:
            raise ValueError("Invalid Signal Type {:}".format(self.signal_generator_type))
        return signal_generator
        
    @property
    def samples_per_acquire(self):
        """Property returning the samples per acquisition step given the overlap"""
        return int(self.samples_per_frame*(1-self.overlap))
    
    @property
    def frame_time(self):
        """Property returning the time per measurement frame"""
        return self.samples_per_frame/self.sample_rate
    
    @property
    def nyquist_frequency(self):
        """Property returning half the sample rate"""
        return self.sample_rate/2
    
    @property
    def fft_lines(self):
        """Property returning the frequency lines given the sampling parameters"""
        return self.samples_per_frame//2 + 1
    
    @property
    def skip_frames(self):
        return int(np.ceil(
            self.wait_for_steady_state
            *self.sample_rate/
            (self.samples_per_frame
             *(1-self.overlap))
            ))
    
    @property
    def frequency_spacing(self):
        """Property returning frequency line spacing given the sampling parameters"""
        return self.sample_rate/self.samples_per_frame
    
    @property
    def overlapped_output_samples(self):
        """Property returning the number of output samples that are overlapped."""
        return self.samples_per_frame - self.samples_per_output
    
    def get_trigger_levels(self,channels):
        channel = channels[self.trigger_channel]
        try:
            volt_range = float(channel.maximum_value)
            if volt_range == 0.0:
                volt_range = 10.0
        except (ValueError,TypeError):
            volt_range = 10.0
        try:
            mv_per_eu = float(channel.sensitivity)
            if mv_per_eu == 0.0:
                mv_per_eu = 1000.0
        except (ValueError,TypeError):
            mv_per_eu = 1000.0
        v_per_eu = mv_per_eu/1000.0
        trigger_level_v = self.trigger_level*volt_range
        trigger_level_eu = trigger_level_v/v_per_eu
        hysterisis_level_v = self.hysteresis_level*volt_range
        hysterisis_level_eu = hysterisis_level_v/v_per_eu
        return trigger_level_v,trigger_level_eu,hysterisis_level_v,hysterisis_level_eu
    
    @property
    def disabled_signals(self):
        return [i for i,index in enumerate(self.output_channel_indices) 
                if not (index in self.response_channel_indices 
                        or index in self.reference_channel_indices)]
    
    @property
    def hysteresis_samples(self):
        return int(self.hysteresis_length*self.samples_per_frame)
    
    def store_to_netcdf(self,netcdf_group_handle : nc4._netCDF4.Group):
        """Store parameters to a group in a netCDF streaming file.
        
        This function stores parameters from the environment into the netCDF
        file in a group with the environment's name as its name.  The function
        will receive a reference to the group within the dataset and should
        store the environment's parameters into that group in the form of
        attributes, dimensions, or variables.
        
        This function is the "write" counterpart to the retrieve_metadata 
        function in the AbstractUI class, which will read parameters from
        the netCDF file to populate the parameters in the user interface.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            A reference to the Group within the netCDF dataset where the
            environment's metadata is stored.

        """
        netcdf_group_handle.samples_per_frame = self.samples_per_frame
        netcdf_group_handle.averaging_type = self.averaging_type
        netcdf_group_handle.num_averages = self.num_averages
        netcdf_group_handle.averaging_coefficient = self.averaging_coefficient
        netcdf_group_handle.frf_technique = self.frf_technique
        netcdf_group_handle.frf_window = self.frf_window
        netcdf_group_handle.overlap = self.overlap
        netcdf_group_handle.trigger_type = self.trigger_type
        netcdf_group_handle.accept_type = self.accept_type
        netcdf_group_handle.wait_for_steady_state = self.wait_for_steady_state
        netcdf_group_handle.trigger_channel = self.trigger_channel
        netcdf_group_handle.pretrigger = self.pretrigger
        netcdf_group_handle.trigger_slope_positive = 1 if self.trigger_slope_positive else 0
        netcdf_group_handle.trigger_level = self.trigger_level
        netcdf_group_handle.hysteresis_level = self.hysteresis_level
        netcdf_group_handle.hysteresis_length = self.hysteresis_length
        netcdf_group_handle.signal_generator_type = self.signal_generator_type
        netcdf_group_handle.signal_generator_level = self.signal_generator_level
        netcdf_group_handle.signal_generator_min_frequency = self.signal_generator_min_frequency
        netcdf_group_handle.signal_generator_max_frequency = self.signal_generator_max_frequency
        netcdf_group_handle.signal_generator_on_fraction = self.signal_generator_on_fraction
        netcdf_group_handle.exponential_window_value_at_frame_end = self.exponential_window_value_at_frame_end
        netcdf_group_handle.acceptance_function = self.acceptance_function[0]+':'+self.acceptance_function[1] if not self.acceptance_function is None else 'None'
        # Reference channels
        netcdf_group_handle.createDimension('reference_channels',len(self.reference_channel_indices))
        var = netcdf_group_handle.createVariable('reference_channel_indices','i4',('reference_channels'))
        var[...] = self.reference_channel_indices
        # Response channels
        netcdf_group_handle.createDimension('response_channels',len(self.response_channel_indices))
        var = netcdf_group_handle.createVariable('response_channel_indices','i4',('response_channels'))
        var[...] = self.response_channel_indices
    
    @classmethod
    def from_ui(cls,ui):
        """
        Creates a ModalMetadata object from the user interface

        Parameters
        ----------
        ui : ModalUI
            A Modal User Interface.

        Returns
        -------
        test_parameters : ModalMetadata
            Parameters corresponding to the data in the user interface

        """
        signal_generator_level = 0
        signal_generator_min_frequency = 0
        signal_generator_max_frequency = 0
        signal_generator_on_percent = 0
        if ui.definition_widget.signal_generator_selector.currentIndex() == 0: # None
            signal_generator_type = 'none'
        elif ui.definition_widget.signal_generator_selector.currentIndex() == 1: # Random
            signal_generator_type = 'random'
            signal_generator_level = ui.definition_widget.random_rms_selector.value()
            signal_generator_min_frequency = ui.definition_widget.random_min_frequency_selector.value()
            signal_generator_max_frequency = ui.definition_widget.random_max_frequency_selector.value()
        elif ui.definition_widget.signal_generator_selector.currentIndex() == 2: # Burst Random
            signal_generator_type = 'burst'
            signal_generator_level = ui.definition_widget.burst_rms_selector.value()
            signal_generator_min_frequency = ui.definition_widget.burst_min_frequency_selector.value()
            signal_generator_max_frequency = ui.definition_widget.burst_max_frequency_selector.value()
            signal_generator_on_percent = ui.definition_widget.burst_on_percentage_selector.value()
        elif ui.definition_widget.signal_generator_selector.currentIndex() == 3: # Pseudorandom
            signal_generator_type = 'pseudorandom'
            signal_generator_level = ui.definition_widget.pseudorandom_rms_selector.value()
            signal_generator_min_frequency = ui.definition_widget.pseudorandom_min_frequency_selector.value()
            signal_generator_max_frequency = ui.definition_widget.pseudorandom_max_frequency_selector.value()
        elif ui.definition_widget.signal_generator_selector.currentIndex() == 4: # Chirp
            signal_generator_type = 'chirp'
            signal_generator_level = ui.definition_widget.chirp_level_selector.value()
            signal_generator_min_frequency = ui.definition_widget.chirp_min_frequency_selector.value()
            signal_generator_max_frequency = ui.definition_widget.chirp_max_frequency_selector.value()
        elif ui.definition_widget.signal_generator_selector.currentIndex() == 5: # Square
            signal_generator_type = 'square'
            signal_generator_level = ui.definition_widget.square_level_selector.value()
            signal_generator_min_frequency = ui.definition_widget.square_frequency_selector.value()
            signal_generator_on_percent = ui.definition_widget.square_percent_on_selector.value()
        elif ui.definition_widget.signal_generator_selector.currentIndex() == 6: # Sine
            signal_generator_type = 'sine'
            signal_generator_level = ui.definition_widget.sine_level_selector.value()
            signal_generator_min_frequency = ui.definition_widget.sine_frequency_selector.value()
        return cls(
            ui.definition_widget.sample_rate_display.value(),
            ui.definition_widget.samples_per_frame_selector.value(),
            ui.definition_widget.system_id_averaging_scheme_selector.itemText(ui.definition_widget.system_id_averaging_scheme_selector.currentIndex()),
            ui.definition_widget.system_id_frames_to_average_selector.value(),
            ui.definition_widget.system_id_averaging_coefficient_selector.value(),
            ui.definition_widget.system_id_frf_technique_selector.itemText(ui.definition_widget.system_id_frf_technique_selector.currentIndex()),
            ui.definition_widget.system_id_transfer_function_computation_window_selector.itemText(ui.definition_widget.system_id_transfer_function_computation_window_selector.currentIndex()).lower(),
            ui.definition_widget.system_id_overlap_percentage_selector.value(),
            ui.definition_widget.triggering_type_selector.itemText(ui.definition_widget.triggering_type_selector.currentIndex()),
            ui.definition_widget.acceptance_selector.itemText(ui.definition_widget.acceptance_selector.currentIndex()),
            ui.definition_widget.wait_for_steady_selector.value(),
            ui.definition_widget.trigger_channel_selector.currentIndex(),
            ui.definition_widget.pretrigger_selector.value(),
            ui.definition_widget.trigger_slope_selector.currentIndex() == 0,
            ui.definition_widget.trigger_level_selector.value(),
            ui.definition_widget.hysteresis_selector.value(),
            ui.definition_widget.hysteresis_length_selector.value(),
            signal_generator_type,
            signal_generator_level,
            signal_generator_min_frequency,
            signal_generator_max_frequency,
            signal_generator_on_percent,
            ui.acceptance_function,
            ui.reference_indices,
            ui.response_indices,
            ui.all_output_channel_indices,
            ui.data_acquisition_parameters,
            ui.definition_widget.window_value_selector.value()/100
            )

    def generate_signal(self):
        if self.signal_generator is None:
            return np.zeros((
                len(self.output_channel_indices),self.samples_per_frame*self.output_oversample))
        else:
            return self.signal_generator.generate_frame()[0]

from .spectral_processing import (spectral_processing_process,
                                  SpectralProcessingCommands,
                                  SpectralProcessingMetadata,
                                  AveragingTypes,Estimator)
from .signal_generation_process import (signal_generation_process,
                                        SignalGenerationCommands,
                                        SignalGenerationMetadata)
from .data_collector import (data_collector_process,DataCollectorCommands,CollectorMetadata,
                             AcquisitionType,Acceptance,TriggerSlope,Window)

class ModalUI(AbstractUI):
    """Modal User Interface class defining the interface with the controller
    
    This class is used to define the interface between the User Interface of the
    Modal environment in the controller and the main controller."""
    
    def __init__(self,
                 environment_name : str,
                 definition_tabwidget : QtWidgets.QTabWidget,
                 system_id_tabwidget : QtWidgets.QTabWidget,
                 test_predictions_tabwidget : QtWidgets.QTabWidget,
                 run_tabwidget : QtWidgets.QTabWidget,
                 environment_command_queue : VerboseMessageQueue,
                 controller_communication_queue : VerboseMessageQueue,
                 log_file_queue : Queue):
        """
        Constructs a Modal User Interface
        
        Given the tab widgets from the main interface as well as communication
        queues, this class assembles the user interface components specific to
        the Modal Environment

        Parameters
        ----------
        definition_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Control
            Definition main tab
        system_id_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the System
            Identification main tab
        test_predictions_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Test Predictions
            main tab
        run_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Run
            main tab.
        environment_command_queue : VerboseMessageQueue
            Queue for sending commands to the Modal Environment
        controller_communication_queue : VerboseMessageQueue
            Queue for sending global commands to the controller
        log_file_queue : Queue
            Queue where log file messages can be written.
    
        """
        super().__init__(environment_name,
                         environment_command_queue,
                         controller_communication_queue,
                         log_file_queue)
        # Add the page to the control definition tabwidget
        self.definition_widget = QtWidgets.QWidget()
        uic.loadUi(environment_definition_ui_paths[control_type],self.definition_widget)
        definition_tabwidget.addTab(self.definition_widget,self.environment_name)
        # Add the page to the run tabwidget
        self.run_widget = QtWidgets.QWidget()
        uic.loadUi(environment_run_ui_paths[control_type],self.run_widget)
        run_tabwidget.addTab(self.run_widget,self.environment_name)
    
        self.trigger_widgets = [
            self.definition_widget.trigger_channel_selector,
            self.definition_widget.pretrigger_selector,
            self.definition_widget.trigger_slope_selector,
            self.definition_widget.trigger_level_selector,
            self.definition_widget.trigger_level_voltage_display,
            self.definition_widget.trigger_level_eu_display,
            self.definition_widget.hysteresis_selector,
            self.definition_widget.hysteresis_voltage_display,
            self.definition_widget.hysteresis_eu_display,
            self.definition_widget.hysteresis_length_selector,
            self.definition_widget.hysteresis_samples_display,
            self.definition_widget.hysteresis_time_display,]
    
        self.signal_generator_widgets = [
            self.definition_widget.random_rms_selector,
            self.definition_widget.random_min_frequency_selector,
            self.definition_widget.random_max_frequency_selector,
            self.definition_widget.burst_rms_selector,
            self.definition_widget.burst_min_frequency_selector,
            self.definition_widget.burst_max_frequency_selector,
            self.definition_widget.burst_on_percentage_selector,
            self.definition_widget.pseudorandom_rms_selector,
            self.definition_widget.pseudorandom_min_frequency_selector,
            self.definition_widget.pseudorandom_max_frequency_selector,
            self.definition_widget.chirp_level_selector,
            self.definition_widget.chirp_min_frequency_selector,
            self.definition_widget.chirp_max_frequency_selector,
            self.definition_widget.square_level_selector,
            self.definition_widget.square_frequency_selector,
            self.definition_widget.square_percent_on_selector,
            self.definition_widget.sine_level_selector,
            self.definition_widget.sine_frequency_selector
            ]
    
        self.window_parameter_widgets = [
            self.definition_widget.window_value_label,
            self.definition_widget.window_value_selector,
            ]
    
        self.definition_widget.reference_channels_selector.setColumnCount(3)
        self.definition_widget.reference_channels_selector.setVerticalHeaderLabels(['Enabled','Reference','Channel'])
    
        self.data_acquisition_parameters = None
        self.environment_parameters = None
        self.channel_names = None
        self.acceptance_function = None
        self.plot_data_items = {}
        self.reference_channel_indices = None
        self.all_output_channel_indices = None
        self.response_channel_indices = None
        self.last_frame = None
        self.last_frf = None
        self.last_coherence = None
        self.last_response_cpsd = None
        self.last_reference_cpsd = None
        self.last_condition = None
        self.acquiring = False
        self.netcdf_handle = None
        self.override_table = {}
        self.reciprocal_responses = []
        
        # Store some information into the channel display so the plots have
        # access to it
        self.run_widget.channel_display_area.time_abscissa = None
        self.run_widget.channel_display_area.frequency_abscissa = None
        self.run_widget.channel_display_area.window_function = None
        self.run_widget.channel_display_area.last_frame = None
        self.run_widget.channel_display_area.last_spectrum = None
        self.run_widget.channel_display_area.last_autospectrum = None
        self.run_widget.channel_display_area.last_frf = None
        self.run_widget.channel_display_area.last_coh = None
        self.run_widget.channel_display_area.channel_names = None
        self.run_widget.channel_display_area.reference_channel_indices = None
        self.run_widget.channel_display_area.response_channel_indices = None
    
        self.complete_ui()
        self.connect_callbacks()
    
    @property
    def reference_indices(self):
        return [i for i in range(
            self.definition_widget.reference_channels_selector.rowCount()) if 
            self.definition_widget.reference_channels_selector.cellWidget(i,0).isChecked() and 
            self.definition_widget.reference_channels_selector.cellWidget(i,1).isChecked()]
    
    @property
    def response_indices(self):
        return [i for i in range(
            self.definition_widget.reference_channels_selector.rowCount()) if 
            self.definition_widget.reference_channels_selector.cellWidget(i,0).isChecked() and 
            not self.definition_widget.reference_channels_selector.cellWidget(i,1).isChecked()]
    
    @property
    def output_channel_indices(self):
        return [i for i in self.all_output_channel_indices if self.definition_widget.reference_channels_selector.cellWidget(i,0).isChecked()]
    
    @property
    def initialized_response_names(self):
        return [self.channel_names[i] for i in range(len(self.channel_names)) if i not in self.environment_parameters.reference_channel_indices]
    
    @property
    def initialized_reference_names(self):
        return [self.channel_names[i] for i in self.environment_parameters.reference_channel_indices]
    
    def complete_ui(self):
        self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        for widget in self.trigger_widgets:
            widget.setEnabled(False)
        
        # Set common look and feel for plots
        plotWidgets = [self.definition_widget.output_signal_plot]
        for plotWidget in plotWidgets:
            plot_item = plotWidget.getPlotItem()
            plot_item.showGrid(True,True,0.25)
            plot_item.enableAutoRange()
            plot_item.getViewBox().enableAutoRange(enable=True)
            
        # Disable the currently inactive portions of the definition layout
        self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        for widget in self.window_parameter_widgets:
            widget.hide()
        
    def connect_callbacks(self):
        # Definition Callbacks
        self.definition_widget.samples_per_frame_selector.valueChanged.connect(self.update_parameters)
        self.definition_widget.system_id_overlap_percentage_selector.valueChanged.connect(self.update_parameters)
        self.definition_widget.triggering_type_selector.currentIndexChanged.connect(self.activate_trigger_options)
        self.definition_widget.acceptance_selector.currentIndexChanged.connect(self.select_acceptance)
        self.definition_widget.trigger_channel_selector.currentIndexChanged.connect(self.update_trigger_levels)
        self.definition_widget.trigger_level_selector.valueChanged.connect(self.update_trigger_levels)
        self.definition_widget.hysteresis_selector.valueChanged.connect(self.update_trigger_levels)
        self.definition_widget.regenerate_signal_button.clicked.connect(self.generate_signal)
        self.definition_widget.signal_generator_selector.currentChanged.connect(self.update_signal)
        for widget in self.signal_generator_widgets:
            widget.valueChanged.connect(self.update_signal)
        self.definition_widget.check_selected_button.clicked.connect(self.check_selected_reference_channels)
        self.definition_widget.uncheck_selected_button.clicked.connect(self.uncheck_selected_reference_channels)
        self.definition_widget.enable_selected_button.clicked.connect(self.enable_selected_channels)
        self.definition_widget.disable_selected_button.clicked.connect(self.disable_selected_channels)
        self.definition_widget.hysteresis_length_selector.valueChanged.connect(self.update_hysteresis_length)
        self.definition_widget.system_id_averaging_scheme_selector.currentIndexChanged.connect(self.update_averaging_type)
        self.definition_widget.system_id_transfer_function_computation_window_selector.currentIndexChanged.connect(self.update_window)
        # Run Callbacks
        self.run_widget.preview_test_button.clicked.connect(self.preview_acquisition)
        self.run_widget.start_test_button.clicked.connect(self.start_control)
        self.run_widget.stop_test_button.clicked.connect(self.stop_control)
        self.run_widget.select_file_button.clicked.connect(self.select_file)
        self.run_widget.accept_average_button.clicked.connect(self.accept_frame)
        self.run_widget.reject_average_button.clicked.connect(self.reject_frame)
        self.run_widget.new_window_button.clicked.connect(self.new_window)
        self.run_widget.new_from_template_combobox.currentIndexChanged.connect(self.new_window_from_template)
        self.run_widget.tile_layout_button.clicked.connect(self.run_widget.channel_display_area.tileSubWindows)
        self.run_widget.close_all_button.clicked.connect(self.close_windows)
        self.run_widget.decrement_channels_button.clicked.connect(self.decrement_channels)
        self.run_widget.increment_channels_button.clicked.connect(self.increment_channels)
        self.run_widget.dof_override_table.itemChanged.connect(self.update_override_table)
        self.run_widget.add_override_button.clicked.connect(self.add_override_channel)
        self.run_widget.remove_override_button.clicked.connect(self.remove_override_channel)
    
    # Definition Callbacks
    def update_parameters(self):
        if self.definition_widget.samples_per_frame_selector.value() % 2 == 1:
            self.definition_widget.samples_per_frame_selector.blockSignals(True)
            self.definition_widget.samples_per_frame_selector.setValue(self.definition_widget.samples_per_frame_selector.value()+1)
            self.definition_widget.samples_per_frame_selector.blockSignals(False)
        data = self.collect_environment_definition_parameters()
        self.definition_widget.samples_per_acquire_display.setValue(data.samples_per_acquire)
        self.definition_widget.frame_time_display.setValue(data.frame_time)
        self.definition_widget.nyquist_frequency_display.setValue(data.nyquist_frequency)
        self.definition_widget.fft_lines_display.setValue(data.fft_lines)
        self.definition_widget.frequency_spacing_display.setValue(data.frequency_spacing)
        if self.definition_widget.regenerate_signal_auto_checkbox.isChecked():
            self.generate_signal()

    def update_reference_channels(self):
        self.definition_widget.response_channels_display.setValue(len(self.response_indices))
        self.definition_widget.reference_channels_display.setValue(len(self.reference_indices))
        self.definition_widget.output_channels_display.setValue(len(self.output_channel_indices))
        if self.definition_widget.regenerate_signal_auto_checkbox.isChecked():
            self.generate_signal()

    def check_selected_reference_channels(self):
        select = self.definition_widget.reference_channels_selector.selectionModel()
        rows = select.selectedRows()
        for row in rows:
            index = row.row()
            self.definition_widget.reference_channels_selector.cellWidget(index,1).setChecked(True)
        
    def uncheck_selected_reference_channels(self):
        select = self.definition_widget.reference_channels_selector.selectionModel()
        rows = select.selectedRows()
        for row in rows:
            index = row.row()
            self.definition_widget.reference_channels_selector.cellWidget(index,1).setChecked(False)

    def enable_selected_channels(self):
        select = self.definition_widget.reference_channels_selector.selectionModel()
        rows = select.selectedRows()
        for row in rows:
            index = row.row()
            self.definition_widget.reference_channels_selector.cellWidget(index,0).setChecked(True)
        
    def disable_selected_channels(self):
        select = self.definition_widget.reference_channels_selector.selectionModel()
        rows = select.selectedRows()
        for row in rows:
            index = row.row()
            self.definition_widget.reference_channels_selector.cellWidget(index,0).setChecked(False)

    def activate_trigger_options(self):
        if self.definition_widget.triggering_type_selector.currentIndex() == 0:
            for widget in self.trigger_widgets:
                widget.setEnabled(False)
        else:
            for widget in self.trigger_widgets:
                widget.setEnabled(True)
    
    def select_acceptance(self):
        if self.definition_widget.acceptance_selector.currentIndex() == 2:
            # Open up a file dialog
            filename,file_filter = QtWidgets.QFileDialog.getOpenFileName(self.definition_widget,'Select Python Module',filter='Python Modules (*.py)')
            if filename == '':
                self.definition_widget.acceptance_selector.setCurrentIndex(0)
                return
            module = load_python_module(filename)
            functions = [function for function in inspect.getmembers(module)
                     if inspect.isfunction(function[1])]
            item,okPressed = QtWidgets.QInputDialog.getItem(self.definition_widget,"Select Acceptance Function",
                                                            "Function Name:",[function[0] for function in functions],0,False)
            if okPressed:
                self.acceptance_function=[filename,item]
            else:
                self.definition_widget.acceptance_selector.setCurrentIndex(0)
                return
        else:
            self.acceptance_function = None

    def update_trigger_levels(self):
        data = self.collect_environment_definition_parameters()
        t_v, t_eu, h_v, h_eu = data.get_trigger_levels(self.data_acquisition_parameters.channel_list)
        self.definition_widget.trigger_level_voltage_display.setValue(t_v)
        self.definition_widget.trigger_level_eu_display.setValue(t_eu)
        self.definition_widget.hysteresis_voltage_display.setValue(h_v)
        self.definition_widget.hysteresis_eu_display.setValue(h_eu)
        eu_suffix = self.data_acquisition_parameters.channel_list[data.trigger_channel].unit
        self.definition_widget.hysteresis_eu_display.setSuffix((' '+eu_suffix) if not (eu_suffix == '' or eu_suffix == None) else '')
        self.definition_widget.trigger_level_eu_display.setSuffix((' '+eu_suffix) if not (eu_suffix == '' or eu_suffix == None) else '')

    def update_hysteresis_length(self):
        data = self.collect_environment_definition_parameters()
        self.definition_widget.hysteresis_samples_display.setValue(data.hysteresis_samples)
        self.definition_widget.hysteresis_time_display.setValue(data.hysteresis_samples/data.sample_rate)

    def update_signal(self):
        if self.definition_widget.regenerate_signal_auto_checkbox.isChecked():
            self.generate_signal()

    def generate_signal(self):
        if self.data_acquisition_parameters is None:
            return
        output_oversample = self.data_acquisition_parameters.output_oversample
        output_rate = self.data_acquisition_parameters.output_sample_rate
        data = self.collect_environment_definition_parameters()
        frame_output_samples = int(data.samples_per_frame*output_oversample)
        signal = data.generate_signal()
        # Reduce down to just one frame
        while signal.shape[-1] < frame_output_samples:
            signal = np.concatenate((signal,data.generate_signal()),axis=-1)
        signal = signal[...,:frame_output_samples]
        signal[data.disabled_signals] = 0
        times = np.arange(frame_output_samples)/output_rate
        for s,plot in zip(signal,self.plot_data_items['signal_representation']):
            plot.setData(times,s)

    def update_averaging_type(self):
        if self.definition_widget.system_id_averaging_scheme_selector.currentIndex() == 0:
            self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        else:
            self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(True)
    
    def update_window(self):
        if self.definition_widget.system_id_transfer_function_computation_window_selector.currentIndex() == 2:
            for widget in self.window_parameter_widgets:
                widget.show()
        else:
            for widget in self.window_parameter_widgets:
                widget.hide()

    # Run Callbacks
    def preview_acquisition(self):
        self.run_widget.stop_test_button.setEnabled(True)
        self.run_widget.preview_test_button.setEnabled(False)
        self.run_widget.start_test_button.setEnabled(False)
        self.run_widget.select_file_button.setEnabled(False)
        self.controller_communication_queue.put(self.log_name,(GlobalCommands.START_ENVIRONMENT,self.environment_name))
        self.environment_command_queue.put(self.log_name,(ModalCommands.START_CONTROL,None))
        self.run_widget.dof_override_table.setEnabled(False)
        self.run_widget.add_override_button.setEnabled(False)
        self.run_widget.remove_override_button.setEnabled(False)
        
    def start_control(self):
        self.acquiring = True
        # Create the output file
        filename = self.run_widget.data_file_selector.text()
        if filename == '':
            error_message_qt('Invalid File', 'Please select a file in which to store modal data')
            return
        if self.run_widget.autoincrement_checkbox.isChecked():
            # Add the file increment
            path,ext = os.path.splitext(filename)
            index = len(glob(path+'*'+ext))
            filename = path+'_{:04d}'.format(index)+ext
        self.create_netcdf_file(filename)
        self.preview_acquisition()
        
    def stop_control(self):
        self.environment_command_queue.put(self.log_name,(ModalCommands.STOP_CONTROL,None))
    
    def select_file(self):
        filename,file_filter = QtWidgets.QFileDialog.getSaveFileName(self.run_widget,'Select NetCDF File to Save Modal Data',filter='NetCDF File (*.nc4)')
        if filename == '':
            return
        self.run_widget.data_file_selector.setText(filename)
        
    def accept_frame(self):
        self.environment_command_queue.put(self.log_name,(ModalCommands.ACCEPT_FRAME,True))
        self.run_widget.accept_average_button.setEnabled(False)
        self.run_widget.reject_average_button.setEnabled(False)
    
    def reject_frame(self):
        self.environment_command_queue.put(self.log_name,(ModalCommands.ACCEPT_FRAME,False))
        self.run_widget.accept_average_button.setEnabled(False)
        self.run_widget.reject_average_button.setEnabled(False)
    
    def new_window(self):
        widget = ModalMDISubWindow(self.run_widget.channel_display_area)
        self.run_widget.channel_display_area.addSubWindow(widget)
        widget.show()
        return widget
        # print('Windows: {:}'.format(self.run_widget.channel_display_area.subWindowList()))
        
    def new_window_from_template(self):
        if self.run_widget.new_from_template_combobox.currentIndex() == 0:
            return
        elif self.run_widget.new_from_template_combobox.currentIndex() == 6:
            # 3x3 channel grid
            for i in range(9):
                widget = self.new_window()
                widget.signal_selector.setCurrentIndex(0)
                widget.response_coordinate_selector.setCurrentIndex(i)
                widget.lock_response_checkbox.setChecked(False)
        elif self.run_widget.new_from_template_combobox.currentIndex() == 5:
            # Reference autospectra
            for index in self.run_widget.channel_display_area.reference_channel_indices:
                widget = self.new_window()
                widget.signal_selector.setCurrentIndex(3)
                widget.response_coordinate_selector.setCurrentIndex(index)
                widget.lock_response_checkbox.setChecked(True)
        else:
            corresponding_drive_responses = self.run_widget.channel_display_area.reciprocal_responses                
            if self.run_widget.new_from_template_combobox.currentIndex() == 1:
                # Create drive point FRFs in magnitude
                for i,index in enumerate(corresponding_drive_responses):
                    widget = self.new_window()
                    widget.signal_selector.setCurrentIndex(4)
                    widget.response_coordinate_selector.setCurrentIndex(index)
                    widget.reference_coordinate_selector.setCurrentIndex(i)
                    widget.data_type_selector.setCurrentIndex(0)
                    widget.lock_response_checkbox.setChecked(True)
            elif self.run_widget.new_from_template_combobox.currentIndex() == 2:
                # Create drive point FRFs in imaginary
                for i,index in enumerate(corresponding_drive_responses):
                    widget = self.new_window()
                    widget.signal_selector.setCurrentIndex(4)
                    widget.response_coordinate_selector.setCurrentIndex(index)
                    widget.reference_coordinate_selector.setCurrentIndex(i)
                    widget.data_type_selector.setCurrentIndex(3)
                    widget.lock_response_checkbox.setChecked(True)
            elif self.run_widget.new_from_template_combobox.currentIndex() == 3:
                # Create drive point Coherence
                for i,index in enumerate(corresponding_drive_responses):
                    widget = self.new_window()
                    widget.signal_selector.setCurrentIndex(6)
                    widget.response_coordinate_selector.setCurrentIndex(index)
                    widget.reference_coordinate_selector.setCurrentIndex(i)
                    widget.lock_response_checkbox.setChecked(True)
            elif self.run_widget.new_from_template_combobox.currentIndex() == 4:
                # Create drive point Coherence
                for i,index in enumerate(corresponding_drive_responses):
                    for j, index in enumerate(corresponding_drive_responses):
                        if i <= j:
                            continue
                        widget = self.new_window()
                        widget.signal_selector.setCurrentIndex(7)
                        widget.response_coordinate_selector.setCurrentIndex(i)
                        widget.reference_coordinate_selector.setCurrentIndex(j)
                        widget.lock_response_checkbox.setChecked(True)
        self.run_widget.new_from_template_combobox.setCurrentIndex(0)
    
    def close_windows(self):
        for window in self.run_widget.channel_display_area.subWindowList():
            window.close()
            
    def decrement_channels(self):
        number = -self.run_widget.increment_channels_number.value()
        for window in self.run_widget.channel_display_area.subWindowList():
            window.widget().increment_channel(number)
    
    def increment_channels(self):
        number = self.run_widget.increment_channels_number.value()
        for window in self.run_widget.channel_display_area.subWindowList():
            window.widget().increment_channel(number)

    def add_override_channel(self):
        selected_row = self.run_widget.dof_override_table.blockSignals(True)
        selected_row = self.run_widget.dof_override_table.rowCount()
        self.run_widget.dof_override_table.insertRow(selected_row)
        channel_combobox = QtWidgets.QComboBox()
        for channel_name in self.channel_names:
            channel_combobox.addItem(channel_name)
        channel_combobox.currentIndexChanged.connect(self.update_override_table)
        self.run_widget.dof_override_table.setCellWidget(selected_row,0,channel_combobox)
        data_item = QtWidgets.QTableWidgetItem()
        data_item.setText('1')
        self.run_widget.dof_override_table.setItem(selected_row,1,data_item)
        data_item = QtWidgets.QTableWidgetItem()
        data_item.setText('X+')
        self.run_widget.dof_override_table.setItem(selected_row,2,data_item)
        selected_row = self.run_widget.dof_override_table.blockSignals(False)
        self.update_override_table()
    
    def remove_override_channel(self):
        selected_row = self.run_widget.dof_override_table.currentRow()
        if selected_row >= 0:
            self.run_widget.dof_override_table.removeRow(selected_row)
        self.update_override_table()

    def update_override_table(self):
        self.override_table = {}
        for row in range(self.run_widget.dof_override_table.rowCount()):
            index = self.run_widget.dof_override_table.cellWidget(row,0).currentIndex()
            new_node = self.run_widget.dof_override_table.item(row,1).text()
            new_direction = self.run_widget.dof_override_table.item(row,2).text()
            self.override_table[index] = [new_node,new_direction]
        self.update_channel_names()
        self.run_widget.channel_display_area.reciprocal_responses = self.get_reciprocal_measurements()
        # Go through and update all the existing windows in the MDI display
        for window in self.run_widget.channel_display_area.subWindowList():
            widget = window.widget()
            current_response = widget.response_coordinate_selector.currentIndex()
            current_reference = widget.reference_coordinate_selector.currentIndex()
            current_data_type = widget.data_type_selector.currentIndex()
            widget.channel_names = self.channel_names
            widget.reference_names = [self.channel_names[i] for i in self.run_widget.channel_display_area.reference_channel_indices]
            widget.response_names = [self.channel_names[i] for i in self.run_widget.channel_display_area.response_channel_indices]
            widget.reciprocal_responses = self.run_widget.channel_display_area.reciprocal_responses
            widget.update_ui()
            widget.response_coordinate_selector.setCurrentIndex(current_response)
            widget.reference_coordinate_selector.setCurrentIndex(current_reference)
            widget.data_type_selector.setCurrentIndex(current_data_type)
        
    def get_reciprocal_measurements(self):
        node_numbers = np.array([channel.node_number if not i in self.override_table else self.override_table[i][0] 
                                 for i,channel in enumerate(self.data_acquisition_parameters.channel_list)])
        node_directions = np.array([''.join([char for char in (
            channel.node_direction if not i in self.override_table else self.override_table[i][1])
            if not char in '+-']) for i,channel in enumerate(self.data_acquisition_parameters.channel_list)])
        reference_node_numbers = node_numbers[self.environment_parameters.reference_channel_indices]
        reference_node_directions = node_directions[self.environment_parameters.reference_channel_indices]
        response_node_numbers = node_numbers[self.environment_parameters.response_channel_indices]
        response_node_directions = node_directions[self.environment_parameters.response_channel_indices]
        corresponding_drive_responses = []
        for node,direction in zip(reference_node_numbers,reference_node_directions):
            index = np.where((response_node_numbers == node) & (response_node_directions == direction))[0]
            if len(index) == 0:
                corresponding_drive_responses.append(None)
            elif len(index) > 1:
                corresponding_drive_responses.append(None)
            else:
                corresponding_drive_responses.append(index[0])
        # print(corresponding_drive_responses)
        return corresponding_drive_responses
        
    def create_netcdf_file(self,filename):
        self.netcdf_handle = nc4.Dataset(filename,'w',format='NETCDF4',clobber=True)
        # Create dimensions
        self.netcdf_handle.createDimension('response_channels',len(self.data_acquisition_parameters.channel_list))
        self.netcdf_handle.createDimension('output_channels',len([channel for channel in self.data_acquisition_parameters.channel_list if not channel.feedback_device is None]))
        self.netcdf_handle.createDimension('num_environments',len(self.data_acquisition_parameters.environment_names))
        self.netcdf_handle.createDimension('time_samples',None)
        # Create attributes
        self.netcdf_handle.sample_rate = self.data_acquisition_parameters.sample_rate
        self.netcdf_handle.time_per_write = self.data_acquisition_parameters.samples_per_write/self.data_acquisition_parameters.output_sample_rate
        self.netcdf_handle.time_per_read = self.data_acquisition_parameters.samples_per_read/self.data_acquisition_parameters.sample_rate
        self.netcdf_handle.hardware = self.data_acquisition_parameters.hardware
        self.netcdf_handle.hardware_file = 'None' if self.data_acquisition_parameters.hardware_file is None else self.data_acquisition_parameters.hardware_file
        self.netcdf_handle.maximum_acquisition_processes = self.data_acquisition_parameters.maximum_acquisition_processes
        self.netcdf_handle.output_oversample = self.data_acquisition_parameters.output_oversample
        # Create Variables
        self.netcdf_handle.createVariable('time_data','f8',('response_channels','time_samples'))
        var = self.netcdf_handle.createVariable('environment_names',str,('num_environments',))
        this_environment_index = None
        for i,name in enumerate(self.data_acquisition_parameters.environment_names):
            var[i] = name
            if name == self.environment_name:
                this_environment_index = i
        var = self.netcdf_handle.createVariable('environment_active_channels','i1',('response_channels','num_environments'))
        var[...] = self.data_acquisition_parameters.environment_active_channels.astype('int8')[
            self.data_acquisition_parameters.environment_active_channels[:,this_environment_index],:]
        # Create channel table variables
        labels = [['node_number',str],
                  ['node_direction',str],
                  ['comment',str],
                  ['serial_number',str],
                  ['triax_dof',str],
                  ['sensitivity',str],
                  ['unit',str],
                  ['make',str],
                  ['model',str],
                  ['expiration',str],
                  ['physical_device',str],
                  ['physical_channel',str],
                  ['channel_type',str],
                  ['minimum_value',str],
                  ['maximum_value',str],
                  ['coupling',str],
                  ['excitation_source',str],
                  ['excitation',str],
                  ['feedback_device',str],
                  ['feedback_channel',str],
                  ['warning_level',str],
                  ['abort_level',str],
                  ]
        for (label,netcdf_datatype) in labels:
            var = self.netcdf_handle.createVariable('/channels/'+label,netcdf_datatype,('response_channels',))
            channel_data = [getattr(channel,label) for channel in self.data_acquisition_parameters.channel_list]
            if netcdf_datatype == 'i1':
                channel_data = np.array([1 if val else 0 for val in channel_data])
            else:
                channel_data = ['' if val is None else val for val in channel_data]
            for i,cd in enumerate(channel_data):
                if label == 'node_number' and i in self.override_table:
                    var[i] = self.override_table[i][0]
                elif label == 'node_direction' and i in self.override_table:
                    var[i] = self.override_table[i][1]
                else:
                    var[i] = cd
        group_handle = self.netcdf_handle.createGroup(self.environment_name)
        self.environment_parameters.store_to_netcdf(group_handle)
        group_handle.createDimension('fft_lines',self.environment_parameters.fft_lines)
        group_handle.createVariable('frf_data_real','f8',('fft_lines','response_channels','reference_channels'))
        group_handle.createVariable('frf_data_imag','f8',('fft_lines','response_channels','reference_channels'))
        group_handle.createVariable('coherence','f8',('fft_lines','response_channels'))
        
    def collect_environment_definition_parameters(self) -> AbstractMetadata:
        """
        Collect the parameters from the user interface defining the environment

        Returns
        -------
        ModalMetadata
            A metadata or parameters object containing the parameters defining
            the corresponding environment.

        """
        return ModalMetadata.from_ui(self)

    def update_channel_names(self):
        self.channel_names = []
        for i,channel in enumerate(self.data_acquisition_parameters.channel_list):
            self.channel_names.append(
                '{:} {:} {:}'.format('' if channel.channel_type is None else channel.channel_type,
                                     channel.node_number if not i in self.override_table else self.override_table[i][0],
                                     channel.node_direction if not i in self.override_table else self.override_table[i][1])[:maximum_name_length])
        self.run_widget.channel_display_area.channel_names = self.channel_names
                              
    def initialize_data_acquisition(self,data_acquisition_parameters : DataAcquisitionParameters):
        """Update the user interface with data acquisition parameters
        
        This function is called when the Data Acquisition parameters are
        initialized.  This function should set up the environment user interface
        accordingly.

        Parameters
        ----------
        data_acquisition_parameters : DataAcquisitionParameters :
            Container containing the data acquisition parameters, including
            channel table and sampling information.

        """
        self.data_acquisition_parameters = data_acquisition_parameters
        self.definition_widget.sample_rate_display.setValue(data_acquisition_parameters.sample_rate)
        self.all_output_channel_indices = [index for index,channel in enumerate(self.data_acquisition_parameters.channel_list) if not channel.feedback_device is None]
        self.update_channel_names()
        self.definition_widget.reference_channels_selector.setRowCount(0)
        self.definition_widget.trigger_channel_selector.blockSignals(True)
        self.definition_widget.trigger_channel_selector.clear()
        for i,channel_name in enumerate(self.channel_names):
            self.definition_widget.trigger_channel_selector.addItem(channel_name)
            self.definition_widget.reference_channels_selector.insertRow(i)
            item = QtWidgets.QTableWidgetItem()
            item.setText(channel_name)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.definition_widget.reference_channels_selector.setItem(i,2,item)
            ref_checkbox = QtWidgets.QCheckBox()
            ref_checkbox.stateChanged.connect(self.update_reference_channels)
            self.definition_widget.reference_channels_selector.setCellWidget(i,1,ref_checkbox)
            enabled_checkbox = QtWidgets.QCheckBox()
            enabled_checkbox.setChecked(True)
            enabled_checkbox.stateChanged.connect(self.update_reference_channels)
            self.definition_widget.reference_channels_selector.setCellWidget(i,0,enabled_checkbox)
        self.definition_widget.trigger_channel_selector.blockSignals(False)
        self.update_trigger_levels()
            
        checked_state = self.definition_widget.regenerate_signal_auto_checkbox.isChecked()
        self.definition_widget.regenerate_signal_auto_checkbox.setChecked(False)
        self.definition_widget.signal_generator_selector.setCurrentIndex(0)
        self.definition_widget.samples_per_frame_selector.setValue(data_acquisition_parameters.sample_rate)
        self.definition_widget.random_max_frequency_selector.setValue(data_acquisition_parameters.sample_rate/2)
        self.definition_widget.random_min_frequency_selector.setValue(0)
        self.definition_widget.burst_max_frequency_selector.setValue(data_acquisition_parameters.sample_rate/2)
        self.definition_widget.burst_min_frequency_selector.setValue(0)
        self.definition_widget.chirp_max_frequency_selector.setValue(data_acquisition_parameters.sample_rate/2)
        self.definition_widget.chirp_min_frequency_selector.setValue(0)
        self.definition_widget.pseudorandom_max_frequency_selector.setValue(data_acquisition_parameters.sample_rate/2)
        self.definition_widget.pseudorandom_min_frequency_selector.setValue(0)
            
        self.definition_widget.response_channels_display.setValue(len(self.channel_names))
        self.definition_widget.reference_channels_display.setValue(0)
        num_outputs = len(self.output_channel_indices)
        self.definition_widget.output_channels_display.setValue(num_outputs)
        if num_outputs == 0:
            for i in range(self.definition_widget.signal_generator_selector.count()-1):
                self.definition_widget.signal_generator_selector.setTabEnabled(i+1,False)
        
        self.definition_widget.output_signal_plot.getPlotItem().clear()
        self.plot_data_items['signal_representation'] = multiline_plotter(
                (0,1),
                np.zeros((len(self.all_output_channel_indices),2)),
                widget=self.definition_widget.output_signal_plot,
                other_pen_options={'width':1},
                names = ['Output {:}'.format(i+1) for i in range(len(self.all_output_channel_indices))])
        self.definition_widget.regenerate_signal_auto_checkbox.setChecked(checked_state)
        if checked_state:
            self.generate_signal()
        
        for widget in [
            self.definition_widget.random_min_frequency_selector,
            self.definition_widget.random_max_frequency_selector,
            self.definition_widget.burst_min_frequency_selector,
            self.definition_widget.burst_max_frequency_selector,
            self.definition_widget.pseudorandom_min_frequency_selector,
            self.definition_widget.pseudorandom_max_frequency_selector,
            self.definition_widget.chirp_min_frequency_selector,
            self.definition_widget.chirp_max_frequency_selector,
            self.definition_widget.square_frequency_selector,
            self.definition_widget.sine_frequency_selector,
            ]:
            widget.setMaximum(self.data_acquisition_parameters.sample_rate/2)

    def initialize_environment(self) -> AbstractMetadata:
        """
        Update the user interface with environment parameters
        
        This function is called when the Environment parameters are initialized.
        This function should set up the user interface accordingly.  It must
        return the parameters class of the environment that inherits from
        AbstractMetadata.

        Returns
        ModalMetadata
            An AbstractMetadata-inheriting object that contains the parameters
            defining the environment.

        """
        self.environment_parameters = self.collect_environment_definition_parameters()
        self.reference_channel_indices = self.environment_parameters.reference_channel_indices
        self.response_channel_indices = self.environment_parameters.response_channel_indices
        self.run_widget.channel_display_area.reference_channel_indices = self.reference_channel_indices
        self.run_widget.channel_display_area.response_channel_indices = self.response_channel_indices
        for window in self.run_widget.channel_display_area.subWindowList():
            widget = window.widget()
            current_response = widget.response_coordinate_selector.currentIndex()
            current_reference = widget.reference_coordinate_selector.currentIndex()
            current_data_type = widget.data_type_selector.currentIndex()
            widget.reference_names = np.array([widget.channel_names[i] for i in self.run_widget.channel_display_area.reference_channel_indices])
            widget.response_names = np.array([widget.channel_names[i] for i in self.run_widget.channel_display_area.response_channel_indices])
            widget.update_ui()
            widget.response_coordinate_selector.setCurrentIndex(current_response)
            widget.reference_coordinate_selector.setCurrentIndex(current_reference)
            widget.data_type_selector.setCurrentIndex(current_data_type)
        self.run_widget.total_averages_display.setValue(self.environment_parameters.num_averages)
        self.run_widget.channel_display_area.time_abscissa = np.arange(self.environment_parameters.samples_per_frame)/self.environment_parameters.sample_rate
        self.run_widget.channel_display_area.frequency_abscissa = np.fft.rfftfreq(self.environment_parameters.samples_per_frame,1/self.environment_parameters.sample_rate)
        if self.environment_parameters.frf_window == 'rectangle':
            window = 1
        elif self.environment_parameters.frf_window == 'exponential':
            window_parameter = -(self.environment_parameters.samples_per_frame) / np.log(self.environment_parameters.exponential_window_value_at_frame_end)
            window = sig.get_window(('exponential',0,window_parameter),self.environment_parameters.samples_per_frame,fftbins=True)
        else:
            window = sig.get_window(self.environment_parameters.frf_window,self.environment_parameters.samples_per_frame,fftbins=True)
        self.run_widget.channel_display_area.window_function = window
        self.run_widget.channel_display_area.reciprocal_responses = self.get_reciprocal_measurements()
        
        return self.environment_parameters
        
    def retrieve_metadata(self, netcdf_handle : nc4._netCDF4.Dataset):
        """Collects environment parameters from a netCDF dataset.

        This function retrieves parameters from a netCDF dataset that was written
        by the controller during streaming.  It must populate the widgets
        in the user interface with the proper information.

        This function is the "read" counterpart to the store_to_netcdf 
        function in the ModalMetadata class, which will write parameters to
        the netCDF file to document the metadata.
        
        Note that the entire dataset is passed to this function, so the function
        should collect parameters pertaining to the environment from a Group
        in the dataset sharing the environment's name, e.g.
        
        ``group = netcdf_handle.groups[self.environment_name]``
        ``self.definition_widget.parameter_selector.setValue(group.parameter)``
        
        Parameters
        ----------
        netcdf_handle : nc4._netCDF4.Dataset :
            The netCDF dataset from which the data will be read.  It should have
            a group name with the enviroment's name.

        """
        netcdf_group_handle = netcdf_handle[self.environment_name]
        self.definition_widget.samples_per_frame_selector.setValue(netcdf_group_handle.samples_per_frame)
        self.definition_widget.system_id_averaging_scheme_selector.setCurrentIndex(self.definition_widget.system_id_averaging_scheme_selector.findText(netcdf_group_handle.averaging_type))
        self.definition_widget.system_id_frames_to_average_selector.setValue(netcdf_group_handle.num_averages)
        self.definition_widget.system_id_averaging_coefficient_selector.setValue(netcdf_group_handle.averaging_coefficient)
        self.definition_widget.system_id_frf_technique_selector.setCurrentIndex(self.definition_widget.system_id_frf_technique_selector.findText(netcdf_group_handle.frf_technique))
        self.definition_widget.system_id_transfer_function_computation_window_selector.setCurrentIndex(self.definition_widget.system_id_transfer_function_computation_window_selector.findText(netcdf_group_handle.frf_window.capitalize()))
        self.definition_widget.system_id_overlap_percentage_selector.setValue(netcdf_group_handle.overlap*100)
        self.definition_widget.triggering_type_selector.setCurrentIndex(self.definition_widget.triggering_type_selector.findText(netcdf_group_handle.trigger_type))
        acceptance = netcdf_group_handle.accept_type
        self.definition_widget.acceptance_selector.blockSignals(True)
        self.definition_widget.acceptance_selector.setCurrentIndex(self.definition_widget.acceptance_selector.findText(acceptance))
        self.definition_widget.acceptance_selector.blockSignals(False)
        if acceptance == 'Autoreject...':
            self.acceptance_function = netcdf_group_handle.acceptance_function.split(':')
        else:
            self.acceptance_function = None
        self.definition_widget.wait_for_steady_selector.setValue(netcdf_group_handle.wait_for_steady_state)
        self.definition_widget.trigger_channel_selector.setCurrentIndex(netcdf_group_handle.trigger_channel)
        self.definition_widget.pretrigger_selector.setValue(netcdf_group_handle.pretrigger*100)
        self.definition_widget.trigger_slope_selector.setCurrentIndex(0 if netcdf_group_handle.trigger_slope_positive == 1 else 1)
        self.definition_widget.trigger_level_selector.setValue(100*netcdf_group_handle.trigger_level)
        self.definition_widget.hysteresis_selector.setValue(100*netcdf_group_handle.hysteresis_level)
        self.definition_widget.hysteresis_length_selector.setValue(100*netcdf_group_handle.hysteresis_length)
        self.definition_widget.signal_generator_selector.setCurrentIndex(['none','random','burst','pseudorandom','chirp','square','sine'].index(netcdf_group_handle.signal_generator_type))
        self.definition_widget.random_rms_selector.setValue(netcdf_group_handle.signal_generator_level)
        self.definition_widget.random_min_frequency_selector.setValue(netcdf_group_handle.signal_generator_min_frequency)
        self.definition_widget.random_max_frequency_selector.setValue(netcdf_group_handle.signal_generator_max_frequency)
        self.definition_widget.burst_rms_selector.setValue(netcdf_group_handle.signal_generator_level)
        self.definition_widget.burst_min_frequency_selector.setValue(netcdf_group_handle.signal_generator_min_frequency)
        self.definition_widget.burst_max_frequency_selector.setValue(netcdf_group_handle.signal_generator_max_frequency)
        self.definition_widget.burst_on_percentage_selector.setValue(100*netcdf_group_handle.signal_generator_on_fraction)
        self.definition_widget.pseudorandom_rms_selector.setValue(netcdf_group_handle.signal_generator_level)
        self.definition_widget.pseudorandom_min_frequency_selector.setValue(netcdf_group_handle.signal_generator_min_frequency)
        self.definition_widget.pseudorandom_max_frequency_selector.setValue(netcdf_group_handle.signal_generator_max_frequency)
        self.definition_widget.chirp_level_selector.setValue(netcdf_group_handle.signal_generator_level)
        self.definition_widget.chirp_min_frequency_selector.setValue(netcdf_group_handle.signal_generator_min_frequency)
        self.definition_widget.chirp_max_frequency_selector.setValue(netcdf_group_handle.signal_generator_max_frequency)
        self.definition_widget.square_level_selector.setValue(netcdf_group_handle.signal_generator_level)
        self.definition_widget.square_frequency_selector.setValue(netcdf_group_handle.signal_generator_min_frequency)
        self.definition_widget.square_percent_on_selector.setValue(100*netcdf_group_handle.signal_generator_on_fraction)
        self.definition_widget.sine_level_selector.setValue(netcdf_group_handle.signal_generator_level)
        self.definition_widget.sine_frequency_selector.setValue(netcdf_group_handle.signal_generator_min_frequency)
        self.definition_widget.window_value_selector.setValue(netcdf_group_handle.exponential_window_value_at_frame_end*100)
        response_inds = netcdf_group_handle.variables['response_channel_indices'][...]
        reference_inds = netcdf_group_handle.variables['reference_channel_indices'][...]
        for row in range(self.definition_widget.reference_channels_selector.rowCount()):
            if row in reference_inds:
                widget = self.definition_widget.reference_channels_selector.cellWidget(row,1)
                widget.setChecked(True)
                widget = self.definition_widget.reference_channels_selector.cellWidget(row,0)
                widget.setChecked(True)
            elif row in response_inds:
                widget = self.definition_widget.reference_channels_selector.cellWidget(row,0)
                widget.setChecked(True)
                widget = self.definition_widget.reference_channels_selector.cellWidget(row,1)
                widget.setChecked(False)
            else:
                widget = self.definition_widget.reference_channels_selector.cellWidget(row,0)
                widget.setChecked(False)
                widget = self.definition_widget.reference_channels_selector.cellWidget(row,1)
                widget.setChecked(False)
                
    def update_gui(self,queue_data : tuple):
        """Update the environment's graphical user interface
        
        This function will receive data from the gui_update_queue that
        specifies how the user interface should be updated.  Data will usually
        be received as ``(instruction,data)`` pairs, where the ``instruction`` notes
        what operation should be taken or which widget should be modified, and
        the ``data`` notes what data should be used in the update.

        Parameters
        ----------
        queue_data : tuple
            A tuple containing ``(instruction,data)`` pairs where ``instruction``
            defines and operation or widget to be modified and ``data`` contains
            the data used to perform the operation.
        """
        # print('Got GUI Update {:}'.format(queue_data[0]))
        message,data = queue_data
        if message == 'spectral_update':
            (frames,total_frames,frequencies,
             self.last_frf,self.last_coherence,
             last_response_cpsd,last_reference_cpsd,
             self.last_condition) = data
            self.run_widget.channel_display_area.last_frf = self.last_frf
            self.run_widget.channel_display_area.last_coh = self.last_coherence.T
            if last_response_cpsd.ndim == 3:
                self.last_response_cpsd = np.einsum('fii->fi',last_response_cpsd)
            else:
                self.last_response_cpsd = last_response_cpsd
            if last_reference_cpsd.ndim == 3:
                self.last_reference_cpsd = np.einsum('fii->fi',last_reference_cpsd)
            else:
                self.last_reference_cpsd = last_reference_cpsd
            # Assemble autospectrum
            self.run_widget.channel_display_area.last_autospectrum = np.zeros((len(self.data_acquisition_parameters.channel_list),self.last_response_cpsd.shape[0]))
            for i,index in enumerate(self.environment_parameters.reference_channel_indices):
                self.run_widget.channel_display_area.last_autospectrum[index,:] = self.last_reference_cpsd[:,i].real
            for i,index in enumerate(self.environment_parameters.response_channel_indices):
                self.run_widget.channel_display_area.last_autospectrum[index,:] = self.last_response_cpsd[:,i].real
            self.run_widget.current_average_display.setValue(frames)
            for window in self.run_widget.channel_display_area.subWindowList():
                widget = window.widget()
                if widget.signal_selector.currentIndex() in [3,4,5,6,7]:
                    widget.update_data()
            if self.acquiring and not self.netcdf_handle is None:
                group = self.netcdf_handle.groups[self.environment_name]
                group.variables['frf_data_real'][:] = np.real(self.last_frf)
                group.variables['frf_data_imag'][:] = np.imag(self.last_frf)
                group.variables['coherence'][:] = self.last_coherence
            if self.acquiring and frames >= self.environment_parameters.num_averages:
                # print('Stopping Control')
                self.stop_control()
                self.acquiring = False
            # else:
            #     print('Continuing Control')
        elif message == 'time_frame':
            frame,accepted = data
            self.run_widget.channel_display_area.last_frame = frame
            self.run_widget.channel_display_area.last_spectrum = np.abs(np.fft.rfft(frame,axis=-1))
            for window in self.run_widget.channel_display_area.subWindowList():
                widget = window.widget()
                if widget.signal_selector.currentIndex() not in [3,4,5,6,7]:
                    widget.update_data()
            if self.netcdf_handle is not None and accepted:
                # Get current timestep
                num_timesteps = self.netcdf_handle.dimensions['time_samples'].size
                current_frame = num_timesteps//self.environment_parameters.samples_per_frame
                if current_frame < self.environment_parameters.num_averages:
                    timesteps = slice(num_timesteps,None,None)
                    self.netcdf_handle.variables['time_data'][:,timesteps] = frame
            if self.environment_parameters.accept_type == 'Manual' and not accepted:
                self.run_widget.accept_average_button.setEnabled(True)
                self.run_widget.reject_average_button.setEnabled(True)
            
        elif message == 'finished':
            self.run_widget.stop_test_button.setEnabled(False)
            self.run_widget.preview_test_button.setEnabled(True)
            self.run_widget.start_test_button.setEnabled(True)
            self.run_widget.select_file_button.setEnabled(True)
            self.run_widget.dof_override_table.setEnabled(True)
            self.run_widget.add_override_button.setEnabled(True)
            self.run_widget.remove_override_button.setEnabled(True)
            if not self.netcdf_handle is None:
                self.netcdf_handle.close()
                self.netcdf_handle = None
            
        elif message == 'enable':
            widget = None
            for parent in [self.definition_widget,self.system_id_widget,self.prediction_widget,self.run_widget]:
                try:
                    widget = getattr(parent,data)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Enable Widget {:}: not found in UI'.format(data))
            widget.setEnabled(True)
        elif message == 'disable':
            widget = None
            for parent in [self.definition_widget,self.system_id_widget,self.prediction_widget,self.run_widget]:
                try:
                    widget = getattr(parent,data)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Disable Widget {:}: not found in UI'.format(data))
            widget.setEnabled(False)
        else:
            widget = None
            for parent in [self.definition_widget,self.run_widget]:
                try:
                    widget = getattr(parent,message)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError('Cannot Update Widget {:}: not found in UI'.format(message))
            if type(widget) is QtWidgets.QDoubleSpinBox:
                widget.setValue(data)
            elif type(widget) is QtWidgets.QSpinBox:
                widget.setValue(data)
            elif type(widget) is QtWidgets.QLineEdit:
                widget.setText(data)
            elif type(widget) is QtWidgets.QListWidget:
                widget.clear()
                widget.addItems(['{:.3f}'.format(d) for d in data])
    
    @staticmethod
    def create_environment_template(environment_name : str, workbook : openpyxl.workbook.workbook.Workbook):
        """Creates a template worksheet in an Excel workbook defining the
        environment.
        
        This function creates a template worksheet in an Excel workbook that
        when filled out could be read by the controller to re-create the 
        environment.
        
        This function is the "write" counterpart to the 
        ``set_parameters_from_template`` function in the ``ModalUI`` class,
        which reads the values from the template file to populate the user
        interface.

        Parameters
        ----------
        environment_name : str :
            The name of the environment that will specify the worksheet's name
        workbook : openpyxl.workbook.workbook.Workbook :
            A reference to an ``openpyxl`` workbook.

        """
        worksheet = workbook.create_sheet(environment_name)
        worksheet.cell(1,1,'Control Type')
        worksheet.cell(1,2,'Modal')
        worksheet.cell(2,1,'Samples Per Frame:')
        worksheet.cell(2,2,'# Number of Samples per Measurement Frame')
        worksheet.cell(3,1,'Averaging Type:')
        worksheet.cell(3,2,'# Averaging Type')
        worksheet.cell(4,1,'Number of Averages:')
        worksheet.cell(4,2,'# Number of Averages used when computing the FRF')
        worksheet.cell(5,1,'Averaging Coefficient:')
        worksheet.cell(5,2,'# Averaging Coefficient for Exponential Averaging')
        worksheet.cell(6,1,'FRF Technique:')
        worksheet.cell(6,2,'# FRF Technique')
        worksheet.cell(7,1,'FRF Window:')
        worksheet.cell(7,2,'# Window used to compute FRF')
        worksheet.cell(8,1,'Exponential Window End Value:')
        worksheet.cell(8,2,'# Exponential Window Value at the end of the measurement frame (0.5 or 50%, not 50)')
        worksheet.cell(9,1,'FRF Overlap:')
        worksheet.cell(9,2,'# Overlap for FRF calculations (0.5 or 50%, not 50)')
        worksheet.cell(10,1,'Triggering Type:')
        worksheet.cell(10,2,'# One of "Free Run", "First Frame", or "Every Frame"')
        worksheet.cell(11,1,'Average Acceptance:')
        worksheet.cell(11,2,'# One of "Accept All", "Manual", or "Autoreject"')
        worksheet.cell(12,1,'Trigger Channel')
        worksheet.cell(12,2,'# Channel number (1-based) to use for triggering')
        worksheet.cell(13,1,'Pretrigger')
        worksheet.cell(13,2,'# Amount of frame to use as pretrigger (0.5 or 50%, not 50)')
        worksheet.cell(14,1,'Trigger Slope')
        worksheet.cell(14,2,'# One of "Positive" or "Negative"')
        worksheet.cell(15,1,'Trigger Level')
        worksheet.cell(15,2,'# Level to use to trigger the test as a fraction of the total range of the channel (0.5 or 50%, not 50)')
        worksheet.cell(16,1,'Hysteresis Level')
        worksheet.cell(16,2,'# Level that a channel must fall below before another trigger can be considered (0.5 or 50%, not 50)')
        worksheet.cell(17,1,'Hysteresis Frame Fraction')
        worksheet.cell(17,2,'# Fraction of the frame that a channel maintain hysteresis condition before another trigger can be considered (0.5 or 50%, not 50)')
        worksheet.cell(18,1,'Signal Generator Type')
        worksheet.cell(18,2,'# One of "None", "Random", "Burst Random", "Pseudorandom", "Chirp", "Square", or "Sine"')
        worksheet.cell(19,1,'Signal Generator Level')
        worksheet.cell(19,2,'# RMS voltage level for random signals, Peak voltage level for chirp, sine, and square pulse')
        worksheet.cell(20,1,'Signal Generator Frequency 1')
        worksheet.cell(20,2,'# Minimum frequency for broadband signals or frequency for sine and square  pulse')
        worksheet.cell(21,1,'Signal Generator Frequency 2')
        worksheet.cell(21,2,'# Maximum frequency for broadband signals.  Ignored for sine and square pulse')
        worksheet.cell(22,1,'Signal Generator On Fraction')
        worksheet.cell(22,2,'# Fraction of time that the burst or square wave is on (0.5 or 50%, not 50)')
        worksheet.cell(23,1,'Wait Time for Steady State')
        worksheet.cell(23,2,'# Time to wait after output starts to allow the system to reach steady state')
        worksheet.cell(24,1,'Autoaccept Script')
        worksheet.cell(24,2,'# File in which an autoacceptance function is defined')
        worksheet.cell(25,1,'Autoaccept Function')
        worksheet.cell(25,2,'# Function name in which the autoacceptance function is defined')
        worksheet.cell(26,1,'Reference Channels')
        worksheet.cell(26,2,'# List of channels, one per cell on this row')
        worksheet.cell(27,1,'Disabled Channels')
        worksheet.cell(27,2,'# List of channels, one per cell on this row')
        
    def set_parameters_from_template(self, worksheet : openpyxl.worksheet.worksheet.Worksheet):
        """
        Collects parameters for the user interface from the Excel template file
        
        This function reads a filled out template worksheet to create an
        environment.  Cells on this worksheet contain parameters needed to
        specify the environment, so this function should read those cells and
        update the UI widgets with those parameters.
        
        This function is the "read" counterpart to the 
        ``create_environment_template`` function in the ``ModalUI`` class,
        which writes a template file that can be filled out by a user.
        

        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            An openpyxl worksheet that contains the environment template.
            Cells on this worksheet should contain the parameters needed for the
            user interface.

        """
        self.definition_widget.samples_per_frame_selector.setValue(worksheet.cell(2,2).value)
        self.definition_widget.system_id_averaging_scheme_selector.setCurrentIndex(self.definition_widget.system_id_averaging_scheme_selector.findText(worksheet.cell(3,2).value))
        self.definition_widget.system_id_frames_to_average_selector.setValue(worksheet.cell(4,2).value)
        self.definition_widget.system_id_averaging_coefficient_selector.setValue(worksheet.cell(5,2).value)
        self.definition_widget.system_id_frf_technique_selector.setCurrentIndex(self.definition_widget.system_id_frf_technique_selector.findText(worksheet.cell(6,2).value))
        self.definition_widget.system_id_transfer_function_computation_window_selector.setCurrentIndex(self.definition_widget.system_id_transfer_function_computation_window_selector.findText(worksheet.cell(7,2).value))
        self.definition_widget.window_value_selector.setValue(worksheet.cell(8,2).value*100)
        self.definition_widget.system_id_overlap_percentage_selector.setValue(worksheet.cell(9,2).value*100)
        self.definition_widget.triggering_type_selector.setCurrentIndex(self.definition_widget.triggering_type_selector.findText(worksheet.cell(10,2).value))
        acceptance = worksheet.cell(11,2).value
        self.definition_widget.acceptance_selector.blockSignals(True)
        if acceptance == 'Autoreject':
            self.definition_widget.acceptance_selector.setCurrentIndex(2)
            self.acceptance_function = [worksheet.cell(24,2).value,worksheet.cell(25,2).value]
        else:
            self.definition_widget.acceptance_selector.setCurrentIndex(self.definition_widget.acceptance_selector.findText(acceptance))
            self.acceptance_function = None
        self.definition_widget.acceptance_selector.blockSignals(False)
        self.definition_widget.trigger_channel_selector.setCurrentIndex(worksheet.cell(12,2).value-1)
        self.definition_widget.pretrigger_selector.setValue(worksheet.cell(13,2).value*100)
        self.definition_widget.trigger_slope_selector.setCurrentIndex(self.definition_widget.trigger_slope_selector.findText(worksheet.cell(14,2).value))
        self.definition_widget.trigger_level_selector.setValue(worksheet.cell(15,2).value*100)
        self.definition_widget.hysteresis_selector.setValue(worksheet.cell(16,2).value*100)
        self.definition_widget.hysteresis_length_selector.setValue(worksheet.cell(17,2).value*100)
        signal_index = ["None", "Random", "Burst Random", "Pseudorandom", "Chirp", "Square", "Sine"].index(worksheet.cell(18,2).value)
        self.definition_widget.signal_generator_selector.setCurrentIndex(signal_index)
        level = worksheet.cell(19,2).value
        freq_1 = worksheet.cell(20,2).value
        freq_2 = worksheet.cell(21,2).value
        sig_on = worksheet.cell(22,2).value*100
        for widget in [self.definition_widget.random_rms_selector,
                       self.definition_widget.burst_rms_selector,
                       self.definition_widget.pseudorandom_rms_selector,
                       self.definition_widget.chirp_level_selector,
                       self.definition_widget.square_level_selector,
                       self.definition_widget.sine_level_selector]:
            widget.setValue(level)
        for widget in [self.definition_widget.random_min_frequency_selector,
                       self.definition_widget.burst_min_frequency_selector,
                       self.definition_widget.pseudorandom_min_frequency_selector,
                       self.definition_widget.chirp_min_frequency_selector,
                       self.definition_widget.square_frequency_selector,
                       self.definition_widget.sine_frequency_selector]:
            widget.setValue(freq_1)
        for widget in [self.definition_widget.random_max_frequency_selector,
                       self.definition_widget.burst_max_frequency_selector,
                       self.definition_widget.pseudorandom_max_frequency_selector,
                       self.definition_widget.chirp_max_frequency_selector]:
            widget.setValue(freq_2)
        for widget in [self.definition_widget.burst_on_percentage_selector,
                       self.definition_widget.square_percent_on_selector]:
            widget.setValue(sig_on)
        self.definition_widget.wait_for_steady_selector.setValue(worksheet.cell(23,2).value)
        column_index = 2
        while True:
            value = worksheet.cell(26,column_index).value
            if value is None or (isinstance(value,str) and value.strip() == ''):
                break
            widget = self.definition_widget.reference_channels_selector.cellWidget(int(value)-1,1)
            widget.setChecked(True)
            column_index += 1
        for i in range(self.definition_widget.reference_channels_selector.rowCount()):
            widget = self.definition_widget.reference_channels_selector.cellWidget(int(i),0)
            widget.setChecked(True)
        column_index = 2
        while True:
            value = worksheet.cell(27,column_index).value
            if value is None or (isinstance(value,str) and value.strip() == ''):
                break
            widget = self.definition_widget.reference_channels_selector.cellWidget(int(value)-1,0)
            widget.setChecked(False)
            column_index += 1

class ModalEnvironment(AbstractEnvironment):
    """Modal Environment class defining the interface with the controller"""
    
    def __init__(self,
                 environment_name : str,
                 queues : ModalQueues,
                 acquisition_active : mp.Value,
                 output_active : mp.Value):
        super().__init__(environment_name, queues.environment_command_queue,
                         queues.gui_update_queue,
                         queues.controller_communication_queue, queues.log_file_queue,
                         queues.data_in_queue, queues.data_out_queue,
                         acquisition_active,
                         output_active)
        self.queue_container = queues
        self.data_acquisition_parameters = None
        self.environment_parameters = None
        self.frame_number = 0
        self.siggen_shutdown_achieved = False
        self.collector_shutdown_achieved = False
        self.spectral_shutdown_achieved = False
        
        # Map commands
        self.map_command(ModalCommands.ACCEPT_FRAME,self.accept_frame)
        self.map_command(ModalCommands.START_CONTROL,self.start_environment)
        self.map_command(ModalCommands.RUN_CONTROL,self.run_control)
        self.map_command(ModalCommands.STOP_CONTROL,self.stop_environment)
        self.map_command(ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN,self.check_for_shutdown)
        self.map_command(SignalGenerationCommands.SHUTDOWN_ACHIEVED,self.siggen_shutdown_achieved_fn)
        self.map_command(DataCollectorCommands.SHUTDOWN_ACHIEVED,self.collector_shutdown_achieved_fn)
        self.map_command(SpectralProcessingCommands.SHUTDOWN_ACHIEVED,self.spectral_shutdown_achieved_fn)

    def initialize_data_acquisition_parameters(self,data_acquisition_parameters : DataAcquisitionParameters):
        """Initialize the data acquisition parameters in the environment.
        
        The environment will receive the global data acquisition parameters from
        the controller, and must set itself up accordingly.

        Parameters
        ----------
        data_acquisition_parameters : DataAcquisitionParameters :
            A container containing data acquisition parameters, including
            channels active in the environment as well as sampling parameters.
        """
        self.data_acquisition_parameters = data_acquisition_parameters
        
    def initialize_environment_test_parameters(self,environment_parameters : ModalMetadata):
        """
        Initialize the environment parameters specific to this environment
        
        The environment will recieve parameters defining itself from the
        user interface and must set itself up accordingly.

        Parameters
        ----------
        environment_parameters : ModalMetadata
            A container containing the parameters defining the environment

        """
        self.environment_parameters = environment_parameters
        
        # Set up the collector
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (DataCollectorCommands.INITIALIZE_COLLECTOR,
             self.get_data_collector_metadata()))
        # Set up the signal generation
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.INITIALIZE_PARAMETERS,
             self.get_signal_generation_metadata()))
        # Set up the spectral processing
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.INITIALIZE_PARAMETERS,
             self.get_spectral_processing_metadata()))

    def get_data_collector_metadata(self) -> CollectorMetadata:
        num_channels =  len(self.data_acquisition_parameters.channel_list)
        reference_channel_indices = self.environment_parameters.reference_channel_indices
        response_channel_indices = self.environment_parameters.response_channel_indices
        if self.environment_parameters.trigger_type == 'Free Run':
            acquisition_type = AcquisitionType.FREE_RUN
        elif self.environment_parameters.trigger_type == 'First Frame':
            acquisition_type = AcquisitionType.TRIGGER_FIRST_FRAME
        elif self.environment_parameters.trigger_type == 'Every Frame':
            acquisition_type = AcquisitionType.TRIGGER_EVERY_FRAME
        else:
            raise ValueError('Invalid Acquisition Type: {:}'.format(self.environment_parameters.trigger_type))
        if self.environment_parameters.accept_type == 'Accept All':
            acceptance = Acceptance.AUTOMATIC
            acceptance_function = None
        elif self.environment_parameters.accept_type == 'Manual':
            acceptance = Acceptance.MANUAL
            acceptance_function = None
        elif self.environment_parameters.accept_type == 'Autoreject...':
            acceptance = Acceptance.AUTOMATIC
            acceptance_function = self.environment_parameters.acceptance_function
        else:
            raise ValueError('Invalid Acceptance Type: {:}'.format(self.environment_parameters.accept_type))
        overlap_fraction = self.environment_parameters.overlap
        trigger_channel_index = self.environment_parameters.trigger_channel
        trigger_slope = TriggerSlope.POSITIVE if self.environment_parameters.trigger_slope_positive else TriggerSlope.NEGATIVE
        (trigger_level_v, trigger_level, hysterisis_level_v, trigger_hysteresis
         ) =  self.environment_parameters.get_trigger_levels(self.data_acquisition_parameters.channel_list)
        trigger_hysteresis_samples = self.environment_parameters.hysteresis_samples
        pretrigger_fraction = self.environment_parameters.pretrigger
        frame_size = self.environment_parameters.samples_per_frame
        if self.environment_parameters.frf_window == 'hann':
            window = Window.HANN
        elif self.environment_parameters.frf_window == 'rectangle':
            window =  Window.RECTANGLE
        elif self.environment_parameters.frf_window == 'exponential':
            window =  Window.EXPONENTIAL
        else:
            raise ValueError('Invalid Window Type: {:}'.format(self.environment_parameters.frf_window))
        window_parameter = -(frame_size) / np.log(self.environment_parameters.exponential_window_value_at_frame_end)
        return CollectorMetadata(
            num_channels,
            response_channel_indices,
            reference_channel_indices,
            acquisition_type,
            acceptance,
            acceptance_function,
            overlap_fraction,
            trigger_channel_index,
            trigger_slope,
            trigger_level,
            trigger_hysteresis,
            trigger_hysteresis_samples,
            pretrigger_fraction,
            frame_size,
            window,
            response_transformation_matrix = None,
            reference_transformation_matrix = None,
            window_parameter_2 = window_parameter)
    
    def get_spectral_processing_metadata(self) -> SpectralProcessingMetadata:
        averaging_type = AveragingTypes.LINEAR if self.environment_parameters.averaging_type == 'Linear' else AveragingTypes.EXPONENTIAL
        averages = self.environment_parameters.num_averages
        exponential_averaging_coefficient = self.environment_parameters.averaging_coefficient
        if self.environment_parameters.frf_technique == 'H1':
            frf_estimator = Estimator.H1
        elif self.environment_parameters.frf_technique == 'H2':
            frf_estimator = Estimator.H2
        elif self.environment_parameters.frf_technique == 'H3':
            frf_estimator = Estimator.H3
        elif self.environment_parameters.frf_technique == 'Hv':
            frf_estimator = Estimator.HV
        num_response_channels = len(self.environment_parameters.response_channel_indices)
        num_reference_channels = len(self.environment_parameters.reference_channel_indices)
        frequency_spacing = self.environment_parameters.frequency_spacing
        sample_rate = self.environment_parameters.sample_rate
        num_frequency_lines = self.environment_parameters.fft_lines
        return SpectralProcessingMetadata(
            averaging_type, 
            averages, 
            exponential_averaging_coefficient, 
            frf_estimator, 
            num_response_channels, 
            num_reference_channels, 
            frequency_spacing, 
            sample_rate,
            num_frequency_lines,
            compute_cpsd = False,
            compute_apsd = True)
    
    def get_signal_generation_metadata(self) -> SignalGenerationMetadata:
        return SignalGenerationMetadata(
            samples_per_write = self.data_acquisition_parameters.samples_per_write,
            level_ramp_samples = 1,
            output_transformation_matrix = None,
            disabled_signals = self.environment_parameters.disabled_signals
            )
    
    def get_signal_generator(self):
        return self.environment_parameters.get_signal_generator()

    def start_environment(self,data):
        self.log('Starting Modal')
        self.siggen_shutdown_achieved = False
        self.collector_shutdown_achieved = False
        self.spectral_shutdown_achieved = False
        
        # Set up the collector
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (DataCollectorCommands.FORCE_INITIALIZE_COLLECTOR,
             self.get_data_collector_metadata()))
        
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (DataCollectorCommands.SET_TEST_LEVEL,
             (self.environment_parameters.skip_frames,1)))
        time.sleep(0.01)
        
        # Set up the signal generation
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.INITIALIZE_PARAMETERS,
             self.get_signal_generation_metadata()))
        
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
             self.get_signal_generator()))
        
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.MUTE,None))
        
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.ADJUST_TEST_LEVEL,1.0))
        
        # Tell the collector to start acquiring data
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (DataCollectorCommands.ACQUIRE,None))
        
        # Tell the signal generation to start generating signals
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.GENERATE_SIGNALS,None))
        
        # Set up the spectral processing
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.INITIALIZE_PARAMETERS,
             self.get_spectral_processing_metadata()))
        
        # Tell the spectral analysis to clear and start acquiring
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.CLEAR_SPECTRAL_PROCESSING,None))
        
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.RUN_SPECTRAL_PROCESSING,None))
        
        self.queue_container.environment_command_queue.put(
            self.environment_name,
            (ModalCommands.RUN_CONTROL,None))
    
    def run_control(self,data):
        # Pull data off the spectral queue
        spectral_data = flush_queue(self.queue_container.updated_spectral_quantities_queue,timeout = WAIT_TIME)
        if len(spectral_data) > 0:
            self.log('Received Data')
            (frames,
             frequencies,
             frf,
             coherence,
             response_cpsd,
             reference_cpsd,
             condition) = spectral_data[-1]
            self.gui_update_queue.put((self.environment_name,
                                       ('spectral_update',
                                        (frames,
                                        self.environment_parameters.num_averages,
                                        frequencies,
                                        frf,
                                        coherence,
                                        response_cpsd,
                                        reference_cpsd,
                                        condition))))
        else:
            time.sleep(WAIT_TIME)
        self.queue_container.environment_command_queue.put(
            self.environment_name,
            (ModalCommands.RUN_CONTROL,None))
    
    def siggen_shutdown_achieved_fn(self,data):
        self.siggen_shutdown_achieved = True
        
    def collector_shutdown_achieved_fn(self,data):
        self.collector_shutdown_achieved = True
        
    def spectral_shutdown_achieved_fn(self,data):
        self.spectral_shutdown_achieved = True
        
    def check_for_shutdown(self,data):
        if (self.siggen_shutdown_achieved and self.collector_shutdown_achieved
            and self.spectral_shutdown_achieved):
            self.log('Shutdown Achieved')
            self.gui_update_queue.put((self.environment_name,('finished',None)))
        else:
            # print('self.siggen_shutdown_achieved: {:}'.format(self.siggen_shutdown_achieved))
            # print('self.collector_shutdown_achieved: {:}'.format(self.collector_shutdown_achieved))
            # print('self.spectral_shutdown_achieved: {:}'.format(self.spectral_shutdown_achieved))
            # Recheck some time later
            time.sleep(1)
            self.environment_command_queue.put(
                self.environment_name,
                (ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN,None))
    
    def accept_frame(self,data):
        self.queue_container.collector_command_queue.put(self.environment_name,(DataCollectorCommands.ACCEPT,data))
    
    def stop_environment(self,data):
        """Stop the environment gracefully
        
        This function defines the operations to shut down the environment
        gracefully so there is no hard stop that might damage test equipment
        or parts.

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        """
        self.log('Stopping Control')
        flush_queue(self.queue_container.environment_command_queue)
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (DataCollectorCommands.SET_TEST_LEVEL,
             (1000,1)))
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (SignalGenerationCommands.START_SHUTDOWN,None))
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.STOP_SPECTRAL_PROCESSING,None))
        self.queue_container.environment_command_queue.put(
            self.environment_name,
            (ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN,None))
    
    def quit(self,data):
        """Returns True to stop the ``run`` while loop and exit the process

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        Returns
        -------
        True :
            This function returns True to signal to the ``run`` while loop
            that it is time to close down the environment.

        """
        for queue in [self.queue_container.spectral_command_queue,
              self.queue_container.signal_generation_command_queue,
              self.queue_container.collector_command_queue]:
            queue.put(self.environment_name,(GlobalCommands.QUIT,None))
        return True
    
def modal_process(environment_name : str,
                 input_queue : VerboseMessageQueue,
                 gui_update_queue : Queue,
                 controller_communication_queue : VerboseMessageQueue,
                 log_file_queue : Queue,
                 data_in_queue : Queue,
                 data_out_queue : Queue,
                 acquisition_active : mp.Value,
                 output_active : mp.Value):
    """Modal environment process function called by multiprocessing
    
    This function defines the Modal Environment process that
    gets run by the multiprocessing module when it creates a new process.  It
    creates a ModalEnvironment object and runs it.

    Parameters
    ----------
    environment_name : str :
        Name of the environment
    input_queue : VerboseMessageQueue :
        Queue containing instructions for the environment
    gui_update_queue : Queue :
        Queue where GUI updates are put
    controller_communication_queue : Queue :
        Queue for global communications with the controller
    log_file_queue : Queue :
        Queue for writing log file messages
    data_in_queue : Queue :
        Queue from which data will be read by the environment
    data_out_queue : Queue :
        Queue to which data will be written that will be output by the hardware.

    """
    queue_container = ModalQueues(environment_name,input_queue, gui_update_queue,
                               controller_communication_queue,
                               data_in_queue, data_out_queue, log_file_queue)

    spectral_proc = mp.Process(target=spectral_processing_process,
                               args=(environment_name,
                                     queue_container.spectral_command_queue,
                                     queue_container.data_for_spectral_computation_queue,
                                     queue_container.updated_spectral_quantities_queue,
                                     queue_container.environment_command_queue,
                                     queue_container.gui_update_queue,
                                     queue_container.log_file_queue))
    spectral_proc.start()
    siggen_proc = mp.Process(target=signal_generation_process,args=(environment_name,
                                                                    queue_container.signal_generation_command_queue,
                                                                    queue_container.signal_generation_update_queue,
                                                                    queue_container.data_out_queue,
                                                                    queue_container.environment_command_queue,
                                                                    queue_container.log_file_queue,
                                                                    queue_container.gui_update_queue))
    siggen_proc.start()
    collection_proc = mp.Process(target=data_collector_process,
                                 args=(environment_name,
                                       queue_container.collector_command_queue,
                                       queue_container.data_in_queue,
                                       [queue_container.data_for_spectral_computation_queue],
                                       queue_container.environment_command_queue,
                                       queue_container.log_file_queue,
                                       queue_container.gui_update_queue))
    
    collection_proc.start()
    
    process_class = ModalEnvironment(
            environment_name,
            queue_container,
            acquisition_active,
            output_active)
    process_class.run()
    
    # Rejoin all the processes
    process_class.log('Joining Subprocesses')
    process_class.log('Joining Spectral Computation')
    spectral_proc.join()
    process_class.log('Joining Signal Generation')
    siggen_proc.join()
    process_class.log('Joining Data Collection')
    collection_proc.join()