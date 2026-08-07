import numpy as np
import netCDF4 as nc4
import openpyxl

import numpy as np

from rattlesnake.environment.environment_registry import (
    ENVIRONMENT_COMMANDS,
    ENVIRONMENT_METADATA,
)
from rattlesnake.environment.environment_utilities import EnvironmentType
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.hardware.hardware_registry import HARDWARE_METADATA
from rattlesnake.hardware.hardware_utilities import HardwareType
from rattlesnake.profile_manager import ProfileEvent
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.utilities import GlobalCommands, RattlesnakeError


def load_metadata_from_netcdf(dataset):
    """Loads hardware and environment metadata from an open netCDF4 dataset

    Parameters
    ----------
    dataset : netCDF4.Dataset
        An open netCDF4 dataset containing Rattlesnake hardware and
        environment metadata

    Returns
    -------
    hardware_metadata : HardwareMetadata
        The hardware metadata loaded from the dataset
    environment_metadata_list : list
        A list of environment metadata objects, one per environment stored
        in the dataset
    """
    hardware_type = HardwareType(dataset.hardware)
    hardware_metadata_class = HARDWARE_METADATA[hardware_type]
    hardware_metadata = hardware_metadata_class.load_metadata_from_netcdf(dataset)

    # Environments
    environment_metadata_list = []
    for environment_index, environment_name in enumerate(
        dataset.variables["environment_names"][...],
    ):
        # Discover environment type
        environment_group = dataset.groups[environment_name]
        try:
            environment_type_int = dataset.variables["environment_types"][
                environment_index
            ]
            environment_type = EnvironmentType(environment_type_int)
        except (KeyError, ValueError):
            environment_type = discover_environment_type_in_old_netcdf(
                environment_group
            )

        # Create metadata class and append to list
        environment_metadata_class = ENVIRONMENT_METADATA[environment_type]
        channel_list_bools = dataset.variables["environment_active_channels"][
            :, environment_index
        ]
        environment_metadata = environment_metadata_class.load_metadata_from_netcdf(
            environment_group,
            environment_name,
            channel_list_bools,
            hardware_metadata,
        )
        environment_metadata_list.append(environment_metadata)

    return (hardware_metadata, environment_metadata_list)


_LEGACY_ENVIRONMENT_TYPE_ATTRIBUTES = {
    "tracking_filter_type": EnvironmentType.SINE,
    "update_tf_during_control": EnvironmentType.RANDOM,
    "num_averages": EnvironmentType.MODAL,
    "test_level_ramp_time": EnvironmentType.TRANSIENT,
    "cancel_rampdown_time": EnvironmentType.TIME,
}


def discover_environment_type_in_old_netcdf(environment_group):
    """Infers the environment type from attributes in a legacy netCDF4 file

    Older netCDF4 files do not store an explicit environment type, so this
    function infers it from the presence of environment-specific attributes.

    Parameters
    ----------
    environment_group : netCDF4.Group
        The netCDF4 group corresponding to a single environment

    Returns
    -------
    EnvironmentType
        The inferred environment type

    Raises
    ------
    RattlesnakeError
        If the environment type cannot be determined from the group's
        attributes
    """
    for attribute, environment_type in _LEGACY_ENVIRONMENT_TYPE_ATTRIBUTES.items():
        if hasattr(environment_group, attribute):
            return environment_type
    raise RattlesnakeError("Invalid netcdf4 file")


def load_metadata_from_workbook(workbook):
    """Loads hardware, environment, and profile metadata from an Excel workbook

    Parameters
    ----------
    workbook : openpyxl.workbook.Workbook
        An open Excel workbook containing a "Hardware" sheet, a channel
        sheet, one sheet per environment, and a "Test Profile" sheet

    Returns
    -------
    hardware_metadata : HardwareMetadata
        The hardware metadata loaded from the workbook
    environment_metadata_list : list
        A list of environment metadata objects, one per environment sheet
    profile_event_list : list[ProfileEvent]
        The list of profile events loaded from the "Test Profile" sheet
    """
    hardware_sheet = workbook["Hardware"]
    for row in hardware_sheet.rows:
        name = str(row[0].value).lower().strip().replace(" ", "_")
        value = row[1].value
        if value is None or value == "":
            continue
        match name:
            case "hardware_type":
                hardware_type_int = int(value)
                break
    hardware_type = HardwareType(hardware_type_int)

    hardware_metadata_class = HARDWARE_METADATA[hardware_type]
    hardware_metadata = hardware_metadata_class.load_metadata_from_workbook(workbook)

    environment_names = []
    environment_channel_list_bools = {}
    sheets = workbook.sheetnames
    if len(sheets) > 1:
        sheets = [sheet for sheet in sheets if "channel" in sheet.lower()]
    channel_sheet = workbook[sheets[0]]
    col = 24
    num_channels = len(hardware_metadata.channel_list)
    while True:
        environment_name = channel_sheet.cell(row=2, column=col).value

        # Stop if empty or None
        if environment_name is None or str(environment_name).strip() == "":
            break

        # Build environment channel list
        environment_active_channels = [False] * num_channels
        for i in range(num_channels):
            row = 3 + i
            value = channel_sheet.cell(row=row, column=col).value

            if value is not None and str(value).strip() != "":
                environment_active_channels[i] = True

        environment_names.append(environment_name)
        environment_channel_list_bools[environment_name] = environment_active_channels
        col += 1

    environment_metadata_list = []
    environment_types = {"Global": "Global"}
    for environment_name in environment_names:
        environment_sheet = workbook[environment_name]
        environment_type_name = environment_sheet.cell(row=1, column=2).value
        environment_type_name = str(environment_type_name).upper()
        environment_type = EnvironmentType[environment_type_name]
        environment_types[environment_name] = environment_type

        environment_metadata_class = ENVIRONMENT_METADATA[environment_type]
        channel_list_bools = environment_channel_list_bools[environment_name]
        environment_metadata = environment_metadata_class.load_metadata_from_worksheet(
            environment_sheet,
            environment_name,
            channel_list_bools,
            hardware_metadata,
        )
        environment_metadata_list.append(environment_metadata)

    profile_event_list = load_profile_from_workbook(workbook, environment_types)

    workbook.close()

    return (hardware_metadata, environment_metadata_list, profile_event_list)


def save_rattlesnake_to_workbook(
    workbook,
    hardware_metadata=None,
    environment_metadata_dict=None,
    profile_event_list=None,
):
    """Saves hardware, environment, and profile metadata to an Excel workbook

    Parameters
    ----------
    workbook : openpyxl.workbook.Workbook
        The workbook to save the metadata to. The active sheet is used as
        the channel sheet and additional sheets are created for each
        environment and for the test profile.
    hardware_metadata : HardwareMetadata, optional
        The hardware metadata to save
    environment_metadata_dict : dict, optional
        A dictionary where the keys are environment names and the values are
        either an EnvironmentType (to create a blank template) or an
        environment metadata object to save
    profile_event_list : list[ProfileEvent], optional
        The list of profile events to save to the "Test Profile" sheet
    """
    # Open workbook and save to blank template
    channel_worksheet = workbook.active
    HardwareMetadata.save_blank_hardware_to_workbook(workbook)
    # Save hardware metadata values
    if hardware_metadata:
        hardware_metadata.save_metadata_to_workbook(workbook)

    channel_worksheet.cell(row=1, column=24, value="Environments")
    # Save environment metadata values
    if environment_metadata_dict:
        for col, (environment_name, environment_metadata) in enumerate(
            environment_metadata_dict.items()
        ):
            col_idx = col + 24
            channel_worksheet.cell(row=2, column=col_idx, value=environment_name)
            environment_worksheet = workbook.create_sheet(environment_name)
            if isinstance(environment_metadata, EnvironmentType):
                ENVIRONMENT_METADATA[
                    environment_metadata
                ].create_blank_worksheet_template(environment_worksheet)
            else:
                for row in environment_metadata.channel_indices:
                    row_idx = row + 3
                    channel_worksheet.cell(row=row_idx, column=col_idx, value="x")
                environment_metadata.save_metadata_to_worksheet(environment_worksheet)

    # Save profile event list
    profile_sheet = workbook.create_sheet("Test Profile")
    save_profile_to_workbook(profile_sheet, profile_event_list)


def load_profile_from_workbook(workbook, environment_types):
    """Loads the test profile event list from an Excel workbook

    Parameters
    ----------
    workbook : openpyxl.workbook.Workbook
        An open Excel workbook containing a "Test Profile" sheet
    environment_types : dict
        A dictionary mapping environment names to their EnvironmentType,
        used to resolve environment-specific commands

    Returns
    -------
    list[ProfileEvent]
        The list of profile events loaded from the "Test Profile" sheet
    """
    profile_sheet = workbook["Test Profile"]
    index = 2
    profile_event_list = []
    while True:
        timestamp = profile_sheet.cell(index, 1).value
        if timestamp is None or (
            isinstance(timestamp, str) and timestamp.strip() == ""
        ):
            break
        timestamp = float(timestamp)

        environment_name = profile_sheet.cell(index, 2).value
        environment_type = environment_types[environment_name]

        # I have to conver the command string to an actual command
        command = profile_sheet.cell(index, 3).value
        command = str(command).upper().strip().replace(" ", "_")
        if command in GlobalCommands.__members__:
            command = GlobalCommands[command]
        elif command in ENVIRONMENT_COMMANDS[environment_type].__members__:
            command = ENVIRONMENT_COMMANDS[environment_type][command]
        elif (
            command == "SET_ENVIRONMENT_INSTRUCTIONS"
        ):  # This should not be set by workbook.
            index += 1
            continue
        else:
            raise RattlesnakeError(
                f"Invalid command: {command} for {environment_name} | "
                f"{environment_type}"
            )

        data = profile_sheet.cell(index, 4).value
        data = None if isinstance(data, str) and not data.strip() else data

        event = ProfileEvent(timestamp, environment_name, command, data)
        profile_event_list.append(event)
        index += 1
    workbook.close()

    return profile_event_list


def save_profile_to_workbook(profile_sheet, profile_event_list):
    """Saves a test profile event list to an Excel worksheet

    Parameters
    ----------
    profile_sheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet to write the profile event list to
    profile_event_list : list[ProfileEvent]
        The list of profile events to save. Events with the
        SET_ENVIRONMENT_INSTRUCTIONS command are skipped.
    """
    profile_sheet.cell(1, 1, "Time (s)")
    profile_sheet.cell(1, 2, "Environment")
    profile_sheet.cell(1, 3, "Operation")
    profile_sheet.cell(1, 4, "Data")
    # Fill out values
    if profile_event_list:
        row_idx = 2
        for event in profile_event_list:
            if event.command == UICommands.SET_ENVIRONMENT_INSTRUCTIONS:
                continue
            profile_sheet.cell(row_idx, 1, str(event.timestamp))
            profile_sheet.cell(row_idx, 2, event.environment_name)
            profile_sheet.cell(row_idx, 3, event.command.label)
            profile_sheet.cell(row_idx, 4, str(event.data))
            row_idx = row_idx + 1
